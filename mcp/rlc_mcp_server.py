"""
rlc_mcp — read-only MCP server for the RLC-Agent database and workbook files.

Gives a local LLM (Ollama via mcphost) safe, structured access to:
  * The RLC-Agent PostgreSQL database (bronze/silver/gold/reports schemas) — SELECT only.
  * A single allowed directory of spreadsheet files (.xlsx/.xlsm/.csv) — read only.

Safety is enforced structurally, not by prompt:
  * Connection sets default_transaction_read_only=on and a statement timeout.
  * SQL is validated (single statement, must begin SELECT/WITH).
  * Row counts and cell widths are capped so a bad query can't flood the model's context.
  * Workbook access is confined to RLC_WORKBOOK_ROOT (path-traversal safe).

Configuration (environment variables):
  RLC_PG_DSN         e.g. postgresql://rlc_readonly:PASSWORD@localhost:5432/rlc
  RLC_WORKBOOK_ROOT  e.g. C:\\Users\\torem\\Documents\\RLC\\workbooks
  RLC_MAX_ROWS       optional, default 200
  RLC_STMT_TIMEOUT_MS optional, default 15000

Run directly for a smoke test:  python rlc_mcp_server.py --selftest
Normally launched by an MCP host (mcphost / Claude Desktop) over stdio.
"""

from __future__ import annotations

import csv
import os
import re
import sys
from pathlib import Path
from typing import Any, Optional

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field

# --- MCP SDK compatibility: 2.x renamed FastMCP -> MCPServer -----------------
try:  # SDK >= 2.0
    from mcp.server.mcpserver import MCPServer as _Server
except ImportError:  # SDK 1.x
    from mcp.server.fastmcp import FastMCP as _Server  # type: ignore

mcp = _Server("rlc_mcp")

# --- Configuration -----------------------------------------------------------

PG_DSN = os.environ.get("RLC_PG_DSN", "")
WORKBOOK_ROOT = Path(os.environ.get("RLC_WORKBOOK_ROOT", "")).resolve() if os.environ.get("RLC_WORKBOOK_ROOT") else None
MAX_ROWS = int(os.environ.get("RLC_MAX_ROWS", "200"))
STMT_TIMEOUT_MS = int(os.environ.get("RLC_STMT_TIMEOUT_MS", "15000"))
MAX_CELL_CHARS = 200          # truncate any single value beyond this
MAX_WORKBOOK_CELLS = 2000     # hard cap on cells returned from a sheet read
DEFAULT_SCHEMAS = ("bronze", "silver", "gold", "reports")
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}

# --- Database helpers --------------------------------------------------------


def _connect():
    """Open a read-only, time-limited Postgres connection."""
    if not PG_DSN:
        raise RuntimeError(
            "RLC_PG_DSN is not set. Set it to a read-only connection string, e.g. "
            "postgresql://rlc_readonly:PASSWORD@localhost:5432/rlc"
        )
    conn = psycopg2.connect(PG_DSN)
    conn.autocommit = True
    with conn.cursor() as cur:
        # Belt-and-suspenders: even if the role isn't SELECT-only, this session is.
        cur.execute("SET default_transaction_read_only = on")
        cur.execute(f"SET statement_timeout = {STMT_TIMEOUT_MS}")
    return conn


_SQL_FORBIDDEN = re.compile(
    r"\b(insert|update|delete|drop|truncate|alter|create|grant|revoke|copy|vacuum|call|do)\b",
    re.IGNORECASE,
)


def _validate_sql(sql: str) -> str:
    """Allow exactly one SELECT/WITH statement. Return cleaned SQL or raise."""
    cleaned = sql.strip().rstrip(";").strip()
    if not cleaned:
        raise ValueError("Empty query.")
    if ";" in cleaned:
        raise ValueError("Multiple statements are not allowed. Send one SELECT at a time.")
    first_word = cleaned.split(None, 1)[0].lower()
    if first_word not in ("select", "with"):
        raise ValueError(
            f"Only SELECT queries are allowed (query began with '{first_word.upper()}'). "
            "This server is read-only."
        )
    if _SQL_FORBIDDEN.search(cleaned):
        # WITH ... INSERT etc. — the read-only session would reject it anyway,
        # but fail fast with a clear message.
        raise ValueError("Query contains a write/DDL keyword. This server is read-only.")
    return cleaned


def _truncate(value: Any) -> str:
    s = "" if value is None else str(value)
    return s if len(s) <= MAX_CELL_CHARS else s[: MAX_CELL_CHARS] + "…"


def _rows_to_table(columns: list[str], rows: list[tuple]) -> str:
    """Compact pipe-delimited table — cheap on tokens, easy for small models."""
    lines = [" | ".join(columns)]
    lines.append("-" * min(len(lines[0]), 120))
    for row in rows:
        lines.append(" | ".join(_truncate(v) for v in row))
    return "\n".join(lines)


# --- Workbook helpers --------------------------------------------------------


def _resolve_workbook_path(relative: str) -> Path:
    """Resolve a user-supplied path safely inside WORKBOOK_ROOT."""
    if WORKBOOK_ROOT is None:
        raise RuntimeError(
            "RLC_WORKBOOK_ROOT is not set. Set it to the directory containing your workbooks."
        )
    if not WORKBOOK_ROOT.is_dir():
        raise RuntimeError(f"Workbook root does not exist: {WORKBOOK_ROOT}")
    candidate = (WORKBOOK_ROOT / relative).resolve()
    if WORKBOOK_ROOT not in candidate.parents and candidate != WORKBOOK_ROOT:
        raise ValueError("Path escapes the allowed workbook directory.")
    if candidate.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Only {sorted(ALLOWED_EXTENSIONS)} files can be read.")
    if not candidate.is_file():
        raise FileNotFoundError(
            f"File not found: {relative}. Use rlc_list_workbooks to see available files."
        )
    return candidate


# --- Tool input models -------------------------------------------------------


class ListTablesInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True, extra="forbid")
    schema_name: Optional[str] = Field(
        default=None,
        description="Restrict to one schema (bronze, silver, gold, reports). Omit for all four.",
    )


class DescribeTableInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True, extra="forbid")
    table: str = Field(
        ...,
        description="Schema-qualified table name, e.g. 'gold.curve_term' or 'silver.price_mark'.",
        min_length=3,
        max_length=200,
    )


class QueryInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True, extra="forbid")
    sql: str = Field(
        ...,
        description=(
            "A single SELECT (or WITH...SELECT) statement. Always schema-qualify tables "
            "(e.g. gold.curve_term). Add your own LIMIT for large tables; results are "
            f"capped at {MAX_ROWS} rows regardless."
        ),
        min_length=6,
    )


class ListWorkbooksInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True, extra="forbid")
    pattern: Optional[str] = Field(
        default=None,
        description="Optional case-insensitive substring filter on filename, e.g. 'tallow'.",
    )


class ReadWorkbookInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True, extra="forbid")
    path: str = Field(
        ...,
        description="Filename or relative path inside the workbook directory, e.g. 'tallow_balance.xlsx'.",
    )
    sheet: Optional[str] = Field(
        default=None,
        description="Sheet name. Omit to list sheets and their dimensions instead of reading data.",
    )
    max_rows: int = Field(
        default=50,
        description="Rows to read from the top of the sheet (header included).",
        ge=1,
        le=500,
    )


# --- Tools -------------------------------------------------------------------


@mcp.tool(
    name="rlc_list_tables",
    annotations={"readOnlyHint": True, "destructiveHint": False, "idempotentHint": True, "openWorldHint": False},
)
def rlc_list_tables(params: ListTablesInput) -> str:
    """List tables and views in the RLC database (bronze/silver/gold/reports schemas)
    with approximate row counts. Start here to see what data exists.

    Returns one line per table: schema.table (kind, ~rows).
    """
    try:
        schemas = (params.schema_name,) if params.schema_name else DEFAULT_SCHEMAS
        with _connect() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT n.nspname, c.relname,
                       CASE c.relkind WHEN 'r' THEN 'table' WHEN 'v' THEN 'view'
                            WHEN 'm' THEN 'matview' ELSE c.relkind::text END,
                       c.reltuples::bigint
                FROM pg_class c
                JOIN pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = ANY(%s) AND c.relkind IN ('r','v','m')
                ORDER BY n.nspname, c.relname
                """,
                (list(schemas),),
            )
            rows = cur.fetchall()
        if not rows:
            return f"No tables found in schemas: {', '.join(schemas)}."
        lines = [f"{s}.{t} ({kind}, ~{max(n, 0):,} rows)" for s, t, kind, n in rows]
        return f"{len(lines)} tables/views:\n" + "\n".join(lines)
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="rlc_describe_table",
    annotations={"readOnlyHint": True, "destructiveHint": False, "idempotentHint": True, "openWorldHint": False},
)
def rlc_describe_table(params: DescribeTableInput) -> str:
    """Show the columns, types, and nullability of one table, plus its primary key.
    Always call this before writing a query against an unfamiliar table —
    do not guess column names.

    Args: table — schema-qualified name like 'gold.curve_term'.
    """
    try:
        if "." not in params.table:
            return "Error: use a schema-qualified name, e.g. 'gold.curve_term'."
        schema, table = params.table.split(".", 1)
        with _connect() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT column_name, data_type, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                ORDER BY ordinal_position
                """,
                (schema, table),
            )
            cols = cur.fetchall()
            if not cols:
                return (
                    f"Error: table '{params.table}' not found. "
                    "Call rlc_list_tables to see valid names."
                )
            cur.execute(
                """
                SELECT a.attname
                FROM pg_index i
                JOIN pg_attribute a ON a.attrelid = i.indrelid AND a.attnum = ANY(i.indkey)
                WHERE i.indrelid = %s::regclass AND i.indisprimary
                """,
                (params.table,),
            )
            pk = [r[0] for r in cur.fetchall()]
        lines = [f"{params.table} — {len(cols)} columns"]
        for name, dtype, nullable, default in cols:
            bits = [dtype]
            if nullable == "NO":
                bits.append("not null")
            if default:
                bits.append(f"default {_truncate(default)}")
            lines.append(f"  {name}: {', '.join(bits)}")
        if pk:
            lines.append(f"primary key: ({', '.join(pk)})")
        return "\n".join(lines)
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="rlc_query_database",
    annotations={"readOnlyHint": True, "destructiveHint": False, "idempotentHint": True, "openWorldHint": False},
)
def rlc_query_database(params: QueryInput) -> str:
    """Run one read-only SELECT against the RLC PostgreSQL database and return the
    result as a compact table. Writes are impossible: the session is read-only and
    non-SELECT statements are rejected.

    Workflow: rlc_list_tables -> rlc_describe_table -> this tool.
    Results are capped at a fixed row limit; if the output says it was truncated,
    add filters or aggregate instead of asking for more rows.
    """
    try:
        sql = _validate_sql(params.sql)
        with _connect() as conn, conn.cursor() as cur:
            cur.execute(sql)
            if cur.description is None:
                return "Query ran but returned no result set."
            columns = [d.name for d in cur.description]
            rows = cur.fetchmany(MAX_ROWS + 1)
        truncated = len(rows) > MAX_ROWS
        rows = rows[:MAX_ROWS]
        out = _rows_to_table(columns, rows)
        footer = f"\n({len(rows)} rows"
        footer += f", TRUNCATED at {MAX_ROWS} — refine the query)" if truncated else ")"
        return out + footer
    except psycopg2.errors.QueryCanceled:
        return (
            f"Error: query exceeded the {STMT_TIMEOUT_MS} ms timeout. "
            "Add a WHERE clause or LIMIT, or aggregate instead of selecting raw rows."
        )
    except psycopg2.Error as e:
        # Postgres error messages are genuinely useful to the model — pass them through.
        return f"SQL error: {e.pgerror or e}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="rlc_list_workbooks",
    annotations={"readOnlyHint": True, "destructiveHint": False, "idempotentHint": True, "openWorldHint": False},
)
def rlc_list_workbooks(params: ListWorkbooksInput) -> str:
    """List spreadsheet files (.xlsx/.xlsm/.csv) available in the allowed workbook
    directory, with sizes. Use before rlc_read_workbook to find exact filenames.
    """
    try:
        if WORKBOOK_ROOT is None or not WORKBOOK_ROOT.is_dir():
            return "Error: RLC_WORKBOOK_ROOT is not set or does not exist."
        files = sorted(
            p for p in WORKBOOK_ROOT.rglob("*")
            if p.is_file() and p.suffix.lower() in ALLOWED_EXTENSIONS
            and not p.name.startswith("~$")  # skip Excel lock files
        )
        if params.pattern:
            needle = params.pattern.lower()
            files = [p for p in files if needle in p.name.lower()]
        if not files:
            return "No matching workbook files found."
        lines = [
            f"{p.relative_to(WORKBOOK_ROOT)} ({p.stat().st_size / 1024:.0f} KB)"
            for p in files[:100]
        ]
        return f"{len(files)} files:\n" + "\n".join(lines)
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="rlc_read_workbook",
    annotations={"readOnlyHint": True, "destructiveHint": False, "idempotentHint": True, "openWorldHint": False},
)
def rlc_read_workbook(params: ReadWorkbookInput) -> str:
    """Read a spreadsheet from the allowed workbook directory.

    Two modes:
      * Omit 'sheet' -> returns the list of sheet names with their dimensions.
      * Provide 'sheet' -> returns the top max_rows rows of that sheet as a table.
    For CSV files the 'sheet' argument is ignored and rows are read directly.
    """
    try:
        path = _resolve_workbook_path(params.path)

        if path.suffix.lower() == ".csv":
            with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
                reader = csv.reader(f)
                rows = []
                for i, row in enumerate(reader):
                    if i >= params.max_rows:
                        break
                    rows.append(tuple(row))
            if not rows:
                return "CSV file is empty."
            header = [str(c) for c in rows[0]]
            return _rows_to_table(header, rows[1:]) + f"\n({len(rows) - 1} data rows shown)"

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if params.sheet is None:
                lines = [f"{path.name} — {len(wb.sheetnames)} sheets:"]
                for name in wb.sheetnames:
                    ws = wb[name]
                    lines.append(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")
                lines.append("Call again with a 'sheet' name to read data.")
                return "\n".join(lines)

            if params.sheet not in wb.sheetnames:
                return (
                    f"Error: sheet '{params.sheet}' not found. "
                    f"Available: {', '.join(wb.sheetnames)}"
                )
            ws = wb[params.sheet]
            n_cols = min(ws.max_column or 1, 30)
            n_rows = min(params.max_rows, MAX_WORKBOOK_CELLS // max(n_cols, 1))
            data = []
            for i, row in enumerate(ws.iter_rows(max_row=n_rows, max_col=n_cols, values_only=True)):
                data.append(row)
            if not data:
                return "Sheet is empty."
            columns = [
                str(c) if c is not None else get_column_letter(j + 1)
                for j, c in enumerate(data[0])
            ]
            body = data[1:]
            note = ""
            if ws.max_row and ws.max_row > n_rows:
                note = f" of {ws.max_row} total — raise max_rows or use the database instead"
            return _rows_to_table(columns, body) + f"\n({len(body)} data rows shown{note})"
        finally:
            wb.close()
    except Exception as e:
        return f"Error: {e}"


# --- Entrypoint --------------------------------------------------------------


def _selftest() -> int:
    """Quick local smoke test without an MCP host."""
    print("rlc_mcp selftest")
    print(f"  RLC_PG_DSN set: {'yes' if PG_DSN else 'NO — set it'}")
    print(f"  RLC_WORKBOOK_ROOT: {WORKBOOK_ROOT or 'NOT SET'}")
    if PG_DSN:
        try:
            print(rlc_list_tables(ListTablesInput())[:2000])
        except Exception as e:
            print(f"  DB check failed: {e}")
    if WORKBOOK_ROOT:
        print(rlc_list_workbooks(ListWorkbooksInput())[:2000])
    return 0


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        sys.exit(_selftest())
    mcp.run(transport="stdio")
