"""
Iowa DNR Title V Permit Capacity Extractor

Downloads Title V operating permits from Iowa DNR and extracts
equipment-level capacity data for soybean/oilseed processing facilities.

Data Sources:
  - Draft & Final Permits page: facility -> media ID mapping
  - PDF permits: equipment descriptions, rated capacities

Usage:
    python iowa_capacity_collector.py                    # Download + parse all Iowa soy crushers
    python iowa_capacity_collector.py --download-only    # Just download PDFs
    python iowa_capacity_collector.py --parse-only       # Parse already-downloaded PDFs
    python iowa_capacity_collector.py --facility "Eagle Grove"  # Single facility
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print('ERROR: pdfplumber required. pip install pdfplumber')
    sys.exit(1)

try:
    import requests
except ImportError:
    print('ERROR: requests required. pip install requests')
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print('ERROR: openpyxl required. pip install openpyxl')
    sys.exit(1)


# =============================================================================
# IOWA SOYBEAN/OILSEED FACILITY PERMIT MAPPING
# =============================================================================

# Facility number -> media ID mapping from Iowa DNR Draft & Final Permits page
# URL: https://www.iowadnr.gov/environmental-protection/air-quality/operating-permits/draft-final-permits
IOWA_SOY_PERMITS = {
    # AG Processing facilities
    'agp_algona': {
        'name': 'Ag Processing Inc. - Algona',
        'facility_number': '55-01-032',
        'city': 'Algona',
        'permit_number': '15-TV-008R2',
        'media_id': 2707,
    },
    'agp_eagle_grove': {
        'name': 'Ag Processing, Inc. - Eagle Grove',
        'facility_number': '99-01-001',
        'city': 'Eagle Grove',
        'permit_number': '05-TV-005R3',
        'media_id': 2826,
    },
    'agp_emmetsburg': {
        'name': 'Ag Processing, Inc. - Emmetsburg',
        'facility_number': '74-01-012',
        'city': 'Emmetsburg',
        'permit_number': '04-TV-013R2',
        'media_id': 2761,
    },
    'agp_manning': {
        'name': 'Ag Processing, Inc. - Manning',
        'facility_number': '14-02-003',
        'city': 'Manning',
        'permit_number': '11-TV-004R2',
        'media_id': 2808,
    },
    'agp_mason_city': {
        'name': 'Ag Processing, Inc. - Mason City',
        'facility_number': '17-01-027',
        'city': 'Mason City',
        'permit_number': '12-TV-003R2',
        'media_id': 2776,
    },
    'agp_sergeant_bluff': {
        'name': 'Ag Processing, Inc. - Sergeant Bluff',
        'facility_number': '97-04-005',
        'city': 'Sergeant Bluff',
        'permit_number': '99-TV-004R3',
        'media_id': 2909,
    },
    'agp_sheldon': {
        'name': 'Ag Processing, Inc. - Sheldon',
        'facility_number': '71-01-001',
        'city': 'Sheldon',
        'permit_number': '12-TV-001R2',
        'media_id': 2818,
    },
    # ADM
    'adm_des_moines': {
        'name': 'Archer Daniels Midland - Des Moines',
        'facility_number': '77-01-045',
        'city': 'Des Moines',
        'permit_number': '04-TV-020R1',
        'media_id': 2879,
    },
    # Bunge
    'bunge_council_bluffs': {
        'name': 'Bunge North America, Inc. - Council Bluffs',
        'facility_number': '78-01-085',
        'city': 'Council Bluffs',
        'permit_number': '02-TV-017R3-M002',
        'media_id': 2736,
    },
    # Cargill
    'cargill_cedar_rapids_57004': {
        'name': 'Cargill, Inc. - Cedar Rapids (57-01-004)',
        'facility_number': '57-01-004',
        'city': 'Cedar Rapids',
        'permit_number': '07-TV-006R3',
        'media_id': 2925,
    },
    'cargill_cedar_rapids_east': {
        'name': 'Cargill, Inc. - Soybean East Plant',
        'facility_number': '57-01-003',
        'city': 'Cedar Rapids',
        'permit_number': '99-TV-044R4',
        'media_id': 2777,
    },
    'cargill_cedar_rapids_west': {
        'name': 'Cargill Cedar Rapids West',
        'facility_number': '57-01-002',
        'city': 'Cedar Rapids',
        'permit_number': '07-TV-010R3',
        'media_id': 7165,
    },
    'cargill_eddyville': {
        'name': 'Cargill, Inc. - Eddyville',
        'facility_number': '68-09-001',
        'city': 'Eddyville',
        'permit_number': '06-TV-006R1',
        'media_id': 2795,
    },
    'cargill_fort_dodge': {
        'name': 'Cargill, Inc. - Fort Dodge',
        'facility_number': '94-01-080',
        'city': 'Fort Dodge',
        'permit_number': '17-TV-003R1',
        'media_id': 2881,
    },
    'cargill_iowa_falls': {
        'name': 'Cargill, Inc. - Iowa Falls',
        'facility_number': '42-01-003',
        'city': 'Iowa Falls',
        'permit_number': '99-TV-050R5',
        'media_id': 2708,
    },
    'cargill_sioux_city': {
        'name': 'Cargill, Inc. - Sioux City',
        'facility_number': '97-01-001',
        'city': 'Sioux City',
        'permit_number': '99-TV-013R5',
        'media_id': 2694,
    },
    'cargill_vitamin_e': {
        'name': 'Cargill - Vitamin E - Eddyville',
        'facility_number': '68-09-005',
        'city': 'Eddyville',
        'permit_number': '04-TV-004R3',
        'media_id': 2723,
    },
}

# Equipment keywords that identify soybean processing equipment
CRUSH_KEYWORDS = [
    'extraction', 'extractor', 'solvent extraction', 'oil extraction',
    'soy oil extraction', 'vegetable oil processing',
    'cracking', 'dehulling', 'dehull', 'flaker', 'flaking',
    'conditioning', 'bean conditioning', 'soybean conditioning',
    'prep aspiration', 'hull grinding', 'hull',
    'meal dryer', 'meal drying', 'meal grinding', 'meal cooling',
    'spent flake', 'pellet cooler',
    'grain receiving', 'bean receiving', 'truck unload', 'rail unload',
    'grain cleaning', 'grain cleaner',
    'meal loadout', 'meal storage',
]

REFINERY_KEYWORDS = [
    'oil refin', 'refinery', 'refining',
    'degum', 'bleach', 'deodor',
    'clay handling', 'refinery clay',
    'lecithin', 'gum drying',
    'caustic', 'neutrali',
    'hydrogenat',
    'winteriz',
]

BIODIESEL_KEYWORDS = [
    'biodiesel', 'transesterif', 'methanol',
    'ester wash', 'ester dry', 'glycerin',
    'renewable diesel', 'saf ',
]


# =============================================================================
# PDF CAPACITY PARSER
# =============================================================================

class TitleVCapacityParser:
    """
    Parses Iowa DNR Title V operating permits to extract
    equipment-level capacity data for soybean/oilseed processing.
    """

    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.facility_name = ''
        self.permit_number = ''
        self.facility_description = ''
        self.emission_units = []
        self.capacity_summary = {}

    def parse(self):
        """Parse the PDF and extract capacity data."""
        pdf = pdfplumber.open(str(self.pdf_path))
        all_text = []
        for page in pdf.pages:
            text = page.extract_text() or ''
            all_text.append(text)

        full_text = '\n'.join(all_text)

        # Extract facility metadata
        self._extract_facility_info(full_text)

        # Extract emission unit details with capacities
        self._extract_emission_units(all_text)

        # Classify and summarize
        self._classify_equipment()

        pdf.close()
        return self.capacity_summary

    def _extract_facility_info(self, text):
        """Extract facility name, permit number, and description."""
        # Facility name
        m = re.search(r'Facility Name:\s*(.+)', text)
        if m:
            self.facility_name = m.group(1).strip()

        # Permit number
        m = re.search(r'Permit Number:\s*(\S+)', text)
        if m:
            self.permit_number = m.group(1).strip()

        # Facility description (SIC codes)
        m = re.search(r'Facility Description:\s*(.+?)(?:\n|$)', text)
        if m:
            self.facility_description = m.group(1).strip()

    def _extract_emission_units(self, page_texts):
        """
        Extract emission unit descriptions and rated capacities from
        the Emission Point-Specific Conditions sections.

        Iowa DNR Title V permits have a consistent format:
        - Each emission point starts with "Emission Point ID Number: EP-XX"
        - Equipment table has: EU ID, Description, Control Equipment,
          Raw Material, Rated Capacity, Construction Permit
        - Operating requirements often contain throughput limits
        """
        for page_idx, text in enumerate(page_texts):
            lines = text.split('\n')
            in_equipment_section = False

            for i, line in enumerate(lines):
                line_stripped = line.strip()
                line_lower = line_stripped.lower()

                # Detect emission point sections
                if 'emission point id number' in line_lower:
                    ep_match = re.search(r'EP-(\S+)', line_stripped)
                    current_ep = ep_match.group(1) if ep_match else ''
                    in_equipment_section = True
                    continue

                # Look for capacity values in any line
                # Note: Iowa DNR permits use abbreviations: bu/hr, bu./hr., tons/hr
                capacity_match = re.search(
                    r'(\d[\d,]*\.?\d*)\s*(tons?/h(?:ou)?r|(?:bushels?|bu\.?)/h(?:ou)?r|'
                    r'(?:bushels?|bu\.?)/day|(?:bushels?|bu\.?)\s*/\s*hr|'
                    r'tons?/day|tons?/yr|tons?/year|MMgal(?:lons)?/year|Kgal/hour|gallons?/year|'
                    r'MMBtu/hr|gal/ton|lb\.?\s*soybeans?/hr)',
                    line_stripped, re.IGNORECASE
                )

                if capacity_match:
                    value_str = capacity_match.group(1).replace(',', '')
                    try:
                        value = float(value_str)
                    except ValueError:
                        continue
                    unit = capacity_match.group(2)

                    # Get context: look at surrounding lines for equipment description
                    context_lines = []
                    start = max(0, i - 5)
                    end = min(len(lines), i + 3)
                    for j in range(start, end):
                        context_lines.append(lines[j].strip())
                    context = ' '.join(context_lines)

                    # Try to find EU number
                    eu_match = re.search(r'(?:EU-?|EU\s)(\d+[\.\d]*)', context)
                    eu_id = eu_match.group(1) if eu_match else ''

                    # Try to get description from context
                    desc = self._extract_description(context, line_stripped)

                    self.emission_units.append({
                        'page': page_idx + 1,
                        'eu_id': eu_id,
                        'description': desc,
                        'capacity_value': value,
                        'capacity_unit': unit,
                        'raw_line': line_stripped,
                        'context': context[:300],
                    })

                # Also look for throughput limits in operating requirements
                throughput_match = re.search(
                    r'(?:not exceed|maximum|shall not exceed|limit)\D{0,30}'
                    r'(\d[\d,]*\.?\d*)\s*(bushels?\s*per\s*day|tons?\s*per\s*(?:hour|day|year)|'
                    r'gallons?\s*(?:per|of)\s*(?:\w+\s*){0,3}(?:per|/)?\s*'
                    r'(?:twelve.month|12.month|rolling|year|annual))',
                    line_stripped, re.IGNORECASE
                )

                if throughput_match:
                    value_str = throughput_match.group(1).replace(',', '')
                    try:
                        value = float(value_str)
                    except ValueError:
                        continue
                    unit_text = throughput_match.group(2)

                    context_lines = []
                    start = max(0, i - 3)
                    end = min(len(lines), i + 2)
                    for j in range(start, end):
                        context_lines.append(lines[j].strip())
                    context = ' '.join(context_lines)

                    self.emission_units.append({
                        'page': page_idx + 1,
                        'eu_id': '',
                        'description': 'Operating Limit',
                        'capacity_value': value,
                        'capacity_unit': unit_text.strip(),
                        'raw_line': line_stripped,
                        'context': context[:300],
                    })

    def _extract_description(self, context, current_line):
        """Try to extract equipment description from context."""
        # Common patterns in Iowa DNR permits
        desc_patterns = [
            r'(?:EU-?\d+[\.\d]*)\s+(.+?)(?:\d+[\d,]*\.?\d*\s*(?:tons|bushels|gal))',
            r'(?:Emission\s+Unit\s+Description)\s*(.+?)(?:Raw\s+Material|Control)',
        ]
        for pat in desc_patterns:
            m = re.search(pat, context, re.IGNORECASE)
            if m:
                desc = m.group(1).strip()
                desc = re.sub(r'\s+', ' ', desc)
                if len(desc) > 5 and len(desc) < 100:
                    return desc

        # Fallback: look for process-related keywords near the capacity
        keywords_found = []
        context_lower = context.lower()
        for kw in CRUSH_KEYWORDS + REFINERY_KEYWORDS + BIODIESEL_KEYWORDS:
            if kw in context_lower:
                keywords_found.append(kw)
        if keywords_found:
            return '; '.join(keywords_found[:3])

        return ''

    def _classify_equipment(self):
        """Classify emission units and build capacity summary."""
        crush_units = []
        receiving_units = []  # Grain receiving/handling — NOT crush capacity
        refinery_units = []
        biodiesel_units = []
        boiler_units = []
        other_units = []

        # Keywords that indicate handling/storage (NOT processing rate)
        receiving_keywords = [
            'grain receiving', 'grain handling', 'grain unload',
            'bean receiving', 'bean rail unload', 'truck unload',
            'truck dump', 'rail unload', 'bean truck dump',
            'grain transfer', 'grain convey',
            'meal storage', 'meal loadout', 'meal transfer',
            'meal sifter', 'meal screening',
            'hull grinding', 'hull storage', 'hull loadout',
            'pellet transfer', 'pellet cooler',
            'product loadout', 'haul road',
            'grain clean', 'grain dryer',
            'rail loadout', 'truck loadout', 'rail meal',
            'rail bean', 'truck bean', 'conveyance',
            'truck soybean', 'rail soybean',
            'soybean storage', 'bean storage',
        ]

        # Keywords that indicate actual soybean PROCESSING (the real crush rate)
        processing_keywords = [
            'extraction', 'extractor', 'solvent extraction',
            'oil extraction', 'vegetable oil processing',
            'flak', 'conditioning', 'bean conditioning',
            'soybean conditioning', 'cracking', 'dehull',
            'prep aspiration', 'spent flake',
        ]

        for eu in self.emission_units:
            ctx = (eu['description'] + ' ' + eu['context']).lower()

            if any(kw in ctx for kw in REFINERY_KEYWORDS):
                refinery_units.append(eu)
            elif any(kw in ctx for kw in BIODIESEL_KEYWORDS):
                biodiesel_units.append(eu)
            elif any(kw in ctx for kw in ['boiler', 'mmbtu']):
                boiler_units.append(eu)
            elif any(kw in ctx for kw in receiving_keywords):
                receiving_units.append(eu)
            elif any(kw in ctx for kw in CRUSH_KEYWORDS):
                crush_units.append(eu)
            else:
                other_units.append(eu)

        # Find the primary crush capacity
        # Priority: extraction/flaking rates > conditioning > general processing > receiving
        crush_capacity_tph = 0
        crush_capacity_bpd = 0
        crush_capacity_bph = 0
        crush_description = ''

        # First pass: look for extraction/flaking/conditioning (actual process rate)
        for eu in crush_units:
            ctx = (eu['description'] + ' ' + eu['context']).lower()
            is_processing = any(kw in ctx for kw in processing_keywords)
            unit = eu['capacity_unit'].lower()
            val = eu['capacity_value']

            if is_processing:
                if 'tons/h' in unit or 'tons/hour' in unit:
                    if val > crush_capacity_tph:
                        crush_capacity_tph = val
                        crush_description = eu['description'] or eu['raw_line']
                elif 'bu' in unit and '/h' in unit:
                    if val > crush_capacity_bph:
                        crush_capacity_bph = val
                        if not crush_description:
                            crush_description = eu['description'] or eu['raw_line']
                elif 'bu' in unit and '/day' in unit or 'bushels per day' in unit:
                    if val > crush_capacity_bpd:
                        crush_capacity_bpd = val
                elif 'lb' in unit and 'soybean' in unit:
                    # Convert lb soybeans/hr to tons/hr
                    tph = val / 2000.0
                    if tph > crush_capacity_tph:
                        crush_capacity_tph = tph
                        crush_description = eu['description'] or eu['raw_line']

        # If no processing-specific capacity found, use general crush units
        if not crush_capacity_tph and not crush_capacity_bph:
            for eu in crush_units:
                unit = eu['capacity_unit'].lower()
                val = eu['capacity_value']
                if 'tons/h' in unit or 'tons/hour' in unit:
                    if val > crush_capacity_tph:
                        crush_capacity_tph = val
                        crush_description = eu['description'] or eu['raw_line']
                elif 'bu' in unit and '/h' in unit:
                    if val > crush_capacity_bph:
                        crush_capacity_bph = val
                elif 'bu' in unit and ('/day' in unit or 'per day' in unit):
                    if val > crush_capacity_bpd:
                        crush_capacity_bpd = val

        # Convert bu/hr to tons/hr if we only have bu/hr
        # 1 bushel soybeans = 60 lbs = 0.03 short tons = 0.027216 MT
        if crush_capacity_bph and not crush_capacity_tph:
            crush_capacity_tph = crush_capacity_bph * 60.0 / 2000.0  # short tons

        # Convert bu/hr to bu/day if we only have hourly
        if crush_capacity_bph and not crush_capacity_bpd:
            crush_capacity_bpd = crush_capacity_bph * 24.0

        # Find refinery capacity
        refinery_capacity_tph = 0
        refinery_capacity_tpy = 0
        refinery_description = ''
        has_refinery = len(refinery_units) > 0
        for eu in refinery_units:
            unit = eu['capacity_unit'].lower()
            val = eu['capacity_value']
            if 'tons/h' in unit or 'tons/hour' in unit:
                if val > refinery_capacity_tph:
                    refinery_capacity_tph = val
                    refinery_description = eu['description'] or eu['raw_line']
            elif 'tons/yr' in unit or 'tons/year' in unit or 'per year' in unit:
                if val > refinery_capacity_tpy:
                    refinery_capacity_tpy = val

        # Determine refining type from keywords
        refining_type = ''
        if has_refinery:
            all_refinery_context = ' '.join(
                eu['context'].lower() for eu in refinery_units
            )
            if 'deodor' in all_refinery_context:
                refining_type = 'RBD'  # Refined, Bleached, Deodorized
            elif 'bleach' in all_refinery_context or 'clay' in all_refinery_context:
                refining_type = 'RB'  # Refined, Bleached
            elif 'degum' in all_refinery_context:
                refining_type = 'Degummed'
            elif 'caustic' in all_refinery_context or 'neutrali' in all_refinery_context:
                refining_type = 'Caustic Refined'
            else:
                refining_type = 'Unknown'

        # Biodiesel capacity
        biodiesel_capacity_mgy = 0
        for eu in biodiesel_units:
            unit = eu['capacity_unit'].lower()
            val = eu['capacity_value']
            if 'mmgal' in unit:
                if val > biodiesel_capacity_mgy:
                    biodiesel_capacity_mgy = val
            elif 'gallons/year' in unit or 'gallons per' in unit:
                mgy = val / 1e6
                if mgy > biodiesel_capacity_mgy:
                    biodiesel_capacity_mgy = mgy

        self.capacity_summary = {
            'facility_name': self.facility_name,
            'permit_number': self.permit_number,
            'facility_description': self.facility_description,
            'crush_capacity_tons_per_hour': crush_capacity_tph,
            'crush_capacity_bushels_per_day': crush_capacity_bpd,
            'crush_description': crush_description,
            'has_refinery': has_refinery,
            'refinery_capacity_tons_per_hour': refinery_capacity_tph,
            'refinery_capacity_tons_per_year': refinery_capacity_tpy,
            'refinery_description': refinery_description,
            'refining_type': refining_type,
            'has_biodiesel': len(biodiesel_units) > 0,
            'biodiesel_capacity_mgy': biodiesel_capacity_mgy,
            'total_emission_units_found': len(self.emission_units),
            'crush_units': len(crush_units),
            'refinery_units': len(refinery_units),
            'biodiesel_units': len(biodiesel_units),
            'all_emission_units': self.emission_units,
        }


# =============================================================================
# DOWNLOADER
# =============================================================================

class IowaPermitDownloader:
    """Downloads Title V permit PDFs from Iowa DNR."""

    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        })

    def download_permit(self, key, permit_info):
        """Download a single Title V permit PDF."""
        media_id = permit_info['media_id']
        filename = '{}_titlev.pdf'.format(key)
        filepath = self.output_dir / filename

        if filepath.exists() and filepath.stat().st_size > 10000:
            print('  [SKIP] {} already downloaded ({} KB)'.format(
                filename, filepath.stat().st_size // 1024))
            return str(filepath)

        url = 'https://www.iowadnr.gov/media/{}/download'.format(media_id)
        print('  [DOWNLOAD] {} (media {})...'.format(
            permit_info['name'], media_id), end=' ')

        try:
            r = self.session.get(url, timeout=60)
            if r.status_code == 200 and len(r.content) > 10000:
                with open(str(filepath), 'wb') as f:
                    f.write(r.content)
                print('OK - {} KB'.format(len(r.content) // 1024))
                time.sleep(1.5)  # Be polite
                return str(filepath)
            else:
                print('FAILED (status={}, size={})'.format(
                    r.status_code, len(r.content)))
                return None
        except Exception as e:
            print('ERROR: {}'.format(str(e)[:80]))
            return None

    def download_all(self, permits=None):
        """Download all permits. Returns dict of key -> filepath."""
        permits = permits or IOWA_SOY_PERMITS
        results = {}
        print('\nDownloading {} Title V permits from Iowa DNR...'.format(
            len(permits)))
        for key, info in permits.items():
            filepath = self.download_permit(key, info)
            if filepath:
                results[key] = filepath
        return results


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def create_capacity_excel(results, output_path):
    """Create Excel workbook with capacity data."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = 'Capacity Summary'

    headers = [
        'Facility', 'City', 'Facility #', 'Permit #',
        'Crush (tons/hr)', 'Crush (bu/day)',
        'Has Refinery', 'Refinery (tons/hr)', 'Refinery (tons/yr)',
        'Refining Type',
        'Has Biodiesel', 'Biodiesel (MGY)',
        'Facility Description',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

    row = 2
    for key, data in sorted(results.items()):
        permit_info = IOWA_SOY_PERMITS.get(key, {})
        summary = data.get('summary', {})

        values = [
            summary.get('facility_name') or permit_info.get('name', ''),
            permit_info.get('city', ''),
            permit_info.get('facility_number', ''),
            summary.get('permit_number') or permit_info.get('permit_number', ''),
            summary.get('crush_capacity_tons_per_hour', 0) or '',
            summary.get('crush_capacity_bushels_per_day', 0) or '',
            'Yes' if summary.get('has_refinery') else 'No',
            summary.get('refinery_capacity_tons_per_hour', 0) or '',
            summary.get('refinery_capacity_tons_per_year', 0) or '',
            summary.get('refining_type', ''),
            'Yes' if summary.get('has_biodiesel') else 'No',
            summary.get('biodiesel_capacity_mgy', 0) or '',
            summary.get('facility_description', ''),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col in (5, 6, 8, 9, 12) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

        # Highlight refineries
        if summary.get('has_refinery'):
            for col in range(7, 11):
                ws.cell(row=row, column=col).fill = PatternFill(
                    'solid', fgColor='E2EFDA')

        row += 1

    # Column widths
    widths = [35, 18, 12, 22, 14, 14, 12, 14, 14, 14, 12, 14, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # --- Sheet 2: Equipment Detail ---
    ws2 = wb.create_sheet('Equipment Detail')
    detail_headers = [
        'Facility', 'Page', 'EU ID', 'Description', 'Category',
        'Capacity Value', 'Capacity Unit', 'Raw Line',
    ]
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for key, data in sorted(results.items()):
        permit_info = IOWA_SOY_PERMITS.get(key, {})
        summary = data.get('summary', {})
        fac_name = summary.get('facility_name') or permit_info.get('name', '')

        for eu in summary.get('all_emission_units', []):
            ctx = (eu.get('description', '') + ' ' + eu.get('context', '')).lower()
            if any(kw in ctx for kw in REFINERY_KEYWORDS):
                category = 'Refinery'
            elif any(kw in ctx for kw in BIODIESEL_KEYWORDS):
                category = 'Biodiesel'
            elif any(kw in ctx for kw in ['boiler', 'mmbtu']):
                category = 'Boiler'
            elif any(kw in ctx for kw in CRUSH_KEYWORDS):
                category = 'Crush/Processing'
            else:
                category = 'Other'

            values = [
                fac_name,
                eu.get('page', ''),
                eu.get('eu_id', ''),
                eu.get('description', ''),
                category,
                eu.get('capacity_value', ''),
                eu.get('capacity_unit', ''),
                eu.get('raw_line', ''),
            ]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.border = thin_border
            row += 1

    # Column widths
    detail_widths = [35, 8, 10, 40, 16, 14, 18, 60]
    for col, w in enumerate(detail_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(str(output_path))
    print('\nExcel saved: {}'.format(output_path))


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Iowa DNR Title V Permit Capacity Extractor'
    )
    parser.add_argument('--download-only', action='store_true',
                        help='Only download PDFs, do not parse')
    parser.add_argument('--parse-only', action='store_true',
                        help='Parse already-downloaded PDFs')
    parser.add_argument('--facility', '-f', default=None,
                        help='Filter to a single facility (name substring match)')
    parser.add_argument('--output', '-o', default=None,
                        help='Output Excel path')
    args = parser.parse_args()

    project_root = Path(__file__).parent.parent.parent
    raw_dir = Path(__file__).parent / 'raw'
    output_dir = Path(__file__).parent / 'output'
    output_dir.mkdir(exist_ok=True)

    # Filter permits if requested
    permits = IOWA_SOY_PERMITS
    if args.facility:
        search = args.facility.lower()
        permits = {
            k: v for k, v in permits.items()
            if search in k.lower() or search in v['name'].lower()
            or search in v['city'].lower()
        }
        if not permits:
            print('No facilities matching "{}"'.format(args.facility))
            return

    print('Iowa DNR Title V Permit Capacity Extractor')
    print('==========================================')
    print('Facilities: {}'.format(len(permits)))
    print()

    # Step 1: Download PDFs
    downloaded = {}
    if not args.parse_only:
        downloader = IowaPermitDownloader(str(raw_dir))
        downloaded = downloader.download_all(permits)
    else:
        # Find already-downloaded files
        for key in permits:
            filepath = raw_dir / '{}_titlev.pdf'.format(key)
            if filepath.exists():
                downloaded[key] = str(filepath)
        print('Found {} already-downloaded PDFs'.format(len(downloaded)))

    if args.download_only:
        print('\nDownload complete. {} PDFs available.'.format(len(downloaded)))
        return

    # Step 2: Parse PDFs
    print('\nParsing {} Title V permits for capacity data...\n'.format(
        len(downloaded)))

    results = {}
    for key, filepath in sorted(downloaded.items()):
        permit_info = permits.get(key, {})
        print('Parsing: {} ({})'.format(
            permit_info.get('name', key),
            permit_info.get('city', '')))

        try:
            parser_obj = TitleVCapacityParser(filepath)
            summary = parser_obj.parse()
            results[key] = {
                'filepath': filepath,
                'summary': summary,
            }

            # Print quick summary
            crush = summary.get('crush_capacity_tons_per_hour', 0)
            bpd = summary.get('crush_capacity_bushels_per_day', 0)
            ref = summary.get('has_refinery', False)
            ref_tph = summary.get('refinery_capacity_tons_per_hour', 0)
            ref_type = summary.get('refining_type', '')
            bio = summary.get('has_biodiesel', False)
            bio_mgy = summary.get('biodiesel_capacity_mgy', 0)

            parts = []
            if crush:
                parts.append('Crush: {:.1f} tph'.format(crush))
            if bpd:
                parts.append('{:,.0f} bu/day'.format(bpd))
            if ref:
                parts.append('Refinery: {:.1f} tph ({})'.format(
                    ref_tph, ref_type))
            if bio:
                parts.append('Biodiesel: {:.0f} MGY'.format(bio_mgy))
            if not parts:
                parts.append('No capacity data extracted')

            print('  -> {}'.format(' | '.join(parts)))
            print()

        except Exception as e:
            print('  -> ERROR: {}'.format(str(e)[:100]))
            results[key] = {'filepath': filepath, 'summary': {}}
            print()

    # Step 3: Output Excel
    if results:
        timestamp = datetime.now().strftime('%Y-%m-%d')
        output_path = args.output or str(
            output_dir / 'iowa_soy_capacity_{}.xlsx'.format(timestamp))
        create_capacity_excel(results, output_path)

        # Also save JSON
        json_path = output_path.replace('.xlsx', '.json')
        json_results = {}
        for key, data in results.items():
            summary = data.get('summary', {})
            # Remove the full emission_units list for the JSON (too verbose)
            clean = {k: v for k, v in summary.items()
                     if k != 'all_emission_units'}
            json_results[key] = clean
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_results, f, indent=2, default=str)
        print('JSON saved: {}'.format(json_path))

    # Print summary table
    print('\n' + '=' * 90)
    print('IOWA SOYBEAN/OILSEED PROCESSING CAPACITY SUMMARY')
    print('=' * 90)
    print('{:<35s} {:>10s} {:>10s} {:>8s} {:>10s} {:>8s}'.format(
        'Facility', 'Crush tph', 'Crush bpd', 'Refine?', 'Ref tph', 'Type'))
    print('-' * 90)

    for key, data in sorted(results.items()):
        s = data.get('summary', {})
        permit_info = permits.get(key, {})
        name = (s.get('facility_name') or permit_info.get('name', ''))[:34]
        crush = s.get('crush_capacity_tons_per_hour', 0)
        bpd = s.get('crush_capacity_bushels_per_day', 0)
        has_ref = 'YES' if s.get('has_refinery') else 'No'
        ref_tph = s.get('refinery_capacity_tons_per_hour', 0)
        ref_type = s.get('refining_type', '')[:8]

        print('{:<35s} {:>10s} {:>10s} {:>8s} {:>10s} {:>8s}'.format(
            name,
            '{:.1f}'.format(crush) if crush else '-',
            '{:,.0f}'.format(bpd) if bpd else '-',
            has_ref,
            '{:.1f}'.format(ref_tph) if ref_tph else '-',
            ref_type,
        ))

    print('=' * 90)


if __name__ == '__main__':
    main()
