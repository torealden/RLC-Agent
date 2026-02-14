"""
Nebraska NDEE Title V Permit Capacity Collector

Downloads permit metadata from the Nebraska Department of Environment and
Energy (NDEE) and extracts equipment-level capacity data from Title V
operating permit PDFs for soybean/oilseed processing facilities.

Data Sources:
  - NDEE Permit Search JSON API: permit metadata, facility IDs, permit history
  - Title V Operating Permit PDFs: equipment descriptions, rated capacities

NDEE System Notes:
  - The NDEE web portal (deq-iis.ne.gov) is metadata-only; permit PDFs are
    NOT available for download from the portal.
  - PDFs must be obtained via public records request to NDEE Air Quality
    Division (402-471-2186) or dee.nebraska.gov/about/about-agency/request-public-records
  - Once PDFs are placed in the raw/ directory, the parser extracts capacity data.
  - NDEE PRGID1 field (e.g., "053 00018") maps to ECHO source_id
    (e.g., NE0000003105300018) via county FIPS + facility number.

Usage:
    python nebraska_capacity_collector.py                     # Full pipeline
    python nebraska_capacity_collector.py --download-only     # Just download metadata + PDFs
    python nebraska_capacity_collector.py --parse-only        # Parse existing PDFs
    python nebraska_capacity_collector.py --metadata-only     # Just download NDEE metadata
    python nebraska_capacity_collector.py --facility "ADM"    # Single facility
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print('ERROR: pdfplumber required. pip install pdfplumber')
    sys.exit(1)

try:
    import requests
except ImportError:
    print('ERROR: requests required. pip install requests')
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print('ERROR: openpyxl required. pip install openpyxl')
    sys.exit(1)


# =============================================================================
# NEBRASKA SOYBEAN/OILSEED FACILITY MAPPING
# =============================================================================

# Nebraska target facilities from EPA ECHO database (bronze.epa_echo_facility)
# The 'prgid1' field is derived from the ECHO source_id:
#   source_id NE0000003105300018 -> PRGID1 "053 00018"
#   source_id NELLC0003110900011 -> PRGID1 "109 00011"
NE_SOY_PERMITS = {
    # --- Operating Facilities (Priority) ---
    'adm_lincoln_extraction': {
        'name': 'ADM Soybean Oil Extraction Plant',
        'city': 'Lincoln',
        'county': 'Lancaster',
        'frs_registry_id': '110002382680',
        'source_id': 'NELLC0003110900011',
        'prgid1': '109 00011',
        'sic_codes': '2075, 2079',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'adm_milling_lincoln': {
        'name': 'ADM Milling',
        'city': 'Lincoln',
        'county': 'Lancaster',
        'frs_registry_id': '110015682260',
        'source_id': 'NELLC0003110900003',
        'prgid1': '109 00003',
        'sic_codes': '2041, 5153, 2047, 2048, 2075',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'adm_fremont': {
        'name': 'Archer Daniels Midland Co',
        'city': 'Fremont',
        'county': 'Dodge',
        'frs_registry_id': '110000447437',
        'source_id': 'NE0000003105300018',
        'prgid1': '053 00018',
        'sic_codes': '2075',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'bruning_grain': {
        'name': 'Bruning Grain & Feed Co',
        'city': 'Bruning',
        'county': 'Thayer',
        'frs_registry_id': '110000724306',
        'source_id': 'NE0000003116900005',
        'prgid1': '169 00005',
        'sic_codes': '5153, 4221, 2075',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'eco_energy_beatrice': {
        'name': 'Eco-Energy Distrb-Beatrice LLC',
        'city': 'Beatrice',
        'county': 'Gage',
        'frs_registry_id': '110045416974',
        'source_id': 'NE0000003106700085',
        'prgid1': '067 00085',
        'sic_codes': '2075, 2869',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'frontier_cooperative': {
        'name': 'Frontier Cooperative',
        'city': 'Columbus',
        'county': 'Platte',
        'frs_registry_id': '110041346843',
        'source_id': 'NE0000003114100103',
        'prgid1': '141 00103',
        'sic_codes': '2075',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'ingredion_ssc': {
        'name': 'Ingredion Incorporated',
        'city': 'South Sioux City',
        'county': 'Dakota',
        'frs_registry_id': '110040498235',
        'source_id': 'NE0000003104300032',
        'prgid1': '043 00032',
        'sic_codes': '2075, 2076',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'kansas_organic': {
        'name': 'Kansas Organic Producers Assn',
        'city': 'Du Bois',
        'county': 'Pawnee',
        'frs_registry_id': '110007131004',
        'source_id': 'NE0000003113300018',
        'prgid1': '133 00018',
        'sic_codes': '2075',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'ne_soybean_processing': {
        'name': 'Nebraska Soybean Processing',
        'city': 'Scribner',
        'county': 'Dodge',
        'frs_registry_id': '110002441947',
        'source_id': 'NE0000003105300070',
        'prgid1': '053 00070',
        'sic_codes': '2075, 4221',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'richardson_milling': {
        'name': 'Richardson Milling Inc',
        'city': 'South Sioux City',
        'county': 'Dakota',
        'frs_registry_id': '110000497686',
        'source_id': 'NE0000003104300020',
        'prgid1': '043 00020',
        'sic_codes': '2041, 2043, 2076, 5141, 4221, 2099',
        'status': 'Operating',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    # --- Planned Facilities ---
    'agp_david_city': {
        'name': 'AG Processing Inc',
        'city': 'David City',
        'county': 'Butler',
        'frs_registry_id': '110071344603',
        'source_id': 'NE0000003102300054',
        'prgid1': '023 00054',
        'sic_codes': '2075',
        'status': 'Planned',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
    'norfolk_crush': {
        'name': 'Norfolk Crush LLC',
        'city': 'Norfolk',
        'county': 'Madison',
        'frs_registry_id': '110071274379',
        'source_id': 'NE0000003111900119',
        'prgid1': '119 00119',
        'sic_codes': '2075',
        'status': 'Planned',
        'permit_number': '',
        'modsq': '',
        'ndee_appno': '',
    },
}


# Equipment keywords that identify soybean processing equipment
CRUSH_KEYWORDS = [
    'extraction', 'extractor', 'solvent extraction', 'oil extraction',
    'soy oil extraction', 'vegetable oil processing',
    'cracking', 'dehulling', 'dehull', 'flaker', 'flaking',
    'conditioning', 'bean conditioning', 'soybean conditioning',
    'prep aspiration', 'hull grinding', 'hull',
    'meal dryer', 'meal drying', 'meal grinding', 'meal cooling',
    'spent flake', 'pellet cooler',
    'grain receiving', 'bean receiving', 'truck unload', 'rail unload',
    'grain cleaning', 'grain cleaner',
    'meal loadout', 'meal storage',
]

REFINERY_KEYWORDS = [
    'oil refin', 'refinery', 'refining',
    'degum', 'bleach', 'deodor',
    'clay handling', 'refinery clay',
    'lecithin', 'gum drying',
    'caustic', 'neutrali',
    'hydrogenat',
    'winteriz',
]

BIODIESEL_KEYWORDS = [
    'biodiesel', 'transesterif', 'methanol',
    'ester wash', 'ester dry', 'glycerin',
    'renewable diesel', 'saf ',
]


# =============================================================================
# NDEE API CLIENT
# =============================================================================

class NDEEPermitClient:
    """
    Client for the Nebraska DEE permit search system at deq-iis.ne.gov.

    The NDEE system uses a Dojo-based JavaScript UI. The JSON API has two
    access patterns:
      - POST to result_result.php: returns ALL ~13,000 records (ignores params)
      - GET to result_result.php with query params: returns filtered results
      - GET to result_detail.php?modsq={MODSQ}: returns full permit history

    IMPORTANT: The NDEE portal does NOT serve permit PDF documents. It is
    metadata-only. Actual permit PDFs must be obtained via FOIA/public records
    request to NDEE Air Quality Division.
    """

    BASE_URL = 'https://deq-iis.ne.gov/zs/permit'

    def __init__(self, raw_dir):
        self.raw_dir = Path(raw_dir)
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                          'AppleWebKit/537.36',
        })
        self._all_permits = None

    def download_all_permits(self, force=False):
        """
        Download the full NDEE permit database (~13,000 records).

        The POST endpoint returns a nested list-of-lists grouped by program.
        We flatten it and cache the result.
        """
        cache_path = self.raw_dir / 'ndee_all_permits.json'

        if cache_path.exists() and not force:
            age_hours = (time.time() - cache_path.stat().st_mtime) / 3600
            if age_hours < 24:
                print('  [CACHE] Using cached NDEE permits ({:.1f} hrs old)'.format(
                    age_hours))
                with open(str(cache_path), 'r', encoding='utf-8') as f:
                    self._all_permits = json.load(f)
                return self._all_permits

        print('  [DOWNLOAD] Fetching full NDEE permit database...')
        url = '{}/result_result.php'.format(self.BASE_URL)

        try:
            r = self.session.post(url, data={}, timeout=120)
            r.raise_for_status()
            raw_data = r.json()
        except Exception as e:
            print('  ERROR downloading NDEE permits: {}'.format(e))
            # Try loading from cache as fallback
            if cache_path.exists():
                print('  [FALLBACK] Using stale cache')
                with open(str(cache_path), 'r', encoding='utf-8') as f:
                    self._all_permits = json.load(f)
                return self._all_permits
            return []

        # Flatten nested list-of-lists
        flat = []
        if isinstance(raw_data, list):
            for group in raw_data:
                if isinstance(group, list):
                    flat.extend(group)
                elif isinstance(group, dict):
                    flat.append(group)

        self._all_permits = flat
        print('  [OK] {} total NDEE permit records'.format(len(flat)))

        # Save cache
        with open(str(cache_path), 'w', encoding='utf-8') as f:
            json.dump(flat, f, indent=2)

        return flat

    def get_air_operating_permits(self):
        """Filter for Clean Air Act Operating (Title V) permits."""
        if self._all_permits is None:
            self.download_all_permits()

        return [
            p for p in self._all_permits
            if (p.get('DBNAME', '').strip() == 'Clean Air Act'
                and p.get('PERMIT', '').strip() == 'Operating')
        ]

    def get_all_air_permits(self):
        """Get all Clean Air Act permits (Operating + Construction + Low Emitter)."""
        if self._all_permits is None:
            self.download_all_permits()

        return [
            p for p in self._all_permits
            if p.get('DBNAME', '').strip() == 'Clean Air Act'
        ]

    def get_facility_details(self, modsq):
        """
        Fetch full permit history for a facility from the detail endpoint.

        Returns a flat list of all permit records (across all NDEE programs)
        for the given MODSQ.
        """
        url = '{}/result_detail.php?modsq={}'.format(self.BASE_URL, modsq)

        try:
            r = self.session.get(url, timeout=30)
            r.raise_for_status()
            raw_data = r.json()
        except Exception as e:
            print('  WARNING: Could not fetch details for MODSQ {}: {}'.format(
                modsq, e))
            return []

        # Flatten nested list-of-lists
        flat = []
        if isinstance(raw_data, list):
            for group in raw_data:
                if isinstance(group, list):
                    flat.extend(group)
                elif isinstance(group, dict):
                    flat.append(group)

        time.sleep(0.5)  # Be polite
        return flat

    def match_facilities(self, target_permits):
        """
        Match NDEE permit records to our target facilities.

        Primary match: PRGID1 from NDEE details matches our derived prgid1
        (from ECHO source_id). This is the most reliable method.

        Secondary match: facility NAME + CITY (fuzzy, case-insensitive).

        Returns updated target_permits dict with ndee fields populated.
        """
        air_permits = self.get_all_air_permits()
        print('\n  Matching {} NDEE air permits to {} target facilities...'.format(
            len(air_permits), len(target_permits)))

        # Build lookup by (city, county) for name matching
        by_location = {}
        for p in air_permits:
            key = (p.get('CITY', '').strip().upper(),
                   p.get('COUNTY', '').strip().upper())
            by_location.setdefault(key, []).append(p)

        # Build lookup by MODSQ for deduplication
        modsq_seen = {}
        for p in air_permits:
            modsq = p.get('MODSQ', '').strip()
            if modsq and modsq not in modsq_seen:
                modsq_seen[modsq] = p

        matched = 0
        for key, info in target_permits.items():
            target_prgid = info.get('prgid1', '')

            # Strategy 1: Match by PRGID1 via detail records
            # First, find candidate MODSQs by city/county
            city_upper = info['city'].upper()
            county_upper = info['county'].upper()
            candidates = by_location.get((city_upper, county_upper), [])

            # Also check nearby cities (e.g., "SOUTH SIOUX CITY" vs "S SIOUX CITY")
            for loc_key, loc_permits in by_location.items():
                if county_upper == loc_key[1]:
                    loc_city = loc_key[0]
                    if (city_upper in loc_city or loc_city in city_upper
                            or _city_fuzzy_match(city_upper, loc_city)):
                        candidates.extend(loc_permits)

            # Deduplicate by MODSQ
            seen = set()
            unique_candidates = []
            for p in candidates:
                modsq = p.get('MODSQ', '').strip()
                if modsq and modsq not in seen:
                    seen.add(modsq)
                    unique_candidates.append(p)

            # For each candidate MODSQ, check if its details contain our PRGID1
            best_match = None
            best_appno = ''
            best_modsq = ''

            for cand in unique_candidates:
                modsq = cand.get('MODSQ', '').strip()
                if not modsq:
                    continue

                # Fetch details to check PRGID1
                details = self.get_facility_details(modsq)
                air_details = [
                    d for d in details
                    if d.get('DBNAME', '').strip() == 'Clean Air Act'
                ]

                for d in air_details:
                    d_prgid = d.get('PRGID1', '').strip()
                    if d_prgid == target_prgid:
                        # Found PRGID1 match! Get the latest Operating permit
                        operating = [
                            dd for dd in air_details
                            if dd.get('PERMIT', '').strip() == 'Operating'
                        ]
                        if operating:
                            # Sort by PDDATE descending to get latest
                            operating.sort(
                                key=lambda x: x.get('PDDATE', ''),
                                reverse=True)
                            best_match = operating[0]
                            best_appno = best_match.get('APPNO', '').strip()
                            best_modsq = modsq
                        else:
                            # No Operating permit, use latest Construction
                            construction = [
                                dd for dd in air_details
                                if dd.get('PERMIT', '').strip() == 'Construction'
                            ]
                            if construction:
                                construction.sort(
                                    key=lambda x: x.get('PDDATE', ''),
                                    reverse=True)
                                best_match = construction[0]
                                best_appno = best_match.get('APPNO', '').strip()
                                best_modsq = modsq
                            elif air_details:
                                best_match = air_details[0]
                                best_appno = best_match.get('APPNO', '').strip()
                                best_modsq = modsq
                        break

                if best_match:
                    break

            # Strategy 2: If no PRGID1 match, try name fuzzy match
            if not best_match:
                name_lower = info['name'].lower()
                name_words = set(
                    name_lower.replace(',', '').replace('.', '').split()
                ) - {'inc', 'llc', 'co', 'corp', 'the', '-'}

                for cand in unique_candidates:
                    cand_name = cand.get('NAME', '').strip().lower()
                    cand_words = set(
                        cand_name.replace(',', '').replace('.', '').split()
                    ) - {'inc', 'llc', 'co', 'corp', 'the', '-'}

                    overlap = name_words & cand_words
                    if len(overlap) >= 2 or name_lower in cand_name or cand_name in name_lower:
                        modsq = cand.get('MODSQ', '').strip()
                        best_modsq = modsq

                        details = self.get_facility_details(modsq)
                        air_details = [
                            d for d in details
                            if d.get('DBNAME', '').strip() == 'Clean Air Act'
                        ]
                        operating = [
                            d for d in air_details
                            if d.get('PERMIT', '').strip() == 'Operating'
                        ]
                        if operating:
                            operating.sort(
                                key=lambda x: x.get('PDDATE', ''),
                                reverse=True)
                            best_match = operating[0]
                            best_appno = best_match.get('APPNO', '').strip()
                        elif air_details:
                            air_details.sort(
                                key=lambda x: x.get('PDDATE', ''),
                                reverse=True)
                            best_match = air_details[0]
                            best_appno = best_match.get('APPNO', '').strip()
                        break

            if best_match:
                info['modsq'] = best_modsq
                info['ndee_appno'] = best_appno
                info['permit_number'] = best_appno
                info['ndee_prgid1'] = best_match.get('PRGID1', '').strip()
                info['ndee_permit_type'] = best_match.get('PERMIT', '').strip()
                info['ndee_issued_date'] = best_match.get('PDDATE', '')
                info['ndee_expiry_date'] = best_match.get('PTDATE', '')
                info['ndee_name'] = best_match.get('NAME', '').strip()
                matched += 1
                print('    [MATCH] {} -> MODSQ {} APPNO {} ({})'.format(
                    info['name'], best_modsq, best_appno,
                    info.get('ndee_permit_type', '')))
            else:
                print('    [MISS]  {} ({}, {})'.format(
                    info['name'], info['city'], info['county']))

        print('\n  Matched {} of {} target facilities'.format(
            matched, len(target_permits)))
        return target_permits


def _city_fuzzy_match(city1, city2):
    """Check if two city names are fuzzy matches."""
    # Handle abbreviations: "S SIOUX CITY" vs "SOUTH SIOUX CITY"
    abbrevs = {
        'S ': 'SOUTH ', 'N ': 'NORTH ', 'E ': 'EAST ', 'W ': 'WEST ',
        'FT ': 'FORT ', 'ST ': 'SAINT ', 'MT ': 'MOUNT ',
    }
    for short, long in abbrevs.items():
        if city1.startswith(short):
            if city2.startswith(long) or city2 == city1:
                return True
        if city2.startswith(short):
            if city1.startswith(long) or city1 == city2:
                return True
    return False


# =============================================================================
# PDF CAPACITY PARSER (adapted from Iowa for Nebraska NDEE permits)
# =============================================================================

class TitleVCapacityParser:
    """
    Parses Title V operating permits to extract equipment-level capacity data
    for soybean/oilseed processing.

    Nebraska NDEE permits follow a similar federal Title V format to Iowa DNR,
    but may have state-specific section headers, equipment table layouts, and
    capacity unit conventions.
    """

    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.facility_name = ''
        self.permit_number = ''
        self.facility_description = ''
        self.emission_units = []
        self.capacity_summary = {}

    def parse(self):
        """Parse the PDF and extract capacity data."""
        pdf = pdfplumber.open(str(self.pdf_path))
        all_text = []
        for page in pdf.pages:
            text = page.extract_text() or ''
            all_text.append(text)

        full_text = '\n'.join(all_text)

        # Extract facility metadata
        self._extract_facility_info(full_text)

        # Extract emission unit details with capacities
        self._extract_emission_units(all_text)

        # Classify and summarize
        self._classify_equipment()

        pdf.close()
        return self.capacity_summary

    def _extract_facility_info(self, text):
        """Extract facility name, permit number, and description."""
        # Try multiple patterns for facility name (varies by state)
        for pattern in [
            r'Facility Name:\s*(.+)',
            r'Facility:\s*(.+)',
            r'Source Name:\s*(.+)',
            r'Company Name:\s*(.+)',
        ]:
            m = re.search(pattern, text)
            if m:
                self.facility_name = m.group(1).strip()
                break

        # Permit number
        for pattern in [
            r'Permit Number:\s*(\S+)',
            r'Operating Permit\s*(?:No\.?|Number|#)\s*[:.]?\s*(\S+)',
            r'Permit\s*(?:No\.?|#)\s*[:.]?\s*(\S+)',
        ]:
            m = re.search(pattern, text)
            if m:
                self.permit_number = m.group(1).strip()
                break

        # Facility description (SIC codes)
        for pattern in [
            r'Facility Description:\s*(.+?)(?:\n|$)',
            r'SIC Code[s]?:\s*(.+?)(?:\n|$)',
            r'Standard Industrial Classification:\s*(.+?)(?:\n|$)',
        ]:
            m = re.search(pattern, text)
            if m:
                self.facility_description = m.group(1).strip()
                break

    def _extract_emission_units(self, page_texts):
        """
        Extract emission unit descriptions and rated capacities.

        Looks for capacity values with unit patterns in the permit text,
        along with surrounding context for equipment classification.
        Nebraska NDEE permits may use slightly different section headers
        but the core capacity format is similar to other states.
        """
        for page_idx, text in enumerate(page_texts):
            lines = text.split('\n')

            for i, line in enumerate(lines):
                line_stripped = line.strip()

                # Look for capacity values in any line
                capacity_match = re.search(
                    r'(\d[\d,]*\.?\d*)\s*(tons?/h(?:ou)?r|(?:bushels?|bu\.?)/h(?:ou)?r|'
                    r'(?:bushels?|bu\.?)/day|(?:bushels?|bu\.?)\s*/\s*hr|'
                    r'tons?/day|tons?/yr|tons?/year|MMgal(?:lons)?/year|Kgal/hour|'
                    r'gallons?/year|gallons?/hr|gallons?\s*per\s*(?:year|hour|day)|'
                    r'MMBtu/hr|gal/ton|lb\.?\s*(?:soybeans?|beans?)/hr|'
                    r'lb\.?/h(?:ou)?r|lb\.?\s*/\s*hr|'
                    r'tons?\s*per\s*(?:hour|day|year)|'
                    r'(?:bushels?|bu\.?)\s*per\s*(?:hour|day))',
                    line_stripped, re.IGNORECASE
                )

                if capacity_match:
                    value_str = capacity_match.group(1).replace(',', '')
                    try:
                        value = float(value_str)
                    except ValueError:
                        continue
                    unit = capacity_match.group(2)

                    # Get context: look at surrounding lines
                    context_lines = []
                    start = max(0, i - 5)
                    end = min(len(lines), i + 3)
                    for j in range(start, end):
                        context_lines.append(lines[j].strip())
                    context = ' '.join(context_lines)

                    # Try to find EU number
                    eu_match = re.search(
                        r'(?:EU-?|EU\s|EP-?|EP\s|Emission\s*(?:Unit|Point)\s*(?:ID\s*)?'
                        r'(?:Number\s*)?[:.]?\s*)(\d+[\.\d]*)',
                        context, re.IGNORECASE)
                    eu_id = eu_match.group(1) if eu_match else ''

                    # Try to get description from context
                    desc = self._extract_description(context, line_stripped)

                    self.emission_units.append({
                        'page': page_idx + 1,
                        'eu_id': eu_id,
                        'description': desc,
                        'capacity_value': value,
                        'capacity_unit': unit,
                        'raw_line': line_stripped,
                        'context': context[:300],
                    })

                # Also look for throughput limits in operating requirements
                throughput_match = re.search(
                    r'(?:not exceed|maximum|shall not exceed|limit|'
                    r'maximum rated capacity|rated capacity)\D{0,30}'
                    r'(\d[\d,]*\.?\d*)\s*(bushels?\s*per\s*day|'
                    r'tons?\s*per\s*(?:hour|day|year)|'
                    r'gallons?\s*(?:per|of)\s*(?:\w+\s*){0,3}'
                    r'(?:per|/)?\s*(?:twelve.month|12.month|rolling|year|annual))',
                    line_stripped, re.IGNORECASE
                )

                if throughput_match:
                    value_str = throughput_match.group(1).replace(',', '')
                    try:
                        value = float(value_str)
                    except ValueError:
                        continue
                    unit_text = throughput_match.group(2)

                    context_lines = []
                    start = max(0, i - 3)
                    end = min(len(lines), i + 2)
                    for j in range(start, end):
                        context_lines.append(lines[j].strip())
                    context = ' '.join(context_lines)

                    self.emission_units.append({
                        'page': page_idx + 1,
                        'eu_id': '',
                        'description': 'Operating Limit',
                        'capacity_value': value,
                        'capacity_unit': unit_text.strip(),
                        'raw_line': line_stripped,
                        'context': context[:300],
                    })

    def _extract_description(self, context, current_line):
        """Try to extract equipment description from context."""
        desc_patterns = [
            r'(?:EU-?\d+[\.\d]*|EP-?\d+[\.\d]*)\s+(.+?)(?:\d+[\d,]*\.?\d*\s*(?:tons|bushels|gal))',
            r'(?:Emission\s+(?:Unit|Point)\s+Description)\s*(.+?)(?:Raw\s+Material|Control)',
            r'(?:Source\s+Description)\s*[:.]?\s*(.+?)(?:Control|Rated)',
        ]
        for pat in desc_patterns:
            m = re.search(pat, context, re.IGNORECASE)
            if m:
                desc = m.group(1).strip()
                desc = re.sub(r'\s+', ' ', desc)
                if 5 < len(desc) < 100:
                    return desc

        # Fallback: look for process-related keywords near the capacity
        keywords_found = []
        context_lower = context.lower()
        for kw in CRUSH_KEYWORDS + REFINERY_KEYWORDS + BIODIESEL_KEYWORDS:
            if kw in context_lower:
                keywords_found.append(kw)
        if keywords_found:
            return '; '.join(keywords_found[:3])

        return ''

    def _classify_equipment(self):
        """Classify emission units and build capacity summary."""
        crush_units = []
        receiving_units = []
        refinery_units = []
        biodiesel_units = []
        boiler_units = []
        other_units = []

        # Keywords that indicate handling/storage (NOT processing rate)
        receiving_keywords = [
            'grain receiving', 'grain handling', 'grain unload',
            'bean receiving', 'bean rail unload', 'truck unload',
            'truck dump', 'rail unload', 'bean truck dump',
            'grain transfer', 'grain convey',
            'meal storage', 'meal loadout', 'meal transfer',
            'meal sifter', 'meal screening',
            'hull grinding', 'hull storage', 'hull loadout',
            'pellet transfer', 'pellet cooler',
            'product loadout', 'haul road',
            'grain clean', 'grain dryer',
            'rail loadout', 'truck loadout', 'rail meal',
            'rail bean', 'truck bean', 'conveyance',
            'truck soybean', 'rail soybean',
            'soybean storage', 'bean storage',
        ]

        # Keywords that indicate actual soybean PROCESSING (the real crush rate)
        processing_keywords = [
            'extraction', 'extractor', 'solvent extraction',
            'oil extraction', 'vegetable oil processing',
            'flak', 'conditioning', 'bean conditioning',
            'soybean conditioning', 'cracking', 'dehull',
            'prep aspiration', 'spent flake',
        ]

        for eu in self.emission_units:
            ctx = (eu['description'] + ' ' + eu['context']).lower()

            if any(kw in ctx for kw in REFINERY_KEYWORDS):
                refinery_units.append(eu)
            elif any(kw in ctx for kw in BIODIESEL_KEYWORDS):
                biodiesel_units.append(eu)
            elif any(kw in ctx for kw in ['boiler', 'mmbtu']):
                boiler_units.append(eu)
            elif any(kw in ctx for kw in receiving_keywords):
                receiving_units.append(eu)
            elif any(kw in ctx for kw in CRUSH_KEYWORDS):
                crush_units.append(eu)
            else:
                other_units.append(eu)

        # Find the primary crush capacity
        crush_capacity_tph = 0
        crush_capacity_bpd = 0
        crush_capacity_bph = 0
        crush_description = ''

        # First pass: look for extraction/flaking/conditioning (actual process rate)
        for eu in crush_units:
            ctx = (eu['description'] + ' ' + eu['context']).lower()
            is_processing = any(kw in ctx for kw in processing_keywords)
            unit = eu['capacity_unit'].lower()
            val = eu['capacity_value']

            if is_processing:
                if 'tons/h' in unit or 'tons/hour' in unit or 'tons per hour' in unit:
                    if val > crush_capacity_tph:
                        crush_capacity_tph = val
                        crush_description = eu['description'] or eu['raw_line']
                elif 'bu' in unit and '/h' in unit:
                    if val > crush_capacity_bph:
                        crush_capacity_bph = val
                        if not crush_description:
                            crush_description = eu['description'] or eu['raw_line']
                elif ('bu' in unit and '/day' in unit) or 'bushels per day' in unit:
                    if val > crush_capacity_bpd:
                        crush_capacity_bpd = val
                elif 'lb' in unit and ('soybean' in unit or 'bean' in unit):
                    tph = val / 2000.0
                    if tph > crush_capacity_tph:
                        crush_capacity_tph = tph
                        crush_description = eu['description'] or eu['raw_line']
                elif 'lb' in unit and '/h' in unit:
                    # Generic lb/hr — could be soybeans if in processing context
                    tph = val / 2000.0
                    if tph > crush_capacity_tph:
                        crush_capacity_tph = tph
                        crush_description = eu['description'] or eu['raw_line']

        # If no processing-specific capacity found, use general crush units
        if not crush_capacity_tph and not crush_capacity_bph:
            for eu in crush_units:
                unit = eu['capacity_unit'].lower()
                val = eu['capacity_value']
                if 'tons/h' in unit or 'tons/hour' in unit or 'tons per hour' in unit:
                    if val > crush_capacity_tph:
                        crush_capacity_tph = val
                        crush_description = eu['description'] or eu['raw_line']
                elif 'bu' in unit and '/h' in unit:
                    if val > crush_capacity_bph:
                        crush_capacity_bph = val
                elif 'bu' in unit and ('/day' in unit or 'per day' in unit):
                    if val > crush_capacity_bpd:
                        crush_capacity_bpd = val

        # Convert bu/hr to tons/hr if we only have bu/hr
        # 1 bushel soybeans = 60 lbs = 0.03 short tons
        if crush_capacity_bph and not crush_capacity_tph:
            crush_capacity_tph = crush_capacity_bph * 60.0 / 2000.0

        # Convert bu/hr to bu/day if we only have hourly
        if crush_capacity_bph and not crush_capacity_bpd:
            crush_capacity_bpd = crush_capacity_bph * 24.0

        # Find refinery capacity
        refinery_capacity_tph = 0
        refinery_capacity_tpy = 0
        refinery_description = ''
        has_refinery = len(refinery_units) > 0
        for eu in refinery_units:
            unit = eu['capacity_unit'].lower()
            val = eu['capacity_value']
            if 'tons/h' in unit or 'tons/hour' in unit or 'tons per hour' in unit:
                if val > refinery_capacity_tph:
                    refinery_capacity_tph = val
                    refinery_description = eu['description'] or eu['raw_line']
            elif 'tons/yr' in unit or 'tons/year' in unit or 'per year' in unit:
                if val > refinery_capacity_tpy:
                    refinery_capacity_tpy = val

        # Determine refining type from keywords
        refining_type = ''
        if has_refinery:
            all_refinery_context = ' '.join(
                eu['context'].lower() for eu in refinery_units
            )
            if 'deodor' in all_refinery_context:
                refining_type = 'RBD'
            elif 'bleach' in all_refinery_context or 'clay' in all_refinery_context:
                refining_type = 'RB'
            elif 'degum' in all_refinery_context:
                refining_type = 'Degummed'
            elif 'caustic' in all_refinery_context or 'neutrali' in all_refinery_context:
                refining_type = 'Caustic Refined'
            else:
                refining_type = 'Unknown'

        # Biodiesel capacity
        biodiesel_capacity_mgy = 0
        for eu in biodiesel_units:
            unit = eu['capacity_unit'].lower()
            val = eu['capacity_value']
            if 'mmgal' in unit:
                if val > biodiesel_capacity_mgy:
                    biodiesel_capacity_mgy = val
            elif 'gallons' in unit and ('year' in unit or 'annual' in unit):
                mgy = val / 1e6
                if mgy > biodiesel_capacity_mgy:
                    biodiesel_capacity_mgy = mgy

        self.capacity_summary = {
            'facility_name': self.facility_name,
            'permit_number': self.permit_number,
            'facility_description': self.facility_description,
            'crush_capacity_tons_per_hour': crush_capacity_tph,
            'crush_capacity_bushels_per_day': crush_capacity_bpd,
            'crush_description': crush_description,
            'has_refinery': has_refinery,
            'refinery_capacity_tons_per_hour': refinery_capacity_tph,
            'refinery_capacity_tons_per_year': refinery_capacity_tpy,
            'refinery_description': refinery_description,
            'refining_type': refining_type,
            'has_biodiesel': len(biodiesel_units) > 0,
            'biodiesel_capacity_mgy': biodiesel_capacity_mgy,
            'total_emission_units_found': len(self.emission_units),
            'crush_units': len(crush_units),
            'refinery_units': len(refinery_units),
            'biodiesel_units': len(biodiesel_units),
            'all_emission_units': self.emission_units,
        }


# =============================================================================
# PERMIT PDF DOWNLOADER
# =============================================================================

class NebraskaPermitDownloader:
    """
    Downloads Title V permit PDFs for Nebraska facilities.

    NOTE: The NDEE web portal does NOT serve permit PDFs. This downloader
    attempts multiple strategies:
      1. Check for manually-placed PDFs in raw/ directory
      2. Try known NDEE URL patterns (unlikely to work)

    For actual PDF acquisition, submit a public records request to:
      NDEE Air Quality Division: 402-471-2186
      https://dee.nebraska.gov/about/about-agency/request-public-records
    """

    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                          'AppleWebKit/537.36',
        })

    def download_permit(self, key, permit_info):
        """
        Attempt to download or locate a permit PDF for a facility.
        Returns filepath if found, None otherwise.
        """
        filename = 'ne_{}_titlev.pdf'.format(key)
        filepath = self.output_dir / filename

        # Check if already downloaded/placed manually
        if filepath.exists() and filepath.stat().st_size > 10000:
            print('  [FOUND] {} ({} KB)'.format(
                filename, filepath.stat().st_size // 1024))
            return str(filepath)

        # Also check for any PDF with this key prefix
        for f in self.output_dir.glob('ne_{}*.pdf'.format(key)):
            if f.stat().st_size > 10000:
                print('  [FOUND] {} ({} KB)'.format(
                    f.name, f.stat().st_size // 1024))
                return str(f)

        appno = permit_info.get('ndee_appno', '')
        modsq = permit_info.get('modsq', '')
        prgid = permit_info.get('ndee_prgid1', '') or permit_info.get('prgid1', '')

        if not appno and not modsq:
            print('  [SKIP] {} - no NDEE permit ID found'.format(
                permit_info['name']))
            return None

        # NOTE: NDEE portal does not serve PDFs. Log the info for manual retrieval.
        print('  [NEED] {} - PDF not available from NDEE portal'.format(
            permit_info['name']))
        print('         APPNO: {}  PRGID1: {}  MODSQ: {}'.format(
            appno, prgid, modsq))
        print('         Request from NDEE Air Quality: 402-471-2186')

        return None

    def download_all(self, permits=None):
        """Check for/download all permits. Returns dict of key -> filepath."""
        permits = permits or NE_SOY_PERMITS
        results = {}
        found = 0
        needed = 0

        print('\nChecking {} Nebraska permit PDFs...'.format(len(permits)))
        for key, info in permits.items():
            filepath = self.download_permit(key, info)
            if filepath:
                results[key] = filepath
                found += 1
            else:
                needed += 1

        print('\n  {} PDFs found, {} need manual acquisition'.format(
            found, needed))
        return results


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def create_capacity_excel(results, permits, output_path):
    """Create Excel workbook with capacity data."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = 'Capacity Summary'

    headers = [
        'Facility', 'City', 'County', 'NDEE APPNO', 'NDEE Permit Type',
        'Crush (tons/hr)', 'Crush (bu/day)',
        'Has Refinery', 'Refinery (tons/hr)', 'Refinery (tons/yr)',
        'Refining Type',
        'Has Biodiesel', 'Biodiesel (MGY)',
        'FRS Registry ID', 'ECHO Source ID', 'SIC Codes',
        'NDEE MODSQ', 'NDEE PRGID1', 'Status',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

    row = 2
    for key in sorted(permits.keys()):
        info = permits[key]
        data = results.get(key, {})
        summary = data.get('summary', {})

        values = [
            summary.get('facility_name') or info.get('name', ''),
            info.get('city', ''),
            info.get('county', ''),
            info.get('ndee_appno', '') or info.get('permit_number', ''),
            info.get('ndee_permit_type', ''),
            summary.get('crush_capacity_tons_per_hour', 0) or '',
            summary.get('crush_capacity_bushels_per_day', 0) or '',
            'Yes' if summary.get('has_refinery') else 'No',
            summary.get('refinery_capacity_tons_per_hour', 0) or '',
            summary.get('refinery_capacity_tons_per_year', 0) or '',
            summary.get('refining_type', ''),
            'Yes' if summary.get('has_biodiesel') else 'No',
            summary.get('biodiesel_capacity_mgy', 0) or '',
            info.get('frs_registry_id', ''),
            info.get('source_id', ''),
            info.get('sic_codes', ''),
            info.get('modsq', ''),
            info.get('ndee_prgid1', '') or info.get('prgid1', ''),
            info.get('status', ''),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col in (6, 7, 9, 10, 13) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

        # Highlight refineries
        if summary.get('has_refinery'):
            for c in range(8, 12):
                ws.cell(row=row, column=c).fill = PatternFill(
                    'solid', fgColor='E2EFDA')

        # Highlight planned facilities
        if info.get('status') == 'Planned':
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = PatternFill(
                    'solid', fgColor='FCE4D6')

        row += 1

    # Column widths
    widths = [35, 18, 14, 16, 16, 14, 14, 12, 14, 14, 14, 12, 14,
              16, 22, 24, 10, 14, 12]
    for col, w in enumerate(widths, 1):
        if col <= len(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # --- Sheet 2: NDEE Permit Metadata ---
    ws2 = wb.create_sheet('NDEE Permit Metadata')
    meta_headers = [
        'Facility', 'City', 'County', 'MODSQ', 'PRGID1',
        'Permit Type', 'APPNO', 'Issued Date', 'Expiry Date',
        'NDEE Name', 'ECHO Source ID', 'FRS ID',
    ]
    for col, header in enumerate(meta_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for key in sorted(permits.keys()):
        info = permits[key]
        values = [
            info.get('name', ''),
            info.get('city', ''),
            info.get('county', ''),
            info.get('modsq', ''),
            info.get('ndee_prgid1', '') or info.get('prgid1', ''),
            info.get('ndee_permit_type', ''),
            info.get('ndee_appno', ''),
            info.get('ndee_issued_date', ''),
            info.get('ndee_expiry_date', ''),
            info.get('ndee_name', ''),
            info.get('source_id', ''),
            info.get('frs_registry_id', ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border
        row += 1

    meta_widths = [35, 18, 14, 10, 14, 16, 16, 14, 14, 35, 22, 16]
    for col, w in enumerate(meta_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # --- Sheet 3: Equipment Detail (if any PDFs were parsed) ---
    ws3 = wb.create_sheet('Equipment Detail')
    detail_headers = [
        'Facility', 'Page', 'EU ID', 'Description', 'Category',
        'Capacity Value', 'Capacity Unit', 'Raw Line',
    ]
    for col, header in enumerate(detail_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for key in sorted(results.keys()):
        data = results[key]
        info = permits.get(key, {})
        summary = data.get('summary', {})
        fac_name = summary.get('facility_name') or info.get('name', '')

        for eu in summary.get('all_emission_units', []):
            ctx = (eu.get('description', '') + ' ' + eu.get('context', '')).lower()
            if any(kw in ctx for kw in REFINERY_KEYWORDS):
                category = 'Refinery'
            elif any(kw in ctx for kw in BIODIESEL_KEYWORDS):
                category = 'Biodiesel'
            elif any(kw in ctx for kw in ['boiler', 'mmbtu']):
                category = 'Boiler'
            elif any(kw in ctx for kw in CRUSH_KEYWORDS):
                category = 'Crush/Processing'
            else:
                category = 'Other'

            values = [
                fac_name,
                eu.get('page', ''),
                eu.get('eu_id', ''),
                eu.get('description', ''),
                category,
                eu.get('capacity_value', ''),
                eu.get('capacity_unit', ''),
                eu.get('raw_line', ''),
            ]
            for col, val in enumerate(values, 1):
                cell = ws3.cell(row=row, column=col, value=val)
                cell.border = thin_border
            row += 1

    detail_widths = [35, 8, 10, 40, 16, 14, 18, 60]
    for col, w in enumerate(detail_widths, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(str(output_path))
    print('\nExcel saved: {}'.format(output_path))


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Nebraska NDEE Title V Permit Capacity Collector'
    )
    parser.add_argument('--download-only', action='store_true',
                        help='Only download metadata + PDFs, do not parse')
    parser.add_argument('--parse-only', action='store_true',
                        help='Parse already-downloaded PDFs')
    parser.add_argument('--metadata-only', action='store_true',
                        help='Only download NDEE permit metadata (no PDFs)')
    parser.add_argument('--facility', '-f', default=None,
                        help='Filter to a single facility (name substring match)')
    parser.add_argument('--output', '-o', default=None,
                        help='Output Excel path')
    parser.add_argument('--force-refresh', action='store_true',
                        help='Force re-download of NDEE permit database')
    args = parser.parse_args()

    raw_dir = Path(__file__).parent / 'raw'
    output_dir = Path(__file__).parent / 'output'
    output_dir.mkdir(exist_ok=True)

    # Filter permits if requested
    permits = dict(NE_SOY_PERMITS)  # copy
    if args.facility:
        search = args.facility.lower()
        permits = {
            k: v for k, v in permits.items()
            if search in k.lower() or search in v['name'].lower()
            or search in v['city'].lower()
        }
        if not permits:
            print('No facilities matching "{}"'.format(args.facility))
            return

    print('Nebraska NDEE Title V Permit Capacity Collector')
    print('================================================')
    print('Facilities: {}'.format(len(permits)))
    print()

    # Step 1: Download NDEE metadata and match facilities
    if not args.parse_only:
        print('Step 1: Downloading NDEE permit metadata...')
        client = NDEEPermitClient(str(raw_dir))
        client.download_all_permits(force=args.force_refresh)
        permits = client.match_facilities(permits)

        # Save metadata results
        metadata_path = output_dir / 'nebraska_ndee_metadata_{}.json'.format(
            datetime.now().strftime('%Y-%m-%d'))
        meta_out = {}
        for key, info in permits.items():
            meta_out[key] = {
                k: v for k, v in info.items()
                if not k.startswith('_')
            }
        with open(str(metadata_path), 'w', encoding='utf-8') as f:
            json.dump(meta_out, f, indent=2)
        print('\nMetadata saved: {}'.format(metadata_path))

        if args.metadata_only:
            _print_metadata_summary(permits)
            return

    # Step 2: Download/locate PDFs
    downloaded = {}
    if not args.parse_only:
        print('\nStep 2: Locating permit PDFs...')
        downloader = NebraskaPermitDownloader(str(raw_dir))
        downloaded = downloader.download_all(permits)
    else:
        # Find already-downloaded PDFs
        for key in permits:
            filepath = raw_dir / 'ne_{}_titlev.pdf'.format(key)
            if filepath.exists() and filepath.stat().st_size > 10000:
                downloaded[key] = str(filepath)
            else:
                # Check for any PDF with this key
                for f in raw_dir.glob('ne_{}*.pdf'.format(key)):
                    if f.stat().st_size > 10000:
                        downloaded[key] = str(f)
                        break
        print('Found {} already-downloaded PDFs'.format(len(downloaded)))

    if args.download_only:
        print('\nDownload complete. {} PDFs available.'.format(len(downloaded)))
        _print_metadata_summary(permits)
        return

    # Step 3: Parse PDFs
    results = {}
    if downloaded:
        print('\nStep 3: Parsing {} Title V permits for capacity data...\n'.format(
            len(downloaded)))

        for key, filepath in sorted(downloaded.items()):
            info = permits.get(key, {})
            print('Parsing: {} ({})'.format(
                info.get('name', key), info.get('city', '')))

            try:
                parser_obj = TitleVCapacityParser(filepath)
                summary = parser_obj.parse()
                results[key] = {
                    'filepath': filepath,
                    'summary': summary,
                }

                # Print quick summary
                crush = summary.get('crush_capacity_tons_per_hour', 0)
                bpd = summary.get('crush_capacity_bushels_per_day', 0)
                ref = summary.get('has_refinery', False)
                ref_tph = summary.get('refinery_capacity_tons_per_hour', 0)
                ref_type = summary.get('refining_type', '')
                bio = summary.get('has_biodiesel', False)
                bio_mgy = summary.get('biodiesel_capacity_mgy', 0)

                parts = []
                if crush:
                    parts.append('Crush: {:.1f} tph'.format(crush))
                if bpd:
                    parts.append('{:,.0f} bu/day'.format(bpd))
                if ref:
                    parts.append('Refinery: {:.1f} tph ({})'.format(
                        ref_tph, ref_type))
                if bio:
                    parts.append('Biodiesel: {:.0f} MGY'.format(bio_mgy))
                if not parts:
                    parts.append('No capacity data extracted')

                print('  -> {}'.format(' | '.join(parts)))
                print()

            except Exception as e:
                print('  -> ERROR: {}'.format(str(e)[:100]))
                import traceback
                traceback.print_exc()
                results[key] = {'filepath': filepath, 'summary': {}}
                print()
    else:
        print('\nStep 3: No PDFs available to parse.')
        print('  Place permit PDFs in: {}'.format(raw_dir))
        print('  Naming: ne_<facility_key>_titlev.pdf')
        print('  Example: ne_adm_fremont_titlev.pdf')

    # Step 4: Output Excel + JSON
    timestamp = datetime.now().strftime('%Y-%m-%d')
    output_path = args.output or str(
        output_dir / 'nebraska_soy_capacity_{}.xlsx'.format(timestamp))

    create_capacity_excel(results, permits, output_path)

    # Save JSON (capacity data for facilities with parsed PDFs,
    # plus metadata for all facilities)
    json_path = output_path.replace('.xlsx', '.json')
    json_results = {}
    for key in sorted(permits.keys()):
        info = permits[key]
        data = results.get(key, {})
        summary = data.get('summary', {})

        # Remove the full emission_units list for the JSON (too verbose)
        clean = {k: v for k, v in summary.items()
                 if k != 'all_emission_units'} if summary else {}

        # Always include facility metadata even without parsed capacity
        if not clean:
            clean = {
                'facility_name': info.get('name', ''),
                'permit_number': info.get('permit_number', '')
                                 or info.get('ndee_appno', ''),
                'facility_description': 'SIC: {}'.format(
                    info.get('sic_codes', '')),
                'crush_capacity_tons_per_hour': 0,
                'crush_capacity_bushels_per_day': 0,
                'crush_description': '',
                'has_refinery': False,
                'refinery_capacity_tons_per_hour': 0,
                'refinery_capacity_tons_per_year': 0,
                'refinery_description': '',
                'refining_type': '',
                'has_biodiesel': False,
                'biodiesel_capacity_mgy': 0,
                'total_emission_units_found': 0,
                'crush_units': 0,
                'refinery_units': 0,
                'biodiesel_units': 0,
            }

        # Add NE-specific metadata
        clean['frs_registry_id'] = info.get('frs_registry_id', '')
        clean['source_id'] = info.get('source_id', '')
        clean['ndee_modsq'] = info.get('modsq', '')
        clean['ndee_appno'] = info.get('ndee_appno', '')
        clean['ndee_prgid1'] = info.get('ndee_prgid1', '') or info.get('prgid1', '')
        clean['ndee_permit_type'] = info.get('ndee_permit_type', '')
        clean['echo_status'] = info.get('status', '')
        clean['pdf_parsed'] = key in results and bool(
            results[key].get('summary', {}).get('total_emission_units_found'))

        json_results[key] = clean

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_results, f, indent=2, default=str)
    print('JSON saved: {}'.format(json_path))

    # Print summary table
    _print_capacity_summary(results, permits)
    _print_metadata_summary(permits)


def _print_capacity_summary(results, permits):
    """Print a formatted capacity summary table."""
    parsed = {k: v for k, v in results.items()
              if v.get('summary', {}).get('total_emission_units_found')}

    if not parsed:
        print('\nNo PDFs were parsed. Capacity data requires Title V permit PDFs.')
        print('See NDEE metadata summary below for permit identification.')
        return

    print('\n' + '=' * 90)
    print('NEBRASKA SOYBEAN/OILSEED PROCESSING CAPACITY SUMMARY')
    print('=' * 90)
    print('{:<35s} {:>10s} {:>10s} {:>8s} {:>10s} {:>8s}'.format(
        'Facility', 'Crush tph', 'Crush bpd', 'Refine?', 'Ref tph', 'Type'))
    print('-' * 90)

    for key, data in sorted(parsed.items()):
        s = data.get('summary', {})
        info = permits.get(key, {})
        name = (s.get('facility_name') or info.get('name', ''))[:34]
        crush = s.get('crush_capacity_tons_per_hour', 0)
        bpd = s.get('crush_capacity_bushels_per_day', 0)
        has_ref = 'YES' if s.get('has_refinery') else 'No'
        ref_tph = s.get('refinery_capacity_tons_per_hour', 0)
        ref_type = s.get('refining_type', '')[:8]

        print('{:<35s} {:>10s} {:>10s} {:>8s} {:>10s} {:>8s}'.format(
            name,
            '{:.1f}'.format(crush) if crush else '-',
            '{:,.0f}'.format(bpd) if bpd else '-',
            has_ref,
            '{:.1f}'.format(ref_tph) if ref_tph else '-',
            ref_type,
        ))

    print('=' * 90)


def _print_metadata_summary(permits):
    """Print NDEE metadata matching summary."""
    print('\n' + '=' * 100)
    print('NDEE PERMIT METADATA SUMMARY')
    print('=' * 100)
    print('{:<30s} {:>12s} {:>10s} {:>14s} {:>12s} {:>12s}'.format(
        'Facility', 'City', 'MODSQ', 'APPNO', 'Permit Type', 'Status'))
    print('-' * 100)

    for key in sorted(permits.keys()):
        info = permits[key]
        name = info.get('name', '')[:29]
        city = info.get('city', '')[:12]
        modsq = info.get('modsq', '') or '-'
        appno = info.get('ndee_appno', '') or '-'
        ptype = info.get('ndee_permit_type', '') or '-'
        status = info.get('status', '')

        print('{:<30s} {:>12s} {:>10s} {:>14s} {:>12s} {:>12s}'.format(
            name, city, modsq, appno, ptype, status))

    matched = sum(1 for v in permits.values() if v.get('modsq'))
    total = len(permits)
    print('-' * 100)
    print('Matched: {} / {}  |  Unmatched: {} (may need manual NDEE lookup)'.format(
        matched, total, total - matched))
    print('=' * 100)

    # Remind about PDF acquisition
    unmatched_pdf = [
        k for k, v in permits.items()
        if v.get('modsq') and not (
            Path(__file__).parent / 'raw' / 'ne_{}_titlev.pdf'.format(k)
        ).exists()
    ]
    if unmatched_pdf:
        print('\nPDFs needed for {} matched facilities:'.format(
            len(unmatched_pdf)))
        print('  Request from NDEE Air Quality Division: 402-471-2186')
        print('  Or: https://dee.nebraska.gov/about/about-agency/'
              'request-public-records')
        print('  Place PDFs in: {}'.format(Path(__file__).parent / 'raw'))
        for k in sorted(unmatched_pdf):
            info = permits[k]
            print('    ne_{}_titlev.pdf  ({}, PRGID1: {})'.format(
                k, info['name'],
                info.get('ndee_prgid1', '') or info.get('prgid1', '')))


if __name__ == '__main__':
    main()
