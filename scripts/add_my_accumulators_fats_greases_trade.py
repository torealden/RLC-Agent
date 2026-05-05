"""
Add Oct-Sep marketing year accumulator formulas to every trade tab in
models/Fats and Greases/us_fats_greases_trade.xlsm.

What it does:
  - For each sheet, scans row 2 headers
  - Identifies date-typed headers (the actual month columns) and string
    headers in the form "YYYY/YY" (the MY accumulator columns)
  - For each MY column, writes =SUM(<Oct-cell>:<Sep-cell>) into every
    data row (4..217). Regional subtotal rows get the same SUM formula
    on the same row — their shipped totals already aggregate vertically,
    and SUM-Oct-through-Sep-on-the-same-row is the right horizontal
    aggregation for the MY accumulator.
  - Skips any MY header whose Oct-YYYY or Sep-(YYYY+1) date column is
    missing from the sheet (logs which ones).

Convention (per project_calendar_year_vs_marketing_year.md):
  Fats/oils/meals use Oct-Sep MY. Display as "YYYY/YY".
  e.g. "2024/25" = Oct 2024 - Sep 2025.

The hardened TradeUpdaterSQL macro (FindColumnForDate / IsMonthHeaderCell
now reject String headers) prevents these formula columns from being
clobbered by future updates. But the .bas must be re-imported via VBE
for that fix to take effect — see scripts/README or memory notes.

Usage:
    python scripts/add_my_accumulators_fats_greases_trade.py
"""
from __future__ import annotations

import datetime
import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


WB_PATH = Path("models/Fats and Greases/us_fats_greases_trade.xlsm")
HEADER_ROW = 2
DATA_START_ROW = 4
DATA_END_ROW = 217


def parse_my_header(h) -> int | None:
    """Parse 'YYYY/YY' marketing year header -> start year. Returns None
    if the header is not a MY accumulator label."""
    if not isinstance(h, str):
        return None
    s = h.strip()
    if "/" not in s:
        return None
    try:
        a, b = s.split("/", 1)
        if not (a.isdigit() and b.isdigit()):
            return None
        start_year = int(a)
        if start_year < 1990 or start_year > 2050:
            return None
        return start_year
    except (ValueError, IndexError):
        return None


def add_my_accumulators(ws) -> tuple[int, int, list[str]]:
    """Write Oct-Sep SUM formulas in every data row for each MY column.

    Returns (n_my_cols_processed, n_formulas_written, warnings).
    """
    date_to_col: dict[tuple[int, int], int] = {}
    my_cols: list[tuple[int, int, str]] = []

    last_col = ws.max_column or 0
    for col in range(2, last_col + 1):
        h = ws.cell(row=HEADER_ROW, column=col).value
        if isinstance(h, datetime.datetime):
            date_to_col[(h.year, h.month)] = col
        else:
            start = parse_my_header(h)
            if start is not None:
                my_cols.append((col, start, str(h).strip()))

    warnings: list[str] = []
    n_formulas = 0
    n_processed = 0
    for col, start_year, label in my_cols:
        oct_col = date_to_col.get((start_year, 10))
        sep_col = date_to_col.get((start_year + 1, 9))
        if oct_col is None or sep_col is None:
            warnings.append(
                f"col {col} ({label!r}): missing date col "
                f"oct-{start_year}={oct_col} sep-{start_year + 1}={sep_col} - skipped"
            )
            continue
        oct_letter = get_column_letter(oct_col)
        sep_letter = get_column_letter(sep_col)
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            formula = f"=SUM({oct_letter}{row}:{sep_letter}{row})"
            ws.cell(row=row, column=col).value = formula
            n_formulas += 1
        n_processed += 1
    return n_processed, n_formulas, warnings


def main():
    if not WB_PATH.exists():
        raise SystemExit(f"Workbook not found: {WB_PATH}")

    backup = WB_PATH.with_name(
        WB_PATH.stem + "_backup_my_" +
        datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + WB_PATH.suffix
    )
    shutil.copy2(WB_PATH, backup)
    print(f"Backed up to: {backup.name}")

    wb = openpyxl.load_workbook(WB_PATH, keep_vba=True, data_only=False)
    print(f"Sheets to process: {len(wb.sheetnames)}")
    print()

    grand_cols = grand_formulas = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n_cols, n_formulas, warnings = add_my_accumulators(ws)
        suffix = ""
        if warnings:
            suffix = "  (" + "; ".join(w[:60] for w in warnings[:2]) + ")"
        print(f"  {sheet_name:30s}  MY-cols={n_cols:>3}  formulas={n_formulas:>5,}{suffix}")
        grand_cols += n_cols
        grand_formulas += n_formulas

    print()
    print(f"Total: {grand_cols} MY columns updated across all tabs, "
          f"{grand_formulas:,} formulas written")

    wb.save(WB_PATH)
    print(f"Saved: {WB_PATH}")
    print()
    print("Next steps:")
    print("  1. Open the workbook in Excel - formulas will compute on first")
    print("     recalc, populating MY columns from the date columns.")
    print("  2. Re-import src/tools/TradeUpdaterSQL.bas into the workbook via")
    print("     VBE (Alt+F11 -> File -> Import File) so the IsMonthHeaderCell")
    print("     and FindColumnForDate fixes take effect. Until that re-import")
    print("     happens, Ctrl+I may still wipe these formulas on update.")


if __name__ == "__main__":
    main()
