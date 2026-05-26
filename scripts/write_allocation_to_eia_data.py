"""
Write allocation data from bronze.historical_feedstock_allocation
to eia_data.xlsm.

Replaces the manual feedstock_allocation_output.xlsx → eia_data.xlsm copy
step. Run this after the allocation engine updates the DB (or on a schedule).

For each (period, fuel_type, feedstock_code) row in the DB, write
quantity_mil_lbs into the matching column in the appropriate fuel tab.

Feedstock code → column mapping (DB code: eia_data column header):

  Biodiesel tab (no subtotal cols):       RD / SAF / CO-PROC tabs:
    SBO  → c2  Soybean Oil                   SBO  → c2
    CO   → c3  Canola/Rapeseed Oil           CO   → c3
    DCO  → c4  Corn Oil                      DCO  → c4
    CSO  → c5  Cottonseed Oil                CSO  → c5
    PALM → c6  Palm Oil                      PALM → c6
    PF   → c8  Poultry Fat                   PF   → c9
    BFT  → c9  Tallow                        BFT  → c10
    CWG  → c10 White Grease                  CWG  → c11
    YG   → c12 Yellow Grease                 YG   → c13
    UCO  → c13 Other Grease                  UCO  → c14
    (Total written to c16)                   (Total written to c17)
"""
from __future__ import annotations

import shutil
import sys
import tempfile
import time
from collections import defaultdict
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from dotenv import load_dotenv
load_dotenv('.env')

from openpyxl import load_workbook
import psycopg2.extras

from src.services.database.db_config import get_connection


# Primary location (in-repo). Falls back to the legacy Dropbox path if
# someone happens to be running against the synced Dropbox copy instead.
REPO_EIA_DATA = Path(r"C:/dev/RLC-Agent/models/Biofuels/eia_data.xlsm")
DROPBOX_EIA_DATA = Path(r"C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models/Biofuels/new_models/eia_data.xlsm")
EIA_DATA = REPO_EIA_DATA if REPO_EIA_DATA.exists() else DROPBOX_EIA_DATA

# Fuel type → sheet name in eia_data.xlsm. Allocator output uses
# 'coprocessing' (no underscore); legacy bronze.historical_feedstock_allocation
# uses 'co_processing'. Accept both spellings.
FUEL_TO_SHEET = {
    'biodiesel':        'biodiesel_monthly',
    'renewable_diesel': 'renewable_diesel_monthly',
    'saf':              'sustainable_aviation_monthly',
    'co_processing':    'co_processing_monthly',
    'coprocessing':     'co_processing_monthly',
}

# Feedstock column maps per tab layout
# Biodiesel has no subtotal cols (c8 Veg Oil Total, c16 Fat&Grease Total)
BIODIESEL_COLS = {
    'SBO':   2, 'CO':    3, 'DCO':   4, 'CSO':   5, 'PALM':  6,
    'PF':    8, 'BFT':   9, 'CWG':   10, 'YG':   12, 'UCO':  13,
}
BIODIESEL_TOTAL_COL = 16

# RD / SAF / CO-PROC layouts include subtotals (all shifted +1 after feedstocks)
OTHER_COLS = {
    'SBO':   2, 'CO':    3, 'DCO':   4, 'CSO':   5, 'PALM':  6,
    'PF':    9, 'BFT':   10, 'CWG':  11, 'YG':   13, 'UCO':  14,
}
OTHER_TOTAL_COL = 17


def fetch_allocation() -> dict:
    """Return {(fuel_type, period): {feedstock_code: quantity_mil_lbs}}.

    Source priority per (period, fuel, feedstock):
      1. eia_form819     — EIA canon (BD national totals)
      2. rlc_allocator_v1 — our allocator output (RD + coprocessing)
      3. fastmarkets     — legacy reference (RD + SAF forward projections)

    Highest-priority source wins; lower-priority sources do NOT add to
    the total (avoiding double-counting where allocator + fastmarkets
    both have RD).

    Also collapses tallow grade split (EBFT + IBFT) back to BFT for
    sheet display, since the xlsm has a single Tallow column.

    IP rule (memory: feedback_fastmarkets_keep_dont_show): FM-era rows
    must never appear in CLIENT-FACING material. eia_data.xlsm is
    currently Tore's INTERNAL modeling tool so fastmarkets is included
    here for forward-period coverage. When this artifact starts going
    to clients, drop 'fastmarkets' from the source_priority CTE.
    """
    data: dict = defaultdict(lambda: defaultdict(float))
    with get_connection() as conn:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("""
            WITH agg AS (
                SELECT
                    period,
                    fuel_type,
                    -- Collapse tallow grade split back to BFT for xlsm display
                    CASE WHEN feedstock_code IN ('EBFT', 'IBFT')
                         THEN 'BFT' ELSE feedstock_code END AS feedstock_code,
                    source,
                    SUM(quantity_mil_lbs) AS qty
                FROM bronze.historical_feedstock_allocation
                WHERE fuel_type IN ('biodiesel', 'renewable_diesel', 'saf',
                                    'co_processing', 'coprocessing')
                  AND quantity_mil_lbs IS NOT NULL
                GROUP BY period, fuel_type,
                         CASE WHEN feedstock_code IN ('EBFT', 'IBFT')
                              THEN 'BFT' ELSE feedstock_code END,
                         source
            ),
            ranked AS (
                SELECT *,
                       ROW_NUMBER() OVER (
                           PARTITION BY period, fuel_type, feedstock_code
                           ORDER BY CASE source
                               WHEN 'eia_form819'      THEN 1
                               WHEN 'rlc_allocator_v1' THEN 2
                               WHEN 'fastmarkets'      THEN 3
                               ELSE 99 END
                       ) AS rn
                FROM agg
            )
            SELECT period, fuel_type, feedstock_code, qty AS total_mil_lbs, source
            FROM ranked
            WHERE rn = 1
        """)
        for r in cur.fetchall():
            key = (r['fuel_type'], r['period'])
            data[key][r['feedstock_code']] = float(r['total_mil_lbs'])
    return data


def write_to_eia_data(allocation: dict):
    print(f">> opening {EIA_DATA}", flush=True)
    wb = load_workbook(EIA_DATA, keep_vba=True)

    stats = defaultdict(lambda: {'rows': 0, 'cells': 0, 'skipped_codes': set()})
    for (fuel_type, period), codes in allocation.items():
        sheet_name = FUEL_TO_SHEET.get(fuel_type)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Build date→row map once per sheet (cache)
        if not hasattr(ws, '_date_row_cache'):
            cache = {}
            for r in range(5, ws.max_row + 1):
                v = ws.cell(r, 1).value
                if hasattr(v, 'year'):
                    d = v.date() if hasattr(v, 'hour') else v
                    cache[d] = r
            ws._date_row_cache = cache
        date_to_row = ws._date_row_cache

        if period not in date_to_row:
            continue
        target_row = date_to_row[period]
        col_map = BIODIESEL_COLS if fuel_type == 'biodiesel' else OTHER_COLS
        total_col = BIODIESEL_TOTAL_COL if fuel_type == 'biodiesel' else OTHER_TOTAL_COL

        row_total = 0.0
        wrote_any = False
        for code, val in codes.items():
            if code in col_map:
                ws.cell(target_row, col_map[code], val)
                row_total += val
                stats[sheet_name]['cells'] += 1
                wrote_any = True
            else:
                stats[sheet_name]['skipped_codes'].add(code)
        # Write aggregated Total (sum of all feedstocks we have for this period)
        if wrote_any:
            ws.cell(target_row, total_col, row_total)
            stats[sheet_name]['cells'] += 1
            stats[sheet_name]['rows'] += 1

    for sheet, s in stats.items():
        skipped = ', '.join(sorted(s['skipped_codes'])) or '-'
        print(f"   {sheet}: {s['rows']} months, {s['cells']} cells, "
              f"skipped codes: {skipped}", flush=True)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm") as tmp:
        tmp_path = Path(tmp.name)
    t0 = time.time()
    wb.save(tmp_path)
    wb.close()
    print(f">> save: {time.time()-t0:.1f}s  size={tmp_path.stat().st_size/1024:.0f} KB", flush=True)
    shutil.copy2(tmp_path, EIA_DATA)
    tmp_path.unlink(missing_ok=True)
    print(f">> wrote {EIA_DATA}", flush=True)


def main():
    # Skip silently if the Dropbox workbook isn't synced on this machine —
    # this script is best-effort, not blocking. The allocator wraps the
    # subprocess call in try/except, so a sys.exit(0) here means "no work
    # to do" rather than a noisy failure.
    if not EIA_DATA.exists():
        print(f">> eia_data.xlsm not found at {EIA_DATA} — skipping sync (no-op)")
        return

    print("=" * 70)
    print("Writing bronze.historical_feedstock_allocation -> eia_data.xlsm")
    print("=" * 70)
    allocation = fetch_allocation()
    if not allocation:
        print("No allocation rows found in bronze.historical_feedstock_allocation.")
        return
    months = sorted({p for (_, p) in allocation.keys()})
    print(f"{len(allocation)} (fuel × period) records across "
          f"{len(months)} months ({months[0]} to {months[-1]})", flush=True)
    write_to_eia_data(allocation)


if __name__ == "__main__":
    main()
