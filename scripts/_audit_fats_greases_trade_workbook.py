"""Audit models/Fats and Greases/us_fats_greases_trade.xlsm:
- list every tab
- show row 2 header types per tab (date vs MY accumulator vs other)
- flag whether MY headers parse as dates (which would let IsMonthHeaderCell
  in the macro misfire)
"""
from __future__ import annotations
import datetime
from pathlib import Path
import openpyxl

WB_PATH = Path("models/Fats and Greases/us_fats_greases_trade.xlsm")
HEADER_ROW = 2


def header_kind(v):
    if v is None:
        return "EMPTY"
    if isinstance(v, datetime.datetime):
        return f"DATE({v:%Y-%m})"
    if isinstance(v, str):
        s = v.strip()
        if "/" in s:
            try:
                # MY header parse "YYYY/YY" or "YYYY/YYYY"
                a, b = s.split("/", 1)
                if a.isdigit() and b.isdigit():
                    return f"MY({s})"
            except ValueError:
                pass
        return f"STR({s[:25]})"
    return f"OTHER({type(v).__name__}={v!r})"


def main():
    wb = openpyxl.load_workbook(WB_PATH, keep_vba=True, data_only=False)
    print(f"Workbook: {WB_PATH}")
    print(f"Sheets: {len(wb.sheetnames)}")
    print()

    for name in wb.sheetnames:
        ws = wb[name]
        # Skip sheets that don't look like trade data (no row 2 headers)
        last_col = ws.max_column or 0
        if last_col < 2:
            print(f"  {name:35s}  (no data, max_col={last_col})")
            continue

        # Categorize first 30 columns + total counts
        n_dates = n_my = n_str = n_other = n_empty = 0
        my_examples = []
        for c in range(2, last_col + 1):
            v = ws.cell(row=HEADER_ROW, column=c).value
            k = header_kind(v)
            if k.startswith("DATE"):
                n_dates += 1
            elif k.startswith("MY"):
                n_my += 1
                if len(my_examples) < 4:
                    my_examples.append((c, v))
            elif k == "EMPTY":
                n_empty += 1
            elif k.startswith("STR"):
                n_str += 1
                if len(my_examples) < 4 and "/" in str(v):
                    my_examples.append((c, v))
            else:
                n_other += 1

        my_str = ", ".join(f"col{c}={v!r}" for c, v in my_examples) if my_examples else ""
        print(f"  {name:35s}  cols={last_col:>3}  date={n_dates:>3}  "
              f"my={n_my:>2}  str={n_str:>2}  empty={n_empty:>2}  other={n_other:>2}")
        if my_str:
            print(f"      MY/str sample: {my_str}")


if __name__ == "__main__":
    main()
