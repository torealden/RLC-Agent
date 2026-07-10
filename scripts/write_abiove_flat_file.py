"""Write the Abiove Brazil soy-complex flat file from gold.abiove_soy_complex_monthly.

Output: models/Oilseeds/brazil_soy_complex.xlsx
  - soy_complex : 13-column LONG (flat_file_contract.md), sorted ascending, thousand MT
  - _meta       : one row per series (coverage, source, unit, vintage set, notes)

Desktop links the balance-sheet workbook to soy_complex via MAXIFS/SUMIFS on the key
columns (series, marketing_year, period, MAX vintage_rank). Units: thousand metric tons.

Usage:  python scripts/write_abiove_flat_file.py
"""
import sys
from pathlib import Path
ROOT = Path(r"C:/dev/RLC-Agent"); sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
import openpyxl
from openpyxl.styles import Font, PatternFill
from src.services.database.db_config import get_connection

OUT = ROOT / "models/Oilseeds/brazil_soy_complex.xlsx"
COLS = ['commodity', 'class', 'series', 'marketing_year', 'period_type', 'period',
        'vintage', 'vintage_rank', 'value', 'unit', 'source', 'release_date', 'revision']
HDR_FILL = PatternFill("solid", fgColor="3C7D22")   # internal-model green
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri")

SERIES_NOTES = {
    'crush':           'Soybean crush (Processamento), Tabela history 2012+',
    'meal_production': 'Soybean meal production (Balanco_Brasil); pre-2025 monthly TBD from legacy sheets',
    'oil_production':  'Soybean oil production, crude+refined (Balanco_Brasil); pre-2025 TBD',
    'seed_stocks':     'Soybean (grão) industry final stocks (Estoques_Finais) 2021+',
    'meal_stocks':     'Soybean meal (farelo) industry final stocks 2021+',
    'oil_stocks':      'Soybean oil (óleo) industry final stocks 2021+',
}

def fetch():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT {', '.join(COLS)} FROM gold.abiove_soy_complex_monthly")
        return [dict(r) for r in cur.fetchall()]

def sort_key(r):
    return (r['series'], r['class'], r['marketing_year'], int(r['period']), r['vintage_rank'])

def main():
    rows = fetch()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "soy_complex"
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        ws.cell(1, c).fill = HDR_FILL; ws.cell(1, c).font = HDR_FONT
    for r in sorted(rows, key=sort_key):
        ws.append([r[c] for c in COLS])

    # _meta: one row per series
    wm = wb.create_sheet("_meta")
    wm.append(['series', 'n', 'first', 'last', 'source', 'unit', 'vintages', 'notes'])
    for c in range(1, 9):
        wm.cell(1, c).fill = HDR_FILL; wm.cell(1, c).font = HDR_FONT
    agg = {}
    for r in rows:
        a = agg.setdefault(r['series'], {'n': 0, 'periods': [], 'vint': set()})
        a['n'] += 1
        a['periods'].append(r['marketing_year'] * 100 + int(r['period']))
        a['vint'].add(r['vintage'])
    for s in SERIES_NOTES:
        if s not in agg: continue
        a = agg[s]; ps = sorted(a['periods'])
        fmt = lambda p: f"{p//100}-{p % 100:02d}"
        wm.append([s, a['n'], fmt(ps[0]), fmt(ps[-1]), 'ABIOVE', '1000 MT',
                   ', '.join(sorted(a['vint'])), SERIES_NOTES[s]])
    wm.append([])
    wm.append(['Generated from', 'gold.abiove_soy_complex_monthly'])
    wm.append(['Units', 'thousand metric tons (1.000 t) — Abiove native'])
    wm.append(['Source', 'Abiove monthly soy-complex Power BI (operator-extracted)'])
    wm.append(['Update process', 'docs/runbooks/abiove_update_runbook.md'])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  soy_complex: {len(rows)} data rows, {len(SERIES_NOTES)} series")
    for s in SERIES_NOTES:
        if s in agg:
            ps = sorted(agg[s]['periods'])
            print(f"    {s:16} n={agg[s]['n']:3}  {ps[0]//100}-{ps[0]%100:02d} -> {ps[-1]//100}-{ps[-1]%100:02d}")

if __name__ == "__main__":
    main()
