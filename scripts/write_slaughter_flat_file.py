"""Regenerate models/Fats and Greases/us_livestock_slaughter.xlsx from gold.livestock_slaughter_flat.

The slaughter data is a supporting input to the fats/greases balance sheets: live-weight production
drives the tallow/CWG/poultry-fat yield calc (tallow_ruling §3). The workbook existed but had NO
writer (orphaned from the pipeline). This reproduces the exact legacy WIDE layout Desktop links to
(3-row grouped header: species / measure / unit, ascending by date) so existing formulas keep
resolving, and adds a _meta tab per the flat-file conventions.

Source view is refreshed by scripts/collect_nass_livestock_slaughter.py -> silver builders -> gold.
Run after a slaughter collection. Idempotent (full rewrite).
"""
import sys
from pathlib import Path

ROOT = Path(r"C:/dev/RLC-Agent"); sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from src.services.database.db_config import get_connection

OUT = ROOT / "models" / "Fats and Greases" / "us_livestock_slaughter.xlsx"
HEADER_FILL = PatternFill("solid", fgColor="3C7D22")   # internal-xlsx green (reference_excel_color_conventions)
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri")

# Column spec: (group_label, measure_label, unit, db_field) — None row = blank separator column.
SPEC = [
    ("",                    "Date",                 "",         "price_date"),
    ("Hogs",                "Commercial Slaughter", "head",     "hog_slaughter_head"),
    ("Hogs",                "Avg Live Weight",      "lbs/head", "hog_avg_wt"),
    ("Hogs",                "Live Weight Production","lbs",     "hog_production_lbs"),
    ("Hogs",                "FI Slaughter",         "head",     "hog_fi_head"),
    ("Hogs",                "FI Barrows & Gilts",   "head",     "hog_fi_bg_head"),
    ("Hogs",                "FI Sows",              "head",     "hog_fi_sows_head"),
    (None, None, None, None),
    ("Cattle (GE 500 lbs)", "Commercial Slaughter", "head",     "cattle_slaughter_head"),
    ("Cattle (GE 500 lbs)", "Avg Live Weight",      "lbs/head", "cattle_avg_wt"),
    ("Cattle (GE 500 lbs)", "Live Weight Production","lbs",     "cattle_production_lbs"),
    (None, None, None, None),
    ("Calves",              "Commercial Slaughter", "head",     "calves_slaughter_head"),
    ("Calves",              "Live Weight Production","lbs",     "calves_production_lbs"),
    (None, None, None, None),
    ("Chickens (Total FI)", "Slaughter",            "head",     "chickens_slaughter_head"),
    ("Chickens (Total FI)", "Avg Live Weight",      "lbs/head", "chickens_avg_wt"),
    ("Chickens (Total FI)", "Live Weight Production","lbs",     "chickens_production_lbs"),
    (None, None, None, None),
    ("Broilers (Young FI)", "Slaughter",            "head",     "broilers_slaughter_head"),
    ("Broilers (Young FI)", "Avg Live Weight",      "lbs/head", "broilers_avg_wt"),
    ("Broilers (Young FI)", "Live Weight Production","lbs",     "broilers_production_lbs"),
    (None, None, None, None),
    ("Turkeys (Total FI)",  "Slaughter",            "head",     "turkeys_slaughter_head"),
    ("Turkeys (Total FI)",  "Avg Live Weight",      "lbs/head", "turkeys_avg_wt"),
    ("Turkeys (Total FI)",  "Live Weight Production","lbs",     "turkeys_production_lbs"),
]


def main():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM gold.livestock_slaughter_flat ORDER BY price_date")  # ascending: latest at stable bottom
        rows = cur.fetchall()
    if not rows:
        raise SystemExit("gold.livestock_slaughter_flat returned no rows")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "US Livestock Slaughter"
    # 3-row header
    ws.append([c[0] for c in SPEC])   # group
    ws.append([c[1] for c in SPEC])   # measure
    ws.append([c[2] for c in SPEC])   # unit
    for r in range(1, 4):
        for c in range(1, len(SPEC) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                cell.fill = HEADER_FILL; cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
    # data (ascending)
    for rec in rows:
        out = []
        for _g, _m, _u, field in SPEC:
            if field is None:
                out.append(None)
            elif field == "price_date":
                out.append(rec["price_date"])
            else:
                v = rec[field]
                out.append(float(v) if v is not None else None)
        ws.append(out)
    ws.freeze_panes = "A4"

    # _meta tab
    wm = wb.create_sheet("_meta")
    wm.append(["key", "value"])
    latest = rows[-1]["price_date"]
    for k, v in [
        ("source", "gold.livestock_slaughter_flat (NASS Livestock & Poultry Slaughter)"),
        ("writer", "scripts/write_slaughter_flat_file.py"),
        ("collector", "nass_livestock_slaughter (dispatcher, monthly days 20-31)"),
        ("purpose", "live-weight production -> tallow/CWG/poultry-fat yield calc (tallow_ruling §3)"),
        ("row_count", len(rows)),
        ("period_min", str(rows[0]["price_date"])),
        ("period_max", str(latest)),
        ("sort", "ascending (latest row at stable bottom address for VLOOKUP)"),
    ]:
        wm.append([k, v])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(rows)} rows, {rows[0]['price_date']} .. {latest}")


if __name__ == "__main__":
    main()
