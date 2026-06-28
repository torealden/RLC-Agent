"""
Parse models/Biofuels/RD Feedstock Build Up.xlsx -> per-facility assumed feedstock MIX (%).

This is the CONSUMPTION prior (what a plant actually runs), distinct from CARB eligibility
(what it's allowed to run). Slots below CARB in the FFA hierarchy: CARB says allowed, this
says run, allocator economics adjusts from there.

The workbook is TRANSPOSED: facilities are COLUMNS, attributes ('Feedstock Mix', 'Name Plate
Capac', ...) are ROWS keyed by column A. Mix is a free-text % string ("23% YG - 10%CWG - ...",
"100% DCO", "60% Tallow - 40% ...").

VALIDATE mode (default): print each facility's raw->parsed mix, unmapped tokens, and master
match coverage. --write persists to reference.facility_assumed_mix (created if absent).
"""
import sys, argparse, re
from pathlib import Path
import openpyxl
ROOT = Path(__file__).resolve().parents[1]; sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
from src.services.database.db_config import get_connection
# site-aware matcher (KEEP city tokens — both sides have cities here, unlike CARB).
# strip only corp suffixes + most-generic words so the distinctive city/company drives the match.
_STRIP = re.compile(r'\b(llc|lp|inc|corp|corporation|company|co|ltd|holdings|energy|fuels?'
                    r'|renewables?|renewable|biodiesel|biofuels?|bio|bioenergy|diesel|products'
                    r'|refining|refinery|group|stage|expansion|ventures?|the)\b')
def _toks(name):
    if not name: return set()
    s = re.sub(r'[^a-z0-9 ]', ' ', str(name).lower())
    s = _STRIP.sub(' ', s)
    return {t for t in s.split() if len(t) > 2}
def site_score(a, b):
    A, B = _toks(a), _toks(b)
    if not A or not B: return 0.0
    return len(A & B) / min(len(A), len(B))

XLSX = ROOT / "models" / "Biofuels" / "RD Feedstock Build Up.xlsx"
SHEETS = ['Renewable Diesel', 'Co-Processing', 'Canada Renewable Diesel']

# feedstock-name -> master code. Order matters (check longer/specific first).
NAME2CODE = [
    ('distillers corn', 'DCO'), ('dco', 'DCO'),
    ('tech tallow', 'BFT'), ('beef tallow', 'BFT'), ('tallow', 'BFT'),
    ('choice white', 'CWG'), ('white grease', 'CWG'), ('cwg', 'CWG'),
    ('yellow grease', 'YG'), ('yg', 'YG'),
    ('used cooking', 'UCO'), ('uco', 'UCO'),
    ('soybean', 'SBO'), ('sbo', 'SBO'),
    ('canola', 'CAN'), ('rapeseed', 'CAN'),
    ('corn oil', 'CO'),
    ('poultry', 'PLT'), ('camelina', 'CAM'), ('carinata', 'CAR'), ('fish', 'FSH'),
    ('low ci', 'LCI'),   # unspecified low-CI placeholder (HollyFrontier) — keep visible as calib gap
]
# pct, optional separator dash, then feedstock name. Handles "33% DCO", "33% - UCO", "10%CWG".
TOK = re.compile(r'(\d+\.?\d*)\s*%\s*-?\s*([A-Za-z][A-Za-z0-9 ./&]*)')

def to_code(name):
    n = name.strip().lower()
    for key, code in NAME2CODE:
        if key in n:
            return code
    return None

TOK_REV = re.compile(r'([A-Za-z][A-Za-z0-9 ./&]*?)\s*(\d+\.?\d*)\s*%')   # "Tallow 70%" name-first
def _collect(pairs):
    mix, unmapped = {}, []
    for name, pct in pairs:
        code = to_code(name)
        if code: mix[code] = mix.get(code, 0.0) + float(pct)
        elif name.strip(): unmapped.append(name.strip())
    return mix, unmapped

def parse_mix(s):
    """Parse a % mix string. Handles pct-first ('23% YG') and name-first ('Tallow 70%');
    picks whichever sums closer to 100. Returns ({code:pct}, [unmapped tokens])."""
    if not s: return {}, []
    fwd = _collect((nm, pc) for pc, nm in TOK.findall(str(s)))      # pct-first
    rev = _collect((nm, pc) for nm, pc in TOK_REV.findall(str(s)))  # name-first
    pick = fwd if abs(sum(fwd[0].values()) - 100) <= abs(sum(rev[0].values()) - 100) else rev
    return pick

def g(r, k, i):
    try: return r[k]
    except Exception: return r[i]

def read_sheet(ws):
    """Return list of {name, mix_raw, nameplate} per facility column."""
    rows = list(ws.iter_rows(values_only=True))
    # locate label rows by column-A text
    def find(label):
        for row in rows:
            if row and row[0] and label.lower() in str(row[0]).lower():
                return row
        return None
    name_row = rows[1] if len(rows) > 1 else None         # facility names live in row 2
    mix_row, cap_row = find('Feedstock Mix'), find('Name Plate')
    out = []
    if not name_row or not mix_row: return out
    for c in range(1, len(name_row)):
        nm = name_row[c]
        if not nm or not str(nm).strip(): continue
        out.append({'name': str(nm).strip(),
                    'mix_raw': mix_row[c] if c < len(mix_row) else None,
                    'nameplate': cap_row[c] if (cap_row and c < len(cap_row)) else None})
    return out

def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--write", action="store_true")
    ap.add_argument("--show", action="store_true", help="print every facility parse"); args = ap.parse_args()
    if not XLSX.exists(): print(f"NOT FOUND: {XLSX}"); return
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    facs = []
    for sn in SHEETS:
        if sn in wb.sheetnames:
            for f in read_sheet(wb[sn]):
                f['sheet'] = sn; facs.append(f)

    # master for name-matching
    with get_connection() as c:
        cur = c.cursor()
        cur.execute("SELECT facility_id, company, facility_name, state FROM reference.biofuel_facilities")
        master = cur.fetchall()

    parsed = matched = 0; all_unmapped = {}; results = []
    for f in facs:
        mix, unmapped = parse_mix(f['mix_raw'])
        for u in unmapped: all_unmapped[u] = all_unmapped.get(u, 0) + 1
        if mix: parsed += 1
        # best master match — site-aware (keeps city tokens to disambiguate multi-site companies)
        best, bs = None, 0.0
        for m in master:
            s = site_score(f['name'], g(m,'facility_name',2))
            if s > bs: best, bs = m, s
        mid = g(best,'facility_id',0) if (best and bs >= 0.6) else None
        if mid: matched += 1
        tot = sum(mix.values())
        results.append({**f, 'mix': mix, 'unmapped': unmapped, 'mid': mid,
                        'mname': g(best,'facility_name',2) if best else None, 'mscore': round(bs,2), 'total': tot})

    print(f"=== RD Feedstock Build Up parse — {len(facs)} facility-columns across {len(SHEETS)} sheets ===")
    print(f"  parsed a mix: {parsed} | name-matched to master (>=0.6): {matched}\n")
    print("Unmapped feedstock tokens (extend NAME2CODE):")
    for tok, n in sorted(all_unmapped.items(), key=lambda x: -x[1]):
        print(f"    {tok!r:32s} x{n}")
    if not all_unmapped: print("    (none — all tokens mapped)")

    # flag mixes that don't sum to ~100
    bad = [r for r in results if r['mix'] and abs(r['total'] - 100) > 1]
    print(f"\nMixes not summing to 100% ({len(bad)}):")
    for r in bad[:12]: print(f"    {str(r['name'])[:30]:30s} sum={r['total']:.0f}  {r['mix']}  raw={str(r['mix_raw'])[:40]}")

    if args.show:
        print("\nPer-facility parse:")
        for r in sorted(results, key=lambda x: x['sheet']):
            mt = f"-> {str(r['mname'])[:26]} [{r['mscore']}]" if r['mid'] else "-> UNMATCHED"
            print(f"  [{r['sheet'][:4]}] {str(r['name'])[:28]:28s} {r['mix']}  {mt}")

    writable = [r for r in results if r['mid'] and r['mix']]
    if args.write:
        with get_connection() as c:
            cur = c.cursor()
            cur.execute("""CREATE TABLE IF NOT EXISTS reference.facility_assumed_mix (
                facility_id   integer NOT NULL,
                feedstock_code text   NOT NULL,
                pct           numeric NOT NULL,
                source        text    NOT NULL DEFAULT 'rd_buildup_xlsx',
                loaded_at     timestamptz DEFAULT now(),
                PRIMARY KEY (facility_id, feedstock_code, source))""")
            cur.execute("DELETE FROM reference.facility_assumed_mix WHERE source='rd_buildup_xlsx'")
            n = 0
            for r in writable:
                for code, pct in r['mix'].items():
                    cur.execute("""INSERT INTO reference.facility_assumed_mix (facility_id, feedstock_code, pct, source)
                                   VALUES (%s,%s,%s,'rd_buildup_xlsx')""", (r['mid'], code, pct))
                    n += 1
            c.commit()
        print(f"\n[WROTE] reference.facility_assumed_mix: {n} rows across {len(writable)} facilities (source=rd_buildup_xlsx)")
    else:
        print(f"\n[VALIDATE] no writes. {len(writable)} matched facilities with a mix ready to write. "
              f"Re-run --write to persist to reference.facility_assumed_mix.")

if __name__ == "__main__":
    main()
