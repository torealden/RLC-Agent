#!/usr/bin/env python3
"""
Create Power BI Dashboard Export

Creates a professionally formatted Excel workbook optimized for Power BI
with pre-calculated measures and analysis-ready data.

Usage:
    python scripts/create_powerbi_export.py
"""

import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference
except ImportError:
    print("Installing required packages...")
    os.system("pip install pandas openpyxl")
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference


# Cache directory
CACHE_DIR = Path(__file__).parent.parent / 'data' / 'cache'
EXPORT_DIR = Path(__file__).parent.parent / 'exports'


def load_all_cache_data():
    """Load all cached data and identify data types"""
    datasets = {}

    if not CACHE_DIR.exists():
        print(f"Cache directory not found: {CACHE_DIR}")
        return datasets

    for cache_file in CACHE_DIR.glob('*.json'):
        try:
            with open(cache_file, 'r') as f:
                data = json.load(f)

            # Handle different formats
            if isinstance(data, list):
                records = data
            elif isinstance(data, dict):
                records = data.get('data', [])
            else:
                continue

            if not records:
                continue

            # Identify dataset type
            sample = records[0] if records else {}
            fields = set(sample.keys()) if isinstance(sample, dict) else set()

            dataset_name = None
            if 'series_id' in fields and 'value' in fields:
                # Check for ethanol vs petroleum
                if any('PET' in str(r.get('series_id', '')) for r in records[:5]):
                    dataset_name = 'petroleum'
                else:
                    dataset_name = 'ethanol'
            elif 'commodity_desc' in fields or 'statisticcat_desc' in fields:
                dataset_name = 'crop_data'
            elif 'period' in fields and 'value' in fields:
                dataset_name = 'eia_data'

            if dataset_name:
                if dataset_name not in datasets:
                    datasets[dataset_name] = []
                datasets[dataset_name].extend(records)
            else:
                # Generic data
                datasets[cache_file.stem[:20]] = records

        except Exception as e:
            print(f"Error loading {cache_file}: {e}")

    return datasets


def process_eia_data(records):
    """Process EIA data into analysis-ready format"""
    df = pd.DataFrame(records)

    # Try to find and convert date column
    date_col = None
    for col in ['period', 'date', 'Date', 'PERIOD', 'report_date', 'week_ending']:
        if col in df.columns:
            date_col = col
            break

    if date_col:
        # Convert to datetime
        df['date'] = pd.to_datetime(df[date_col], errors='coerce')

        # Check if conversion was successful (has valid dates)
        valid_dates = df['date'].notna().sum()
        if valid_dates > 0:
            # Add week number and year only if we have valid dates
            df['year'] = df['date'].dt.year
            df['week'] = df['date'].dt.isocalendar().week
            df['month'] = df['date'].dt.month
            df['month_name'] = df['date'].dt.strftime('%B')
            # Sort by date
            df = df.sort_values('date')
        else:
            # No valid dates, drop the date column we created
            df = df.drop(columns=['date'])

    if 'value' in df.columns:
        df['value'] = pd.to_numeric(df['value'], errors='coerce')

    return df


def calculate_weekly_changes(df, value_col='value', date_col='date'):
    """Calculate week-over-week and year-over-year changes"""
    if date_col not in df.columns or value_col not in df.columns:
        return df

    df = df.sort_values(date_col)

    # Week-over-week change
    df['wow_change'] = df[value_col].diff()
    df['wow_change_pct'] = df[value_col].pct_change() * 100

    # 4-week rolling average
    df['rolling_4wk_avg'] = df[value_col].rolling(window=4).mean()

    # Year-over-year (52 weeks)
    df['yoy_change'] = df[value_col].diff(52)
    df['yoy_change_pct'] = df[value_col].pct_change(52) * 100

    return df


def create_summary_stats(df, value_col='value', group_col=None):
    """Create summary statistics"""
    stats = {}

    if value_col in df.columns:
        stats['current'] = df[value_col].iloc[-1] if len(df) > 0 else None
        stats['previous'] = df[value_col].iloc[-2] if len(df) > 1 else None
        stats['wow_change'] = stats['current'] - stats['previous'] if stats['current'] and stats['previous'] else None
        stats['wow_pct'] = (stats['wow_change'] / stats['previous'] * 100) if stats['previous'] else None
        stats['avg_52wk'] = df[value_col].tail(52).mean()
        stats['max_52wk'] = df[value_col].tail(52).max()
        stats['min_52wk'] = df[value_col].tail(52).min()

    return stats


def create_excel_workbook(datasets, output_path):
    """Create formatted Excel workbook for Power BI"""

    EXPORT_DIR.mkdir(exist_ok=True)

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    number_format = '#,##0.00'
    date_format = 'YYYY-MM-DD'

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    sheets_created = []

    for dataset_name, records in datasets.items():
        if not records:
            continue

        print(f"Processing: {dataset_name} ({len(records)} records)")

        # Create DataFrame
        df = pd.DataFrame(records)

        # Process based on dataset type - look for any date-like columns
        date_cols = ['period', 'date', 'Date', 'PERIOD', 'report_date', 'week_ending']
        has_date_col = any(col in df.columns for col in date_cols)

        if has_date_col:
            df = process_eia_data(records)
            if 'value' in df.columns and 'date' in df.columns:
                df = calculate_weekly_changes(df)

        # Create sheet
        sheet_name = dataset_name[:31]  # Excel limit
        ws = wb.create_sheet(title=sheet_name)
        sheets_created.append(sheet_name)

        # Convert date columns to string format for clean Excel export
        for col in df.columns:
            if df[col].dtype == 'datetime64[ns]' or 'date' in col.lower():
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
                except:
                    pass

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Format header
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Format numbers
                    if isinstance(value, float):
                        cell.number_format = number_format

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Create Summary sheet
    ws_summary = wb.create_sheet(title='Summary', index=0)

    summary_data = [
        ['HigbyBarrett Commodity Dashboard', '', ''],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M'), ''],
        ['', '', ''],
        ['Dataset', 'Records', 'Description'],
    ]

    for dataset_name, records in datasets.items():
        summary_data.append([dataset_name, len(records), f'{dataset_name} data'])

    summary_data.append(['', '', ''])
    summary_data.append(['Total Records:', sum(len(r) for r in datasets.values()), ''])

    for row in summary_data:
        ws_summary.append(row)

    # Format summary
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:C1')

    # Save
    wb.save(output_path)
    print(f"\n✓ Saved to: {output_path}")
    print(f"  Sheets created: {', '.join(sheets_created)}")

    return sheets_created


def main():
    print("="*60)
    print("CREATING POWER BI DASHBOARD EXPORT")
    print("="*60 + "\n")

    # Load data
    datasets = load_all_cache_data()

    if not datasets:
        print("No cached data found!")
        return

    print(f"Found {len(datasets)} datasets:\n")
    for name, records in datasets.items():
        print(f"  - {name}: {len(records):,} records")

    # Create output
    output_path = EXPORT_DIR / 'HigbyBarrett_Dashboard.xlsx'
    sheets = create_excel_workbook(datasets, output_path)

    print("\n" + "="*60)
    print("POWER BI IMPORT INSTRUCTIONS")
    print("="*60)
    print(f"""
1. Open Power BI Desktop

2. Click 'Get Data' → 'Excel'

3. Navigate to:
   {output_path}

4. Select all sheets and click 'Load'

5. Go to 'Model' view to create relationships

6. Create your visualizations:

   RECOMMENDED CHARTS:

   a) Line Chart - Weekly Trend
      - X-axis: date
      - Y-axis: value
      - Legend: series_id (if multiple series)

   b) Card - Current Value
      - Value: MAX(value) filtered to latest date

   c) Card - Week-over-Week Change
      - Value: wow_change_pct (last row)
      - Conditional formatting: Green if positive

   d) Area Chart - Rolling Average
      - X-axis: date
      - Y-axis: rolling_4wk_avg

   e) Table - Recent Data
      - Columns: date, value, wow_change, wow_change_pct

7. Add slicers for filtering:
   - Date range slicer
   - Series/commodity slicer (if applicable)

QUICK WINS FOR IMPRESSIVE DEMO:
- Use Card visuals with big numbers
- Add sparklines in tables
- Use conditional formatting (green/red)
- Add a title with dynamic date

""")


if __name__ == '__main__':
    main()
