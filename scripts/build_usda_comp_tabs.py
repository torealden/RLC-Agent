"""
Build / refresh `usda_comp` tabs in the oilseed balance-sheet workbooks.

Extends the wasde_comp pattern from us_soybean_complex_bal_sheets.xlsm to
every balance-sheet workbook, reading gold.psd_wasde_vintages (all countries,
all commodities, shared vintage ladder -- higher vintage_rank = more recent).

Layout per member block (mirrors the hand-built wasde_comp exactly):

    col A  row labels
    col B  USDA, current vintage, first active MY
    col C  delta vs prior vintage (formula =B-I)
    col D  RLC (formula link into the member balance-sheet tab)
    col E  USDA, current vintage, second active MY
    col F  delta (formula =E-J)
    col G  RLC
    col I  USDA, prior vintage, first active MY
    col J  USDA, prior vintage, second active MY

Write engines (ruling for the "Python vs VBA/ODBC per file type" decision in
docs/handoffs/2026-08-01_market_dashboard.md):
  * Generated country books (.xlsx outside `United States/`): openpyxl --
    same engine copy_legacy_monthly_blocks.py already uses on these files.
  * Hand-maintained US books (everything in `United States/`, .xlsx or
    .xlsm): Excel COM -- preserves VBA projects, charts and formatting that
    openpyxl would drop. Macros are force-disabled while open so
    Workbook_Open handlers (shortcut banners) don't fire.
  The us_soybean_complex book keeps its existing VBA wasde_comp and is
  skipped here.

Unit handling: PSD values are 1000 MT (area 1000 HA). The factor into book
units is NOT assumed -- it is derived per member sheet by cross-checking the
latest FINAL (closed-MY, vintage_rank 90) PSD value against the number the
sheet already carries for that MY, then snapped to a known conversion
factor. No clean snap -> the member is skipped loudly, never written wrong.

Usage:
    python scripts/build_usda_comp_tabs.py                 # all eligible books
    python scripts/build_usda_comp_tabs.py --only brazil   # filename filter
    python scripts/build_usda_comp_tabs.py --dry-run       # report, no writes
    python scripts/build_usda_comp_tabs.py --no-us         # skip COM/US books
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.services.database.db_config import get_connection  # noqa: E402

MODELS_DIR = PROJECT_ROOT / "models" / "Oilseeds"
ARCHIVE_DIR = MODELS_DIR / "Archive"

# ---------------------------------------------------------------------------
# Static maps
# ---------------------------------------------------------------------------

# Folder name -> PSD country code. PSD's own (FIPS-style) codes, verified
# against /api/psd/countries 2026-08-01. bronze.fas_psd also carries a few
# stray ISO-coded rows (CN, AU, ZA...) from an old mis-coded pull -- these
# PSD codes are the ones with real history; do not "correct" them to ISO.
COUNTRY_CODES = {
    "United States": "US",
    "Argentina": "AR",
    "Australia": "AS",
    "Brazil": "BR",
    "Canada": "CA",
    "China": "CH",
    "EU": "E4",
    "India": "IN",
    "Indonesia": "ID",
    "Japan": "JA",
    "Malaysia": "MY",
    "Mexico": "MX",
    "Paraguay": "PA",
    "Philippines": "RP",
    "Russia": "RS",
    "Ukraine": "UP",
    "Uruguay": "UY",
}

# In-sheet title fragment -> PSD commodity slug. Ordered most-specific-first;
# matching uses the SHEET TITLE (A2), never the tab name -- seven rapeseed
# country books still carry soy_* tab names from the template clone.
TITLE_TO_COMMODITY = [
    ("PALM KERNEL CAKE", "palm_kernel_meal"),
    ("PALM KERNEL MEAL", "palm_kernel_meal"),
    ("PALM KERNEL OIL", "palm_kernel_oil"),
    ("PALM KERNEL", "palm_kernel"),
    ("PALM OIL", "palm_oil"),
    ("COPRA MEAL", "copra_meal"),
    ("COPRA", "copra"),
    ("COCONUT OIL", "coconut_oil"),
    # US coconut book titles its copra (seed) sheet "US COCONUT S&D"
    ("COCONUT", "copra"),
    ("RAPESEED MEAL", "rapeseed_meal"),
    ("CANOLA MEAL", "rapeseed_meal"),
    ("RAPESEED OIL", "rapeseed_oil"),
    ("CANOLA/RAPESEED OIL", "rapeseed_oil"),
    ("CANOLA OIL", "rapeseed_oil"),
    ("RAPESEED", "rapeseed"),
    ("CANOLA", "rapeseed"),
    ("SUNFLOWERSEED MEAL", "sunflowerseed_meal"),
    ("SUNFLOWER MEAL", "sunflowerseed_meal"),
    ("SUNFLOWERSEED OIL", "sunflowerseed_oil"),
    ("SUNFLOWER OIL", "sunflowerseed_oil"),
    ("SUNFLOWERSEED", "sunflowerseed"),
    ("SUNFLOWER", "sunflowerseed"),
    ("COTTONSEED MEAL", "cottonseed_meal"),
    ("COTTONSEED OIL", "cottonseed_oil"),
    ("COTTONSEED", "cottonseed"),
    # plain COTTON (after all COTTONSEED entries): the lint balance sheet,
    # kept in million 480-lb bales exactly as the WASDE table prints it.
    # PSD 'cotton' is natively 1000 480-lb bales -- no mass conversion.
    ("COTTON", "cotton"),
    ("PEANUT MEAL", "peanut_meal"),
    ("PEANUT OIL", "peanut_oil"),
    ("PEANUT", "peanuts"),
    ("SOYBEAN MEAL", "soybean_meal"),
    ("SOYBEAN OIL", "soybean_oil"),
    ("SOYBEAN", "soybeans"),
]

SEED_COMMODITIES = {
    "soybeans", "rapeseed", "sunflowerseed", "cottonseed",
    "peanuts", "copra", "palm_kernel",
}

# Books that must not get a usda_comp tab at all.
SKIP_BOOKS = {
    # has the hand-built wasde_comp + WASDECompUpdater VBA already
    "us_soybean_complex_bal_sheets.xlsm": "existing wasde_comp (VBA)",
    # stale soybean clone, Tore is rebuilding it (handoff 2026-08-01 item 3b)
    "us_lauric_oils_bal_sheets.xlsm": "stale clone, being rebuilt",
}

# PSD does not publish these commodities at all (verified vs
# /api/psd/commodities 2026-08-01). Ruled by Tore 2026-08-01: these books
# get a NOTE-ONLY usda_comp tab saying no comp is available, rather than
# nothing (no fabricated #N/A rows in bronze -- the note lives in the book).
NO_PSD_BOOKS = {
    "us_corn_oil_balance_sheets.xlsx": "corn oil",
    "us_flaxseed_balance_sheets.xlsx": "flaxseed",
    "us_safflower_balance_sheets.xlsx": "safflower",
}

# Known 1000-MT -> book-unit factors the magnitude cross-check may snap to.
KNOWN_FACTORS = [
    (1.0, "thousand tonnes"),
    (2.204623, "million pounds"),
    (1.102311, "thousand short tons"),
    (0.0367437, "million bushels (60 lb)"),
    (0.0393683, "million bushels (56 lb)"),
    (0.001, "million tonnes"),
]
# NOTE: no bale factors here on purpose. Ruled by Tore 2026-08-02: bales
# belong to the cotton LINT sheet only (COTTON_* tables below); a bales
# label on any seed/meal/oil sheet is a mislabel and must skip loudly.

# Row-label parenthetical -> factor. Used when the annual block is still
# empty (most generated country books carry the template structure but no
# data yet), where the magnitude cross-check has nothing to bite on.
LABEL_UNIT_FACTORS = {
    "thousand tonnes": (1.0, "thousand tonnes"),
    "thousand metric tons": (1.0, "thousand tonnes"),
    "million pounds": (2.204623, "million pounds"),
    "thousand short tons": (1.102311, "thousand short tons"),
    # US-convention "tons" = short tons (Tore, 2026-08-02, cottonseed).
    # Metric sheets must say "tonnes" -- the country books all do.
    "thousand tons": (1.102311, "thousand short tons"),
    # among the oilseed books only soybeans are kept in bushels -> 60 lb
    "million bushels": (0.0367437, "million bushels (60 lb)"),
}
AREA_FACTORS = {
    # book area unit -> factor from PSD 1000 HA
    "million hectares": 0.001,
    "million acres": 0.00247105,
    "thousand hectares": 1.0,
    "thousand acres": 2.47105,
}

# Comp rows: (label, psd_field, kind) where kind drives which member types
# carry the row. psd_field None => formula row.
SEED_ROWS = [
    ("Harvested Area", "area_harvested"),
    ("Beginning Stocks", "beginning_stocks"),
    ("Production", "production"),
    ("Imports", "imports"),
    ("Total Supply", "SUM_SUPPLY"),
    ("Crush", "crush"),
    ("Exports", "exports"),
    # everything that isn't crush or exports: food, seed, feed/waste,
    # residual -- composition varies by seed (soybeans: seed+residual;
    # peanuts: mostly food; cottonseed: feed+residual)
    ("Other Domestic Use", "RESIDUAL"),
    ("Total Demand", "DEMAND"),
    ("Ending Stocks", "ending_stocks"),
    ("Stocks-to-Use", "STU"),
]
PRODUCT_ROWS = [
    ("Beginning Stocks", "beginning_stocks"),
    ("Production", "production"),
    ("Imports", "imports"),
    ("Total Supply", "SUM_SUPPLY"),
    ("Domestic Use", "domestic_consumption"),
    ("Exports", "exports"),
    ("Total Demand", "DEMAND"),
    ("Ending Stocks", "ending_stocks"),
    ("Stocks-to-Use", "STU"),
]

# Cotton (lint): WASDE-table shape -- area/yield up top, no crush.
# Total Demand here = Total Supply - Ending Stocks, which absorbs WASDE's
# "Unaccounted" line into demand (PSD carries no unaccounted column).
COTTON_ROWS = [
    ("Harvested Area", "area_harvested"),
    ("Beginning Stocks", "beginning_stocks"),
    ("Production", "production"),
    ("Imports", "imports"),
    ("Total Supply", "SUM_SUPPLY"),
    ("Domestic Use", "domestic_consumption"),
    ("Exports", "exports"),
    ("Total Demand", "DEMAND"),
    ("Ending Stocks", "ending_stocks"),
    ("Stocks-to-Use", "STU"),
]

# PSD serves cotton in 1000 480-lb bales (not 1000 MT), so its factor
# tables are bale-relative, not mass-relative.
COTTON_KNOWN_FACTORS = [
    (0.001, "million 480-lb bales"),
    (1.0, "thousand 480-lb bales"),
]
COTTON_LABEL_UNIT_FACTORS = {
    "million 480-lb bales": (0.001, "million 480-lb bales"),
    "million 480 lb bales": (0.001, "million 480-lb bales"),
    "million 480 pound bales": (0.001, "million 480-lb bales"),
    "thousand 480-lb bales": (1.0, "thousand 480-lb bales"),
    "thousand 480 lb bales": (1.0, "thousand 480-lb bales"),
}

# Comp label -> member-sheet row label patterns (lowercased, ordered by
# priority) for the RLC link columns.
RLC_LINK_PATTERNS = {
    "Harvested Area": ["harvested area"],
    "Beginning Stocks": ["beginning stocks"],
    "Production": ["production"],
    "Imports": ["imports"],
    "Total Supply": ["total supply"],
    "Crush": ["crush"],
    "Domestic Use": ["total domestic use", "domestic use", "domestic demand",
                     "domestic disappearance"],
    "Exports": ["exports"],
    "Total Demand": ["total demand", "total use"],
    "Ending Stocks": ["ending stocks"],
    "Stocks-to-Use": ["stocks-to-use", "stocks to use"],
}

MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November",
               "December"]


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def load_vintages(cur, commodity: str, country_code: str):
    """Return {my: [newest_row, prior_row_or_None]} for active MYs, plus the
    newest FINAL (closed-MY) row for the unit cross-check."""
    cur.execute(
        """
        SELECT marketing_year, report_date, psd_cycle, vintage, vintage_rank,
               area_harvested, beginning_stocks, production, imports,
               total_supply, crush, domestic_consumption, exports,
               ending_stocks
        FROM gold.psd_wasde_vintages
        WHERE commodity = %s AND country_code = %s AND is_active_my
        -- psd_cycle breaks rank ties: with the mig-168 archive union, an active
        -- MY can carry >19 cycles and everything past the 19th caps at rank 79
        ORDER BY marketing_year, vintage_rank DESC, psd_cycle DESC
        """,
        (commodity, country_code),
    )
    by_my: dict[int, list] = {}
    for row in cur.fetchall():
        by_my.setdefault(row["marketing_year"], []).append(row)
    active = {my: (rows[0], rows[1] if len(rows) > 1 else None)
              for my, rows in by_my.items()}

    cur.execute(
        """
        SELECT marketing_year, area_harvested, beginning_stocks, production,
               imports, crush, domestic_consumption, exports, ending_stocks
        FROM gold.psd_wasde_vintages
        WHERE commodity = %s AND country_code = %s
          AND NOT is_active_my AND vintage = 'FINAL'
        ORDER BY marketing_year DESC
        LIMIT 3
        """,
        (commodity, country_code),
    )
    finals = cur.fetchall()
    return active, finals


# ---------------------------------------------------------------------------
# Workbook inspection (openpyxl, read-only pass, cached values)
# ---------------------------------------------------------------------------

def inspect_book(path: Path):
    """Read member-sheet structure: title, MY->column map, label->row map,
    cached values for the unit cross-check. Also returns the existing
    usda_comp A1 sheet title when Tore has hand-set one (donor: 'ARGENTINA
    OILSEEDS COMPLEX') so a rebuild preserves it -- generated banners
    (recognizable by 'USDA (PSD/WASDE)') are NOT preserved."""
    wb_v = openpyxl.load_workbook(path, data_only=True)   # cached values
    existing_title = None
    if "usda_comp" in wb_v.sheetnames:
        v = wb_v["usda_comp"].cell(row=1, column=1).value
        if isinstance(v, str) and v.strip() and "USDA (PSD/WASDE)" not in v:
            existing_title = v.strip()
    members = []
    for name in wb_v.sheetnames:
        if not name.endswith("_balance_sheet"):
            continue
        ws = wb_v[name]
        title = str(ws.cell(row=2, column=1).value or "").upper()
        commodity = None
        for frag, slug in TITLE_TO_COMMODITY:
            if frag in title:
                commodity = slug
                break

        # MY -> column. Row 3 carries '1990/91'-style labels in every book;
        # row 2 start-year ints exist only in the populated ones (Brazil, US).
        my_col = {}
        for col in range(2, 90):
            v3 = ws.cell(row=3, column=col).value
            if isinstance(v3, str) and "/" in v3:
                head = v3.split("/")[0].strip()
                if head.isdigit() and 1980 <= int(head) <= 2060:
                    my_col[int(head)] = col
                    continue
            v2 = ws.cell(row=2, column=col).value
            if isinstance(v2, (int, float)) and 1980 <= v2 <= 2060:
                my_col[int(v2)] = col

        # annual block: rows 4.. until the first ALL-CAPS section title
        label_row: dict[str, int] = {}
        values: dict[tuple[str, int], float] = {}
        for r in range(4, 40):
            raw = ws.cell(row=r, column=1).value
            if raw is None:
                continue
            label = str(raw).strip()
            if len(label) > 10 and label == label.upper():
                break  # monthly-section banner => annual block ended
            key = label.lower()
            if key not in label_row:
                label_row[key] = r
            for my, col in my_col.items():
                v = ws.cell(row=r, column=col).value
                if isinstance(v, (int, float)):
                    values[(key, my)] = float(v)

        members.append({
            "tab": name,
            "title": title,
            "commodity": commodity,
            "my_col": my_col,
            "label_row": label_row,
            "values": values,
        })
    wb_v.close()
    return members, existing_title


def find_label_row(label_row: dict, patterns: list[str]):
    for pat in patterns:
        for key, r in label_row.items():
            base = key.split("(")[0].strip()
            if base == pat or base.startswith(pat):
                return r
    return None


def derive_unit_factor(member, finals, known_factors=None):
    """Snap sheet_value / psd_value to known conversion factors. Returns
    {factor: (psd_v, unit_name, evidence)} — the strongest snap PER FACTOR,
    because different fields can snap to different factors when the sheet's
    source differs from PSD (india soy oil 2026-08-02: production tied PSD
    in tonnes to 0.2%% while domestic use sat 2.5%% off the short-tons
    factor — a source difference, not a unit difference). The caller must
    treat conflicting snaps as ambiguity, not take the largest blindly."""
    known_factors = known_factors or KNOWN_FACTORS
    fields = [(["production"], "production"), (["imports"], "imports"),
              (["exports"], "exports"),
              (["total domestic use", "domestic use"], "domestic_consumption"),
              (["ending stocks"], "ending_stocks")]
    snaps = {}  # factor -> (psd_v, unit_name, evidence)
    for final in finals:
        my = final["marketing_year"]
        if my not in member["my_col"]:
            continue
        for sheet_patterns, psd_field in fields:
            psd_v = final.get(psd_field)
            if psd_v is None or abs(float(psd_v)) < 50:
                continue  # too small to discriminate factors
            r = find_label_row(member["label_row"], sheet_patterns)
            if r is None:
                continue
            key = [k for k in member["label_row"] if member["label_row"][k] == r][0]
            sheet_v = member["values"].get((key, my))
            if sheet_v is None or sheet_v == 0:
                continue
            ratio = sheet_v / float(psd_v)
            for factor, unit_name in known_factors:
                if abs(ratio / factor - 1.0) < 0.05:
                    if factor not in snaps or float(psd_v) > snaps[factor][0]:
                        snaps[factor] = (float(psd_v), unit_name,
                                         (sheet_patterns[0], my, sheet_v,
                                          float(psd_v)))
    return snaps


def unit_factor_from_labels(member, label_factors=None):
    """Read the unit out of the row-label parentheticals, preferring the
    stock/production rows ('Beginning Stocks (thousand tonnes)')."""
    label_factors = label_factors or LABEL_UNIT_FACTORS
    preferred = ["beginning stocks", "production", "ending stocks"]
    keys = sorted(member["label_row"],
                  key=lambda k: next((i for i, p in enumerate(preferred)
                                      if k.startswith(p)), 99))
    for key in keys:
        if "(" not in key:
            continue
        unit_txt = key.split("(", 1)[1].rstrip(")").strip()
        if unit_txt in label_factors:
            return label_factors[unit_txt]
    return None, None


def derive_area_factor(member):
    """Area factor from the label text of the Harvested/Planted Area row."""
    for key in member["label_row"]:
        if "area" in key and "(" in key:
            unit_txt = key.split("(", 1)[1].rstrip(")").strip()
            if unit_txt in AREA_FACTORS:
                return AREA_FACTORS[unit_txt]
    # generated country books label area "(million hectares)" on the
    # Planted Area row only; Harvested Area inherits it
    return None


# ---------------------------------------------------------------------------
# Comp-tab content assembly (engine-neutral styled-cell map)
# ---------------------------------------------------------------------------
# Styling replicates the donor tab Tore hand-formatted 2026-08-03 in
# argentina_soybean_complex_bal_sheets.xlsm/usda_comp: Aptos Display
# throughout, internal-green MY headers (#3C7D22, white text) merged across
# B:D / E:G / I:J, every 2nd data row banded light gray, medium box outline
# with thin column separators, centered values with red-parenthesis
# negatives, notes in Aptos Narrow 8. Column widths follow the donor
# exactly: A ~40.7, B ~12.7, C:J left at the Excel default (Tore's call --
# long "Δ from September"-style headers will clip; flagged in handoff).

GREEN = "3C7D22"     # internal green (reference_excel_color_conventions)
BAND = "D0CECE"      # donor band fill: Background 2, darker 10%
GRAY = "757171"      # donor unit-line text: Background 2, darker 50%
WHITE = "FFFFFF"
F_MAIN = "Aptos Display"   # donor A1 says "Bierstadt Display" = Aptos' old name
F_NOTE = "Aptos Narrow"

NF_INT = "#,##0_);[Red](#,##0)"
NF_1DP = "#,##0.0_);[Red](#,##0.0)"
NF_AREA = "#,##0.000_);[Red](#,##0.000)"
NF_PCT = "0.0%"

# data-grid columns -> (left, right) border weight
VCOLS = {2: ("medium", "thin"), 3: ("thin", "thin"), 4: ("thin", "thin"),
         5: ("thin", "thin"), 6: ("thin", "thin"), 7: ("thin", "medium"),
         9: ("medium", "thin"), 10: ("thin", "medium")}


def _cell(cells, r, c, v=None, f=False, nf=None, st=None):
    """Merge a value / number-format / style into the (r, c) cell record."""
    cur = cells.setdefault((r, c), {"v": None, "f": False, "nf": None, "st": {}})
    if v is not None:
        cur["v"], cur["f"] = v, f
    if nf:
        cur["nf"] = nf
    if st:
        border = {**cur["st"].get("border", {}), **st.get("border", {})}
        cur["st"].update(st)
        if border:
            cur["st"]["border"] = border


def build_block(member, active, finals, country_name, start_row):
    """Return (cells, merges, n_rows, notes). cells = {(row, col): record},
    merges = [(row, col_first, col_last)]."""
    commodity = member["commodity"]
    is_cotton = commodity == "cotton"
    is_seed = commodity in SEED_COMMODITIES
    rows_spec = (COTTON_ROWS if is_cotton
                 else SEED_ROWS if is_seed else PRODUCT_ROWS)
    known_factors = COTTON_KNOWN_FACTORS if is_cotton else None
    label_factors = COTTON_LABEL_UNIT_FACTORS if is_cotton else None

    mys = sorted(active.keys())[:2]
    if len(mys) < 2:
        return None, None, 0, [f"{commodity}: fewer than 2 active MYs"]
    my1, my2 = mys

    # Unit factor: the row-label unit is authoritative when present (it is
    # part of the template). The magnitude cross-check can only overrule it
    # on STRONG evidence (a snapped field with PSD value >= 500) that is
    # also UNANIMOUS: if any field snaps in agreement with the label, a
    # conflicting snap on another field is a source difference, not a unit
    # difference (india soy oil: production tied the label's tonnes exactly
    # while domestic use coincidentally sat near the short-tons factor).
    # Small denominators can't tell a ~10% source difference from the
    # short-tons factor. No label and no snap -> loud skip, never a guess.
    label_factor, label_unit = unit_factor_from_labels(member, label_factors)
    snaps = derive_unit_factor(member, finals, known_factors)
    notes = []
    evidence = None
    if label_factor is not None:
        agree = [(pv, u, ev) for f, (pv, u, ev) in snaps.items()
                 if abs(label_factor / f - 1.0) <= 0.05]
        conflict = [(pv, u, ev) for f, (pv, u, ev) in snaps.items()
                    if abs(label_factor / f - 1.0) > 0.05]
        factor, unit_name = label_factor, label_unit
        if agree:
            pv, u, ev = max(agree)
            evidence = ev  # label confirmed by values
            if conflict:
                cpv, cu, cev = max(conflict)
                notes.append(f"{commodity}: mixed snaps -- label {label_unit} "
                             f"confirmed by {ev[0]} (psd={pv:,.0f}); "
                             f"conflicting {cu} snap on {cev[0]} "
                             f"(psd={cpv:,.0f}) treated as source "
                             f"difference, label kept")
        elif conflict:
            cpv, cu, cev = max(conflict)
            if cpv >= 500:
                return None, None, 0, [f"{commodity}: label says {label_unit} "
                                       f"but values strongly snap to {cu} "
                                       f"(psd={cpv:,.0f}) -- skipped"]
            notes.append(f"{commodity}: weak snap to {cu} "
                         f"(psd={cpv:,.0f}) ignored; label "
                         f"{label_unit} used")
    elif snaps:
        # No label: values may only decide the unit when they agree with
        # each other AND the evidence is strong.
        if len(snaps) > 1:
            return None, None, 0, [f"{commodity}: no unit label and values "
                                   f"snap to {len(snaps)} different factors "
                                   f"-- ambiguous, skipped"]
        (factor, (strength, unit_name, ev)), = snaps.items()
        if strength < 500:
            return None, None, 0, [f"{commodity}: no recognized unit label "
                                   f"and no strong value snap -- skipped, "
                                   f"never guessed"]
        evidence = ev
    else:
        # No recognized label AND no snap at all. A weak label-less snap
        # is not enough: mid-conversion sheets (cottonseed 2026-08-02)
        # can coincidentally snap on one small field.
        return None, None, 0, [f"{commodity}: no recognized unit label and "
                               f"no strong value snap -- skipped, never "
                               f"guessed"]
    area_factor = derive_area_factor(member) or 0.001

    cur1, prior1 = active[my1]
    cur2, prior2 = active[my2]

    def conv(rowdict, field):
        if rowdict is None:
            return None
        v = rowdict.get(field)
        if v is None:
            return None
        f = area_factor if field == "area_harvested" else factor
        return round(float(v) * f, 2)

    tab = member["tab"]

    cells: dict = {}
    merges = []
    r0 = start_row
    my_label = lambda my: f"{my}/{str(my + 1)[-2:]}"
    short = lambda my: f"{str(my)[-2:]}/{str(my + 1)[-2:]}"

    # Month labels come from psd_cycle (the WASDE cycle the values belong to),
    # NOT report_date (the pull that happened to carry them) — mig 166.
    cur_month = MONTH_NAMES[cur1["psd_cycle"].month]
    cur_year = cur1["psd_cycle"].year
    prior_month = (MONTH_NAMES[prior1["psd_cycle"].month]
                   if prior1 is not None else "")

    fmt_val = NF_INT if factor >= 1 else NF_1DP

    has_area = any(f == "area_harvested" for _, f in rows_spec)
    area_unit_name = next((k for k, v in AREA_FACTORS.items()
                           if v == area_factor), "million hectares")
    unit_text = (f"({area_unit_name}, {unit_name})" if has_area
                 else f"({unit_name})")

    # Block title: the member sheet's own title, as in the donor.
    _cell(cells, r0, 1, member["title"],
          st={"font": F_MAIN, "bold": True})

    # MY header row: green merged headers B:D / E:G, prior month I:J.
    r1 = r0 + 1
    my_hdr = {"font": F_MAIN, "bold": True, "center": True,
              "fill": GREEN, "color": WHITE}
    _cell(cells, r1, 2, my_label(my1), st={**my_hdr, "border": {"left": "medium"}})
    _cell(cells, r1, 4, st={"border": {"right": "thin"}})
    _cell(cells, r1, 5, my_label(my2), st={**my_hdr, "border": {"left": "thin"}})
    _cell(cells, r1, 7, st={"border": {"right": "thin"}})
    _cell(cells, r1, 9, prior_month or " ",
          st={**my_hdr, "border": {"top": "medium", "left": "medium"}})
    _cell(cells, r1, 10, st={"border": {"top": "medium", "right": "medium"}})
    merges += [(r1, 2, 4), (r1, 5, 7), (r1, 9, 10)]

    # Column header row + unit line. Delta headers use the ABBREVIATED month
    # ("Δ from Sep") so the longest months don't clip at grid width — the
    # merged I:J prior-month header keeps the full name like the donor.
    hdr = r0 + 2
    delta_txt = f"Δ from {prior_month[:3]}" if prior_month else "Δ (no prior)"
    _cell(cells, hdr, 1, unit_text,
          st={"font": F_MAIN, "size": 8, "color": GRAY})
    col_hdr = {"font": F_MAIN, "bold": True, "center": True}
    for col, txt in [(2, "USDA"), (3, delta_txt), (4, "RLC"),
                     (5, "USDA"), (6, delta_txt), (7, "RLC")]:
        left, right = VCOLS[col]
        _cell(cells, hdr, col, txt,
              st={**col_hdr, "border": {"left": left, "right": right}})
    _cell(cells, hdr, 9, short(my1),
          st={**col_hdr, "border": {"left": "medium", "right": "thin",
                                    "bottom": "medium"}})
    _cell(cells, hdr, 10, short(my2),
          st={**col_hdr, "border": {"left": "thin", "right": "medium",
                                    "bottom": "medium"}})

    data0 = hdr + 1
    row_of = {}
    for i, (label, _f) in enumerate(rows_spec):
        row_of[label] = data0 + i

    for i, (label, field) in enumerate(rows_spec):
        r = row_of[label]
        band = i % 2 == 1
        fill = BAND if band else None
        fmt = (NF_PCT if field == "STU"
               else NF_AREA if field == "area_harvested" else fmt_val)

        # Row label (units live in the unit line, per the donor) + the
        # styled grid: every B..G / I..J cell gets border/fill/format even
        # when empty, so banding and the box stay continuous.
        _cell(cells, r, 1, label,
              st={"font": F_MAIN, "size": 10, "fill": fill})
        for col, (left, right) in VCOLS.items():
            border = {"left": left, "right": right}
            if i == 0:
                border["top"] = "medium"
            _cell(cells, r, col, nf=fmt,
                  st={"font": F_MAIN, "size": 9, "center": True,
                      "fill": fill, "border": border})

        # RLC link columns D/G
        link_row = find_label_row(member["label_row"],
                                  RLC_LINK_PATTERNS.get(label, []))
        if link_row is not None:
            if my1 in member["my_col"]:
                c1 = get_column_letter(member["my_col"][my1])
                _cell(cells, r, 4, f"='{tab}'!{c1}{link_row}", f=True)
            if my2 in member["my_col"]:
                c2 = get_column_letter(member["my_col"][my2])
                _cell(cells, r, 7, f"='{tab}'!{c2}{link_row}", f=True)

        # formula rows land in B/E (current vintage) and, when a prior
        # vintage exists, in I/J too -- the hand-built wasde_comp carries
        # =I17/I16 etc. in the prior columns the same way
        formula_cols = [2, 5]
        if prior1 is not None:
            formula_cols.append(9)
        if prior2 is not None:
            formula_cols.append(10)

        if field == "SUM_SUPPLY":
            top, bot = row_of["Beginning Stocks"], r - 1
            for col in formula_cols:
                L = get_column_letter(col)
                _cell(cells, r, col, f"=SUM({L}{top}:{L}{bot})", f=True)
        elif field == "RESIDUAL":
            for col in formula_cols:
                L = get_column_letter(col)
                _cell(cells, r, col,
                      f"={L}{row_of['Total Demand']}-{L}{row_of['Crush']}-{L}{row_of['Exports']}",
                      f=True)
        elif field == "DEMAND":
            for col in formula_cols:
                L = get_column_letter(col)
                _cell(cells, r, col,
                      f"={L}{row_of['Total Supply']}-{L}{row_of['Ending Stocks']}",
                      f=True)
        elif field == "STU":
            for col in formula_cols:
                L = get_column_letter(col)
                _cell(cells, r, col,
                      f"={L}{row_of['Ending Stocks']}/{L}{row_of['Total Demand']}",
                      f=True)
        else:
            v1, v2 = conv(cur1, field), conv(cur2, field)
            p1, p2 = conv(prior1, field), conv(prior2, field)
            if v1 is not None:
                _cell(cells, r, 2, v1)
            if v2 is not None:
                _cell(cells, r, 5, v2)
            if p1 is not None:
                _cell(cells, r, 9, p1)
            if p2 is not None:
                _cell(cells, r, 10, p2)

        # deltas only when a prior vintage exists
        if prior1 is not None:
            _cell(cells, r, 3, f"=B{r}-I{r}", f=True)
        if prior2 is not None:
            _cell(cells, r, 6, f"=E{r}-J{r}", f=True)

    # closing edge under the last data row
    r_close = data0 + len(rows_spec)
    for col in VCOLS:
        _cell(cells, r_close, col, st={"border": {"top": "medium"}})

    note_r = r_close + 1
    vtag = cur1["vintage"]
    note = (f"USDA columns: {vtag} ({cur_month} {cur_year} cycle, "
            f"pulled {cur1['report_date']:%Y-%m-%d}) from "
            f"gold.psd_wasde_vintages"
            + (f"; Δ vs {prior1['vintage']}" if prior1 is not None
               else "; no prior vintage yet -- Δ columns blank")
            + f". RLC columns link to {tab}. "
            + (f"Unit check: {evidence[0]} MY{evidence[1]} sheet "
               f"{evidence[2]:,.0f} vs PSD {evidence[3]:,.0f} x{factor:g}."
               if evidence is not None else
               f"Unit from row labels ({unit_name}); sheet has no values "
               f"yet to cross-check."))
    _cell(cells, note_r, 1, note, st={"font": F_NOTE, "size": 8})

    n_rows = (note_r - r0) + 3
    return cells, merges, n_rows, notes


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def write_openpyxl(path: Path, sheet, dry_run: bool):
    wb = openpyxl.load_workbook(path)
    if "usda_comp" in wb.sheetnames:
        del wb["usda_comp"]
    ws = wb.create_sheet("usda_comp", 0)
    ws.column_dimensions["A"].width = 40.71
    for col in "BCDEFGIJ":   # grid columns uniform; H spacer stays default
        ws.column_dimensions[col].width = 12.71
    ws.row_dimensions[1].height = 21
    white = PatternFill("solid", fgColor=WHITE)
    for row in ws.iter_rows(min_row=1, max_row=sheet["max_row"],
                            min_col=1, max_col=11):
        for c in row:
            c.fill = white
    for (r, col), rec in sheet["cells"].items():
        cell = ws.cell(row=r, column=col)
        if rec["v"] is not None:
            cell.value = rec["v"]
        if rec["nf"]:
            cell.number_format = rec["nf"]
        st = rec["st"]
        cell.font = Font(name=st.get("font", F_MAIN),
                         size=st.get("size", 11),
                         bold=st.get("bold", False),
                         color=st.get("color") or "FF000000")
        if st.get("fill"):
            cell.fill = PatternFill("solid", fgColor=st["fill"])
        if st.get("center"):
            cell.alignment = Alignment(horizontal="center")
        if st.get("border"):
            cell.border = Border(**{edge: Side(style=w)
                                    for edge, w in st["border"].items()})
    for (r, c1, c2) in sheet["merges"]:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    if not dry_run:
        wb.save(path)
    wb.close()


XL_EDGE = {"left": 7, "top": 8, "bottom": 9, "right": 10}
XL_WEIGHT = {"thin": 2, "medium": -4138}  # xlThin, xlMedium


def _bgr(hexrgb: str) -> int:
    """'RRGGBB' -> the BGR long Excel COM color properties expect."""
    r, g, b = (int(hexrgb[i:i + 2], 16) for i in (0, 2, 4))
    return r + g * 256 + b * 65536


def write_com(path: Path, sheet, dry_run: bool):
    if dry_run:
        return
    import pythoncom
    import win32com.client as win32
    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    # 3 = msoAutomationSecurityForceDisable: don't run Workbook_Open macros
    excel.AutomationSecurity = 3
    try:
        wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0)
        for sh in list(wb.Sheets):
            if sh.Name == "usda_comp":
                sh.Delete()
        ws = wb.Sheets.Add(Before=wb.Sheets(1))
        ws.Name = "usda_comp"
        ws.Columns("A").ColumnWidth = 40.71
        ws.Columns("B:G").ColumnWidth = 12.71   # grid uniform; H spacer default
        ws.Columns("I:J").ColumnWidth = 12.71
        ws.Rows(1).RowHeight = 21
        ws.Range(ws.Cells(1, 1),
                 ws.Cells(sheet["max_row"], 11)).Interior.Color = _bgr(WHITE)
        for (r, col), rec in sorted(sheet["cells"].items()):
            cell = ws.Cells(r, col)
            if rec["v"] is not None:
                if rec["f"]:
                    cell.Formula = rec["v"]
                else:
                    cell.Value = rec["v"]
            if rec["nf"]:
                cell.NumberFormat = rec["nf"]
            st = rec["st"]
            cell.Font.Name = st.get("font", F_MAIN)
            cell.Font.Size = st.get("size", 11)
            if st.get("bold"):
                cell.Font.Bold = True
            if st.get("color"):
                cell.Font.Color = _bgr(st["color"])
            if st.get("fill"):
                cell.Interior.Color = _bgr(st["fill"])
            if st.get("center"):
                cell.HorizontalAlignment = -4108  # xlCenter
            for edge, w in (st.get("border") or {}).items():
                b = cell.Borders(XL_EDGE[edge])
                b.LineStyle = 1  # xlContinuous
                b.Weight = XL_WEIGHT[w]
        for (r, c1, c2) in sheet["merges"]:
            ws.Range(ws.Cells(r, c1), ws.Cells(r, c2)).Merge()
        wb.Save()
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def default_book_title(path: Path) -> str:
    """'argentina_soybean_complex_bal_sheets' -> 'ARGENTINA SOYBEAN COMPLEX'
    (country display name from the folder, descriptor from the filename)."""
    stem = path.stem
    for suf in ("_bal_sheets", "_balance_sheets", "_bal_sheet",
                "_balance_sheet"):
        if stem.endswith(suf):
            stem = stem[: -len(suf)]
            break
    descriptor = " ".join(stem.split("_")[1:]).upper()
    return f"{path.parent.name.upper()} {descriptor}".strip()


def process_book(path: Path, cur, dry_run: bool, engine: str = "com"):
    country_folder = path.parent.name
    code = COUNTRY_CODES.get(country_folder)
    if code is None:
        return f"SKIP {path.name}: unknown country folder {country_folder!r}"

    members, existing_title = inspect_book(path)
    cells, merges, notes, built = {}, [], [], 0
    next_row = 2   # A1 carries the sheet title; first block right below it
    for m in members:
        if m["commodity"] is None:
            notes.append(f"  {m['tab']}: unrecognized title {m['title'][:50]!r}")
            continue
        active, finals = load_vintages(cur, m["commodity"], code)
        if not active:
            notes.append(f"  {m['tab']}: no PSD data for "
                         f"({m['commodity']}, {code})")
            continue
        bcells, bmerges, n_rows, errs = build_block(m, active, finals,
                                                    country_folder, next_row)
        if bcells is None:
            notes.append("  " + "; ".join(errs))
            continue
        for w in errs:
            notes.append("  " + w)
        cells.update(bcells)
        merges.extend(bmerges)
        next_row += n_rows
        built += 1

    if built == 0:
        return (f"SKIP {path.name}: no member with usable PSD data\n"
                + "\n".join(notes))

    # Sheet title (preserved if hand-set) + the single refresh stamp line.
    # Both this builder and USDACompUpdater.bas rewrite the stamp on every
    # touch (feedback_timestamp_every_touch).
    _cell(cells, 1, 1, existing_title or default_book_title(path),
          st={"font": F_MAIN, "size": 16, "bold": True})
    stamp_row = next_row
    _cell(cells, stamp_row, 1,
          f"USDA (PSD/WASDE) vs RLC comparison — refreshed "
          f"{datetime.now():%Y-%m-%d %H:%M} by "
          f"scripts/build_usda_comp_tabs.py — source gold.psd_wasde_vintages",
          st={"font": F_NOTE, "size": 8})
    sheet = {"cells": cells, "merges": merges, "max_row": stamp_row + 2}

    if not dry_run:
        ARCHIVE_DIR.mkdir(exist_ok=True)
        bak = ARCHIVE_DIR / f"{path.name}.bak_{datetime.now():%Y%m%d_%H%M%S}"
        shutil.copy2(path, bak)

    # Ruled by Tore 2026-08-01: ALL books are (or will be) hand-maintained,
    # so COM is the default everywhere -- openpyxl re-saves would drop
    # charts/objects added by hand. openpyxl remains as an explicit
    # escape hatch (--engine openpyxl) for headless/no-Excel environments,
    # safe only on books that are still generated shells.
    if engine == "openpyxl" and path.suffix.lower() != ".xlsm":
        write_openpyxl(path, sheet, dry_run)
    else:
        write_com(path, sheet, dry_run)

    tag = "DRY-RUN " if dry_run else ""
    out = f"{tag}OK   {path.name}: {built} member blocks"
    if notes:
        out += "\n" + "\n".join(notes)
    return out


def write_note_tab(path: Path, commodity_name: str, dry_run: bool):
    """usda_comp tab containing only an explanatory note, for books whose
    commodity PSD does not publish."""
    cells: dict = {}
    _cell(cells, 1, 1, default_book_title(path),
          st={"font": F_MAIN, "size": 16, "bold": True})
    _cell(cells, 3, 1,
          f"No USDA comparison available: PSD does not publish "
          f"{commodity_name} (verified against /api/psd/commodities "
          f"2026-08-01). This tab is a placeholder so the absence is "
          f"deliberate, not an oversight.",
          st={"font": F_NOTE, "size": 8})
    _cell(cells, 5, 1,
          f"USDA (PSD/WASDE) vs RLC comparison — refreshed "
          f"{datetime.now():%Y-%m-%d %H:%M} by "
          f"scripts/build_usda_comp_tabs.py",
          st={"font": F_NOTE, "size": 8})
    sheet = {"cells": cells, "merges": [], "max_row": 7}
    if not dry_run:
        ARCHIVE_DIR.mkdir(exist_ok=True)
        bak = ARCHIVE_DIR / f"{path.name}.bak_{datetime.now():%Y%m%d_%H%M%S}"
        shutil.copy2(path, bak)
    write_com(path, sheet, dry_run)
    return f"{'DRY-RUN ' if dry_run else ''}OK   {path.name}: note-only tab (no PSD coverage)"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", help="substring filter on filename")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-us", action="store_true",
                    help="skip hand-maintained US books (COM writer)")
    ap.add_argument("--engine", choices=["com", "openpyxl"], default="com",
                    help="com (default, safe on hand-maintained books) or "
                         "openpyxl (headless; generated .xlsx shells only)")
    args = ap.parse_args()

    books = sorted(
        p for p in MODELS_DIR.glob("*/*_bal*_sheets.xls[xm]")
        if p.parent.name != "Archive" and not p.name.startswith("~$")
    )

    with get_connection() as conn:
        cur = conn.cursor()
        for path in books:
            if args.only and args.only.lower() not in path.name.lower():
                continue
            if path.name in SKIP_BOOKS:
                print(f"SKIP {path.name}: {SKIP_BOOKS[path.name]}")
                continue
            if args.no_us and path.parent.name == "United States":
                print(f"SKIP {path.name}: --no-us")
                continue
            try:
                if path.name in NO_PSD_BOOKS:
                    print(write_note_tab(path, NO_PSD_BOOKS[path.name],
                                         args.dry_run))
                else:
                    print(process_book(path, cur, args.dry_run, args.engine))
            except Exception as e:  # noqa: BLE001
                print(f"FAIL {path.name}: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
