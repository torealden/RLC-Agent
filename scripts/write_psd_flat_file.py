"""PSD -> flat file, in the frozen flat-file contract schema (docs/specs/flat_file_contract_v1.md).

Parameterized annual-grain writer: given (commodity, country_code) it pulls the LATEST PSD vintage
per marketing year from bronze.fas_psd and emits a contract-shaped flat file:
  <commodity>_supply / _demand   (long/tidy tabs — the MAXIFS/SUMIFS target)
  <commodity>_supply_wide / _demand_wide + _wide_index   (display + anchoring mirror)
  _meta                          (provenance)

This is the Tuesday backbone: annual S&D for every Tier-A cell, from data already in the DB. Monthly
national-source rows upgrade the SAME sheets later via the vintage ladder (no rework). Fundamentals
only — no price layer.

Run:  python scripts/write_psd_flat_file.py soybean_oil BR
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

ROOT = Path(r"C:/dev/RLC-Agent")
sys.path.insert(0, str(ROOT))
from src.services.database.db_config import get_connection  # noqa: E402
OILSEEDS = ROOT / "models" / "Oilseeds"
PSD_RANK = 70  # WASDE/PSD band 61-90; any national/actual row outranks it (contract sec 5)

CODE_TO_FOLDER = {
    "US": "United States", "BR": "Brazil", "AR": "Argentina", "CA": "Canada",
    "E4": "Europe", "EU": "Europe", "AS": "Australia", "AU": "Australia",
    "UP": "Ukraine", "UA": "Ukraine", "RS": "Russia", "RU": "Russia",
    "MY": "Malaysia", "ID": "Indonesia", "CH": "China", "CN": "China",
    "IN": "India", "TU": "Turkey", "TR": "Turkey", "MX": "Mexico",
}

# PSD numeric column -> contract series name. Same map serves oils, meals, seeds; a column that is
# null/absent for a commodity simply yields no rows for that series.
SERIES_MAP = {
    "beginning_stocks": "beginning_stocks",
    "production": "production",
    "imports": "imports",
    "crush": "crush",
    "feed_dom_consumption": "feed_use",
    "fsi_consumption": "fsi_use",
    "domestic_consumption": "domestic_use",
    "exports": "exports",
    "ending_stocks": "ending_stocks",
}
SUPPLY_SERIES = {"beginning_stocks", "production", "imports", "ending_stocks"}
SERIES_ORDER = ["beginning_stocks", "production", "imports", "crush",
                "feed_use", "fsi_use", "domestic_use", "exports", "ending_stocks"]
LONG_COLS = ["commodity", "class", "series", "marketing_year", "period_type", "period",
             "vintage", "vintage_rank", "value", "unit", "source"]


def fetch(commodity, code):
    """Latest PSD vintage per marketing_year for (commodity, country_code)."""
    sql = """
        SELECT DISTINCT ON (marketing_year)
            marketing_year, report_date, unit,
            beginning_stocks, production, imports, crush,
            feed_dom_consumption, fsi_consumption, domestic_consumption, exports, ending_stocks
        FROM bronze.fas_psd
        WHERE commodity = %s AND country_code = %s
        ORDER BY marketing_year, report_date DESC NULLS LAST
    """
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (commodity, code))
            data = cur.fetchall()
            if data and isinstance(data[0], dict):     # RealDictCursor
                return [dict(r) for r in data]
            cols = [d[0] for d in cur.description]       # tuple cursor
            return [dict(zip(cols, r)) for r in data]


def to_long(commodity, rows):
    """Emit contract long rows. One row per (series, marketing_year), annual grain."""
    out = []
    for r in rows:
        my = r["marketing_year"]
        unit = r["unit"] or "1000 MT"
        for col, series in SERIES_MAP.items():
            v = r.get(col)
            if v is None:
                continue
            out.append(dict(commodity=commodity, **{"class": "ALL"}, series=series,
                            marketing_year=my, period_type="annual", period="ANNUAL",
                            vintage="PSD", vintage_rank=PSD_RANK, value=float(v),
                            unit=unit, source="USDA_FAS_PSD"))
    return out


def write_long_tab(ws, rows):
    for j, h in enumerate(LONG_COLS, 1):
        c = ws.cell(1, j, h); c.font = Font(bold=True)
    for i, row in enumerate(rows, 2):
        for j, h in enumerate(LONG_COLS, 1):
            ws.cell(i, j, row["class"] if h == "class" else row[h])


def write_wide(ws, rows, commodity, side_series):
    """Annual wide: one block per series — title row, MY header row, single ANNUAL data row.
    Returns index entries (tab, series, title_row, header_row, first_month_row, last_month_row,
    total_row, first_my_col, first_my, last_my)."""
    mys = sorted({r["marketing_year"] for r in rows})
    unit = rows[0]["unit"] if rows else "1000 MT"
    by_series = {}
    for r in rows:
        by_series.setdefault(r["series"], {})[r["marketing_year"]] = r["value"]
    index = []
    rr = 1
    for series in [s for s in SERIES_ORDER if s in side_series and s in by_series]:
        title_row = rr
        ws.cell(rr, 1, f"{commodity.upper().replace('_',' ')} {series.upper().replace('_',' ')}")
        ws.cell(rr, 1).font = Font(bold=True)
        rr += 1
        header_row = rr
        ws.cell(rr, 1, f"({unit})")
        for j, my in enumerate(mys, 2):
            ws.cell(rr, j, f"{my}/{str(my+1)[-2:]}")
        rr += 1
        data_row = rr
        ws.cell(rr, 1, "Annual")
        for j, my in enumerate(mys, 2):
            ws.cell(rr, j, by_series[series].get(my))
        rr += 2  # blank spacer
        index.append((ws.title, series, title_row, header_row, data_row, data_row, data_row,
                      "B", mys[0] if mys else None, mys[-1] if mys else None))
    return index


def build(commodity, code):
    folder = CODE_TO_FOLDER.get(code)
    if not folder:
        raise SystemExit(f"Unknown country_code {code!r} — add to CODE_TO_FOLDER")
    rows = fetch(commodity, code)
    if not rows:
        raise SystemExit(f"No PSD rows for {commodity} / {code}")
    longrows = to_long(commodity, rows)
    supply = [r for r in longrows if r["series"] in SUPPLY_SERIES]
    demand = [r for r in longrows if r["series"] not in SUPPLY_SERIES]

    wb = openpyxl.Workbook()
    ws_s = wb.active; ws_s.title = f"{commodity}_supply"
    ws_d = wb.create_sheet(f"{commodity}_demand")
    write_long_tab(ws_s, supply)
    write_long_tab(ws_d, demand)

    ws_sw = wb.create_sheet(f"{commodity}_supply_wide")
    ws_dw = wb.create_sheet(f"{commodity}_demand_wide")
    idx = write_wide(ws_sw, supply, commodity, SUPPLY_SERIES)
    idx += write_wide(ws_dw, demand, commodity, set(SERIES_MAP.values()) - SUPPLY_SERIES)

    ws_i = wb.create_sheet("_wide_index")
    ihdr = ["tab", "series", "title_row", "header_row", "first_month_row", "last_month_row",
            "total_row", "first_my_col", "first_my", "last_my"]
    for j, h in enumerate(ihdr, 1):
        ws_i.cell(1, j, h).font = Font(bold=True)
    for i, e in enumerate(idx, 2):
        for j, v in enumerate(e, 1):
            ws_i.cell(i, j, v)

    ws_m = wb.create_sheet("_meta")
    for j, h in enumerate(["series", "source", "unit", "last_updated", "notes"], 1):
        ws_m.cell(1, j, h).font = Font(bold=True)
    latest = max((r["report_date"] for r in rows if r["report_date"]), default=None)
    for i, series in enumerate([s for s in SERIES_ORDER if any(x["series"] == s for x in longrows)], 2):
        ws_m.cell(i, 1, series); ws_m.cell(i, 2, "USDA_FAS_PSD")
        ws_m.cell(i, 3, rows[0]["unit"] or "1000 MT")
        ws_m.cell(i, 4, str(latest) if latest else None)
        ws_m.cell(i, 5, f"PSD latest vintage (rank {PSD_RANK}); annual grain — monthly national "
                        "sources upgrade via the ladder")

    out = OILSEEDS / folder / f"{folder.lower().replace(' ','_')}_{commodity}_supply_demand.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out, len(longrows), sorted({r["marketing_year"] for r in rows})


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: python scripts/write_psd_flat_file.py <commodity> <country_code>")
    commodity, code = sys.argv[1], sys.argv[2]
    out, n, mys = build(commodity, code)
    print(f"Wrote {out}")
    print(f"  {n} long rows · MY {mys[0]}..{mys[-1]} ({len(mys)} years)")


if __name__ == "__main__":
    main()
