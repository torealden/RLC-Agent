"""Load Abiove Brazilian soy-complex data into bronze.abiove_soy_complex.

Parses the operator-extracted Power BI workbook
(data/raw/oilseeds_fats_greases/brazil_crushing_data.xlsx) — four tabs:

  Tabela                 monthly soybean CRUSH history (years across cols, months down)
  Estoques_Finais        monthly FINAL STOCKS, 3 blocks (Soja / Farelo / Óleo), 2021+
  Balanco_Brasil         monthly processing-sector balance, one year, 3 commodity blocks x 6 items
  Balanco_Complexo_Anual annual whole-complex balance, 2014+, 3 commodity blocks

Units: thousand metric tons (native Abiove). Long-format upsert; the unique key
includes source_tab so overlapping series from different tabs coexist — silver
picks the canonical tab per series.

Usage:  python scripts/load_abiove_crushing_data.py
        python scripts/load_abiove_crushing_data.py --file <path>
"""
from __future__ import annotations
import argparse, sys, unicodedata
from pathlib import Path

ROOT = Path(r"C:/dev/RLC-Agent"); sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
import openpyxl
from psycopg2.extras import execute_values
from src.services.database.db_config import get_connection

DEFAULT_FILE = ROOT / "data/raw/oilseeds_fats_greases/brazil_crushing_data.xlsx"

MONTHS = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
          "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}

def _norm(s):
    """lower, strip accents, collapse spaces — robust PT label matching."""
    if s is None: return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode()
    return " ".join(s.lower().split())

def _num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    t = str(v).strip().replace(".","").replace(",",".")
    if t in ("","na","n/a","-","nd"): return None
    try: return float(t)
    except ValueError: return None

def _month(v):
    return MONTHS.get(_norm(v)[:3])

# generic-line classifier from a PT label
def classify(label):
    n = _norm(label)
    if "estoque inicial" in n: return "initial_stock"
    if "estoque final"   in n: return "final_stock"
    if "processamento"   in n: return "crush"
    if "aquisicao"       in n: return "grain_acquisition"
    if "sementes"        in n: return "seeds_other"
    if "producao"        in n: return "production"
    if "importacao"      in n: return "imports"
    if "exportacao"      in n: return "exports"
    if "consumo interno" in n or "consumo aparente" in n: return "domestic_consumption"
    return None

def commodity_of(label):
    n = _norm(label)
    if n.startswith("1.") and "soja" in n or n == "1. soja": return "soybeans"
    if "farelo" in n: return "meal"
    if "oleo"   in n: return "oil"
    if "soja"   in n: return "soybeans"
    return None

def item_code_of(label):
    tok = str(label).strip().split()[0] if label else ""
    return tok.rstrip(".") if tok and tok[0].isdigit() else None


def parse_tabela(ws, src):
    """Monthly crush: row1 year headers (B..), rows Jan..Dez down col A."""
    rows = []
    years = {}
    for c in range(2, ws.max_column + 1):
        y = _num(ws.cell(1, c).value)
        if y and 1990 < y < 2100: years[c] = int(y)
    for r in range(2, ws.max_row + 1):
        m = _month(ws.cell(r, 1).value)
        if not m: continue
        for c, y in years.items():
            v = _num(ws.cell(r, c).value)
            if v is None: continue
            rows.append((f"{y}-{m:02d}-01", y, m, "monthly", "processing_sector",
                         "soybeans", "crush", None, "Crush (Tabela)", v, False, "Tabela", src))
    return rows


def parse_estoques(ws, src):
    """Monthly final stocks. Row4 block labels (Soja/Farelo/Óleo), row5 year headers per block."""
    rows = []
    # locate block start columns from row 4
    blocks = []  # (start_col, commodity)
    for c in range(2, ws.max_column + 1):
        n = _norm(ws.cell(4, c).value)
        if not n: continue
        if "soja"   in n: blocks.append((c, "soybeans"))
        elif "farelo" in n: blocks.append((c, "meal"))
        elif "oleo"  in n: blocks.append((c, "oil"))
    # for each block, read consecutive year columns from row 5
    for bi, (start, commodity) in enumerate(blocks):
        end = blocks[bi + 1][0] if bi + 1 < len(blocks) else ws.max_column + 1
        yearcols = {}
        for c in range(start, end):
            y = _num(ws.cell(5, c).value)
            if y and 1990 < y < 2100: yearcols[c] = int(y)
        for r in range(6, ws.max_row + 1):
            m = _month(ws.cell(r, 1).value)
            if not m: continue
            for c, y in yearcols.items():
                v = _num(ws.cell(r, c).value)
                if v is None: continue
                rows.append((f"{y}-{m:02d}-01", y, m, "monthly", "processing_sector",
                             commodity, "final_stock", None, "Estoque Final (Estoques_Finais)",
                             v, False, "Estoques_Finais", src))
    return rows


def _parse_block_grid(ws, src, tab, scope, freq, header_row, first_data_row):
    """Shared parser for Balanco_Brasil (monthly, 1 yr) & Balanco_Complexo_Anual (annual, many yrs).
    Header row: period columns (months for monthly single-year handled by caller override)."""
    rows = []
    # period columns from header_row
    periodcols = {}  # col -> (year, month_or_None, is_projection)
    for c in range(2, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        n = _norm(raw)
        if not n: continue
        proj = "(p)" in n or n.startswith("p ") or "proj" in n
        # annual: a 4-digit year; monthly single-year handled by caller
        yr = None
        for tok in n.replace("(", " ").replace(")", " ").split():
            if tok.isdigit() and len(tok) == 4: yr = int(tok)
        if yr: periodcols[c] = (yr, None, proj)
    return rows, periodcols  # caller fills rows


def parse_balanco_brasil(ws, src):
    """Monthly processing-sector balance. Multiple side-by-side year-blocks: row 3 carries
    the block year (e.g. C='2025 – mensal', O='2026 – mensal'); row 4 carries the month
    headers per block. The '(amostra)' year is a partial sample → flagged preliminary."""
    rows = []
    # 1. block anchors from row 3 (col -> year)
    anchors = {}
    for c in range(2, ws.max_column + 1):
        for tok in _norm(ws.cell(3, c).value).replace("(", " ").replace(")", " ").split():
            if tok.isdigit() and len(tok) == 4: anchors[c] = int(tok)
    if not anchors:  # fallback: single year from row 4 col B
        for tok in _norm(ws.cell(4, 2).value).split():
            if tok.isdigit() and len(tok) == 4: anchors[3] = int(tok)
    anchor_cols = sorted(anchors)
    def year_for_col(c):
        yr = None
        for ac in anchor_cols:
            if ac <= c: yr = anchors[ac]
        return yr
    # 2. amostra (sample) years -> preliminary
    prelim_years = set()
    for c in range(2, ws.max_column + 1):
        if "amostra" in _norm(ws.cell(4, c).value) or "amostra" in _norm(ws.cell(3, c).value):
            y = year_for_col(c)
            if y: prelim_years.add(y)
    # 3. month columns -> (year, month)
    monthcols = {}
    for c in range(2, ws.max_column + 1):
        m = _month(ws.cell(4, c).value)
        if m:
            y = year_for_col(c)
            if y: monthcols[c] = (y, m)
    # 4. walk rows
    commodity = None
    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None: continue
        n = _norm(label)
        if n[:2] in ("1.", "2.", "3.") and "-" not in n:
            commodity = commodity_of(label); continue
        attr = classify(label)
        if not attr or commodity is None: continue
        code = item_code_of(label)
        for c, (y, m) in monthcols.items():
            v = _num(ws.cell(r, c).value)
            if v is None: continue
            rows.append((f"{y}-{m:02d}-01", y, m, "monthly", "processing_sector",
                         commodity, attr, code, str(label).strip(), v,
                         y in prelim_years, "Balanco_Brasil", src))
    return rows


def parse_complexo_anual(ws, src):
    """Annual whole-complex balance, many years across row4."""
    rows = []
    yearcols = {}
    for c in range(2, ws.max_column + 1):
        n = _norm(ws.cell(4, c).value)
        if not n: continue
        proj = "(p)" in n or " p " in f" {n} "
        yr = None
        for tok in n.replace("(", " ").replace(")", " ").split():
            if tok.isdigit() and len(tok) == 4: yr = int(tok)
        if yr: yearcols[c] = (yr, proj)
    commodity = None
    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None: continue
        n = _norm(label)
        if n[:2] in ("1.","2.","3.") and "-" not in n:
            commodity = commodity_of(label); continue
        attr = classify(label)
        if not attr or commodity is None: continue
        code = item_code_of(label)
        for c, (yr, proj) in yearcols.items():
            v = _num(ws.cell(r, c).value)
            if v is None: continue
            rows.append((f"{yr}-01-01", yr, None, "annual", "complex",
                         commodity, attr, code, str(label).strip(), v, proj,
                         "Balanco_Complexo_Anual", src))
    return rows


UPSERT = """
INSERT INTO bronze.abiove_soy_complex
  (period, year, month, frequency, scope, commodity, attribute, item_code,
   item_label_pt, value_1000t, is_projection, source_tab, source_file)
VALUES %s
ON CONFLICT (period, frequency, scope, commodity, attribute, source_tab)
DO UPDATE SET value_1000t=EXCLUDED.value_1000t, item_code=EXCLUDED.item_code,
   item_label_pt=EXCLUDED.item_label_pt, is_projection=EXCLUDED.is_projection,
   source_file=EXCLUDED.source_file, collected_at=NOW()
"""

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=str(DEFAULT_FILE))
    args = ap.parse_args()
    src = Path(args.file).name
    # NOTE: read_only=False (normal mode). read_only streams forward per sheet and
    # ws.cell(r,c) random access across multiple sheets returns stale/wrong cells —
    # it silently corrupted Jan-Apr values. Files are small; normal mode is safe.
    wb = openpyxl.load_workbook(args.file, data_only=True, read_only=False)

    allrows = []
    allrows += parse_tabela(wb["Tabela"], src)
    allrows += parse_estoques(wb["Estoques_Finais"], src)
    allrows += parse_balanco_brasil(wb["Balanco_Brasil"], src)
    allrows += parse_complexo_anual(wb["Balanco_Complexo_Anual"], src)

    # de-dup within batch on the unique key
    seen = {}
    for row in allrows:
        key = (row[0], row[3], row[4], row[5], row[6], row[11])
        seen[key] = row
    rows = list(seen.values())

    with get_connection() as conn:
        cur = conn.cursor()
        execute_values(cur, UPSERT, rows)
        conn.commit()
        print(f"Loaded {len(rows)} rows into bronze.abiove_soy_complex")
        # summary by tab
        cur.execute("""SELECT source_tab, frequency, count(*) n,
                              min(period) mn, max(period) mx
                       FROM bronze.abiove_soy_complex GROUP BY 1,2 ORDER BY 1""")
        for r in cur.fetchall():
            print(f"  {r['source_tab']:24} {r['frequency']:8} n={r['n']:4}  {r['mn']} -> {r['mx']}")
        # reconciliation: Tabela crush vs Balanco_Brasil crush for the overlap year
        cur.execute("""
            SELECT extract(year from period)::int y, extract(month from period)::int m,
                   max(value_1000t) FILTER (WHERE source_tab='Tabela') tabela,
                   max(value_1000t) FILTER (WHERE source_tab='Balanco_Brasil') balanco
            FROM bronze.abiove_soy_complex
            WHERE commodity='soybeans' AND attribute='crush' AND frequency='monthly'
            GROUP BY 1,2 HAVING max(value_1000t) FILTER (WHERE source_tab='Balanco_Brasil') IS NOT NULL
            ORDER BY 1,2 LIMIT 12""")
        recon = cur.fetchall()
        if recon:
            print("\n  Crush reconciliation Tabela vs Balanco_Brasil (overlap year):")
            for r in recon:
                d = (r['tabela'] or 0) - (r['balanco'] or 0)
                print(f"    {r['y']}-{r['m']:02d}  Tabela={r['tabela']}  Balanco={r['balanco']}  diff={d:+.0f}")

if __name__ == "__main__":
    main()
