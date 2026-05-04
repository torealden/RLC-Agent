"""
One-time cleanup: clear stale data from DCO Imports and DCO Exports sheets
in models/Fats and Greases/us_fats_greases_trade.xlsm.

Why: The macro currently filters by country at write-time, so non-DCO
country rows (Chile, Aruba, etc.) never get OVERWRITTEN by new data.
But they often retain stale data from prior macro runs that used
different filters or row mappings. The fix is two-part:
  (a) clear all data cells on the DCO sheets ONCE (this script)
  (b) tighten the macro's clear logic to clear-all on refill (separate change)

What it does:
- Opens the workbook with macros preserved (keep_vba=True)
- For each DCO sheet, finds the data range (DATA_START_ROW..DATA_END_ROW ×
  data columns from col 2 onward where row 2 has a date header)
- Clears values in every NON-regional-subtotal row
- Saves the workbook back

Regional subtotal rows (preserved by the macro itself) are also preserved
here. Their values come from sheet formulas, not cell values.

Usage:
    python scripts/clear_dco_stale_data.py
    # then run the macro (Ctrl+Shift+I or whatever you use) to refill
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl


WB_PATH = Path("models/Fats and Greases/us_fats_greases_trade.xlsm")
SHEETS = ["DCO Imports", "DCO Exports"]
DATA_START_ROW = 4   # matches the macro
DATA_END_ROW = 217   # matches the macro

# Regional subtotal rows — preserved by the macro and we preserve them too.
# Verbatim from TradeUpdaterSQL.bas line 35:
#   Private Const REGIONAL_ROWS As String = "4,33,47,61,108,165,216"
REGIONAL_ROW_NUMBERS = {4, 33, 47, 61, 108, 165, 216}
# Note: row 217 (WORLD TOTAL) is NOT in the protected set — the macro
# writes to it. For DCO mode it copies row 216 → row 217 explicitly.


def clear_data_cells(ws, max_col: int) -> int:
    """Clear values in non-regional rows for COLUMNS THAT HAVE A MONTH-YEAR
    DATE IN ROW 2. Skips marketing-year accumulator columns (which often
    hold formulas referencing the data range). Mirrors the macro's
    IsMonthHeaderCell rule from TradeUpdaterSQL.bas.
    Returns number of cells cleared (had non-None values)."""
    import datetime
    cleared = 0
    cleared_cols = 0
    for col in range(2, max_col + 1):
        header = ws.cell(row=2, column=col).value
        # Match the macro's rule: only clear if header is a month-year date
        if not isinstance(header, datetime.datetime):
            continue
        cleared_cols += 1
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            if row in REGIONAL_ROW_NUMBERS:
                continue
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None
                cleared += 1
    return cleared, cleared_cols


def main():
    if not WB_PATH.exists():
        raise SystemExit(f"Workbook not found: {WB_PATH}")

    # Backup the workbook first — touching xlsm files is risky
    backup = WB_PATH.with_name(
        WB_PATH.stem + "_backup_" + datetime.now().strftime("%Y%m%d_%H%M%S") + WB_PATH.suffix
    )
    shutil.copy2(WB_PATH, backup)
    print(f"Backed up to: {backup.name}")

    wb = openpyxl.load_workbook(WB_PATH, keep_vba=True, data_only=False)

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        # max_column tells us where data actually goes
        max_col = ws.max_column
        cleared, cleared_cols = clear_data_cells(ws, max_col)
        print(f"  {sheet_name}: cleared {cleared} cells across {cleared_cols} "
              f"month-header columns (rows {DATA_START_ROW}-{DATA_END_ROW}, "
              f"preserved {len(REGIONAL_ROW_NUMBERS)} regional rows + "
              f"all marketing-year aggregate columns)")

    wb.save(WB_PATH)
    print(f"\nSaved: {WB_PATH}")
    print("Next step: open in Excel and run the DCO macro (Ctrl+Shift+I) to refill.")


if __name__ == "__main__":
    main()
