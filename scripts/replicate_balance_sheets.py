"""
Replicate CWG balance sheet template for other fats/greases commodities.
Replaces NASS columns, trade sheet names, allocation columns, and labels.
"""
import openpyxl
import re

COMMODITIES = {
    'inedible_tallow': {
        'sheet_name': 'Inedible Tallow',
        'block_prefix': 'INEDIBLE TALLOW',
        'nass_prod_col': 'AE', 'nass_stocks_col': 'AG',
        'import_sheet': 'Inedible Tallow Imports',
        'export_sheet': 'Inedible Tallow Exports',
        'alloc_bd': 'F', 'alloc_rd': 'O', 'alloc_saf': 'X', 'alloc_copro': 'AG',
        'slaughter_label': 'COMMERCIAL CATTLE SLAUGHTER',
        'liveweight_label': 'COMMERCIAL CATTLE PRODUCTION LIVE WEIGHT',
        'price_label_1': 'INEDIBLE TALLOW AVERAGE PRICE',
        'price_sub_1': '(Chicago - cents per pound)',
    },
    'edible_tallow': {
        'sheet_name': 'Edible Tallow',
        'block_prefix': 'EDIBLE TALLOW',
        'nass_prod_col': 'AA', 'nass_stocks_col': 'AC',
        'import_sheet': 'Edible Tallow Imports',
        'export_sheet': 'Edible Tallow Exports',
        'alloc_bd': 'F', 'alloc_rd': 'O', 'alloc_saf': 'X', 'alloc_copro': 'AG',
        'slaughter_label': 'COMMERCIAL CATTLE SLAUGHTER',
        'liveweight_label': 'COMMERCIAL CATTLE PRODUCTION LIVE WEIGHT',
        'price_label_1': 'EDIBLE TALLOW AVERAGE PRICE',
        'price_sub_1': '(Chicago - cents per pound)',
    },
    'yellow_grease': {
        'sheet_name': 'Yellow Grease',
        'block_prefix': 'YELLOW GREASE',
        'nass_prod_col': 'AM', 'nass_stocks_col': 'AO',
        'import_sheet': 'Yellow Grease Imports',
        'export_sheet': 'Yellow Grease Exports',
        'alloc_bd': 'H', 'alloc_rd': 'Q', 'alloc_saf': 'Z', 'alloc_copro': 'AI',
        'slaughter_label': None, 'liveweight_label': None,
        'price_label_1': 'YELLOW GREASE AVERAGE PRICE',
        'price_sub_1': '(IL/WI - cents per pound)',
    },
    'poultry_fat': {
        'sheet_name': 'Poultry Fat',
        'block_prefix': 'POULTRY FAT',
        'nass_prod_col': 'T', 'nass_stocks_col': 'V',
        'import_sheet': 'Poultry Fat Imports',
        'export_sheet': 'Poultry Fat Exports',
        'alloc_bd': 'E', 'alloc_rd': 'N', 'alloc_saf': 'W', 'alloc_copro': 'AF',
        'slaughter_label': 'COMMERCIAL CHICKEN SLAUGHTER',
        'liveweight_label': 'COMMERCIAL CHICKEN PRODUCTION LIVE WEIGHT',
        'price_label_1': 'POULTRY FAT AVERAGE PRICE',
        'price_sub_1': '(Southeast - cents per pound)',
    },
    'lard': {
        'sheet_name': 'Lard',
        'block_prefix': 'LARD',
        'nass_prod_col': 'I', 'nass_stocks_col': 'K',
        'import_sheet': 'Lard Imports',
        'export_sheet': 'Lard Exports',
        'alloc_bd': None, 'alloc_rd': None, 'alloc_saf': None, 'alloc_copro': None,
        'slaughter_label': 'COMMERCIAL HOG SLAUGHTER',
        'liveweight_label': 'COMMERCIAL HOG PRODUCTION LIVE WEIGHT',
        'price_label_1': 'LARD AVERAGE PRICE',
        'price_sub_1': '(Chicago - cents per pound)',
    },
}


def replace_nass_col(formula, old_col, new_col):
    """Replace NASS column references like '!B437' -> '!AE437'"""
    # Match the pattern: '!{old_col} followed by digits
    pattern = f"'!{old_col}(\\d)"
    replacement = f"'!{new_col}\\1"
    return re.sub(pattern, replacement, formula)


def replace_alloc_col(formula, old_col, new_col):
    """Replace allocation column refs like '$G5' -> '$F5'"""
    if new_col is None:
        return '=0'
    pattern = f"\\${old_col}(\\d+)"
    replacement = f"${new_col}\\1"
    return re.sub(pattern, replacement, formula)


def process_commodity(code, cfg):
    template_path = (
        r'C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models/'
        r'Fats and Greases/new_models/us_choice_white_grease_balance.xlsx'
    )
    output_dir = (
        r'C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models/'
        r'Fats and Greases/new_models'
    )

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Choice White Grease']
    ws.title = cfg['sheet_name']

    for row in ws.iter_rows(min_row=1, max_row=276, max_col=35, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue

            if isinstance(v, str):
                # --- Label replacements ---
                v = v.replace('US CHOICE WHITE GREASE SUPPLY AND DEMAND',
                              f"US {cfg['block_prefix']} SUPPLY AND DEMAND")
                v = v.replace('CHOICE WHITE GREASE', cfg['block_prefix'])

                if cfg.get('slaughter_label'):
                    v = v.replace('COMMERCIAL HOG SLAUGHTER', cfg['slaughter_label'])
                    v = v.replace('COMMERCIAL HOG PRODUCTION LIVE WEIGHT',
                                  cfg['liveweight_label'])

                if cfg.get('price_label_1'):
                    v = v.replace('CHOICE WHITE GREASE AVERAGE PRICE',
                                  cfg['price_label_1'])
                if cfg.get('price_sub_1'):
                    v = v.replace('(Chicago - cents per pound)', cfg['price_sub_1'])
                    v = v.replace('(Missouri River - cents per pound)',
                                  cfg['price_sub_1'].replace(
                                      cfg['price_sub_1'].split('(')[1].split('-')[0].strip(),
                                      'West Coast'))

                # --- Formula replacements ---
                if v.startswith('='):
                    # Production: NASS col B -> new col
                    v = replace_nass_col(v, 'B', cfg['nass_prod_col'])

                    # Stocks: NASS col D -> new col
                    v = replace_nass_col(v, 'D', cfg['nass_stocks_col'])

                    # Trade sheet names
                    v = v.replace('CWG Imports', cfg['import_sheet'])
                    v = v.replace('CWG Exports', cfg['export_sheet'])

                    # Allocation columns
                    if '[2]Allocation' in v or '[4]Allocation' in v:
                        if cfg['alloc_bd'] is None:
                            # Lard: no allocation, zero out
                            v = '=0'
                        else:
                            v = replace_alloc_col(v, 'G', cfg['alloc_bd'])
                            v = replace_alloc_col(v, 'P', cfg['alloc_rd'])
                            v = replace_alloc_col(v, 'Y', cfg['alloc_saf'])
                            v = replace_alloc_col(v, 'AH', cfg['alloc_copro'])

                cell.value = v

    # Clear hardcoded price data - these need commodity-specific values
    for row_num in range(216, 228):
        for col_num in range(2, 35):
            c = ws.cell(row=row_num, column=col_num)
            if c.value is not None and isinstance(c.value, (int, float)):
                c.value = None
    for row_num in range(232, 244):
        for col_num in range(2, 35):
            c = ws.cell(row=row_num, column=col_num)
            if c.value is not None and isinstance(c.value, (int, float)):
                c.value = None

    outpath = f'{output_dir}/us_{code}_balance.xlsx'
    wb.save(outpath)
    print(f'Created: us_{code}_balance.xlsx ({cfg["sheet_name"]})')


if __name__ == '__main__':
    for code, cfg in COMMODITIES.items():
        process_commodity(code, cfg)
    print('\nDone - 5 balance sheets created from CWG template')
