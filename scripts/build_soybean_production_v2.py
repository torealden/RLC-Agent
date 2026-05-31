"""V2 production workbook: one file per commodity, one tab per statistic,
vintage columns. Soybeans as the prototype.

Per Tore 2026-05-30 (after the wide-format v1 was flagged for layout +
NASS short_desc contamination):
  - File: models/Oilseeds/us_soybean_production.xlsx (NEW)
  - Tabs: area_planted, area_harvested, production, yield, _meta
  - Rows: years ascending (latest at bottom — VLOOKUP convention)
  - Columns: vintage trail
      area_planted    : Year | PP (Mar) | Acreage (Jun) | Aug | Sep | Oct | Nov | Final (Jan)
      area_harvested  : Year | Aug | Sep | Oct | Nov | Final (Jan)
      production      : same as area_harvested
      yield           : same as area_harvested
  - National only (US Total). State-level revival is a follow-up.
  - Data filter: pin short_desc to the canonical ACRES PLANTED / ACRES
    HARVESTED / YIELD MEASURED IN BU/ACRE / PRODUCTION MEASURED IN BU rows
    so percentages and farm-counts don't sneak in.

Why this redesign:
  v1 picked "latest release_date" across ALL short_descs sharing a
  statisticcat_desc. NASS publishes per-method percentages and
  irrigated-farm counts under the same statisticcat — so the AH 2022 row
  got "2,979 operations" instead of 86.17M acres, etc. V2 pins to exact
  short_descs.
"""
from __future__ import annotations

import sys, logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / '.env')

from src.services.database.db_config import get_connection

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
log = logging.getLogger('build_soybean_production_v2')

# Per-commodity short_desc map. For soybeans these are the canonical rows
# that DEFINE area planted / harvested / yield / production. Anything else
# under the same statisticcat_desc (irrigated, biotech, double-cropped, etc.)
# is excluded.
SHORT_DESC = {
    'area_planted':   'SOYBEANS - ACRES PLANTED',
    'area_harvested': 'SOYBEANS - ACRES HARVESTED',
    'yield':          'SOYBEANS - YIELD, MEASURED IN BU / ACRE',
    'production':     'SOYBEANS - PRODUCTION, MEASURED IN BU',
}

# Vintage column definitions per statistic.
# Each entry: (label, ref_periods accepted, release-date filter, pick mode)
#   ref_periods: list of NASS reference_period_desc values to accept
#   release filter: callable (release_date, crop_year) -> bool
#   pick mode: 'earliest' picks the original release (first publication of
#              this vintage); 'latest' picks the most current revision
#
# Tore's preference (2026-05-31): "hold the original PP planting number
# in that column and as the estimate changes, reflect the changes in the
# current non-PP columns." So vintage columns use 'earliest' (original
# release) where possible. Final uses 'latest' so it always reflects
# the most current canonical 'YEAR' value.

def _in_month(target_month):
    return lambda rel, yr: rel.year == yr and rel.month == target_month
def _any(rel, yr): return True
def _final_year(rel, yr):
    # Final = post-harvest canonical 'YEAR' release. Must be released in
    # crop_year+1 or later (skips intra-year rp='YEAR' releases like the
    # Mar PP and Jun Acreage when they use the bare 'YEAR' periodicity).
    # Also avoids Census-of-Ag re-publications (which release values for
    # historical crop_years in subsequent off-cycle months under the same
    # short_desc but with anomalous magnitudes).
    return rel.year >= yr + 1

AP_COLS = [
    # Tore confirmed (2026-05-31) that NASS does NOT actually revise the
    # original PP/Acreage values — they just reset load_time in a
    # 2018-01-23 database migration. Spot-checked 2012/2013/2014 against
    # the actual USDA press releases; values match to the unit. So we
    # trust reference_period as the discriminator (no release-month filter).
    ('PP (Mar)',      ['YEAR - MAR ACREAGE'],          _any,         'earliest'),
    ('Acreage (Jun)', ['YEAR - JUN ACREAGE'],          _any,         'earliest'),
    ('Aug WASDE',     ['YEAR - AUG FORECAST'],         _any,         'earliest'),
    ('Sep',           ['YEAR - SEP FORECAST'],         _any,         'earliest'),
    ('Oct',           ['YEAR - OCT FORECAST'],         _any,         'earliest'),
    ('Nov',           ['YEAR - NOV FORECAST'],         _any,         'earliest'),
    ('Final (Jan)',   ['YEAR'],                        _final_year,  'earliest'),
]

SEASON_COLS = [
    ('Aug WASDE',     ['YEAR - AUG FORECAST'],         _any,         'earliest'),
    ('Sep',           ['YEAR - SEP FORECAST'],         _any,         'earliest'),
    ('Oct',           ['YEAR - OCT FORECAST'],         _any,         'earliest'),
    ('Nov',           ['YEAR - NOV FORECAST'],         _any,         'earliest'),
    ('Final (Jan)',   ['YEAR'],                        _final_year,  'earliest'),
]

# Brand styling
HEADER_FILL = PatternFill(start_color='3C7D22', end_color='3C7D22', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
YEAR_FONT   = Font(name='Calibri', size=10, bold=True)
DATA_FONT   = Font(name='Calibri', size=10)
ALT_FILL    = PatternFill(start_color='F4F8F1', end_color='F4F8F1', fill_type='solid')


def fetch_vintage(commodity_db: str, class_db: str, statistic: str,
                  short_desc: str) -> list:
    """Pull all rows. Returns list of (crop_year, ref_period, release_date,
    value, unit) tuples — no deduplication, caller picks per-vintage rule."""
    sql = """
        SELECT crop_year, reference_period, release_date, value, unit
        FROM silver.crop_production
        WHERE commodity = %s AND class = %s
          AND statistic = %s AND short_desc = %s
          AND agg_level = 'NATIONAL'
          AND value IS NOT NULL
        ORDER BY crop_year, release_date
    """
    out = []
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (commodity_db, class_db, statistic, short_desc))
            for r in cur.fetchall():
                d = dict(r)
                out.append((int(d['crop_year']), d['reference_period'],
                            d['release_date'], float(d['value']), d['unit']))
    return out


def match_value(rows: list, year: int, ref_periods: list,
                rel_filter, mode: str) -> tuple | None:
    """Find a row matching (year, ref_period in ref_periods, rel_filter passes).
    Returns (value, unit, release_date) per pick mode ('earliest' or 'latest').
    Tries each ref_period in order; once any matches, that's the column's source."""
    for ref_period in ref_periods:
        candidates = [
            (rel, val, unit) for (cy, rp, rel, val, unit) in rows
            if cy == year and rp == ref_period and rel_filter(rel, year)
        ]
        if not candidates:
            continue
        candidates.sort(key=lambda x: x[0])
        chosen = candidates[0] if mode == 'earliest' else candidates[-1]
        rel, val, unit = chosen
        return (val, unit, rel)
    return None


def build_tab(ws, statistic: str, short_desc: str, col_specs: list,
              year_ge: int, year_le: int):
    """Populate one tab. Returns row count."""
    rows = fetch_vintage('soybeans', 'all_classes', statistic, short_desc)

    # Determine units (should be consistent within one short_desc)
    units = sorted({r[4] for r in rows})
    unit_label = units[0] if len(units) == 1 else f'mixed: {units}'

    # Headers
    ws.cell(1, 1, value='Year').font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    for i, (label, _, _, _) in enumerate(col_specs, start=2):
        c = ws.cell(1, i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')

    # Data rows, ascending years
    row = 2
    for year in range(year_ge, year_le + 1):
        c = ws.cell(row, 1, value=year)
        c.font = YEAR_FONT
        if year % 2 == 1:
            c.fill = ALT_FILL
        for i, (_, ref_periods, rel_filter, mode) in enumerate(col_specs, start=2):
            match = match_value(rows, year, ref_periods, rel_filter, mode)
            cell = ws.cell(row, i)
            if year % 2 == 1:
                cell.fill = ALT_FILL
            if match is None:
                continue
            value, unit, rel_date = match
            cell.value = value
            cell.font = DATA_FONT
            if statistic == 'yield':
                cell.number_format = '#,##0.0'
            else:
                cell.number_format = '#,##0'
        row += 1

    # Column widths + freeze
    ws.column_dimensions['A'].width = 8
    for c in range(2, 2 + len(col_specs)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    ws.freeze_panes = 'B2'

    # Footer note with unit
    foot_row = row + 1
    ws.cell(foot_row, 1, value=f'Units: {unit_label}').font = Font(name='Calibri', size=9, italic=True)

    return row - 2  # data rows written (excluding header)


def build_meta(ws, file_meta: dict):
    ws['A1'] = 'Meta'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws['B1'] = 'Value'
    ws['B1'].font = HEADER_FONT
    ws['B1'].fill = HEADER_FILL

    ws['A2'] = 'Generated'
    ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A3'] = 'Source'
    ws['B3'] = 'silver.crop_production (NASS Crop Production + Annual CPS)'
    ws['A4'] = 'Layout'
    ws['B4'] = 'One row per crop year, columns = vintage release. Pinned to canonical short_desc to exclude percentages/farm-counts.'
    ws['A5'] = 'Vintage definitions'
    ws['B5'] = 'PP=Prospective Plantings (Mar 31); Acreage=Jun 30 USDA Acreage report; Aug/Sep/Oct/Nov=Crop Production monthly forecast; Final=Annual Crop Production Summary (Jan)'

    row = 7
    ws[f'A{row}'] = 'Tab'
    ws[f'B{row}'] = 'Rows'
    ws[f'C{row}'] = 'Short desc'
    for c in (f'A{row}', f'B{row}', f'C{row}'):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
    row += 1
    for tab, info in file_meta.items():
        ws.cell(row, 1, value=tab)
        ws.cell(row, 2, value=info['rows'])
        ws.cell(row, 3, value=info['short_desc'])
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 60


def main():
    out_path = PROJECT_ROOT / 'models/Oilseeds/us_soybean_production.xlsx'
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    file_meta = {}

    tabs = [
        ('area_planted',   AP_COLS),
        ('area_harvested', SEASON_COLS),
        ('production',     SEASON_COLS),
        ('yield',          SEASON_COLS),
    ]
    for stat, col_specs in tabs:
        sd = SHORT_DESC[stat]
        log.info(f'Building tab: {stat}  short_desc={sd}')
        ws = wb.create_sheet(title=stat)
        n = build_tab(ws, stat, sd, col_specs, year_ge=2000, year_le=2026)
        file_meta[stat] = {'rows': n, 'short_desc': sd}

    ws_meta = wb.create_sheet(title='_meta')
    build_meta(ws_meta, file_meta)

    wb.save(out_path)
    log.info(f'Wrote {out_path}')


if __name__ == '__main__':
    main()
