"""Build the four US BBD price workbooks per Tore's spec (2026-05-31):
  - models/Oilseeds/us_veg_oil_prices.xlsx
  - models/Fats and Greases/us_animal_fat_prices.xlsx
  - models/Biofuels/us_fuel_prices.xlsx
  - models/Biofuels/us_credit_prices.xlsx

Each tab = one product, rows = weekly Friday dates ascending,
columns = locations. Data sources: silver.feedstock_prices_consolidated
(AMS + fastmarkets) for feedstock + animal fat tabs; bronze.fuel_prices
and bronze.credit_prices for fuel + credit.

Stale data treatment (Tore 2026-05-31): include fastmarkets rows with
date stamp + source per column. _meta tab notes the last update per
column. The Feedstock Report generator filters source LIKE 'fastmarkets%'
before client-facing rendering per the keep-don't-show rule.
"""
from __future__ import annotations

import logging
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / '.env')

from src.services.database.db_config import get_connection

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
log = logging.getLogger('build_price_workbooks')


# ============================================================
# Workbook definitions
# ============================================================
# Each tab definition: {col_label: (feedstock_code, region)} for the
# feedstock view path. For fuel/credit we use a different fetch path.

VEG_OIL_TABS = {
    'sbo': {
        'Central IL (cash)':     ('SBO', 'central_il'),
        'Central IL (basis)':    ('SBO', 'central_il_basis'),
        'Central IL (crude)':    ('SBO', 'central_il_crude'),
        'Central IL (RBD basis)':('SBO', 'central_il_rbd_basis'),
        'US Gulf':               ('SBO', 'us_gulf'),
        'US Gulf (basis)':       ('SBO', 'us_gulf_basis'),
        'US Gulf (crude)':       ('SBO', 'us_gulf_crude'),
        'West Coast (RBD)':      ('SBO', 'west_coast_rbd'),
        'Argentina (upriver)':   ('SBO', 'argentina_upriver'),
        'Brazil (FOB)':          ('SBO', 'brazil_fob'),
    },
    'canola_oil': {
        'Central US':            ('CO', 'central_us'),
        'Los Angeles':           ('CO', 'los_angeles'),
        'Canada (CNF)':          ('CO', 'canada_cnf'),
        'Dutch (FOB)':           ('CO', 'dutch_fob'),
    },
    'dco': {
        'IL / WI':               ('DCO', 'il_wi'),
        'West Coast':            ('DCO', 'west_coast'),
        'Delivered CA':          ('DCO', 'delivered_ca'),
    },
    'palm_oil': {
        'US Gulf':               ('PALM', 'us_gulf'),
        'Malaysia (futures)':    ('PALM', 'malaysia_futures'),
        'Indonesia (FOB)':       ('PALM', 'indonesia_fob'),
        'NYC (CNO)':             ('PALM', 'nyc_cno'),
        'NYC (PKO)':             ('PALM', 'nyc_pko'),
    },
}

ANIMAL_FAT_TABS = {
    'tallow_inedible': {
        'Chicago (AMS)':         ('BFT', 'chicago'),
        'West Coast (FM)':       ('BFT', 'west_coast'),
        'US Gulf (FM)':          ('BFT', 'us_gulf'),
        'Chicago - LCFS':        ('BFT', 'chicago_lcfs'),
        'Delivered CA (LCFS)':   ('BFT', 'delivered_ca'),
    },
    'cwg': {
        'Chicago (AMS)':         ('CWG', 'chicago'),
        'Central US (AMS)':      ('CWG', 'central_us'),
        'Minnesota (AMS)':       ('CWG', 'minnesota'),
        'Missouri River (FM)':   ('CWG', 'missouri_river'),
        'West Coast (FM)':       ('CWG', 'west_coast'),
    },
    'yellow_grease': {
        'Minnesota (AMS)':       ('YG', 'minnesota'),
        'CA Central Coast (AMS)':('YG', 'ca_central_coast'),
        'CA SJV (AMS)':          ('YG', 'ca_sjv'),
        'CA South (AMS)':        ('YG', 'ca_south'),
        'Eastern Cornbelt (AMS)':('YG', 'eastern_corn_belt'),
        'IL / WI (FM)':          ('YG', 'il_wi'),
        'Los Angeles (FM)':      ('YG', 'los_angeles'),
        'LA - LCFS':             ('YG', 'la_lcfs'),
    },
    'lard': {
        'Chicago (AMS)':         ('LARD', 'chicago'),
        'Central US (AMS)':      ('LARD', 'central_us'),
    },
    'uco': {
        'IL / WI (FM)':          ('UCO', 'il_wi'),
        'SoCal (FM)':            ('UCO', 'socal'),
        'SoCal - LCFS':          ('UCO', 'socal_lcfs'),
    },
    'poultry_fat': {
        'Southeast (FM)':        ('PF', 'southeast'),
        'West Coast (FM)':       ('PF', 'west_coast'),
    },
    'mbm': {
        'Central US':            ('MBM', 'central_us'),
        'Minnesota':             ('MBM', 'minnesota'),
        'US PNW':                ('MBM', 'us_pnw'),
        'Southern Plains':       ('MBM', 'southern_plains'),
        'Eastern Cornbelt':      ('MBM', 'eastern_corn_belt'),
        'Panhandle':             ('MBM', 'panhandle'),
        'CA SJV':                ('MBM', 'ca_sjv'),
        'CA South':              ('MBM', 'ca_south'),
    },
    'blood_meal': {
        'Central US':            ('BM', 'central_us'),
        'Minnesota':             ('BM', 'minnesota'),
        'California':            ('BM', 'california'),
        'CA SJV':                ('BM', 'ca_sjv'),
        'Eastern Cornbelt':      ('BM', 'eastern_corn_belt'),
        'Southern Plains':       ('BM', 'southern_plains'),
        'Panhandle':             ('BM', 'panhandle'),
    },
    'feathermeal': {
        'Minnesota':             ('FM', 'minnesota'),
        'KC Region':             ('FM', 'kc'),
        'Arkansas':              ('FM', 'arkansas'),
        'Mississippi':           ('FM', 'mississippi'),
    },
}

# Fuel + credit use a different schema — columnar bronze tables, not the
# long/tidy feedstock view. Map: {tab: {col_label: column_in_bronze_table}}
FUEL_TABS = {
    'ulsd': {
        'US Gulf':              'ulsd_gulf',
        'NY Harbor':            'ulsd_nyharbor',
    },
    'biodiesel': {
        'National (B100)':      'b100_national',
        'Upper Midwest':        'b100_upper_midwest',
        'Lower Midwest':        'b100_lower_midwest',
        'South Central':        'b100_south_central',
        'Southeast':            'b100_southeast',
        'Northeast':            'b100_northeast',
        'Rocky Mountain':       'b100_rocky_mountain',
    },
    'renewable_diesel': {
        'California':           'rd_california',
    },
    'jet_a': {
        'Spot':                 'jet_a_spot',
    },
    'crude_and_heating': {
        'WTI Crude':            'wti_crude',
        'Heating Oil (futures)':'heating_oil_futures',
    },
}

CREDIT_TABS = {
    'rin_credits': {
        'D3 RIN':               'd3_rin',
        'D4 RIN':               'd4_rin',
        'D5 RIN':               'd5_rin',
        'D6 RIN':               'd6_rin',
        'D4-D6 spread':         'd4_d6_spread',
        'D4-D5 spread':         'd4_d5_spread',
    },
    'lcfs_and_state': {
        'California LCFS':      'lcfs_ca',
        'Oregon CFP':           'cfp_or',
        'Washington CFS':       'cfs_wa',
    },
}


# ============================================================
# Styling
# ============================================================
HEADER_FILL = PatternFill(start_color='3C7D22', end_color='3C7D22', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
DATE_FONT   = Font(name='Calibri', size=10, bold=True)
DATA_FONT   = Font(name='Calibri', size=10)
ALT_FILL    = PatternFill(start_color='F4F8F1', end_color='F4F8F1', fill_type='solid')
NOTE_FONT   = Font(name='Calibri', size=9, italic=True, color='666666')


# ============================================================
# Data fetchers
# ============================================================
def fetch_feedstock_series(fs_code: str, region: str) -> dict:
    """Returns dict[price_date] -> (price_per_lb, source)."""
    sql = """
        SELECT price_date, price_per_lb, source
        FROM silver.feedstock_prices_consolidated
        WHERE feedstock_code = %s AND region = %s
          AND price_per_lb IS NOT NULL AND price_per_lb > 0
        ORDER BY price_date
    """
    out = {}
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (fs_code, region))
            for r in cur.fetchall():
                d = dict(r)
                out[d['price_date']] = (float(d['price_per_lb']), d['source'])
    return out


def fetch_table_column(table: str, column: str) -> dict:
    """Generic fetcher for bronze.fuel_prices and bronze.credit_prices.
    Returns dict[price_date] -> (value, source).

    NOTE: credit_prices includes forward-curve data going to 2050 (used for
    45Z economics modeling). Filter to price_date <= today so it doesn't
    bleed into the historical weekly grid.
    """
    # Forward-curve rows in credit_prices have frequency='monthly' and run
    # through 2050 (used for 45Z economics modeling). Historical observations
    # have frequency='weekly'. Filter to weekly to drop the projections.
    sql = f"""
        SELECT price_date, {column} AS val, source
        FROM bronze.{table}_prices
        WHERE {column} IS NOT NULL
          AND price_date <= CURRENT_DATE
          AND frequency = 'weekly'
        ORDER BY price_date
    """
    out = {}
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql)
            for r in cur.fetchall():
                d = dict(r)
                out[d['price_date']] = (float(d['val']), d['source'])
    return out


# ============================================================
# Weekly Friday resampling
# ============================================================
def friday_of_week(d: date) -> date:
    """Return the Friday of the week containing d (Monday=0)."""
    return d + timedelta(days=(4 - d.weekday()))


def resample_to_fridays(series: dict, start: date, end: date) -> dict:
    """For each Friday in [start, end], pick the latest observation on or
    before that Friday within the trailing week (Sat-Fri). Returns
    dict[friday_date] -> (value, source, observation_date)."""
    if not series:
        return {}
    obs = sorted(series.items())  # [(date, (val, src)), ...]
    out = {}
    first_friday = friday_of_week(start)
    if first_friday < start:
        first_friday += timedelta(days=7)
    fri = first_friday
    while fri <= end:
        # Find the latest observation in window (fri - 6 days, fri]
        window_start = fri - timedelta(days=6)
        candidates = [(d, v) for d, v in obs if window_start <= d <= fri]
        if candidates:
            obs_date, (val, src) = candidates[-1]
            out[fri] = (val, src, obs_date)
        fri += timedelta(days=7)
    return out


# ============================================================
# Tab builder
# ============================================================
def build_tab(ws, tab_name: str, columns: dict, fetch_fn,
              start: date, end: date) -> dict:
    """columns = {label: param_for_fetch_fn}. Returns metadata."""
    # Fetch + resample each column
    series_per_col = {}
    src_per_col = {}
    last_per_col = {}
    for label, params in columns.items():
        if isinstance(params, tuple):
            raw = fetch_fn(*params)
        else:
            raw = fetch_fn(params)
        series_per_col[label] = resample_to_fridays(raw, start, end)
        # Determine dominant source + last observation date
        if raw:
            srcs = {s for _, s in raw.values()}
            src_per_col[label] = ' / '.join(sorted(srcs))
            last_per_col[label] = max(raw.keys())
        else:
            src_per_col[label] = 'no data'
            last_per_col[label] = None

    # Build the grid: rows = all Fridays, cols = labels
    all_fridays = set()
    for s in series_per_col.values():
        all_fridays.update(s.keys())
    fridays = sorted(all_fridays)

    # Header row
    ws.cell(1, 1, value='Date').font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    for i, label in enumerate(columns.keys(), start=2):
        c = ws.cell(1, i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')

    # Source row (row 2)
    ws.cell(2, 1, value='Source').font = NOTE_FONT
    for i, label in enumerate(columns.keys(), start=2):
        c = ws.cell(2, i, value=src_per_col[label])
        c.font = NOTE_FONT

    # Last update row (row 3)
    ws.cell(3, 1, value='Last update').font = NOTE_FONT
    for i, label in enumerate(columns.keys(), start=2):
        lu = last_per_col[label]
        c = ws.cell(3, i, value=str(lu) if lu else '-')
        c.font = NOTE_FONT

    # Data rows
    row = 4
    for i_fr, fri in enumerate(fridays):
        c = ws.cell(row, 1, value=fri)
        c.font = DATE_FONT
        c.number_format = 'yyyy-mm-dd'
        if i_fr % 2 == 1:
            c.fill = ALT_FILL
        for i, label in enumerate(columns.keys(), start=2):
            cell = ws.cell(row, i)
            if i_fr % 2 == 1:
                cell.fill = ALT_FILL
            match = series_per_col[label].get(fri)
            if match is None:
                continue
            val, src, obs_dt = match
            cell.value = val
            cell.font = DATA_FONT
            cell.number_format = '#,##0.0000'
        row += 1

    # Column widths + freeze panes
    ws.column_dimensions['A'].width = 12
    for c in range(2, 2 + len(columns)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 20
    ws.freeze_panes = 'B4'

    return {
        'rows': len(fridays),
        'columns': len(columns),
        'sources': src_per_col,
        'last_updates': last_per_col,
    }


# ============================================================
# Meta tab
# ============================================================
def build_meta(ws, file_meta: dict, workbook_name: str):
    ws['A1'] = 'Meta'; ws['B1'] = 'Value'
    for c in ('A1', 'B1'):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL

    ws['A2'] = 'Workbook'
    ws['B2'] = workbook_name
    ws['A3'] = 'Generated'
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A4'] = 'Frequency'
    ws['B4'] = 'Weekly Friday (latest observation in trailing Sat-Fri window)'
    ws['A5'] = 'Note'
    ws['B5'] = ('Source column on row 2 shows feed origin. Fastmarkets rows '
                'are for INTERNAL TRIANGULATION ONLY — never publish in '
                'client-facing material (per memory:feedback_fastmarkets_'
                'keep_dont_show).')

    r = 7
    ws[f'A{r}'] = 'Tab'
    ws[f'B{r}'] = 'Column'
    ws[f'C{r}'] = 'Source(s)'
    ws[f'D{r}'] = 'Last update'
    for col in (f'A{r}', f'B{r}', f'C{r}', f'D{r}'):
        ws[col].font = HEADER_FONT
        ws[col].fill = HEADER_FILL
    r += 1
    for tab, meta in file_meta.items():
        for col_label in meta['sources'].keys():
            ws.cell(r, 1, value=tab)
            ws.cell(r, 2, value=col_label)
            ws.cell(r, 3, value=meta['sources'][col_label])
            lu = meta['last_updates'][col_label]
            ws.cell(r, 4, value=str(lu) if lu else '-')
            r += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 14


# ============================================================
# Top-level
# ============================================================
def build_workbook(out_path: Path, tabs: dict, fetch_fn,
                   start: date, end: date, name: str):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    file_meta = {}
    for tab_name, cols in tabs.items():
        ws = wb.create_sheet(title=tab_name[:31])
        meta = build_tab(ws, tab_name, cols, fetch_fn, start, end)
        file_meta[tab_name] = meta
    ws_meta = wb.create_sheet(title='_meta')
    build_meta(ws_meta, file_meta, name)
    wb.save(out_path)
    total_rows = sum(m['rows'] for m in file_meta.values())
    log.info(f'Wrote {out_path.relative_to(PROJECT_ROOT)} ({len(tabs)} tabs, '
             f'{total_rows} Fridays)')


def main():
    start = date(2019, 1, 1)
    end = date.today()

    # Feedstock view-based workbooks
    build_workbook(
        PROJECT_ROOT / 'models/Oilseeds/us_veg_oil_prices.xlsx',
        VEG_OIL_TABS, fetch_feedstock_series, start, end,
        'US vegetable oil prices',
    )
    build_workbook(
        PROJECT_ROOT / 'models/Fats and Greases/us_animal_fat_prices.xlsx',
        ANIMAL_FAT_TABS, fetch_feedstock_series, start, end,
        'US animal fat + protein prices',
    )

    # Fuel + credit use a different fetch function (closure capturing table)
    def fuel_fetch(column): return fetch_table_column('fuel', column)
    def credit_fetch(column): return fetch_table_column('credit', column)

    build_workbook(
        PROJECT_ROOT / 'models/Biofuels/us_fuel_prices.xlsx',
        FUEL_TABS, fuel_fetch, start, end,
        'US fuel prices (ULSD / B100 / RD / Jet A)',
    )
    build_workbook(
        PROJECT_ROOT / 'models/Biofuels/us_credit_prices.xlsx',
        CREDIT_TABS, credit_fetch, start, end,
        'US biofuel credit prices (RIN / LCFS / CFP / CFS)',
    )


if __name__ == '__main__':
    main()
