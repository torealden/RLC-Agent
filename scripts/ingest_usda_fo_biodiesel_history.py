"""
USDA Fats & Oils — Biodiesel Feedstock History Ingest
======================================================

One-off ingest of pre-2012 biodiesel feedstock consumption data from
Tore's models/Oilseeds/us_oilseed_crush.xlsm (Census Crush tab).

The Census Crush tab cross-references USDA Fats & Oils Selected End-Use
data, which reports methyl-ester (biodiesel) consumption by feedstock
back to 2006-2007 — earlier than EIA's old biofuels report covers.

Coverage filled:
  - Soybean Oil → biodiesel: 2006-01 onward
  - Inedible Tallow / Yellow Grease / Other Grease → biodiesel: 2007-01 onward
  - Cottonseed Oil → biodiesel: 2007 onward (often '(D)' = disclosed)

Stops at July 2011 (end of the xlsm sheet); EIA old_table3 takes over
from Jan 2012.

Upserts into bronze.eia_feedstock_monthly with:
  - source_sheet = 'usda_fo'
  - plant_type   = 'biodiesel'

Usage:
  python scripts/ingest_usda_fo_biodiesel_history.py
"""

from __future__ import annotations

import logging
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / '.env')

from src.services.database.db_config import get_connection

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger('ingest_usda_fo')

XLSM_PATH = PROJECT_ROOT / 'models' / 'Oilseeds' / 'us_oilseed_crush.xlsm'

# Column index -> (feedstock_name) mapping (1-indexed to match openpyxl)
# All map to plant_type='biodiesel'
COL_MAP = {
    221: 'Soybean Oil',       # Total methyl esters from SBO
    243: 'Tallow',            # Methyl Esters - Inedible Tallow (a.k.a. IBFT)
    244: 'Yellow Grease',     # Methyl Esters - Yellow Grease
    245: 'Other Grease',      # Methyl Esters - Other Grease
    257: 'Cottonseed Oil',    # Methyl esters (under cottonseed section)
}


def _parse_value(val):
    """Convert xlsm cell value -> (quantity_mil_lbs, is_withheld, is_no_data)."""
    if val is None or val == '':
        return None, False, True
    if isinstance(val, str):
        s = val.strip()
        # USDA uses (D) for disclosed/withheld, (Z) for less than half unit,
        # (NA) for not available. Treat like EIA's 'W'.
        if s in ('(D)', 'D', '(d)'):
            return None, True, False
        if s in ('(Z)', 'Z'):
            return 0.25, False, False  # "less than half"
        if s in ('(NA)', 'NA', 'N/A', '-', '--'):
            return None, False, True
        if s.startswith('#'):  # #VALUE!, #N/A, etc.
            return None, False, True
        try:
            return float(s.replace(',', '')), False, False
        except ValueError:
            return None, False, True
    try:
        return float(val), False, False
    except (ValueError, TypeError):
        return None, False, True


UPSERT_SQL = """
    INSERT INTO bronze.eia_feedstock_monthly
        (year, month, source_sheet, feedstock_name, plant_type,
         quantity_mil_lbs, is_withheld, is_no_data, source_file, collected_at)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
    ON CONFLICT (year, month, source_sheet, feedstock_name, plant_type)
    DO UPDATE SET
        quantity_mil_lbs = EXCLUDED.quantity_mil_lbs,
        is_withheld = EXCLUDED.is_withheld,
        is_no_data = EXCLUDED.is_no_data,
        source_file = EXCLUDED.source_file,
        collected_at = NOW()
    RETURNING (xmax = 0) AS is_insert
"""


def main():
    if not XLSM_PATH.exists():
        logger.error(f"File not found: {XLSM_PATH}")
        sys.exit(1)

    logger.info(f"Opening {XLSM_PATH}")
    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, read_only=True,
                                keep_vba=False)
    ws = wb['Census Crush']
    logger.info(f"Census Crush: {ws.max_row} rows, {ws.max_column} cols")

    records = []
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if date_val is None or not hasattr(date_val, 'year'):
            continue
        year = date_val.year
        month = date_val.month

        # Skip rows >= 2012 to avoid colliding with EIA old_table3 history.
        # 2012+ EIA data is more authoritative since it's the published source.
        if year >= 2012:
            continue

        for col_idx, feedstock_name in COL_MAP.items():
            if col_idx > ws.max_column:
                continue
            val = ws.cell(r, col_idx).value
            qty, is_withheld, is_no_data = _parse_value(val)

            # Skip pure no-data rows (don't pollute bronze with rows that
            # are just "no observation yet")
            if qty is None and not is_withheld and is_no_data:
                continue

            records.append({
                'year': year,
                'month': month,
                'source_sheet': 'usda_fo',
                'feedstock_name': feedstock_name,
                'plant_type': 'biodiesel',
                'quantity_mil_lbs': qty,
                'is_withheld': is_withheld,
                'is_no_data': is_no_data,
            })

    wb.close()
    logger.info(f"Parsed {len(records)} records (pre-2012, biodiesel feedstock)")

    if not records:
        logger.warning("No records to ingest — exiting")
        return

    # Year coverage summary
    years = sorted({r['year'] for r in records})
    logger.info(f"Year coverage: {years[0]}-{years[-1]} ({len(years)} years)")
    feedstocks = sorted({r['feedstock_name'] for r in records})
    logger.info(f"Feedstocks: {feedstocks}")

    inserted = updated = errors = 0
    with get_connection() as conn:
        with conn.cursor() as cur:
            for rec in records:
                try:
                    cur.execute("SAVEPOINT sp_usda")
                    cur.execute(UPSERT_SQL, (
                        rec['year'], rec['month'], rec['source_sheet'],
                        rec['feedstock_name'], rec['plant_type'],
                        rec['quantity_mil_lbs'],
                        rec['is_withheld'], rec['is_no_data'],
                        'us_oilseed_crush.xlsm:Census Crush',
                    ))
                    result = cur.fetchone()
                    is_insert = result[0] if isinstance(result, tuple) else result['is_insert']
                    if is_insert:
                        inserted += 1
                    else:
                        updated += 1
                    cur.execute("RELEASE SAVEPOINT sp_usda")
                except Exception as e:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_usda")
                    logger.error(f"Save error {rec}: {e}")
                    errors += 1
            conn.commit()

    logger.info(f"Done: {inserted} inserted, {updated} updated, {errors} errors")


if __name__ == '__main__':
    main()
