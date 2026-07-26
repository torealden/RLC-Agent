"""Brazil soybean-complex MONTHLY block -> merged into the contract flat files.

The monthly block is the actual deliverable (annual PSD is only the control total it rakes to). This
assembles the monthly ACTUALS that exist and merges them into the annual brazil_<commodity>_flat.xlsx
long tabs, so annual (control) + monthly (detail) coexist via period_type. The vintage ladder resolves
overlaps; the balance sheet reads the long tab through its ff_ mirror.

Sources (verified 2026-07-26):
  - gold.abiove_soy_complex_monthly : crush (2012+), oil/meal production (2025+), oil/meal/seed stocks
    (2021+). ABIOVE labels marketing_year = CALENDAR year, period = calendar month.
  - bronze.comexstat_trade          : Brazil exports by commodity/month (THIN — ~2025-2026, exports
    only, no imports/history yet; closing this = backfilling the comexstat collector).
  - derived: oil/meal production pre-2025 = crush x extraction yield (vintage MODEL_YIELD, rank 30, so
    any ABIOVE actual outranks it).

Crop-MY convention: Brazil soybean MY starts February -> crop_my(y,m) = y if m>=2 else y-1. All 12
months of a crop year share one marketing_year in column D (the US flat-file convention), period = MNN
(calendar month).

Run:  python scripts/write_brazil_soy_monthly.py
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

ROOT = Path(r"C:/dev/RLC-Agent")
sys.path.insert(0, str(ROOT))
from src.services.database.db_config import get_connection  # noqa: E402

FOLDER = ROOT / "models" / "Oilseeds" / "Brazil"
LONG_COLS = ["commodity", "class", "series", "marketing_year", "period_type", "period",
             "vintage", "vintage_rank", "value", "unit", "source"]
SUPPLY_SERIES = {"beginning_stocks", "production", "imports", "ending_stocks"}
OIL_YIELD, MEAL_YIELD = 0.185, 0.770   # per MT soybeans crushed (Brazil); ~4.5% hull/loss

# ABIOVE series -> (target commodity, target series)
ABIOVE_MAP = {
    "crush":           ("soybeans", "crush"),
    "seed_stocks":     ("soybeans", "ending_stocks"),
    "oil_production":  ("soybean_oil", "production"),
    "oil_stocks":      ("soybean_oil", "ending_stocks"),
    "meal_production": ("soybean_meal", "production"),
    "meal_stocks":     ("soybean_meal", "ending_stocks"),
}


def crop_my(cal_year, month):
    return cal_year if month >= 2 else cal_year - 1


def q(sql, args=()):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, args)
            data = cur.fetchall()
            if data and isinstance(data[0], dict):
                return [dict(r) for r in data]
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, r)) for r in data]


def build_rows():
    """Return {commodity: [long-row dicts]} of monthly rows to append."""
    out = {"soybeans": [], "soybean_oil": [], "soybean_meal": []}

    # 1) ABIOVE actuals
    ab = q("""SELECT marketing_year AS cal_year, period AS month, series, vintage, vintage_rank,
                     value, unit
              FROM gold.abiove_soy_complex_monthly
              WHERE period_type='month' AND series = ANY(%s)""",
           (list(ABIOVE_MAP),))
    crush_by = {}   # (cal_year, month) -> crush value, to derive production
    for r in ab:
        commodity, series = ABIOVE_MAP[r["series"]]
        y, m = r["cal_year"], r["month"]
        out[commodity].append(dict(
            commodity=commodity, **{"class": "ALL"}, series=series, marketing_year=crop_my(y, m),
            period_type="cal_month", period=f"M{m:02d}", vintage=r["vintage"],
            vintage_rank=r["vintage_rank"], value=float(r["value"]), unit=r["unit"] or "1000 MT",
            source="ABIOVE"))
        if r["series"] == "crush":
            crush_by[(y, m)] = float(r["value"])

    # 2) derived oil/meal production where ABIOVE production is absent (pre-2025)
    have = {(commodity, r["cal_year"], r["month"])
            for r in ab if r["series"] in ("oil_production", "meal_production")
            for commodity in [ABIOVE_MAP[r["series"]][0]]}
    for (y, m), crush in crush_by.items():
        for commodity, yld in (("soybean_oil", OIL_YIELD), ("soybean_meal", MEAL_YIELD)):
            if (commodity, y, m) in have:
                continue
            out[commodity].append(dict(
                commodity=commodity, **{"class": "ALL"}, series="production",
                marketing_year=crop_my(y, m), period_type="cal_month", period=f"M{m:02d}",
                vintage="MODEL_YIELD", vintage_rank=30, value=round(crush * yld, 2),
                unit="1000 MT", source=f"ABIOVE_crush x {yld:.3f}"))

    # 3) comexstat exports (thin; exports only)
    cx = q("""SELECT commodity, year AS cal_year, month, SUM(weight_kg) AS kg
              FROM bronze.comexstat_trade
              WHERE flow='export' AND commodity = ANY(%s)
              GROUP BY commodity, year, month""",
           (["soybeans", "soybean_oil", "soybean_meal"],))
    for r in cx:
        y, m = r["cal_year"], r["month"]
        out[r["commodity"]].append(dict(
            commodity=r["commodity"], **{"class": "ALL"}, series="exports",
            marketing_year=crop_my(y, m), period_type="cal_month", period=f"M{m:02d}",
            vintage="COMEXSTAT", vintage_rank=90, value=round(float(r["kg"]) / 1e6, 2),
            unit="1000 MT", source="ComexStat"))
    return out


def append_to_flat(commodity, rows):
    fn = FOLDER / f"brazil_{commodity}_flat.xlsx"
    if not fn.exists():
        raise SystemExit(f"missing annual flat file {fn} — run write_psd_flat_file.py {commodity} BR first")
    wb = openpyxl.load_workbook(fn)
    tabs = {"supply": wb[f"{commodity}_supply"], "demand": wb[f"{commodity}_demand"]}
    added = {"supply": 0, "demand": 0}
    for row in rows:
        side = "supply" if row["series"] in SUPPLY_SERIES else "demand"
        ws = tabs[side]
        r = ws.max_row + 1
        for j, h in enumerate(LONG_COLS, 1):
            ws.cell(r, j, row["class"] if h == "class" else row[h])
        added[side] += 1
    # meta note
    ms = wb["_meta"]; mr = ms.max_row + 1
    ms.cell(mr, 1, "MONTHLY BLOCK").font = Font(bold=True)
    ms.cell(mr, 5, "ABIOVE crush/prod/stocks (rank 99) + comexstat exports (rank 90) + derived "
                   "oil/meal production pre-2025 (MODEL_YIELD rank 30). GAP: monthly IMPORTS + trade "
                   "history absent (comexstat thin 2025-26, exports only) -> backfill comexstat collector.")
    wb.save(fn)
    return added


def main():
    rows = build_rows()
    for commodity in ("soybeans", "soybean_oil", "soybean_meal"):
        added = append_to_flat(commodity, rows[commodity])
        mys = sorted({r["marketing_year"] for r in rows[commodity]})
        print(f"brazil_{commodity}_flat.xlsx  +{added['supply']} supply / +{added['demand']} demand "
              f"monthly rows  | MY {mys[0]}..{mys[-1]}")


if __name__ == "__main__":
    main()
