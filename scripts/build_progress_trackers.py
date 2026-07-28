"""Build the per-group BUILD-PROGRESS trackers + a master summary.

New process (2026-07-28): we copy data into new workbooks first, then backfill the automation. Each
country x complex build walks 10 steps. This tool emits one workbook per commodity-complex GROUP
(Food Grains, Feed Grains, Oilseeds, Energy). In each group workbook:
  - a "Summary" tab FIRST: one row per country x complex tab, per-step % + overall %, rolled up.
  - one tab per country x complex: a GRID of that complex's balance sheets (rows) x the 10 steps
    (cols). The analyst marks a cell with the checkbox glyph when a step is done; row/tab/summary
    percentages flow up by formula.

A master `models/_Progress_Summary.xlsx` links to each group workbook's Summary and rolls the four
groups into one comprehensive view.

Checkbox mechanism: a cell holds the DONE glyph (✓) or is blank, via a data-validation dropdown +
conditional formatting (green when ✓). openpyxl cannot write native Excel-365 checkbox controls; this
convention behaves the same and rolls up cleanly with COUNTIF. (An analyst may drop a native 365
checkbox on top later; if they do, switch the glyph convention to TRUE/FALSE and the COUNTIFs with it.)

Oilseeds tabs are derived from `build_pepsi_coverage_tracker.COMPLEXES` so the universe stays in sync
with the coverage matrix. Food/Feed/Energy carry a STARTER tab set to prune (no canonical source yet).

Run:  python scripts/build_progress_trackers.py
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import importlib.util
ROOT = Path(r"C:/dev/RLC-Agent")
MODELS = ROOT / "models"

# import the coverage universe without running its __main__
_spec = importlib.util.spec_from_file_location("pct", ROOT / "scripts" / "build_pepsi_coverage_tracker.py")
PCT = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(PCT)

DONE_GLYPH = "✓"   # ✓
STEPS = [
    ("1", "Copy data"),
    ("2", "Add formulas"),
    ("3", "Extend data"),
    ("4", "DB schema"),
    ("5", "Import history"),
    ("6", "Wire API / update proc"),
    ("7", "Flat files"),
    ("8", "Update script"),
    ("9", "Link flat → balance"),
    ("10", "Forecasts"),
]
NSTEPS = len(STEPS)

# --- palette (internal green + neutrals) ---
GREEN = "3C7D22"; INK = "1B2A4A"
HDR_FILL = PatternFill("solid", fgColor=INK)
GRP_FILL = PatternFill("solid", fgColor=GREEN)
DONE_FILL = PatternFill("solid", fgColor="E2EFD9")
TOT_FILL = PatternFill("solid", fgColor="F2F2F0")
THIN = Side(style="thin", color="D4D4CE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_F = Font(bold=True, size=13, color=GREEN, name="Calibri")
SUB_F = Font(italic=True, size=9, color="6E7178", name="Calibri")
HDR_F = Font(bold=True, color="FFFFFF", name="Calibri")
BOLD = Font(bold=True, name="Calibri")
PLAIN = Font(name="Calibri")

# --- short display names for tabs ---
COUNTRY_SHORT = {"United States": "US", "Europe": "EU"}
COMPLEX_SHORT = {"Rapeseed / Canola": "Canola", "Sunflower": "Sun", "Corn Oil": "Corn Oil"}


def short_country(c): return COUNTRY_SHORT.get(c, c)
def short_complex(cx): return COMPLEX_SHORT.get(cx, cx)


# ---- assemble the tab universe per group ---------------------------------------------------------
# a "tab" = dict(country, complex, tier, sheets=[grid-row labels], note)

def oilseed_tabs():
    tabs = []
    for cx, cfg in PCT.COMPLEXES.items():
        for tier in ("A", "B", "D"):
            for country in cfg.get(tier, []):
                sheets = [s for s in PCT.SHEET_TYPES if s in PCT.sheets_for(cx, country, tier)]
                note = PCT.TIER_LABEL[tier]
                if (cx, country) in PCT.DONE:
                    note += " · US template built"
                tabs.append(dict(country=country, complex=cx, tier=tier, sheets=sheets, note=note))
    return tabs


GRAIN_SHEETS = ["S&D", "Trade", "Stocks"]
CORN_SHEETS = ["S&D", "Trade", "Stocks", "Ethanol / DDGS"]
ENERGY_SHEETS = ["Balance (S&D)", "Feedstock / Trade"]


def _mk(country, cx, sheets, tier="", note=""):
    return dict(country=country, complex=cx, tier=tier, sheets=sheets, note=note or "starter — prune")


def foodgrain_tabs():
    wheat = ["United States", "Europe", "Russia", "Canada", "Australia", "Argentina", "Ukraine", "India", "China"]
    rice = ["India", "Thailand", "Vietnam", "United States", "China"]
    return ([_mk(c, "Wheat", GRAIN_SHEETS) for c in wheat] +
            [_mk(c, "Rice", GRAIN_SHEETS) for c in rice])


def feedgrain_tabs():
    corn = ["United States", "Brazil", "Argentina", "Ukraine", "China", "Europe"]
    sorghum = ["United States", "Argentina"]
    barley = ["Europe", "Russia", "Australia", "Canada"]
    return ([_mk(c, "Corn", CORN_SHEETS) for c in corn] +
            [_mk(c, "Sorghum", GRAIN_SHEETS) for c in sorghum] +
            [_mk(c, "Barley", GRAIN_SHEETS) for c in barley])


def energy_tabs():
    return [
        _mk("US", "Ethanol", ENERGY_SHEETS),
        _mk("US", "Biodiesel", ENERGY_SHEETS),
        _mk("US", "Renewable Diesel", ENERGY_SHEETS),
        _mk("US", "SAF", ENERGY_SHEETS),
        _mk("US", "Petroleum", ENERGY_SHEETS),
        _mk("Brazil", "Ethanol", ENERGY_SHEETS),
        _mk("EU", "Biodiesel", ENERGY_SHEETS),
    ]


GROUPS = {
    "Oilseeds":    dict(file="_Progress_Oilseeds.xlsx",   tabs=oilseed_tabs,  scaffold=False),
    "Food Grains": dict(file="_Progress_FoodGrains.xlsx", tabs=foodgrain_tabs, scaffold=True),
    "Feed Grains": dict(file="_Progress_FeedGrains.xlsx", tabs=feedgrain_tabs, scaffold=True),
    "Energy":      dict(file="_Progress_Energy.xlsx",      tabs=energy_tabs,    scaffold=True),
}

MASTER = MODELS / "_Progress_Summary.xlsx"


def tab_title(t):
    """<=31-char unique sheet name."""
    name = f"{short_country(t['country'])} {short_complex(t['complex'])}"
    return name[:31]


# ---- one country x complex grid tab --------------------------------------------------------------
GRID_HDR_ROW = 4
GRID_START = 5
OVERALL_CELL = "C2"          # tab overall % lives here; the Summary tab reads it
STEP_COL0 = 2                # first step column (B)
PCT_COL = STEP_COL0 + NSTEPS # column after the 10 steps = "% complete"


def build_grid_tab(wb, t):
    ws = wb.create_sheet(tab_title(t))
    ws["A1"] = f"{t['country'].upper()} {t['complex'].upper()} — build progress"
    ws["A1"].font = TITLE_F
    ws["A2"] = "Overall:"; ws["A2"].font = BOLD
    nrows = len(t["sheets"])
    grid_end = GRID_START + nrows - 1
    first = f"{get_column_letter(STEP_COL0)}{GRID_START}"
    last = f"{get_column_letter(STEP_COL0 + NSTEPS - 1)}{grid_end}"
    ws[OVERALL_CELL] = f'=IF({nrows}=0,0,COUNTIF({first}:{last},"{DONE_GLYPH}")/({nrows}*{NSTEPS}))'
    ws[OVERALL_CELL].number_format = "0%"; ws[OVERALL_CELL].font = Font(bold=True, color=GREEN, name="Calibri")
    ws["D2"] = f"mark cells with {DONE_GLYPH} as each step completes"; ws["D2"].font = SUB_F

    # header
    ws.cell(GRID_HDR_ROW, 1, "Balance Sheet").font = HDR_F
    ws.cell(GRID_HDR_ROW, 1).fill = HDR_FILL; ws.cell(GRID_HDR_ROW, 1).border = BORDER
    for i, (num, label) in enumerate(STEPS):
        c = ws.cell(GRID_HDR_ROW, STEP_COL0 + i, num)
        c.font = HDR_F; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        ws.cell(GRID_HDR_ROW, STEP_COL0 + i).comment = None
    pc = ws.cell(GRID_HDR_ROW, PCT_COL, "%"); pc.font = HDR_F; pc.fill = HDR_FILL
    pc.border = BORDER; pc.alignment = Alignment(horizontal="center")

    # data-validation dropdown ("", ✓) applied to the whole grid
    dv = DataValidation(type="list", formula1=f'"{DONE_GLYPH}"', allow_blank=True)
    ws.add_data_validation(dv)

    for r_off, sheet in enumerate(t["sheets"]):
        r = GRID_START + r_off
        ws.cell(r, 1, sheet).font = PLAIN; ws.cell(r, 1).border = BORDER
        for i in range(NSTEPS):
            cell = ws.cell(r, STEP_COL0 + i)
            cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
            dv.add(cell)
        rowrange = f"{get_column_letter(STEP_COL0)}{r}:{get_column_letter(STEP_COL0 + NSTEPS - 1)}{r}"
        pcell = ws.cell(r, PCT_COL, f'=COUNTIF({rowrange},"{DONE_GLYPH}")/{NSTEPS}')
        pcell.number_format = "0%"; pcell.border = BORDER; pcell.alignment = Alignment(horizontal="center")

    # TAB TOTAL row: per-step % across the sheets + overall
    tr = grid_end + 1
    ws.cell(tr, 1, "TAB TOTAL").font = BOLD; ws.cell(tr, 1).fill = TOT_FILL; ws.cell(tr, 1).border = BORDER
    for i in range(NSTEPS):
        col = get_column_letter(STEP_COL0 + i)
        colrange = f"{col}{GRID_START}:{col}{grid_end}"
        c = ws.cell(tr, STEP_COL0 + i, f'=IF({nrows}=0,0,COUNTIF({colrange},"{DONE_GLYPH}")/{nrows})')
        c.number_format = "0%"; c.fill = TOT_FILL; c.border = BORDER; c.font = BOLD
        c.alignment = Alignment(horizontal="center")
    tc = ws.cell(tr, PCT_COL, f"={OVERALL_CELL}"); tc.number_format = "0%"
    tc.fill = TOT_FILL; tc.border = BORDER; tc.font = BOLD; tc.alignment = Alignment(horizontal="center")

    # conditional formatting: green fill where a grid cell = ✓
    grid_range = f"{first}:{last}"
    ws.conditional_formatting.add(grid_range, CellIsRule(operator="equal",
        formula=[f'"{DONE_GLYPH}"'], fill=DONE_FILL, font=Font(color=GREEN, bold=True, name="Calibri")))

    ws.column_dimensions["A"].width = 18
    for i in range(NSTEPS):
        ws.column_dimensions[get_column_letter(STEP_COL0 + i)].width = 5
    ws.column_dimensions[get_column_letter(PCT_COL)].width = 8
    ws.freeze_panes = f"B{GRID_START}"
    return ws.title, tr


# ---- the per-group Summary tab --------------------------------------------------------------------
def build_group_summary(wb, group, tabs, tab_meta):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = f"{group} — build progress summary"; ws["A1"].font = TITLE_F
    ws["A2"] = (f"{len(tabs)} country × complex tabs. Each cell = COUNTIF of {DONE_GLYPH} on that "
                "tab. Fill the checkboxes on the country tabs; this rolls up automatically.")
    ws["A2"].font = SUB_F

    headers = ["Complex", "Country", "Tier / note"] + [n for n, _ in STEPS] + ["Overall"]
    hr = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(hr, j, h); c.fill = HDR_FILL; c.font = HDR_F; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # group total row (row 5), then one row per tab
    gtr = hr + 1
    ws.cell(gtr, 1, "GROUP TOTAL").font = Font(bold=True, color="FFFFFF", name="Calibri")
    ws.cell(gtr, 1).fill = GRP_FILL
    r0 = gtr + 1
    overall_cells = []
    for k, (t, (title, totrow)) in enumerate(zip(tabs, tab_meta)):
        r = r0 + k
        ws.cell(r, 1, t["complex"]).font = PLAIN; ws.cell(r, 1).border = BORDER
        ws.cell(r, 2, t["country"]).font = PLAIN; ws.cell(r, 2).border = BORDER
        ws.cell(r, 3, t["note"]).font = Font(size=9, color="6E7178", name="Calibri"); ws.cell(r, 3).border = BORDER
        q = f"'{title}'"
        for i in range(NSTEPS):
            src = f"{q}!{get_column_letter(STEP_COL0 + i)}{totrow}"
            c = ws.cell(r, 4 + i, f"={src}"); c.number_format = "0%"; c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        oc = ws.cell(r, 4 + NSTEPS, f"={q}!{OVERALL_CELL}"); oc.number_format = "0%"
        oc.border = BORDER; oc.font = BOLD; oc.alignment = Alignment(horizontal="center")
        overall_cells.append(f"{get_column_letter(4 + NSTEPS)}{r}")

    # group total = average of the per-tab step %s and overall
    last_r = r0 + len(tabs) - 1
    for i in range(NSTEPS):
        col = get_column_letter(4 + i)
        c = ws.cell(gtr, 4 + i, f"=IFERROR(AVERAGE({col}{r0}:{col}{last_r}),0)")
        c.number_format = "0%"; c.fill = GRP_FILL; c.font = HDR_F; c.alignment = Alignment(horizontal="center")
    oc = ws.cell(gtr, 4 + NSTEPS, f"=IFERROR(AVERAGE({get_column_letter(4+NSTEPS)}{r0}:{get_column_letter(4+NSTEPS)}{last_r}),0)")
    oc.number_format = "0%"; oc.fill = GRP_FILL; oc.font = HDR_F; oc.alignment = Alignment(horizontal="center")

    widths = [18, 15, 30] + [5] * NSTEPS + [9]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "D5"
    return gtr, 4 + NSTEPS  # (group-total row, overall column) for the master to link


def build_group(group, cfg, force=False):
    """Create the group workbook. Clobber-safe: an EXISTING file is left untouched (it holds the
    analyst's checkmarks) unless force=True. Returns (path, ntabs, created?)."""
    out = MODELS / cfg["file"]
    if out.exists() and not force:
        return out, None, False
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    tabs = cfg["tabs"]()
    tab_meta = [build_grid_tab(wb, t) for t in tabs]
    build_group_summary(wb, group, tabs, tab_meta)
    wb.save(out)
    return out, len(tabs), True


def _group_progress(path):
    """Read a group workbook (openpyxl) and return (ntabs, overall%) computed statically from the ✓
    glyphs — never rewrites the file. overall = mean of per-tab overalls (matches the in-file AVERAGE)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tab_overalls = []
    ntabs = 0
    for ws in wb.worksheets:
        if ws.title == "Summary":
            continue
        ntabs += 1
        done = total = 0
        for r in range(GRID_START, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if label in (None, "", "TAB TOTAL"):
                break
            for i in range(NSTEPS):
                total += 1
                if ws.cell(r, STEP_COL0 + i).value == DONE_GLYPH:
                    done += 1
        tab_overalls.append(done / total if total else 0.0)
    wb.close()
    overall = sum(tab_overalls) / len(tab_overalls) if tab_overalls else 0.0
    return ntabs, overall


def refresh_master(stamp):
    """Static snapshot across the four group files — READS them, writes values (no live links, which
    openpyxl can't register and which are the fragility this project is removing)."""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"
    ws["A1"] = "RLC Model Build — Master Progress"; ws["A1"].font = TITLE_F
    ws["A2"] = (f"Snapshot across the four group workbooks, refreshed {stamp}. Re-run "
                "scripts/build_progress_trackers.py to refresh (it reads the group files, never "
                "overwrites your checkmarks).")
    ws["A2"].font = SUB_F
    for j, h in enumerate(["Group", "Tabs", "Overall %", "Workbook"], 1):
        c = ws.cell(4, j, h); c.fill = HDR_FILL; c.font = HDR_F; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    r = 5; ov = []
    for group, cfg in GROUPS.items():
        p = MODELS / cfg["file"]
        ntabs, overall = _group_progress(p) if p.exists() else (0, 0.0)
        ws.cell(r, 1, group).font = BOLD; ws.cell(r, 1).border = BORDER
        ws.cell(r, 2, ntabs).font = PLAIN; ws.cell(r, 2).border = BORDER
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        c = ws.cell(r, 3, overall); c.number_format = "0%"; c.border = BORDER
        c.font = Font(bold=True, color=GREEN, name="Calibri"); c.alignment = Alignment(horizontal="center")
        ws.cell(r, 4, cfg["file"]).font = Font(size=9, color="6E7178", name="Calibri"); ws.cell(r, 4).border = BORDER
        ov.append(overall); r += 1
    ws.cell(r, 1, "ALL GROUPS").font = HDR_F; ws.cell(r, 1).fill = GRP_FILL
    tc = ws.cell(r, 3, sum(ov) / len(ov) if ov else 0.0); tc.number_format = "0%"
    tc.fill = GRP_FILL; tc.font = HDR_F; tc.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCD", (16, 8, 12, 30)):
        ws.column_dimensions[col].width = w
    wb.save(MASTER)


def main():
    import sys, datetime
    force = "--force" in sys.argv
    for group, cfg in GROUPS.items():
        out, ntabs, created = build_group(group, cfg, force=force)
        tag = "scaffold" if cfg["scaffold"] else "FULL"
        state = f"{ntabs:2d} tabs -> {out.name}" if created else f"exists, kept (use --force to rebuild)"
        print(f"  {group:12s} [{tag:8s}] {state}")
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    refresh_master(stamp)
    print(f"  Master  refreshed {stamp} -> {MASTER.name}")


if __name__ == "__main__":
    main()
