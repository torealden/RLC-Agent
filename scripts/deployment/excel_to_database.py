#!/usr/bin/env python3
"""
Excel to Database ETL Pipeline

Loads commodity data from Excel spreadsheets into SQLite database.
Designed to work with RLC's model files (balance sheets, prices, etc.)

Features:
- Auto-detects data type (time series, balance sheet, price data)
- Handles various Excel formats and structures
- Tracks import history to avoid duplicates
- Can be run incrementally or full refresh

Usage:
    python deployment/excel_to_database.py --scan          # Scan and list Excel files
    python deployment/excel_to_database.py --load          # Load all Excel files
    python deployment/excel_to_database.py --file "path"   # Load specific file
    python deployment/excel_to_database.py --status        # Show database status
"""

import argparse
import sqlite3
import logging
import json
import hashlib
import re
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field

# Try to import openpyxl and pandas
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Project paths
PROJECT_ROOT = Path(__file__).parent.parent
MODELS_DIR = PROJECT_ROOT / "Models"
DATA_DIR = PROJECT_ROOT / "data"
DB_PATH = DATA_DIR / "rlc_commodities.db"


@dataclass
class SheetInfo:
    """Information about an Excel sheet"""
    name: str
    rows: int
    cols: int
    data_type: str  # 'time_series', 'balance_sheet', 'price_data', 'reference', 'unknown'
    headers: List[str] = field(default_factory=list)
    date_column: Optional[str] = None
    value_columns: List[str] = field(default_factory=list)
    preview: List[Dict] = field(default_factory=list)


@dataclass
class ExcelFileInfo:
    """Information about an Excel file"""
    path: Path
    category: str  # biofuels, oilseeds, feed_grains, etc.
    subcategory: str
    file_hash: str
    modified_time: datetime
    sheets: List[SheetInfo] = field(default_factory=list)
    total_rows: int = 0
    loadable: bool = True
    load_errors: List[str] = field(default_factory=list)


class ExcelAnalyzer:
    """Analyzes Excel files to understand their structure"""

    # Common date-like column names
    DATE_PATTERNS = [
        r'^date$', r'^dates?$', r'^period$', r'^month$', r'^year$',
        r'^week$', r'^day$', r'^time$', r'^marketing.?year$', r'^my$',
        r'^crop.?year$', r'^\d{4}$', r'^\d{4}/\d{2}$'
    ]

    # Patterns indicating time series data
    TIME_SERIES_HEADERS = ['date', 'period', 'month', 'year', 'week', 'day']

    # Patterns indicating balance sheet data
    BALANCE_SHEET_METRICS = [
        'production', 'imports', 'exports', 'crush', 'feed', 'food',
        'domestic', 'consumption', 'ending stocks', 'beginning stocks',
        'total supply', 'total demand', 'stocks/use', 'area', 'yield',
        'harvested', 'planted'
    ]

    # Price-related headers
    PRICE_HEADERS = ['price', 'bid', 'offer', 'settle', 'close', 'open', 'high', 'low']

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

    def get_file_hash(self, filepath: Path) -> str:
        """Get hash of file for change detection"""
        hasher = hashlib.md5()
        with open(filepath, 'rb') as f:
            # Read in chunks for large files
            for chunk in iter(lambda: f.read(8192), b''):
                hasher.update(chunk)
        return hasher.hexdigest()

    def categorize_file(self, filepath: Path) -> Tuple[str, str]:
        """Categorize file based on its path"""
        parts = filepath.parts

        # Find category from path
        category = 'general'
        subcategory = 'unknown'

        for i, part in enumerate(parts):
            part_lower = part.lower()
            if 'biofuel' in part_lower:
                category = 'biofuels'
            elif 'oilseed' in part_lower:
                category = 'oilseeds'
            elif 'feed grain' in part_lower or 'feedgrain' in part_lower:
                category = 'feed_grains'
            elif 'food grain' in part_lower or 'foodgrain' in part_lower:
                category = 'food_grains'
            elif 'fats' in part_lower or 'grease' in part_lower:
                category = 'fats_greases'
            elif 'macro' in part_lower:
                category = 'macro'

        # Subcategory from filename
        filename = filepath.stem.lower()
        if 'balance' in filename:
            subcategory = 'balance_sheet'
        elif 'price' in filename:
            subcategory = 'prices'
        elif 'trade' in filename:
            subcategory = 'trade'
        elif 'crush' in filename:
            subcategory = 'crush'
        elif 'production' in filename:
            subcategory = 'production'
        elif 'export' in filename:
            subcategory = 'exports'
        elif 'import' in filename:
            subcategory = 'imports'
        else:
            subcategory = 'other'

        return category, subcategory

    def detect_data_type(self, headers: List[str], sample_rows: List[List]) -> str:
        """Detect what type of data the sheet contains"""
        headers_lower = [str(h).lower() if h else '' for h in headers]

        # Check for time series (has date column)
        has_date = any(
            any(re.match(pattern, h, re.IGNORECASE) for pattern in self.DATE_PATTERNS)
            for h in headers_lower
        )

        # Check for balance sheet metrics
        balance_keywords = sum(
            1 for h in headers_lower
            for keyword in self.BALANCE_SHEET_METRICS
            if keyword in h
        )

        # Check for price data
        price_keywords = sum(
            1 for h in headers_lower
            for keyword in self.PRICE_HEADERS
            if keyword in h
        )

        # Decision logic
        if balance_keywords >= 3:
            return 'balance_sheet'
        elif price_keywords >= 2:
            return 'price_data'
        elif has_date:
            return 'time_series'
        elif len(headers) > 0:
            return 'reference'
        else:
            return 'unknown'

    def analyze_sheet(self, sheet, max_rows: int = 1000) -> SheetInfo:
        """Analyze a single sheet"""
        # Find headers (usually first row with data)
        headers = []
        header_row = 1

        for row_idx in range(1, min(10, sheet.max_row + 1)):
            row = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(20, sheet.max_column + 1))]
            non_empty = [c for c in row if c is not None]
            if len(non_empty) >= 3:
                headers = row
                header_row = row_idx
                break

        # Get sample data (rows after header)
        sample_rows = []
        for row_idx in range(header_row + 1, min(header_row + 6, sheet.max_row + 1)):
            row = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(20, sheet.max_column + 1))]
            sample_rows.append(row)

        # Detect data type
        data_type = self.detect_data_type(headers, sample_rows)

        # Find date column
        date_column = None
        for i, h in enumerate(headers):
            if h and any(re.match(p, str(h), re.IGNORECASE) for p in self.DATE_PATTERNS):
                date_column = str(h)
                break

        # Find value columns (numeric)
        value_columns = []
        for i, h in enumerate(headers):
            if h and i < len(sample_rows[0]) if sample_rows else False:
                sample_val = sample_rows[0][i] if sample_rows else None
                if isinstance(sample_val, (int, float)):
                    value_columns.append(str(h))

        # Create preview
        preview = []
        if headers and sample_rows:
            for row in sample_rows[:3]:
                row_dict = {}
                for i, h in enumerate(headers):
                    if h and i < len(row):
                        row_dict[str(h)] = row[i]
                preview.append(row_dict)

        return SheetInfo(
            name=sheet.title,
            rows=sheet.max_row,
            cols=sheet.max_column,
            data_type=data_type,
            headers=[str(h) for h in headers if h],
            date_column=date_column,
            value_columns=value_columns[:10],  # Limit to 10
            preview=preview
        )

    def analyze_file(self, filepath: Path) -> ExcelFileInfo:
        """Analyze an Excel file"""
        try:
            category, subcategory = self.categorize_file(filepath)
            file_hash = self.get_file_hash(filepath)
            modified_time = datetime.fromtimestamp(filepath.stat().st_mtime)

            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

            sheets = []
            total_rows = 0

            for sheet_name in wb.sheetnames:
                try:
                    sheet = wb[sheet_name]
                    sheet_info = self.analyze_sheet(sheet)
                    sheets.append(sheet_info)
                    total_rows += sheet_info.rows
                except Exception as e:
                    logger.warning(f"Error analyzing sheet {sheet_name}: {e}")

            wb.close()

            return ExcelFileInfo(
                path=filepath,
                category=category,
                subcategory=subcategory,
                file_hash=file_hash,
                modified_time=modified_time,
                sheets=sheets,
                total_rows=total_rows,
                loadable=True
            )

        except Exception as e:
            logger.error(f"Error analyzing {filepath}: {e}")
            return ExcelFileInfo(
                path=filepath,
                category='error',
                subcategory='error',
                file_hash='',
                modified_time=datetime.now(),
                loadable=False,
                load_errors=[str(e)]
            )


class DatabaseManager:
    """Manages the SQLite database"""

    def __init__(self, db_path: Path = DB_PATH):
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

    def get_connection(self) -> sqlite3.Connection:
        """Get database connection"""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        return conn

    def init_database(self):
        """Initialize database with required tables"""
        conn = self.get_connection()
        cursor = conn.cursor()

        # Core tables for Excel imports
        cursor.executescript("""
            -- Track imported Excel files
            CREATE TABLE IF NOT EXISTS excel_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path TEXT UNIQUE NOT NULL,
                file_hash TEXT,
                category TEXT,
                subcategory TEXT,
                sheets_count INTEGER,
                rows_imported INTEGER,
                import_status TEXT DEFAULT 'pending',
                last_imported TIMESTAMP,
                error_message TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            -- Generic time series data from Excel
            CREATE TABLE IF NOT EXISTS excel_time_series (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                series_name TEXT NOT NULL,
                date_value TEXT,
                numeric_value REAL,
                text_value TEXT,
                unit TEXT,
                category TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(source_file, sheet_name, series_name, date_value)
            );

            CREATE INDEX IF NOT EXISTS idx_ts_series ON excel_time_series(series_name);
            CREATE INDEX IF NOT EXISTS idx_ts_date ON excel_time_series(date_value);
            CREATE INDEX IF NOT EXISTS idx_ts_category ON excel_time_series(category);

            -- Balance sheet data (commodity, country, year, metric, value)
            CREATE TABLE IF NOT EXISTS balance_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT NOT NULL,
                commodity TEXT NOT NULL,
                country TEXT,
                region TEXT,
                marketing_year TEXT,
                calendar_year INTEGER,
                metric TEXT NOT NULL,
                value REAL,
                unit TEXT,
                source TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(source_file, commodity, country, marketing_year, metric)
            );

            CREATE INDEX IF NOT EXISTS idx_bs_commodity ON balance_sheets(commodity);
            CREATE INDEX IF NOT EXISTS idx_bs_country ON balance_sheets(country);
            CREATE INDEX IF NOT EXISTS idx_bs_year ON balance_sheets(marketing_year);
            CREATE INDEX IF NOT EXISTS idx_bs_metric ON balance_sheets(metric);

            -- Price data
            CREATE TABLE IF NOT EXISTS price_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT NOT NULL,
                symbol TEXT NOT NULL,
                commodity TEXT,
                price_type TEXT,  -- spot, futures, cash, basis
                location TEXT,
                date_value DATE NOT NULL,
                open REAL,
                high REAL,
                low REAL,
                close REAL,
                settle REAL,
                volume INTEGER,
                unit TEXT,
                currency TEXT DEFAULT 'USD',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(source_file, symbol, date_value, price_type)
            );

            CREATE INDEX IF NOT EXISTS idx_price_symbol ON price_history(symbol);
            CREATE INDEX IF NOT EXISTS idx_price_date ON price_history(date_value);
            CREATE INDEX IF NOT EXISTS idx_price_commodity ON price_history(commodity);

            -- Data catalog - tracks all available data series
            CREATE TABLE IF NOT EXISTS data_catalog (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                series_id TEXT UNIQUE NOT NULL,
                series_name TEXT NOT NULL,
                description TEXT,
                source_file TEXT,
                sheet_name TEXT,
                data_table TEXT,  -- which table holds this data
                commodity TEXT,
                country TEXT,
                category TEXT,
                frequency TEXT,  -- daily, weekly, monthly, annual
                unit TEXT,
                start_date TEXT,
                end_date TEXT,
                row_count INTEGER,
                last_updated TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE INDEX IF NOT EXISTS idx_catalog_commodity ON data_catalog(commodity);
            CREATE INDEX IF NOT EXISTS idx_catalog_category ON data_catalog(category);

            -- Relationships between data series
            CREATE TABLE IF NOT EXISTS data_relationships (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                series_id_1 TEXT NOT NULL,
                series_id_2 TEXT NOT NULL,
                relationship_type TEXT,  -- 'derived_from', 'related_to', 'component_of'
                formula TEXT,
                description TEXT,
                correlation REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(series_id_1, series_id_2, relationship_type)
            );
        """)

        conn.commit()
        conn.close()
        logger.info(f"Database initialized: {self.db_path}")

    def record_import(self, file_info: ExcelFileInfo, rows_imported: int, status: str, error: str = None):
        """Record an import attempt"""
        conn = self.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT OR REPLACE INTO excel_imports
            (file_path, file_hash, category, subcategory, sheets_count,
             rows_imported, import_status, last_imported, error_message)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            str(file_info.path),
            file_info.file_hash,
            file_info.category,
            file_info.subcategory,
            len(file_info.sheets),
            rows_imported,
            status,
            datetime.now().isoformat(),
            error
        ))

        conn.commit()
        conn.close()

    def get_import_status(self) -> Dict[str, Any]:
        """Get overall import status"""
        conn = self.get_connection()
        cursor = conn.cursor()

        # Get import counts
        cursor.execute("""
            SELECT import_status, COUNT(*) as count, SUM(rows_imported) as rows
            FROM excel_imports
            GROUP BY import_status
        """)
        status_counts = {row['import_status']: {'count': row['count'], 'rows': row['rows']}
                        for row in cursor.fetchall()}

        # Get table counts
        tables = ['excel_time_series', 'balance_sheets', 'price_history', 'data_catalog']
        table_counts = {}
        for table in tables:
            try:
                cursor.execute(f"SELECT COUNT(*) as count FROM {table}")
                table_counts[table] = cursor.fetchone()['count']
            except:
                table_counts[table] = 0

        conn.close()

        return {
            'import_status': status_counts,
            'table_counts': table_counts,
            'database_path': str(self.db_path),
            'database_size_mb': self.db_path.stat().st_size / (1024*1024) if self.db_path.exists() else 0
        }


class ExcelLoader:
    """Loads Excel data into database"""

    def __init__(self, db_manager: DatabaseManager):
        self.db = db_manager
        self.analyzer = ExcelAnalyzer()

    def load_time_series(self, filepath: Path, sheet_info: SheetInfo) -> int:
        """Load time series data from a sheet"""
        if not PANDAS_AVAILABLE:
            logger.warning("pandas required for time series loading")
            return 0

        try:
            df = pd.read_excel(filepath, sheet_name=sheet_info.name)

            if df.empty:
                return 0

            # Find date column
            date_col = None
            for col in df.columns:
                if any(re.match(p, str(col), re.IGNORECASE) for p in self.analyzer.DATE_PATTERNS):
                    date_col = col
                    break

            if date_col is None:
                # Try first column
                date_col = df.columns[0]

            # Melt to long format
            value_cols = [c for c in df.columns if c != date_col]
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()

            rows_inserted = 0
            conn = self.db.get_connection()
            cursor = conn.cursor()

            source_file = filepath.name

            for col in numeric_cols:
                if col == date_col:
                    continue

                for idx, row in df.iterrows():
                    try:
                        cursor.execute("""
                            INSERT OR IGNORE INTO excel_time_series
                            (source_file, sheet_name, series_name, date_value, numeric_value, category)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (
                            source_file,
                            sheet_info.name,
                            str(col),
                            str(row[date_col]) if pd.notna(row[date_col]) else None,
                            float(row[col]) if pd.notna(row[col]) else None,
                            filepath.parent.name  # Use parent folder as category
                        ))
                        rows_inserted += cursor.rowcount
                    except Exception as e:
                        pass  # Skip problematic rows

            conn.commit()
            conn.close()

            return rows_inserted

        except Exception as e:
            logger.error(f"Error loading time series from {filepath}/{sheet_info.name}: {e}")
            return 0

    def load_balance_sheet(self, filepath: Path, sheet_info: SheetInfo) -> int:
        """Load balance sheet data from a sheet"""
        if not PANDAS_AVAILABLE:
            return 0

        try:
            df = pd.read_excel(filepath, sheet_name=sheet_info.name)

            if df.empty:
                return 0

            # Detect commodity from filename
            filename_lower = filepath.stem.lower()
            commodity = 'unknown'
            for comm in ['soybean', 'corn', 'wheat', 'canola', 'palm', 'biodiesel', 'ethanol', 'tallow']:
                if comm in filename_lower:
                    commodity = comm
                    break

            # Detect country from filename or sheet name
            country = 'unknown'
            for c in ['us', 'brazil', 'argentina', 'china', 'eu', 'canada', 'world']:
                if c in filename_lower or c in sheet_info.name.lower():
                    country = c.upper() if len(c) <= 2 else c.title()
                    break

            rows_inserted = 0
            conn = self.db.get_connection()
            cursor = conn.cursor()

            source_file = filepath.name

            # Try to identify year columns (e.g., "2024/25" or "2024")
            year_cols = []
            for col in df.columns:
                col_str = str(col)
                if re.match(r'^\d{4}(/\d{2})?$', col_str):
                    year_cols.append(col)

            if year_cols:
                # Pivot table format - metrics in first column, years across
                metric_col = df.columns[0]

                for _, row in df.iterrows():
                    metric = str(row[metric_col]) if pd.notna(row[metric_col]) else ''
                    if not metric or metric == 'nan':
                        continue

                    for year_col in year_cols:
                        try:
                            value = row[year_col]
                            if pd.notna(value) and isinstance(value, (int, float)):
                                cursor.execute("""
                                    INSERT OR IGNORE INTO balance_sheets
                                    (source_file, commodity, country, marketing_year, metric, value)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                """, (
                                    source_file,
                                    commodity,
                                    country,
                                    str(year_col),
                                    metric,
                                    float(value)
                                ))
                                rows_inserted += cursor.rowcount
                        except:
                            pass

            conn.commit()
            conn.close()

            return rows_inserted

        except Exception as e:
            logger.error(f"Error loading balance sheet from {filepath}/{sheet_info.name}: {e}")
            return 0

    def load_file(self, filepath: Path) -> Tuple[int, str]:
        """Load all sheets from an Excel file"""
        try:
            file_info = self.analyzer.analyze_file(filepath)

            if not file_info.loadable:
                self.db.record_import(file_info, 0, 'error', '; '.join(file_info.load_errors))
                return 0, 'error'

            total_rows = 0

            for sheet in file_info.sheets:
                if sheet.data_type == 'balance_sheet':
                    rows = self.load_balance_sheet(filepath, sheet)
                elif sheet.data_type in ['time_series', 'price_data']:
                    rows = self.load_time_series(filepath, sheet)
                else:
                    rows = self.load_time_series(filepath, sheet)  # Default to time series

                total_rows += rows

            status = 'success' if total_rows > 0 else 'empty'
            self.db.record_import(file_info, total_rows, status)

            return total_rows, status

        except Exception as e:
            logger.error(f"Error loading {filepath}: {e}")
            return 0, 'error'


def scan_excel_files(directory: Path = MODELS_DIR) -> List[Path]:
    """Find all Excel files in directory"""
    files = []
    for pattern in ['**/*.xlsx', '**/*.xls']:
        files.extend(directory.glob(pattern))

    # Filter out temp files and archive
    files = [f for f in files if not f.name.startswith('~') and 'archive' not in str(f).lower()]

    return sorted(files)


def main():
    parser = argparse.ArgumentParser(description='Excel to Database ETL')
    parser.add_argument('--scan', action='store_true', help='Scan and list Excel files')
    parser.add_argument('--load', action='store_true', help='Load all Excel files')
    parser.add_argument('--file', type=str, help='Load specific file')
    parser.add_argument('--status', action='store_true', help='Show database status')
    parser.add_argument('--init', action='store_true', help='Initialize database')
    parser.add_argument('--analyze', type=str, help='Analyze a specific file')

    args = parser.parse_args()

    # Check dependencies
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        return

    if not PANDAS_AVAILABLE:
        print("WARNING: pandas recommended for full functionality. Install with: pip install pandas")

    db = DatabaseManager()

    if args.init:
        db.init_database()
        print(f"Database initialized: {DB_PATH}")
        return

    if args.status:
        status = db.get_import_status()
        print("\n" + "="*60)
        print("  DATABASE STATUS")
        print("="*60)
        print(f"\n  Path: {status['database_path']}")
        print(f"  Size: {status['database_size_mb']:.2f} MB")
        print("\n  Import Status:")
        for s, info in status['import_status'].items():
            print(f"    {s}: {info['count']} files, {info['rows'] or 0} rows")
        print("\n  Table Counts:")
        for table, count in status['table_counts'].items():
            print(f"    {table}: {count:,} rows")
        print()
        return

    if args.scan:
        files = scan_excel_files()
        print(f"\n{'='*60}")
        print(f"  Found {len(files)} Excel files")
        print("="*60)

        analyzer = ExcelAnalyzer()
        categories = {}

        for f in files:
            info = analyzer.analyze_file(f)
            cat = info.category
            categories.setdefault(cat, []).append(info)

        for cat, files in sorted(categories.items()):
            print(f"\n  {cat.upper()} ({len(files)} files)")
            for info in files[:5]:  # Show first 5
                sheets_summary = ', '.join(f"{s.name}({s.data_type})" for s in info.sheets[:3])
                print(f"    - {info.path.name}")
                print(f"      Sheets: {sheets_summary}")
            if len(files) > 5:
                print(f"    ... and {len(files)-5} more")
        print()
        return

    if args.analyze:
        filepath = Path(args.analyze)
        if not filepath.exists():
            print(f"File not found: {filepath}")
            return

        analyzer = ExcelAnalyzer()
        info = analyzer.analyze_file(filepath)

        print(f"\n{'='*60}")
        print(f"  FILE ANALYSIS: {filepath.name}")
        print("="*60)
        print(f"\n  Category: {info.category}")
        print(f"  Subcategory: {info.subcategory}")
        print(f"  Total Rows: {info.total_rows}")
        print(f"  Sheets: {len(info.sheets)}")

        for sheet in info.sheets:
            print(f"\n  Sheet: {sheet.name}")
            print(f"    Type: {sheet.data_type}")
            print(f"    Size: {sheet.rows} x {sheet.cols}")
            print(f"    Headers: {sheet.headers[:5]}...")
            if sheet.date_column:
                print(f"    Date Column: {sheet.date_column}")
            if sheet.preview:
                print(f"    Preview: {sheet.preview[0]}")
        print()
        return

    if args.file:
        filepath = Path(args.file)
        if not filepath.exists():
            print(f"File not found: {filepath}")
            return

        db.init_database()
        loader = ExcelLoader(db)
        rows, status = loader.load_file(filepath)
        print(f"Loaded {rows} rows from {filepath.name} (status: {status})")
        return

    if args.load:
        db.init_database()
        loader = ExcelLoader(db)
        files = scan_excel_files()

        print(f"\n{'='*60}")
        print(f"  LOADING {len(files)} EXCEL FILES")
        print("="*60)

        total_rows = 0
        success_count = 0

        for i, filepath in enumerate(files, 1):
            print(f"\n  [{i}/{len(files)}] {filepath.name}...")
            rows, status = loader.load_file(filepath)
            print(f"    → {rows} rows ({status})")
            total_rows += rows
            if status == 'success':
                success_count += 1

        print(f"\n{'='*60}")
        print(f"  COMPLETE")
        print("="*60)
        print(f"  Files loaded: {success_count}/{len(files)}")
        print(f"  Total rows: {total_rows:,}")
        print(f"  Database: {DB_PATH}")
        print()
        return

    # Default: show help
    parser.print_help()


if __name__ == '__main__':
    main()
