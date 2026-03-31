"""
Balance Sheet Template Generator
Creates Excel workbook templates for all commodity/country combinations.

Each workbook follows the layout of us_soybean_complex_bal_sheets.xlsx:
- One tab per commodity in the complex
- Standard balance sheet rows (area, yield, production, imports, crush, etc.)
- Marketing year columns
- Headers, formatting, and structure ready for data entry or formula linking

Usage:
    python scripts/generate_balance_sheet_templates.py
    python scripts/generate_balance_sheet_templates.py --group oilseeds
    python scripts/generate_balance_sheet_templates.py --workbook brazil_soybean_complex
"""

import argparse
import logging
import os
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("template_gen")

OUTPUT_DIR = Path("output/balance_sheet_templates")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CONFIG_PATH = Path("config/balance_sheet_workbooks.yaml")

# ── Standard balance sheet row layouts ──────────────────────────────
# These define what rows appear in each tab, varying by commodity type

OILSEED_ROWS = [
    ("", "header"),
    ("Supply", "section"),
    ("Area Planted", "data"),
    ("Area Harvested", "data"),
    ("Yield", "data"),
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Demand", "section"),
    ("Crush", "data"),
    ("Food, Seed & Industrial", "data"),
    ("Feed & Residual", "data"),
    ("Domestic Consumption", "formula"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "data"),
    ("Stocks-to-Use %", "formula"),
    ("", "spacer"),
    ("Memo Items", "section"),
    ("Stocks-to-Use Ratio", "data"),
    ("YoY Production Change %", "data"),
    ("YoY Exports Change %", "data"),
]

GRAIN_ROWS = [
    ("", "header"),
    ("Supply", "section"),
    ("Area Planted", "data"),
    ("Area Harvested", "data"),
    ("Yield", "data"),
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Demand", "section"),
    ("Feed & Residual", "data"),
    ("Food, Seed & Industrial", "data"),
    ("Domestic Consumption", "formula"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "data"),
    ("Stocks-to-Use %", "formula"),
]

OIL_ROWS = [
    ("", "header"),
    ("Supply", "section"),
    ("Beginning Stocks", "data"),
    ("Production — Crude", "data"),
    ("Production — Refined", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Demand", "section"),
    ("Domestic Disappearance", "data"),
    ("   Biodiesel/RD Use", "data"),
    ("   Food Use", "data"),
    ("   Other Industrial", "data"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "data"),
    ("Stocks-to-Use %", "formula"),
]

MEAL_ROWS = [
    ("", "header"),
    ("Supply", "section"),
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Demand", "section"),
    ("Domestic Disappearance", "data"),
    ("   Feed Use", "data"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "data"),
    ("Stocks-to-Use %", "formula"),
]

FAT_ROWS = [
    ("", "header"),
    ("Supply", "section"),
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Demand", "section"),
    ("Domestic Disappearance", "data"),
    ("   Biofuel Use", "data"),
    ("   Food/Industrial Use", "data"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "data"),
]

PALM_ROWS = [
    ("", "header"),
    ("Supply", "section"),
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Demand", "section"),
    ("Domestic Disappearance", "data"),
    ("   Biodiesel Use", "data"),
    ("   Food Use", "data"),
    ("   Oleochemical Use", "data"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "data"),
    ("Stocks-to-Use %", "formula"),
]

ETHANOL_ROWS = [
    ("", "header"),
    ("Supply", "section"),
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Demand", "section"),
    ("Fuel Use", "data"),
    ("Export Use", "data"),
    ("Other Use", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "data"),
]

# Map commodity types to row layouts
ROW_LAYOUTS = {
    'oilseed': OILSEED_ROWS,
    'grain': GRAIN_ROWS,
    'oil': OIL_ROWS,
    'meal': MEAL_ROWS,
    'fat': FAT_ROWS,
    'palm': PALM_ROWS,
    'ethanol': ETHANOL_ROWS,
}

# Map specific commodities to their row layout type
COMMODITY_LAYOUT = {
    'soybeans': 'oilseed', 'rapeseed': 'oilseed', 'sunflowerseed': 'oilseed',
    'cottonseed': 'oilseed', 'peanuts': 'oilseed', 'flaxseed': 'oilseed',
    'safflower': 'oilseed', 'copra': 'oilseed',
    'corn': 'grain', 'wheat': 'grain', 'sorghum': 'grain',
    'barley': 'grain', 'oats': 'grain', 'rice': 'grain', 'cotton': 'grain',
    'sugar': 'grain',
    'soybean_oil': 'oil', 'rapeseed_oil': 'oil', 'sunflowerseed_oil': 'oil',
    'cottonseed_oil': 'oil', 'peanut_oil': 'oil', 'corn_oil': 'oil',
    'linseed_oil': 'oil', 'coconut_oil': 'oil',
    'palm_oil': 'palm', 'palm_kernel_oil': 'palm',
    'soybean_meal': 'meal', 'rapeseed_meal': 'meal', 'sunflowerseed_meal': 'meal',
    'cottonseed_meal': 'meal', 'peanut_meal': 'meal',
    'tallow': 'fat', 'edible_tallow': 'fat', 'inedible_tallow': 'fat',
    'uco': 'fat', 'yellow_grease': 'fat', 'lard': 'fat',
    'cwg': 'fat', 'poultry_fat': 'fat', 'dco': 'fat', 'other_grease': 'fat',
    'ethanol': 'ethanol', 'biodiesel': 'ethanol', 'renewable_diesel': 'ethanol',
}

# ── Styling ─────────────────────────────────────────────────────────
NAVY = '1F4E79'
GOLD = 'C8963E'
LIGHT_BLUE = 'D6E4F0'
LIGHT_GRAY = 'F2F2F2'

TITLE_FONT = Font(name='Calibri', size=14, bold=True, color=NAVY)
SECTION_FONT = Font(name='Calibri', size=11, bold=True, color=NAVY)
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=10)
FORMULA_FONT = Font(name='Calibri', size=10, bold=True)

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
SECTION_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

THIN_BORDER = Border(
    bottom=Side(style='thin', color='D0D0D0'),
)

COUNTRY_NAMES = {
    'US': 'United States', 'BR': 'Brazil', 'AR': 'Argentina',
    'CH': 'China', 'E4': 'EU-27', 'IN': 'India', 'RS': 'Russia',
    'CA': 'Canada', 'UP': 'Ukraine', 'AS': 'Australia', 'WD': 'World',
    'ID': 'Indonesia', 'MY': 'Malaysia', 'PA': 'Paraguay', 'UY': 'Uruguay',
    'JA': 'Japan', 'MX': 'Mexico', 'KS': 'South Korea', 'TW': 'Taiwan',
    'TH': 'Thailand', 'EG': 'Egypt', 'SA': 'Saudi Arabia', 'KZ': 'Kazakhstan',
    'PK': 'Pakistan', 'ZA': 'South Africa', 'TU': 'Turkey', 'SF': 'South Africa',
    'PH': 'Philippines', 'RP': 'Philippines',
}

MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# ── Monthly block definitions per commodity type ────────────────────
# Each block becomes a 14-row section: title, unit, 12 months, MY total

OILSEED_MONTHLY_BLOCKS = [
    ('IMPORTS', 'million bushels'),
    ('EXPORTS', 'million bushels'),
    ('CRUSH', 'million bushels'),
    ('SEED USE', 'million bushels'),
    ('RESIDUAL USE', 'million bushels'),
    ('STOCKS', 'million bushels'),
]

OIL_MONTHLY_BLOCKS = [
    ('PRODUCTION', 'million pounds'),
    ('YIELD', 'pounds per bushel'),
    ('IMPORTS', 'million pounds'),
    ('EXPORTS', 'million pounds'),
    ('DOMESTIC USE', 'million pounds'),
    ('BIOMASS-BASED DIESEL USE', 'million pounds'),
    ('BIODIESEL USE', 'million pounds'),
    ('RENEWABLE DIESEL USE', 'million pounds'),
    ('CO-PROCESSING USE', 'million pounds'),
    ('NON-BIODIESEL USE', 'million pounds'),
    ('STOCKS', 'million pounds'),
]

MEAL_MONTHLY_BLOCKS = [
    ('PRODUCTION', 'thousand short tons'),
    ('YIELD', 'pounds per bushel'),
    ('IMPORTS', 'thousand short tons'),
    ('EXPORTS', 'thousand short tons'),
    ('DOMESTIC USE', 'thousand short tons'),
    ('STOCKS', 'thousand short tons'),
]

FAT_MONTHLY_BLOCKS = OIL_MONTHLY_BLOCKS  # Same structure as vegetable oils

GRAIN_MONTHLY_BLOCKS = [
    ('IMPORTS', '1,000 MT'),
    ('EXPORTS', '1,000 MT'),
    ('FEED & RESIDUAL USE', '1,000 MT'),
    ('FSI USE', '1,000 MT'),
    ('STOCKS', '1,000 MT'),
]

PALM_MONTHLY_BLOCKS = [
    ('PRODUCTION', '1,000 MT'),
    ('IMPORTS', '1,000 MT'),
    ('EXPORTS', '1,000 MT'),
    ('DOMESTIC USE', '1,000 MT'),
    ('BIODIESEL USE', '1,000 MT'),
    ('STOCKS', '1,000 MT'),
]

ETHANOL_MONTHLY_BLOCKS = [
    ('PRODUCTION', 'million gallons'),
    ('IMPORTS', 'million gallons'),
    ('EXPORTS', 'million gallons'),
    ('STOCKS', 'million gallons'),
]

MONTHLY_BLOCKS = {
    'oilseed': OILSEED_MONTHLY_BLOCKS,
    'oil': OIL_MONTHLY_BLOCKS,
    'meal': MEAL_MONTHLY_BLOCKS,
    'fat': FAT_MONTHLY_BLOCKS,
    'grain': GRAIN_MONTHLY_BLOCKS,
    'palm': PALM_MONTHLY_BLOCKS,
    'ethanol': ETHANOL_MONTHLY_BLOCKS,
}

BLOCK_FONT = Font(name='Calibri', size=11, bold=True, color=NAVY)
BLOCK_UNIT_FONT = Font(name='Calibri', size=9, italic=True, color='808080')
MONTH_FONT = Font(name='Calibri', size=10)
TOTAL_FONT = Font(name='Calibri', size=10, bold=True)


def get_my_months(my_start):
    """Return month names in marketing year order."""
    return [MONTH_NAMES[(my_start - 1 + i) % 12] for i in range(12)]


def get_my_label(my_start, year):
    """Generate marketing year column headers."""
    if my_start == 1:
        return str(year)
    end_yr = year + 1
    return f"{year}/{str(end_yr)[-2:]}"


def add_monthly_block(ws, start_row, block_title, unit_label, my_start,
                      country_name, commodity_display, num_my_cols):
    """
    Add one monthly block (14 rows):
    Row 1: blank spacer
    Row 2: "{COUNTRY} {COMMODITY} {BLOCK_TITLE}"
    Row 3: "(unit_label)"
    Rows 4-15: 12 months in MY order
    Row 16: "  Marketing-year Total" or "  Marketing-year Average" for yields
    """
    row = start_row

    # Spacer
    row += 1

    # Block title
    title = f"{country_name.upper()} {commodity_display.upper()} {block_title}"
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = BLOCK_FONT
    row += 1

    # Unit
    ws.cell(row=row, column=1, value=f"({unit_label})").font = BLOCK_UNIT_FONT
    row += 1

    # 12 months in MY order
    months = get_my_months(my_start)
    for month_name in months:
        ws.cell(row=row, column=1, value=month_name).font = MONTH_FONT
        row += 1

    # MY Total or Average
    is_yield = 'YIELD' in block_title.upper()
    total_label = "  Marketing-year Average" if is_yield else "  Marketing-year Total"
    cell = ws.cell(row=row, column=1, value=total_label)
    cell.font = TOTAL_FONT
    row += 1

    return row


def create_balance_sheet_tab(wb, tab_name, commodity, country_code, my_start,
                             start_year=2015, end_year=2026):
    """Create one balance sheet tab with annual section + monthly blocks."""
    ws = wb.create_sheet(title=tab_name)

    # Determine row layout
    layout_type = COMMODITY_LAYOUT.get(commodity, 'grain')
    rows = ROW_LAYOUTS.get(layout_type, GRAIN_ROWS)

    country_name = COUNTRY_NAMES.get(country_code, country_code)
    commodity_display = commodity.replace('_', ' ').title()

    # ── Title row ──────────────────────────────────────────────────
    ws.merge_cells('A1:M1')
    ws['A1'] = f"{country_name} {commodity_display.upper()} SUPPLY AND DEMAND"
    ws['A1'].font = TITLE_FONT

    # ── Column headers (marketing years) ───────────────────────────
    header_row = 3
    ws.cell(row=header_row, column=1, value="Item").font = HEADER_FONT
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal='left')

    col = 2
    for year in range(start_year, end_year + 1):
        label = get_my_label(my_start, year)
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        col += 1

    num_my_cols = col - 2

    # ── Annual balance sheet rows ──────────────────────────────────
    current_row = header_row + 1
    for i, (label, row_type) in enumerate(rows):
        cell = ws.cell(row=current_row, column=1, value=label)

        if row_type == 'section':
            cell.font = SECTION_FONT
            for c in range(1, col):
                ws.cell(row=current_row, column=c).fill = SECTION_FILL
        elif row_type == 'formula':
            cell.font = FORMULA_FONT
            cell.border = THIN_BORDER
        elif row_type == 'data':
            cell.font = DATA_FONT
            if i % 2 == 0:
                for c in range(1, col):
                    ws.cell(row=current_row, column=c).fill = ALT_FILL
        elif row_type == 'spacer':
            current_row += 1
            continue

        current_row += 1

    # ── RLC estimate note ──────────────────────────────────────────
    current_row += 1
    ws.cell(row=current_row, column=1,
            value="Bold, green numbers are RLC estimates and predictions").font = Font(
        name='Calibri', size=9, italic=True, color='548235')
    current_row += 1

    # ── Monthly blocks ─────────────────────────────────────────────
    monthly_blocks = MONTHLY_BLOCKS.get(layout_type, GRAIN_MONTHLY_BLOCKS)

    # Determine the MY start for monthly blocks
    # Oils and meals use Oct-Sep even if the seed uses Sep-Aug
    block_my_start = my_start
    if layout_type in ('oil', 'meal', 'fat'):
        block_my_start = 10  # Oct-Sep for products

    for block_title, unit in monthly_blocks:
        current_row = add_monthly_block(
            ws, current_row, block_title, unit, block_my_start,
            country_name, commodity_display, num_my_cols
        )

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions['A'].width = 38
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Freeze panes
    ws.freeze_panes = 'B4'

    return ws


def create_workbook(config):
    """Create one complete workbook from configuration."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    filename = config['filename']
    country = config.get('country', 'US')
    commodities = config.get('commodities', [])
    my_starts = config.get('my_start', {})

    for commodity in commodities:
        my_start = my_starts.get(commodity, 9)
        commodity_display = commodity.replace('_', ' ').title()
        tab_name = commodity_display[:31]  # Excel max tab name length

        create_balance_sheet_tab(
            wb, tab_name, commodity, country, my_start
        )

    # Save
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    logger.info(f"Created {filepath} ({len(commodities)} tabs)")
    return filepath


def main():
    parser = argparse.ArgumentParser(description="Generate balance sheet templates")
    parser.add_argument("--group", help="Only generate for this group (oilseeds, grains, etc.)")
    parser.add_argument("--workbook", help="Only generate this specific workbook")
    args = parser.parse_args()

    # Load config
    with open(CONFIG_PATH, encoding='utf-8') as f:
        config = yaml.safe_load(f)

    workbooks = config.get('workbooks', {})

    count = 0
    for name, wb_config in workbooks.items():
        if wb_config.get('status') == 'EXISTS':
            logger.info(f"Skipping {name} (already exists)")
            continue

        if args.workbook and args.workbook != name:
            continue

        try:
            create_workbook(wb_config)
            count += 1
        except Exception as e:
            logger.error(f"Failed to create {name}: {e}")

    logger.info(f"\nGenerated {count} workbook templates in {OUTPUT_DIR}/")


if __name__ == '__main__':
    main()
