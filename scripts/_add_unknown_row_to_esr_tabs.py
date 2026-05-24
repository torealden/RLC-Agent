"""Add row 218 = UNKNOWN to each of the 12 ESR tabs in
us_soy_complex_trade.xlsm, and update the row 216 SUM OF REGIONAL TOTALS
formula in every date column (35..1738) to include row 218 so the
sum balances with the new WORLD TOTAL coming from the gold view."""

import openpyxl
import shutil
from datetime import datetime
from openpyxl.utils import get_column_letter
from pathlib import Path

WB_PATH = Path('models/Oilseeds/us_soy_complex_trade.xlsm')
ESR_TABS = [
    'Weekly Soybean Export Sales', 'Weekly Soybean Export Shipments',
    'Weekly Soybean Export Commits', 'Weekly Soybean Export NMY Sales',
    'Weekly Meal Export Sales', 'Weekly Meal Export Shipments',
    'Weekly Meal Export Commits', 'Weekly Meal Export NMY Sales',
    'Weekly SBO Export Sales', 'Weekly SBO Export Shipments',
    'Weekly SBO Export Commits', 'Weekly SBO Export NMY Sales',
]
DATE_COL_START = 35
DATE_COL_END = 1738   # set by _set_esr_date_headers.py — 1,704 weekly columns from col 35

# Regional totals: row 4 (EU), 33 (Other Europe), 47 (FSU), 61 (Asia/Ocean),
# 108 (Africa), 165 (Western Hemisphere), 218 (UNKNOWN).
REGIONAL_ROWS = [4, 33, 47, 61, 108, 165, 218]

# Backup
backup = WB_PATH.with_suffix(f'.xlsm.bak.{datetime.now().strftime("%Y%m%d-%H%M%S")}')
shutil.copy2(WB_PATH, backup)
print(f'Backup -> {backup.name}')

wb = openpyxl.load_workbook(WB_PATH, keep_vba=True)
for tab_name in ESR_TABS:
    if tab_name not in wb.sheetnames:
        print(f'  SKIP {tab_name}')
        continue
    ws = wb[tab_name]

    # 1. Label row 218
    ws.cell(row=218, column=1).value = 'UNKNOWN'

    # 2. Update SUM formula in row 216 across all date columns
    formulas_updated = 0
    for c in range(DATE_COL_START, DATE_COL_END + 1):
        col_letter = get_column_letter(c)
        new_formula = '=' + '+'.join(f'{col_letter}{r}' for r in REGIONAL_ROWS)
        ws.cell(row=216, column=c).value = new_formula
        formulas_updated += 1
    print(f'  {tab_name}: A218=UNKNOWN, {formulas_updated} SUM formulas updated to include row 218')

wb.save(WB_PATH)
print(f'\nSaved {WB_PATH.name}')

# Verify one tab
wb2 = openpyxl.load_workbook(WB_PATH, read_only=True, keep_vba=False)
ws2 = wb2['Weekly Soybean Export Sales']
print('\nVerify Weekly Soybean Export Sales:')
print(f'  A217: {ws2.cell(217,1).value}')
print(f'  A218: {ws2.cell(218,1).value}')
print(f'  Row 216 col 35 formula: {ws2.cell(216,35).value}')
print(f'  Row 216 col 100 formula: {ws2.cell(216,100).value}')
