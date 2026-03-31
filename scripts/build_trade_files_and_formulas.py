"""
Build trade data files and populate balance sheet formulas.

1. Creates trade files (same structure as us_soy_complex_trade.xlsm)
2. Writes external reference formulas into balance sheet monthly blocks

Usage:
    python scripts/build_trade_files_and_formulas.py --trade-file fats_greases
    python scripts/build_trade_files_and_formulas.py --formulas us_soybean
    python scripts/build_trade_files_and_formulas.py --all
"""

import argparse
import logging
import os
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("trade_builder")

DROPBOX_PATH = Path("C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models")
TEMPLATE_PATH = Path("output/balance_sheet_templates")

# ── Trade file constants ────────────────────────────────────────────
HEADER_ROW = 2
DATA_START_ROW = 4
REGIONAL_ROWS = [4, 33, 47, 61, 108, 165, 216]  # Regional subtotals
WORLD_TOTAL_ROW = 217

# Standard country list (matches existing trade files exactly)
COUNTRIES = {
    # EU-27 (rows 4-31, row 4 = regional total)
    4: "EUROPEAN UNION-27",
    5: "AUSTRIA", 6: "BELGIUM", 7: "BULGARIA", 8: "CROATIA", 9: "CYPRUS",
    10: "CZECH REPUBLIC", 11: "DENMARK", 12: "ESTONIA", 13: "FINLAND",
    14: "FRANCE", 15: "GERMANY", 16: "GREECE", 17: "HUNGARY", 18: "IRELAND",
    19: "ITALY", 20: "LATVIA", 21: "LITHUANIA", 22: "LUXEMBOURG", 23: "MALTA",
    24: "NETHERLANDS", 25: "POLAND", 26: "PORTUGAL", 27: "ROMANIA",
    28: "SLOVAKIA", 29: "SLOVENIA", 30: "SPAIN", 31: "SWEDEN",
    # Other Europe (rows 33-46, row 33 = regional total)
    33: "OTHER EUROPE TOTAL",
    34: "ALBANIA", 35: "ANDORRA", 36: "BOSNIA AND HERZEGOVINA", 37: "ICELAND",
    38: "KOSOVO", 39: "MONTENEGRO", 40: "NORTH MACEDONIA", 41: "NORWAY",
    42: "SERBIA", 43: "SWITZERLAND", 44: "TURKEY", 45: "UNITED KINGDOM",
    # FSU (rows 47-60, row 47 = regional total)
    47: "FORMER SOVIET UNION",
    48: "ARMENIA", 49: "AZERBAIJAN", 50: "BELARUS", 51: "GEORGIA",
    52: "KAZAKHSTAN", 53: "KYRGYZSTAN", 54: "MOLDOVA", 55: "RUSSIA",
    56: "TAJIKISTAN", 57: "TURKMENISTAN", 58: "UKRAINE", 59: "UZBEKISTAN",
    # Asia/Oceania (rows 61-107, row 61 = regional total)
    61: "ASIA & OCEANIA",
    62: "AFGHANISTAN", 63: "AUSTRALIA", 64: "BANGLADESH", 65: "BRUNEI",
    66: "BURMA (MYANMAR)", 67: "CAMBODIA", 68: "CHINA", 69: "FIJI",
    70: "FRENCH POLYNESIA", 71: "HONG KONG", 72: "INDIA", 73: "INDONESIA",
    74: "IRAN", 75: "IRAQ", 76: "ISRAEL", 77: "JAPAN",
    78: "JORDAN", 79: "KOREA, NORTH", 80: "KOREA, SOUTH", 81: "KUWAIT",
    82: "LAOS", 83: "LEBANON", 84: "MACAU", 85: "MALAYSIA",
    86: "MALDIVES", 87: "MONGOLIA", 88: "NEPAL", 89: "NEW ZEALAND",
    90: "OMAN", 91: "PAKISTAN", 92: "PAPUA NEW GUINEA", 93: "PHILIPPINES",
    94: "QATAR", 95: "SAMOA", 96: "SAUDI ARABIA", 97: "SINGAPORE",
    98: "SRI LANKA", 99: "SYRIA", 100: "TAIWAN", 101: "THAILAND",
    102: "TIMOR-LESTE", 103: "TONGA", 104: "UNITED ARAB EMIRATES",
    105: "VIETNAM", 106: "YEMEN",
    # Africa (rows 108-164, row 108 = regional total)
    108: "AFRICA",
    109: "ALGERIA", 110: "ANGOLA", 111: "BENIN", 112: "BOTSWANA",
    113: "BURKINA FASO", 114: "BURUNDI", 115: "CAMEROON",
    116: "CAPE VERDE", 117: "CENTRAL AFRICAN REPUBLIC", 118: "CHAD",
    119: "COMOROS", 120: "CONGO (BRAZZAVILLE)", 121: "CONGO (KINSHASA)",
    122: "DJIBOUTI", 123: "EGYPT", 124: "EQUATORIAL GUINEA",
    125: "ERITREA", 126: "ESWATINI", 127: "ETHIOPIA", 128: "GABON",
    129: "GAMBIA", 130: "GHANA", 131: "GUINEA", 132: "GUINEA-BISSAU",
    133: "IVORY COAST", 134: "KENYA", 135: "LESOTHO", 136: "LIBERIA",
    137: "LIBYA", 138: "MADAGASCAR", 139: "MALAWI", 140: "MALI",
    141: "MAURITANIA", 142: "MAURITIUS", 143: "MOROCCO", 144: "MOZAMBIQUE",
    145: "NAMIBIA", 146: "NIGER", 147: "NIGERIA", 148: "REUNION",
    149: "RWANDA", 150: "SAO TOME AND PRINCIPE", 151: "SENEGAL",
    152: "SEYCHELLES", 153: "SIERRA LEONE", 154: "SOMALIA",
    155: "SOUTH AFRICA", 156: "SOUTH SUDAN", 157: "SUDAN", 158: "TANZANIA",
    159: "TOGO", 160: "TUNISIA", 161: "UGANDA", 162: "ZAMBIA",
    163: "ZIMBABWE",
    # Western Hemisphere (rows 165-215, row 165 = regional total)
    165: "WESTERN HEMISPHERE",
    166: "ANTIGUA AND BARBUDA", 167: "ARGENTINA", 168: "ARUBA",
    169: "BAHAMAS", 170: "BARBADOS", 171: "BELIZE", 172: "BERMUDA",
    173: "BOLIVIA", 174: "BRAZIL", 175: "BRITISH VIRGIN ISLANDS",
    176: "CANADA", 177: "CAYMAN ISLANDS", 178: "CHILE", 179: "COLOMBIA",
    180: "COSTA RICA", 181: "CUBA", 182: "CURACAO", 183: "DOMINICA",
    184: "DOMINICAN REPUBLIC", 185: "ECUADOR", 186: "EL SALVADOR",
    187: "FRENCH GUIANA", 188: "GRENADA", 189: "GUADELOUPE",
    190: "GUATEMALA", 191: "GUYANA", 192: "HAITI", 193: "HONDURAS",
    194: "JAMAICA", 195: "MARTINIQUE", 196: "MEXICO", 197: "NICARAGUA",
    198: "PANAMA", 199: "PARAGUAY", 200: "PERU", 201: "ST KITTS AND NEVIS",
    202: "ST LUCIA", 203: "ST VINCENT AND GRENADINES",
    204: "SURINAME", 205: "TRINIDAD AND TOBAGO", 206: "TURKS AND CAICOS",
    207: "UNITED STATES", 208: "URUGUAY", 209: "VENEZUELA",
    210: "VIRGIN ISLANDS US",
    # Totals
    216: "SUM OF REGIONAL TOTALS",
    217: "WORLD TOTAL",
}

# Styling
NAVY = '1F4E79'
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
DATA_FONT = Font(name='Calibri', size=10)
REGIONAL_FONT = Font(name='Calibri', size=10, bold=True)
TITLE_FONT = Font(name='Calibri', size=11, bold=True, color=NAVY)

# ── Trade file definitions ──────────────────────────────────────────
TRADE_FILES = {
    'fats_greases': {
        'filename': 'us_fats_greases_trade.xlsx',
        'output_dir': DROPBOX_PATH / 'Oilseeds' / 'new_models',
        'sheets': [
            ('Edible Tallow Exports', 'exports', 'Oct-Sep'),
            ('Edible Tallow Imports', 'imports', 'Oct-Sep'),
            ('Inedible Tallow Exports', 'exports', 'Oct-Sep'),
            ('Inedible Tallow Imports', 'imports', 'Oct-Sep'),
            ('Yellow Grease Exports', 'exports', 'Oct-Sep'),
            ('Yellow Grease Imports', 'imports', 'Oct-Sep'),
            ('CWG Exports', 'exports', 'Oct-Sep'),
            ('CWG Imports', 'imports', 'Oct-Sep'),
            ('UCO Exports', 'exports', 'Oct-Sep'),
            ('UCO Imports', 'imports', 'Oct-Sep'),
            ('Lard Exports', 'exports', 'Oct-Sep'),
            ('Lard Imports', 'imports', 'Oct-Sep'),
            ('Poultry Fat Exports', 'exports', 'Oct-Sep'),
            ('Poultry Fat Imports', 'imports', 'Oct-Sep'),
            ('DCO Exports', 'exports', 'Oct-Sep'),
            ('DCO Imports', 'imports', 'Oct-Sep'),
        ],
    },
}


def col_to_letter(col_num):
    """Convert 1-based column number to Excel column letter(s)."""
    return get_column_letter(col_num)


def create_trade_sheet(wb, sheet_name, flow, my_label,
                       start_year=2013, end_year=2026):
    """Create one trade data sheet with the standard country layout."""
    ws = wb.create_sheet(title=sheet_name)

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"{my_label} MY {flow.upper()}").font = TITLE_FONT

    # Row 2: Headers — Column 1 = unit, then MY annual labels, then monthly dates
    ws.cell(row=2, column=1, value="1,000 MT").font = Font(
        name='Calibri', size=9, italic=True, color='808080')

    col = 2
    # Annual MY columns first
    for year in range(start_year, end_year + 1):
        end_yr = year + 1
        label = f"{year}/{str(end_yr)[-2:]}"
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        col += 1

    # Monthly columns — from Jan 2013 to current
    current_date = date(start_year, 1, 1)
    end_date = date(end_year, 12, 1)
    monthly_start_col = col

    while current_date <= end_date:
        cell = ws.cell(row=2, column=col, value=current_date)
        cell.number_format = 'MMM-YY'
        cell.font = Font(name='Calibri', size=9)
        cell.alignment = Alignment(horizontal='center')
        col += 1
        # Advance to next month
        if current_date.month == 12:
            current_date = date(current_date.year + 1, 1, 1)
        else:
            current_date = date(current_date.year, current_date.month + 1, 1)

    # Row 3: blank

    # Rows 4+: Countries
    for row_num, country_name in sorted(COUNTRIES.items()):
        cell = ws.cell(row=row_num, column=1, value=country_name)
        if row_num in REGIONAL_ROWS or row_num == WORLD_TOTAL_ROW:
            cell.font = REGIONAL_FONT
        else:
            cell.font = DATA_FONT

    # Regional subtotal formulas (SUM of the countries in each region)
    # Row 4 (EU): =SUM(rows 5:31)
    # etc. — leave as empty for now, user will add SUM formulas

    # Column widths
    ws.column_dimensions['A'].width = 28
    for c in range(2, col):
        ws.column_dimensions[col_to_letter(c)].width = 10

    # Freeze panes
    ws.freeze_panes = 'B3'

    return ws, monthly_start_col


def create_trade_file(config):
    """Create a complete trade file with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    filename = config['filename']
    output_dir = config.get('output_dir', TEMPLATE_PATH)

    monthly_start_cols = {}
    for sheet_name, flow, my_label in config['sheets']:
        ws, start_col = create_trade_sheet(wb, sheet_name, flow, my_label)
        monthly_start_cols[sheet_name] = start_col

    filepath = output_dir / filename
    wb.save(filepath)
    logger.info(f"Created trade file: {filepath} ({len(config['sheets'])} sheets)")
    return filepath, monthly_start_cols


# ══════════════════════════════════════════════════════════════════════
# FORMULA WRITER
# ══════════════════════════════════════════════════════════════════════

def find_trade_col_for_date(target_year, target_month, start_year=2013):
    """
    Find the column number in the trade file for a given year/month.
    Monthly columns start after the annual MY columns.

    Annual MYs: cols 2 through (2 + num_years - 1)
    Monthly: starts at (2 + num_years), one per month from Jan start_year
    """
    num_annual_cols = 2026 - start_year + 1  # 14 annual columns
    monthly_start = 2 + num_annual_cols  # First monthly column

    # Months from Jan start_year to target
    months_offset = (target_year - start_year) * 12 + (target_month - 1)
    return monthly_start + months_offset


def write_formulas_for_block(ws, block_start_row, trade_file_name, trade_sheet_name,
                             my_start_month, bal_sheet_start_col, num_my_cols,
                             trade_row=217, divisor=1000,
                             trade_start_year=2013, bal_start_year=2015):
    """
    Write external reference formulas into a monthly block.

    Args:
        ws: Balance sheet worksheet
        block_start_row: Row of the first month in the block
        trade_file_name: e.g., 'us_soy_complex_trade.xlsm'
        trade_sheet_name: e.g., 'Soybean Imports'
        my_start_month: Month number where MY starts (9=Sep, 10=Oct)
        bal_sheet_start_col: First data column in balance sheet
        num_my_cols: Number of marketing year columns
        trade_row: Row in trade file to reference (217 = WORLD TOTAL)
        divisor: Division factor (1000 for bushels→million bushels)
        trade_start_year: First year in trade file
        bal_start_year: First MY in balance sheet columns
    """
    formulas_written = 0

    for my_offset in range(num_my_cols):
        my_year = bal_start_year + my_offset
        bal_col = bal_sheet_start_col + my_offset

        for month_idx in range(12):
            # Calculate the calendar month/year for this MY month
            cal_month = ((my_start_month - 1 + month_idx) % 12) + 1
            cal_year = my_year + (1 if (my_start_month - 1 + month_idx) >= 12 else 0)

            # Find the column in the trade file
            trade_col = find_trade_col_for_date(cal_year, cal_month, trade_start_year)
            trade_col_letter = col_to_letter(trade_col)

            # Build the formula
            if divisor != 1:
                formula = f"='[{trade_file_name}]{trade_sheet_name}'!{trade_col_letter}${trade_row}/{divisor}"
            else:
                formula = f"='[{trade_file_name}]{trade_sheet_name}'!{trade_col_letter}${trade_row}"

            # Write to balance sheet
            row = block_start_row + month_idx
            ws.cell(row=row, column=bal_col, value=formula)
            formulas_written += 1

    return formulas_written


def main():
    parser = argparse.ArgumentParser(description="Build trade files and write formulas")
    parser.add_argument("--trade-file", help="Create a trade file (e.g., fats_greases)")
    parser.add_argument("--all-trade", action="store_true", help="Create all trade files")
    args = parser.parse_args()

    if args.trade_file:
        config = TRADE_FILES.get(args.trade_file)
        if config:
            create_trade_file(config)
        else:
            logger.error(f"Unknown trade file: {args.trade_file}")
            logger.info(f"Available: {list(TRADE_FILES.keys())}")

    elif args.all_trade:
        for name, config in TRADE_FILES.items():
            create_trade_file(config)

    else:
        parser.print_help()


if __name__ == '__main__':
    main()
