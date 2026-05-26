"""
Link feedstock balance sheet workbooks to eia_data.xlsm
=======================================================

For each feedstock balance sheet file, populate the monthly BD/RD/SAF/
coproc blocks in the feedstock-specific tab with cell-to-cell formulas
pointing at eia_data.xlsm.

Two layouts supported:

OILSEEDS layout (models/Oilseeds/* — soy, canola, corn-oil, cotton,
palm, sunflower):
  - Col 2 = MY 1990/91, col 57 = MY 2045/46
  - Section header rows: BD=108, RD=124, SAF=140, CoProc=156
  - Data rows: BD 110-121, RD 126-137, SAF 142-153, CoProc 158-169
  - First MY column with eia_data: col 10 = MY 1998/99

FATS/GREASES layout (models/Fats and Greases/* — DCO, tallow, CWG, PF,
YG, UCO):
  - Col 2 = MY 1993/94, col 35 = MY 2026/27
  - Section header rows: BD=119, RD=135, CoProc=151, SAF=167
  - Data rows: BD 121-132, RD 137-148, CoProc 153-164, SAF 169-180
  - First MY column with eia_data: col 7 = MY 1998/99
  - **Note: CoProc and SAF are SWAPPED in section order vs Oilseeds.**

eia_data!*_monthly row layout (same for all 4 fuel tabs):
  row 5 = Oct 1998, row 17 = Oct 1999, ...
  Generally:  row = 5 + 12 * (MY_start_year - 1998) + month_offset

Feedstock to eia_data column letter (varies by fuel tab; RD/SAF/CoProc
tabs include a "Vegetable Oils Total" subtotal that BD doesn't):

                   BD    RD    SAF   CoProc
  Soybean Oil      B     B     B     B
  Canola Oil       C     C     C     C
  Corn Oil (DCO)   D     D     D     D
  Cottonseed Oil   E     E     E     E
  Palm Oil         F     F     F     F
  Sunflower/Other  G     G     G     G
  Poultry Fat      H     I     I     I
  Tallow           I     J     J     J
  White Grease     J     K     K     K
  Yellow Grease    L     M     M     M
  Other (UCO)      M     N     N     N

Usage:
  python scripts/link_feedstock_balance_sheets_to_eia.py
  python scripts/link_feedstock_balance_sheets_to_eia.py --dry-run
  python scripts/link_feedstock_balance_sheets_to_eia.py --file tallow
  python scripts/link_feedstock_balance_sheets_to_eia.py --revert PATH TAB
"""

from __future__ import annotations

import argparse
import logging
import os
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger('link_bs')


# ---------------------------------------------------------------------------
# Layout configs
# ---------------------------------------------------------------------------

OILSEEDS_LAYOUT = {
    'col_2_my_start_year': 1990,  # col 2 = MY 1990/91
    'first_bs_col':        10,    # col 10 = MY 1998/99
    'last_bs_col':         57,    # col 57 = MY 2045/46
    'section_oct_row': {
        'bd':  110,
        'rd':  126,
        'saf': 142,
        'cop': 158,
    },
}

FATS_LAYOUT = {
    'col_2_my_start_year': 1993,  # col 2 = MY 1993/94
    'first_bs_col':        7,     # col 7 = MY 1998/99
    'last_bs_col':         35,    # col 35 = MY 2026/27
    'section_oct_row': {
        'bd':  121,
        'rd':  137,
        'cop': 153,   # **swapped vs Oilseeds**
        'saf': 169,   # **swapped vs Oilseeds**
    },
}

EIA_DATA_OCT_1998_ROW = 5
MONTHS_IN_MY = 12

# Feedstock to eia_data column letter per fuel tab
FEEDSTOCK_COLS = {
    'soybean_oil':     {'bd': 'B', 'rd': 'B', 'saf': 'B', 'cop': 'B'},
    'canola_oil':      {'bd': 'C', 'rd': 'C', 'saf': 'C', 'cop': 'C'},
    'corn_oil':        {'bd': 'D', 'rd': 'D', 'saf': 'D', 'cop': 'D'},
    'cottonseed_oil':  {'bd': 'E', 'rd': 'E', 'saf': 'E', 'cop': 'E'},
    'palm_oil':        {'bd': 'F', 'rd': 'F', 'saf': 'F', 'cop': 'F'},
    'sunflower_oil':   {'bd': 'G', 'rd': 'G', 'saf': 'G', 'cop': 'G'},
    'poultry_fat':     {'bd': 'H', 'rd': 'I', 'saf': 'I', 'cop': 'I'},
    'tallow':          {'bd': 'I', 'rd': 'J', 'saf': 'J', 'cop': 'J'},
    'white_grease':    {'bd': 'J', 'rd': 'K', 'saf': 'K', 'cop': 'K'},
    'yellow_grease':   {'bd': 'L', 'rd': 'M', 'saf': 'M', 'cop': 'M'},
    'uco_other_grease':{'bd': 'M', 'rd': 'N', 'saf': 'N', 'cop': 'N'},
}

FUEL_TABS_ORDER = [
    ('bd',  'biodiesel_monthly'),
    ('rd',  'renewable_diesel_monthly'),
    ('saf', 'sustainable_aviation_monthly'),
    ('cop', 'co_processing_monthly'),
]

# (file_path, tab_in_file, feedstock_key, layout_name)
FILES_TO_LINK = [
    # Oilseeds (already linked previously except corn_oil reverted below)
    (r'C:\dev\RLC-Agent\models\Oilseeds\us_canola_balance_sheets.xlsx',
     'canola_oil_balance_sheet', 'canola_oil', 'oilseeds'),
    (r'C:\dev\RLC-Agent\models\Oilseeds\us_cottonseed_balance_sheets.xlsx',
     'cotton_oil_balance_sheet', 'cottonseed_oil', 'oilseeds'),
    (r'C:\dev\RLC-Agent\models\Oilseeds\us_palm_complex_balance_sheets.xlsm',
     'palm_oil_balance_sheet', 'palm_oil', 'oilseeds'),
    (r'C:\dev\RLC-Agent\models\Oilseeds\us_sunflower_balance_sheets.xlsx',
     'sunflower_oil_balance_sheet', 'sunflower_oil', 'oilseeds'),

    # Fats and Greases — DCO goes here (the Fats&Greases version is the
    # active one, not the Oilseeds dco_balance_sheet tab)
    (r'C:\dev\RLC-Agent\models\Fats and Greases\us_distillers_corn_oil_balance.xlsx',
     'Distillers Corn Oil', 'corn_oil', 'fats'),
    (r'C:\dev\RLC-Agent\models\Fats and Greases\us_tallow_complex_balance.xlsx',
     'Inedible Tallow', 'tallow', 'fats'),
    (r'C:\dev\RLC-Agent\models\Fats and Greases\us_choice_white_grease_balance.xlsx',
     'Choice White Grease', 'white_grease', 'fats'),
    (r'C:\dev\RLC-Agent\models\Fats and Greases\us_poultry_fat_balance.xlsx',
     'Poultry Fat', 'poultry_fat', 'fats'),
    (r'C:\dev\RLC-Agent\models\Fats and Greases\us_used_cooking_oil_balance.xlsx',
     'Yellow Grease', 'yellow_grease', 'fats'),
    (r'C:\dev\RLC-Agent\models\Fats and Greases\us_used_cooking_oil_balance.xlsx',
     'Used Cooking Oil', 'uco_other_grease', 'fats'),
]

LAYOUTS = {'oilseeds': OILSEEDS_LAYOUT, 'fats': FATS_LAYOUT}


def find_eia_link_index(wb: openpyxl.Workbook) -> int | None:
    """Find which external link index points at eia_data.xlsm."""
    eia_signature = {'feedstock_monthly', 'capacity_monthly',
                     'biodiesel_monthly', 'renewable_diesel_monthly'}
    for i, link in enumerate(wb._external_links, start=1):
        sheet_names = link.externalBook.sheetNames.sheetName if link.externalBook.sheetNames else []
        if eia_signature.issubset(set(sheet_names)):
            return i
    return None


def eia_data_row(bs_col: int, layout: dict, month_offset: int) -> int:
    """Given a BS column and month offset, return eia_data row number."""
    my_start_year = layout['col_2_my_start_year'] + (bs_col - 2)
    row_offset = (my_start_year - 1998) * 12 + month_offset
    return EIA_DATA_OCT_1998_ROW + row_offset


def revert_section(ws, layout: dict, fuel_key: str) -> int:
    """Clear formulas in one section's data block. Returns cells cleared."""
    oct_row = layout['section_oct_row'][fuel_key]
    first_col = layout['first_bs_col']
    last_col = layout['last_bs_col']
    cleared = 0
    for month_offset in range(MONTHS_IN_MY):
        for c in range(first_col, last_col + 1):
            ws.cell(oct_row + month_offset, c, None)
            cleared += 1
    return cleared


def link_file(file_path: str, tab_name: str, feedstock_key: str,
              layout_name: str, dry_run: bool = False) -> dict:
    name = os.path.basename(file_path)
    if not os.path.exists(file_path):
        logger.error(f"  {name}: file not found")
        return {'cells': 0, 'errors': 1}

    if feedstock_key not in FEEDSTOCK_COLS:
        logger.error(f"  {name}: unknown feedstock {feedstock_key}")
        return {'cells': 0, 'errors': 1}

    layout = LAYOUTS[layout_name]
    logger.info(f"\n=== {name} :: {tab_name} ({feedstock_key}, {layout_name}) ===")
    wb = openpyxl.load_workbook(file_path, data_only=False,
                                keep_vba=file_path.endswith('.xlsm'))

    if tab_name not in wb.sheetnames:
        logger.error(f"  tab {tab_name} not found")
        wb.close()
        return {'cells': 0, 'errors': 1}

    eia_idx = find_eia_link_index(wb)
    if eia_idx is None:
        logger.error(f"  no eia_data external link found")
        wb.close()
        return {'cells': 0, 'errors': 1}
    logger.info(f"  eia_data link index = [{eia_idx}]")

    ws = wb[tab_name]
    cells_written = 0
    cols_used = FEEDSTOCK_COLS[feedstock_key]

    for fuel_key, eia_tab_name in FUEL_TABS_ORDER:
        eia_col_letter = cols_used[fuel_key]
        section_oct_row = layout['section_oct_row'][fuel_key]
        first_col = layout['first_bs_col']
        last_col = layout['last_bs_col']
        section_cells = 0
        for month_offset in range(MONTHS_IN_MY):
            target_row = section_oct_row + month_offset
            for bs_col in range(first_col, last_col + 1):
                eia_row = eia_data_row(bs_col, layout, month_offset)
                if eia_row > 703:
                    continue
                formula = f"=[{eia_idx}]{eia_tab_name}!${eia_col_letter}${eia_row}"
                if not dry_run:
                    ws.cell(target_row, bs_col, formula)
                section_cells += 1
                cells_written += 1
        logger.info(f"  {eia_tab_name:30}  col={eia_col_letter}  cells={section_cells}")

    if not dry_run:
        backup_path = file_path + f'.bak_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        shutil.copy2(file_path, backup_path)
        logger.info(f"  Backup: {os.path.basename(backup_path)}")
        wb.save(file_path)
        logger.info(f"  SAVED: {name}")
    wb.close()
    return {'cells': cells_written, 'errors': 0}


def revert_file_section(file_path: str, tab_name: str, layout_name: str,
                        dry_run: bool = False) -> int:
    """Clear all 4 section blocks (BD/RD/SAF/CoProc) in a tab. For
    rollback of a previously-linked file."""
    name = os.path.basename(file_path)
    if not os.path.exists(file_path):
        logger.error(f"  {name}: not found")
        return 0
    layout = LAYOUTS[layout_name]
    wb = openpyxl.load_workbook(file_path, data_only=False,
                                keep_vba=file_path.endswith('.xlsm'))
    if tab_name not in wb.sheetnames:
        logger.error(f"  {name}: tab {tab_name} not found")
        wb.close()
        return 0
    ws = wb[tab_name]
    total = 0
    for fuel_key in ['bd', 'rd', 'saf', 'cop']:
        n = revert_section(ws, layout, fuel_key)
        total += n
    if not dry_run:
        backup_path = file_path + f'.bak_{datetime.now().strftime("%Y%m%d_%H%M%S")}_revert'
        shutil.copy2(file_path, backup_path)
        wb.save(file_path)
        logger.info(f"  REVERTED {name} :: {tab_name} ({total} cells cleared)")
    wb.close()
    return total


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--file', help='Limit to filename substring')
    parser.add_argument('--revert', nargs=3, metavar=('PATH', 'TAB', 'LAYOUT'),
                        help='Clear all 4 sections in a tab')
    args = parser.parse_args()

    if args.revert:
        revert_file_section(args.revert[0], args.revert[1], args.revert[2],
                            dry_run=args.dry_run)
        return

    total_cells = total_errors = 0
    for fp, tab, fs_key, layout in FILES_TO_LINK:
        if args.file and args.file.lower() not in os.path.basename(fp).lower():
            continue
        result = link_file(fp, tab, fs_key, layout, dry_run=args.dry_run)
        total_cells += result['cells']
        total_errors += result['errors']

    logger.info(f"\nDone: {total_cells} cells written, {total_errors} errors")
    if args.dry_run:
        logger.info("--dry-run: no files modified")


if __name__ == '__main__':
    main()
