"""Build the four production tracking workbooks from silver.crop_production.

Per Tore's spec (2026-05-30):
  - One file per commodity group, matching existing models/ directory layout
  - Wide format: year-by-year columns, one row per state
  - Latest vintage only (max release_date per natural-key minus release_date)

Output files:
  - models/Oilseeds/us_oilseed_production.xlsx          (overwrites existing blank)
  - models/Feed Grains/us_feed_grain_production.xlsx    (new)
  - models/Food Grains/us_food_grain_production.xlsx    (new)
  - models/Cotton/us_cotton_production.xlsx             (new)

Each tab layout:
  Row 1: 'State' | 'AP_YYYY' 'AH_YYYY' 'Y_YYYY' 'P_YYYY' for each year
  Row 2: 'US Total' (NATIONAL)
  Row 3+: states alphabetical, with state codes
  Last col group includes the unit per statistic in a footer note tab _meta.
"""
from __future__ import annotations

import sys, logging
from collections import defaultdict
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / '.env')

from src.services.database.db_config import get_connection

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
log = logging.getLogger('build_production_workbooks')

# --- Files and tabs --------------------------------------------------
# Each tab = (commodity_db, class_db, tab_name).
# Tab name follows Tore's existing oilseed file convention (singular forms).
FILES = {
    PROJECT_ROOT / 'models/Oilseeds/us_oilseed_production.xlsx': [
        ('soybeans',  'all_classes', 'soybeans'),
        ('canola',    'all_classes', 'canola'),
        ('sunflower', 'all_classes', 'sunflower'),
        ('sunflower', 'oil_type',    'sunflower_oil'),
        ('sunflower', 'confection',  'sunflower_confection'),
        # cottonseed is derived from cotton — left blank for now (TODO)
        ('peanuts',   'all_classes', 'peanut'),
        ('peanuts',   'runner',      'peanut_runner'),
        ('peanuts',   'spanish',     'peanut_spanish'),
        ('peanuts',   'virginia_valencia', 'peanut_vv'),
        # flaxseed, safflower not yet in collector — leave blank
    ],
    PROJECT_ROOT / 'models/Feed Grains/us_feed_grain_production.xlsx': [
        ('corn',     'all_classes', 'corn'),
        ('corn',     'grain',       'corn_grain'),
        ('corn',     'silage',      'corn_silage'),
        ('sorghum',  'all_classes', 'sorghum'),
        ('sorghum',  'grain',       'sorghum_grain'),
        ('sorghum',  'silage',      'sorghum_silage'),
    ],
    PROJECT_ROOT / 'models/Food Grains/us_food_grain_production.xlsx': [
        ('wheat',    'all_classes', 'wheat'),
        ('wheat',    'winter',      'wheat_winter'),
        ('wheat',    'spring',      'wheat_spring'),
        ('wheat',    'spring_durum','wheat_durum'),
    ],
    PROJECT_ROOT / 'models/Cotton/us_cotton_production.xlsx': [
        ('cotton',   'all_classes', 'cotton'),
        ('cotton',   'upland',      'cotton_upland'),
        ('cotton',   'pima',        'cotton_pima'),
        # cottonseed: TODO once NASS cottonseed series is wired
    ],
}

# Tabs that have NO collector source yet (will be created empty with a note).
# Tore: flaxseed + safflower are not in the NASS commodities list we configured.
EMPTY_TABS = {
    PROJECT_ROOT / 'models/Oilseeds/us_oilseed_production.xlsx': [
        ('cottonseed', 'TODO: derived from cotton or NASS COTTONSEED series'),
        ('flaxseed',   'TODO: add to nass_crop_production_collector COMMODITIES'),
        ('safflower',  'TODO: add to nass_crop_production_collector COMMODITIES'),
    ],
}

# Statistic short-codes for the column headers.
STAT_CODES = [
    ('area_planted',   'AP'),
    ('area_harvested', 'AH'),
    ('yield',          'Y'),
    ('production',     'P'),
]

# Brand color (Tore's internal-file convention)
HEADER_FILL = PatternFill(start_color='3C7D22', end_color='3C7D22', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SUB_FILL    = PatternFill(start_color='E8F0E2', end_color='E8F0E2', fill_type='solid')
NORMAL_FONT = Font(name='Calibri', size=10)
TOTAL_FONT  = Font(name='Calibri', size=10, bold=True)

THIN = Side(border_style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fetch_latest_rows(commodity: str, class_: str, year_ge: int = 2000):
    """Fetch latest-release-date rows per (state, year, statistic) for this
    (commodity, class). Returns dict[(year, statistic, state_fips, state_alpha)]
    -> (value, unit)."""
    sql = """
        WITH ranked AS (
            SELECT
                crop_year, statistic, state_fips, state_alpha, agg_level,
                value, unit, release_date,
                ROW_NUMBER() OVER (
                    PARTITION BY crop_year, statistic, agg_level, state_fips
                    ORDER BY release_date DESC, load_ts DESC
                ) AS rn
            FROM silver.crop_production
            WHERE commodity = %s AND class = %s AND crop_year >= %s
              AND value IS NOT NULL
              AND agg_level IN ('NATIONAL', 'STATE')
        )
        SELECT crop_year, statistic, state_fips, state_alpha, agg_level, value, unit
        FROM ranked WHERE rn = 1
    """
    out = {}
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (commodity, class_, year_ge))
            for r in cur.fetchall():
                d = dict(r)
                key = (int(d['crop_year']), d['statistic'],
                       d['state_fips'], d['state_alpha'])
                out[key] = (float(d['value']), d['unit'], d['agg_level'])
    return out


def build_tab(ws, commodity: str, class_: str, year_ge: int):
    """Populate one worksheet with wide-format production data."""
    rows = fetch_latest_rows(commodity, class_, year_ge)
    if not rows:
        ws.cell(1, 1, value=f'(no data yet for {commodity}/{class_})')
        return 0, {}

    # Year range that actually has data
    years = sorted({k[0] for k in rows.keys()})
    # Distinct states (collect, then sort with US Total first)
    state_fips_to_alpha = {}
    for k in rows.keys():
        crop_yr, stat, fips, alpha = k
        state_fips_to_alpha[fips] = alpha
    state_fips_sorted = sorted(state_fips_to_alpha.keys(),
                               key=lambda f: (f != '99', state_fips_to_alpha[f]))

    # Track units per statistic (for the _meta footer)
    unit_per_stat = {}
    for v in rows.values():
        unit_per_stat.setdefault(v[2] if False else None, None)  # placeholder
    # Better: derive units per statistic from any row
    unit_per_stat = {}
    for (yr, stat, fips, alpha), (val, unit, agg) in rows.items():
        unit_per_stat.setdefault(stat, unit)

    # --- Header rows ---
    ws.cell(1, 1, value='State').font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    col = 2
    for year in years:
        for stat, code in STAT_CODES:
            ws.cell(1, col, value=f'{code}_{year}')
            ws.cell(1, col).font = HEADER_FONT
            ws.cell(1, col).fill = HEADER_FILL
            ws.cell(1, col).alignment = Alignment(horizontal='center')
            col += 1

    # --- Data rows ---
    row = 2
    for fips in state_fips_sorted:
        alpha = state_fips_to_alpha[fips]
        label = 'US Total' if fips == '99' else alpha
        c = ws.cell(row, 1, value=label)
        c.font = TOTAL_FONT if fips == '99' else NORMAL_FONT
        if fips == '99':
            c.fill = SUB_FILL
        col = 2
        for year in years:
            for stat, code in STAT_CODES:
                v = rows.get((year, stat, fips, alpha))
                if v is not None:
                    cell = ws.cell(row, col, value=v[0])
                    cell.font = TOTAL_FONT if fips == '99' else NORMAL_FONT
                    if fips == '99':
                        cell.fill = SUB_FILL
                    # Format: yield gets decimals, others rounded
                    if stat == 'yield':
                        cell.number_format = '#,##0.0'
                    else:
                        cell.number_format = '#,##0'
                col += 1
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 16
    for c in range(2, 2 + len(years) * len(STAT_CODES)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12

    # Freeze the top row + state column
    ws.freeze_panes = 'B2'

    return len(rows), unit_per_stat


def build_meta(ws, file_meta: dict):
    """Write a _meta tab with units, source, generation timestamp."""
    ws['A1'] = 'Meta'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws['B1'] = 'Value'
    ws['B1'].font = HEADER_FONT
    ws['B1'].fill = HEADER_FILL

    ws['A2'] = 'Generated'
    ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A3'] = 'Source'
    ws['B3'] = 'silver.crop_production (mig 122)'
    ws['A4'] = 'Vintage rule'
    ws['B4'] = 'Latest release_date per (state, year, statistic)'
    ws['A5'] = 'Statistic codes'
    ws['B5'] = 'AP=area_planted  AH=area_harvested  Y=yield  P=production'

    row = 7
    ws[f'A{row}'] = 'Tab'
    ws[f'B{row}'] = 'Rows'
    ws[f'C{row}'] = 'Units (per statistic)'
    for c in (f'A{row}', f'B{row}', f'C{row}'):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
    row += 1
    for tab_name, info in file_meta.items():
        ws.cell(row, 1, value=tab_name)
        ws.cell(row, 2, value=info['rows'])
        units_str = '; '.join(f'{s}={u}' for s, u in info.get('units', {}).items())
        ws.cell(row, 3, value=units_str)
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 90


def build_file(filepath: Path, tab_specs):
    """Build one xlsx file with the given tab specs."""
    filepath.parent.mkdir(parents=True, exist_ok=True)

    # Start with empty workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default Sheet

    file_meta = {}
    total_rows_written = 0

    for commodity, class_, tab_name in tab_specs:
        log.info(f'  {filepath.name} :: {tab_name}  ({commodity}/{class_})')
        ws = wb.create_sheet(title=tab_name[:31])  # Excel tab name limit
        n, units = build_tab(ws, commodity, class_, year_ge=2000)
        file_meta[tab_name] = {'rows': n, 'units': units}
        total_rows_written += n

    # Empty/TODO tabs
    for tab_name, note in EMPTY_TABS.get(filepath, []):
        log.info(f'  {filepath.name} :: {tab_name}  (empty: {note})')
        ws = wb.create_sheet(title=tab_name[:31])
        ws['A1'] = note
        ws['A1'].font = Font(name='Calibri', size=11, italic=True, color='808080')
        ws.column_dimensions['A'].width = 80
        file_meta[tab_name] = {'rows': 0, 'units': {}}

    # _meta tab last
    ws_meta = wb.create_sheet(title='_meta')
    build_meta(ws_meta, file_meta)

    wb.save(filepath)
    log.info(f'  -> {filepath.name}: {total_rows_written} total data rows, {len(wb.sheetnames)} tabs')


def main():
    for filepath, tabs in FILES.items():
        log.info(f'Building {filepath}...')
        build_file(filepath, tabs)
    log.info('Done.')


if __name__ == '__main__':
    main()
