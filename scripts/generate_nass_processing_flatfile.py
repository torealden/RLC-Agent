"""
Generate NASS Processing Flat Files for Fats & Oils and Grain Crushings data.

Creates two flat files:
1. us_fats_oils_nass_data.xlsx — Monthly oilseed crush, oil production, stocks
2. us_grain_crush_nass_data.xlsx — Monthly corn grind, DCO production, co-products

These flat files hold the historical NASS data that balance sheets link to.
"""

import logging
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("nass_flatfile")

OUTPUT_DIR = Path("C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models")

NAVY = '1F4E79'
GOLD = 'C8963E'
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
DATA_FONT = Font(name='Calibri', size=10)
TITLE_FONT = Font(name='Calibri', size=12, bold=True, color=NAVY)
SECTION_FONT = Font(name='Calibri', size=10, bold=True, color=NAVY)


def create_monthly_columns(ws, start_col, start_year=2015, end_year=2026):
    """Create monthly date columns from start_year to end_year."""
    col = start_col
    current = date(start_year, 1, 1)
    end = date(end_year, 12, 1)

    while current <= end:
        cell = ws.cell(row=2, column=col, value=current)
        cell.number_format = 'MMM-YY'
        cell.font = Font(name='Calibri', size=9)
        cell.alignment = Alignment(horizontal='center')
        col += 1
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return col - 1


def build_fats_oils_flatfile():
    """
    Fats & Oils NASS data flat file.

    Sections for each oilseed reported in the NASS Fats & Oils report:
    - Soybeans, Canola, Cottonseed, Sunflower, Corn, Peanut
    Each section has: Crush, Oil Prod Crude, Oil Prod Refined, Oil Stocks,
                      Meal Production, Hull Production
    Plus aggregate sections: Total Veg Oil, Total Protein Meal

    Also includes fats & greases from the same report:
    - Edible Tallow, Inedible Tallow, Lard, Yellow Grease, CWG,
      Poultry Fat, Other Grease
    Each with: Production, Stocks
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Fats and Oils Data"

    ws.cell(row=1, column=1, value="NASS FATS & OILS REPORT DATA").font = TITLE_FONT

    # Column A = item labels, row 2 = date headers
    ws.cell(row=2, column=1, value="Month →").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL

    last_col = create_monthly_columns(ws, 2, 2015, 2026)

    # Build sections
    oilseeds = ['Soybeans', 'Canola', 'Cottonseed', 'Sunflower', 'Corn', 'Peanut']
    oil_items = [
        'Crush',
        'Oil Production — Crude',
        'Oil Production — Refined',
        'Oil Yield (lbs per unit)',
        'Oil Stocks — Crude',
        'Oil Stocks — Refined',
        'Oil Stocks — Total',
        'Meal Production',
        'Hull/Cake Production',
    ]

    fat_commodities = [
        'Edible Tallow', 'Inedible Tallow', 'Lard', 'Yellow Grease',
        'Choice White Grease', 'Poultry Fat', 'Other Grease',
        'Used Cooking Oil', 'Fish Oil',
    ]
    fat_items = ['Production', 'Stocks', 'Domestic Disappearance']

    row = 4

    # Oilseed sections
    for oilseed in oilseeds:
        ws.cell(row=row, column=1, value=f"{oilseed.upper()}").font = SECTION_FONT
        for c in range(1, last_col + 1):
            ws.cell(row=row, column=c).fill = PatternFill(
                start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        row += 1

        for item in oil_items:
            ws.cell(row=row, column=1, value=f"  {item}").font = DATA_FONT
            row += 1

        row += 1  # Spacer

    # Fats & Greases sections
    ws.cell(row=row, column=1, value="FATS & GREASES").font = SECTION_FONT
    for c in range(1, last_col + 1):
        ws.cell(row=row, column=c).fill = PatternFill(
            start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    row += 1

    for fat in fat_commodities:
        ws.cell(row=row, column=1, value=f"  {fat}").font = Font(
            name='Calibri', size=10, bold=True)
        row += 1
        for item in fat_items:
            ws.cell(row=row, column=1, value=f"    {item}").font = DATA_FONT
            row += 1
        row += 1

    # Aggregate sections
    aggregates = [
        'TOTAL VEGETABLE OIL',
        '  Production — Crude',
        '  Production — Refined',
        '  Stocks — Total',
        '  Domestic Disappearance',
        '',
        'TOTAL PROTEIN MEAL',
        '  Production',
        '  Stocks',
        '  Domestic Disappearance',
        '',
        'TOTAL FATS & GREASES',
        '  Production',
        '  Stocks',
        '  Domestic Disappearance',
    ]

    row += 1
    for agg in aggregates:
        if not agg:
            row += 1
            continue
        is_header = not agg.startswith(' ')
        ws.cell(row=row, column=1, value=agg).font = SECTION_FONT if is_header else DATA_FONT
        if is_header:
            for c in range(1, last_col + 1):
                ws.cell(row=row, column=c).fill = PatternFill(
                    start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        row += 1

    # Column width
    ws.column_dimensions['A'].width = 35
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    ws.freeze_panes = 'B3'

    filepath = OUTPUT_DIR / "Oilseeds" / "new_models" / "us_fats_oils_nass_data.xlsx"
    wb.save(filepath)
    logger.info(f"Created {filepath} ({row} rows, {last_col} columns)")


def build_grain_crush_flatfile():
    """
    Grain Crushings NASS data flat file.

    Sections:
    - Corn grind: Total, Beverage Alcohol, Fuel Alcohol, Dry Mill, Wet Mill,
                  Industrial, Wet Mill Other
    - Sorghum grind
    - Co-products: DCO, DDG, DDGS, DWG, CDS, CGM, CGF
    - Derived: DCO production (dry mill + wet mill), DCO yield per bushel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Grain Crush Data"

    ws.cell(row=1, column=1, value="NASS GRAIN CRUSHINGS & CO-PRODUCTS DATA").font = TITLE_FONT

    ws.cell(row=2, column=1, value="Month →").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL

    last_col = create_monthly_columns(ws, 2, 2015, 2026)

    row = 4

    # Corn Grind section
    corn_items = [
        ('CORN GRIND', True),
        ('  Total Corn Grind (mil bu)', False),
        ('  Beverage Alcohol (mil bu)', False),
        ('  Fuel Alcohol (mil bu)', False),
        ('    Dry Mill (mil bu)', False),
        ('    Wet Mill (mil bu)', False),
        ('  Industrial Alcohol (mil bu)', False),
        ('  Wet Mill — Other Products (mil bu)', False),
        ('', False),
        ('SORGHUM', True),
        ('  Sorghum for Fuel Alcohol (1,000 CWT)', False),
        ('', False),
        ('CO-PRODUCTS (1,000 short tons)', True),
        ('  Distillers Corn Oil (DCO)', False),
        ('    DCO from Dry Mill', False),
        ('    DCO from Wet Mill', False),
        ('  DCO Yield (lbs per bushel)', False),
        ('  DDG (Distillers Dried Grains)', False),
        ('  DDGS (Distillers Dried Grains w/ Solubles)', False),
        ('  DWG (Distillers Wet Grains)', False),
        ('  CDS (Condensed Distillers Solubles / Syrup)', False),
        ('  Corn Gluten Meal (CGM)', False),
        ('  Corn Gluten Feed (CGF)', False),
        ('', False),
        ('DERIVED CALCULATIONS', True),
        ('  DCO Production — Total (mil lbs)', False),
        ('  DCO Production = (1000 tons × 2000) / 1e6', False),
        ('  Implied Ethanol Production (mil gal)', False),
        ('  Ethanol Yield (gal per bushel)', False),
        ('  Corn Grind Pace vs USDA Annual (cumulative %)', False),
    ]

    for label, is_section in corn_items:
        if not label:
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).font = SECTION_FONT if is_section else DATA_FONT
        if is_section:
            for c in range(1, last_col + 1):
                ws.cell(row=row, column=c).fill = PatternFill(
                    start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        row += 1

    ws.column_dimensions['A'].width = 45
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    ws.freeze_panes = 'B3'

    filepath = OUTPUT_DIR / "Feed Grains" / "new_models" / "us_grain_crush_nass_data.xlsx"
    wb.save(filepath)
    logger.info(f"Created {filepath} ({row} rows, {last_col} columns)")


if __name__ == '__main__':
    build_fats_oils_flatfile()
    build_grain_crush_flatfile()
    print("Done.")
