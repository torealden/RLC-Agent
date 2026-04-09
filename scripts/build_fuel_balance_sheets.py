"""
Build Fuel Balance Sheet Excel Workbooks
Creates standardized balance sheet templates for biofuels and fossil fuels,
populated with data from the database and legacy Excel files.

Generates 6 workbooks in output/balance_sheet_templates/:
  - us_biodiesel_bal_sheets.xlsx
  - us_renewable_diesel_bal_sheets.xlsx
  - us_saf_bal_sheets.xlsx
  - us_distillate_fuel_oil_bal_sheets.xlsx
  - us_ethanol_bal_sheets.xlsx
  - us_bbd_combined_bal_sheets.xlsx

Usage:
    python scripts/build_fuel_balance_sheets.py
    python scripts/build_fuel_balance_sheets.py --fuel biodiesel
    python scripts/build_fuel_balance_sheets.py --no-populate   # empty templates only
"""

import argparse
import logging
import os
import sys
from datetime import date, datetime
from pathlib import Path

import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv(PROJECT_ROOT / '.env')

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("fuel_bal_builder")

OUTPUT_DIR = PROJECT_ROOT / "output" / "balance_sheet_templates"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

LEGACY_DIR = PROJECT_ROOT / "models" / "Biofuels"

# ── Styling (matches generate_balance_sheet_templates.py) ──────────────
NAVY = '1F4E79'
GOLD = 'C8963E'
LIGHT_BLUE = 'D6E4F0'
LIGHT_GRAY = 'F2F2F2'
FORECAST_GREEN = '548235'

TITLE_FONT = Font(name='Calibri', size=14, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name='Calibri', size=12, bold=True, color=NAVY)
SECTION_FONT = Font(name='Calibri', size=11, bold=True, color=NAVY)
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=10)
FORMULA_FONT = Font(name='Calibri', size=10, bold=True)
BLOCK_FONT = Font(name='Calibri', size=11, bold=True, color=NAVY)
BLOCK_UNIT_FONT = Font(name='Calibri', size=9, italic=True, color='808080')
MONTH_FONT = Font(name='Calibri', size=10)
TOTAL_FONT = Font(name='Calibri', size=10, bold=True)
NOTE_FONT = Font(name='Calibri', size=9, italic=True, color=FORECAST_GREEN)

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
SECTION_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

THIN_BORDER = Border(bottom=Side(style='thin', color='D0D0D0'))

MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# ── Annual S&D row definitions ─────────────────────────────────────────

BIOFUEL_ANNUAL_ROWS = [
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Domestic Consumption", "data"),
    ("Losses and Co-Products", "data"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "formula"),
    ("", "spacer"),
    ("Capacity (MMGY)", "data"),
    ("Operating Rate (%)", "formula"),
    ("Stocks/Use Ratio", "formula"),
    ("Average Price ($/gal)", "data"),
]

FOSSIL_ANNUAL_ROWS = [
    ("Beginning Stocks", "data"),
    ("Production", "data"),
    ("Imports", "data"),
    ("Total Supply", "formula"),
    ("", "spacer"),
    ("Domestic Consumption", "data"),
    ("Exports", "data"),
    ("Total Use", "formula"),
    ("", "spacer"),
    ("Ending Stocks", "formula"),
    ("Stocks/Use Ratio", "formula"),
    ("Average Price ($/gal)", "data"),
]

# ── Monthly block definitions ──────────────────────────────────────────

BIOFUEL_MONTHLY_BLOCKS = [
    ('PRODUCTION', 'million gallons'),
    ('IMPORTS', 'million gallons'),
    ('EXPORTS', 'million gallons'),
    ('CONSUMPTION', 'million gallons'),
    ('STOCKS', 'million gallons'),
    ('CAPACITY', 'MMGY'),
]

FOSSIL_MONTHLY_BLOCKS = [
    ('PRODUCTION', 'million gallons'),
    ('IMPORTS', 'million gallons'),
    ('EXPORTS', 'million gallons'),
    ('CONSUMPTION', 'million gallons'),
    ('STOCKS', 'million gallons'),
]

# ── Fuel type configurations ──────────────────────────────────────────

FUEL_CONFIGS = {
    'biodiesel': {
        'filename': 'us_biodiesel_bal_sheets.xlsx',
        'sheet_name': 'Biodiesel',
        'title': 'US BIODIESEL SUPPLY AND DEMAND',
        'header': 'US BIOFUEL AND FUEL COMPLEX',
        'unit': 'Million Gallons',
        'start_year': 2001,
        'end_year': 2030,
        'annual_rows': BIOFUEL_ANNUAL_ROWS,
        'monthly_blocks': BIOFUEL_MONTHLY_BLOCKS,
        'db_fuel_type': 'biodiesel',
        'legacy_sheet': 'BD',
        'legacy_blocks': {
            'PRODUCTION': {'start_row': 25, 'year_row': 24},
            'IMPORTS': {'start_row': 41, 'year_row': 40},
            'EXPORTS': {'start_row': 57, 'year_row': 56},
            'CONSUMPTION': {'start_row': 73, 'year_row': 72},
            'STOCKS': {'start_row': 89, 'year_row': 88},
        },
        'legacy_annual': {
            'beg_stocks': 4, 'production': 5, 'imports': 6,
            'consumption': 8, 'losses': 9, 'exports': 10, 'end_stocks': 12,
        },
    },
    'renewable_diesel': {
        'filename': 'us_renewable_diesel_bal_sheets.xlsx',
        'sheet_name': 'Renewable Diesel',
        'title': 'US RENEWABLE DIESEL SUPPLY AND DEMAND',
        'header': 'US BIOFUEL AND FUEL COMPLEX',
        'unit': 'Million Gallons',
        'start_year': 2011,
        'end_year': 2030,
        'annual_rows': BIOFUEL_ANNUAL_ROWS,
        'monthly_blocks': BIOFUEL_MONTHLY_BLOCKS,
        'db_fuel_type': 'renewable_diesel',
        'legacy_sheet': 'RD',
        'legacy_blocks': {
            'PRODUCTION': {'start_row': 24, 'year_row': 23},
            'IMPORTS': {'start_row': 40, 'year_row': 39},
            'CONSUMPTION': {'start_row': 56, 'year_row': 55},
            'STOCKS': {'start_row': 73, 'year_row': 72},
        },
        'legacy_annual': {
            'beg_stocks': 4, 'production': 5, 'imports': 6,
            'consumption': 8, 'exports': 9, 'end_stocks': 11,
        },
    },
    'saf': {
        'filename': 'us_saf_bal_sheets.xlsx',
        'sheet_name': 'SAF',
        'title': 'US SUSTAINABLE AVIATION FUEL SUPPLY AND DEMAND',
        'header': 'US BIOFUEL AND FUEL COMPLEX',
        'unit': 'Million Gallons',
        'start_year': 2020,
        'end_year': 2035,
        'annual_rows': BIOFUEL_ANNUAL_ROWS,
        'monthly_blocks': BIOFUEL_MONTHLY_BLOCKS,
        'db_fuel_type': 'saf',
        'legacy_sheet': None,
        'legacy_blocks': {},
        'legacy_annual': {},
    },
    'distillate_fuel_oil': {
        'filename': 'us_distillate_fuel_oil_bal_sheets.xlsx',
        'sheet_name': 'Distillate',
        'title': 'US DISTILLATE FUEL OIL SUPPLY AND DEMAND',
        'header': 'US FOSSIL FUEL COMPLEX',
        'unit': 'Million Gallons',
        'start_year': 2001,
        'end_year': 2030,
        'annual_rows': FOSSIL_ANNUAL_ROWS,
        'monthly_blocks': FOSSIL_MONTHLY_BLOCKS,
        'db_fuel_type': None,
        'legacy_sheet': 'Diesel',
        'legacy_blocks': {
            'PRODUCTION': {'start_row': 21, 'year_row': 20},
            'IMPORTS': {'start_row': 37, 'year_row': 36},
            'EXPORTS': {'start_row': 53, 'year_row': 52},
            'CONSUMPTION': {'start_row': 69, 'year_row': 68},
            'STOCKS': {'start_row': 85, 'year_row': 84},
        },
        'legacy_annual': {
            'beg_stocks': 4, 'production': 5, 'imports': 6,
            'consumption': 10, 'exports': 11, 'end_stocks': 13,
        },
    },
    'ethanol': {
        'filename': 'us_ethanol_bal_sheets.xlsx',
        'sheet_name': 'Ethanol',
        'title': 'US FUEL ETHANOL SUPPLY AND DEMAND',
        'header': 'US BIOFUEL AND FUEL COMPLEX',
        'unit': 'Million Gallons',
        'start_year': 2010,
        'end_year': 2030,
        'annual_rows': BIOFUEL_ANNUAL_ROWS,
        'monthly_blocks': BIOFUEL_MONTHLY_BLOCKS,
        'db_fuel_type': None,
        'legacy_sheet': None,
        'legacy_blocks': {},
        'legacy_annual': {},
    },
}


# ── Database helpers ───────────────────────────────────────────────────

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv('RLC_PG_HOST'),
        dbname='rlc_commodities',
        user=os.getenv('RLC_PG_USER'),
        password=os.getenv('RLC_PG_PASSWORD'),
        sslmode='require',
    )


def load_annual_sd(fuel_type):
    """Load annual S&D from bronze.bbd_balance_sheet."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT year, is_forecast, beginning_stocks, production, imports,
                   total_supply, domestic_consumption, exports, total_use,
                   ending_stocks, capacity_mmgy, operating_rate_pct
            FROM bronze.bbd_balance_sheet
            WHERE fuel_type = %s
            ORDER BY year
        """, (fuel_type,))
        rows = cur.fetchall()
        conn.close()
        result = {}
        for r in rows:
            result[int(r[0])] = {
                'is_forecast': r[1],
                'beg_stocks': float(r[2]) if r[2] else None,
                'production': float(r[3]) if r[3] else None,
                'imports': float(r[4]) if r[4] else None,
                'total_supply': float(r[5]) if r[5] else None,
                'consumption': float(r[6]) if r[6] else None,
                'exports': float(r[7]) if r[7] else None,
                'total_use': float(r[8]) if r[8] else None,
                'end_stocks': float(r[9]) if r[9] else None,
                'capacity': float(r[10]) if r[10] else None,
                'operating_rate': float(r[11]) if r[11] else None,
            }
        logger.info(f"  Loaded {len(result)} annual S&D records for {fuel_type}")
        return result
    except Exception as e:
        logger.warning(f"  Could not load annual S&D for {fuel_type}: {e}")
        return {}


def load_monthly_production(fuel_type):
    """Load monthly production from bronze.bbd_capacity_history."""
    col_map = {
        'biodiesel': 'bd_production_mmgal',
        'renewable_diesel': 'rd_production_mmgal',
    }
    col = col_map.get(fuel_type)
    if not col:
        return {}

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f"""
            SELECT period, {col}, bd_capacity_mmgy, rd_capacity_mmgy
            FROM bronze.bbd_capacity_history
            WHERE {col} IS NOT NULL AND {col} > 0
            ORDER BY period
        """)
        rows = cur.fetchall()
        conn.close()

        cap_col = 'bd_capacity_mmgy' if fuel_type == 'biodiesel' else 'rd_capacity_mmgy'
        result = {}
        for r in rows:
            period = r[0]
            year = period.year
            month = period.month
            if year not in result:
                result[year] = {}
            result[year][month] = {
                'production': float(r[1]) if r[1] else None,
                'capacity': float(r[2] if fuel_type == 'biodiesel' else r[3]) if (r[2] or r[3]) else None,
            }
        logger.info(f"  Loaded monthly production for {fuel_type}: {len(rows)} months")
        return result
    except Exception as e:
        logger.warning(f"  Could not load monthly production for {fuel_type}: {e}")
        return {}


def load_legacy_monthly(config):
    """Load monthly block data from legacy US Fuel Balance Sheets.xlsx."""
    legacy_sheet = config.get('legacy_sheet')
    legacy_blocks = config.get('legacy_blocks', {})
    if not legacy_sheet or not legacy_blocks:
        return {}

    legacy_file = LEGACY_DIR / "US Fuel Balance Sheets.xlsx"
    if not legacy_file.exists():
        logger.warning(f"  Legacy file not found: {legacy_file}")
        return {}

    try:
        wb = load_workbook(legacy_file, data_only=True, read_only=True)
        ws = wb[legacy_sheet]

        # Build year -> column mapping from year header row
        # Use the first block's year_row to find headers
        first_block = next(iter(legacy_blocks.values()))
        year_row = first_block['year_row']

        year_cols = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=year_row, column=col).value
            if val is not None:
                try:
                    year_cols[int(val)] = col
                except (ValueError, TypeError):
                    pass

        result = {}
        for block_name, block_info in legacy_blocks.items():
            start_row = block_info['start_row']
            block_data = {}
            for year, col in year_cols.items():
                monthly = {}
                for m in range(12):
                    val = ws.cell(row=start_row + m, column=col).value
                    if val is not None:
                        try:
                            monthly[m + 1] = float(val)
                        except (ValueError, TypeError):
                            pass
                if monthly:
                    block_data[year] = monthly
            result[block_name] = block_data

        wb.close()
        logger.info(f"  Loaded legacy monthly data from {legacy_sheet}: {list(result.keys())}")
        return result
    except Exception as e:
        logger.warning(f"  Could not load legacy data from {legacy_sheet}: {e}")
        return {}


def load_legacy_annual(config):
    """Load annual S&D from legacy spreadsheet rows."""
    legacy_sheet = config.get('legacy_sheet')
    annual_map = config.get('legacy_annual', {})
    if not legacy_sheet or not annual_map:
        return {}

    legacy_file = LEGACY_DIR / "US Fuel Balance Sheets.xlsx"
    if not legacy_file.exists():
        return {}

    try:
        wb = load_workbook(legacy_file, data_only=True, read_only=True)
        ws = wb[legacy_sheet]

        # Year columns from row 2 (BD) or row 2 (RD/Diesel)
        year_cols = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val is not None:
                try:
                    year_cols[int(val)] = col
                except (ValueError, TypeError):
                    pass

        result = {}
        for year, col in year_cols.items():
            row_data = {}
            for key, row_num in annual_map.items():
                val = ws.cell(row=row_num, column=col).value
                if val is not None:
                    try:
                        row_data[key] = float(val)
                    except (ValueError, TypeError):
                        pass
            if row_data:
                result[year] = row_data

        wb.close()
        logger.info(f"  Loaded legacy annual data from {legacy_sheet}: {len(result)} years")
        return result
    except Exception as e:
        logger.warning(f"  Could not load legacy annual: {e}")
        return {}


# ── Excel builder ──────────────────────────────────────────────────────

def add_monthly_block(ws, start_row, block_title, unit_label, fuel_display,
                      num_year_cols, start_year):
    """
    Add one monthly block (16 rows):
    Row 1: blank spacer
    Row 2: "US {FUEL} {BLOCK_TITLE}"
    Row 3: "(unit)"
    Rows 4-15: Jan-Dec
    Row 16: "  Calendar-year Total"
    Returns (next_row, month_start_row) for data population.
    """
    row = start_row + 1  # spacer

    # Block title
    title = f"US {fuel_display.upper()} {block_title}"
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = BLOCK_FONT
    title_row = row
    row += 1

    # Unit + year headers
    ws.cell(row=row, column=1, value=f"({unit_label})").font = BLOCK_UNIT_FONT
    for i in range(num_year_cols):
        yr = start_year + i
        cell = ws.cell(row=row, column=2 + i, value=yr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # 12 months (Jan-Dec)
    month_start = row
    for m, month_name in enumerate(MONTH_NAMES):
        cell = ws.cell(row=row, column=1, value=month_name)
        cell.font = MONTH_FONT
        if m % 2 == 0:
            cell.fill = ALT_FILL
            for c in range(2, 2 + num_year_cols):
                ws.cell(row=row, column=c).fill = ALT_FILL
        row += 1

    # Calendar-year Total with SUM formulas
    cell = ws.cell(row=row, column=1, value="  Calendar-year Total")
    cell.font = TOTAL_FONT
    cell.border = THIN_BORDER
    for i in range(num_year_cols):
        col = 2 + i
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{month_start}:{col_letter}{month_start + 11})"
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
    row += 1

    return row, month_start


def create_fuel_workbook(fuel_key, config, populate=True):
    """Create one fuel balance sheet workbook."""
    logger.info(f"Building {config['filename']}...")

    wb = Workbook()
    ws = wb.active
    ws.title = config['sheet_name']

    fuel_display = config['sheet_name']
    start_year = config['start_year']
    end_year = config['end_year']
    num_years = end_year - start_year + 1
    annual_rows = config['annual_rows']
    monthly_blocks = config['monthly_blocks']

    # ── Row 1: Header ──────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(12, num_years + 1))
    ws['A1'] = config['header']
    ws['A1'].font = TITLE_FONT

    # ── Row 2: Title ───────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(12, num_years + 1))
    ws['A2'] = f"{config['title']} ({config['unit']})"
    ws['A2'].font = SUBTITLE_FONT

    # ── Row 3: Year headers ────────────────────────────────────────
    header_row = 3
    for i in range(num_years):
        yr = start_year + i
        cell = ws.cell(row=header_row, column=2 + i, value=yr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # ── Annual S&D block ───────────────────────────────────────────
    current_row = 5
    annual_row_map = {}  # label -> row number, for formula references

    for i, (label, row_type) in enumerate(annual_rows):
        if row_type == 'spacer':
            current_row += 1
            continue

        cell = ws.cell(row=current_row, column=1, value=label)
        annual_row_map[label] = current_row

        if row_type == 'formula':
            cell.font = FORMULA_FONT
            cell.border = THIN_BORDER
            # Add formulas
            for ci in range(num_years):
                col = 2 + ci
                col_letter = get_column_letter(col)
                if label == 'Total Supply':
                    # = Beginning Stocks + Production + Imports
                    bs_row = annual_row_map.get('Beginning Stocks', current_row - 3)
                    pr_row = annual_row_map.get('Production', current_row - 2)
                    im_row = annual_row_map.get('Imports', current_row - 1)
                    formula = f"={col_letter}{bs_row}+{col_letter}{pr_row}+{col_letter}{im_row}"
                elif label == 'Total Use':
                    # = Domestic Consumption + [Losses] + Exports
                    dc_row = annual_row_map.get('Domestic Consumption')
                    ex_row = annual_row_map.get('Exports')
                    lo_row = annual_row_map.get('Losses and Co-Products')
                    if dc_row and ex_row:
                        if lo_row:
                            formula = f"={col_letter}{dc_row}+{col_letter}{lo_row}+{col_letter}{ex_row}"
                        else:
                            formula = f"={col_letter}{dc_row}+{col_letter}{ex_row}"
                    else:
                        formula = ""
                elif label == 'Ending Stocks':
                    ts_row = annual_row_map.get('Total Supply')
                    tu_row = annual_row_map.get('Total Use')
                    if ts_row and tu_row:
                        formula = f"={col_letter}{ts_row}-{col_letter}{tu_row}"
                    else:
                        formula = ""
                elif label == 'Operating Rate (%)':
                    pr_row = annual_row_map.get('Production')
                    ca_row = annual_row_map.get('Capacity (MMGY)')
                    if pr_row and ca_row:
                        formula = f"=IF({col_letter}{ca_row}>0,{col_letter}{pr_row}/{col_letter}{ca_row}*100,\"\")"
                    else:
                        formula = ""
                elif label == 'Stocks/Use Ratio':
                    es_row = annual_row_map.get('Ending Stocks')
                    tu_row = annual_row_map.get('Total Use')
                    if es_row and tu_row:
                        formula = f"=IF({col_letter}{tu_row}>0,{col_letter}{es_row}/{col_letter}{tu_row},\"\")"
                    else:
                        formula = ""
                else:
                    formula = ""

                if formula:
                    ws.cell(row=current_row, column=col, value=formula).font = FORMULA_FONT
        else:
            # data row
            cell.font = DATA_FONT
            if i % 2 == 0:
                cell.fill = ALT_FILL
                for c in range(2, 2 + num_years):
                    ws.cell(row=current_row, column=c).fill = ALT_FILL

        current_row += 1

    # ── RLC forecast note ──────────────────────────────────────────
    current_row += 1
    ws.cell(row=current_row, column=1,
            value="Bold, green numbers in shaded cells are RLC forecasts and estimates.").font = NOTE_FONT
    current_row += 1

    # ── Monthly blocks ─────────────────────────────────────────────
    block_positions = {}  # block_name -> month_start_row

    for block_title, unit in monthly_blocks:
        current_row, month_start = add_monthly_block(
            ws, current_row, block_title, unit, fuel_display,
            num_years, start_year
        )
        block_positions[block_title] = month_start

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions['A'].width = 42
    for c in range(2, 2 + num_years):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Freeze panes
    ws.freeze_panes = 'B4'

    # ── Data population ────────────────────────────────────────────
    if populate:
        _populate_workbook(ws, config, annual_row_map, block_positions,
                          start_year, end_year, num_years)

    # Save
    filepath = OUTPUT_DIR / config['filename']
    wb.save(filepath)
    logger.info(f"  Saved: {filepath}")
    return filepath


def _populate_workbook(ws, config, annual_row_map, block_positions,
                       start_year, end_year, num_years):
    """Populate workbook with data from DB and legacy files."""

    db_fuel_type = config.get('db_fuel_type')

    # Load data sources
    db_annual = load_annual_sd(db_fuel_type) if db_fuel_type else {}
    db_monthly_prod = load_monthly_production(db_fuel_type) if db_fuel_type else {}
    legacy_monthly = load_legacy_monthly(config)
    legacy_annual = load_legacy_annual(config)

    # Merge annual data: prefer DB, fall back to legacy
    merged_annual = {}
    for year in range(start_year, end_year + 1):
        if year in db_annual:
            merged_annual[year] = db_annual[year]
        elif year in legacy_annual:
            merged_annual[year] = legacy_annual[year]

    # ── Populate annual S&D ────────────────────────────────────────
    field_to_label = {
        'beg_stocks': 'Beginning Stocks',
        'production': 'Production',
        'imports': 'Imports',
        'consumption': 'Domestic Consumption',
        'losses': 'Losses and Co-Products',
        'exports': 'Exports',
        'capacity': 'Capacity (MMGY)',
    }

    for year, data in merged_annual.items():
        col_idx = year - start_year
        if col_idx < 0 or col_idx >= num_years:
            continue
        col = 2 + col_idx

        for field, label in field_to_label.items():
            if field in data and data[field] is not None and label in annual_row_map:
                row = annual_row_map[label]
                cell = ws.cell(row=row, column=col, value=round(data[field], 2))
                cell.font = DATA_FONT
                cell.number_format = '#,##0.00'

    # ── Populate monthly blocks ────────────────────────────────────
    # Monthly production from DB (bbd_capacity_history)
    if db_monthly_prod and 'PRODUCTION' in block_positions:
        month_start_row = block_positions['PRODUCTION']
        for year, months in db_monthly_prod.items():
            col_idx = year - start_year
            if col_idx < 0 or col_idx >= num_years:
                continue
            col = 2 + col_idx
            for m, vals in months.items():
                if vals['production'] is not None:
                    row = month_start_row + (m - 1)  # Jan=row+0, Feb=row+1, etc.
                    cell = ws.cell(row=row, column=col, value=round(vals['production'], 2))
                    cell.number_format = '#,##0.00'

    # Monthly capacity from DB
    if db_monthly_prod and 'CAPACITY' in block_positions:
        month_start_row = block_positions['CAPACITY']
        for year, months in db_monthly_prod.items():
            col_idx = year - start_year
            if col_idx < 0 or col_idx >= num_years:
                continue
            col = 2 + col_idx
            for m, vals in months.items():
                if vals.get('capacity') is not None:
                    row = month_start_row + (m - 1)
                    cell = ws.cell(row=row, column=col, value=round(vals['capacity'], 0))
                    cell.number_format = '#,##0'

    # Legacy monthly blocks (imports, exports, consumption, stocks)
    for block_name, block_data in legacy_monthly.items():
        if block_name not in block_positions:
            continue
        month_start_row = block_positions[block_name]
        for year, months in block_data.items():
            col_idx = year - start_year
            if col_idx < 0 or col_idx >= num_years:
                continue
            col = 2 + col_idx
            for m, val in months.items():
                row = month_start_row + (m - 1)
                # Don't overwrite DB data (production already populated from DB)
                existing = ws.cell(row=row, column=col).value
                if existing is None:
                    cell = ws.cell(row=row, column=col, value=round(val, 2))
                    cell.number_format = '#,##0.00'

    logger.info(f"  Populated with DB + legacy data")


# ── BBD Combined workbook ──────────────────────────────────────────────

def create_bbd_combined():
    """Create combined BBD workbook with aggregated view + capacity sheet."""
    logger.info("Building us_bbd_combined_bal_sheets.xlsx...")

    wb = Workbook()

    # ── Sheet 1: BBD Combined S&D ──────────────────────────────────
    ws = wb.active
    ws.title = "BBD Combined"

    start_year = 2011
    end_year = 2030
    num_years = end_year - start_year + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws['A1'] = "US BIOFUEL AND FUEL COMPLEX"
    ws['A1'].font = TITLE_FONT

    ws['A2'] = "US BIOMASS-BASED DIESEL + SAF COMBINED (Million Gallons)"
    ws['A2'].font = SUBTITLE_FONT

    # Year headers
    for i in range(num_years):
        cell = ws.cell(row=3, column=2 + i, value=start_year + i)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Combined S&D rows
    combined_rows = [
        "Biodiesel Production",
        "Renewable Diesel Production",
        "SAF Production",
        "Total BBD+SAF Production",
        "",
        "Biodiesel Capacity (MMGY)",
        "Renewable Diesel Capacity (MMGY)",
        "Total Capacity (MMGY)",
        "",
        "Biodiesel Operating Rate (%)",
        "Renewable Diesel Operating Rate (%)",
        "Combined Operating Rate (%)",
    ]

    row = 5
    for label in combined_rows:
        if not label:
            row += 1
            continue
        cell = ws.cell(row=row, column=1, value=label)
        if 'Total' in label:
            cell.font = FORMULA_FONT
            cell.border = THIN_BORDER
        else:
            cell.font = DATA_FONT
        row += 1

    # Populate from DB
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        for fuel_type in ['biodiesel', 'renewable_diesel', 'saf']:
            cur.execute("""
                SELECT year, production, capacity_mmgy, operating_rate_pct
                FROM bronze.bbd_balance_sheet
                WHERE fuel_type = %s ORDER BY year
            """, (fuel_type,))
            for r in cur.fetchall():
                year = int(r[0])
                col_idx = year - start_year
                if col_idx < 0 or col_idx >= num_years:
                    continue
                col = 2 + col_idx

                if fuel_type == 'biodiesel':
                    if r[1]: ws.cell(row=5, column=col, value=float(r[1])).number_format = '#,##0.0'
                    if r[2]: ws.cell(row=10, column=col, value=float(r[2])).number_format = '#,##0.0'
                    if r[3]: ws.cell(row=14, column=col, value=float(r[3])).number_format = '0.0'
                elif fuel_type == 'renewable_diesel':
                    if r[1]: ws.cell(row=6, column=col, value=float(r[1])).number_format = '#,##0.0'
                    if r[2]: ws.cell(row=11, column=col, value=float(r[2])).number_format = '#,##0.0'
                    if r[3]: ws.cell(row=15, column=col, value=float(r[3])).number_format = '0.0'
                elif fuel_type == 'saf':
                    if r[1]: ws.cell(row=7, column=col, value=float(r[1])).number_format = '#,##0.0'

        # Total formulas
        for i in range(num_years):
            col = 2 + i
            cl = get_column_letter(col)
            ws.cell(row=8, column=col, value=f"={cl}5+{cl}6+{cl}7").font = FORMULA_FONT
            ws.cell(row=12, column=col, value=f"={cl}10+{cl}11").font = FORMULA_FONT
            ws.cell(row=16, column=col, value=f"=IF({cl}12>0,{cl}8/{cl}12*100,\"\")").font = FORMULA_FONT

        conn.close()
    except Exception as e:
        logger.warning(f"  Could not populate BBD combined: {e}")

    # ── Sheet 2: Monthly Capacity ──────────────────────────────────
    ws2 = wb.create_sheet("Capacity")
    ws2['A1'] = "US BIOMASS-BASED DIESEL MONTHLY CAPACITY AND PRODUCTION"
    ws2['A1'].font = TITLE_FONT

    headers = ['Period', 'BD Capacity (MMGY)', 'RD Capacity (MMGY)',
               'Combined Capacity', 'BD Production (MMgal)', 'RD Production (MMgal)',
               'BD Utilization %', 'RD Utilization %']
    for i, h in enumerate(headers):
        cell = ws2.cell(row=3, column=1 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT period, bd_capacity_mmgy, rd_capacity_mmgy, combined_capacity_mmgy,
                   bd_production_mmgal, rd_production_mmgal,
                   bd_utilization_pct, rd_utilization_pct
            FROM bronze.bbd_capacity_history
            ORDER BY period
        """)
        row = 4
        for r in cur.fetchall():
            ws2.cell(row=row, column=1, value=r[0]).number_format = 'YYYY-MM'
            for c in range(1, 8):
                if r[c] is not None:
                    ws2.cell(row=row, column=1 + c, value=float(r[c])).number_format = '#,##0.0'
            row += 1
        conn.close()
        logger.info(f"  Capacity sheet: {row - 4} rows")
    except Exception as e:
        logger.warning(f"  Could not populate capacity: {e}")

    # Column widths
    ws.column_dimensions['A'].width = 38
    ws2.column_dimensions['A'].width = 14
    for c in range(2, 2 + num_years):
        ws.column_dimensions[get_column_letter(c)].width = 12
    for c in range(2, 9):
        ws2.column_dimensions[get_column_letter(c)].width = 16

    ws.freeze_panes = 'B4'
    ws2.freeze_panes = 'B4'

    filepath = OUTPUT_DIR / "us_bbd_combined_bal_sheets.xlsx"
    wb.save(filepath)
    logger.info(f"  Saved: {filepath}")
    return filepath


# ── Main ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build fuel balance sheet Excel workbooks")
    parser.add_argument("--fuel", help="Only build this fuel type (biodiesel, renewable_diesel, saf, distillate_fuel_oil, ethanol)")
    parser.add_argument("--no-populate", action="store_true", help="Create empty templates without data")
    args = parser.parse_args()

    populate = not args.no_populate

    built = 0
    for fuel_key, config in FUEL_CONFIGS.items():
        if args.fuel and args.fuel != fuel_key:
            continue
        try:
            create_fuel_workbook(fuel_key, config, populate=populate)
            built += 1
        except Exception as e:
            logger.error(f"Failed to build {fuel_key}: {e}", exc_info=True)

    # Build combined BBD workbook
    if not args.fuel or args.fuel == 'bbd_combined':
        try:
            create_bbd_combined()
            built += 1
        except Exception as e:
            logger.error(f"Failed to build bbd_combined: {e}", exc_info=True)

    logger.info(f"\nBuilt {built} fuel balance sheet workbooks in {OUTPUT_DIR}/")


if __name__ == '__main__':
    main()
