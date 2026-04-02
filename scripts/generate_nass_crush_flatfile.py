"""
Generate comprehensive NASS Crush & Fats/Oils flat file.

Replicates the Census Crush tab structure from us_oilseed_crush.xlsm
but adds fats & greases columns and extends to cover all NASS-reported
commodities.

Data goes down (dates as rows), variables across (columns).
This is the master flat file that balance sheets link to.
"""

import logging
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("crush_flatfile")

OUTPUT_DIR = Path("C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models/Oilseeds/new_models")

NAVY = '1F4E79'
HEADER_FONT = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
SECTION_FONT = Font(name='Calibri', size=9, bold=True, color=NAVY)
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
DATA_FONT = Font(name='Calibri', size=9)
DATE_FONT = Font(name='Calibri', size=9)
TITLE_FONT = Font(name='Calibri', size=12, bold=True, color=NAVY)


def build_master_crush_file():
    """
    Master NASS crush and fats/oils data file.

    Structure: Dates down rows (Sep 1993 → current), variables across columns.
    Grouped by commodity complex with section headers.

    This file holds the ACTUAL reported data from NASS.
    Balance sheets link to specific cells in this file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "NASS Crush and F&O Data"

    # ── Column definitions ─────────────────────────────────────────
    # Each column is one data series. Grouped by commodity.

    columns = []

    # Column A = Date
    columns.append(('Date', '', 'date'))

    # SOYBEANS
    columns.append(('SOYBEANS', '', 'section'))
    columns.append(('Soybean Crush', '000 ST', 'data'))
    columns.append(('Soybean Stocks at Mills', '000 ST', 'data'))
    columns.append(('Crude SBO Production', 'mil lbs', 'data'))
    columns.append(('Refined SBO Production', 'mil lbs', 'data'))
    columns.append(('SBO Stocks — Crude', 'mil lbs', 'data'))
    columns.append(('SBO Stocks — Refined', 'mil lbs', 'data'))
    columns.append(('SBO Stocks — Total', 'mil lbs', 'data'))
    columns.append(('SBO Yield', 'lbs/bu', 'data'))
    columns.append(('Soy Meal Production', '000 ST', 'data'))
    columns.append(('Soy Hull Production', '000 ST', 'data'))
    columns.append(('Soy Meal Stocks', '000 ST', 'data'))

    # COTTONSEED
    columns.append(('COTTONSEED', '', 'section'))
    columns.append(('Cottonseed Crush', '000 ST', 'data'))
    columns.append(('Cottonseed Stocks at Mills', '000 ST', 'data'))
    columns.append(('Crude CSO Production', 'mil lbs', 'data'))
    columns.append(('CSO Stocks — Crude', 'mil lbs', 'data'))
    columns.append(('CSO Stocks — Refined', 'mil lbs', 'data'))
    columns.append(('CSO Stocks — Total', 'mil lbs', 'data'))
    columns.append(('Cottonseed Meal Production', '000 ST', 'data'))

    # CANOLA/RAPESEED
    columns.append(('CANOLA', '', 'section'))
    columns.append(('Canola Crush', '000 ST', 'data'))
    columns.append(('Canola Stocks at Mills', '000 ST', 'data'))
    columns.append(('Crude Canola Oil Production', 'mil lbs', 'data'))
    columns.append(('Canola Oil Stocks — Total', 'mil lbs', 'data'))
    columns.append(('Canola Meal Production', '000 ST', 'data'))

    # SUNFLOWER
    columns.append(('SUNFLOWER', '', 'section'))
    columns.append(('Sunflower Crush', '000 ST', 'data'))
    columns.append(('Sunflower Stocks at Mills', '000 ST', 'data'))
    columns.append(('Crude Sunflower Oil Production', 'mil lbs', 'data'))
    columns.append(('Sunflower Oil Stocks — Total', 'mil lbs', 'data'))
    columns.append(('Sunflower Meal Production', '000 ST', 'data'))

    # PEANUT
    columns.append(('PEANUT', '', 'section'))
    columns.append(('Peanut Crush', '000 ST', 'data'))
    columns.append(('Crude Peanut Oil Production', 'mil lbs', 'data'))
    columns.append(('Peanut Oil Stocks', 'mil lbs', 'data'))
    columns.append(('Peanut Meal Production', '000 ST', 'data'))

    # CORN OIL (from wet milling, reported in Fats & Oils)
    columns.append(('CORN OIL', '', 'section'))
    columns.append(('Corn Oil Production — Crude', 'mil lbs', 'data'))
    columns.append(('Corn Oil Production — Refined', 'mil lbs', 'data'))
    columns.append(('Corn Oil Stocks', 'mil lbs', 'data'))

    # FATS & GREASES (from Fats & Oils report)
    columns.append(('EDIBLE TALLOW', '', 'section'))
    columns.append(('Edible Tallow Production', 'mil lbs', 'data'))
    columns.append(('Edible Tallow Stocks', 'mil lbs', 'data'))

    columns.append(('INEDIBLE TALLOW', '', 'section'))
    columns.append(('Inedible Tallow Production', 'mil lbs', 'data'))
    columns.append(('Inedible Tallow Stocks', 'mil lbs', 'data'))

    columns.append(('LARD', '', 'section'))
    columns.append(('Lard Production', 'mil lbs', 'data'))
    columns.append(('Lard Stocks', 'mil lbs', 'data'))

    columns.append(('YELLOW GREASE', '', 'section'))
    columns.append(('Yellow Grease Production', 'mil lbs', 'data'))
    columns.append(('Yellow Grease Stocks', 'mil lbs', 'data'))

    columns.append(('CHOICE WHITE GREASE', '', 'section'))
    columns.append(('CWG Production', 'mil lbs', 'data'))
    columns.append(('CWG Stocks', 'mil lbs', 'data'))

    columns.append(('POULTRY FAT', '', 'section'))
    columns.append(('Poultry Fat Production', 'mil lbs', 'data'))
    columns.append(('Poultry Fat Stocks', 'mil lbs', 'data'))

    columns.append(('OTHER GREASE', '', 'section'))
    columns.append(('Other Grease Production', 'mil lbs', 'data'))
    columns.append(('Other Grease Stocks', 'mil lbs', 'data'))

    columns.append(('FISH OIL', '', 'section'))
    columns.append(('Fish Oil Production', 'mil lbs', 'data'))
    columns.append(('Fish Oil Stocks', 'mil lbs', 'data'))

    # AGGREGATES
    columns.append(('TOTALS', '', 'section'))
    columns.append(('Total Veg Oil Production', 'mil lbs', 'data'))
    columns.append(('Total Veg Oil Stocks', 'mil lbs', 'data'))
    columns.append(('Total Animal Fat Production', 'mil lbs', 'data'))
    columns.append(('Total Animal Fat Stocks', 'mil lbs', 'data'))
    columns.append(('Total Protein Meal Production', '000 ST', 'data'))

    # ── Write headers ──────────────────────────────────────────────
    # Row 1: Section headers (merged/colored)
    # Row 2: Column names
    # Row 3: Units
    # Row 4+: Data

    for col_idx, (name, unit, col_type) in enumerate(columns):
        col = col_idx + 1

        if col_type == 'section':
            # Section header
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            ws.cell(row=2, column=col).fill = SECTION_FILL
            ws.cell(row=3, column=col).fill = SECTION_FILL
        elif col_type == 'date':
            ws.cell(row=2, column=col, value='Date').font = HEADER_FONT
            ws.cell(row=2, column=col).fill = HEADER_FILL
        else:
            ws.cell(row=2, column=col, value=name).font = HEADER_FONT
            ws.cell(row=2, column=col).fill = HEADER_FILL
            ws.cell(row=3, column=col, value=unit).font = Font(
                name='Calibri', size=8, italic=True, color='808080')

    # ── Write date column (Sep 1979 → Dec 2026) ───────────────────
    row = 4
    current = date(1979, 9, 1)
    end = date(2026, 12, 1)

    while current <= end:
        ws.cell(row=row, column=1, value=current).font = DATE_FONT
        ws.cell(row=row, column=1).number_format = 'MMM-YY'
        row += 1
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)

    total_rows = row - 1
    total_cols = len(columns)

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions['A'].width = 10
    for col_idx, (name, unit, col_type) in enumerate(columns):
        if col_type == 'section':
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 3
        elif col_type == 'data':
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = 12

    ws.freeze_panes = 'B4'

    filepath = OUTPUT_DIR / "us_nass_crush_fats_oils_master.xlsx"
    wb.save(filepath)
    logger.info(f"Created {filepath.name}: {total_cols} columns, {total_rows} rows (Sep 1979 - Dec 2026)")
    return filepath


if __name__ == '__main__':
    build_master_crush_file()
