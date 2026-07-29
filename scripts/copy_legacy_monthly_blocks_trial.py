"""
TRIAL: copy monthly blocks from a legacy wld<crop>bal country tab into the new 3-tab country
workbook, re-bucketing each month into the US marketing year (Tore 2026-07-29).

This is the Australia-rapeseed reference for the eventual comprehensive prompt. It copies MONTHLY
BLOCKS ONLY (production/imports/exports/crush/stocks); Tore builds MY annuals from the monthly sums
and fills implied lines (domestic use, non-biofuel) with formulas himself.

MARKETING-YEAR RE-BUCKETING (the crux):
  Source (Australia rapeseed) monthly blocks run Nov->Oct (local MY). Target uses US MY:
  seed Sep-Aug, meal & oil Oct-Sep. We do NOT copy column-for-column; we map each source cell to its
  real calendar (month, year), then place it in the target column whose US-MY window contains that
  month. So Sep/Oct at the tail of a Nov-Oct year correctly move to the NEXT US year (Tore's "shift
  over a row"), done deterministically.

VALUES ARE COPIED AS-IS (source metric tonnes). Unit conversion, if any, is PENDING Tore's answer
  (the template labels meal 'thousand short tons' but seed/oil 'thousand tonnes' -- flagged).
"""
import os, shutil
import openpyxl

SRC = r"C:/Users/torem/RLC Dropbox/Tore Alden/Soybean Spreadsheets - Copy/wldrapbal.xlsx"
SRC_TAB = "Australia Rapeseed"
TEMPLATE = r"models/Oilseeds/China/china_soybean_complex_bal_sheets.xlsx"
OUT = r"models/Oilseeds/Australia/australia_rapeseed_complex_bal_sheets.xlsx"
COUNTRY, CROP = "Australia", "Rapeseed"   # template is China/Soybean; relabel to these

# Ordered text corrections applied to LABEL cells (never formulas, never tab names -- 816 formulas
# reference the tab name 'soy_balance_sheet'). Artifact fixes MUST precede the CHINA->country swap so
# 'CHINAE'/'SCHINATAINABLE' don't become 'AUSTRALIAE'/'SAUSTRALIA...'. Everything non-US = thousand
# tonnes (no short tons).
def relabel_text(s, country, crop):
    reps = [
        ("SCHINATAINABLE", "SUSTAINABLE"),           # US->CHINA artifact over 'sUStainable'
        ("CHINAE", "USE"),                            # US->CHINA artifact over 'USE'
        ("CHINA", country.upper()), ("China", country), ("Chinese", country + "n"),
        ("SOYBEAN", crop.upper()), ("Soybean", crop), ("soybean", crop.lower()),
        ("thousand short tons", "thousand tonnes"), ("short tons", "tonnes"), ("short ton", "tonne"),
    ]
    for a, b in reps:
        s = s.replace(a, b)
    return s

MONTHNUM = {m: i for i, m in enumerate(
    ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"], 1)}
for full, ab in [("january","jan"),("february","feb"),("march","mar"),("april","apr"),
                 ("june","jun"),("july","jul"),("august","aug"),("september","sep"),
                 ("october","oct"),("november","nov"),("december","dec")]:
    MONTHNUM[full] = MONTHNUM[ab]

# (source header, target tab, target header, product-convention). product: 'seed'=Sep-Aug, else Oct-Sep.
BLOCKS = [
    ("RAPESEED IMPORTS", "soy_balance_sheet", "CHINA SOYBEAN IMPORTS", "seed"),
    ("RAPESEED EXPORTS", "soy_balance_sheet", "CHINA SOYBEAN EXPORTS", "seed"),
    ("RAPESEED CRUSH",   "soy_balance_sheet", "CHINA SOYBEAN CRUSH",   "seed"),
    ("RAPESEED MEAL PRODUCTION", "soymeal_balance_sheet", "CHINA SOYBEAN MEAL PRODUCTION", "meal"),
    ("RAPESEED MEAL IMPORTS",    "soymeal_balance_sheet", "CHINA SOYBEAN MEAL IMPORTS",    "meal"),
    ("RAPESEED MEAL EXPORTS",    "soymeal_balance_sheet", "CHINA SOYBEAN MEAL EXPORTS",    "meal"),
    ("RAPESEED MEAL END-OF-MONTH STOCKS", "soymeal_balance_sheet", "CHINA SOYBEAN MEAL MONTH-ENDING STOCKS", "meal"),
    ("RAPESEED OIL PRODUCTION", "soyoil_balance_sheet", "CHINA SOYBEAN OIL PRODUCTION", "oil"),
    ("RAPESEED OIL IMPORTS",    "soyoil_balance_sheet", "CHINA SOYBEAN OIL IMPORTS",    "oil"),
    ("RAPESEED OIL EXPORTS",    "soyoil_balance_sheet", "CHINA SOYBEAN OIL EXPORTS",    "oil"),
    ("RAPESEED OIL END-OF-MONTH STOCKS", "soyoil_balance_sheet", "CHINA SOYBEAN OIL MONTH-ENDING STOCKS", "oil"),
]


def src_start_year(label):
    yy = int(str(label).strip()[:2])
    return 1900 + yy if yy >= 90 else 2000 + yy


def tgt_start_year(label):
    return int(str(label).strip()[:4])


def cal_year_from_src(month_num, my_start):
    # source MY is Nov-Oct: Nov,Dec belong to my_start; Jan-Oct to my_start+1
    return my_start if month_num in (11, 12) else my_start + 1


def us_my(month_num, cal_year, product):
    if product == "seed":                       # Sep-Aug: Sep..Dec start the year
        return cal_year if month_num >= 9 else cal_year - 1
    else:                                        # meal/oil Oct-Sep: Oct..Dec start the year
        return cal_year if month_num >= 10 else cal_year - 1


def find_header_row(ws, text):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().upper().startswith(text.upper()):
            return r
    raise ValueError(f"header not found: {text}")


def read_block(ws, header_text):
    """Return {(month_num, cal_year): value} for a source monthly block (Nov-Oct)."""
    h = find_header_row(ws, header_text)
    yr_row = h + 1
    # source month rows start h+2, 12 of them
    mstart = h + 2
    # map columns -> src MY start year
    col_my = {}
    for c in range(2, ws.max_column + 1):
        lab = ws.cell(yr_row, c).value
        if lab is None: continue
        try: col_my[c] = src_start_year(lab)
        except Exception: pass
    out = {}
    for rr in range(mstart, mstart + 12):
        mlab = ws.cell(rr, 1).value
        if not isinstance(mlab, str): continue
        mn = MONTHNUM.get(mlab.split()[0].strip(",").lower())
        if not mn: continue
        for c, my in col_my.items():
            val = ws.cell(rr, c).value
            if val is None or val == "": continue
            cy = cal_year_from_src(mn, my)
            out[(mn, cy)] = val
    return out


def tab_year_cols(ws):
    """Column -> US-MY start year, read from the LITERAL year header in row 3 (block year rows are
    formulas =B$3, so read the source-of-truth row 3 once per tab)."""
    out = {}
    for c in range(2, ws.max_column + 1):
        lab = ws.cell(3, c).value
        if isinstance(lab, str) and "/" in lab:
            try: out[tgt_start_year(lab)] = c
            except Exception: pass
    return out


def clear_month_data(ws):
    """Blank hardcoded NUMBERS in month-row data cells (China's raw data) but PRESERVE formulas -- the
    template's derived formulas (yields, cross-tab refs to soy_balance_sheet) are structure that must
    recompute for the new country, not be wiped."""
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.split() and a.split()[0].strip(",").lower() in MONTHNUM:
            for c in range(2, ws.max_column + 1):
                if isinstance(ws.cell(r, c).value, (int, float)):
                    ws.cell(r, c).value = None


def target_month_rows(ws, header_text):
    h = find_header_row(ws, header_text)
    row_by_month = {}
    for rr in range(h + 2, h + 14):
        mlab = ws.cell(rr, 1).value
        if isinstance(mlab, str):
            mn = MONTHNUM.get(mlab.split()[0].strip(",").lower())
            if mn: row_by_month[mn] = rr
    return row_by_month


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    shutil.copyfile(TEMPLATE, OUT)
    src = openpyxl.load_workbook(SRC, data_only=True)[SRC_TAB]
    wb = openpyxl.load_workbook(OUT)

    # clear China data + cache the literal year columns per tab
    year_cols = {}
    for tab in ("soy_balance_sheet", "soymeal_balance_sheet", "soyoil_balance_sheet"):
        clear_month_data(wb[tab])
        year_cols[tab] = tab_year_cols(wb[tab])

    total_written = 0
    traces = []
    for src_hdr, tgt_tab, tgt_hdr, product in BLOCKS:
        block = read_block(src, src_hdr)
        ws = wb[tgt_tab]
        col_by_year = year_cols[tgt_tab]
        row_by_month = target_month_rows(ws, tgt_hdr)
        n = 0
        for (mn, cy), val in block.items():
            my = us_my(mn, cy, product)
            col = col_by_year.get(my)
            row = row_by_month.get(mn)
            if col and row:
                ws.cell(row, col).value = val
                n += 1
        total_written += n
        # trace one Sep + one Nov value for the seed import block to show the boundary shift
        if src_hdr == "RAPESEED IMPORTS":
            for (mn, cy) in sorted(block):
                if mn in (9, 11) and cy in (1994, 1993):
                    traces.append((mn, cy, us_my(mn, cy, "seed"), round(block[(mn, cy)], 4)))
        print(f"  {src_hdr:36} -> {tgt_tab:22} {tgt_hdr[:32]:32}  {n} cells")

    # relabel LABEL cells only (skip formulas; tab names untouched -> 816 formulas depend on them)
    relabeled = 0
    for tab in wb.sheetnames:
        ws = wb[tab]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and not v.startswith("="):
                    nv = relabel_text(v, COUNTRY, CROP)
                    if nv != v:
                        c.value = nv; relabeled += 1
    print(f"relabeled {relabeled} label cells (China->{COUNTRY}, Soybean->{CROP}, artifacts/units fixed)")
    wb.save(OUT)
    print(f"\nTotal monthly cells written: {total_written}")
    print("Trace (seed imports, calendar month/year -> US MY start, value):")
    for mn, cy, my, v in traces:
        print(f"  month {mn:2}/{cy} -> US MY {my}/{my+1}  value {v}")
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
