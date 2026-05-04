"""
Add Oct-Sep marketing year accumulator formulas to DCO Imports / DCO Exports
sheets in models/Fats and Greases/us_fats_greases_trade.xlsm.

Convention (per project_calendar_year_vs_marketing_year.md):
  Fats/oils/meals use Oct-Sep MY. Display as start-year/end-year.
  e.g. "2024/25" = Oct 2024 - Sep 2025

For each non-date header column with format "YYYY/YY", write
  =SUM(<Oct-cell>:<Sep-cell>)
into every data row (4..217). Regional rows and row 217 get the same
formula treatment — their contents are already row-aggregated, so
summing Oct..Sep on the same row gives the right answer.

The macro NEVER writes to MY columns (FindColumnForDate matches on
datetime headers only), so these formulas are safe from clobbering.

Usage:
    python scripts/add_my_accumulators_dco.py
"""
from __future__ import annotations

import datetime
import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


WB_PATH = Path("models/Fats and Greases/us_fats_greases_trade.xlsm")
SHEETS = ["DCO Imports", "DCO Exports"]
HEADER_ROW = 2
DATA_START_ROW = 4
DATA_END_ROW = 217


def parse_my_header(h) -> int | None:
    """Parse 'YYYY/YY' marketing year header → start year. Returns None if
    not a MY header."""
    if not isinstance(h, str):
        return None
    if "/" not in h:
        return None
    try:
        parts = h.split("/")
        if len(parts) != 2:
            return None
        start_year = int(parts[0])
        if start_year < 1990 or start_year > 2050:
            return None
        return start_year
    except (ValueError, IndexError):
        return None


def add_my_accumulators(ws) -> tuple[int, int]:
    """For each MY column, write =SUM(Oct..Sep) formula in all data rows.
    Returns (n_my_cols_processed, n_formulas_written)."""
    # Build map: (year, month) → column number
    date_to_col = {}
    my_cols = []
    for col in range(2, ws.max_column + 1):
        h = ws.cell(row=HEADER_ROW, column=col).value
        if isinstance(h, datetime.datetime):
            date_to_col[(h.year, h.month)] = col
        else:
            start = parse_my_header(h)
            if start is not None:
                my_cols.append((col, start, h))

    n_formulas = 0
    n_processed = 0
    for col, start_year, label in my_cols:
        oct_col = date_to_col.get((start_year, 10))
        sep_col = date_to_col.get((start_year + 1, 9))
        if oct_col is None or sep_col is None:
            print(f"    SKIP col {col} ({label!r}): missing date col "
                  f"oct={oct_col} sep={sep_col}")
            continue
        oct_letter = get_column_letter(oct_col)
        sep_letter = get_column_letter(sep_col)
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            formula = f"=SUM({oct_letter}{row}:{sep_letter}{row})"
            ws.cell(row=row, column=col).value = formula
            n_formulas += 1
        print(f"    col {col} ({label!r}): "
              f"=SUM({oct_letter}<row>:{sep_letter}<row>) "
              f"covering Oct-{start_year} through Sep-{start_year+1}")
        n_processed += 1
    return n_processed, n_formulas


def main():
    if not WB_PATH.exists():
        raise SystemExit(f"Workbook not found: {WB_PATH}")

    backup = WB_PATH.with_name(
        WB_PATH.stem + "_backup_my_" +
        datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + WB_PATH.suffix
    )
    shutil.copy2(WB_PATH, backup)
    print(f"Backed up to: {backup.name}")

    wb = openpyxl.load_workbook(WB_PATH, keep_vba=True, data_only=False)

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet '{sheet_name}' not found, skipping")
            continue
        print(f"\n  {sheet_name}:")
        ws = wb[sheet_name]
        n_cols, n_formulas = add_my_accumulators(ws)
        print(f"  -> {n_formulas} formulas written across {n_cols} MY columns")

    wb.save(WB_PATH)
    print(f"\nSaved: {WB_PATH}")
    print("Open in Excel and run the DCO macro to populate the date columns; "
          "MY accumulators will compute automatically on Excel's next recalc.")


if __name__ == "__main__":
    main()
