"""
Build peanut_food_use_subbalance tab in us_peanut_bal_sheets.xlsm.

Tab layout matches the soybean-complex balance sheet convention:
  Row 1: title
  Row 2: subtitle
  Row 3: MY column headers (MY 1990/91 -> MY 2025/26)
  Row 4: units note
  Rows 6-12: annual rollup by sub-flow (top section)
  Rows 14-110: monthly sections, one per sub-flow:
    Section header row, units row, then 12 month rows (Sep-Aug
    for peanut MY convention per existing tab), then MY total row.

Sub-flows (per Tier 2B):
  Peanut Butter, Peanut Candy, Snack Peanuts, Other Edible,
  Clean In-shell

Data sources:
  - Annual canon: silver.peanut_food_use_annual (ERS Yearbook T12,
    1980/81-2023/24)
  - Monthly:      silver.peanut_food_use_monthly (NASS Peanut Stocks
    & Processing). Currently covers Jan 2025-Mar 2026 only — the
    NASS Peanut historical backfill is a separate task. Missing
    cells stay blank.

NOT-formulas (static values). Re-run after each NASS / ERS update.
Idempotent — removes the tab before re-adding.
"""

from __future__ import annotations

import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / '.env')
from src.services.database.db_config import get_connection

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger('peanut_food_tab')

XLSM = PROJECT_ROOT / 'models' / 'Oilseeds' / 'us_peanut_bal_sheets.xlsm'
TAB = 'peanut_food_use_subbalance'

HEADER_FILL = PatternFill('solid', fgColor='3C7D22')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri')
SECTION_FONT = Font(bold=True, name='Calibri')
ITALIC_FONT = Font(italic=True, color='666666', name='Calibri')

# Peanut MY = Sep-Aug per the existing peanut_balance_sheet tab convention.
# (ERS T11 metadata says Aug-Jul, but the spreadsheet template uses Sep-Aug.
# Matching the tab convention so consistency with existing balance sheet
# tabs is preserved.)
MY_MONTHS = [('Sep', 9), ('Oct', 10), ('Nov', 11), ('Dec', 12),
             ('Jan', 1), ('Feb', 2), ('Mar', 3), ('Apr', 4),
             ('May', 5), ('Jun', 6), ('Jul', 7), ('Aug', 8)]

# MY column layout: col 2 = MY 1990/91, col 57 = MY 2045/46
FIRST_MY_START = 1990
FIRST_BS_COL = 2
LAST_MY_START = 2045

# Sub-flow rows in monthly sections + DB column name
SUB_FLOWS = [
    ('Peanut Butter Use',     'peanut_butter_mil_lbs'),
    ('Peanut Candy Use',      'peanut_candy_mil_lbs'),
    ('Snack Peanut Use',      'snack_peanuts_mil_lbs'),
    ('Other Edible Use',      'other_food_mil_lbs'),
    ('Clean In-shell Use',    'clean_in_shell_mil_lbs'),
]

# Section spacing: each section is 16 rows (1 header + 1 unit + 12 months + total + 1 blank)
SECTION_SPACING = 16
FIRST_MONTHLY_SECTION_ROW = 30  # leaves room for annual block above


def main():
    if not XLSM.exists():
        logger.error(f"{XLSM} not found"); sys.exit(1)

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT marketing_year, peanut_butter_mil_lbs, peanut_candy_mil_lbs,
                       snack_peanuts_mil_lbs, other_food_mil_lbs,
                       clean_in_shell_mil_lbs, total_food_use_mil_lbs
                FROM silver.peanut_food_use_annual
                ORDER BY marketing_year
            """)
            annual_by_my = {r['marketing_year']: dict(r) for r in cur.fetchall()}
            cur.execute("""
                SELECT period, year, month, marketing_year_start,
                       peanut_butter_mil_lbs, peanut_candy_mil_lbs,
                       snack_peanuts_mil_lbs, other_food_mil_lbs,
                       clean_in_shell_mil_lbs
                FROM silver.peanut_food_use_monthly
                ORDER BY period
            """)
            monthly_by_year_mo = {(r['year'], r['month']): dict(r) for r in cur.fetchall()}

    logger.info(f"Annual: {len(annual_by_my)} MYs, Monthly: {len(monthly_by_year_mo)} months")

    backup = XLSM.parent / (XLSM.name + f'.bak_{datetime.now():%Y%m%d_%H%M%S}')
    shutil.copy2(XLSM, backup)
    logger.info(f"Backup: {backup.name}")

    wb = openpyxl.load_workbook(XLSM, keep_vba=True)
    if TAB in wb.sheetnames:
        del wb[TAB]
    ws = wb.create_sheet(TAB)

    # ── Header rows ──
    ws.cell(1, 1, 'US PEANUT FOOD USE SUB-BALANCE').font = Font(bold=True, size=14, name='Calibri')
    ws.cell(2, 1, 'Tier 2B — shelled basis, million pounds. Annual = ERS Oil Crops Yearbook Table 12. Monthly = NASS Peanut Stocks & Processing.').font = Font(italic=True, name='Calibri')

    # MY year labels in row 3, units in row 4
    my_cols = {}  # 'YYYY/YY' -> col_idx
    for col_idx in range(FIRST_BS_COL, FIRST_BS_COL + (LAST_MY_START - FIRST_MY_START + 1)):
        my_start = FIRST_MY_START + (col_idx - FIRST_BS_COL)
        my_label = f"{my_start}/{(my_start + 1) % 100:02d}"
        c = ws.cell(3, col_idx, my_label)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        my_cols[my_label] = col_idx
    ws.cell(4, 1, '(million pounds, shelled basis)').font = ITALIC_FONT

    # ── Annual rollup block (rows 6-12) ──
    ws.cell(6, 1, 'ANNUAL ROLLUP — ERS Yearbook Table 12 canon').font = SECTION_FONT
    annual_rows = [
        ('Peanut Butter',     'peanut_butter_mil_lbs',  7),
        ('Peanut Candy',      'peanut_candy_mil_lbs',   8),
        ('Snack Peanuts',     'snack_peanuts_mil_lbs',  9),
        ('Other Edible',      'other_food_mil_lbs',     10),
        ('Clean In-shell',    'clean_in_shell_mil_lbs', 11),
        ('Total Food Use',    'total_food_use_mil_lbs', 12),
    ]
    for label, field, row in annual_rows:
        c = ws.cell(row, 1, label)
        if label == 'Total Food Use':
            c.font = SECTION_FONT
        for my_label, col_idx in my_cols.items():
            v = annual_by_my.get(my_label, {}).get(field)
            if v is not None:
                cell = ws.cell(row, col_idx, float(v))
                if label == 'Total Food Use':
                    cell.font = SECTION_FONT

    # ── Monthly sections (one per sub-flow) ──
    for s_idx, (label, field) in enumerate(SUB_FLOWS):
        section_row = FIRST_MONTHLY_SECTION_ROW + s_idx * SECTION_SPACING
        ws.cell(section_row, 1, f'US PEANUT {label.upper()}').font = SECTION_FONT
        ws.cell(section_row + 1, 1, '(million pounds, shelled basis)').font = ITALIC_FONT

        # 12 month rows
        for m_idx, (m_label, m_num) in enumerate(MY_MONTHS):
            r = section_row + 2 + m_idx
            ws.cell(r, 1, m_label)
            for my_label, col_idx in my_cols.items():
                my_start = int(my_label.split('/')[0])
                # Peanut MY Sep-Aug: month >= 9 belongs to start year, < 9 belongs to start year + 1
                cal_year = my_start if m_num >= 9 else my_start + 1
                mo_data = monthly_by_year_mo.get((cal_year, m_num))
                if mo_data:
                    v = mo_data.get(field)
                    if v is not None:
                        ws.cell(r, col_idx, float(v))

        # MY total row
        total_row = section_row + 2 + 12
        c = ws.cell(total_row, 1, 'Marketing-year Total')
        c.font = SECTION_FONT
        # Use Excel formula to sum 12 month rows
        for col_idx in my_cols.values():
            col_letter = get_column_letter(col_idx)
            ws.cell(total_row, col_idx,
                    f'=SUM({col_letter}{section_row+2}:{col_letter}{section_row+13})').font = SECTION_FONT

    # Column widths
    ws.column_dimensions['A'].width = 32
    for col_idx in my_cols.values():
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    wb.save(XLSM)
    logger.info(f"Saved: {XLSM.name}")


if __name__ == '__main__':
    main()
