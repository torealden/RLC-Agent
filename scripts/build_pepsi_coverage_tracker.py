"""Build the Pepsi (Helios pilot) coverage tracker + country folders.

Tore's model is a TRADE-BALANCED CLOSED LOOP: build the bulk of world producers, crushers, AND
importers for each complex, wire them through a trade matrix (imports<->exports), and the loop
answers the first-order price question — is there enough to meet world demand?

Scope is SOW No. 1: FIVE complexes (palm, rapeseed/canola, sunflower, soybean oil, corn oil), not
three. Each complex carries its OWN sheet set; forcing the oilseed 5-sheet grid on everything
produces columns nobody can fill.

PALM IS A FULL CRUSH COMPLEX WITH TWO OILS — corrected 2026-07-21 against the World Lauric Oils
template (docs/specs/palm_lauric_balance_sheet_template.md). FFB -> mesocarp gives PALM OIL; the
KERNEL is the "seed", crushed into PALM KERNEL OIL + PALM KERNEL CAKE (the meal). So palm carries
FOUR balance sheets plus an area/yield block and is the LARGEST per-country build of the five, not
the smallest. Corn oil is the genuinely small one — a derived co-product, no crush complex.

Tiering (see docs/specs/helios_pepsi_spreadsheet_gameplan_internal.md):
  A  price-setting exporter — sets the reference series we quote. Full sheet set, no shortcuts.
  B  swing importer        — ONE workbook per country, tab per oil + a shared allocation tab that
                             splits total veg-oil import demand across palm/sun/rape/soy on relative
                             price. Build once, serves all five complexes AND the substitution
                             scenario promised to Helios 2026-07-13.
  C  World rollup          — straight from bronze.fas_psd, automated, no manual build.
  D  scenario stub         — production + trade + a shock coefficient. Enough for a what-if; no
                             balance sheet. Covers SOW s2's scenario-only origins.
  E  loop fill             — deferred. Needed to fully close the trade matrix, NOT for bare bones.

Bare bones = A + B + C + D. Tier E is explicitly out of the six-week build.

Run:  python scripts/build_pepsi_coverage_tracker.py
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(r"C:/dev/RLC-Agent")
OILSEEDS = ROOT / "models" / "Oilseeds"
TRACKER = OILSEEDS / "_Pepsi_Coverage_Tracker.xlsx"

# Union of every sheet type across complexes; each complex declares the subset it actually uses and
# the rest are pre-marked N/A so the unfillable cells are visible rather than silently blank.
SHEET_TYPES = ["Plantation", "Seed S&D", "Crush", "Oil S&D", "Kernel Oil S&D",
               "Meal S&D", "Trade", "Stocks"]

OILSEED_SET = {"Seed S&D", "Crush", "Oil S&D", "Meal S&D", "Trade"}
# Palm: Plantation (immature/mature area, oil yield) + CPO + palm kernel (seed) + PKO (second oil)
# + PKC (meal) + trade + monthly stocks. "Kernel Oil S&D" exists ONLY for palm — no other complex
# has a second oil, so PKO cannot be folded into Oil S&D without losing it.
PALM_SET = {"Plantation", "Seed S&D", "Crush", "Oil S&D", "Kernel Oil S&D",
            "Meal S&D", "Trade", "Stocks"}
CORNOIL_SET = {"Oil S&D", "Trade"}          # derived co-product: DCO / wet-mill oil, no crush complex

# What a country needs when it does NOT grow the crop. For the oilseed complexes this is still the
# full set — China, the EU, and Turkey genuinely crush imported seed, so seed/crush/meal all apply.
# For PALM it is not: nobody outside the tropics grows oil palm or crushes kernels, so importers
# need the two oil balance sheets and trade, and nothing else.
IMPORTER_SHEETS = {
    "Palm": {"Oil S&D", "Kernel Oil S&D", "Trade"},
    "Corn Oil": {"Oil S&D", "Trade"},
}
# Tier-D scenario origins that actually GROW the crop (so they keep the production-side sheets);
# everything else in tier D is a demand-side stub.
TIER_D_PRODUCERS = {"Palm": {"Colombia", "Guatemala"}}

# Per complex: the sheet set that applies, then countries by tier.
# Origins named in SOW s2 are marked (SOW) in the notes column.
COMPLEXES = {
    "Palm": {
        "sheets": PALM_SET,
        "A": ["Malaysia", "Indonesia"],
        "B": ["India", "China", "Europe"],
        "D": ["Colombia", "Guatemala", "Mexico"],
        "E": ["Thailand", "Nigeria", "Pakistan", "Bangladesh"],
    },
    "Rapeseed / Canola": {
        "sheets": OILSEED_SET,
        "A": ["Europe", "Canada", "Australia", "Russia"],
        "B": ["China", "India", "Turkey"],
        "D": ["Brazil", "Mexico"],
        "E": ["Ukraine", "Japan", "UAE", "United States"],
    },
    "Sunflower": {
        "sheets": OILSEED_SET,
        "A": ["Ukraine", "Russia", "Argentina"],
        "B": ["Europe", "Turkey", "India", "China"],
        "D": ["Colombia", "Mexico"],
        "E": ["Kazakhstan", "Egypt", "Moldova", "United States"],
    },
    "Soybean": {
        "sheets": OILSEED_SET,
        "A": ["United States", "Brazil", "Argentina"],
        "B": ["China", "India", "Europe"],
        "D": ["Mexico"],
        "E": ["Paraguay", "Canada", "Egypt", "Bangladesh", "Japan", "Indonesia", "Ukraine", "Russia"],
    },
    "Corn Oil": {
        "sheets": CORNOIL_SET,
        "A": ["United States", "Brazil"],
        "B": [],
        "D": ["Mexico"],
        "E": [],
    },
}

# SOW s2 origins in scope, for the notes column.
SOW_ORIGINS = {
    "Palm": {"Malaysia", "Indonesia", "Colombia", "Guatemala", "Mexico"},
    "Rapeseed / Canola": {"Europe", "Russia", "Turkey", "Canada", "Brazil"},
    "Sunflower": {"Europe", "Russia", "Ukraine", "Turkey", "Argentina"},
    "Soybean": {"United States", "Brazil", "Argentina"},
    "Corn Oil": {"Mexico", "Brazil"},
}

TIER_LABEL = {
    "A": "A — exporter",
    "B": "B — importer",
    "C": "C — world",
    "D": "D — scenario",
    "E": "E — loop fill",
}
TIER_COLOR = {"A": "3C7D22", "B": "1F6F8B", "C": "6E7178", "D": "8A6D1F", "E": "A8ADB5"}

# Already built. US soy complex is the template; US corn oil exists via the feedstock/DCO layer.
DONE = {
    ("Soybean", "United States"): {"Seed S&D", "Crush", "Oil S&D", "Meal S&D", "Trade"},
    ("Rapeseed / Canola", "United States"): {"Seed S&D", "Crush", "Oil S&D", "Trade"},
    ("Sunflower", "United States"): {"Seed S&D"},
    ("Corn Oil", "United States"): {"Oil S&D", "Trade"},
}

GREEN = "3C7D22"
THIN = Side(style="thin", color="D4D4CE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STATUSES = ["", "Planned", "In Progress", "Done", "N/A"]
NA_FILL = PatternFill("solid", fgColor="F2F2F0")
DONE_FILL = PatternFill("solid", fgColor="E2EFD9")


def sheets_for(cx, country, tier):
    """Sheet set actually applicable to this cell. Producers get the full complex set; countries
    that don't grow the crop get the demand-side subset (matters for palm, not for the oilseeds)."""
    full = COMPLEXES[cx]["sheets"]
    if tier == "A" or (tier == "D" and country in TIER_D_PRODUCERS.get(cx, set())):
        return full
    if tier in ("B", "D"):
        return IMPORTER_SHEETS.get(cx, full) & full
    return full


def tiered_rows():
    """Yield (complex, country, tier) in build order: A, B, C(World), D, E."""
    for cx, cfg in COMPLEXES.items():
        for tier in ("A", "B"):
            for c in cfg.get(tier, []):
                yield cx, c, tier
        yield cx, "World", "C"
        for tier in ("D", "E"):
            for c in cfg.get(tier, []):
                yield cx, c, tier


def make_folders():
    created = []
    for c in sorted({c for _, c, _ in tiered_rows()}):
        d = OILSEEDS / c
        if not d.exists():
            d.mkdir(parents=True, exist_ok=True)
            created.append(c)
    return created


def build_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coverage"

    ws["A1"] = "Pepsi / Helios Pilot — Country × Complex Coverage Tracker"
    ws["A1"].font = Font(bold=True, size=13, color=GREEN, name="Calibri")
    ws["A2"] = ("SOW No. 1 scope: five complexes. Sheet set varies by complex (palm has no crush/meal; "
                "corn oil is derived). Bare bones = tiers A+B+C+D; tier E is deferred loop fill.")
    ws["A2"].font = Font(italic=True, size=9, color="6E7178", name="Calibri")

    headers = ["Complex", "Country", "Tier"] + SHEET_TYPES + ["Notes"]
    hr = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=j, value=h)
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(s for s in STATUSES if s), allow_blank=True)
    ws.add_data_validation(dv)

    r = hr + 1
    last_cx = None
    for cx, country, tier in tiered_rows():
        if last_cx is not None and cx != last_cx:
            r += 1
        last_cx = cx
        applicable = sheets_for(cx, country, tier)
        done = DONE.get((cx, country), set())

        ws.cell(row=r, column=1, value=cx).font = Font(name="Calibri")
        cc = ws.cell(row=r, column=2, value=country)
        cc.font = Font(bold=(country in done or bool(done)), name="Calibri")
        tc = ws.cell(row=r, column=3, value=TIER_LABEL[tier])
        tc.font = Font(size=9, name="Calibri", color=TIER_COLOR[tier])
        tc.border = BORDER

        for j, st in enumerate(SHEET_TYPES, start=4):
            cell = ws.cell(row=r, column=j)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if st not in applicable:
                cell.value = "N/A"
                cell.fill = NA_FILL
                cell.font = Font(size=8, color="A8ADB5", name="Calibri")
                continue
            dv.add(cell)
            if tier == "C":
                cell.value = "Planned"      # automated from bronze.fas_psd
            elif st in done:
                cell.value = "Done"
                cell.fill = DONE_FILL

        note = ws.cell(row=r, column=len(headers))
        note.border = BORDER
        bits = []
        if country in SOW_ORIGINS.get(cx, set()):
            bits.append("SOW origin")
        if done:
            bits.append("built — models/Oilseeds/United States/")
        if tier == "B":
            bits.append("shared importer workbook: tab per oil + allocation tab")
        if tier == "C":
            bits.append("auto from bronze.fas_psd — no manual build")
        if tier == "D":
            bits.append("stub only: production + trade + shock coefficient")
        if tier == "E":
            bits.append("deferred — not in the six-week build")
        if bits:
            note.value = " · ".join(bits)
            note.font = Font(italic=True, size=9, color="6E7178", name="Calibri")
        r += 1

    widths = [18, 16, 14] + [11] * len(SHEET_TYPES) + [52]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "D5"

    lr = r + 1
    ws.cell(row=lr, column=1, value="Tiers").font = Font(bold=True, size=9, name="Calibri")
    for i, (t, desc) in enumerate([
        ("A — exporter", "Price-setting exporter. Sets the reference series we quote. Full sheet set."),
        ("B — importer", "Swing importer. ONE workbook per country, tab per oil + shared allocation tab "
                         "(splits veg-oil demand across palm/sun/rape/soy on relative price)."),
        ("C — world", "World rollup, automated from bronze.fas_psd."),
        ("D — scenario", "Scenario-only origin. Production + trade + shock coefficient, no balance sheet."),
        ("E — loop fill", "Closes the trade matrix. DEFERRED — explicitly out of the six-week build."),
    ], start=1):
        ws.cell(row=lr + i, column=1, value=t).font = Font(size=9, color=TIER_COLOR[t[0]], name="Calibri")
        ws.cell(row=lr + i, column=2, value=desc).font = Font(size=9, color="6E7178", name="Calibri")

    sr = lr + 7
    ws.cell(row=sr, column=1, value="Sheet sets by complex").font = Font(bold=True, size=9, name="Calibri")
    for i, (cx, cfg) in enumerate(COMPLEXES.items(), start=1):
        ws.cell(row=sr + i, column=1, value=cx).font = Font(size=9, name="Calibri")
        ws.cell(row=sr + i, column=2,
                value=" · ".join(s for s in SHEET_TYPES if s in cfg["sheets"])
                ).font = Font(size=9, color="6E7178", name="Calibri")

    TRACKER.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TRACKER)


def main():
    created = make_folders()
    build_tracker()
    rows = list(tiered_rows())
    print(f"Tracker: {TRACKER}")
    print(f"  {len(rows)} country x complex rows across {len(COMPLEXES)} complexes")
    bare = full = 0
    for cx, cfg in COMPLEXES.items():
        n = {t: len(cfg.get(t, [])) for t in ("A", "B", "D", "E")}
        cells = sum(len(sheets_for(cx, c, t))
                    for t in ("A", "B", "D") for c in cfg.get(t, []))
        allc = cells + sum(len(sheets_for(cx, c, "E")) for c in cfg.get("E", []))
        bare += cells
        full += allc
        print(f"  {cx:20s} A={n['A']} B={n['B']} C=1 D={n['D']} E={n['E']}  "
              f"| sheets: {len(cfg['sheets'])} | bare-bones cells: {cells}")
    print(f"  BARE BONES (A+B+D): {bare} sheets   |   with tier E loop fill: {full}")
    print(f"  Folders created: {created or 'none (all existed)'}")


if __name__ == "__main__":
    main()
