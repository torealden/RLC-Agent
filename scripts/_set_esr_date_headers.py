"""Replace row 2 columns 35+ in each of the 12 ESR tabs in
us_soy_complex_trade.xlsm with weekly Thursday dates from 1993-10-07
onward. Real Date values so VBA IsDate() recognizes them."""

import openpyxl
import shutil
from datetime import date, timedelta
from pathlib import Path
from datetime import datetime

WB_PATH = Path('models/Oilseeds/us_soy_complex_trade.xlsm')

ESR_TABS = [
    'Weekly Soybean Export Sales',
    'Weekly Soybean Export Shipments',
    'Weekly Soybean Export Commits',
    'Weekly Soybean Export NMY Sales',
    'Weekly Meal Export Sales',
    'Weekly Meal Export Shipments',
    'Weekly Meal Export Commits',
    'Weekly Meal Export NMY Sales',
    'Weekly SBO Export Sales',
    'Weekly SBO Export Shipments',
    'Weekly SBO Export Commits',
    'Weekly SBO Export NMY Sales',
]

# Generate weekly Thursdays from 1993-10-07 onward
# Cap at the next Thursday after today so we have room for incoming data.
start = date(1993, 10, 7)
today = date.today()
# Next Thursday from today (or today if today is a Thursday)
days_to_thu = (3 - today.weekday()) % 7  # 3 = Thursday in Monday-0 indexing
end = today + timedelta(days=days_to_thu)
weeks = []
d = start
while d <= end:
    weeks.append(d)
    d += timedelta(days=7)
print(f'Generated {len(weeks)} Thursday week-endings from {weeks[0]} to {weeks[-1]}')

# Where to start writing in the sheet
START_COL = 35  # column AI — first date column in the existing layout
HEADER_ROW = 2

# Backup first
backup = WB_PATH.with_suffix(f'.xlsm.bak.{datetime.now().strftime("%Y%m%d-%H%M%S")}')
shutil.copy2(WB_PATH, backup)
print(f'Backup -> {backup.name}')

wb = openpyxl.load_workbook(WB_PATH, keep_vba=True)
for tab_name in ESR_TABS:
    if tab_name not in wb.sheetnames:
        print(f'  SKIP {tab_name} (not found)')
        continue
    ws = wb[tab_name]
    # Write the dates
    for i, dt in enumerate(weeks):
        cell = ws.cell(row=HEADER_ROW, column=START_COL + i)
        cell.value = dt
        # Format as Date so Excel/VBA recognize it
        cell.number_format = 'm/d/yyyy'
    # Clear any stale monthly date cells past the last weekly column we wrote
    last_col = START_COL + len(weeks) - 1
    for c in range(last_col + 1, ws.max_column + 1):
        ws.cell(HEADER_ROW, c).value = None
    print(f'  {tab_name}: wrote {len(weeks)} dates cols {START_COL}..{last_col}, cleared {ws.max_column - last_col} stale')

wb.save(WB_PATH)
print(f'\nSaved {WB_PATH.name}')

# Quick verification
wb2 = openpyxl.load_workbook(WB_PATH, read_only=True, keep_vba=False)
ws2 = wb2['Weekly Soybean Export Sales']
print('\nSpot-check Weekly Soybean Export Sales row 2:')
for c in [START_COL, START_COL+1, START_COL+10, START_COL+len(weeks)-1]:
    v = ws2.cell(HEADER_ROW, c).value
    print(f'  col {c}: {v} ({type(v).__name__})')
