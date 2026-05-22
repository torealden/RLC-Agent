"""
Ingest FAO Population (domain OA) from the bulk ZIP on disk.

Two outputs:
  1. bronze.faostat_data — long-format (area, element, year, value), domain='OA'
  2. models/Population/population_by_country.xlsx — wide flat-file for
     reference from any balance sheet (rows=years descending, columns=country).

Idempotent: re-running upserts on (country, element, year, domain).

Source: data/raw/oilseeds_fats_greases/Population_E_All_Data.zip (FAO bulk).
Refresh cadence: quarterly via the calendar reminder (15th of Mar/Jun/Sep/Dec).
"""
from __future__ import annotations

import argparse
import io
import logging
import os
import sys
import zipfile
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)

from dotenv import load_dotenv
load_dotenv()
import psycopg2
import psycopg2.extras
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


SRC_ZIP = PROJECT_ROOT / "data" / "raw" / "oilseeds_fats_greases" / "Population_E_All_Data.zip"
INTERNAL_CSV = "Population_E_All_Data_NOFLAG.csv"

OUT_DIR = PROJECT_ROOT / "models" / "Population"
OUT_XLSX = OUT_DIR / "population_by_country.xlsx"

# Countries we surface as columns in the flat file. ISO-2 keys map to FAO Area
# names (FAO uses inconsistent naming so we map a few variants per code).
# Add new entries here as project expansion brings new country balance sheets.
COUNTRY_MAP: dict[str, list[str]] = {
    "US":   ["United States of America", "United States"],
    "BR":   ["Brazil"],
    "AR":   ["Argentina"],
    "CN":   ["China, mainland", "China"],
    "EU":   ["European Union (27)", "European Union"],
    "IN":   ["India"],
    "ID":   ["Indonesia"],
    "MY":   ["Malaysia"],
    "RU":   ["Russian Federation", "Russia"],
    "UA":   ["Ukraine"],
    "AU":   ["Australia"],
    "CA":   ["Canada"],
    "MX":   ["Mexico"],
    "JP":   ["Japan"],
    "KR":   ["Republic of Korea", "Korea, Republic of"],
    "EG":   ["Egypt"],
    "TR":   ["Türkiye", "Turkey"],
    "VN":   ["Viet Nam", "Vietnam"],
    "PH":   ["Philippines"],
    "TH":   ["Thailand"],
    "PK":   ["Pakistan"],
    "BD":   ["Bangladesh"],
    "NG":   ["Nigeria"],
    "ZA":   ["South Africa"],
    "WORLD": ["World"],
}


# ---------------------------------------------------------------------
# Step 1 — Load CSV from ZIP
# ---------------------------------------------------------------------

def load_csv() -> pd.DataFrame:
    print(f"Reading {SRC_ZIP}...")
    with zipfile.ZipFile(SRC_ZIP) as z:
        with z.open(INTERNAL_CSV) as f:
            df = pd.read_csv(f, encoding="latin-1", low_memory=False)
    print(f"  loaded {len(df):,} rows, {len(df.columns)} columns")
    return df


# ---------------------------------------------------------------------
# Step 2 — Wide → long melt
# ---------------------------------------------------------------------

def melt_long(df: pd.DataFrame) -> pd.DataFrame:
    year_cols = [c for c in df.columns if c.startswith("Y") and c[1:].isdigit()]
    id_cols = [c for c in df.columns if c not in year_cols]
    long = df.melt(id_vars=id_cols, value_vars=year_cols, var_name="year_col", value_name="value")
    long = long.dropna(subset=["value"])
    long["year"] = long["year_col"].str[1:].astype(int)
    long = long.drop(columns=["year_col"])
    print(f"  melted to {len(long):,} long rows ({long['year'].min()}-{long['year'].max()})")
    return long


# ---------------------------------------------------------------------
# Step 3 — Upsert to bronze.faostat_data
# ---------------------------------------------------------------------

def get_conn():
    return psycopg2.connect(
        host=os.getenv("RLC_PG_HOST"), port=os.getenv("RLC_PG_PORT", "5432"),
        dbname=os.getenv("RLC_PG_DB", "rlc_commodities"),
        user=os.getenv("RLC_PG_USER"), password=os.getenv("RLC_PG_PASSWORD"),
        sslmode="require",
    )


def ensure_unique_index(conn):
    """Make sure we have an upsert key on bronze.faostat_data."""
    cur = conn.cursor()
    cur.execute("""
        SELECT 1 FROM pg_indexes
        WHERE schemaname='bronze' AND tablename='faostat_data'
          AND indexname='faostat_data_unique_grain'
    """)
    if not cur.fetchone():
        print("  creating unique index on (country, domain, element, year, commodity)...")
        cur.execute("""
            CREATE UNIQUE INDEX faostat_data_unique_grain
            ON bronze.faostat_data (country, domain, element, year, commodity)
        """)
        conn.commit()


def upsert_bronze(long_df: pd.DataFrame) -> int:
    rows = []
    for _, r in long_df.iterrows():
        rows.append((
            str(r["Area Code"]),                          # country (FAO area code as string)
            r["Area"],                                    # country_name
            "OA",                                         # commodity (using as item code)
            r.get("Item", "Population - Est. & Proj."),   # commodity_name
            "OA",                                         # domain
            r["Element"],                                 # element
            int(r["year"]),                               # year
            float(r["value"]) * 1000,                     # value (FAO unit 1000 No -> raw count)
            "people",                                     # unit (normalized)
            "",                                           # flag (NOFLAG file)
        ))

    with get_conn() as conn:
        ensure_unique_index(conn)
        cur = conn.cursor()
        # Truncate-and-reload pattern: faostat_data has 0 rows for OA today,
        # and reloads are cheap — 90k rows in ~10s.
        cur.execute("DELETE FROM bronze.faostat_data WHERE domain = 'OA'")
        n_deleted = cur.rowcount
        print(f"  deleted {n_deleted:,} existing OA rows")

        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO bronze.faostat_data
                (country, country_name, commodity, commodity_name, domain,
                 element, year, value, unit, flag)
               VALUES %s""",
            rows,
            page_size=2000,
        )
        conn.commit()
        print(f"  inserted {len(rows):,} OA rows to bronze.faostat_data")
    return len(rows)


# ---------------------------------------------------------------------
# Step 4 — Generate the flat file
# ---------------------------------------------------------------------

def build_flat_file(long_df: pd.DataFrame):
    """Pivot to wide for the Excel reference file. Total population only."""
    # Filter to Total Population, Both sexes
    tot = long_df[long_df["Element"] == "Total Population - Both sexes"].copy()
    tot["value_raw"] = tot["value"] * 1000

    # Build country lookup: FAO Area → ISO code we use
    rev_map = {}
    for iso, fao_names in COUNTRY_MAP.items():
        for fao_name in fao_names:
            rev_map[fao_name] = iso
    tot["iso"] = tot["Area"].map(rev_map)
    tot = tot.dropna(subset=["iso"])

    # Pivot: rows = years, columns = ISO country
    pivot = tot.pivot_table(index="year", columns="iso", values="value_raw", aggfunc="first")
    # Order columns by our COUNTRY_MAP order
    iso_order = [iso for iso in COUNTRY_MAP if iso in pivot.columns]
    pivot = pivot[iso_order]
    # Years descending (most recent on top — matches your existing balance sheet conventions)
    pivot = pivot.sort_index(ascending=False)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Population"

    # Header
    ws.cell(row=1, column=1, value="Year")
    for col_idx, iso in enumerate(iso_order, start=2):
        ws.cell(row=1, column=col_idx, value=iso)

    # Header fill color matches Tore's internal balance-sheet convention
    # (us_soybean_complex_bal_sheets.xlsm and family). Note: this is the
    # *internal* color. Public/client-facing artifacts (IFVS widget, Helios
    # deck) use the brand-kit INK #1B2A4A instead. See
    # reference_excel_color_conventions.md.
    header_fill = PatternFill(start_color="3C7D22", end_color="3C7D22", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    for col in range(1, len(iso_order) + 2):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, year in enumerate(pivot.index, start=2):
        ws.cell(row=row_idx, column=1, value=int(year))
        for col_idx, iso in enumerate(iso_order, start=2):
            val = pivot.loc[year, iso]
            ws.cell(row=row_idx, column=col_idx, value=float(val) if pd.notna(val) else None)
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"

    # Column widths
    ws.column_dimensions["A"].width = 8
    for col_idx in range(2, len(iso_order) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    # Metadata sheet
    meta = wb.create_sheet("_meta")
    meta["A1"] = "Source"
    meta["B1"] = "FAO FAOSTAT Annual Population (domain OA)"
    meta["A2"] = "Source zip"
    meta["B2"] = str(SRC_ZIP.name)
    meta["A3"] = "Generated"
    meta["B3"] = datetime.utcnow().isoformat(timespec="seconds") + " UTC"
    meta["A4"] = "Element shown"
    meta["B4"] = "Total Population - Both sexes (people, not thousands)"
    meta["A5"] = "Refresh"
    meta["B5"] = "Run scripts/ingest_fao_population.py after pulling a fresh ZIP"
    meta["A6"] = "Years covered"
    meta["B6"] = f"{int(pivot.index.min())} – {int(pivot.index.max())} (FAO includes UN projections)"
    meta["A7"] = "Countries"
    meta["B7"] = f"{len(iso_order)} columns: {', '.join(iso_order)}"
    for col in ["A", "B"]:
        meta.column_dimensions[col].width = 32
    for r in range(1, 8):
        meta[f"A{r}"].font = Font(bold=True)

    wb.save(OUT_XLSX)
    print(f"  wrote {OUT_XLSX}")
    print(f"    rows: {len(pivot)} years")
    print(f"    cols: {len(iso_order)} countries ({', '.join(iso_order)})")
    return OUT_XLSX


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--skip-bronze", action="store_true", help="Only write the flat file")
    p.add_argument("--skip-xlsx", action="store_true", help="Only refresh bronze")
    args = p.parse_args()

    df = load_csv()
    long_df = melt_long(df)

    if not args.skip_bronze:
        upsert_bronze(long_df)
    if not args.skip_xlsx:
        build_flat_file(long_df)
    print("done.")


if __name__ == "__main__":
    main()
