#!/usr/bin/env python3
"""Improved Census trade downloader.

This version focuses on:
- Resilient API calls (requests Session + retries + flag handling)
- Explicit unit conversion to the units used in the US Soybean Trade workbook
- Bronze-layer persistence with raw + converted values
- Spreadsheet updates that infer commodity/flow from sheet names

The original script is left untouched; this file can be used as a safer drop-in
that keeps the earlier API parameters but hardens error handling and unit
transforms.
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3 import Retry

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv(PROJECT_ROOT / ".env")
api_manager_env = PROJECT_ROOT / "api Manager" / ".env"
if api_manager_env.exists():
    load_dotenv(api_manager_env)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

CENSUS_API_BASE = "https://api.census.gov/data/timeseries/intltrade"


@dataclass(frozen=True)
class CommodityConfig:
    name: str
    hs_codes: Tuple[str, ...]
    flow_sheets: Dict[str, str]
    # multiplier applied to kg to reach the workbook units
    kg_to_target_multiplier: float
    target_unit: str


COMMODITIES: Dict[str, CommodityConfig] = {
    "SOYBEANS": CommodityConfig(
        name="Soybeans",
        hs_codes=("120110", "120190"),
        flow_sheets={"exports": "Soybean Exports", "imports": "Soybean Imports"},
        kg_to_target_multiplier=0.0367437 / 1000,  # kg -> thousand bushels
        target_unit="1000 bushels",
    ),
    "SOYBEAN_MEAL": CommodityConfig(
        name="Soybean Meal",
        hs_codes=("120810", "230400", "230499"),
        flow_sheets={"exports": "Soymeal Exports", "imports": "Soymeal Imports"},
        kg_to_target_multiplier=1 / 907.185,  # kg -> short tons
        target_unit="short tons",
    ),
    "SOYBEAN_OIL": CommodityConfig(
        name="Soybean Oil",
        hs_codes=("150710", "150790"),
        flow_sheets={"exports": "Soyoil Exports", "imports": "Soyoil Imports"},
        kg_to_target_multiplier=2.20462 / 1000,  # kg -> thousand lbs
        target_unit="1000 lbs",
    ),
}

WORKBOOK_PATH = PROJECT_ROOT / "Models" / "Oilseeds" / "US Soybean Trade.xlsx"


def build_session() -> requests.Session:
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=1.5,
        status_forcelist=(500, 502, 503, 504),
        allowed_methods=("GET",),
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def census_params(flow: str, hs_code: str, period: date, api_key: Optional[str]) -> Dict[str, str]:
    commodity_field = "I_COMMODITY" if flow == "imports" else "E_COMMODITY"
    value_field = "GEN_VAL_MO" if flow == "imports" else "ALL_VAL_MO"
    qty_fields = ["GEN_QY1_MO", "CON_QY1_MO"] if flow == "imports" else ["QTY_1_MO", "QTY_2_MO"]
    flag_fields = [f"{name}_FLAG" for name in qty_fields]
    get_fields = ",".join([value_field, *qty_fields, *flag_fields, "UNIT_QY1", "CTY_CODE", "CTY_NAME"])
    params = {
        "get": get_fields,
        commodity_field: hs_code,
        "time": period.strftime("%Y-%m"),
    }
    if api_key:
        params["key"] = api_key
    return params


def parse_quantity(record: List[str], headers: List[str]) -> Tuple[Optional[float], Optional[str]]:
    header_map = {name: idx for idx, name in enumerate(headers)}
    unit = record[header_map.get("UNIT_QY1", -1)] if "UNIT_QY1" in header_map else None
    for qty_field in ("GEN_QY1_MO", "CON_QY1_MO", "QTY_1_MO", "QTY_2_MO"):
        flag_field = f"{qty_field}_FLAG"
        if qty_field in header_map and flag_field in header_map:
            qty_value = record[header_map[qty_field]]
            flag_value = record[header_map[flag_field]]
            if qty_value and flag_value == "":  # empty flag means the value is valid
                try:
                    return float(qty_value), unit
                except ValueError:
                    continue
    return None, unit


def fetch_month(session: requests.Session, commodity: CommodityConfig, flow: str, period: date, api_key: Optional[str]) -> List[Dict[str, object]]:
    url = f"{CENSUS_API_BASE}/{flow}/hs"
    records: List[Dict[str, object]] = []
    for hs_code in commodity.hs_codes:
        params = census_params(flow, hs_code, period, api_key)
        resp = session.get(url, params=params, timeout=30)
        resp.raise_for_status()
        raw = resp.json()
        if not raw or len(raw) <= 1:
            continue
        headers = raw[0]
        for row in raw[1:]:
            qty_kg, unit = parse_quantity(row, headers)
            value_usd = row[headers.index("GEN_VAL_MO" if flow == "imports" else "ALL_VAL_MO")]
            try:
                value_usd = float(value_usd)
            except (TypeError, ValueError):
                value_usd = None
            converted_qty = qty_kg * commodity.kg_to_target_multiplier if qty_kg is not None else None
            records.append(
                {
                    "flow": flow,
                    "hs_code": hs_code,
                    "period": period,
                    "country_code": row[headers.index("CTY_CODE")],
                    "country_name": row[headers.index("CTY_NAME")],
                    "quantity_kg": qty_kg,
                    "quantity_converted": converted_qty,
                    "converted_unit": commodity.target_unit,
                    "value_usd": value_usd,
                    "reported_unit": unit,
                }
            )
    return records


def iterate_periods(years_back: int) -> Iterable[date]:
    today = date.today().replace(day=1)
    for offset in range(years_back * 12):
        month = today.month - offset
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        yield date(year, month, 1)


def write_bronze(records: List[Dict[str, object]], table: str = "census_trade_bronze") -> None:
    import psycopg2
    from psycopg2.extras import execute_values
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        logger.warning("DATABASE_URL not set; skipping bronze load")
        return
    conn = psycopg2.connect(database_url)
    conn.autocommit = True
    with conn, conn.cursor() as cur:
        cur.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {table} (
                flow TEXT,
                hs_code TEXT,
                period DATE,
                country_code TEXT,
                country_name TEXT,
                quantity_kg DOUBLE PRECISION,
                quantity_converted DOUBLE PRECISION,
                converted_unit TEXT,
                value_usd DOUBLE PRECISION,
                reported_unit TEXT,
                inserted_at TIMESTAMPTZ DEFAULT NOW()
            )
            """
        )
        rows = [
            (
                r["flow"],
                r["hs_code"],
                r["period"],
                r["country_code"],
                r["country_name"],
                r["quantity_kg"],
                r["quantity_converted"],
                r["converted_unit"],
                r["value_usd"],
                r["reported_unit"],
            )
            for r in records
        ]
        execute_values(
            cur,
            f"INSERT INTO {table} (flow, hs_code, period, country_code, country_name, quantity_kg, quantity_converted, converted_unit, value_usd, reported_unit) VALUES %s",
            rows,
        )
    conn.close()


def update_workbook(records: List[Dict[str, object]], commodity: CommodityConfig) -> None:
    if not records:
        logger.info("No records to write to Excel")
        return
    if not WORKBOOK_PATH.exists():
        logger.warning("Workbook path not found; skipping Excel update")
        return
    wb = load_workbook(WORKBOOK_PATH)
    by_sheet: Dict[str, List[Dict[str, object]]] = {"exports": [], "imports": []}
    for record in records:
        by_sheet[record["flow"]].append(record)
    for flow, sheet_records in by_sheet.items():
        if not sheet_records:
            continue
        sheet_name = commodity.flow_sheets[flow]
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet %s not found in workbook", sheet_name)
            continue
        ws = wb[sheet_name]
        start_row = ws.max_row + 1
        ws.cell(row=1, column=ws.max_column + 1, value="(appended)")
        for idx, record in enumerate(sorted(sheet_records, key=lambda r: r["period"])):
            row_idx = start_row + idx
            ws.cell(row=row_idx, column=1, value=record["period"].strftime("%Y-%m"))
            ws.cell(row=row_idx, column=2, value=record["country_name"])
            ws.cell(row=row_idx, column=3, value=record["quantity_converted"])
            ws.cell(row=row_idx, column=4, value=record["converted_unit"])
            ws.cell(row=row_idx, column=5, value=record["value_usd"])
    wb.save(WORKBOOK_PATH)
    logger.info("Workbook updated at %s", WORKBOOK_PATH)


def main() -> None:
    parser = argparse.ArgumentParser(description="Pull Census trade data with improved reliability.")
    parser.add_argument("--commodity", choices=sorted(COMMODITIES.keys()) + ["ALL"], default="SOYBEANS")
    parser.add_argument("--years", type=int, default=1, help="How many years back to pull (monthly granularity)")
    parser.add_argument("--flow", choices=["imports", "exports", "both"], default="both")
    parser.add_argument("--api-key", dest="api_key", default=os.getenv("CENSUS_API_KEY"))
    parser.add_argument("--save-to-db", action="store_true")
    parser.add_argument("--update-excel", action="store_true")
    args = parser.parse_args()

    commodities = COMMODITIES.values() if args.commodity == "ALL" else [COMMODITIES[args.commodity]]
    flows = ["imports", "exports"] if args.flow == "both" else [args.flow]

    session = build_session()
    all_records: List[Dict[str, object]] = []
    for commodity in commodities:
        logger.info("Pulling Census data for %s", commodity.name)
        for period in iterate_periods(args.years):
            for flow in flows:
                try:
                    month_records = fetch_month(session, commodity, flow, period, args.api_key)
                except requests.HTTPError as exc:
                    logger.warning("HTTP error for %s %s %s: %s", commodity.name, flow, period, exc)
                    continue
                all_records.extend(month_records)
        if args.save_to_db:
            write_bronze(all_records)
        if args.update_excel:
            update_workbook(all_records, commodity)
        all_records.clear()


if __name__ == "__main__":
    main()
