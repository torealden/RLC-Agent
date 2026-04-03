"""
Extend us_oilseed_crush.xlsm with NASS Low CI (fats/greases),
Other Veg Oils, and Peanut crush tabs.

Matches the structure of the existing World Crush file's NASS Low CI
and NASS Other tabs.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

fp = "C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models/Oilseeds/new_models/us_oilseed_crush.xlsm"
wb = load_workbook(fp, keep_vba=True)

print(f"Existing sheets: {wb.sheetnames}")

NAVY = '1F4E79'
HEADER_FONT = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
SECTION_FONT = Font(name='Calibri', size=10, bold=True, color=NAVY)
DATE_FONT = Font(name='Calibri', size=9)


def create_tab(wb, tab_name, commodities):
    """Create a tab with dates down rows, commodities across columns."""
    ws = wb.create_sheet(title=tab_name)

    col = 2
    for commodity in commodities:
        items = commodity.get('items', ['Production', 'Processing Use', 'End-of-Month Stocks'])
        span = len(items)

        ws.cell(row=1, column=col, value=commodity['name']).font = SECTION_FONT

        for i, item in enumerate(items):
            cell = ws.cell(row=2, column=col + i, value=item)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        unit = commodity.get('unit', 'mil lbs')
        for i in range(span):
            ws.cell(row=3, column=col + i, value=unit).font = Font(
                name='Calibri', size=8, italic=True, color='808080')

        col += span + 1

    ws.cell(row=2, column=1, value='Date').font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL

    row = 4
    current = date(1979, 9, 1)
    end = date(2026, 12, 1)
    while current <= end:
        ws.cell(row=row, column=1, value=current).font = DATE_FONT
        ws.cell(row=row, column=1).number_format = 'MMM-YY'
        row += 1
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)

    ws.column_dimensions['A'].width = 10
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.freeze_panes = 'B4'
    return ws, row - 1


# ── NASS Low CI (Fats & Greases) ───────────────────────────────────
fats = [
    {'name': 'Choice White Grease', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
    {'name': 'Feather Meal', 'items': ['Production', 'End-of-Month Stocks']},
    {'name': 'Lard', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
    {'name': 'Meat and Bone Meal', 'items': ['Production', 'End-of-Month Stocks']},
    {'name': 'Other Grease', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
    {'name': 'Poultry Fats', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
    {'name': 'Poultry By-Product', 'items': ['Production', 'End-of-Month Stocks']},
    {'name': 'Edible Tallow', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
    {'name': 'Inedible Tallow', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
    {'name': 'Technical Tallow', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
    {'name': 'Yellow Grease', 'items': ['Production', 'Processing Use', 'End-of-Month Stocks']},
]

ws1, r1 = create_tab(wb, 'NASS Low CI', fats)
print(f"Created 'NASS Low CI': {r1} rows, 11 commodities")

# ── NASS Other Veg Oils ────────────────────────────────────────────
oils = [
    {'name': 'Palm Kernel Oil', 'items': ['Refined Consumption', 'Stocks']},
    {'name': 'Palm Oil', 'items': ['Refined Consumption', 'Edible Use', 'Inedible Use', 'Stocks']},
    {'name': 'Safflower Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Removed for Processing', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Sunflower Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Consumption', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Coconut Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Consumption', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Corn Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Consumption', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Canola Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Consumption', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Cottonseed Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Consumption', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Peanut Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Consumption', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Linseed Oil', 'items': ['Crude Processed', 'Refined Produced', 'Crude Stocks', 'Refined Stocks']},
    {'name': 'Soybean Oil', 'items': ['Crude Processed', 'Refined Produced', 'Refined Consumption', 'Edible Use', 'Inedible Use', 'Crude Stocks', 'Refined Stocks']},
]

ws2, r2 = create_tab(wb, 'NASS Other Veg Oils', oils)
print(f"Created 'NASS Other Veg Oils': {r2} rows, 11 oils")

# ── Peanut Crush ───────────────────────────────────────────────────
peanut = [
    {'name': 'Peanuts', 'items': ['Crush', 'Stocks at Mills'], 'unit': '000 ST'},
    {'name': 'Peanut Meal', 'items': ['Production', 'Stocks', 'Yield'], 'unit': '000 ST'},
    {'name': 'Peanut Oil', 'items': ['Crude Production', 'Crude Stocks', 'Refined Stocks', 'Total Stocks', 'Yield'], 'unit': 'mil lbs'},
]

ws3, r3 = create_tab(wb, 'peanut_crush', peanut)
print(f"Created 'peanut_crush': {r3} rows")

# Save
wb.save(fp)
print(f"\nSaved. Sheets: {wb.sheetnames}")
