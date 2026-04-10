"""Load US oilseed crush facilities into reference.oilseed_crush_facilities."""

import os
import sys
from pathlib import Path

import openpyxl
import psycopg2
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(PROJECT_ROOT / '.env')

DROPBOX = Path("C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models/Oilseeds")


def sf(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def main():
    # Read North American file (main capacity data)
    wb = openpyxl.load_workbook(
        DROPBOX / "North American Oilseed Crushing Capacity.xlsx", data_only=True
    )
    ws = wb['US Soy Crush']

    na_facilities = []
    for row in range(2, ws.max_row + 1):
        company = ws.cell(row=row, column=1).value
        if not company:
            continue
        na_facilities.append({
            'company': str(company).strip(),
            'status': str(ws.cell(row=row, column=2).value or 'operating').strip().lower(),
            'street': str(ws.cell(row=row, column=3).value or '').strip(),
            'city': str(ws.cell(row=row, column=4).value or '').strip(),
            'state': str(ws.cell(row=row, column=5).value or '').strip(),
            'zip_code': str(ws.cell(row=row, column=6).value or '').strip(),
            'daily_tons': sf(ws.cell(row=row, column=7).value),
            'yearly_tons': sf(ws.cell(row=row, column=8).value),
            'yearly_bushels': sf(ws.cell(row=row, column=9).value),
            'oil_cap_tons': sf(ws.cell(row=row, column=10).value),
            'oil_cap_mil_lbs': sf(ws.cell(row=row, column=11).value),
            'expansion': str(ws.cell(row=row, column=12).value or '').strip(),
        })
    wb.close()
    print(f"Read {len(na_facilities)} facilities from North American file")

    # Read master file for enrichment
    wb2 = openpyxl.load_workbook(
        DROPBOX / "new_models/us_oilseed_crushing_capacity.xlsm",
        data_only=True, keep_vba=True
    )
    ws2 = wb2['Facility Database']
    master = {}
    for row in range(2, ws2.max_row + 1):
        company = ws2.cell(row=row, column=3).value
        city = ws2.cell(row=row, column=2).value
        state = ws2.cell(row=row, column=1).value
        if company:
            key = (str(company).strip().lower(), str(city).strip().lower(),
                   str(state).strip().upper())
            yr_built = ws2.cell(row=row, column=8).value
            last_exp = ws2.cell(row=row, column=9).value
            try:
                yr_built = int(yr_built)
            except (TypeError, ValueError):
                yr_built = None
            try:
                last_exp = int(last_exp)
            except (TypeError, ValueError):
                last_exp = None
            master[key] = {
                'facility': str(ws2.cell(row=row, column=4).value or ''),
                'cap_bpd': sf(ws2.cell(row=row, column=5).value),
                'cap_mbu_yr': sf(ws2.cell(row=row, column=6).value),
                'year_built': yr_built,
                'last_expansion': last_exp,
                'exp_details': str(ws2.cell(row=row, column=10).value or ''),
                'refining': str(ws2.cell(row=row, column=11).value or ''),
                'nopa': str(ws2.cell(row=row, column=13).value or ''),
                'status': str(ws2.cell(row=row, column=14).value or ''),
                'investment': sf(ws2.cell(row=row, column=15).value),
                'notes': str(ws2.cell(row=row, column=17).value or ''),
            }
    wb2.close()
    print(f"Read {len(master)} facilities from master file")

    # Insert into DB
    conn = psycopg2.connect(
        host=os.getenv('RLC_PG_HOST'), dbname='rlc_commodities',
        user=os.getenv('RLC_PG_USER'), password=os.getenv('RLC_PG_PASSWORD'),
        sslmode='require'
    )
    cur = conn.cursor()
    inserted = 0

    for f in na_facilities:
        s = f['status']
        if s == 'open':
            s = 'operating'

        mkey = (f['company'].lower(), f['city'].strip().lower(),
                f['state'].strip().upper())
        m = master.get(mkey, {})

        has_refining = bool(m.get('refining') and m['refining'] not in ('', 'None', 'No'))
        nopa = m.get('nopa', '').upper() in ('YES', 'Y', 'TRUE')

        cur.execute("""
            INSERT INTO reference.oilseed_crush_facilities
                (country, company, facility_name, city, state, zip_code, street_address,
                 crush_capacity_tons_per_day, crush_capacity_tons_per_year,
                 crush_capacity_bushels_per_year,
                 oil_production_capacity_tons_per_year, oil_production_capacity_mil_lbs,
                 has_oil_refining, primary_oilseed, status,
                 year_built, last_expansion_year, expansion_details,
                 investment_mil_usd, nopa_member, notes, source, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT (country, company, city, state) DO UPDATE SET
                crush_capacity_tons_per_day = COALESCE(EXCLUDED.crush_capacity_tons_per_day, reference.oilseed_crush_facilities.crush_capacity_tons_per_day),
                crush_capacity_tons_per_year = COALESCE(EXCLUDED.crush_capacity_tons_per_year, reference.oilseed_crush_facilities.crush_capacity_tons_per_year),
                crush_capacity_bushels_per_year = COALESCE(EXCLUDED.crush_capacity_bushels_per_year, reference.oilseed_crush_facilities.crush_capacity_bushels_per_year),
                oil_production_capacity_tons_per_year = COALESCE(EXCLUDED.oil_production_capacity_tons_per_year, reference.oilseed_crush_facilities.oil_production_capacity_tons_per_year),
                oil_production_capacity_mil_lbs = COALESCE(EXCLUDED.oil_production_capacity_mil_lbs, reference.oilseed_crush_facilities.oil_production_capacity_mil_lbs),
                has_oil_refining = EXCLUDED.has_oil_refining,
                status = EXCLUDED.status,
                updated_at = NOW()
        """, ('US', f['company'], m.get('facility', ''), f['city'], f['state'],
              f['zip_code'], f['street'],
              f['daily_tons'], f['yearly_tons'], f['yearly_bushels'],
              f['oil_cap_tons'], f['oil_cap_mil_lbs'],
              has_refining, 'soybeans', s,
              m.get('year_built'), m.get('last_expansion'),
              m.get('exp_details', ''),
              m.get('investment'), nopa, m.get('notes', ''),
              'NA_Crush_Capacity.xlsx + us_oilseed_crushing_capacity.xlsm'))
        inserted += 1

    # Add expansion/new plants from master not in NA file
    status_map = {
        'Planned 2025': 'planned',
        'Announced': 'announced',
        'Under Construction': 'under_construction',
    }
    for mkey, m in master.items():
        if m['status'] in status_map:
            company = mkey[0].title()
            city = mkey[1].title()
            state = mkey[2]
            cap_bpy = (m.get('cap_mbu_yr') or 0) * 1e6
            db_status = status_map[m['status']]

            cur.execute("""
                INSERT INTO reference.oilseed_crush_facilities
                    (country, company, facility_name, city, state,
                     crush_capacity_bushels_per_year,
                     primary_oilseed, status, expansion_details,
                     investment_mil_usd, notes, source, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                ON CONFLICT (country, company, city, state) DO UPDATE SET
                    status = EXCLUDED.status,
                    expansion_details = EXCLUDED.expansion_details,
                    crush_capacity_bushels_per_year = GREATEST(
                        EXCLUDED.crush_capacity_bushels_per_year,
                        reference.oilseed_crush_facilities.crush_capacity_bushels_per_year
                    ),
                    updated_at = NOW()
            """, ('US', company, m.get('facility', ''), city, state,
                  cap_bpy if cap_bpy > 0 else None,
                  'soybeans', db_status, m.get('exp_details', ''),
                  m.get('investment'), m.get('notes', ''),
                  'us_oilseed_crushing_capacity.xlsm'))
            inserted += 1

    conn.commit()

    # Verify
    cur.execute("""
        SELECT status, COUNT(*),
            COALESCE(SUM(crush_capacity_bushels_per_year), 0) / 1e6 as mil_bu
        FROM reference.oilseed_crush_facilities
        WHERE country = 'US'
        GROUP BY status ORDER BY mil_bu DESC
    """)
    print(f"\n=== reference.oilseed_crush_facilities (US) ===")
    for r in cur.fetchall():
        print(f"  {r[0]:25s}  {r[1]:3d} plants  {float(r[2]):>10,.1f} M bu/yr")

    cur.execute("""
        SELECT COUNT(*), SUM(crush_capacity_bushels_per_year) / 1e6
        FROM reference.oilseed_crush_facilities WHERE country = 'US'
    """)
    r = cur.fetchone()
    print(f"\n  TOTAL: {r[0]} plants, {float(r[1] or 0):,.1f} M bu/yr")
    conn.close()
    print(f"\nInserted/updated {inserted} facilities")


if __name__ == '__main__':
    main()
