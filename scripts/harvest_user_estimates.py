"""
Harvest Tore's balance-sheet projections (green cells) from the model
workbooks into silver.user_sd_estimate.

WHAT COUNTS AS "MINE": cells whose font color is RLC green #3C7D22 — the
workbooks' own legend ("Bold, green numbers are RLC estimates and
predictions"). Black/bold-black cells are history or official estimates
(e.g. CONAB in generated country books) and are NEVER harvested.

MARKETING-YEAR CONVENTION: START year — marketing_year 2024 == MY 2024/25,
parsed from the row-3 labels ('2024/25' -> 2024). This matches
bronze.fas_psd, silver.monthly_realized, and gold.projection_comparison_long.
NOTE: the schema comment on silver.user_sd_estimate says the opposite
(END year); the three legacy CSV rows (estimate_date 2026-01-30) are of
ambiguous vintage and are left untouched.

VINTAGE SEMANTICS: each run that finds changed values for a (commodity,
country, MY) writes a new vintage dated today, flips the old one to
is_current=FALSE via silver.mark_previous_estimates_not_current() (the
correctly-scoped helper — the CSV loader's whole-commodity flip is a known
bug), and upserts on the natural key. Unchanged values are skipped, so
re-running is a no-op and the vintage ladder only grows when the workbook
actually changed.

Usage:
    python scripts/harvest_user_estimates.py --list
    python scripts/harvest_user_estimates.py --dry-run
    python scripts/harvest_user_estimates.py --file "models/Oilseeds/United States/us_soybean_complex_bal_sheets.xlsm"
    python scripts/harvest_user_estimates.py            # harvest everything
"""
import argparse
import glob
import os
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.services.database.db_config import get_connection  # noqa: E402

RLC_GREEN_SUFFIX = '3C7D22'

# Commodity comes from the IN-SHEET title (row 2/1, e.g. 'US SOYBEAN OIL
# SUPPLY AND DEMAND'), not the tab name: us_lauric_oils_bal_sheets.xlsm is a
# stale clone whose tabs are still named soyoil_balance_sheet etc., and tab
# names drift under template cloning. Title slug -> commodity, with aliases
# where the repo already has a convention.
COMMODITY_ALIAS = {
    'soybean': 'soybeans',
    'peanut': 'peanuts',
    'canola': 'rapeseed',
}

# filename token -> words the sheet titles must contain. A workbook whose
# titles don't match its own filename (stale clone) is skipped loudly.
FILENAME_COMPLEX = {
    'soybean': ('SOYBEAN',),
    'palm': ('PALM',),
    'peanut': ('PEANUT',),
    'safflower': ('SAFFLOWER',),
    'sunflower': ('SUNFLOWER',),
    'canola': ('CANOLA', 'RAPESEED'),
    'rapeseed': ('RAPESEED', 'CANOLA'),
    'copra': ('COCONUT', 'COPRA'),
    'coconut': ('COCONUT', 'COPRA'),
    'lauric': ('COCONUT', 'PALM KERNEL', 'LAURIC'),
    'corn_oil': ('CORN',),
    'cottonseed': ('COTTONSEED',),
    'flaxseed': ('FLAX', 'LINSEED'),   # linseed == flaxseed oil/meal
}

TITLE_RE = re.compile(
    r'^(?:US|U\.S\.|[A-Z ]*?)\s*([A-Z][A-Z ]+?)\s+SUPPLY AND DEMAND\s*$')


def commodity_from_title(title: str, country: str) -> str | None:
    """'US SOYBEAN OIL SUPPLY AND DEMAND' -> 'soybean_oil'."""
    if not title:
        return None
    t = str(title).strip().upper()
    for prefix in (country.upper(), 'US', 'U.S.'):
        if t.startswith(prefix + ' '):
            t = t[len(prefix) + 1:]
            break
    if t.endswith('SUPPLY AND DEMAND'):
        t = t[:-len('SUPPLY AND DEMAND')].strip()
    else:
        return None
    slug = t.lower().replace(' ', '_')
    return COMMODITY_ALIAS.get(slug, slug)

# ordered (regex, column) — first match wins; anything unmatched is logged
LABEL_MAP = [
    (r'^Planted Area', 'area_planted'),
    (r'^Harvested Area\b', 'area_harvested'),
    (r'^Average .*Yield.*bushels per acre', 'yield'),
    (r'^Beginning Stocks', 'beginning_stocks'),
    (r'^Production$', 'production'),
    (r'^Imports$', 'imports'),
    (r'^Total Supply$', 'total_supply'),
    (r'^Crush$', 'crush'),
    (r'^Residual$', 'feed_residual'),
    (r'^Domestic Use$', 'domestic_use'),
    (r'^Exports$', 'exports'),
    (r'^Total (Demand|Use)$', 'total_use'),
    (r'^Ending Stocks', 'ending_stocks'),
    (r'^Stocks-to-Use$', 'stocks_use_ratio'),
]

UNIT_MAP = [
    (r'million bushels', 'mil bu'),
    (r'million pounds', 'mil lbs'),
    (r'thousand short tons', '1000 ST'),
    (r'thousand (metric )?tonn?e?s', '1000 MT'),
]

LEGEND_RE = re.compile(r'Bold.*(estimat|number|predict)', re.I)
MY_LABEL_RE = re.compile(r'^(\d{4})/\d{2}')

# round to column precision (NUMERIC(18,2) except these) so a re-run
# compares equal against what the DB actually stored
FIELD_DECIMALS = {'yield': 4, 'stocks_use_ratio': 4}


def is_rlc_green(cell) -> bool:
    f = cell.font
    return bool(f and f.color and f.color.rgb
                and str(f.color.rgb).upper().endswith(RLC_GREEN_SUFFIX))


def parse_tab(ws, tab_name: str, report: dict):
    """Return (unit, {my: {field: value}}) of green-cell values in the
    annual summary block."""
    # column -> marketing year (START year) from row-3 labels
    my_by_col = {}
    for col in range(2, ws.max_column + 1):
        v = ws.cell(3, col).value
        m = MY_LABEL_RE.match(str(v)) if v is not None else None
        if m:
            my_by_col[col] = int(m.group(1))

    unit = None
    rows = {}          # field -> row index
    unmapped = []
    for r in range(4, min(ws.max_row, 60) + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label = str(label).strip()
        if LEGEND_RE.search(label):
            break
        if label.isupper() and len(label) > 8:   # next section header
            break
        field = next((col for rx, col in LABEL_MAP
                      if re.match(rx, label, re.I)), None)
        if field is None:
            unmapped.append(label)
            continue
        rows.setdefault(field, r)
        if unit is None:
            for rx, u in UNIT_MAP:
                if re.search(rx, label, re.I):
                    unit = u
                    break

    estimates = {}     # my -> {field: value}
    green_none = 0
    for field, r in rows.items():
        for col, my in my_by_col.items():
            cell = ws.cell(r, col)
            if not is_rlc_green(cell):
                continue
            if cell.value is None:
                green_none += 1
                continue
            if not isinstance(cell.value, (int, float)):
                continue
            estimates.setdefault(my, {})[field] = round(
                float(cell.value), FIELD_DECIMALS.get(field, 2))

    report[tab_name] = {
        'fields': sorted(rows), 'unmapped': sorted(set(unmapped)),
        'green_none': green_none, 'mys': sorted(estimates),
    }
    if green_none:
        print(f'  WARNING {tab_name}: {green_none} green cells have no '
              'value. If the workbook has NOT been saved by Excel since '
              'generation, open + save + re-run; if it has, the cells are '
              'genuinely empty (formulas never written — e.g. the US canola '
              'meal Stocks-to-Use row) and need filling in the workbook.')
    return unit or 'mil bu', estimates


def expected_words(filename: str):
    for token, words in FILENAME_COMPLEX.items():
        if token in filename.lower():
            return words
    return None


def harvest_workbook(path: Path, only_commodity: str | None):
    country = path.parent.name
    wb = openpyxl.load_workbook(path, data_only=True)
    results = []       # (commodity, country, my, unit, {field: value})
    report = {}
    words = expected_words(path.name)
    for tab in wb.sheetnames:
        if not tab.endswith('_balance_sheet'):
            continue
        ws = wb[tab]
        title = ws.cell(2, 1).value or ws.cell(1, 1).value
        if words and title and not any(w in str(title).upper()
                                       for w in words):
            print(f'  SKIPPING tab {tab}: sheet title {title!r} does not '
                  f'match workbook name {path.name!r} (stale template '
                  'clone?) — fix the workbook, then re-run.')
            continue
        commodity = commodity_from_title(title, country)
        if commodity is None:
            print(f'  SKIPPING tab {tab}: could not parse commodity from '
                  f'title {title!r}.')
            continue
        if only_commodity and commodity != only_commodity:
            continue
        unit, estimates = parse_tab(ws, tab, report)
        for my, fields in sorted(estimates.items()):
            results.append((commodity, country, my, unit, fields))
    wb.close()
    return results, report


def values_match(existing: dict, fields: dict) -> bool:
    for k, v in fields.items():
        ev = existing.get(k)
        if ev is None or abs(float(ev) - v) > 1e-6:
            return False
    return True


def write_estimates(results, source_file: str, dry_run: bool) -> None:
    if not results:
        print('  nothing to write.')
        return
    today = date.today()
    with get_connection() as conn:
        cur = conn.cursor()
        inserted = skipped = 0
        for commodity, country, my, unit, fields in results:
            cur.execute("""
                SELECT * FROM silver.user_sd_estimate
                WHERE commodity=%s AND country=%s AND marketing_year=%s
                  AND is_current
            """, (commodity, country, my))
            existing = cur.fetchone()
            if existing and values_match(dict(existing), fields):
                skipped += 1
                continue
            if dry_run:
                print(f'  WOULD WRITE {commodity} {country} MY{my} '
                      f'({unit}): {fields}')
                inserted += 1
                continue
            cur.execute(
                'SELECT silver.mark_previous_estimates_not_current(%s,%s,%s)',
                (commodity, country, my))
            cols = ['commodity', 'country', 'marketing_year', 'estimate_date',
                    'unit', 'source_file', 'notes', 'is_current'] + list(fields)
            vals = [commodity, country, my, today, unit, source_file,
                    'harvested green cells', True] + list(fields.values())
            updates = ', '.join(
                f'{c}=EXCLUDED.{c}' for c in cols[4:] )
            cur.execute(f"""
                INSERT INTO silver.user_sd_estimate ({', '.join(cols)})
                VALUES ({', '.join(['%s'] * len(vals))})
                ON CONFLICT (commodity, country, marketing_year, estimate_date)
                DO UPDATE SET {updates}, updated_at = NOW()
            """, vals)
            inserted += 1
        verb = 'would write' if dry_run else 'wrote'
        print(f'  {verb} {inserted} vintage rows, {skipped} unchanged '
              '(skipped).')


def discover() -> list[Path]:
    pats = ['models/Oilseeds/*/*_bal_sheets.xls[mx]',
            'models/Oilseeds/*/*_balance_sheets.xls[mx]']
    files = []
    for p in pats:
        files += [Path(f) for f in glob.glob(str(PROJECT_ROOT / p))]
    return sorted(f for f in files
                  if 'archive' not in str(f).lower()
                  and not f.name.startswith('~$'))


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument('--file', help='harvest a single workbook')
    ap.add_argument('--commodity', help='restrict to one commodity')
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--list', action='store_true',
                    help='list discovered workbooks and exit')
    ap.add_argument('--monthly', action='store_true',
                    help='(not implemented) harvest monthly blocks into '
                         'silver.monthly_expectation')
    args = ap.parse_args()

    if args.monthly:
        sys.exit('--monthly is not implemented yet (annual harvest first; '
                 'block grammar is known, see docs/specs).')

    files = ([Path(args.file)] if args.file else discover())
    if args.list:
        for f in files:
            print(f)
        return
    if not files:
        print('No workbooks found under models/Oilseeds/*/.')
        return

    for f in files:
        print(f'{f.relative_to(PROJECT_ROOT) if f.is_absolute() and f.is_relative_to(PROJECT_ROOT) else f}:')
        try:
            results, report = harvest_workbook(f, args.commodity)
        except Exception as e:
            print(f'  ERROR reading workbook: {e}')
            continue
        for tab, r in report.items():
            print(f'  {tab}: fields={r["fields"]} green MYs={r["mys"]}'
                  + (f' unmapped={r["unmapped"]}' if r['unmapped'] else ''))
        if not results:
            print('  no green cells found (nothing marked as RLC estimate).')
            continue
        rel = str(f.relative_to(PROJECT_ROOT)) if f.is_absolute() \
            and f.is_relative_to(PROJECT_ROOT) else str(f)
        write_estimates(results, rel, args.dry_run)


if __name__ == '__main__':
    main()
