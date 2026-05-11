"""
Extract CARB LCFS pathway data → structured JSON/CSV.

Source: https://ww2.arb.ca.gov/sites/default/files/classic/fuels/lcfs/fuelpathways/current-pathways_all.xlsx
Cached locally at: domain_knowledge/external_lists/carb_lcfs/current_pathways_all.xlsx

The full CARB workbook has ~3,725 pathway applications spanning everything
from biodiesel and renewable diesel to electric forklifts and hydrogen
generators. This script filters to LIQUID BIOFUELS (biodiesel,
renewable diesel, SAF/AJF, jet) and emits structured records suitable
for cross-reference against our facility DB.

Output:
    domain_knowledge/external_lists/carb_lcfs/biofuel_pathways.json
    domain_knowledge/external_lists/carb_lcfs/biofuel_pathways.csv

Each pathway record has:
    pathway_id, class (Tier 1 / Tier 2 / Lookup Table),
    applicant_description (full text), facility_name, fuel_producer,
    facility_location (state, sometimes city in description),
    feedstock, fuel_type, ci (current certified carbon intensity),
    certification_date
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Optional

import openpyxl


SRC = Path("domain_knowledge/external_lists/carb_lcfs/current_pathways_all.xlsx")
OUT_DIR = Path("domain_knowledge/external_lists/carb_lcfs")


def is_liquid_biofuel(fuel_type: str) -> bool:
    """Filter out non-liquid-biofuel pathways (forklift, H2, EV charging, etc)."""
    if not fuel_type:
        return False
    ft = fuel_type.lower()
    return any(s in ft for s in [
        "biodiesel", "renewable diesel", "saf", "sustainable aviation",
        "alternative jet", "renewable jet", "jet fuel",
    ])


def parse_facility_from_description(applicant_desc: str) -> tuple[str, str]:
    """
    Extract (fuel_producer, facility_name) from the applicant description.
    Format is typically:
        "Fuel Producer: REG Newton, LLC (3514) ; Facility Name: REG Newton, LLC (80162): ..."
        "Western Iowa Energy (4670) Facility Name: Western Iowa Energy (82630): ..."
    """
    if not applicant_desc:
        return ("", "")
    fp_match = re.search(
        r"Fuel Producer:\s*([^;()]+?)(?:\s*\(\d+\)|\s*;|\s+Facility Name)",
        applicant_desc, re.IGNORECASE
    )
    if not fp_match:
        # Fallback: first text before "Facility Name" or "("
        fp_match = re.search(r"^([^;(]+?)(?:\s*\(\d+\)|\s+Facility Name|;)", applicant_desc)
    fuel_producer = fp_match.group(1).strip() if fp_match else ""

    fn_match = re.search(
        r"Facility Name:\s*([^;()]+?)(?:\s*\(\d+\)|\s*;|:)",
        applicant_desc, re.IGNORECASE
    )
    facility_name = fn_match.group(1).strip() if fn_match else ""

    return (fuel_producer, facility_name)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    ws = wb["All Pathways"]

    # Find header row (row 4)
    header = None
    header_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == "App/Pathway #":
            header = list(row)
            header_idx = i
            break
    if header is None:
        raise SystemExit("Header row not found")
    col = {h: i for i, h in enumerate(header) if h is not None}

    pathways = []
    skipped = 0
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        ft = str(row[col["Fuel Type"]] or "")
        if not is_liquid_biofuel(ft):
            skipped += 1
            continue
        applicant = str(row[col["Applicant & Pathway Description"]] or "")
        fp, fn = parse_facility_from_description(applicant)
        rec = {
            "pathway_id": str(row[col["App/Pathway #"]] or ""),
            "class": str(row[col["Class"]] or ""),
            "calc_version": str(row[col["Calculator Version"]] or ""),
            "fuel_producer": fp,
            "facility_name": fn,
            "facility_location": str(row[col["Facility Location"]] or ""),
            "feedstock": str(row[col["Feedstock"]] or ""),
            "fuel_type": ft,
            "ci": row[col["Current Certified CI"]],
            "fpc": row[col["Current Certified  FPC"]],
            "certification_date": (
                row[col[" Certification Date"]].isoformat()
                if row[col[" Certification Date"]] and hasattr(row[col[" Certification Date"]], "isoformat")
                else str(row[col[" Certification Date"]] or "")
            ),
            "applicant_description": applicant[:500],
        }
        pathways.append(rec)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = OUT_DIR / "biofuel_pathways.json"
    json_path.write_text(json.dumps(pathways, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {json_path} ({len(pathways)} pathways, skipped {skipped})")

    csv_path = OUT_DIR / "biofuel_pathways.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=pathways[0].keys())
        w.writeheader()
        w.writerows(pathways)
    print(f"Wrote {csv_path}")

    # Summary
    from collections import Counter
    print()
    print("Fuel-type breakdown:")
    for ft, n in Counter(p["fuel_type"] for p in pathways).most_common():
        print(f"  {n:>4}  {ft}")
    print()
    print("Top 20 fuel producers (by pathway count):")
    for fp, n in Counter(p["fuel_producer"] for p in pathways).most_common(20):
        print(f"  {n:>4}  {fp}")


if __name__ == "__main__":
    main()
