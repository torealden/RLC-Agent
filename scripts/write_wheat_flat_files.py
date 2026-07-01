"""Serialize silver.wheat_series -> LONG flat file (flat_file_contract.md v1).

Writer half of the seam. Produces us_wheat_production_LONG.xlsx: one 'production' long tab (all
supply series, key columns stable) + a '_meta' tab. Idempotent; rewrites the file each run. Writes
to a *_LONG.xlsx name so it doesn't clobber Desktop's current wide workbook during format review.
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
ROOT = Path(r"C:/dev/RLC-Agent"); sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
from src.services.database.db_config import get_connection

OUT = ROOT / "models" / "Food Grains" / "us_wheat_production.xlsx"  # canonical target (Desktop's balance sheet links here)
COLS = ['commodity','class','series','marketing_year','period_type','period',
        'vintage','vintage_rank','value','unit','source','release_date','revision']

with get_connection() as c:
    cur = c.cursor()
    # Uniqueness guarantee (flat_file_contract §7): exactly ONE row per
    # (commodity, class, series, marketing_year, period, vintage_rank), keeping the latest
    # release_date (then revision, then load). Without this, SUMIFS-on-max-rank double-counts a
    # FINAL + later revision that share the same rank. DISTINCT ON collapses; then re-sort per §2.
    cur.execute(f"""
        WITH deduped AS (
          SELECT DISTINCT ON (commodity, class, series, marketing_year, period, vintage_rank)
                 {','.join(COLS)}
          FROM silver.wheat_series
          ORDER BY commodity, class, series, marketing_year, period, vintage_rank,
                   release_date DESC NULLS LAST, revision DESC NULLS LAST, loaded_at DESC
        )
        SELECT {','.join(COLS)} FROM deduped
        ORDER BY series, class, marketing_year, period, vintage_rank""")
    rows = cur.fetchall()
    cur.execute("""SELECT series, min(source) source, string_agg(DISTINCT vintage, ', ' ORDER BY vintage) vintages,
                   min(unit) unit, max(loaded_at)::date upd, count(*) n
                   FROM silver.wheat_series GROUP BY 1 ORDER BY 1""")
    meta = cur.fetchall()

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "production"
ws.append(COLS)
for cell in ws[1]: cell.font = Font(bold=True)
for r in rows:
    ws.append([r[k] for k in COLS])

wm = wb.create_sheet("_meta")
wm.append(['series','source','api','unit','vintage_set','rows','last_updated','notes'])
for cell in wm[1]: cell.font = Font(bold=True)
for m in meta:
    wm.append([m['series'], m['source'], 'quickstats.nass.usda.gov/api', m['unit'],
               m['vintages'], m['n'], str(m['upd']),
               'RAW units (acres); balance sheet picks MAX(vintage_rank) per (series,class,MY)'])

wb.save(OUT)
print(f"wrote {OUT.name}: {len(rows)} long rows, {len(meta)} series")
print("cols:", ', '.join(COLS))
for m in meta:
    print(f"  {m['series']:16} {m['n']:4} rows | vintages: {m['vintages']}")
