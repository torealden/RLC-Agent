"""
Build the per-facility profitability template — `models/templates/per_facility_profitability_v1.xlsx`.

Constructs the 10-tab template programmatically. Tabs are formula-driven where
possible so that updating Inputs-Time Series automatically recalculates
Revenue / Cost / Profit / Operating Model / Returns.

Per Iowa Crush Agent spec §12 step 5 + Phase 2 of rat-hole action plan.
Spec: docs/specs/per_facility_profitability_template_v1.md
"""

from __future__ import annotations
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = ROOT / "models" / "templates" / "per_facility_profitability_v1.xlsx"

# Styling constants
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
SUBHEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")        # yellow — user/data overrides
FORMULA_FILL = PatternFill("solid", fgColor="E2EFDA")      # green — calculations
NOTE_FILL = PatternFill("solid", fgColor="FCE4D6")         # orange — provenance
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# Feedstocks (RD-side, default mix for soy crushers — extensible)
FEEDSTOCKS = [
    ("SBO",  "Soybean Oil"),
    ("DCO",  "Distillers Corn Oil"),
    ("UCO",  "Used Cooking Oil"),
    ("TLW",  "Tallow (Inedible)"),
    ("CWG",  "Choice White Grease"),
    ("PFAT", "Poultry Fat"),
    ("CANO", "Canola Oil"),
    ("YG",   "Yellow Grease"),
]

# RD lb-per-gal yields (from Plant Model Project Feedstock Conversions tab)
RD_YIELDS = {
    "SBO": 7.5, "DCO": 9.2, "UCO": 8.01, "TLW": 9.38, "CWG": 9.375,
    "PFAT": 8.12, "CANO": 7.55, "YG": 8.5,
}
# BD lb-per-gal yields
BD_YIELDS = {
    "SBO": 7.5, "DCO": 8.2, "UCO": 8.23, "TLW": 7.75, "CWG": 7.858,
    "PFAT": 7.45, "CANO": 7.45, "YG": 8.23,
}
# Default CI scores (g CO2e/MJ) — KG anchors, override per CARB pathway
DEFAULT_CI = {
    "SBO": 70, "DCO": 35, "UCO": 18, "TLW": 28, "CWG": 30,
    "PFAT": 28, "CANO": 60, "YG": 18,
}

LCFS_BASELINE_CI_DIESEL = 88.62   # CARB diesel-pool baseline 2026 (IFVS ADR IFVS-010)
EQUIV_RATIO_RD = 1.7              # D4 RIN equivalence per gal RD
EQUIV_RATIO_BD = 1.5              # D4 RIN equivalence per gal BD

# Forecast horizon
N_FORECAST_YEARS = 10
START_DATE = date(2024, 1, 1)


def style_header(ws, row, span_cols, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def style_subheader(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ============================================================================
# TAB 1: Identity
# ============================================================================
def build_identity_tab(wb):
    ws = wb.create_sheet("Identity")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50

    style_header(ws, 1, 2, "FACILITY IDENTITY")
    style_subheader(ws, 2, 1, "Field"); style_subheader(ws, 2, 2, "Value")

    fields = [
        "facility_id", "name", "operator", "parent_company",
        "address", "city", "county", "state", "zip", "country",
        "lat", "lon",
        "title_v_permit", "state_dnr_facility_num", "epa_frs_id",
        "epa_rfs_rin_id", "eia_plant_id",
        "naics_code", "primary_oilseed",
        "nameplate_bpd", "nameplate_tpd", "nameplate_mmbu_yr",
        "operating_days_year", "commissioned_year", "last_expansion_year",
        "refining_capability", "refining_capacity",
        "biodiesel_capacity_mgy", "has_co_located_biofuel",
        "nopa_member", "status",
        "draw_radius_miles", "draw_area_class",
        "process_type",
        "data_source", "verified_at",
        "populated_at",      # timestamp at populate
        "template_version",  # v1
    ]
    for i, f in enumerate(fields, start=3):
        ws.cell(row=i, column=1, value=f).font = LABEL_FONT
        ws.cell(row=i, column=2, value="").fill = INPUT_FILL
    return ws


# ============================================================================
# TAB 2: Inputs - Static
# ============================================================================
def build_inputs_static_tab(wb):
    ws = wb.create_sheet("Inputs - Static")
    ws.column_dimensions["A"].width = 32
    for i in range(2, 12):
        ws.column_dimensions[L(i)].width = 14

    style_header(ws, 1, 9, "STATIC INPUTS — per-facility constants")

    # Section: Feedstock yields (lb / gallon RD)
    style_subheader(ws, 3, 1, "RD lb/gal yields")
    for i, (code, _) in enumerate(FEEDSTOCKS):
        ws.cell(row=3, column=i+2, value=code).font = SUBHEADER_FONT
        ws.cell(row=4, column=i+2, value=RD_YIELDS[code]).fill = INPUT_FILL
    ws.cell(row=4, column=1, value="lb_per_gal_RD").font = LABEL_FONT

    # Named ranges for yields
    for i, (code, _) in enumerate(FEEDSTOCKS):
        addr = f"'Inputs - Static'!${L(i+2)}$4"
        wb.defined_names[f"yield_rd_{code}"] = DefinedName(name=f"yield_rd_{code}", attr_text=addr)

    # BD yields
    style_subheader(ws, 6, 1, "BD lb/gal yields")
    for i, (code, _) in enumerate(FEEDSTOCKS):
        ws.cell(row=6, column=i+2, value=code).font = SUBHEADER_FONT
        ws.cell(row=7, column=i+2, value=BD_YIELDS[code]).fill = INPUT_FILL
    ws.cell(row=7, column=1, value="lb_per_gal_BD").font = LABEL_FONT
    for i, (code, _) in enumerate(FEEDSTOCKS):
        addr = f"'Inputs - Static'!${L(i+2)}$7"
        wb.defined_names[f"yield_bd_{code}"] = DefinedName(name=f"yield_bd_{code}", attr_text=addr)

    # CI scores (g CO2e/MJ)
    style_subheader(ws, 9, 1, "CI score (g CO2e/MJ)")
    for i, (code, _) in enumerate(FEEDSTOCKS):
        ws.cell(row=9, column=i+2, value=code).font = SUBHEADER_FONT
        ws.cell(row=10, column=i+2, value=DEFAULT_CI[code]).fill = INPUT_FILL
    ws.cell(row=10, column=1, value="ci_score").font = LABEL_FONT
    for i, (code, _) in enumerate(FEEDSTOCKS):
        addr = f"'Inputs - Static'!${L(i+2)}$10"
        wb.defined_names[f"ci_{code}"] = DefinedName(name=f"ci_{code}", attr_text=addr)

    # Constants
    consts = [
        ("LCFS baseline CI diesel (g/MJ)",  LCFS_BASELINE_CI_DIESEL),
        ("D4 RIN equivalence per gal RD",   EQUIV_RATIO_RD),
        ("D4 RIN equivalence per gal BD",   EQUIV_RATIO_BD),
        ("Variable OPEX (USD/gal)",         0.40),
        ("Fixed OPEX + dep (USD/gal)",      0.10),
        ("Annual operating days",            350),
        ("Throughput utilization (%)",       0.90),
        ("Discount rate for NPV",            0.10),
        ("Tax rate",                         0.25),
    ]
    style_subheader(ws, 13, 1, "Operating constants")
    for i, (lbl, val) in enumerate(consts, start=14):
        ws.cell(row=i, column=1, value=lbl).font = LABEL_FONT
        ws.cell(row=i, column=2, value=val).fill = INPUT_FILL

    # Constants named ranges
    name_map = {
        "lcfs_baseline_ci": 14, "equiv_rd": 15, "equiv_bd": 16,
        "opex_var": 17, "opex_fixed": 18, "op_days": 19, "util": 20,
        "discount_rate": 21, "tax_rate": 22,
    }
    for nm, r in name_map.items():
        wb.defined_names[nm] = DefinedName(name=nm, attr_text=f"'Inputs - Static'!$B${r}")

    # Region / pathway
    style_subheader(ws, 24, 1, "Region & pathway")
    ws.cell(row=25, column=1, value="Region").font = LABEL_FONT
    ws.cell(row=25, column=2, value="").fill = INPUT_FILL  # populated from facility state
    ws.cell(row=26, column=1, value="LCFS pathway certified?").font = LABEL_FONT
    ws.cell(row=26, column=2, value="").fill = INPUT_FILL
    ws.cell(row=27, column=1, value="45Z policy scenario").font = LABEL_FONT
    ws.cell(row=27, column=2, value="extension_2031").fill = INPUT_FILL

    return ws


# ============================================================================
# TAB 3: Inputs - Time Series
# ============================================================================
def build_inputs_timeseries_tab(wb):
    ws = wb.create_sheet("Inputs - Time Series")
    ws.column_dimensions["A"].width = 13

    style_header(ws, 1, 1 + 8 + 4 + 1 + 1 + 1 + 1 + 1, "MONTHLY INPUTS — populated from DB")

    # Headers
    cols = ["Period"]
    cols += [f"FS_{c}_$/lb" for c, _ in FEEDSTOCKS]
    cols += ["RD $/gal", "BD $/gal", "SAF $/gal", "ULSD $/gal"]
    cols += ["D4 RIN $/RIN"]
    cols += ["LCFS $/MT"]
    cols += ["45Z $/gal RD"]   # CI-dependent — populator computes
    cols += ["Freight in $/lb"]
    cols += ["Freight out $/gal"]

    for i, h in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        ws.column_dimensions[L(i)].width = 13

    # Generate monthly periods
    n_months = N_FORECAST_YEARS * 12
    for i in range(n_months):
        d = START_DATE.replace(day=1)
        # add i months
        m = (d.month - 1 + i) % 12 + 1
        y = d.year + (d.month - 1 + i) // 12
        period = date(y, m, 1)
        ws.cell(row=3+i, column=1, value=period).number_format = "yyyy-mm"
        for c in range(2, len(cols) + 1):
            ws.cell(row=3+i, column=c).fill = INPUT_FILL
            ws.cell(row=3+i, column=c).number_format = "0.0000"

    return ws, n_months


# ============================================================================
# TAB 4: Revenue Build
# ============================================================================
def build_revenue_tab(wb, n_months):
    ws = wb.create_sheet("Revenue Build")
    ws.column_dimensions["A"].width = 13

    style_header(ws, 1, 1 + len(FEEDSTOCKS), "REVENUE BUILD — Total Revenue per gallon by feedstock")
    ws.cell(row=2, column=1, value="Period").font = SUBHEADER_FONT
    ws.cell(row=2, column=1).fill = SUBHEADER_FILL

    # Per feedstock columns: each cell formula = base_rd + lcfs + d4_rin + 45z
    for i, (code, name) in enumerate(FEEDSTOCKS):
        col = i + 2
        ws.cell(row=2, column=col, value=f"{code}").font = SUBHEADER_FONT
        ws.cell(row=2, column=col).fill = SUBHEADER_FILL
        ws.column_dimensions[L(col)].width = 11

    # Formulas:
    #   base = RD price ($/gal)                                — Inputs!K row
    #   lcfs = LCFS_$/MT × (CI_baseline - CI_feedstock) / 1000 × 0.135 (energy density factor for diesel ~0.135 MT CO2 per gal at 95.6 g/MJ)
    #   d4_rin = D4_RIN_$/RIN × equiv_rd (1.7)
    #   45z = INDEX(45Z col) × 1
    # The 45Z column is in Inputs - Time Series at column N (computed during populate based on CI score)
    # Let's reference inputs by column letter:
    #   Period=A, FS_SBO=B...FS_YG=I, RD=J, BD=K, SAF=L, ULSD=M, D4=N, LCFS=O, 45Z=P, FrghtIn=Q, FrghtOut=R
    # That's the layout from build_inputs_timeseries_tab where:
    #   Period=1, FS_X=2..9, RD=10, BD=11, SAF=12, ULSD=13, D4=14, LCFS=15, 45Z=16, FrghtIn=17, FrghtOut=18

    for r in range(3, 3 + n_months):
        row_in_inputs = r  # same row layout (header at row 2)
        ws.cell(row=r, column=1, value=f"='Inputs - Time Series'!A{row_in_inputs}").number_format = "yyyy-mm"
        for i, (code, _) in enumerate(FEEDSTOCKS):
            col = i + 2
            # LCFS revenue per gallon = LCFS_$/MT × (baseline_CI - pathway_CI) × MJ/gal × 1e-6
            # CARB regulatory diesel-pool energy density = 134.47 MJ/gal → conversion
            # factor = 134.47 / 1e6 = 0.00013447. Patched 2026-05-20 (IFVS ADR IFVS-010).
            #
            # Total Revenue = base RD price ($/gal)
            #               + LCFS_$/MT × (lcfs_baseline_ci - ci_X) × 0.00013447
            #               + D4 × equiv_rd
            #               + 45Z (already $/gal in inputs)
            f = (
                f"='Inputs - Time Series'!J{row_in_inputs}"  # base RD
                f"+'Inputs - Time Series'!O{row_in_inputs}"   # LCFS $/MT
                f"*(lcfs_baseline_ci-ci_{code})*0.00013447"
                f"+'Inputs - Time Series'!N{row_in_inputs}*equiv_rd"  # D4 × equiv
                f"+'Inputs - Time Series'!P{row_in_inputs}"   # 45Z $/gal
            )
            ws.cell(row=r, column=col, value=f).number_format = "0.000"
            ws.cell(row=r, column=col).fill = FORMULA_FILL

    return ws


# ============================================================================
# TAB 5: Cost Build
# ============================================================================
def build_cost_tab(wb, n_months):
    ws = wb.create_sheet("Cost Build")
    ws.column_dimensions["A"].width = 13

    style_header(ws, 1, 1 + len(FEEDSTOCKS), "COST BUILD — Total Cost per gallon by feedstock")
    ws.cell(row=2, column=1, value="Period").font = SUBHEADER_FONT
    ws.cell(row=2, column=1).fill = SUBHEADER_FILL
    for i, (code, _) in enumerate(FEEDSTOCKS):
        col = i + 2
        ws.cell(row=2, column=col, value=code).font = SUBHEADER_FONT
        ws.cell(row=2, column=col).fill = SUBHEADER_FILL
        ws.column_dimensions[L(col)].width = 11

    # Formula: feedstock_cost($/lb × yield_lb_per_gal) + freight_in × yield + freight_out + opex_var + opex_fixed
    feedstock_input_cols = {code: i+2 for i, (code, _) in enumerate(FEEDSTOCKS)}
    for r in range(3, 3 + n_months):
        ws.cell(row=r, column=1, value=f"='Inputs - Time Series'!A{r}").number_format = "yyyy-mm"
        for i, (code, _) in enumerate(FEEDSTOCKS):
            in_col = L(feedstock_input_cols[code])
            f = (
                f"='Inputs - Time Series'!{in_col}{r}*yield_rd_{code}"   # feedstock cost
                f"+'Inputs - Time Series'!Q{r}*yield_rd_{code}"          # freight in × yield
                f"+'Inputs - Time Series'!R{r}"                          # freight out per gal
                f"+opex_var+opex_fixed"
            )
            ws.cell(row=r, column=i+2, value=f).number_format = "0.000"
            ws.cell(row=r, column=i+2).fill = FORMULA_FILL

    return ws


# ============================================================================
# TAB 6: Profit by Feedstock
# ============================================================================
def build_profit_tab(wb, n_months):
    ws = wb.create_sheet("Profit by Feedstock")
    ws.column_dimensions["A"].width = 13

    style_header(ws, 1, 1 + len(FEEDSTOCKS), "PROFIT — $/gal margin by feedstock (Revenue − Cost)")
    ws.cell(row=2, column=1, value="Period").font = SUBHEADER_FONT
    ws.cell(row=2, column=1).fill = SUBHEADER_FILL
    for i, (code, _) in enumerate(FEEDSTOCKS):
        col = i + 2
        ws.cell(row=2, column=col, value=code).font = SUBHEADER_FONT
        ws.cell(row=2, column=col).fill = SUBHEADER_FILL
        ws.column_dimensions[L(col)].width = 11

    for r in range(3, 3 + n_months):
        ws.cell(row=r, column=1, value=f"='Inputs - Time Series'!A{r}").number_format = "yyyy-mm"
        for i, (code, _) in enumerate(FEEDSTOCKS):
            f = f"='Revenue Build'!{L(i+2)}{r}-'Cost Build'!{L(i+2)}{r}"
            ws.cell(row=r, column=i+2, value=f).number_format = "0.000"
            ws.cell(row=r, column=i+2).fill = FORMULA_FILL

    return ws


# ============================================================================
# TAB 7: Operating Model
# ============================================================================
def build_operating_tab(wb, n_months):
    ws = wb.create_sheet("Operating Model")
    ws.column_dimensions["A"].width = 13
    for i in range(2, 16):
        ws.column_dimensions[L(i)].width = 13

    style_header(ws, 1, 6, "OPERATING MODEL — monthly throughput × weighted margin")

    headers = ["Period", "Throughput (gal/mo)"] + [f"Mix_{c}" for c, _ in FEEDSTOCKS] + ["Wtd $/gal", "EBITDA $"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h).font = SUBHEADER_FONT
        ws.cell(row=2, column=i).fill = SUBHEADER_FILL

    n_fs = len(FEEDSTOCKS)
    mix_start_col = 3
    wtd_col = mix_start_col + n_fs
    ebitda_col = wtd_col + 1

    for r in range(3, 3 + n_months):
        ws.cell(row=r, column=1, value=f"='Inputs - Time Series'!A{r}").number_format = "yyyy-mm"
        # Throughput: capacity from Identity tab, util, op_days/12 (gal/mo)
        # capacity_mmgy is on Identity tab — need to know which row.
        # For v1 the populator writes throughput directly (so user/script can override per-month)
        ws.cell(row=r, column=2, value="").fill = INPUT_FILL
        # Default mix — populator fills (e.g., SBO 60% / DCO 20% / UCO 20%)
        for i in range(n_fs):
            ws.cell(row=r, column=mix_start_col + i, value=0).fill = INPUT_FILL
            ws.cell(row=r, column=mix_start_col + i).number_format = "0.00%"
        # Weighted $/gal = SUMPRODUCT(Profit by Feedstock row, Mix row)
        # = SUMPRODUCT('Profit by Feedstock'!B{r}:I{r}, C{r}:J{r})
        prof_range = f"'Profit by Feedstock'!B{r}:{L(1+n_fs)}{r}"
        mix_range  = f"{L(mix_start_col)}{r}:{L(mix_start_col + n_fs - 1)}{r}"
        ws.cell(row=r, column=wtd_col,
                value=f"=SUMPRODUCT({prof_range},{mix_range})").number_format = "0.000"
        ws.cell(row=r, column=wtd_col).fill = FORMULA_FILL
        # EBITDA = throughput × weighted margin
        ws.cell(row=r, column=ebitda_col,
                value=f"=B{r}*{L(wtd_col)}{r}").number_format = "$#,##0"
        ws.cell(row=r, column=ebitda_col).fill = FORMULA_FILL

    return ws


# ============================================================================
# TAB 8: Returns Summary
# ============================================================================
def build_returns_tab(wb, n_months):
    ws = wb.create_sheet("Returns Summary")
    ws.column_dimensions["A"].width = 32
    for i in range(2, N_FORECAST_YEARS + 4):
        ws.column_dimensions[L(i)].width = 14

    style_header(ws, 1, N_FORECAST_YEARS + 2, "RETURNS SUMMARY — annual rollup + IRR/NPV")

    # Year columns
    ws.cell(row=2, column=1, value="Year").font = SUBHEADER_FONT
    for y in range(N_FORECAST_YEARS):
        col = y + 2
        year_val = START_DATE.year + y
        ws.cell(row=2, column=col, value=year_val).font = SUBHEADER_FONT
        ws.cell(row=2, column=col).fill = SUBHEADER_FILL
    ws.cell(row=2, column=N_FORECAST_YEARS + 2, value="Lifetime").font = SUBHEADER_FONT
    ws.cell(row=2, column=N_FORECAST_YEARS + 2).fill = SUBHEADER_FILL

    # Annual EBITDA — sum monthly EBITDA from Operating Model for each calendar year
    ws.cell(row=3, column=1, value="Annual EBITDA ($)").font = LABEL_FONT
    for y in range(N_FORECAST_YEARS):
        col = y + 2
        first_row = 3 + y * 12
        last_row  = 3 + y * 12 + 11
        ebitda_col_letter = L(3 + len(FEEDSTOCKS) + 2 - 1 + 1)  # last column on Operating Model
        # Operating model: throughput=B, mix=C..J, wtd=K, ebitda=L → column 12 (L)
        ws.cell(row=3, column=col,
                value=f"=SUM('Operating Model'!L{first_row}:L{last_row})").number_format = "$#,##0"
    # Lifetime
    ws.cell(row=3, column=N_FORECAST_YEARS + 2,
            value=f"=SUM('Operating Model'!L3:L{2 + n_months})").number_format = "$#,##0"

    # IRR & NPV (placeholder — needs initial capex which is on Identity)
    ws.cell(row=5, column=1, value="Initial CAPEX ($)").font = LABEL_FONT
    ws.cell(row=5, column=2, value=0).fill = INPUT_FILL
    ws.cell(row=6, column=1, value="Maintenance CAPEX ($/yr)").font = LABEL_FONT
    ws.cell(row=6, column=2, value=0).fill = INPUT_FILL

    ws.cell(row=8, column=1, value="Cash Flow ($)").font = LABEL_FONT
    ws.cell(row=8, column=2, value=f"=B3-B5-B6").number_format = "$#,##0"
    for y in range(1, N_FORECAST_YEARS):
        col = y + 2
        ws.cell(row=8, column=col, value=f"={L(col)}3-$B$6").number_format = "$#,##0"

    ws.cell(row=10, column=1, value="Cumulative CF ($)").font = LABEL_FONT
    ws.cell(row=10, column=2, value=f"=B8").number_format = "$#,##0"
    for y in range(1, N_FORECAST_YEARS):
        col = y + 2
        ws.cell(row=10, column=col, value=f"={L(col-1)}10+{L(col)}8").number_format = "$#,##0"

    ws.cell(row=12, column=1, value="IRR (Unlevered)").font = LABEL_FONT
    ws.cell(row=12, column=2,
            value=f"=IFERROR(IRR(B8:{L(N_FORECAST_YEARS + 1)}8),\"n/a\")").number_format = "0.0%"
    ws.cell(row=13, column=1, value="NPV @ discount_rate").font = LABEL_FONT
    ws.cell(row=13, column=2,
            value=f"=NPV(discount_rate, C8:{L(N_FORECAST_YEARS + 1)}8)+B8").number_format = "$#,##0"
    ws.cell(row=14, column=1, value="Payback (yrs)").font = LABEL_FONT
    ws.cell(row=14, column=2, value=f"=IFERROR(MATCH(0,B10:{L(N_FORECAST_YEARS + 1)}10,1)+1,\"n/a\")")

    return ws


# ============================================================================
# TAB 9: Sensitivity
# ============================================================================
def build_sensitivity_tab(wb):
    ws = wb.create_sheet("Sensitivity")
    ws.column_dimensions["A"].width = 32
    for i in range(2, 8): ws.column_dimensions[L(i)].width = 14

    style_header(ws, 1, 5, "SENSITIVITY — Δ EBITDA / IRR per Δ input")

    headers = ["Variable", "Base", "−20%", "−10%", "+10%", "+20%"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h).font = SUBHEADER_FONT
        ws.cell(row=2, column=i).fill = SUBHEADER_FILL

    vars_list = [
        "D4 RIN $/RIN", "LCFS $/MT", "45Z $/gal",
        "Feedstock $/lb (avg)", "RD selling $/gal",
        "Throughput (utilization)", "Variable OPEX",
    ]
    for i, v in enumerate(vars_list, start=3):
        ws.cell(row=i, column=1, value=v).font = LABEL_FONT
        for c in range(2, 7):
            ws.cell(row=i, column=c, value="").fill = INPUT_FILL  # populated by sensitivity-runner
            ws.cell(row=i, column=c).number_format = "$#,##0"

    style_subheader(ws, 12, 1, "45Z policy scenario")
    scenarios = ["extension_2031", "expiry_2027", "iluc_removed", "domestic_restriction", "none"]
    for i, s in enumerate(scenarios, start=13):
        ws.cell(row=i, column=1, value=s).font = LABEL_FONT
        ws.cell(row=i, column=2, value="").fill = INPUT_FILL  # populator fills total NPV under each scenario
        ws.cell(row=i, column=2).number_format = "$#,##0"

    return ws


# ============================================================================
# TAB 10: Notes & Provenance
# ============================================================================
def build_notes_tab(wb):
    ws = wb.create_sheet("Notes & Provenance")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    style_header(ws, 1, 2, "NOTES & PROVENANCE")
    rows = [
        ("Template version",            "v1 (2026-04-29)"),
        ("Template path",                str(TEMPLATE_PATH)),
        ("Spec",                          "docs/specs/per_facility_profitability_template_v1.md"),
        ("Source workbooks",              "D:/Switch Over/Biomass-Based Diesel/Plant Model Project/"),
        ("Populator script",              "scripts/build_per_facility_workbook.py"),
        ("Populated at",                  ""),  # filled by populator
        ("Git commit (populator)",        ""),  # filled by populator
        ("DB sources used",               ""),  # filled by populator
        ("KG-default fallback inputs",    ""),  # filled by populator
        ("Manual override log",           ""),  # if user edits cells, list them here
        ("Pending verification flags",    ""),
    ]
    for i, (lbl, val) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=lbl).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=val)
        c.fill = NOTE_FILL
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # Legend
    style_subheader(ws, 16, 1, "CELL COLOR LEGEND")
    legend = [
        ("Yellow",  "User/data input — populated from DB or manual"),
        ("Green",   "Formula / calculated"),
        ("Orange",  "Provenance / metadata"),
    ]
    for i, (color, desc) in enumerate(legend, start=17):
        ws.cell(row=i, column=1, value=color).font = LABEL_FONT
        ws.cell(row=i, column=1).fill = (
            INPUT_FILL if color == "Yellow"
            else FORMULA_FILL if color == "Green"
            else NOTE_FILL
        )
        ws.cell(row=i, column=2, value=desc)

    return ws


# ============================================================================
# Main: build template
# ============================================================================
def main():
    print(f"Building per-facility profitability template → {TEMPLATE_PATH}")
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_identity_tab(wb)
    build_inputs_static_tab(wb)
    inputs_ts_ws, n_months = build_inputs_timeseries_tab(wb)
    build_revenue_tab(wb, n_months)
    build_cost_tab(wb, n_months)
    build_profit_tab(wb, n_months)
    build_operating_tab(wb, n_months)
    build_returns_tab(wb, n_months)
    build_sensitivity_tab(wb)
    build_notes_tab(wb)

    # Reorder tabs
    desired = ["Identity", "Inputs - Static", "Inputs - Time Series",
               "Revenue Build", "Cost Build", "Profit by Feedstock",
               "Operating Model", "Returns Summary", "Sensitivity",
               "Notes & Provenance"]
    wb._sheets = [wb[name] for name in desired if name in wb.sheetnames]

    wb.save(TEMPLATE_PATH)
    print(f"Wrote {TEMPLATE_PATH}")
    print(f"  Tabs: {wb.sheetnames}")
    print(f"  Forecast horizon: {N_FORECAST_YEARS} years × 12 months = {n_months} months")
    print(f"  Feedstocks: {len(FEEDSTOCKS)} ({', '.join(c for c, _ in FEEDSTOCKS)})")


if __name__ == "__main__":
    main()
