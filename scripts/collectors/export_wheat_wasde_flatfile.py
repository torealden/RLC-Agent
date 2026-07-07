"""
Export gold.wheat_wasde_vintage_ladder to the 13-col LONG flat-file contract.

Writes us_wheat_wasde.xlsx (production + _meta tabs, matching the layout of
us_wheat_production.xlsx) so the wheat balance sheet can pick up each WASDE
release as a new vintage via MAXIFS(vintage_rank)+SUMIFS(value).

Usage:
    RLC_PG_DATABASE=rlc_commodities_sandbox python export_wheat_wasde_flatfile.py [output_path]
"""

import sys
from pathlib import Path
from datetime import date

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from src.services.database.db_config import get_connection  # noqa: E402

CONTRACT_COLUMNS = [
    "commodity", "class", "series", "marketing_year", "period_type", "period",
    "vintage", "vintage_rank", "value", "unit", "source", "release_date", "revision",
]


def fetch_rows():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT {', '.join(CONTRACT_COLUMNS)}
            FROM gold.wheat_wasde_vintage_ladder
            ORDER BY series, vintage_rank
        """)
        return cur.fetchall()


def build_meta_rows(rows):
    """One _meta row per series: source, vintage_set, rows, last_updated, notes."""
    by_series = {}
    for r in rows:
        by_series.setdefault(r["series"], []).append(r)

    meta = []
    for series in sorted(by_series):
        recs = by_series[series]
        vintages = sorted({r["vintage"] for r in recs})
        meta.append({
            "series": series,
            "source": "USDA_FAS_PSD",
            "api": "api.fas.usda.gov/api/psd",
            "unit": recs[0]["unit"],
            "vintage_set": ", ".join(vintages),
            "rows": len(recs),
            "last_updated": date.today().isoformat(),
            "notes": (
                "RAW native PSD units (no bu/acre conversion); one row per WASDE "
                "release month; balance sheet picks MAX(vintage_rank) per "
                "(series,class,MY). Vintage ladder only accumulates forward from "
                "the first month this collector ran -- PSD API has no historical "
                "revision query."
            ),
        })
    return meta


def write_workbook(rows, output_path: Path):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "production"
    ws.append(CONTRACT_COLUMNS)
    for r in rows:
        ws.append([r[c] for c in CONTRACT_COLUMNS])
    for i, col in enumerate(CONTRACT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    meta_cols = ["series", "source", "api", "unit", "vintage_set", "rows", "last_updated", "notes"]
    ws_meta = wb.create_sheet("_meta")
    ws_meta.append(meta_cols)
    for m in build_meta_rows(rows):
        ws_meta.append([m[c] for c in meta_cols])
    for i, col in enumerate(meta_cols, start=1):
        ws_meta.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    wb.save(output_path)


def main():
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("us_wheat_wasde.xlsx")
    rows = fetch_rows()
    if not rows:
        print("ERROR: no rows returned from gold.wheat_wasde_vintage_ladder -- not writing file.")
        sys.exit(1)
    write_workbook(rows, output_path)
    print(f"Wrote {len(rows)} rows to {output_path}")


if __name__ == "__main__":
    main()
