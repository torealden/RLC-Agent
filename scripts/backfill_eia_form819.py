"""
EIA Form 819 Historical Backfill
================================

Ingests historical Table 1 (capacity) and Table 2 (feedstock) xlsx files
into bronze. Two sources:

  1. Local archived files (data/raw/, anywhere on disk)
  2. Wayback Machine snapshots of eia.gov/biofuels/update/table[1|2].xlsx

Wayback Machine is needed because EIA only publishes a rolling 2-year
window at the live URL — historical data beyond ~2 years back is only
available via archived snapshots.

Usage:
  # Ingest local files
  python scripts/backfill_eia_form819.py --local <path1.xlsx> <path2.xlsx> ...
  python scripts/backfill_eia_form819.py --local-dir data/raw

  # Pull from Wayback Machine
  python scripts/backfill_eia_form819.py --wayback --from 2009 --to 2023
  python scripts/backfill_eia_form819.py --wayback --dry-run  # list snapshots only

  # Both
  python scripts/backfill_eia_form819.py --local-dir data/raw --wayback --from 2009
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / '.env')

from src.tools.eia_biofuels_collector import (
    parse_table1, parse_table2,
    save_capacity_records, save_feedstock_records,
    MONTH_NAMES,
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
)
logger = logging.getLogger('backfill_eia_form819')


CACHE_DIR = PROJECT_ROOT / 'data' / 'eia_biofuels' / 'wayback_cache'


# ---------------------------------------------------------------------------
# Old-report parser (pre-2022 EIA biodiesel report)
# ---------------------------------------------------------------------------
# These are at eia.gov/biofuels/biodiesel/production/tableN.xls and contain
# BIODIESEL-ONLY feedstock data. Form 819 superseded this report in 2022,
# but Wayback Machine has snapshots back to 2012 — stitching gives us
# 2009-2022 biodiesel feedstock history.

# Map old-report column header -> normalized feedstock name (matches new format)
OLD_REPORT_NAME_MAP = {
    'Canola oil': 'Canola Oil',
    'Corn oil': 'Corn Oil',
    'Cottonseed oil': 'Cottonseed Oil',
    'Palm oil': 'Palm Oil',
    'Soybean oil': 'Soybean Oil',
    'Other':       'Other',       # disambiguated below
    'Poultry':     'Poultry',
    'Tallow':      'Tallow',
    'White grease': 'White Grease',
    'Yellow grease': 'Yellow Grease',
    'Algae':       'Algae Oil',
    'Alcohol':     'Alcohol',
    'Catalysts':   'Catalysts',
}


def _parse_old_xls_value(val):
    """Convert a cell value from old-report .xls into (quantity, is_withheld, is_no_data)."""
    if val is None or val == '':
        return None, False, True
    if isinstance(val, str):
        v = val.strip()
        if v.upper() == 'W':
            return None, True, False
        if v in ('-', '--'):
            return None, False, True
        if v == '(s)':
            # "(s)" = less than 0.5 — treat as 0.25
            return 0.25, False, False
        try:
            return float(v.replace(',', '')), False, False
        except ValueError:
            return None, False, True
    try:
        return float(val), False, False
    except (ValueError, TypeError):
        return None, False, True


def parse_old_report_table3(path):
    """
    Parse old EIA biodiesel report Table 3 + Table 3a (BIODIESEL-ONLY feedstock).
    Returns records compatible with bronze.eia_feedstock_monthly with
    plant_type='biodiesel' and source_sheet='old_table3' or 'old_t3a'.
    """
    import xlrd
    wb = xlrd.open_workbook(path)
    records = []

    for sheet_name in wb.sheet_names():
        sh = wb.sheet_by_name(sheet_name)
        if sheet_name not in ('Table 3', 'Table 3a'):
            continue
        sheet_tag = 'old_table3' if sheet_name == 'Table 3' else 'old_t3a'

        # Headers — feedstock names live on row 5 (index 4)
        col_to_name = {}
        for c in range(sh.ncols):
            v = sh.cell_value(4, c)
            if not isinstance(v, str) or not v.strip() or v.strip() == 'Period':
                continue
            name = OLD_REPORT_NAME_MAP.get(v.strip(), v.strip())
            # Disambiguate "Other" by context (row 4 has section labels)
            if name == 'Other':
                section = sh.cell_value(3, c) or sh.cell_value(2, c)
                if isinstance(section, str):
                    s = section.strip().lower()
                    if 'animal' in s: name = 'Other Animal Fat'
                    elif 'recycled' in s: name = 'Other Recycled'
                    elif 'vegetable' in s: name = 'Other Vegetable Oil'
                    elif 'other inputs' in s: name = 'Other Inputs'
            col_to_name[c] = name

        # Iterate rows starting row 7 (index 6)
        current_year = None
        for r in range(6, sh.nrows):
            a = sh.cell_value(r, 0)
            try:
                af = float(a)
                if 2000 <= af <= 2100:
                    current_year = int(af)
                    continue
            except (ValueError, TypeError):
                pass
            if not isinstance(a, str):
                continue
            month_name = a.strip()
            if month_name not in MONTH_NAMES or not current_year:
                continue
            month = MONTH_NAMES[month_name]

            for c, fname in col_to_name.items():
                qty, is_withheld, is_no_data = _parse_old_xls_value(sh.cell_value(r, c))
                records.append({
                    'year': current_year,
                    'month': month,
                    'source_sheet': sheet_tag,
                    'feedstock_name': fname,
                    'plant_type': 'biodiesel',
                    'quantity_mil_lbs': qty,
                    'is_withheld': is_withheld,
                    'is_no_data': is_no_data,
                })
    return records


# ---------------------------------------------------------------------------
# File-type detection
# ---------------------------------------------------------------------------

def detect_xlsx_kind(path: str) -> str:
    """Return 'table1' (capacity), 'table2' (feedstock),
    'old_table3' (pre-2022 biodiesel-only feedstock), or 'unknown'."""
    # .xls = old report (xlrd)
    if path.lower().endswith('.xls'):
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            for sn in wb.sheet_names():
                if sn in ('Table 3', 'Table 3a'):
                    return 'old_table3'
                # Old Table 1 = capacity
                if sn == 'Table 1':
                    sh = wb.sheet_by_name(sn)
                    row3 = [sh.cell_value(2, c) for c in range(sh.ncols)]
                    if any(isinstance(v, str) and 'Production Capacity' in v for v in row3):
                        return 'old_table1'
            return 'unknown'
        except Exception as e:
            logger.warning(f"  {path}: cannot open .xls — {e}")
            return 'unknown'

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets = set(wb.sheetnames)
        wb.close()
    except Exception as e:
        logger.warning(f"  {path}: cannot open — {e}")
        return 'unknown'

    table2_sheets = {'table_2a', 'table_2b', 'table_2c', 'table_2d'}
    if table2_sheets & sheets:
        return 'table2'

    # Table 1 is single-sheet with a Period column. Heuristic: any sheet
    # whose A3 = 'Period' and which has 'Biodiesel' / 'Fuel Ethanol' in row 3.
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        for sn in wb.sheetnames:
            ws = wb[sn]
            row3 = [c.value for c in ws[3]]
            if any(isinstance(v, str) and 'Biodiesel' in v for v in row3):
                wb.close()
                return 'table1'
        wb.close()
    except Exception:
        pass
    return 'unknown'


def ingest_xlsx(path: str) -> dict:
    """Detect type, parse, save. Returns counts dict."""
    kind = detect_xlsx_kind(path)
    fname = os.path.basename(path)
    if kind == 'table1':
        records = parse_table1(path)
        if not records:
            return {'kind': 'table1', 'parsed': 0, 'saved': 0}
        ins, upd, err = save_capacity_records(records, fname)
        logger.info(f"  {fname} [table1]: {len(records)} parsed -> ins={ins} upd={upd} err={err}")
        return {'kind': 'table1', 'parsed': len(records), 'saved': ins + upd, 'errors': err}
    elif kind == 'table2':
        records = parse_table2(path)
        if not records:
            return {'kind': 'table2', 'parsed': 0, 'saved': 0}
        ins, upd, err = save_feedstock_records(records, fname)
        logger.info(f"  {fname} [table2]: {len(records)} parsed -> ins={ins} upd={upd} err={err}")
        return {'kind': 'table2', 'parsed': len(records), 'saved': ins + upd, 'errors': err}
    elif kind == 'old_table3':
        records = parse_old_report_table3(path)
        if not records:
            return {'kind': 'old_table3', 'parsed': 0, 'saved': 0}
        ins, upd, err = save_feedstock_records(records, fname)
        logger.info(f"  {fname} [old_table3]: {len(records)} parsed -> ins={ins} upd={upd} err={err}")
        return {'kind': 'old_table3', 'parsed': len(records), 'saved': ins + upd, 'errors': err}
    elif kind == 'old_table1':
        logger.info(f"  {fname} [old_table1]: skipped (old capacity format, biodiesel-only)")
        return {'kind': 'old_table1', 'parsed': 0, 'saved': 0}
    else:
        logger.warning(f"  {fname}: unknown xlsx format, skipping")
        return {'kind': 'unknown', 'parsed': 0, 'saved': 0}


# ---------------------------------------------------------------------------
# Wayback Machine helpers
# ---------------------------------------------------------------------------

WAYBACK_CDX = "https://web.archive.org/cdx/search/cdx"
WAYBACK_WEB = "https://web.archive.org/web"

# Both current (post-2022 Form 819) and old (pre-2022 biodiesel-only) URLs.
# Wayback Machine has snapshots of the old URL back to Sep 2012, which
# stitched together gives us 2009-2022 biodiesel feedstock history.
EIA_URLS = [
    # Current Form 819 (2022+) — feedstock by type, BD + RD + ethanol + SAF
    "https://www.eia.gov/biofuels/update/table1.xlsx",
    "https://www.eia.gov/biofuels/update/table2.xlsx",
    "http://www.eia.gov/biofuels/update/table1.xlsx",
    "http://www.eia.gov/biofuels/update/table2.xlsx",
    # Old biodiesel-only report (deprecated 2022) — biodiesel feedstock
    # Table 3 + Table 3a have the feedstock-by-type breakdown
    "https://www.eia.gov/biofuels/biodiesel/production/table3.xls",
    "http://www.eia.gov/biofuels/biodiesel/production/table3.xls",
]


def query_wayback_snapshots(url: str, year_from: int, year_to: int) -> list:
    """Query Wayback CDX for available snapshots of a URL."""
    params = {
        'url': url,
        'output': 'json',
        'from': f'{year_from}0101',
        'to': f'{year_to}1231',
        'filter': 'statuscode:200',
        'collapse': 'timestamp:6',  # one snapshot per month
    }
    try:
        r = requests.get(WAYBACK_CDX, params=params, timeout=60)
        r.raise_for_status()
        rows = r.json()
        if not rows or len(rows) < 2:
            return []
        # First row is header
        header = rows[0]
        return [dict(zip(header, row)) for row in rows[1:]]
    except Exception as e:
        logger.warning(f"  Wayback CDX query failed for {url}: {e}")
        return []


def download_wayback_snapshot(timestamp: str, url: str, cache_dir: Path) -> str:
    """Download one Wayback snapshot to local cache; return path."""
    # Use 'if_' to get raw content without Wayback's HTML wrapper
    wayback_url = f"{WAYBACK_WEB}/{timestamp}if_/{url}"
    # Filename: extract table1/table2/table3 from URL + use correct extension
    if 'table1' in url: table = 'table1'
    elif 'table2' in url: table = 'table2'
    elif 'table3' in url: table = 'oldtable3'
    else: table = 'unknown'
    ext = 'xls' if url.lower().endswith('.xls') else 'xlsx'
    cache_dir.mkdir(parents=True, exist_ok=True)
    fname = f"{table}_{timestamp}.{ext}"
    dest = cache_dir / fname
    if dest.exists() and dest.stat().st_size > 1000:
        return str(dest)
    try:
        r = requests.get(wayback_url, timeout=60,
                         headers={'User-Agent': 'Mozilla/5.0 (RLC-Agent Backfill)'})
        r.raise_for_status()
        if len(r.content) < 1000:
            logger.warning(f"  {timestamp}: file too small ({len(r.content)} bytes), likely a 404")
            return ''
        dest.write_bytes(r.content)
        return str(dest)
    except Exception as e:
        logger.warning(f"  Download failed for {timestamp}: {e}")
        return ''


def run_wayback_backfill(year_from: int, year_to: int, dry_run: bool = False,
                         delay: float = 1.5) -> dict:
    """Query Wayback for all EIA Form 819 snapshots and ingest them."""
    all_snapshots = []
    for url in EIA_URLS:
        logger.info(f"Querying Wayback CDX: {url}")
        snaps = query_wayback_snapshots(url, year_from, year_to)
        logger.info(f"  Found {len(snaps)} monthly snapshots")
        for s in snaps:
            s['_source_url'] = url
        all_snapshots.extend(snaps)

    # Dedupe by (year-month, table-kind) — http vs https may overlap
    def table_kind_of(u):
        if 'table1' in u: return 'table1'
        if 'table2' in u: return 'table2'
        if 'table3' in u: return 'old_table3'
        return 'unknown'
    seen = set()
    unique = []
    for s in all_snapshots:
        key = (s.get('timestamp', '')[:6], table_kind_of(s['_source_url']))
        if key not in seen:
            seen.add(key)
            unique.append(s)
    all_snapshots = sorted(unique, key=lambda x: x.get('timestamp', ''))
    logger.info(f"Total unique snapshots (monthly bucket × table): {len(all_snapshots)}")

    if dry_run:
        logger.info("DRY RUN — listing snapshots only:")
        for s in all_snapshots[:20]:
            logger.info(f"  {s['timestamp']} {s['_source_url']}")
        if len(all_snapshots) > 20:
            logger.info(f"  ... and {len(all_snapshots) - 20} more")
        return {'snapshots': len(all_snapshots), 'ingested': 0}

    totals = {'table1_saved': 0, 'table2_saved': 0, 'errors': 0, 'files': 0}
    for i, s in enumerate(all_snapshots, 1):
        ts = s.get('timestamp', '')
        url = s.get('_source_url', '')
        logger.info(f"[{i}/{len(all_snapshots)}] {ts} {url}")

        path = download_wayback_snapshot(ts, url, CACHE_DIR)
        if not path:
            totals['errors'] += 1
            time.sleep(delay)
            continue

        try:
            result = ingest_xlsx(path)
            totals['files'] += 1
            if result['kind'] == 'table1':
                totals['table1_saved'] += result['saved']
            elif result['kind'] == 'table2':
                totals['table2_saved'] += result['saved']
        except Exception as e:
            logger.error(f"  Ingest failed: {e}")
            totals['errors'] += 1

        time.sleep(delay)

    return totals


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='EIA Form 819 historical backfill')
    parser.add_argument('--local', nargs='+', help='Specific xlsx files to ingest')
    parser.add_argument('--local-dir', help='Directory to scan for xlsx files')
    parser.add_argument('--wayback', action='store_true', help='Pull from Wayback Machine')
    parser.add_argument('--from', dest='year_from', type=int, default=2009,
                        help='Wayback: earliest year (default 2009)')
    parser.add_argument('--to', dest='year_to', type=int,
                        default=datetime.now().year,
                        help='Wayback: latest year (default current year)')
    parser.add_argument('--dry-run', action='store_true',
                        help='Wayback: list snapshots without downloading')
    parser.add_argument('--delay', type=float, default=1.5,
                        help='Seconds between Wayback requests (default 1.5)')
    args = parser.parse_args()

    if not (args.local or args.local_dir or args.wayback):
        parser.error("Specify at least one of --local, --local-dir, --wayback")

    # --- Local files ---
    local_files = []
    if args.local:
        local_files.extend(args.local)
    if args.local_dir:
        for root, _, files in os.walk(args.local_dir):
            for f in files:
                if f.lower().endswith('.xlsx') and not f.startswith('~$'):
                    local_files.append(os.path.join(root, f))

    if local_files:
        logger.info(f"=== LOCAL: {len(local_files)} files ===")
        local_totals = {'parsed': 0, 'saved': 0, 'errors': 0}
        for fp in local_files:
            try:
                result = ingest_xlsx(fp)
                local_totals['parsed'] += result.get('parsed', 0)
                local_totals['saved'] += result.get('saved', 0)
                local_totals['errors'] += result.get('errors', 0)
            except Exception as e:
                logger.error(f"  {fp}: {e}")
                local_totals['errors'] += 1
        logger.info(f"Local totals: {local_totals}")

    # --- Wayback ---
    if args.wayback:
        logger.info(f"=== WAYBACK: {args.year_from} → {args.year_to} ===")
        wb_totals = run_wayback_backfill(args.year_from, args.year_to,
                                          dry_run=args.dry_run,
                                          delay=args.delay)
        logger.info(f"Wayback totals: {wb_totals}")


if __name__ == '__main__':
    main()
