#!/usr/bin/env python3
"""
Weekly Export Inspections Data Collection Script

Downloads FGIS export inspection data, aggregates by week and month,
and updates the US Soybean Trade Excel file and RLC_commodities database.

=============================================================================
PROCESS DOCUMENTATION (from video transcript)
=============================================================================

STEP-BY-STEP PROCESS:

1. CHECK CORRECTION DATE (CRITICAL FIRST STEP)
   - Download wa_gr101.txt from https://www.ams.usda.gov/mnreports/wa_gr101.txt
   - Look at the bottom of the report for the correction date for your commodity
   - Example: "SOYBEANS 12-9-2025" means corrections were made through that date
   - Must repull data from correction date forward to capture all corrections
   - Without this step, you will have incorrect historical data!

2. DOWNLOAD CSV FILES
   - URL: https://fgisonline.ams.usda.gov/ExportGrainReport/CY{year}.csv
   - Download current year file (e.g., CY2025.csv)
   - If correction date crosses year boundary, also download previous year
   - Example: In early 2026, corrections may affect 2025 data

3. CREATE PIVOT TABLE (conceptually - we do this in code)
   - Columns: Thursday (week ending date)
   - Filters: Cert Date, Grain (select your commodity)
   - Rows: Destination (countries)
   - Values: Sum of Metric Tons

4. CONVERT TO THOUSAND BUSHELS
   - Formula: =IF(cell="","",cell*36.7437/1000)
   - 36.7437 is conversion factor (MT to bushels for soybeans)
   - Leave blanks as blanks (don't store zeros - takes up space)

5. WEEKLY TOTALS
   - Sum all destinations for each week
   - Paste into "Weekly Export Inspections" tab
   - Countries organized by region in the spreadsheet

6. MONTHLY TOTALS BY DESTINATION
   - Select all cert dates for each month
   - Sum by destination country
   - Paste into "Monthly Export Inspections" tab

7. SPECIAL HANDLING: MEXICO
   - Mexico goes to row 300 (not with other countries at ~266)
   - Reason: Truck shipments may not be inspected
   - Census Bureau data later shows actual total
   - Track difference to estimate unrecorded shipments

AG THEORY NOTES:
- Inspections = first look at actual export shipments
- Weekly Export Sales report (Thursday) shows commitments vs shipments
- Census Bureau data (monthly) is the final reconciliation data
- Inspections help adjust forecast vs actuals before Census data arrives
- Marketing year for soybeans starts September 1

=============================================================================

Data Sources:
- Weekly text report: https://www.ams.usda.gov/mnreports/wa_gr101.txt
- CSV files: https://fgisonline.ams.usda.gov/ExportGrainReport/CY{year}.csv

Usage:
    python scripts/pull_weekly_inspections.py
    python scripts/pull_weekly_inspections.py --commodity SOYBEANS
    python scripts/pull_weekly_inspections.py --year 2024 --year 2025
    python scripts/pull_weekly_inspections.py --update-excel
    python scripts/pull_weekly_inspections.py --save-to-db
"""

import argparse
import csv
import logging
import os
import sys
from datetime import datetime, date, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv is optional

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# =============================================================================
# CONSTANTS
# =============================================================================

# FGIS Data URLs
FGIS_CSV_BASE_URL = "https://fgisonline.ams.usda.gov/ExportGrainReport/"
WEEKLY_REPORT_URL = "https://www.ams.usda.gov/mnreports/wa_gr101.txt"

# Bushel weights by commodity (lbs per bushel)
BUSHEL_WEIGHTS = {
    'SOYBEANS': 60.0,
    'CORN': 56.0,
    'WHEAT': 60.0,
    'SORGHUM': 56.0,
    'BARLEY': 48.0,
    'OATS': 32.0,
    'RYE': 56.0,
    'FLAXSEED': 56.0,
    'SUNFLOWER': 28.0,
}

# Conversion factor: Metric Tons to Bushels
# This is used in the pivot table conversion: MT * 36.7437 / 1000 = thousand bushels
# For soybeans: 1 MT = 36.7437 bushels (2204.62 lbs / 60 lbs per bushel)
MT_TO_BUSHELS = {
    'SOYBEANS': 36.7437,  # 2204.62 / 60
    'CORN': 39.3680,      # 2204.62 / 56
    'WHEAT': 36.7437,     # 2204.62 / 60
    'SORGHUM': 39.3680,   # 2204.62 / 56
    'BARLEY': 45.9296,    # 2204.62 / 48
    'OATS': 68.8944,      # 2204.62 / 32
}

# Marketing year start months
MARKETING_YEAR_START = {
    'SOYBEANS': 9,  # September
    'CORN': 9,
    'WHEAT': 6,  # June
    'SORGHUM': 9,
    'BARLEY': 6,
    'OATS': 6,
}

# Excel file paths (relative to project root)
EXCEL_FILES = {
    'SOYBEANS': 'Models/Oilseeds/US Soybean Trade.xlsx',
    'CORN': 'Models/Feed Grains/US Feed Grain Exports.xlsx',
    'WHEAT': 'Models/Food Grains/US Wheat Trade.xlsx',  # May need to create
}

# Excel sheet names
WEEKLY_SHEET = 'Weekly Export Inspections'
MONTHLY_SHEET = 'Monthly Export Inspections'

# Special handling for Mexico
# Mexico data goes to a different row because truck shipments may not be inspected
# The difference between inspections and Census Bureau data is tracked separately
MEXICO_SPECIAL_HANDLING = {
    'SOYBEANS': {
        'weekly_row': 300,  # Instead of normal position (~266)
        'reason': 'Truck shipments to Mexico may not be inspected'
    }
}

# Countries that need special handling due to inspection vs Census differences
BORDER_COUNTRIES = ['MEXICO', 'CANADA']


# =============================================================================
# DATA DOWNLOAD
# =============================================================================

def create_session() -> requests.Session:
    """Create requests session with retry configuration"""
    session = requests.Session()

    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )

    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("https://", adapter)
    session.mount("http://", adapter)

    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                     '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/csv,text/plain,application/csv,*/*',
    })

    return session


def download_csv(year: int, data_dir: Path, force: bool = False) -> Optional[Path]:
    """
    Download FGIS CSV file for a specific year

    Args:
        year: Calendar year (e.g., 2025)
        data_dir: Directory to save files
        force: Force re-download even if file exists

    Returns:
        Path to downloaded file, or None if failed
    """
    data_dir.mkdir(parents=True, exist_ok=True)
    local_path = data_dir / f"CY{year}.csv"

    if local_path.exists() and not force:
        logger.info(f"File already exists: {local_path}")
        return local_path

    url = f"{FGIS_CSV_BASE_URL}CY{year}.csv"
    logger.info(f"Downloading {url}...")

    session = create_session()

    try:
        response = session.get(url, timeout=300)
        response.raise_for_status()

        local_path.write_bytes(response.content)
        logger.info(f"Downloaded {local_path} ({len(response.content):,} bytes)")
        return local_path

    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to download {url}: {e}")
        return None


def download_weekly_report(data_dir: Path) -> Optional[Path]:
    """
    Download the weekly text report (wa_gr101.txt)

    This report has the latest week's summary data AND the correction dates.
    CRITICAL: Must check correction dates to know how far back to repull data.
    """
    data_dir.mkdir(parents=True, exist_ok=True)
    local_path = data_dir / "wa_gr101.txt"

    logger.info(f"Downloading {WEEKLY_REPORT_URL}...")

    session = create_session()

    try:
        response = session.get(WEEKLY_REPORT_URL, timeout=60)
        response.raise_for_status()

        local_path.write_bytes(response.content)
        logger.info(f"Downloaded weekly report ({len(response.content):,} bytes)")
        return local_path

    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to download weekly report: {e}")
        return None


def parse_correction_date(report_path: Path, commodity: str) -> Optional[date]:
    """
    Parse the wa_gr101.txt report to find the latest correction date for a commodity.

    CRITICAL STEP: The correction date tells us how far back we need to repull data
    to capture any corrections that were made to historical records.

    The report format has correction dates at the bottom, like:
    "SOYBEANS 12-9-2025" meaning corrections were made through that date.

    Args:
        report_path: Path to wa_gr101.txt
        commodity: Commodity name (SOYBEANS, CORN, etc.)

    Returns:
        Date of latest correction, or None if not found
    """
    if not report_path or not report_path.exists():
        logger.warning("Weekly report not available for correction date check")
        return None

    try:
        content = report_path.read_text(encoding='latin-1')

        # Look for pattern like "SOYBEANS    12-9-2025" or similar
        import re

        # The correction dates appear at the bottom of the report
        # Format varies but generally: COMMODITY  MM-DD-YYYY or M-D-YYYY
        pattern = rf'{commodity.upper()}\s+(\d{{1,2}}-\d{{1,2}}-\d{{4}})'
        matches = re.findall(pattern, content, re.IGNORECASE)

        if matches:
            # Take the last match (most recent)
            date_str = matches[-1]
            correction_date = datetime.strptime(date_str, '%m-%d-%Y').date()
            logger.info(f"Found correction date for {commodity}: {correction_date}")
            return correction_date

        # Try alternative patterns
        pattern2 = rf'{commodity.upper()}\s+(\d{{1,2}}/\d{{1,2}}/\d{{4}})'
        matches2 = re.findall(pattern2, content, re.IGNORECASE)

        if matches2:
            date_str = matches2[-1]
            correction_date = datetime.strptime(date_str, '%m/%d/%Y').date()
            logger.info(f"Found correction date for {commodity}: {correction_date}")
            return correction_date

        logger.warning(f"Could not find correction date for {commodity} in report")
        return None

    except Exception as e:
        logger.error(f"Error parsing correction date: {e}")
        return None


def get_required_date_range(correction_date: Optional[date], last_update_date: Optional[date] = None) -> Tuple[date, date]:
    """
    Determine the date range to pull based on correction date and last update.

    Logic:
    - If correction_date exists, start from that date (to capture corrections)
    - Otherwise, start from last_update_date
    - If neither, default to beginning of current marketing year

    Args:
        correction_date: Latest correction date from wa_gr101.txt
        last_update_date: Date of our last update (from database or tracking)

    Returns:
        Tuple of (start_date, end_date) for data pull
    """
    today = date.today()
    end_date = today

    if correction_date:
        # Start from correction date to capture all corrections
        start_date = correction_date
        logger.info(f"Using correction date as start: {start_date}")
    elif last_update_date:
        # Start from last update
        start_date = last_update_date
        logger.info(f"Using last update date as start: {start_date}")
    else:
        # Default to beginning of current marketing year (September 1)
        if today.month >= 9:
            start_date = date(today.year, 9, 1)
        else:
            start_date = date(today.year - 1, 9, 1)
        logger.info(f"Using marketing year start as default: {start_date}")

    return start_date, end_date


# =============================================================================
# CSV PARSING
# =============================================================================

def parse_date(value: str) -> Optional[date]:
    """Parse date from various formats"""
    if not value or not value.strip():
        return None

    formats = ['%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y', '%m/%d/%y']

    for fmt in formats:
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue

    return None


def parse_number(value: str) -> Optional[float]:
    """Parse numeric value, handling commas"""
    if not value or not value.strip():
        return None

    try:
        return float(value.replace(',', '').strip())
    except (ValueError, TypeError):
        return None


def get_marketing_year(commodity: str, record_date: date) -> str:
    """
    Calculate marketing year for a given date and commodity

    Returns format like '24/25' for 2024/25 marketing year
    """
    start_month = MARKETING_YEAR_START.get(commodity.upper(), 9)

    if record_date.month >= start_month:
        my_start = record_date.year
    else:
        my_start = record_date.year - 1

    my_end = my_start + 1
    return f"{str(my_start)[-2:]}/{str(my_end)[-2:]}"


def parse_csv_file(file_path: Path, commodity_filter: str = None,
                   start_date: date = None, end_date: date = None) -> List[Dict]:
    """
    Parse FGIS CSV file and return list of records

    The CSV has ~112 columns including quality metrics like moisture, protein, oil, etc.
    We parse all valuable fields for analysis.

    Args:
        file_path: Path to CSV file
        commodity_filter: Optional filter for specific commodity
        start_date: Optional filter - only include records on/after this date
        end_date: Optional filter - only include records on/before this date

    Returns:
        List of parsed records with quantities and quality metrics
    """
    records = []

    encodings = ['utf-8', 'latin-1', 'cp1252']

    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.DictReader(f)

                for row in reader:
                    # Get grain type
                    grain = row.get('Grain', '').strip().upper()

                    # Apply commodity filter if specified
                    if commodity_filter and grain != commodity_filter.upper():
                        continue

                    # Parse week ending date
                    week_ending = parse_date(row.get('Thursday', ''))
                    if not week_ending:
                        continue

                    # Apply date filters if specified (for efficiency)
                    if start_date and week_ending < start_date:
                        continue
                    if end_date and week_ending > end_date:
                        continue

                    # Parse quantities
                    pounds = parse_number(row.get('Pounds', ''))
                    metric_tons = parse_number(row.get('Metric Ton', ''))

                    if not pounds and not metric_tons:
                        continue

                    # Calculate bushels using proper conversion
                    # From transcript: MT * 36.7437 / 1000 = thousand bushels
                    conversion_factor = MT_TO_BUSHELS.get(grain, 36.7437)
                    if metric_tons:
                        thousand_bushels = metric_tons * conversion_factor / 1000
                    elif pounds:
                        bushel_weight = BUSHEL_WEIGHTS.get(grain, 60.0)
                        bushels = pounds / bushel_weight
                        thousand_bushels = bushels / 1000
                    else:
                        thousand_bushels = None

                    # Get destination and location info
                    destination = row.get('Destination', '').strip()

                    # Calculate marketing year
                    marketing_year = get_marketing_year(grain, week_ending)

                    # Parse cert date (actual inspection date)
                    cert_date = parse_date(row.get('Cert Date', ''))

                    # Build base record
                    record = {
                        'week_ending': week_ending,
                        'cert_date': cert_date,
                        'grain': grain,
                        'destination': destination,
                        'pounds': pounds,
                        'metric_tons': metric_tons,
                        'thousand_bushels': thousand_bushels,
                        'marketing_year': marketing_year,
                        'month': week_ending.replace(day=1),
                        'port': row.get('Port', '').strip(),
                        'grade': row.get('Grade', '').strip(),
                        'commodity_class': row.get('Class', '').strip(),

                        # Quality Metrics - these are valuable for analysis
                        # Moisture (critical for soybeans - affects pricing/grading)
                        'moisture_avg': parse_number(row.get('M AVG', '')),
                        'moisture_high': parse_number(row.get('M HIGH', '')),
                        'moisture_low': parse_number(row.get('M LOW', '')),

                        # Test Weight (density indicator)
                        'test_weight': parse_number(row.get('TW', '')),

                        # Protein (important for meal value)
                        'protein_avg': parse_number(row.get('PRO AVG', '')),
                        'protein_high': parse_number(row.get('PRO HIGH', '')),
                        'protein_low': parse_number(row.get('PRO LOW', '')),

                        # Oil content (important for oil value)
                        'oil_avg': parse_number(row.get('OIL AVG', '')),
                        'oil_high': parse_number(row.get('OIL HIGH', '')),
                        'oil_low': parse_number(row.get('OIL LOW', '')),

                        # Damage metrics
                        'total_damage_avg': parse_number(row.get('DM AVG', '')),
                        'heat_damage_avg': parse_number(row.get('HD AVG', '')),
                        'foreign_material_avg': parse_number(row.get('FM AVG', '')),

                        # Splits (for soybeans)
                        'splits_avg': parse_number(row.get('SPL AVG', '')),

                        # Dockage
                        'dockage_avg': parse_number(row.get('DKG AVG', '')),
                    }

                    records.append(record)

            logger.info(f"Parsed {len(records)} records from {file_path.name}")
            if start_date or end_date:
                logger.info(f"  Date filter: {start_date} to {end_date}")
            return records

        except UnicodeDecodeError:
            continue

    logger.error(f"Failed to parse {file_path} with any encoding")
    return []


# =============================================================================
# DATA AGGREGATION
# =============================================================================

def aggregate_weekly_totals(records: List[Dict], commodity: str) -> Dict[date, float]:
    """
    Aggregate records to weekly totals for a commodity

    Returns:
        Dict mapping week_ending date to thousand bushels
    """
    weekly = defaultdict(float)

    for r in records:
        if r['grain'].upper() == commodity.upper():
            if r['thousand_bushels']:
                weekly[r['week_ending']] += r['thousand_bushels']

    return dict(weekly)


def aggregate_weekly_by_destination(records: List[Dict], commodity: str) -> Dict[date, Dict[str, float]]:
    """
    Aggregate records to weekly totals by destination for Excel update

    Returns:
        Dict mapping week_ending date to Dict[destination, thousand_bushels]
    """
    weekly = defaultdict(lambda: defaultdict(float))

    for r in records:
        if r['grain'].upper() == commodity.upper():
            if r['thousand_bushels']:
                weekly[r['week_ending']][r['destination']] += r['thousand_bushels']

    # Convert nested defaultdict to regular dict
    return {week: dict(dests) for week, dests in weekly.items()}


def aggregate_monthly_by_destination(records: List[Dict], commodity: str) -> Dict[Tuple[date, str], float]:
    """
    Aggregate records to monthly totals by destination country

    Returns:
        Dict mapping (month, destination) to thousand bushels
    """
    monthly = defaultdict(float)

    for r in records:
        if r['grain'].upper() == commodity.upper():
            if r['thousand_bushels']:
                key = (r['month'], r['destination'])
                monthly[key] += r['thousand_bushels']

    return dict(monthly)


def aggregate_marketing_year_by_destination(records: List[Dict], commodity: str) -> Dict[Tuple[str, str], float]:
    """
    Aggregate records to marketing year totals by destination country

    Returns:
        Dict mapping (marketing_year, destination) to million bushels
    """
    yearly = defaultdict(float)

    for r in records:
        if r['grain'].upper() == commodity.upper():
            if r['thousand_bushels']:
                key = (r['marketing_year'], r['destination'])
                # Convert to million bushels
                yearly[key] += r['thousand_bushels'] / 1000

    return dict(yearly)


# =============================================================================
# EXCEL UPDATE
# =============================================================================

# Destination name normalization for matching FGIS data to Excel rows
# FGIS uses different naming conventions than the Excel file
DESTINATION_NAME_MAP = {
    # FGIS name -> Excel name (or None if same)
    'GERMANY': 'Germany, Federal Republic of',
    'UNITED KINGDOM': 'United Kingdom',
    'UK': 'United Kingdom',
    'NETHERLANDS': 'Netherlands',
    'KOREA, REPUBLIC OF': 'South Korea',
    'KOREA, SOUTH': 'South Korea',
    'KOREA REP OF': 'South Korea',
    'TAIWAN': 'Taiwan',
    'CHINA': 'China',
    'CHINA, PEOPLES REPUBLIC OF': 'China',
    'JAPAN': 'Japan',
    'INDONESIA': 'Indonesia',
    'THAILAND': 'Thailand',
    'VIETNAM': 'Vietnam',
    'VIET NAM': 'Vietnam',
    'PHILIPPINES': 'Philippines',
    'EGYPT': 'Egypt',
    'SPAIN': 'Spain',
    'PORTUGAL': 'Portugal',
    'ITALY': 'Italy',
    'FRANCE': 'France',
    'BELGIUM': 'Belgium',
    'TURKEY': 'Turkey',
    'RUSSIA': 'Russia',
    'RUSSIAN FEDERATION': 'Russia',
    'MEXICO': 'Mexico',
    'CANADA': 'Canada',
    'BRAZIL': 'Brazil',
    'ARGENTINA': 'Argentina',
    'COLOMBIA': 'Colombia',
    'VENEZUELA': 'Venezuela',
    'PERU': 'Peru',
    'CHILE': 'Chile',
    'PAKISTAN': 'Pakistan',
    'BANGLADESH': 'Bangladesh',
    'MALAYSIA': 'Malaysia',
    'SINGAPORE': 'Singapore',
    'MYANMAR': 'Burma (Myanmar)',
    'BURMA': 'Burma (Myanmar)',
    'INDIA': 'India',
    'SRI LANKA': 'Sri Lanka',
    'ALGERIA': 'Algeria',
    'MOROCCO': 'Morocco',
    'TUNISIA': 'Tunisia',
    'SOUTH AFRICA': 'South Africa',
    'NIGERIA': 'Nigeria',
    'IRAN': 'Iran',
    'SAUDI ARABIA': 'Saudi Arabia',
    'UNITED ARAB EMIRATES': 'United Arab Emirates',
    'UAE': 'United Arab Emirates',
    'ISRAEL': 'Israel',
    'JORDAN': 'Jordan',
    'HONG KONG': 'Hong Kong',
}


def normalize_destination(dest: str) -> str:
    """
    Normalize destination name to match Excel file conventions

    Args:
        dest: Destination name from FGIS data

    Returns:
        Normalized destination name for Excel matching
    """
    if not dest:
        return dest

    upper_dest = dest.upper().strip()

    # Check explicit mapping
    if upper_dest in DESTINATION_NAME_MAP:
        return DESTINATION_NAME_MAP[upper_dest]

    # Return original with title case if no mapping found
    return dest.strip()


def build_destination_to_row_map(ws, special_rows: Dict[str, int] = None) -> Dict[str, int]:
    """
    Build mapping of destination names to row numbers in worksheet

    Args:
        ws: openpyxl worksheet
        special_rows: Dict of special destination handling (e.g., Mexico -> 300)

    Returns:
        Dict mapping destination name (uppercase) to row number
    """
    dest_to_row = {}

    for row in range(4, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value:
            name = str(cell_value).strip()
            # Skip indented subtotals/headers (those starting with spaces)
            if name and not name.startswith(' '):
                dest_to_row[name.upper()] = row

    # Add special row overrides
    if special_rows:
        dest_to_row.update({k.upper(): v for k, v in special_rows.items()})

    return dest_to_row


def update_excel_file(
    excel_path: Path,
    weekly_data: Dict[date, Dict[str, float]],
    monthly_data: Dict[Tuple[date, str], float],
    commodity: str
) -> bool:
    """
    Update the Excel file with new inspection data

    Args:
        excel_path: Path to Excel file
        weekly_data: Weekly data - Dict[week_date, Dict[destination, value]]
        monthly_data: Monthly totals by (date, destination)
        commodity: Commodity name

    Returns:
        True if successful
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return False

    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        return False

    logger.info(f"Updating {excel_path}...")

    try:
        wb = openpyxl.load_workbook(excel_path)

        # Special row handling for Mexico inspections
        special_rows = {}
        if commodity.upper() == 'SOYBEANS':
            special_rows['MEXICO'] = 300  # Inspections row for Mexico

        # Update Weekly sheet
        if WEEKLY_SHEET in wb.sheetnames:
            ws = wb[WEEKLY_SHEET]
            updated_count = _update_weekly_sheet(ws, weekly_data, special_rows)
            logger.info(f"Updated {WEEKLY_SHEET}: {updated_count} cells updated")
        else:
            logger.warning(f"Sheet '{WEEKLY_SHEET}' not found in workbook")

        # Update Monthly sheet
        if MONTHLY_SHEET in wb.sheetnames:
            ws = wb[MONTHLY_SHEET]
            updated_count = _update_monthly_sheet(ws, monthly_data, special_rows)
            logger.info(f"Updated {MONTHLY_SHEET}: {updated_count} cells updated")
        else:
            logger.warning(f"Sheet '{MONTHLY_SHEET}' not found in workbook")

        wb.save(excel_path)
        logger.info(f"Saved {excel_path}")
        return True

    except Exception as e:
        logger.error(f"Failed to update Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False


def _update_weekly_sheet(ws, weekly_data: Dict[date, Dict[str, float]],
                         special_rows: Dict[str, int] = None) -> int:
    """
    Update the weekly inspections sheet

    Args:
        ws: openpyxl worksheet
        weekly_data: Dict mapping week_date to Dict[destination, thousand_bushels]
        special_rows: Special destination->row mappings

    Returns:
        Number of cells updated
    """
    updated = 0
    date_row = 2  # Dates are in row 2

    # Build mapping of existing dates to columns
    date_to_col = {}
    max_col = ws.max_column

    for col in range(2, max_col + 1):
        cell_value = ws.cell(row=date_row, column=col).value
        if cell_value:
            if isinstance(cell_value, datetime):
                date_to_col[cell_value.date()] = col
            elif isinstance(cell_value, date):
                date_to_col[cell_value] = col

    # Build destination to row mapping
    dest_to_row = build_destination_to_row_map(ws, special_rows)

    # Track new columns needed
    new_dates = []
    for week_date in weekly_data.keys():
        if week_date not in date_to_col:
            new_dates.append(week_date)

    # Add new date columns if needed (at the end)
    if new_dates:
        next_col = max_col + 1
        for new_date in sorted(new_dates):
            ws.cell(row=date_row, column=next_col).value = new_date
            date_to_col[new_date] = next_col
            logger.info(f"Added new date column {next_col} for {new_date}")
            next_col += 1

    # Update cells with data
    for week_date, dest_values in weekly_data.items():
        col = date_to_col.get(week_date)
        if not col:
            logger.warning(f"No column found for date {week_date}")
            continue

        for destination, value in dest_values.items():
            # Normalize destination name
            norm_dest = normalize_destination(destination)

            # Find row for this destination
            row = dest_to_row.get(norm_dest.upper())
            if not row:
                # Try the original name
                row = dest_to_row.get(destination.upper())

            if row:
                # Only write non-zero values (blank for zeros per video instructions)
                if value and value > 0:
                    # Round to reasonable precision
                    ws.cell(row=row, column=col).value = round(value, 3)
                    updated += 1
            else:
                # Only log warning for significant values we couldn't place
                if value and value > 1.0:
                    logger.debug(f"No row found for destination: {destination} ({norm_dest})")

    return updated


def _update_monthly_sheet(ws, monthly_data: Dict[Tuple[date, str], float],
                          special_rows: Dict[str, int] = None) -> int:
    """
    Update the monthly inspections sheet

    Args:
        ws: openpyxl worksheet
        monthly_data: Dict mapping (month_date, destination) to thousand_bushels
        special_rows: Special destination->row mappings

    Returns:
        Number of cells updated
    """
    updated = 0
    date_row = 2  # Dates are in row 2

    # Build mapping of existing dates to columns
    date_to_col = {}
    max_col = ws.max_column

    for col in range(2, max_col + 1):
        cell_value = ws.cell(row=date_row, column=col).value
        if cell_value:
            if isinstance(cell_value, datetime):
                # Normalize to first of month
                month_date = cell_value.date().replace(day=1)
                date_to_col[month_date] = col
            elif isinstance(cell_value, date):
                month_date = cell_value.replace(day=1)
                date_to_col[month_date] = col

    # Build destination to row mapping
    dest_to_row = build_destination_to_row_map(ws, special_rows)

    # Track new month columns needed
    new_months = set()
    for (month_date, dest) in monthly_data.keys():
        month_first = month_date.replace(day=1) if isinstance(month_date, date) else month_date
        if month_first not in date_to_col:
            new_months.add(month_first)

    # Add new month columns if needed
    if new_months:
        next_col = max_col + 1
        for new_month in sorted(new_months):
            ws.cell(row=date_row, column=next_col).value = new_month
            date_to_col[new_month] = next_col
            logger.info(f"Added new month column {next_col} for {new_month}")
            next_col += 1

    # Update cells with data
    for (month_date, destination), value in monthly_data.items():
        # Normalize month date
        month_first = month_date.replace(day=1) if isinstance(month_date, date) else month_date

        col = date_to_col.get(month_first)
        if not col:
            logger.warning(f"No column found for month {month_first}")
            continue

        # Normalize destination name
        norm_dest = normalize_destination(destination)

        # Find row for this destination
        row = dest_to_row.get(norm_dest.upper())
        if not row:
            row = dest_to_row.get(destination.upper())

        if row:
            # Only write non-zero values
            if value and value > 0:
                ws.cell(row=row, column=col).value = round(value, 3)
                updated += 1
        else:
            if value and value > 1.0:
                logger.debug(f"No row found for destination: {destination}")

    return updated


# =============================================================================
# DATABASE OPERATIONS
# =============================================================================

def save_to_database(records: List[Dict], connection_string: str = None) -> int:
    """
    Save inspection records to the RLC_commodities database

    Args:
        records: List of parsed records
        connection_string: Database connection string (uses DATABASE_URL env var if not provided)

    Returns:
        Number of records inserted
    """
    connection_string = connection_string or os.getenv('DATABASE_URL')

    if not connection_string:
        logger.error("No database connection string provided. Set DATABASE_URL environment variable.")
        return 0

    try:
        if connection_string.startswith('postgresql'):
            return _save_to_postgresql(records, connection_string)
        elif connection_string.startswith('sqlite'):
            return _save_to_sqlite(records, connection_string)
        else:
            logger.error(f"Unsupported database type: {connection_string.split(':')[0]}")
            return 0
    except Exception as e:
        logger.error(f"Failed to save to database: {e}")
        return 0


def _save_to_postgresql(records: List[Dict], connection_string: str) -> int:
    """Save records to PostgreSQL"""
    try:
        import psycopg2
        from psycopg2.extras import execute_values
    except ImportError:
        logger.error("psycopg2 not installed. Run: pip install psycopg2-binary")
        return 0

    from urllib.parse import urlparse
    parsed = urlparse(connection_string)

    conn = psycopg2.connect(
        host=parsed.hostname,
        port=parsed.port or 5432,
        database=parsed.path[1:],
        user=parsed.username,
        password=parsed.password
    )

    cursor = conn.cursor()

    # Create table if not exists
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS export_inspections (
            id BIGSERIAL PRIMARY KEY,
            week_ending DATE NOT NULL,
            commodity VARCHAR(50) NOT NULL,
            destination VARCHAR(100) NOT NULL,
            pounds NUMERIC,
            metric_tons NUMERIC,
            thousand_bushels NUMERIC,
            marketing_year VARCHAR(10),
            port VARCHAR(50),
            grade VARCHAR(50),
            collected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (week_ending, commodity, destination, port, grade)
        )
    """)

    # Prepare data for insertion
    values = [
        (
            r['week_ending'],
            r['grain'],
            r['destination'],
            r.get('pounds'),
            r.get('metric_tons'),
            r.get('thousand_bushels'),
            r.get('marketing_year'),
            r.get('port'),
            r.get('grade'),
        )
        for r in records
    ]

    # Insert with upsert
    insert_sql = """
        INSERT INTO export_inspections
        (week_ending, commodity, destination, pounds, metric_tons,
         thousand_bushels, marketing_year, port, grade)
        VALUES %s
        ON CONFLICT (week_ending, commodity, destination, port, grade)
        DO UPDATE SET
            pounds = EXCLUDED.pounds,
            metric_tons = EXCLUDED.metric_tons,
            thousand_bushels = EXCLUDED.thousand_bushels,
            collected_at = CURRENT_TIMESTAMP
    """

    execute_values(cursor, insert_sql, values)
    inserted = cursor.rowcount

    conn.commit()
    cursor.close()
    conn.close()

    logger.info(f"Inserted/updated {inserted} records in PostgreSQL")
    return inserted


def _save_to_sqlite(records: List[Dict], connection_string: str) -> int:
    """Save records to SQLite"""
    import sqlite3

    db_path = connection_string.replace('sqlite:///', '')
    os.makedirs(os.path.dirname(db_path) or '.', exist_ok=True)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Create table if not exists
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS export_inspections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            week_ending TEXT NOT NULL,
            commodity TEXT NOT NULL,
            destination TEXT NOT NULL,
            pounds REAL,
            metric_tons REAL,
            thousand_bushels REAL,
            marketing_year TEXT,
            port TEXT,
            grade TEXT,
            collected_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (week_ending, commodity, destination, port, grade)
        )
    """)

    # Insert records
    inserted = 0
    for r in records:
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO export_inspections
                (week_ending, commodity, destination, pounds, metric_tons,
                 thousand_bushels, marketing_year, port, grade)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                r['week_ending'].isoformat() if r['week_ending'] else None,
                r['grain'],
                r['destination'],
                r.get('pounds'),
                r.get('metric_tons'),
                r.get('thousand_bushels'),
                r.get('marketing_year'),
                r.get('port'),
                r.get('grade'),
            ))
            inserted += 1
        except sqlite3.Error as e:
            logger.warning(f"Failed to insert record: {e}")

    conn.commit()
    cursor.close()
    conn.close()

    logger.info(f"Inserted/updated {inserted} records in SQLite")
    return inserted


# =============================================================================
# MAIN WORKFLOW
# =============================================================================

def run_weekly_update(
    commodity: str = 'SOYBEANS',
    years: List[int] = None,
    data_dir: Path = None,
    update_excel: bool = False,
    save_to_db: bool = False,
    force_download: bool = False,
    check_corrections: bool = True
) -> Dict:
    """
    Run the complete weekly update workflow

    PROCESS (from video transcript):
    1. Download wa_gr101.txt and check correction date for commodity
    2. Download CY{year}.csv files for required years
    3. Parse CSV and filter by commodity
    4. Create pivot: weeks as columns, destinations as rows, sum of metric tons
    5. Convert to thousand bushels: MT * 36.7437 / 1000
    6. Aggregate weekly totals and monthly totals by destination
    7. Update Excel file (Weekly Export Inspections + Monthly Export Inspections tabs)
    8. Save to database

    Args:
        commodity: Commodity to process (SOYBEANS, CORN, WHEAT)
        years: List of years to process (default: current and previous year)
        data_dir: Directory for downloaded files
        update_excel: Whether to update Excel file
        save_to_db: Whether to save to database
        force_download: Force re-download of files
        check_corrections: Check wa_gr101.txt for correction dates (recommended)

    Returns:
        Dict with results summary
    """
    results = {
        'success': False,
        'files_downloaded': 0,
        'records_parsed': 0,
        'weekly_totals': 0,
        'monthly_totals': 0,
        'records_saved': 0,
        'correction_date': None,
    }

    # Set defaults
    if years is None:
        current_year = datetime.now().year
        years = [current_year - 1, current_year]

    if data_dir is None:
        data_dir = Path(__file__).parent.parent / 'data' / 'raw'

    project_root = Path(__file__).parent.parent

    logger.info(f"Starting weekly inspection update for {commodity}")
    logger.info(f"Years: {years}")

    # Step 0: Check correction date from weekly report (CRITICAL STEP)
    correction_date = None
    if check_corrections:
        logger.info("Step 0: Checking correction date from wa_gr101.txt...")
        report_path = download_weekly_report(data_dir)
        if report_path:
            correction_date = parse_correction_date(report_path, commodity)
            results['correction_date'] = str(correction_date) if correction_date else None
            if correction_date:
                logger.info(f"IMPORTANT: Corrections found through {correction_date}")
                logger.info(f"Will repull data from {correction_date} forward to capture corrections")
            else:
                logger.info("No correction date found - using default date range")

    # Step 1: Download files
    all_records = []

    for year in years:
        file_path = download_csv(year, data_dir, force=force_download)
        if file_path:
            results['files_downloaded'] += 1

            # Parse the file
            records = parse_csv_file(file_path, commodity_filter=commodity)
            all_records.extend(records)

    results['records_parsed'] = len(all_records)
    logger.info(f"Parsed {len(all_records)} total records for {commodity}")

    if not all_records:
        logger.warning("No records found")
        return results

    # Step 2: Aggregate data
    weekly_totals = aggregate_weekly_totals(all_records, commodity)
    weekly_by_dest = aggregate_weekly_by_destination(all_records, commodity)
    monthly_by_dest = aggregate_monthly_by_destination(all_records, commodity)
    yearly_by_dest = aggregate_marketing_year_by_destination(all_records, commodity)

    results['weekly_totals'] = len(weekly_totals)
    results['monthly_totals'] = len(monthly_by_dest)

    logger.info(f"Aggregated to {len(weekly_totals)} weeks, {len(monthly_by_dest)} month/destination pairs")

    # Print summary of recent data
    print("\n" + "=" * 60)
    print(f"WEEKLY INSPECTION TOTALS - {commodity}")
    print("=" * 60)

    recent_weeks = sorted(weekly_totals.items(), reverse=True)[:10]
    for week_date, value in recent_weeks:
        print(f"  {week_date}: {value:,.1f} thousand bushels")

    print("\n" + "=" * 60)
    print(f"MARKETING YEAR TOTALS BY DESTINATION - {commodity}")
    print("=" * 60)

    # Get latest marketing year
    latest_my = max(r['marketing_year'] for r in all_records)
    my_totals = {dest: val for (my, dest), val in yearly_by_dest.items() if my == latest_my}

    # Sort by value and show top 10
    for dest, value in sorted(my_totals.items(), key=lambda x: -x[1])[:10]:
        print(f"  {dest}: {value:,.2f} million bushels")

    # Step 3: Update Excel (if requested)
    if update_excel:
        excel_file = EXCEL_FILES.get(commodity.upper())
        if excel_file:
            excel_path = project_root / excel_file
            if update_excel_file(excel_path, weekly_by_dest, monthly_by_dest, commodity):
                logger.info("Excel file updated successfully")
            else:
                logger.warning("Failed to update Excel file")
        else:
            logger.warning(f"No Excel file configured for {commodity}")

    # Step 4: Save to database (if requested)
    if save_to_db:
        inserted = save_to_database(all_records)
        results['records_saved'] = inserted

    results['success'] = True
    return results


def main():
    """Command-line entry point"""
    parser = argparse.ArgumentParser(
        description='Download and process weekly export inspection data'
    )

    parser.add_argument(
        '--commodity', '-c',
        default='SOYBEANS',
        choices=['SOYBEANS', 'CORN', 'WHEAT', 'SORGHUM', 'ALL'],
        help='Commodity to process (default: SOYBEANS)'
    )

    parser.add_argument(
        '--year', '-y',
        type=int,
        action='append',
        help='Year(s) to process (can specify multiple times)'
    )

    parser.add_argument(
        '--update-excel',
        action='store_true',
        help='Update the Excel model file'
    )

    parser.add_argument(
        '--save-to-db',
        action='store_true',
        help='Save data to RLC_commodities database'
    )

    parser.add_argument(
        '--force',
        action='store_true',
        help='Force re-download of files'
    )

    parser.add_argument(
        '--data-dir',
        type=Path,
        help='Directory for downloaded files'
    )

    args = parser.parse_args()

    # Handle ALL commodities
    commodities = ['SOYBEANS', 'CORN', 'WHEAT', 'SORGHUM'] if args.commodity == 'ALL' else [args.commodity]

    for commodity in commodities:
        logger.info(f"\n{'=' * 60}")
        logger.info(f"Processing {commodity}")
        logger.info(f"{'=' * 60}")

        results = run_weekly_update(
            commodity=commodity,
            years=args.year,
            data_dir=args.data_dir,
            update_excel=args.update_excel,
            save_to_db=args.save_to_db,
            force_download=args.force
        )

        if results['success']:
            logger.info(f"Completed: {results['records_parsed']} records, "
                       f"{results['weekly_totals']} weeks, "
                       f"{results['records_saved']} saved to DB")
        else:
            logger.error(f"Failed to process {commodity}")


if __name__ == '__main__':
    main()
