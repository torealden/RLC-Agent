"""Emit contract-compliant (flat_file_contract.md v1.1 LONG) fats/greases/DCO supply+demand flat files.

Six workbooks, each with a supply tab, a demand tab, and a _meta tab:
  1. models/Fats and Greases/us_tallow_supply_demand.xlsx        -> tallow_supply,       tallow_demand,       _meta
  2. models/Fats and Greases/us_uco_supply_demand.xlsx           -> uco_supply,          uco_demand,          _meta
  3. models/Fats and Greases/us_poultry_fat_supply_demand.xlsx   -> poultry_fat_supply,  poultry_fat_demand,  _meta
  4. models/Fats and Greases/us_white_grease_supply_demand.xlsx  -> white_grease_supply, white_grease_demand, _meta
  5. models/Fats and Greases/us_yellow_grease_supply_demand.xlsx -> yellow_grease_supply,yellow_grease_demand,_meta
  6. models/Fats and Greases/us_dco_supply_demand.xlsx           -> dco_supply,          dco_demand,          _meta

DCO (distillers corn oil) lives here rather than with the vegetable oils: functionally it is a
low-CI waste-oil biofuel feedstock grouped with the fats/greases, not a food oil like SBO/canola.

Supply tabs are near-passthroughs / assembly from silver.* / gold.* tables (the vintage ladder is
preserved so Desktop's MAXIFS picks the top-ranked vintage per key). Demand tabs are the raked
allocator output (gold.bbd_feedstock_raked, latest run_day). 13-col LONG schema, value = RAW pounds.

Idempotent: rewrites all six files each run. Does NOT commit/push.
"""
import sys
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

ROOT = Path(r"C:/dev/RLC-Agent"); sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
from src.services.database.db_config import get_connection

OUTDIR = ROOT / "models" / "Fats and Greases"
COLS = ['commodity', 'class', 'series', 'marketing_year', 'period_type', 'period',
        'vintage', 'vintage_rank', 'value', 'unit', 'source', 'release_date', 'revision']

# vintage -> rank map for the UCO assembled series (uco_yg_balance has no vintage_rank column)
UCO_RANK = {'PROXY_FOOD': 40, 'PROXY_FOOD_CARRYFWD': 35, 'CENSUS': 90,
            'DERIVED': 90, 'EIA': 95, 'RULED_AMENDMENT1': 30}

TALLOW_SUPPLY_SERIES = (
    'tallow_production', 'tallow_imports', 'tallow_exports', 'tallow_biofuel_use',
    'non_bio_use', 'feed_use', 'non_bio_trend', 'eia_tallow_comparison',
    'cir_ibft_biodiesel_comparison',
)

# NASS Fats & Oils supply mapping: commodity -> [(matrix_column, output_series), ...]
NASS_SUPPLY_SPEC = {
    'poultry_fat': [('poultry_fat_production', 'production'),
                    ('poultry_fat_processing_use', 'processing_use'),
                    ('poultry_fat_stocks', 'stocks')],
    'white_grease': [('cwg_production', 'production'),
                     ('cwg_processing_use', 'processing_use'),
                     ('cwg_stocks', 'stocks')],
    'yellow_grease': [('yellow_grease_production', 'production'),
                      ('yellow_grease_processing_use', 'processing_use'),
                      ('yellow_grease_stocks', 'stocks')],
}

# demand feedstock_code filter per commodity
DEMAND_CODES = {
    'tallow': ('EBFT', 'IBFT', 'BFT'),
    'uco_yg': ('UCO', 'YG'),
    'poultry_fat': ('PF', 'PLT'),
    'white_grease': ('CWG',),
    'yellow_grease': ('YG',),
    'dco': ('DCO',),
}

STALE_NASS = ('Supply from NASS Fats & Oils via gold.nass_low_ci_matrix; recurring collector live '
              '(2026-07-08) so it refreshes monthly to the latest NASS release. Demand (raked) runs '
              'through the latest allocator vintage. non_biofuel_use: disappearance-based (NASS '
              'processing_use - biofuel) for PF/CWG/YG, tallow=modeled; UCO/DCO are biofuel-dominant '
              'so residual non-bio is ~0 in the current period.')

NOTES = {
    # tallow supply
    'tallow_production': 'NASS/CENSUS_CIR/SLAUGHTER vintage ladder (rank 90/80/60)',
    'tallow_imports': 'Census gross imports by grade (EBFT/IBFT)',
    'tallow_exports': 'Census gross exports by grade (EBFT/IBFT)',
    'tallow_biofuel_use': 'modeled biofuel consumption, all grades',
    'non_bio_use': 'IBFT non-biofuel demand (model)',
    'feed_use': 'IBFT feed demand (model)',
    'non_bio_trend': 'IBFT non-biofuel trend line (model)',
    'eia_tallow_comparison': 'EIA comparison only -- never load-bearing',
    'cir_ibft_biodiesel_comparison': 'CIR-era IBFT biodiesel comparison only',
    # uco supply
    'uco_collection': 'UCO-grade collection, food-oil proxy',
    'yg_collection': 'YG-grade collection, food-oil proxy',
    'uco_imports': 'Census gross UCO imports (TOTAL)',
    'uco_exports': 'Census gross UCO exports (TOTAL)',
    'uco_biofuel_use': 'derived UCO biofuel consumption',
    'yg_biofuel_use': 'ruled 0 per UCO Amendment 1',
    # NASS fats/greases supply (generic series names)
    'production': 'NASS Fats & Oils production (raw lb)',
    'processing_use': 'NASS Fats & Oils consumption/processing use (raw lb)',
    'stocks': 'NASS Fats & Oils ending stocks (raw lb)',
    # dco supply
    'dco_production': 'GCCP distillers corn oil co-product: gold.corn_grind_monthly target_col=K '
                      '"Corn oil (DCO)" (thousand short tons -> LB x 2e6)',
    'dco_imports': 'Census DCO imports: gold.dco_trade_monthly trade_category=DCO (000 lb -> LB x 1000)',
    'dco_exports': 'Census DCO exports: gold.dco_trade_monthly trade_category=DCO (000 lb -> LB x 1000)',
    # demand (all)
    'biofuel_use_biodiesel': 'raked allocator: biodiesel feedstock use',
    'biofuel_use_renewable_diesel': 'raked allocator: renewable diesel feedstock use',
    'biofuel_use_saf': 'raked allocator: SAF feedstock use',
    'biofuel_use_coprocessing': 'raked allocator: co-processing feedstock use',
    'biofuel_use_total': 'raked allocator: sum across all fuel types',
    'non_biofuel_use': 'aggregate non-biofuel use (closes the balance sheet). Disappearance-based '
                       '(NASS processing_use - biofuel) for PF/CWG/YG; residual (prod+imp-bio-exp) for '
                       'DCO/UCO; tallow = its first-class modeled non_bio_use. Ruled 2026-07-13.',
}
# non_bio_use override note for the UCO workbook (different meaning than tallow)
UCO_NONBIO_NOTE = 'YG-grade collection to non-bio, YG biofuel=0'


def pm(month):
    """month int -> 'M01'..'M12'"""
    return f"M{int(month):02d}"


def sort_key(r):
    return (r['series'], r['class'], r['marketing_year'], r['period'], r['vintage_rank'])


def dedupe(rows):
    """§7 uniqueness: one row per (commodity,class,series,marketing_year,period,vintage_rank)."""
    seen = {}
    for r in rows:
        k = (r['commodity'], r['class'], r['series'], r['marketing_year'], r['period'], r['vintage_rank'])
        if k not in seen:
            seen[k] = r
    return list(seen.values())


def write_tab(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in sorted(dedupe(rows), key=sort_key):
        ws.append([r[c] for c in COLS])
    return ws


def write_meta(wb, series_rows, note_override=None, extra_meta=None):
    """series_rows: list of dicts w/ series, source, unit, vintage_set.
    extra_meta: list of full meta dicts (series/source/unit/vintage_set/notes) appended verbatim."""
    wm = wb.create_sheet("_meta")
    wm.append(['series', 'source', 'unit', 'vintage_set', 'last_updated', 'notes'])
    for cell in wm[1]:
        cell.font = Font(bold=True)
    for m in series_rows:
        note = (note_override or {}).get(m['series'], NOTES.get(m['series'], ''))
        wm.append([m['series'], m['source'], m['unit'], m['vintage_set'], '', note])
    for m in (extra_meta or []):
        wm.append([m['series'], m['source'], m['unit'], m['vintage_set'], '', m.get('notes', '')])


def meta_from_rows(rows):
    """Build one _meta row per series from emitted data rows (preserving first-seen order)."""
    agg = {}
    order = []
    for r in rows:
        s = r['series']
        if s not in agg:
            agg[s] = {'series': s, 'source': set(), 'unit': set(), 'vintage_set': set()}
            order.append(s)
        agg[s]['source'].add(r['source'])
        agg[s]['unit'].add(r['unit'])
        agg[s]['vintage_set'].add(r['vintage'])
    out = []
    for s in order:
        a = agg[s]
        out.append({'series': s,
                    'source': ', '.join(sorted(a['source'])),
                    'unit': ', '.join(sorted(a['unit'])),
                    'vintage_set': ', '.join(sorted(a['vintage_set']))})
    return out


# ---------------------------------------------------------------- build data
with get_connection() as conn:
    cur = conn.cursor()
    cur.execute("SELECT max(run_day) m FROM gold.bbd_feedstock_raked")
    max_run = cur.fetchone()['m']

    # ---- TALLOW SUPPLY (near-passthrough of silver.tallow_balance) ----
    cur.execute(
        "SELECT class, series, year, month, value_lbs, vintage, vintage_rank "
        "FROM silver.tallow_balance WHERE series = ANY(%s)",
        (list(TALLOW_SUPPLY_SERIES),))
    tallow_supply = [{
        'commodity': 'tallow', 'class': r['class'], 'series': r['series'],
        'marketing_year': r['year'], 'period_type': 'cal_month', 'period': pm(r['month']),
        'vintage': r['vintage'], 'vintage_rank': r['vintage_rank'], 'value': r['value_lbs'],
        'unit': 'LB', 'source': r['vintage'], 'release_date': '', 'revision': '',
    } for r in cur.fetchall()]

    # ---- UCO SUPPLY (assembled) ----
    uco_supply = []

    def uco_row(series, year, month, value, vintage):
        return {'commodity': 'uco_yg', 'class': 'ALL', 'series': series,
                'marketing_year': year, 'period_type': 'cal_month', 'period': pm(month),
                'vintage': vintage, 'vintage_rank': UCO_RANK[vintage], 'value': value,
                'unit': 'LB', 'source': vintage, 'release_date': '', 'revision': ''}

    cur.execute("SELECT series, year, month, value_lbs, vintage FROM silver.uco_yg_balance "
                "WHERE series IN ('uco_collection','yg_collection','uco_biofuel_use')")
    yg_collection_rows = []
    uco_collection_keys = []
    for r in cur.fetchall():
        uco_supply.append(uco_row(r['series'], r['year'], r['month'], r['value_lbs'], r['vintage']))
        if r['series'] == 'yg_collection':
            yg_collection_rows.append((r['year'], r['month'], r['value_lbs']))
        if r['series'] == 'uco_collection':
            uco_collection_keys.append((r['year'], r['month']))

    cur.execute("SELECT year, month, flow, mil_lbs FROM silver.uco_imports WHERE country='TOTAL'")
    for r in cur.fetchall():
        series = 'uco_imports' if r['flow'] == 'import' else 'uco_exports'
        uco_supply.append(uco_row(series, r['year'], r['month'], float(r['mil_lbs']) * 1e6, 'CENSUS'))

    for (y, m) in uco_collection_keys:
        uco_supply.append(uco_row('yg_biofuel_use', y, m, 0, 'RULED_AMENDMENT1'))
    for (y, m, v) in yg_collection_rows:
        uco_supply.append(uco_row('non_bio_use', y, m, v, 'DERIVED'))

    # ---- NASS FATS & OILS SUPPLY (poultry_fat, white_grease, yellow_grease) ----
    def nass_supply(commodity):
        spec = NASS_SUPPLY_SPEC[commodity]
        matrix_cols = [c for c, _ in spec]
        cur.execute(
            "SELECT calendar_year, month, " + ", ".join(matrix_cols) +
            " FROM gold.nass_low_ci_matrix ORDER BY calendar_year, month")
        rows = []
        for r in cur.fetchall():
            for mcol, series in spec:
                v = r[mcol]
                if v is None:
                    continue
                rows.append({
                    'commodity': commodity, 'class': 'ALL', 'series': series,
                    'marketing_year': r['calendar_year'], 'period_type': 'cal_month',
                    'period': pm(r['month']), 'vintage': 'NASS_FATS_OILS', 'vintage_rank': 90,
                    'value': float(v), 'unit': 'LB', 'source': 'NASS_FATS_OILS',
                    'release_date': '', 'revision': '',
                })
        return rows

    poultry_fat_supply = nass_supply('poultry_fat')
    white_grease_supply = nass_supply('white_grease')
    yellow_grease_supply = nass_supply('yellow_grease')

    # ---- DCO SUPPLY (production from GCCP corn_grind_monthly K; trade from dco_trade_monthly) ----
    dco_supply = []
    # production: target_col='K' "Corn oil (DCO)" thousand short tons -> LB (x 1000 x 2000 = x 2e6)
    cur.execute("SELECT year, month, display_value, display_unit FROM gold.corn_grind_monthly "
                "WHERE target_col = 'K' ORDER BY year, month")
    dco_prod_unit = None
    for r in cur.fetchall():
        if r['display_value'] is None:
            continue
        dco_prod_unit = r['display_unit']
        dco_supply.append({
            'commodity': 'dco', 'class': 'ALL', 'series': 'dco_production',
            'marketing_year': r['year'], 'period_type': 'cal_month', 'period': pm(r['month']),
            'vintage': 'NASS_GRAIN_CRUSH', 'vintage_rank': 90,
            'value': float(r['display_value']) * 2_000_000.0, 'unit': 'LB',
            'source': 'GCCP_corn_grind_monthly.K', 'release_date': '', 'revision': '',
        })
    # trade: gold.dco_trade_monthly trade_category='DCO', total_000_lbs -> LB (x 1000)
    cur.execute("SELECT year, month, flow, total_000_lbs FROM gold.dco_trade_monthly "
                "WHERE trade_category = 'DCO' ORDER BY year, month, flow")
    for r in cur.fetchall():
        if r['total_000_lbs'] is None:
            continue
        series = 'dco_imports' if str(r['flow']).lower().startswith('import') else 'dco_exports'
        dco_supply.append({
            'commodity': 'dco', 'class': 'ALL', 'series': series,
            'marketing_year': r['year'], 'period_type': 'cal_month', 'period': pm(r['month']),
            'vintage': 'CENSUS', 'vintage_rank': 90,
            'value': float(r['total_000_lbs']) * 1000.0, 'unit': 'LB',
            'source': 'CENSUS_dco_trade_monthly', 'release_date': '', 'revision': '',
        })

    # ---- DEMAND (raked allocator) ----
    def demand_rows(commodity, codes):
        cur.execute(
            "SELECT extract(year from period)::int yr, extract(month from period)::int mo, "
            "fuel_type, sum(raked_mil_lbs) mil FROM gold.bbd_feedstock_raked "
            "WHERE run_day = %s AND feedstock_code = ANY(%s) "
            "GROUP BY 1,2,3", (max_run, list(codes)))
        raw = cur.fetchall()
        rows = []
        totals = {}
        for r in raw:
            series = 'biofuel_use_' + str(r['fuel_type']).lower().replace(' ', '_')
            val = float(r['mil']) * 1e6
            rows.append({'commodity': commodity, 'class': 'ALL', 'series': series,
                         'marketing_year': r['yr'], 'period_type': 'cal_month', 'period': pm(r['mo']),
                         'vintage': 'RAKED', 'vintage_rank': 95, 'value': val, 'unit': 'LB',
                         'source': 'ALLOCATOR_RAKED', 'release_date': '', 'revision': ''})
            totals[(r['yr'], r['mo'])] = totals.get((r['yr'], r['mo']), 0.0) + val
        for (yr, mo), tot in totals.items():
            rows.append({'commodity': commodity, 'class': 'ALL', 'series': 'biofuel_use_total',
                         'marketing_year': yr, 'period_type': 'cal_month', 'period': pm(mo),
                         'vintage': 'RAKED', 'vintage_rank': 95, 'value': tot, 'unit': 'LB',
                         'source': 'ALLOCATOR_RAKED', 'release_date': '', 'revision': ''})
        return rows

    tallow_demand = demand_rows('tallow', DEMAND_CODES['tallow'])
    uco_demand = demand_rows('uco_yg', DEMAND_CODES['uco_yg'])
    poultry_fat_demand = demand_rows('poultry_fat', DEMAND_CODES['poultry_fat'])
    white_grease_demand = demand_rows('white_grease', DEMAND_CODES['white_grease'])
    yellow_grease_demand = demand_rows('yellow_grease', DEMAND_CODES['yellow_grease'])
    dco_demand = demand_rows('dco', DEMAND_CODES['dco'])

    # ---- AGGREGATE NON-BIOFUEL USE (closes the balance sheet) ----
    # Ruled 2026-07-13 (Tore): disappearance-based where it exists, residual only as fallback.
    # tallow keeps its first-class modeled non_bio_use (already in the supply tab) — not recomputed.
    def _by_period(rows, series):
        return {(r['marketing_year'], r['period']): float(r['value'] or 0)
                for r in rows if r['series'] == series}

    def nonbio_disappearance(commodity, supply_rows, demand_rows_):
        """non_biofuel_use = NASS processing_use (total inedible consumption incl. biofuel) - biofuel_total.
        Disappearance-based: NASS 'removal for processing' is the measured consumption; the non-bio leg
        is what's left after the raked biofuel portion. Clamp >=0."""
        proc = _by_period(supply_rows, 'processing_use')
        bio = _by_period(demand_rows_, 'biofuel_use_total')
        out = []
        for k, pv in proc.items():
            nb = max(0.0, pv - bio.get(k, 0.0))
            out.append({'commodity': commodity, 'class': 'ALL', 'series': 'non_biofuel_use',
                        'marketing_year': k[0], 'period_type': 'cal_month', 'period': k[1],
                        'vintage': 'DISAPPEARANCE', 'vintage_rank': 90, 'value': nb, 'unit': 'LB',
                        'source': 'NASS_processing_use - raked_biofuel', 'release_date': '', 'revision': ''})
        return out

    def nonbio_residual(commodity, supply_rows, demand_rows_, prod_s, imp_s, exp_s):
        """non_biofuel_use = production + imports - biofuel_total - exports. Residual fallback
        for feedstocks with no NASS disappearance series (DCO, UCO). Clamp >=0."""
        prod = _by_period(supply_rows, prod_s); imp = _by_period(supply_rows, imp_s)
        exp = _by_period(supply_rows, exp_s); bio = _by_period(demand_rows_, 'biofuel_use_total')
        out = []
        # only emit where biofuel data exists — past the biofuel frontier the residual has nothing to
        # subtract and spikes to full production (a data-frontier artifact, not real non-bio demand).
        for k in sorted(set(prod) & set(bio)):
            nb = max(0.0, prod.get(k, 0) + imp.get(k, 0) - bio.get(k, 0) - exp.get(k, 0))
            out.append({'commodity': commodity, 'class': 'ALL', 'series': 'non_biofuel_use',
                        'marketing_year': k[0], 'period_type': 'cal_month', 'period': k[1],
                        'vintage': 'RESIDUAL', 'vintage_rank': 50, 'value': nb, 'unit': 'LB',
                        'source': 'residual: production+imports-biofuel-exports', 'release_date': '', 'revision': ''})
        return out

    def nonbio_components(commodity, demand_rows_):
        """Split each non_biofuel_use TOTAL into end-use component lines via the SEASONAL
        (calendar-month) shares in reference.nonbio_enduse_shares_monthly — each month uses its own
        share so components breathe seasonally instead of a flat percentage. Falls back to the flat
        annual reference.nonbio_enduse_shares if a commodity has no monthly rows. Components sum to
        the total within each month, so the sheet still closes. Shared with the oils writer/Desktop."""
        cur.execute("""SELECT month, end_use, share_pct, measured
                       FROM reference.nonbio_enduse_shares_monthly WHERE commodity=%s""", (commodity,))
        monthly = defaultdict(list)
        for s in cur.fetchall():
            monthly[s['month']].append(s)
        if not monthly:
            cur.execute("""SELECT end_use, share_pct, measured FROM reference.nonbio_enduse_shares
                           WHERE commodity=%s""", (commodity,))
            flat = cur.fetchall()
            monthly = {m: flat for m in range(1, 13)}
        out = []
        for r in demand_rows_:
            if r['series'] != 'non_biofuel_use':
                continue
            tot = float(r['value'] or 0)
            mo = int(r['period'][1:])
            ms = monthly.get(mo, [])
            ssum = sum(float(s['share_pct']) for s in ms) or 1.0  # normalize -> exact closure
            for s in ms:
                measured = s['measured']
                out.append({'commodity': commodity, 'class': 'ALL',
                            'series': 'nonbiofuel_use_' + s['end_use'],
                            'marketing_year': r['marketing_year'], 'period_type': 'cal_month',
                            'period': r['period'],
                            'vintage': 'NONBIO_MEASURED' if measured else 'NONBIO_MODELED',
                            'vintage_rank': r['vintage_rank'],
                            'value': tot * float(s['share_pct']) / ssum, 'unit': 'LB',
                            'source': ('Census 2006-2011 seasonal end-use share x non_biofuel_use'
                                       if measured else
                                       'analog/modeled seasonal end-use share x non_biofuel_use'),
                            'release_date': '', 'revision': ''})
        return out

    poultry_fat_demand += nonbio_disappearance('poultry_fat', poultry_fat_supply, poultry_fat_demand)
    white_grease_demand += nonbio_disappearance('white_grease', white_grease_supply, white_grease_demand)
    yellow_grease_demand += nonbio_disappearance('yellow_grease', yellow_grease_supply, yellow_grease_demand)
    dco_demand += nonbio_residual('dco', dco_supply, dco_demand, 'dco_production', 'dco_imports', 'dco_exports')
    uco_demand += nonbio_residual('uco_yg', uco_supply, uco_demand, 'uco_collection', 'uco_imports', 'uco_exports')
    # tallow: mirror its modeled non_bio_use into the uniform non_biofuel_use series for a consistent shape
    for r in tallow_supply:
        if r['series'] == 'non_bio_use':
            tallow_demand.append({**r, 'series': 'non_biofuel_use', 'commodity': 'tallow'})

    # split each commodity's non_biofuel_use total into end-use component lines (shared shares)
    tallow_demand += nonbio_components('tallow', tallow_demand)
    uco_demand += nonbio_components('uco_yg', uco_demand)
    poultry_fat_demand += nonbio_components('poultry_fat', poultry_fat_demand)
    white_grease_demand += nonbio_components('white_grease', white_grease_demand)
    yellow_grease_demand += nonbio_components('yellow_grease', yellow_grease_demand)
    dco_demand += nonbio_components('dco', dco_demand)

print(f"DCO production source unit read from corn_grind_monthly.K: {dco_prod_unit!r} "
      f"(converted to LB via x 2,000,000)")

# ---------------------------------------------------------------- write files
OUTDIR.mkdir(parents=True, exist_ok=True)


def build_meta_rows(supply_rows, demand_rows_):
    return meta_from_rows(sorted(dedupe(supply_rows), key=sort_key) +
                          sorted(dedupe(demand_rows_), key=sort_key))


written = []

# File 1: tallow
wb = openpyxl.Workbook(); wb.remove(wb.active)
write_tab(wb, 'tallow_supply', tallow_supply)
write_tab(wb, 'tallow_demand', tallow_demand)
write_meta(wb, build_meta_rows(tallow_supply, tallow_demand))
f = OUTDIR / "us_tallow_supply_demand.xlsx"; wb.save(f); written.append(f)

# File 2: uco
wb = openpyxl.Workbook(); wb.remove(wb.active)
write_tab(wb, 'uco_supply', uco_supply)
write_tab(wb, 'uco_demand', uco_demand)
write_meta(wb, build_meta_rows(uco_supply, uco_demand),
           note_override={'non_bio_use': UCO_NONBIO_NOTE})
f = OUTDIR / "us_uco_supply_demand.xlsx"; wb.save(f); written.append(f)

# NASS staleness note override for the three frozen-supply commodities
nass_stale_override = {s: NOTES[s] + ' | ' + STALE_NASS for s in ('production', 'processing_use', 'stocks')}

# File 3: poultry_fat
wb = openpyxl.Workbook(); wb.remove(wb.active)
write_tab(wb, 'poultry_fat_supply', poultry_fat_supply)
write_tab(wb, 'poultry_fat_demand', poultry_fat_demand)
write_meta(wb, build_meta_rows(poultry_fat_supply, poultry_fat_demand),
           note_override=nass_stale_override)
f = OUTDIR / "us_poultry_fat_supply_demand.xlsx"; wb.save(f); written.append(f)

# File 4: white_grease (Choice White Grease)
wb = openpyxl.Workbook(); wb.remove(wb.active)
write_tab(wb, 'white_grease_supply', white_grease_supply)
write_tab(wb, 'white_grease_demand', white_grease_demand)
write_meta(wb, build_meta_rows(white_grease_supply, white_grease_demand),
           note_override=nass_stale_override)
f = OUTDIR / "us_white_grease_supply_demand.xlsx"; wb.save(f); written.append(f)

# File 5: yellow_grease (demand empty by design -- Amendment 1 folds YG biofuel use into UCO pool)
wb = openpyxl.Workbook(); wb.remove(wb.active)
write_tab(wb, 'yellow_grease_supply', yellow_grease_supply)
write_tab(wb, 'yellow_grease_demand', yellow_grease_demand)
yg_extra = []
if not yellow_grease_demand:
    yg_extra = [{'series': 'biofuel_use_total', 'source': 'ALLOCATOR_RAKED', 'unit': 'LB',
                 'vintage_set': '(none)',
                 'notes': 'EMPTY BY DESIGN: UCO Amendment 1 rules YG biofuel use = 0 (folded into the '
                          'UCO pool). No YG feedstock_code emerges from the raked allocator -- an empty '
                          'demand tab is expected, not an error.'}]
write_meta(wb, build_meta_rows(yellow_grease_supply, yellow_grease_demand),
           note_override=nass_stale_override, extra_meta=yg_extra)
f = OUTDIR / "us_yellow_grease_supply_demand.xlsx"; wb.save(f); written.append(f)

# File 6: dco (distillers corn oil)
wb = openpyxl.Workbook(); wb.remove(wb.active)
write_tab(wb, 'dco_supply', dco_supply)
write_tab(wb, 'dco_demand', dco_demand)
write_meta(wb, build_meta_rows(dco_supply, dco_demand))
f = OUTDIR / "us_dco_supply_demand.xlsx"; wb.save(f); written.append(f)

for f in written:
    print(f"wrote {f}")

# ---------------------------------------------------------------- verify
FILES = [
    ("us_tallow_supply_demand.xlsx",        'tallow_supply',        'tallow_demand',        'tallow_production', 4.1),
    ("us_uco_supply_demand.xlsx",           'uco_supply',           'uco_demand',           'uco_collection',    8.6),
    ("us_poultry_fat_supply_demand.xlsx",   'poultry_fat_supply',   'poultry_fat_demand',   'production',        0.2),
    ("us_white_grease_supply_demand.xlsx",  'white_grease_supply',  'white_grease_demand',  'production',        0.7),
    ("us_yellow_grease_supply_demand.xlsx", 'yellow_grease_supply', 'yellow_grease_demand', 'production',        0.0),
    ("us_dco_supply_demand.xlsx",           'dco_supply',           'dco_demand',           'dco_production',    4.3),
]


def cy_sum(ws, series, year=2024):
    tot = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(COLS, row))
        if d['series'] == series and d['marketing_year'] == year:
            tot += float(d['value'] or 0)
    return tot


for fname, sup_tab, dem_tab, main_series, exp_dem in FILES:
    path = OUTDIR / fname
    wb = openpyxl.load_workbook(path)
    print(f"\n===== {fname} =====")
    print("  sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        nrows = ws.max_row - 1
        if name == '_meta':
            print(f"  [{name}] {max(nrows, 0)} rows")
            continue
        header = [c.value for c in ws[1]]
        ok = header == COLS
        data = list(ws.iter_rows(min_row=2, values_only=True))
        series = sorted({row[2] for row in data if row[2] is not None})
        print(f"  [{name}] 13-col header {'OK' if ok else 'MISMATCH!'} | {len(data)} data rows | series: {series}")
        assert ok, f"HEADER MISMATCH on {name}: {header}"
    sup_prod = cy_sum(wb[sup_tab], main_series)
    dem_tot = cy_sum(wb[dem_tab], 'biofuel_use_total')
    print(f"  CY2024 supply '{main_series}': {sup_prod/1e9:.3f} B lb")
    print(f"  CY2024 demand 'biofuel_use_total': {dem_tot/1e9:.3f} B lb  (expect ~{exp_dem} B)")
    if wb[dem_tab].max_row <= 1:
        print(f"  NOTE: {dem_tab} is EMPTY (expected -- see _meta).")
