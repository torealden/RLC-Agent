"""Emit contract-compliant (flat_file_contract.md v1.1 LONG) SBO + canola oil supply/demand flat files.

Companion to write_fats_supply_flat_files.py (which does the 6 fats/greases). The vegetable oils have
a different balance shape than the fats — crush-derived crude production, an edible/industrial
disappearance split (SBO), and canola's single removal-for-processing line — so they get their own
writer, same 13-col LONG contract and same non-bio ruling (disappearance-based, 2026-07-13).

  models/Oilseeds/us_soybean_oil_supply_demand.xlsx -> soybean_oil_supply, soybean_oil_demand, _meta
  models/Oilseeds/us_canola_oil_supply_demand.xlsx  -> canola_oil_supply,  canola_oil_demand,  _meta

Supply: production (NASS crude oil), imports/exports (Census HS 1507=soy oil / 1514=canola oil, all
subcodes summed per flow -> kg x 2.20462 -> LB), ending stocks (NASS). Demand: biofuel (raked
allocator SBO / CO, one line per fuel type incl. coprocessing + SAF), food_use (SBO NASS refined
edible), aggregate non_biofuel_use, and its nonbiofuel_use_<end_use> component split via
reference.nonbio_enduse_shares (components sum to the total). Value = RAW pounds. Idempotent full
rewrite. See docs/specs/nonbio_demand_breakout_categories.md.
"""
import sys
from collections import defaultdict
from pathlib import Path
from statistics import mean

ROOT = Path(r"C:/dev/RLC-Agent"); sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
import openpyxl
from openpyxl.styles import Font
from src.services.database.db_config import get_connection

# Convention (Tore, 2026-07-16): flat files live in commodity\country folders. The United States
# oils file is the ACTIVE one Desktop links; the bare models/Oilseeds/*.xlsx copies are stale.
OUTDIR = ROOT / "models" / "Oilseeds" / "United States"
KG_TO_LB = 2.20462
COLS = ['commodity', 'class', 'series', 'marketing_year', 'period_type', 'period',
        'vintage', 'vintage_rank', 'value', 'unit', 'source', 'release_date', 'revision']

# Which published-forecast series belong to the SUPPLY tab vs the DEMAND tab. Used to route
# retained_forecast_series() rows to the correct flat-file tab (6d). Everything not here is demand.
SUPPLY_SERIES = {'production', 'imports', 'exports', 'stocks'}


def pm(month):
    return f"M{int(month):02d}"


def row(commodity, series, year, month, value, vintage, rank, source):
    return {'commodity': commodity, 'class': 'ALL', 'series': series, 'marketing_year': year,
            'period_type': 'cal_month', 'period': pm(month), 'vintage': vintage, 'vintage_rank': rank,
            'value': float(value), 'unit': 'LB', 'source': source, 'release_date': '', 'revision': ''}


def sort_key(r):
    return (r['series'], r['class'], r['marketing_year'], r['period'], r['vintage_rank'])


def dedupe(rows):
    seen = {}
    for r in rows:
        k = (r['commodity'], r['class'], r['series'], r['marketing_year'], r['period'], r['vintage_rank'])
        seen.setdefault(k, r)
    return list(seen.values())


def write_tab(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in sorted(dedupe(rows), key=sort_key):
        ws.append([r[c] for c in COLS])


# ---------------------------------------------------------------------------------------------
# WIDE RENDER (ruled 2026-07-21). The LONG tab is the contract; Excel cannot read it across a
# closed-workbook boundary, because SUMIFS/COUNTIFS/MAXIFS return #VALUE! against a closed source
# and only PLAIN CELL REFS read the link cache. That constraint is the only reason the balance-sheet
# workbook carried mirror tabs. So we render the balance-sheet shape HERE instead: months down,
# marketing years across, one block per series. The balance sheet then uses plain refs, works with
# this file closed, drops its mirror tabs, and stops paying for 6,720 SUMIFS over 8,000 rows.
#
# The grid is deliberately anchored to the EXISTING balance-sheet layout so refs map 1:1:
#   column B = MY 1990/91 (so col AI = 2023/24, exactly as the soyoil sheet already has it)
#   16 rows per block: title / units+MY header / Oct..Sep (12) / Marketing-year Total / blank
# Values are MILLION POUNDS to match the balance sheet's "(million pounds)" convention; the LONG
# tab stays raw LB. Row anchors are published in the _wide_index tab -- read that, don't count rows.
MY_ANCHOR = 1990          # marketing year sitting in column B
MY_COL0 = 2               # column B
BLOCK_ROWS = 16
MY_MONTHS = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
MONTH_LABEL = {10: 'Oct', 11: 'Nov', 12: 'Dec', 1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
               5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep'}
LB_PER_MIL = 1e6

# Explicit, stable block order. New series append at the end so existing anchors never move.
WIDE_ORDER = [
    'production', 'imports', 'exports', 'stocks',
    'biofuel_use_biodiesel', 'biofuel_use_renewable_diesel', 'biofuel_use_coprocessing',
    'biofuel_use_saf', 'biofuel_use_total',
    'non_biofuel_use',
    'nonbiofuel_use_salad_cooking_oil', 'nonbiofuel_use_baking_frying_fats',
    'nonbiofuel_use_margarine', 'nonbiofuel_use_other_edible',
    'nonbiofuel_use_paint_varnish', 'nonbiofuel_use_resins_plastics',
    'nonbiofuel_use_other_inedible',
    'food_use', 'industrial_use',
]


def my_of(year, month):
    """US oil marketing year: Oct-Sep, labelled by its start year."""
    return year if month >= 10 else year - 1


def best_by_period(rows):
    """One value per (series, MY, month): highest vintage_rank wins (actuals 90-95 beat forecast 40).
    dedupe() keeps rank as part of its key, so the same period CAN carry several vintages."""
    best = {}
    for r in rows:
        mo = int(r['period'][1:])
        k = (r['series'], my_of(r['marketing_year'], mo), mo)
        if k not in best or r['vintage_rank'] > best[k][0]:
            best[k] = (r['vintage_rank'], r['value'])
    return {k: v for k, (_, v) in best.items()}


def write_wide(wb, title, rows, commodity_label):
    data = best_by_period(rows)
    present = {s for s, _, _ in data}
    order = [s for s in WIDE_ORDER if s in present] + sorted(present - set(WIDE_ORDER))
    mys = sorted({my for _, my, _ in data})
    if not mys:
        return []
    lo, hi = min(mys), max(mys)
    if lo < MY_ANCHOR:
        print(f"  WARNING [{title}] MY {lo} precedes the {MY_ANCHOR} grid anchor -- "
              f"those years are NOT rendered wide. Re-anchor before backfilling that far.")
    ws = wb.create_sheet(title)
    ncol = MY_COL0 + (hi - MY_ANCHOR)
    index = []
    r = 1
    for series in order:
        ws.cell(r, 1, f"{commodity_label} {series.upper().replace('_', ' ')}").font = Font(bold=True)
        ws.cell(r + 1, 1, "(million pounds)").font = Font(italic=True)
        for my in range(MY_ANCHOR, hi + 1):
            ws.cell(r + 1, MY_COL0 + (my - MY_ANCHOR), f"{my}/{str(my + 1)[-2:]}").font = Font(bold=True)
        for i, mo in enumerate(MY_MONTHS):
            ws.cell(r + 2 + i, 1, MONTH_LABEL[mo])
            for my in range(MY_ANCHOR, hi + 1):
                v = data.get((series, my, mo))
                if v is not None:
                    ws.cell(r + 2 + i, MY_COL0 + (my - MY_ANCHOR), round(v / LB_PER_MIL, 4))
        ws.cell(r + 14, 1, "  Marketing-year Total").font = Font(bold=True)
        for my in range(MY_ANCHOR, hi + 1):
            c = ws.cell(r + 14, MY_COL0 + (my - MY_ANCHOR))
            col = c.column_letter
            c.value = f"=IF(COUNT({col}{r + 2}:{col}{r + 13})=0,\"\",SUM({col}{r + 2}:{col}{r + 13}))"
            c.font = Font(bold=True)
        index.append({'series': series, 'title_row': r, 'header_row': r + 1,
                      'first_month_row': r + 2, 'last_month_row': r + 13, 'total_row': r + 14})
        r += BLOCK_ROWS
    ws.freeze_panes = ws.cell(3, MY_COL0)
    ws.column_dimensions['A'].width = 26
    return [dict(i, tab=title, first_my_col='B', first_my=MY_ANCHOR, last_my=hi, ncol=ncol) for i in index]


def write_wide_index(wb, blocks):
    wi = wb.create_sheet("_wide_index")
    cols = ['tab', 'series', 'title_row', 'header_row', 'first_month_row', 'last_month_row',
            'total_row', 'first_my_col', 'first_my', 'last_my']
    wi.append(cols)
    for cell in wi[1]:
        cell.font = Font(bold=True)
    for b in blocks:
        wi.append([b.get(c, '') for c in cols])
    wi.append([])
    wi.append(['NOTE', 'Month rows run Oct->Sep. Column B = MY 1990/91, so col AI = MY 2023/24 '
                       '(matches the existing balance-sheet grid). Values are MILLION POUNDS; the '
                       'LONG tabs are raw LB. Point balance-sheet cells at the wide tabs with PLAIN '
                       'refs -- SUMIFS/COUNTIFS/MAXIFS cannot read a closed workbook.'])


def write_meta(wb, rows, notes):
    wm = wb.create_sheet("_meta")
    wm.append(['series', 'source', 'unit', 'last_updated', 'notes'])
    for cell in wm[1]:
        cell.font = Font(bold=True)
    seen = {}
    for r in rows:
        seen.setdefault(r['series'], r['source'])
    for series, source in seen.items():
        wm.append([series, source, 'LB', '', notes.get(series, '')])


NOTES = {
    'production': 'NASS crude oil production (oil_production_crude), raw lb; pre-2015 from ERS OCY '
                  '(identical series -- verified 1.000 on the 2015-18 overlap)',
    'stocks': 'TOTAL US stocks = NASS crude + once-refined (onsite & offsite). Crude+refined is '
              'the total for every vegetable oil; the crusher/refiner/offsite cuts are '
              'subcategories of crude, not extra pools. Pre-2015 from ERS OCY (ties exactly). '
              'MY2011/12-2014/15 modeled from ERS annual ending stocks -- no monthly source exists.',
    'imports': 'Census HS 1507 (soy) / 1514 (canola) imports, all subcodes summed, kg -> lb',
    'exports': 'Census HS 1507 (soy) / 1514 (canola) exports, all subcodes summed, kg -> lb',
    'biofuel_use_biodiesel': 'raked allocator ACTUALS (rank 95) through the last EIA month; forward = '
                             'MODEL_BASE forecast (rank 1) from biofuel_feedstock_use_forecast callable '
                             '(fuel-prod x trailing-12mo intensity), published via silver.soybean_oil_series.',
    'biofuel_use_renewable_diesel': 'raked allocator ACTUALS (rank 95) + forward MODEL_BASE forecast '
                                    '(rank 1), same callable. See biofuel_use_biodiesel.',
    'biofuel_use_total': 'raked allocator ACTUALS (rank 95) + forward MODEL_BASE forecast (rank 1) = '
                         'sum across fuel types. Band (value_low/value_high, in silver.soybean_oil_series) '
                         'is a SCENARIO interval from trailing intensity dispersion, NOT a confidence '
                         'level; the total band is the comonotone sum of the per-fuel bands.',
    'food_use': 'NASS refined edible use (oil_refined_edible_use)',
    'industrial_use': 'NASS refined inedible use (oil_refined_inedible_use)',
    'non_biofuel_use': 'aggregate non-biofuel use (closes the sheet). SBO = food+industrial '
                       '(disappearance); canola = removal_for_processing - biofuel. Ruled 2026-07-13.',
}


def nass_series(cur, commodity, attribute, out_series, out_commodity):
    # source IN NASS_* only — silver.monthly_realized also holds ABIOVE (Brazil) rows under
    # commodity='soybeans' with 0/NA US values that would otherwise clobber the real NASS figure.
    cur.execute("""SELECT calendar_year yr, month, realized_value v FROM silver.monthly_realized
                   WHERE commodity=%s AND attribute=%s AND realized_value IS NOT NULL
                     AND source LIKE 'NASS%%'
                   ORDER BY 1,2""", (commodity, attribute))
    return [row(out_commodity, out_series, r['yr'], r['month'], r['v'], 'NASS_FATS_OILS', 90, 'NASS')
            for r in cur.fetchall()]


def nass_total_stocks(cur, oil_name, out_commodity):
    """TOTAL US vegetable-oil stocks = CRUDE + ONCE REFINED (both onsite & offsite).

    Ruled by Tore 2026-07-21: "crude plus refined always equals total across all countries for
    vegetable oils." The crusher / refiner / offsite cuts NASS also publishes are SUBCATEGORIES of
    crude, not additional stock pools -- adding them would double-count.

    This bypasses silver.monthly_realized deliberately. Its 'oil_stocks' attribute is mapped to the
    ONCE REFINED series ALONE, which is roughly a sixth of the total (Oct 2015: refined 370.5 vs
    total 1,940.4 mil lb). That mis-mapping is why our stocks looked like a different concept from
    ERS's. Summing the two components here reproduces ERS EXACTLY -- 108 overlapping months, worst
    difference 0.01 mil lb -- confirming one concept, not two.

    ** silver.monthly_realized attribute 'oil_stocks' is still wrong at source. Anything else
       reading it is understating stocks by ~6x. Needs a collector/mapping fix separately. **

    Emits only months where BOTH components are present; a month with one side missing would
    silently understate the level (canola crude starts 2016-08 while refined starts 2015-05)."""
    cur.execute("""SELECT year yr, month mo,
             max(CASE WHEN short_desc ILIKE %s THEN value END) crude,
             max(CASE WHEN short_desc ILIKE %s THEN value END) refined
           FROM bronze.nass_processing
           WHERE short_desc ILIKE %s AND statisticcat_desc='STOCKS' AND month IS NOT NULL
           GROUP BY 1,2 ORDER BY 1,2""",
                (f'OIL, {oil_name}, ONSITE & OFFSITE, CRUDE - STOCKS%',
                 f'OIL, {oil_name}, ONSITE & OFFSITE, ONCE REFINED - STOCKS%',
                 f'OIL, {oil_name}, ONSITE & OFFSITE, %- STOCKS%'))
    out, skipped = [], 0
    for r in cur.fetchall():
        if r['crude'] is None or r['refined'] is None:
            skipped += 1
            continue
        out.append(row(out_commodity, 'stocks', r['yr'], r['mo'],
                       float(r['crude']) + float(r['refined']), 'NASS_FATS_OILS', 90,
                       'NASS crude + once-refined (onsite & offsite) = total US stocks'))
    if skipped:
        print(f"  {out_commodity} stocks: {skipped} months skipped (only one of crude/refined present)")
    return out


def ers_monthly(cur, commodity, attribute, out_series, out_commodity):
    """ERS Oil Crops Yearbook monthly, used to extend production/stocks BACK before NASS.

    NASS's monthly soybean crush report only starts May 2015 -- Census discontinued the CIR
    'Fats and Oils' program in July 2011 and nothing published the monthly crush series until NASS
    picked it up in 2015. ERS republishes the same numbers and reaches back to Oct 2007, so it is
    the only route to pre-2015 monthly history. Rank 85 (< NASS 90): where both exist NASS wins,
    and the wide render takes the highest rank per period."""
    cur.execute("""SELECT calendar_year yr, month, realized_value v FROM silver.monthly_realized
                   WHERE commodity=%s AND attribute=%s AND realized_value IS NOT NULL
                     AND source='ERS_OCY'
                   ORDER BY 1,2""", (commodity, attribute))
    return [row(out_commodity, out_series, r['yr'], r['month'], r['v'], 'ERS_OCY', 85,
                'ERS Oil Crops Yearbook monthly (pre-NASS history)') for r in cur.fetchall()]


def census_trade(cur, hs_prefix, out_commodity):
    # country_code='-' is the "TOTAL FOR ALL COUNTRIES" row. The table also carries OVERLAPPING
    # country-group aggregates (LAFTA, OECD, SOUTH AMERICA, USMCA, ...) — summing those double-counts
    # massively (soy-oil exports came out ~6x high). Use only the TOTAL row per subcode.
    cur.execute("""SELECT year yr, month, flow, sum(quantity) kg FROM bronze.census_trade
                   WHERE hs_code LIKE %s AND quantity IS NOT NULL AND country_code='-'
                   GROUP BY 1,2,3 ORDER BY 1,2""",
                (hs_prefix + '%',))
    out = []
    for r in cur.fetchall():
        series = 'imports' if str(r['flow']).lower().startswith('import') else 'exports'
        out.append(row(out_commodity, series, r['yr'], r['month'], float(r['kg']) * KG_TO_LB,
                       'CENSUS', 90, 'CENSUS_trade'))
    return out


def raked_demand(cur, feedstock_code, out_commodity):
    cur.execute("""SELECT extract(year from period)::int yr, extract(month from period)::int mo,
                     fuel_type, sum(raked_mil_lbs) mil FROM gold.bbd_feedstock_raked
                   WHERE run_day=(SELECT max(run_day) FROM gold.bbd_feedstock_raked)
                     AND feedstock_code=%s GROUP BY 1,2,3""", (feedstock_code,))
    rows, totals = [], {}
    for r in cur.fetchall():
        s = 'biofuel_use_' + str(r['fuel_type']).lower().replace(' ', '_')
        v = float(r['mil']) * 1e6
        rows.append(row(out_commodity, s, r['yr'], r['mo'], v, 'RAKED', 95, 'ALLOCATOR_RAKED'))
        totals[(r['yr'], r['mo'])] = totals.get((r['yr'], r['mo']), 0.0) + v
    for (yr, mo), tot in totals.items():
        rows.append(row(out_commodity, 'biofuel_use_total', yr, mo, tot, 'RAKED', 95, 'ALLOCATOR_RAKED'))
    return rows, totals


def by_period(rows, series):
    return {(r['marketing_year'], r['period']): r['value'] for r in rows if r['series'] == series}


def nonbio_components(cur, commodity, demand_rows):
    """Split each non_biofuel_use TOTAL into end-use component lines via the SEASONAL
    (calendar-month) shares in reference.nonbio_enduse_shares_monthly — each month uses its own
    share so the components breathe seasonally instead of sitting at a flat percentage. Falls back
    to the flat annual reference.nonbio_enduse_shares if a commodity has no monthly rows.
    Components sum to the total within each month (shares sum to 1.0), so the sheet still closes.
    Vintage flags a measured Census share vs a modeled/analog assumption."""
    cur.execute("""SELECT month, end_use, share_pct, measured
                   FROM reference.nonbio_enduse_shares_monthly WHERE commodity=%s""", (commodity,))
    monthly = defaultdict(list)
    for s in cur.fetchall():
        monthly[s['month']].append(s)
    if not monthly:  # fallback: flat annual shares across every month
        cur.execute("""SELECT end_use, share_pct, measured FROM reference.nonbio_enduse_shares
                       WHERE commodity=%s""", (commodity,))
        flat = cur.fetchall()
        monthly = {m: flat for m in range(1, 13)}
    out = []
    for r in demand_rows:
        if r['series'] != 'non_biofuel_use':
            continue
        tot = float(r['value'] or 0)
        mo = int(r['period'][1:])
        ms = monthly.get(mo, [])
        ssum = sum(float(s['share_pct']) for s in ms) or 1.0  # normalize -> exact closure
        for s in ms:
            vint = 'NONBIO_MEASURED' if s['measured'] else 'NONBIO_MODELED'
            src = ('Census 2006-2011 seasonal end-use share x non_biofuel_use' if s['measured']
                   else 'analog/modeled seasonal end-use share x non_biofuel_use')
            out.append(row(commodity, 'nonbiofuel_use_' + s['end_use'], r['marketing_year'],
                           mo, tot * float(s['share_pct']) / ssum, vint, r['vintage_rank'], src))
    return out


_MO = {'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6, 'july': 7,
       'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12}


def _my_start(y, m):
    return y if m >= 10 else y - 1   # US oil marketing year starts October


def sbo_supply_gapfill(cur, supply_rows):
    """Fill the MY2011/12-2014/15 monthly hole in soybean-oil PRODUCTION and STOCKS.

    Tore's method (2026-07-21): USDA gives the annual, we spread it seasonally for production,
    and we set stocks off the seasonal LEVEL pattern anchored to USDA's published annual ending
    stocks. The two need different treatment because production is a FLOW and stocks is a LEVEL:

      production  annual x mean(month share of MY total)      -- shares sum to 1
      stocks      published MY-end stocks x mean(month / Sep)  -- an index anchored so Sep lands
                                                                  exactly on the published number

    Donors are the complete marketing years either side of the hole (2007/08-2010/11 from ERS,
    2015/16-2018/19 from NASS). Flagged MODELED_GAPFILL, rank 60: the annual level is published,
    the monthly shape is inferred. It must never be mistaken for an actual."""
    cur.execute("""SELECT my_start_year,
             max(CASE WHEN attribute_desc='Production' THEN amount END) prod,
             max(CASE WHEN attribute_desc='Ending stocks' THEN amount END) endstk
           FROM bronze.ers_oil_crops_yearbook
           WHERE commodity ILIKE '%soybean oil%' AND table_number=5
             AND timeperiod_desc='MY Total' AND my_start_year BETWEEN 2011 AND 2014
           GROUP BY 1 ORDER BY 1""")
    annual = {r['my_start_year']: (r['prod'], r['endstk']) for r in cur.fetchall()}
    if not annual:
        return []

    def monthly(series):
        best = {}
        for r in supply_rows:
            if r['series'] != series:
                continue
            k = (r['marketing_year'], int(r['period'][1:]))
            if k not in best or r['vintage_rank'] > best[k][0]:
                best[k] = (r['vintage_rank'], float(r['value'] or 0))
        return {k: v for k, (_, v) in best.items()}

    prod_m, stk_m = monthly('production'), monthly('stocks')
    DONOR = set(range(2007, 2011)) | set(range(2015, 2019))   # complete MYs either side of the hole

    def donor_my(d):
        by = defaultdict(dict)
        for (y, m), v in d.items():
            by[_my_start(y, m)][m] = v
        return {my: mm for my, mm in by.items() if my in DONOR and len(mm) == 12}

    pd_, sd_ = donor_my(prod_m), donor_my(stk_m)
    if not pd_ or not sd_:
        print("  WARNING: no complete donor marketing years -- supply gap NOT filled")
        return []
    pshare = {m: mean([mm[m] / sum(mm.values()) for mm in pd_.values() if sum(mm.values()) > 0])
              for m in range(1, 13)}
    ptot = sum(pshare.values())
    pshare = {m: pshare[m] / ptot for m in pshare}
    sidx = {m: mean([mm[m] / mm[9] for mm in sd_.values() if mm.get(9)]) for m in range(1, 13)}

    out = []
    for my, (prod, endstk) in sorted(annual.items()):
        for m in range(1, 13):
            y = my if m >= 10 else my + 1
            if prod is not None:
                out.append(row('soybean_oil', 'production', y, m, float(prod) * 1e6 * pshare[m],
                               'MODELED_GAPFILL', 60,
                               'ERS annual production x donor-MY seasonal share (no monthly source exists)'))
            if endstk is not None:
                out.append(row('soybean_oil', 'stocks', y, m, float(endstk) * 1e6 * sidx[m],
                               'MODELED_GAPFILL', 60,
                               'ERS annual ending stocks x donor-MY seasonal level index'))
    print(f"  supply gap-fill: {len(out)} rows for MY{min(annual)}-MY{max(annual)} "
          f"(donors prod={sorted(pd_)}, stocks={sorted(sd_)})")
    return out


def sbo_nonbio_series(cur, supply_rows, bio_totals):
    """SBO non-biofuel use = the balance RESIDUAL after biofuel (ruled 2026-07-20;
    project_nonbio_residual_after_biofuel). Full history + independent 2-MY forecast, replacing the
    old 17-month NASS-edible series:
      2007/08-2024/09  ERS Oil Crops Yearbook monthly 'Total domestic use' - 'Domestic use, Biofuel'.
      2024/10-latest   our balance residual = production + imports - exports - dStocks - raked biofuel.
      forecast         rest of current MY + 2 full MYs: trailing-3-MY-avg annual x 5-yr monthly
                       seasonal share -- a mechanical, INDEPENDENT baseline (no analyst projections
                       seeded), so the gap vs Tore's judged view is the reconciliation signal.
    Returns non_biofuel_use rows (LB)."""
    cur.execute("""SELECT marketing_year, lower(timeperiod_desc) tp,
        max(CASE WHEN attribute_desc='Total domestic use' THEN amount END) tot,
        max(CASE WHEN attribute_desc='Domestic use, Biofuel' THEN amount END) biof
      FROM bronze.ers_oil_crops_yearbook WHERE commodity ILIKE '%soybean oil%'
      AND lower(timeperiod_desc) IN ('january','february','march','april','may','june','july',
                                     'august','september','october','november','december')
      GROUP BY 1,2""")
    nb = {}
    for r in cur.fetchall():
        if r['tot'] is None:
            continue
        m = _MO[r['tp']]; y0 = int(str(r['marketing_year'])[:4])
        nb[(y0 if m >= 10 else y0 + 1, m)] = (float(r['tot']) - float(r['biof'] or 0)) * 1000.0
    last_ers = max(nb) if nb else (2007, 10)

    def _bp(series):
        # highest vintage_rank wins -- supply_rows now carries BOTH NASS (90) and the ERS
        # pre-2015 backfill (85) for the same periods, and a plain dict comprehension would
        # silently let whichever was appended last overwrite the other.
        best = {}
        for r in supply_rows:
            if r['series'] != series:
                continue
            k = (r['marketing_year'], int(r['period'][1:]))
            if k not in best or r['vintage_rank'] > best[k][0]:
                best[k] = (r['vintage_rank'], float(r['value'] or 0))
        return {k: v for k, (_, v) in best.items()}
    prod, stk, imp, exp = _bp('production'), _bp('stocks'), _bp('imports'), _bp('exports')
    for k in sorted(set(prod) & set(bio_totals)):
        if k <= last_ers:
            continue
        y, m = k; pk = (y, m - 1) if m > 1 else (y - 1, 12)
        ds = stk.get(k, 0) - stk.get(pk, stk.get(k, 0))
        nb[k] = prod.get(k, 0) + imp.get(k, 0) - exp.get(k, 0) - ds - bio_totals.get(k, 0)
    # ---- THE 2011/12-2014/15 HOLE -------------------------------------------------------------
    # ERS Table 8 (monthly) simply has no rows for MY2011/12-2014/15, and that is NOT an ingest
    # failure: the US had no publisher for the monthly series. Census killed the CIR 'Fats and
    # Oils' program in Jul 2011 (bronze.census_cir_fats ends 2011-07) and NASS did not start the
    # monthly soybean crush report until May 2015. Four marketing years of monthly data do not
    # exist anywhere.
    #
    # ERS Table 5 (ANNUAL) is continuous across them, so we distribute the annual non-bio figure
    # (Total domestic use - Domestic use, Biofuel) across the 12 months on the average monthly
    # share of the complete MYs either side of the hole. Flagged MODELED_GAPFILL at rank 60 --
    # above the forecast (40) because the annual level is a real published number, below every
    # actual (85-95) because the monthly SHAPE is inferred. Never let this read as an actual.
    cur.execute("""SELECT my_start_year,
             max(CASE WHEN attribute_desc='Total domestic use' THEN amount END) tot,
             max(CASE WHEN attribute_desc='Domestic use, Biofuel' THEN amount END) biof
           FROM bronze.ers_oil_crops_yearbook
           WHERE commodity ILIKE '%soybean oil%' AND table_number=5
             AND timeperiod_desc='MY Total' AND my_start_year BETWEEN 2011 AND 2014
           GROUP BY 1 ORDER BY 1""")
    gap_annual = {r['my_start_year']: (float(r['tot']) - float(r['biof'] or 0)) * 1e6
                  for r in cur.fetchall() if r['tot'] is not None}
    if gap_annual:
        # seasonal shape from the complete MYs bracketing the hole
        gmy_tot, gmy_n = defaultdict(float), defaultdict(int)
        for (y, m), v in nb.items():
            gmy_tot[_my_start(y, m)] += v; gmy_n[_my_start(y, m)] += 1
        donors = [my for my in sorted(gmy_tot)
                  if gmy_n[my] == 12 and (2007 <= my <= 2010 or 2015 <= my <= 2018)]
        shape = defaultdict(list)
        for (y, m), v in nb.items():
            my = _my_start(y, m)
            if my in donors and gmy_tot[my] > 0:
                shape[m].append(v / gmy_tot[my])
        gs = {m: mean(shape[m]) for m in range(1, 13) if shape[m]}
        gtot = sum(gs.values())
        if gs and gtot > 0:
            gs = {m: gs[m] / gtot for m in gs}
            for my, annual in gap_annual.items():
                for m in range(1, 13):
                    y = my if m >= 10 else my + 1
                    nb.setdefault((y, m), annual * gs[m])
            print(f"  gap-filled MY{min(gap_annual)}/{str(min(gap_annual)+1)[-2:]}"
                  f"-MY{max(gap_annual)}/{str(max(gap_annual)+1)[-2:]} from ERS ANNUAL "
                  f"({len(gap_annual)*12} months) on the seasonal shape of {donors}")
    gapfill_keys = {(y, m) for my in gap_annual for m in range(1, 13)
                    for y in [my if m >= 10 else my + 1]}

    last_act = max(nb)

    myt, n_mo = defaultdict(float), defaultdict(int)
    for (y, m), v in nb.items():
        myt[_my_start(y, m)] += v; n_mo[_my_start(y, m)] += 1
    complete = [my for my in sorted(myt) if n_mo[my] == 12]
    seas_src = defaultdict(list)
    for (y, m), v in nb.items():
        my = _my_start(y, m)
        if my in complete[-5:] and myt[my] > 0:
            seas_src[m].append(v / myt[my])
    seas = {m: (mean(seas_src[m]) if seas_src[m] else 1 / 12) for m in range(1, 13)}
    ssum = sum(seas.values()) or 1.0
    seas = {m: seas[m] / ssum for m in seas}
    ann = mean(myt[my] for my in complete[-3:]) if complete else 12 * sum(nb.values()) / max(1, len(nb))

    out = [row('soybean_oil', 'non_biofuel_use', y, m, v, 'RESIDUAL_ACTUAL', 90,
               'ERS total dom - biofuel (07/08-09/24); balance residual thereafter')
           for (y, m), v in nb.items()]
    cur_my = _my_start(*last_act)
    for my in (cur_my, cur_my + 1, cur_my + 2):
        for m in list(range(10, 13)) + list(range(1, 10)):
            y = my if m >= 10 else my + 1
            if (y, m) not in nb:
                out.append(row('soybean_oil', 'non_biofuel_use', y, m, ann * seas[m],
                               'FORECAST_SEASONAL', 40, 'independent: trailing-3MY avg x 5yr seasonal'))
    return out


def canola_nonbio_series(cur, supply_rows, bio_totals):
    """Canola non-biofuel use, modeled to MATCH soybean oil's TWO-VINTAGE split (design D8 decision 2;
    Tore 2026-07-23: 'model canola exactly like soy'). Replaces the old single RESIDUAL(50) that
    collapsed actual-derived history and the forward projection into one vintage — the D8 finding was
    that soy separates them and canola did not.

      RESIDUAL_ACTUAL (90)  actual-input history: the supply residual over the biofuel frontier.
      FORECAST_SEASONAL(40)  forward projection: current-MY remainder + 2 full MYs, trailing-3-MY-avg
                             annual x 5-yr monthly seasonal share (soy's method, independent baseline).

    Canola non-bio is the supply RESIDUAL (production + imports - biofuel - exports, clamped >=0):
    NASS removal_for_processing misses the food disappearance of IMPORTED canola oil and canola is
    import-dominant, so residual is the honest measure (Tore's ruling). Unlike soy there is no ERS
    monthly total-domestic series and no 2011/12-2014/15 monthly hole, so the history is purely the
    supply residual — simpler than soy, same STRUCTURE.

    NOTE (design D8 decision 1 deferred): the forward vintage stays FORECAST_SEASONAL at rank 40 here,
    matching soy as it currently stands, rather than rolling into the 1-9 forecast band now. The D4
    hard gate (Tore 2026-07-23) requires every 1-9 forecast row to carry a value_low/value_high band,
    and soybean_oil/canola_oil have no silver.*_series table or band CHECK yet — so moving these into
    1-9 today would create UNBANDED band-rank rows, violating D4's invariant. At rank 40 vs a 1-9 rank
    the MAXIFS behavior is identical (forward periods carry only the forecast), so the roll is purely
    cosmetic and waits for the oils *_series migration. Gate beats parameter; deferral documented.

    bio_totals: {(year, month): biofuel_total_lb} (the raked biofuel actuals; caller's can_bio_tot).
    Returns non_biofuel_use rows (LB).
    """
    prod = by_period(supply_rows, 'production')
    imp  = by_period(supply_rows, 'imports')
    exp  = by_period(supply_rows, 'exports')
    bio_by_key = {(yr, pm(mo)): t for (yr, mo), t in bio_totals.items()}
    # ---- actual-input history: residual over the biofuel frontier (calendar (y, m) keys) ----
    nb = {}
    for (yr, per) in sorted(set(prod) & set(bio_by_key)):
        m = int(per[1:])
        nb[(yr, m)] = max(0.0, prod.get((yr, per), 0) + imp.get((yr, per), 0)
                               - bio_by_key[(yr, per)] - exp.get((yr, per), 0))
    out = [row('canola_oil', 'non_biofuel_use', y, m, v, 'RESIDUAL_ACTUAL', 90,
               'residual: production+imports-biofuel-exports')
           for (y, m), v in nb.items()]
    if not nb:
        return out
    # ---- forward seasonal projection (mirror of sbo_nonbio_series) ----
    last_act = max(nb)
    myt, n_mo = defaultdict(float), defaultdict(int)
    for (y, m), v in nb.items():
        myt[_my_start(y, m)] += v; n_mo[_my_start(y, m)] += 1
    complete = [my for my in sorted(myt) if n_mo[my] == 12]
    seas_src = defaultdict(list)
    for (y, m), v in nb.items():
        my = _my_start(y, m)
        if my in complete[-5:] and myt[my] > 0:
            seas_src[m].append(v / myt[my])
    seas = {m: (mean(seas_src[m]) if seas_src[m] else 1 / 12) for m in range(1, 13)}
    ssum = sum(seas.values()) or 1.0
    seas = {m: seas[m] / ssum for m in seas}
    ann = mean(myt[my] for my in complete[-3:]) if complete else 12 * sum(nb.values()) / max(1, len(nb))
    cur_my = _my_start(*last_act)
    for my in (cur_my, cur_my + 1, cur_my + 2):
        for m in list(range(10, 13)) + list(range(1, 10)):
            y = my if m >= 10 else my + 1
            if (y, m) not in nb:
                out.append(row('canola_oil', 'non_biofuel_use', y, m, ann * seas[m],
                               'FORECAST_SEASONAL', 40, 'independent: trailing-3MY avg x 5yr seasonal'))
    return out


def trade_forward_gap(supply_rows, commodity):
    """Fill imports/exports forward over the CURRENT partial marketing year only (last census actual
    -> Sep of that oil-MY), so the balance sheet's month-ending-stocks ROLL-FORWARD can compute.

    Why this is needed (6d, verified by recalc): the sheet's stocks block rolls
    `stock[m] = stock[m-1] + production + imports - exports - domestic_use`. For the current-MY tail
    (e.g. Jun-Sep 2026) the trade cells use `IF(wide="","",value)` -- an empty forward trade cell
    returns "" (empty STRING), and `... + "" - "" - ...` = #VALUE!, poisoning the whole forward chain.
    Production alone was not enough; trade must be non-blank too.

    Scope is deliberately just the current partial MY: future full MYs (AL/AM) fall back to the sheet's
    own ruled trade constants (imports 350 / exports 1250), so this does NOT override Tore's forward
    export assumption -- it only bridges the census reporting lag. PLACEHOLDER pending the trade matrix
    (Tore 2026-07-24: "our trade matrix will drive the export numbers"). Rank 40 (< census actual 90),
    seasonal = mean of the last 3 years' actuals for each calendar month.
    """
    out = []
    for series in ('imports', 'exports'):
        act = {(r['marketing_year'], int(r['period'][1:])): float(r['value'] or 0)
               for r in supply_rows if r['series'] == series and r['value'] is not None}
        if not act:
            continue
        last = max(act)
        cur_my = _my_start(*last)                 # oil MY (Oct-Sep) of the last actual
        my_end = (cur_my + 1, 9)                  # Sep of that MY = last forecast month
        # seasonal level = mean of the last 3 years' actual value for each calendar month
        by_month = defaultdict(list)
        for (y, m), v in sorted(act.items()):
            by_month[m].append((y, v))
        seas = {m: mean([v for _, v in vs[-3:]]) for m, vs in by_month.items()}
        # walk the gap months last+1 .. my_end
        y, m = last
        while True:
            m += 1
            if m == 13:
                y, m = y + 1, 1
            if (y, m) > my_end:
                break
            if (y, m) not in act and m in seas:
                out.append(row(commodity, series, y, m, seas[m], 'FORECAST_TRADE_PLACEHOLDER', 40,
                               'placeholder: 3yr seasonal mean, bridges census lag (pending trade matrix)'))
    if out:
        print(f"  {commodity} trade gap-fill: {len(out)} rows "
              f"({sorted(set(r['series'] for r in out))}) through current MY Sep")
    return out


def retained_forecast_series(cur, commodity):
    """Merge PUBLISHED book-(b) forecast rows for `commodity` from silver.<commodity>_series into the
    flat file (design forecast_layer_design_v1.md D1/D5). This is the seam that lets a D5 forecast
    callable's output reach the balance sheet: the callable writes rank 1-9 banded rows into the series
    table with a run_id, and the writer picks them up here so write_wide renders them into the forward
    columns of the demand_wide blocks -- closing the biofuel gap the raked ACTUALS leave open.

    Only rows whose run_id is retain=true in core.forecast_run are eligible (belt-and-suspenders: a
    retain=false scenario run never writes series rows, but the join guarantees no what-if leaks into a
    sheet -- and, per design D6, guarantees no book-(a) LLM row could either). Only rank 1-9 (the
    forecast band). The band (value_low/value_high) is NOT rendered wide -- the balance sheet reads the
    positional point value -- but the point flows through best_by_period exactly like any vintage: since
    the forecast (rank 1) sits BELOW the raked actuals (rank 95) and their periods are disjoint (forecast
    starts the month after the last actual), MAXIFS/best_by_period keeps the actual where it exists and
    the forecast forward, with an automatic handoff and no collision.

    Returns writer-row dicts (raw LB), or [] if the commodity has no *_series table yet (e.g. canola).
    """
    table = f"silver.{commodity}_series"
    cur.execute("SELECT to_regclass(%s) AS t", (table,))
    if cur.fetchone()['t'] is None:
        return []
    cur.execute(f"""
        SELECT s.series, s.marketing_year, s.period, s.vintage, s.vintage_rank, s.value, s.source
        FROM {table} s
        JOIN core.forecast_run fr ON fr.run_id = s.run_id
        WHERE fr.retain AND s.vintage_rank BETWEEN 1 AND 9 AND s.value IS NOT NULL
    """)
    out = []
    for r in cur.fetchall():
        m = int(str(r['period'])[1:])
        out.append(row(commodity, r['series'], r['marketing_year'], m, float(r['value']),
                       r['vintage'], r['vintage_rank'], r['source']))
    return out


with get_connection() as conn:
    cur = conn.cursor()

    # ---------------- SOYBEAN OIL ----------------
    sbo_supply = (nass_series(cur, 'soybeans', 'oil_production_crude', 'production', 'soybean_oil')
                  + nass_total_stocks(cur, 'SOYBEAN', 'soybean_oil')
                  # pre-2015 history: NASS's monthly report starts May 2015, ERS reaches to Oct 2007
                  + ers_monthly(cur, 'soybeans', 'oil_production_crude', 'production', 'soybean_oil')
                  # ERS IS the same concept once NASS crude+refined are summed (verified:
                  # 108 overlapping months, worst diff 0.01 mil lb), so it merges into 'stocks'
                  # as straight pre-2015 history at rank 85.
                  + ers_monthly(cur, 'soybeans', 'oil_stocks_sd', 'stocks', 'soybean_oil')
                  + census_trade(cur, '1507', 'soybean_oil'))
    # MY2011/12-2014/15: no monthly source exists anywhere (Census CIR died Jul 2011, NASS started
    # May 2015). Fill production/stocks from the ERS annual, clearly flagged as modeled.
    sbo_supply += sbo_supply_gapfill(cur, sbo_supply)
    # bridge the census reporting lag on trade so the sheet's stocks roll-forward can compute (6d)
    sbo_supply += trade_forward_gap(sbo_supply, 'soybean_oil')
    sbo_demand, sbo_bio_tot = raked_demand(cur, 'SBO', 'soybean_oil')
    sbo_food = nass_series(cur, 'soybeans', 'oil_refined_edible_use', 'food_use', 'soybean_oil')
    sbo_demand += sbo_food  # informational NASS refined-edible line (a subset of non-bio use)
    # non_biofuel_use = the balance RESIDUAL after biofuel is allocated by fuel, full history +
    # independent 2-MY forecast (ruled 2026-07-20; project_nonbio_residual_after_biofuel). Replaces
    # the old 17-month food-only proxy.
    sbo_demand += sbo_nonbio_series(cur, sbo_supply, sbo_bio_tot)
    # PUBLISHED forecast rows (book b) from silver.soybean_oil_series (rank 1-9, retain=true). Two
    # callables now write here: biofuel_feedstock_use_forecast (6c, DEMAND: biofuel_use_*) and
    # soybean_oil_production_forecast (6d, SUPPLY: production). Route each row to the tab whose SUMIFS
    # blocks read it -- a supply series merged into the demand tab (or vice-versa) would never be found
    # by the balance sheet. See design D5 + docs/specs/soyoil_supply_forecast_6d_findings.md.
    sbo_fc = retained_forecast_series(cur, 'soybean_oil')
    sbo_supply += [r for r in sbo_fc if r['series'] in SUPPLY_SERIES]
    sbo_demand += [r for r in sbo_fc if r['series'] not in SUPPLY_SERIES]

    # ---------------- CANOLA OIL ----------------
    can_supply = (nass_series(cur, 'canola', 'oil_production_crude', 'production', 'canola_oil')
                  + nass_total_stocks(cur, 'CANOLA', 'canola_oil')
                  + census_trade(cur, '1514', 'canola_oil'))
    can_supply += trade_forward_gap(can_supply, 'canola_oil')
    can_demand, can_bio_tot = raked_demand(cur, 'CO', 'canola_oil')
    # non_biofuel_use = the supply RESIDUAL, split into RESIDUAL_ACTUAL(90) history + FORECAST_SEASONAL(40)
    # forward — modeled exactly like soybean oil (D8 decision 2; was a single collapsed RESIDUAL=50).
    can_demand += canola_nonbio_series(cur, can_supply, can_bio_tot)

    # split each commodity's non_biofuel_use total into end-use component lines (shared shares)
    sbo_demand += nonbio_components(cur, 'soybean_oil', sbo_demand)
    can_demand += nonbio_components(cur, 'canola_oil', can_demand)

OUTDIR.mkdir(parents=True, exist_ok=True)
for fname, stab, dtab, supply, demand in [
    ('us_soybean_oil_supply_demand.xlsx', 'soybean_oil_supply', 'soybean_oil_demand', sbo_supply, sbo_demand),
    ('us_canola_oil_supply_demand.xlsx', 'canola_oil_supply', 'canola_oil_demand', can_supply, can_demand),
]:
    label = stab.replace('_supply', '').replace('_', ' ').upper()
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    write_tab(wb, stab, supply)
    write_tab(wb, dtab, demand)
    # wide render: what the balance sheet actually links to (plain refs, closed-workbook safe)
    blocks = (write_wide(wb, stab + '_wide', dedupe(supply), label)
              + write_wide(wb, dtab + '_wide', dedupe(demand), label))
    write_wide_index(wb, blocks)
    write_meta(wb, supply + demand, NOTES)
    wb.save(OUTDIR / fname)
    print(f"wrote {OUTDIR / fname}")
    # LOUD: biofuel_use is raked ACTUALS only -- it stops at the last EIA month while the
    # rest of the sheet forecasts forward. Forward biofuel cells are intentionally BLANK
    # (never 0). Announce the gap every run so no consumer silently reads the hole as zero.
    def _key(r):
        mo = int(str(r['period'])[1:])
        return my_of(r['marketing_year'], mo) * 100 + mo
    bio = [r for r in demand if str(r['series']).startswith('biofuel_use')]
    if bio:
        bl, dl = max(_key(r) for r in bio), max(_key(r) for r in demand)
        if bl < dl:
            print(f"  [BIOFUEL HOLE] {label}: biofuel_use ends MY{bl//100}/{str(bl//100+1)[-2:]} "
                  f"m{bl % 100:02d}; demand forecast runs to MY{dl//100}/{str(dl//100+1)[-2:]} "
                  f"m{dl % 100:02d}. Forward biofuel cells are BLANK, not 0 -- BBD/non-bio "
                  f"consumers must treat blank as MISSING FORECAST, never zero.")
    for tab, rows in [(stab, supply), (dtab, demand)]:
        ser = sorted(set(r['series'] for r in rows))
        print(f"  [{tab}] {len(dedupe(rows))} rows | series: {ser}")
    print(f"  [wide] {len(blocks)} blocks across {stab}_wide + {dtab}_wide "
          f"(16 rows each, col B = MY{MY_ANCHOR}/{str(MY_ANCHOR + 1)[-2:]}); anchors in _wide_index")
