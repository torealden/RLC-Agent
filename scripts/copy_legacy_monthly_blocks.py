"""
Copy legacy wld<crop>bal monthly blocks -> new per-country 3-tab workbooks, re-bucketed to US MY.

Generalized from the verified Australia-rapeseed trial (scripts/copy_legacy_monthly_blocks_trial.py).
This pass: RAPESEED (all country tabs). Soybean is heterogeneous (two meal-prod blocks, MEAL AND HULLS,
Brazil in separate tabs) -> handled once Tore rules on the wrinkles.

Per-country LOCAL marketing year is AUTO-DETECTED per block from its first month row (so Nov-Oct AU,
Jul-Jun EU, etc. all work). Each monthly cell is re-bucketed to the US MY (seed Sep-Aug, meal/oil
Oct-Sep) by true calendar month/year. Copies values only; clears template numbers but keeps formulas;
relabels labels only (artifacts -> country -> crop -> tonnes); never renames tabs (816 formulas).
"""
from __future__ import annotations
import os, shutil
import openpyxl
from openpyxl.descriptors import Integer
import openpyxl.chartsheet.custom as _cc
_cc.CustomChartsheetView.scale = Integer(allow_none=True)  # tolerate a bad chartsheet in wldsoybal

SRC_DIR = r"C:/Users/torem/RLC Dropbox/Tore Alden/Soybean Spreadsheets - Copy/"
TEMPLATE = r"models/Oilseeds/China/china_soybean_complex_bal_sheets.xlsx"
OUT_DIR = r"models/Oilseeds"

MONTHNUM = {m: i for i, m in enumerate(
    ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"], 1)}
for full, ab in [("january","jan"),("february","feb"),("march","mar"),("april","apr"),("june","jun"),
                 ("july","jul"),("august","aug"),("september","sep"),("october","oct"),
                 ("november","nov"),("december","dec")]:
    MONTHNUM[full] = MONTHNUM[ab]

# role -> (source header prefix using {CROP}, target tab, target header, product). product 'seed'=Sep-Aug.
# token is the SOURCE crop word ('RAPESEED', or 'CANOLA' for Canada); target headers are always CHINA SOYBEAN.
def rapeseed_blocks(C="RAPESEED"):
    return [
        (f"{C} IMPORTS", "soy_balance_sheet", "CHINA SOYBEAN IMPORTS", "seed"),
        (f"{C} EXPORTS", "soy_balance_sheet", "CHINA SOYBEAN EXPORTS", "seed"),
        (f"{C} CRUSH",   "soy_balance_sheet", "CHINA SOYBEAN CRUSH",   "seed"),
        (f"{C} MEAL PRODUCTION", "soymeal_balance_sheet", "CHINA SOYBEAN MEAL PRODUCTION", "meal"),
        (f"{C} MEAL IMPORTS",    "soymeal_balance_sheet", "CHINA SOYBEAN MEAL IMPORTS",    "meal"),
        (f"{C} MEAL EXPORTS",    "soymeal_balance_sheet", "CHINA SOYBEAN MEAL EXPORTS",    "meal"),
        (f"{C} MEAL END-OF-MONTH STOCKS", "soymeal_balance_sheet", "CHINA SOYBEAN MEAL MONTH-ENDING STOCKS", "meal"),
        (f"{C} OIL PRODUCTION", "soyoil_balance_sheet", "CHINA SOYBEAN OIL PRODUCTION", "oil"),
        (f"{C} OIL IMPORTS",    "soyoil_balance_sheet", "CHINA SOYBEAN OIL IMPORTS",    "oil"),
        (f"{C} OIL EXPORTS",    "soyoil_balance_sheet", "CHINA SOYBEAN OIL EXPORTS",    "oil"),
        (f"{C} OIL END-OF-MONTH STOCKS", "soyoil_balance_sheet", "CHINA SOYBEAN OIL MONTH-ENDING STOCKS", "oil"),
    ]

def us_my(month, cal_year, product):
    start = 9 if product == "seed" else 10
    return cal_year if month >= start else cal_year - 1

def _src_start_year(lab):
    yy = int(str(lab).strip()[:2]); return 1900 + yy if yy >= 90 else 2000 + yy
def _tgt_start_year(lab): return int(str(lab).strip()[:4])
def _month(v):
    if isinstance(v, str) and v.split():
        return MONTHNUM.get(v.split()[0].strip(",").lower())
    return None

def find_header(ws, text):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().upper().startswith(text.upper()):
            return r
    return None

def read_block(ws, header):
    """Return {(month, cal_year): value}. Local MY start month auto-detected from the block's first
    month row -> cal_year(month) = label-year if month>=start else +1."""
    h = find_header(ws, header)
    if h is None: return None
    yr_row, mstart = h + 1, h + 2
    m_start = _month(ws.cell(mstart, 1).value)
    if not m_start: return None
    col_year = {}
    for c in range(2, ws.max_column + 1):
        lab = ws.cell(yr_row, c).value
        if lab is not None:
            try: col_year[c] = _src_start_year(lab)
            except Exception: pass
    out = {}
    for rr in range(mstart, mstart + 12):
        mn = _month(ws.cell(rr, 1).value)
        if not mn: continue
        for c, yy in col_year.items():
            v = ws.cell(rr, c).value
            if v is None or v == "": continue
            cal = yy if mn >= m_start else yy + 1
            out[(mn, cal)] = v
    return out

def tab_year_cols(ws):
    out = {}
    for c in range(2, ws.max_column + 1):
        lab = ws.cell(3, c).value
        if isinstance(lab, str) and "/" in lab:
            try: out[_tgt_start_year(lab)] = c
            except Exception: pass
    return out

def clear_numbers_in_month_rows(ws):
    for r in range(1, ws.max_row + 1):
        if _month(ws.cell(r, 1).value):
            for c in range(2, ws.max_column + 1):
                if isinstance(ws.cell(r, c).value, (int, float)):
                    ws.cell(r, c).value = None

def target_month_rows(ws, header):
    h = find_header(ws, header)
    if h is None: return {}
    return {_month(ws.cell(rr, 1).value): rr for rr in range(h + 2, h + 14) if _month(ws.cell(rr, 1).value)}

def relabel_text(s, country, crop):
    for a, b in [("SCHINATAINABLE","SUSTAINABLE"),("CHINAE","USE"),
                 ("CHINA",country.upper()),("China",country),("Chinese",country+"n"),
                 ("SOYBEAN",crop.upper()),("Soybean",crop),("soybean",crop.lower()),
                 ("thousand short tons","thousand tonnes"),("short tons","tonnes"),("short ton","tonne")]:
        s = s.replace(a, b)
    return s

def copy_country(src, country, crop, blocks, out_path):
    shutil.copyfile(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path)
    ycols = {}
    for tab in ("soy_balance_sheet", "soymeal_balance_sheet", "soyoil_balance_sheet"):
        clear_numbers_in_month_rows(wb[tab]); ycols[tab] = tab_year_cols(wb[tab])
    written = 0; found = 0; missing = []
    for src_hdr, tab, tgt_hdr, product in blocks:
        block = read_block(src, src_hdr)
        if not block: missing.append(src_hdr); continue
        found += 1
        ws = wb[tab]; col = ycols[tab]; mrow = target_month_rows(ws, tgt_hdr)
        for (mn, cal), val in block.items():
            c = col.get(us_my(mn, cal, product)); r = mrow.get(mn)
            if c and r: ws.cell(r, c).value = val; written += 1
    # relabel
    for tab in wb.sheetnames:
        for row in wb[tab].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and not v.startswith("="):
                    nv = relabel_text(v, country, crop)
                    if nv != v: cell.value = nv
    wb.save(out_path)
    # verify
    chk = openpyxl.load_workbook(out_path)
    bad = 0
    for src_hdr, tab, tgt_hdr, product in blocks:
        block = read_block(src, src_hdr)
        if not block: continue
        ws = chk[tab]; col = tab_year_cols(ws); mrow = target_month_rows(ws, tgt_hdr)
        for (mn, cal), val in block.items():
            c = col.get(us_my(mn, cal, product)); r = mrow.get(mn)
            if c and r:
                got = ws.cell(r, c).value
                if got != val and not (isinstance(got,(int,float)) and isinstance(val,(int,float))
                                       and abs(got-val) <= 1e-6*max(1,abs(val))): bad += 1
    xref = sum(1 for tab in chk.sheetnames for row in chk[tab].iter_rows() for c in row
               if isinstance(c.value,str) and c.value.startswith("=") and "soy_balance_sheet" in c.value)
    cl = country.lower()
    dirty = sum(1 for tab in chk.sheetnames for row in chk[tab].iter_rows() for c in row
                if isinstance(c.value,str) and any(k in c.value.lower()
                for k in ("china","short ton","soybean","schina","chinae")) and cl != "china")
    return {"country": country, "written": written, "blocks_found": found,
            "blocks_missing": missing, "genuine_mismatch": bad, "formulas": xref, "dirty": dirty}

def run_rapeseed():
    from openpyxl.descriptors import Integer as _I
    wb = openpyxl.load_workbook(SRC_DIR + "wldrapbal.xlsx", data_only=True)
    tabs = [t for t in wb.sheetnames if t.endswith(" Rapeseed") and not t.startswith("World")]
    results = []
    for tab in tabs:
        country = tab[:-len(" Rapeseed")].strip()
        ws = wb[tab]
        token = "CANOLA" if find_header(ws, "CANOLA IMPORTS") else "RAPESEED"  # Canada uses CANOLA
        d = os.path.join(OUT_DIR, country); os.makedirs(d, exist_ok=True)
        out = os.path.join(d, f"{country.lower().replace(' ','_')}_rapeseed_complex_bal_sheets.xlsx")
        r = copy_country(ws, country, "Rapeseed", rapeseed_blocks(token), out)
        r["src_token"] = token
        results.append(r); print(r)
    return results

if __name__ == "__main__":
    run_rapeseed()
