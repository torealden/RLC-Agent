"""
Build the RLC balance-sheet model build tracker (checkbox-per-step, country x commodity).

Deliverable Tore asked for 2026-07-29: the checkbox tracker like model_files_punchlist.xlsx but
(a) with the fuller ~10-step process, (b) a country dimension, (c) one workbook per commodity complex
with a tab per country, and (d) ONE summary book with a tab per complex-group across all countries,
enumerating every country x commodity combination we will build a balance sheet for.

Scope is seeded from docs/specs/rlc_model_completion_masterplan_v1.md Part C (the master matrix).
Statuses start UNCHECKED -- this is the scaffold to check off, not a claim of current state.

STEPS: general across all commodities. Step 5 (processing linkage) only APPLIES to complexes that
convert (oilseed crush, palm, corn-oil derivation, fats/greases supply-derivation, biofuel conversion);
for raw grains / fuels / energy it renders 'n/a' and is excluded from the completion %. So a wheat sheet
completes at 9/9, an oilseed at 10/10 -- right boxes for the right commodity, one uniform column set so
the summary still rolls up cleanly.

Run:  python scripts/build_model_build_tracker.py
Out:  models/_build_tracker/RLC_Model_Build_Tracker_SUMMARY.xlsx
      models/_build_tracker/by_commodity/<Commodity>.xlsx   (one per commodity complex)
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join("models", "_build_tracker")
BY_COMMODITY_DIR = os.path.join(OUT_DIR, "by_commodity")

# ---------------------------------------------------------------------------------------------------
# The 10 steps. `applies` = 'all' or a set of complex-types that get the box (else 'n/a').
# ---------------------------------------------------------------------------------------------------
PROCESSING_TYPES = {"oilseed", "palm", "corn_oil", "fat_grease", "biofuel"}
STEPS = [
    ("1. Template",        "all"),
    ("2. Sources/API",     "all"),
    ("3. Flat file",       "all"),
    ("4. Linked",          "all"),
    ("5. Processing link",  PROCESSING_TYPES),   # crush / derivation / conversion; n/a for grains & fuels
    ("6. Tie-out",         "all"),
    ("7. Trade loop",      "all"),
    ("8. Forecast",        "all"),
    ("9. Verified",        "all"),
    ("10. Live test",      "all"),
]

# products per complex type (rows within a country tab)
PRODUCTS = {
    "oilseed":   ["Seed", "Meal", "Oil"],
    "palm":      ["CPO", "PKO", "PK Meal"],
    "corn_oil":  ["Oil"],
    "grain":     ["S&D"],
    "fat_grease":["Supply"],
    "biofuel":   ["S&D"],
    "fuel":      ["S&D"],
    "energy":    ["S&D"],
}
# NOTE (Tore 2026-07-29): FULL complex for EVERY country we cover, regardless of tier or client scope.
# Seed -> Meal + Oil via crush; one feeds the other, so a reduced importer set would miss the complete
# S&D picture. No tier-based product reduction anywhere.

# ---------------------------------------------------------------------------------------------------
# The master matrix (from masterplan Part C). commodity -> (group, ctype, [(country, tier)]).
# Tier A = price-setting producer (full set) · B = swing importer (reduced) · C = world rollup (auto) ·
# D = scenario stub. Seed scope -- Tore prunes/extends.
# ---------------------------------------------------------------------------------------------------
A, B, C, D = "A", "B", "C", "D"
MATRIX = {
    # ---- Oilseeds & Crush Products ----
    "Soybean":   ("Oilseeds & Crush Products", "oilseed",
                  [("US", A), ("Brazil", A), ("Argentina", A), ("Paraguay", D), ("Uruguay", D),
                   ("China", B), ("India", B), ("EU", B)]),
    "Rapeseed-Canola": ("Oilseeds & Crush Products", "oilseed",
                  [("EU", A), ("Canada", A), ("Australia", A), ("Russia", D), ("Ukraine", D),
                   ("China", B), ("Japan", B)]),
    "Sunflower": ("Oilseeds & Crush Products", "oilseed",
                  [("Ukraine", A), ("Russia", A), ("Argentina", A), ("EU", B), ("India", B),
                   ("Turkey", B), ("Colombia", D)]),
    "Cottonseed":("Oilseeds & Crush Products", "oilseed",
                  [("US", A), ("China", A), ("India", A), ("Pakistan", A), ("Brazil", A)]),
    "Peanut":    ("Oilseeds & Crush Products", "oilseed",
                  [("US", A), ("China", A), ("India", A), ("Argentina", A)]),
    "Flaxseed":  ("Oilseeds & Crush Products", "oilseed", [("US", A), ("Canada", A)]),
    "Safflower": ("Oilseeds & Crush Products", "oilseed", [("US", A)]),
    # ---- Other Oils & Fats ----
    "Palm":      ("Other Oils & Fats", "palm",
                  [("Malaysia", A), ("Indonesia", A), ("Colombia", D), ("Guatemala", D), ("Mexico", D),
                   ("China", B), ("India", B), ("EU", B)]),
    "Coconut":   ("Other Oils & Fats", "palm",
                  [("Philippines", A), ("Indonesia", A), ("India", A)]),
    "Corn Oil":  ("Other Oils & Fats", "corn_oil", [("US", A), ("Brazil", A), ("Mexico", D)]),
    "Tallow":    ("Other Oils & Fats", "fat_grease",
                  [("US", A), ("EU", B), ("Canada", B), ("Brazil", B), ("Argentina", B)]),
    "UCO-Yellow Grease": ("Other Oils & Fats", "fat_grease",
                  [("US", A), ("China", A), ("EU", B)]),
    "Choice White Grease": ("Other Oils & Fats", "fat_grease", [("US", A)]),
    "Lard":      ("Other Oils & Fats", "fat_grease", [("US", A), ("China", B), ("EU", B)]),
    "Poultry Fat": ("Other Oils & Fats", "fat_grease", [("US", A)]),
    # ---- Food Grains ----
    "Wheat":     ("Food Grains", "grain",
                  [("US", A), ("Russia", A), ("EU", A), ("Canada", A), ("Australia", A),
                   ("Argentina", A), ("Ukraine", A), ("Egypt", B), ("Mexico", B), ("Indonesia", B)]),
    "Rice":      ("Food Grains", "grain",
                  [("India", A), ("Thailand", A), ("Vietnam", A), ("US", A), ("Pakistan", A)]),
    # ---- Feed Grains ----
    "Corn":      ("Feed Grains", "grain",
                  [("US", A), ("Brazil", A), ("Argentina", A), ("Ukraine", A),
                   ("China", B), ("EU", B), ("Mexico", B), ("Japan", B)]),
    "Sorghum":   ("Feed Grains", "grain", [("US", A), ("Argentina", A), ("Australia", A), ("China", B)]),
    "Barley":    ("Feed Grains", "grain",
                  [("US", A), ("EU", A), ("Australia", A), ("Argentina", A), ("Russia", A), ("Canada", A)]),
    "Oats":      ("Feed Grains", "grain", [("US", A), ("Canada", A), ("EU", A)]),
    "Rye":       ("Feed Grains", "grain", [("EU", A), ("US", A)]),
    # ---- Biofuels ----
    "Ethanol":   ("Biofuels", "biofuel", [("US", A), ("Brazil", A), ("EU", B), ("China", B)]),
    "Biodiesel": ("Biofuels", "biofuel",
                  [("US", A), ("EU", A), ("Brazil", A), ("Argentina", A), ("Indonesia", A), ("Canada", B)]),
    "Renewable Diesel": ("Biofuels", "biofuel", [("US", A), ("EU", B), ("Canada", B), ("Singapore", B)]),
    "Sustainable Aviation Fuel": ("Biofuels", "biofuel", [("US", A), ("EU", A)]),
    # ---- Refined Fuels ----
    "Gasoline":  ("Refined Fuels", "fuel", [("US", A)]),
    "Diesel":    ("Refined Fuels", "fuel", [("US", A)]),
    "Jet Fuel":  ("Refined Fuels", "fuel", [("US", A)]),
    "Bunker Fuel":("Refined Fuels", "fuel", [("US", A)]),
    "Heating Oil":("Refined Fuels", "fuel", [("US", A)]),
    "Gas Oil":   ("Refined Fuels", "fuel", [("US", A)]),
    # ---- Energy ----
    "Natural Gas":("Energy", "energy", [("US", A)]),
    "Electricity":("Energy", "energy", [("US", A)]),
}

GROUP_ORDER = ["Oilseeds & Crush Products", "Other Oils & Fats", "Food Grains", "Feed Grains",
               "Biofuels", "Refined Fuels", "Energy"]

# ---- styling ----
HDR = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill("solid", fgColor="3C7D22")           # RLC internal green
GRP_FILL = PatternFill("solid", fgColor="E2EFDA")
NA_FILL = PatternFill("solid", fgColor="D9D9D9")
TITLE = Font(bold=True, size=13, color="3C7D22")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def tcount(rng: str) -> str:
    """Excel expr counting 'checked' cells in rng, robust to BOTH a native-checkbox boolean TRUE and a
    data-validation text 'TRUE' (and ignoring blanks / FALSE / 'n/a'). Avoids the COUNTIF(,TRUE())
    boolean-vs-text trap."""
    return f'SUMPRODUCT(--(UPPER({rng}&"")="TRUE"))'


def step_applies(ctype: str, applies) -> bool:
    return applies == "all" or ctype in applies


def n_applicable(ctype: str) -> int:
    return sum(1 for _, ap in STEPS if step_applies(ctype, ap))


def _hdr_row(ws, row, cols):
    for j, name in enumerate(cols, 1):
        c = ws.cell(row, j, name); c.font = HDR; c.fill = HDR_FILL; c.alignment = WRAP; c.border = BORDER


def _bool_dv(ws):
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    return dv


def build_commodity_workbook(commodity, group, ctype, countries):
    """One workbook: a tab per country, product rows x 10 step checkboxes + rollup, and a Dashboard."""
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash["A1"] = f"{commodity.upper()} — Build Tracker"; dash["A1"].font = TITLE
    dash["A2"] = f"Group: {group}   ·   complex type: {ctype}   ·   step 5 processing: " \
                 f"{'applies' if ctype in PROCESSING_TYPES else 'n/a for this complex'}"
    dash["A2"].font = Font(italic=True, size=9, color="808080")
    dcols = ["Country", "Tier"] + [s for s, _ in STEPS] + ["Complete", "Done", "%"]
    _hdr_row(dash, 4, dcols)

    napp = n_applicable(ctype)
    for i, (country, tier) in enumerate(countries):
        r = 5 + i
        dash.cell(r, 1, country).border = BORDER
        dash.cell(r, 2, tier).alignment = CENTER
        # per-country stage box = flips when that step is done for ALL products on the country tab
        tab = country
        for j, (sname, applies) in enumerate(STEPS):
            col = 3 + j
            cell = dash.cell(r, col); cell.alignment = CENTER; cell.border = BORDER
            if not step_applies(ctype, applies):
                cell.value = "n/a"; cell.fill = NA_FILL; cell.font = Font(color="808080", size=9)
                continue
            L = get_column_letter(col)
            # products live in rows 4..(3+nprod) on the country tab under the same column
            nprod = len(_products_for(ctype, tier))
            cell.value = f'=IF({tcount(chr(39)+tab+chr(39)+"!"+L+"4:"+L+str(3+nprod))}={nprod},"☑","☐")'
        # Complete / Done / %
        first = get_column_letter(3); last = get_column_letter(2 + len(STEPS))
        done = dash.cell(r, 3 + len(STEPS))
        done.value = f'=COUNTIF({first}{r}:{last}{r},"☑")'; done.alignment = CENTER
        comp = dash.cell(r, 3 + len(STEPS) - 0)  # placeholder fixed below
        # columns: after 10 steps -> Complete, Done, %
        c_complete = 3 + len(STEPS)
        c_done = c_complete + 1
        c_pct = c_complete + 2
        dash.cell(r, c_complete).value = f'=IF({first}{r}:{last}{r}<>"",0,0)'  # replaced below
        # simpler: Complete = all applicable boxes are ☑
        dash.cell(r, c_complete).value = f'=IF(COUNTIF({first}{r}:{last}{r},"☑")={napp},"✅","")'
        dash.cell(r, c_complete).alignment = CENTER
        dash.cell(r, c_done).value = f'=COUNTIF({first}{r}:{last}{r},"☑")&"/"&{napp}'
        dash.cell(r, c_done).alignment = CENTER
        dash.cell(r, c_pct).value = f'=COUNTIF({first}{r}:{last}{r},"☑")/{napp}'
        dash.cell(r, c_pct).number_format = "0%"; dash.cell(r, c_pct).alignment = CENTER
        _build_country_tab(wb, country, tier, commodity, ctype)

    dash.column_dimensions["A"].width = 16
    for col in range(2, len(dcols) + 1):
        dash.column_dimensions[get_column_letter(col)].width = 11
    dash.freeze_panes = "A5"
    os.makedirs(BY_COMMODITY_DIR, exist_ok=True)
    safe = commodity.replace("/", "-").replace(" ", "_")
    wb.save(os.path.join(BY_COMMODITY_DIR, f"{safe}.xlsx"))


def _products_for(ctype, tier):
    # Full complex for every country (see NOTE above) -- tier no longer reduces the product set.
    return PRODUCTS[ctype]


def _build_country_tab(wb, country, tier, commodity, ctype):
    ws = wb.create_sheet(title=country[:31])
    ws["A1"] = f"{commodity} — {country}  (Tier {tier})"; ws["A1"].font = Font(bold=True, size=11)
    ws["A2"] = "Check each step TRUE as it's completed (dropdown or type TRUE). n/a = not applicable."
    ws["A2"].font = Font(italic=True, size=8, color="808080")
    cols = ["Product"] + [s for s, _ in STEPS] + ["Complete", "%"]
    _hdr_row(ws, 3, cols)
    dv = _bool_dv(ws)
    prods = _products_for(ctype, tier)
    napp = n_applicable(ctype)
    for i, prod in enumerate(prods):
        r = 4 + i
        ws.cell(r, 1, prod).border = BORDER
        for j, (sname, applies) in enumerate(STEPS):
            col = 2 + j
            cell = ws.cell(r, col); cell.alignment = CENTER; cell.border = BORDER
            if not step_applies(ctype, applies):
                cell.value = "n/a"; cell.fill = NA_FILL; cell.font = Font(color="808080", size=9)
            else:
                dv.add(cell)  # TRUE/FALSE dropdown; blank = unchecked
        first = get_column_letter(2); last = get_column_letter(1 + len(STEPS))
        cc = 2 + len(STEPS); pc = cc + 1
        ws.cell(r, cc).value = f'=IF({tcount(f"{first}{r}:{last}{r}")}={napp},"☑","☐")'
        ws.cell(r, cc).alignment = CENTER
        ws.cell(r, pc).value = f'={tcount(f"{first}{r}:{last}{r}")}/{napp}'
        ws.cell(r, pc).number_format = "0%"; ws.cell(r, pc).alignment = CENTER
    ws.column_dimensions["A"].width = 14
    for col in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.freeze_panes = "B4"


def build_summary_workbook():
    """One book: Master Matrix (every combo) + a tab per complex-group + a Dashboard rollup."""
    wb = Workbook()
    # ---- Master Matrix ----
    mm = wb.active; mm.title = "Master Matrix"
    mm["A1"] = "RLC MODEL BUILD — MASTER MATRIX (every country × commodity)"; mm["A1"].font = TITLE
    mm["A2"] = ("One row per combination we will build a balance sheet for. Seeded from masterplan "
                "Part C — statuses start blank; fill the per-commodity workbooks, this rolls up. "
                "Tier A=price-setter · B=importer · C=world rollup (auto) · D=scenario stub.")
    mm["A2"].font = Font(italic=True, size=9, color="808080")
    cols = ["Model ID", "Group", "Commodity", "Country", "Tier"] + [s for s, _ in STEPS] + ["Complete", "Done", "%"]
    _hdr_row(mm, 4, cols)
    r = 5
    rows_by_group: dict[str, list[int]] = {g: [] for g in GROUP_ORDER}
    for commodity, (group, ctype, countries) in MATRIX.items():
        napp = n_applicable(ctype)
        for country, tier in countries:
            mm.cell(r, 1, f"{_cc(country)}-{_cc(commodity)}").border = BORDER
            mm.cell(r, 2, group); mm.cell(r, 3, commodity); mm.cell(r, 4, country)
            mm.cell(r, 5, tier).alignment = CENTER
            dv = _bool_dv(mm)
            for j, (sname, applies) in enumerate(STEPS):
                col = 6 + j
                cell = mm.cell(r, col); cell.alignment = CENTER; cell.border = BORDER
                if not step_applies(ctype, applies):
                    cell.value = "n/a"; cell.fill = NA_FILL; cell.font = Font(color="808080", size=9)
                else:
                    dv.add(cell)
            first = get_column_letter(6); last = get_column_letter(5 + len(STEPS))
            cc = 6 + len(STEPS); dc = cc + 1; pc = cc + 2
            mm.cell(r, cc).value = f'=IF({tcount(f"{first}{r}:{last}{r}")}={napp},"✅","")'
            mm.cell(r, cc).alignment = CENTER
            mm.cell(r, dc).value = f'={tcount(f"{first}{r}:{last}{r}")}&"/"&{napp}'
            mm.cell(r, dc).alignment = CENTER
            mm.cell(r, pc).value = f'={tcount(f"{first}{r}:{last}{r}")}/{napp}'
            mm.cell(r, pc).number_format = "0%"; mm.cell(r, pc).alignment = CENTER
            rows_by_group[group].append(r)
            r += 1
    mm.column_dimensions["A"].width = 16
    mm.column_dimensions["B"].width = 22; mm.column_dimensions["C"].width = 18; mm.column_dimensions["D"].width = 13
    for col in range(5, len(cols) + 1):
        mm.column_dimensions[get_column_letter(col)].width = 8.5
    mm.freeze_panes = "A5"
    total = r - 5

    # ---- one tab per complex-group (mirrors the matrix rows for that group) ----
    for group in GROUP_ORDER:
        gs = wb.create_sheet(title=group[:31])
        gs["A1"] = f"{group} — build status across all countries"; gs["A1"].font = TITLE
        gcols = ["Commodity", "Country", "Tier"] + [s for s, _ in STEPS] + ["Complete", "%"]
        _hdr_row(gs, 3, gcols)
        gr = 4
        for commodity, (g, ctype, countries) in MATRIX.items():
            if g != group:
                continue
            napp = n_applicable(ctype)
            for country, tier in countries:
                gs.cell(gr, 1, commodity).border = BORDER
                gs.cell(gr, 2, country); gs.cell(gr, 3, tier).alignment = CENTER
                dv = _bool_dv(gs)
                for j, (sname, applies) in enumerate(STEPS):
                    col = 4 + j
                    cell = gs.cell(gr, col); cell.alignment = CENTER; cell.border = BORDER
                    if not step_applies(ctype, applies):
                        cell.value = "n/a"; cell.fill = NA_FILL; cell.font = Font(color="808080", size=9)
                    else:
                        dv.add(cell)
                first = get_column_letter(4); last = get_column_letter(3 + len(STEPS))
                cc = 4 + len(STEPS); pc = cc + 1
                gs.cell(gr, cc).value = f'=IF({tcount(f"{first}{gr}:{last}{gr}")}={napp},"✅","")'
                gs.cell(gr, cc).alignment = CENTER
                gs.cell(gr, pc).value = f'={tcount(f"{first}{gr}:{last}{gr}")}/{napp}'
                gs.cell(gr, pc).number_format = "0%"; gs.cell(gr, pc).alignment = CENTER
                gr += 1
        gs.column_dimensions["A"].width = 18; gs.column_dimensions["B"].width = 13
        for col in range(3, len(gcols) + 1):
            gs.column_dimensions[get_column_letter(col)].width = 9
        gs.freeze_panes = "A4"

    # ---- Dashboard rollup by group ----
    db = wb.create_sheet(title="Dashboard", index=0)
    db["A1"] = "RLC MODEL BUILD — COMPLETION DASHBOARD"; db["A1"].font = TITLE
    db["A2"] = "Combination counts by complex-group. Fill the per-commodity workbooks; the Master Matrix is the row-level truth."
    db["A2"].font = Font(italic=True, size=9, color="808080")
    _hdr_row(db, 4, ["Complex Group", "Combinations", "Tier A", "Tier B", "Tier C/D"])
    dr = 5
    grand = 0
    for group in GROUP_ORDER:
        combos = [(cty, t) for cm, (g, ct, ctys) in MATRIX.items() if g == group for cty, t in ctys]
        nA = sum(1 for _, t in combos if t == A); nB = sum(1 for _, t in combos if t == B)
        nCD = sum(1 for _, t in combos if t in (C, D))
        db.cell(dr, 1, group).border = BORDER
        db.cell(dr, 2, len(combos)).alignment = CENTER
        db.cell(dr, 3, nA).alignment = CENTER; db.cell(dr, 4, nB).alignment = CENTER
        db.cell(dr, 5, nCD).alignment = CENTER
        grand += len(combos); dr += 1
    db.cell(dr, 1, "TOTAL").font = Font(bold=True)
    db.cell(dr, 2, grand).font = Font(bold=True); db.cell(dr, 2).alignment = CENTER
    db.column_dimensions["A"].width = 26
    for col in "BCDE":
        db.column_dimensions[col].width = 14

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(os.path.join(OUT_DIR, "RLC_Model_Build_Tracker_SUMMARY.xlsx"))
    return total


def _cc(name):  # short code for model_id
    return "".join(w[0] for w in name.replace("-", " ").split())[:4].upper() if " " in name or "-" in name else name[:4].upper()


def main():
    total = build_summary_workbook()
    for commodity, (group, ctype, countries) in MATRIX.items():
        build_commodity_workbook(commodity, group, ctype, countries)
    print(f"Summary + {len(MATRIX)} commodity workbooks written to {OUT_DIR}")
    print(f"Total country×commodity combinations tracked: {total}")


if __name__ == "__main__":
    main()
