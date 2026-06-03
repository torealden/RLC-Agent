"""Generic per-commodity production workbook builder.

One file per commodity, per-statistic tabs (area_planted, area_harvested,
production, yield), vintage columns (PP / Acreage / Aug-Nov / Final).

Per Tore (2026-05-30, 2026-05-31):
  - Original-release values for vintage columns (NASS preserves the value
    even when load_time was reset in a 2018-01-23 bulk migration)
  - Final = earliest 'YEAR' release with rel.year >= crop_year+1
    (excludes Census-of-Ag intra-cycle sub-aggregate republications)
  - National only in v1; states + ag districts come later

Output:
  models/Oilseeds/us_soybean_production.xlsx
  models/Oilseeds/us_canola_production.xlsx
  models/Oilseeds/us_sunflower_production.xlsx
  models/Oilseeds/us_peanut_production.xlsx
  models/Feed Grains/us_corn_production.xlsx
  models/Feed Grains/us_sorghum_production.xlsx
  models/Food Grains/us_wheat_production.xlsx
  models/Cotton/us_cotton_production.xlsx

cottonseed is derived from cotton (×~1.6 kg-per-cwt-lint); separate handling.
flaxseed / safflower not yet in NASS collector COMMODITIES — TODO.
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / '.env')

from src.services.database.db_config import get_connection

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
log = logging.getLogger('build_production_v2')


# ============================================================
# Per-commodity config
# ============================================================
# Each entry maps statistic -> (class_db, short_desc).
# 'class_db' = the silver.crop_production.class value to filter by.
# Different stats can use different classes — corn/sorghum AREA PLANTED is
# at all_classes level, but AREA HARVESTED/yield/production use the 'grain'
# class because NASS reports them per util_practice.
COMMODITY_CONFIG = {
    'soybeans': {
        'file': 'models/Oilseeds/us_soybean_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'SOYBEANS - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'SOYBEANS - ACRES HARVESTED'),
            'yield':          ('all_classes', 'SOYBEANS - YIELD, MEASURED IN BU / ACRE'),
            'production':     ('all_classes', 'SOYBEANS - PRODUCTION, MEASURED IN BU'),
        },
    },
    'canola': {
        'file': 'models/Oilseeds/us_canola_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'CANOLA - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'CANOLA - ACRES HARVESTED'),
            'yield':          ('all_classes', 'CANOLA - YIELD, MEASURED IN LB / ACRE'),
            'production':     ('all_classes', 'CANOLA - PRODUCTION, MEASURED IN LB'),
        },
    },
    'sunflower': {
        'file': 'models/Oilseeds/us_sunflower_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'SUNFLOWER - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'SUNFLOWER - ACRES HARVESTED'),
            'yield':          ('all_classes', 'SUNFLOWER - YIELD, MEASURED IN LB / ACRE'),
            'production':     ('all_classes', 'SUNFLOWER - PRODUCTION, MEASURED IN LB'),
        },
    },
    'peanut': {
        'file': 'models/Oilseeds/us_peanut_production.xlsx',
        'commodity_db': 'peanuts',
        'stats': {
            'area_planted':   ('all_classes', 'PEANUTS - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'PEANUTS - ACRES HARVESTED'),
            'yield':          ('all_classes', 'PEANUTS - YIELD, MEASURED IN LB / ACRE'),
            'production':     ('all_classes', 'PEANUTS - PRODUCTION, MEASURED IN LB'),
        },
    },
    'corn': {
        'file': 'models/Feed Grains/us_corn_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'CORN - ACRES PLANTED'),
            'area_harvested': ('grain',       'CORN, GRAIN - ACRES HARVESTED'),
            'yield':          ('grain',       'CORN, GRAIN - YIELD, MEASURED IN BU / ACRE'),
            'production':     ('grain',       'CORN, GRAIN - PRODUCTION, MEASURED IN BU'),
        },
    },
    'sorghum': {
        'file': 'models/Feed Grains/us_sorghum_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'SORGHUM - ACRES PLANTED'),
            'area_harvested': ('grain',       'SORGHUM, GRAIN - ACRES HARVESTED'),
            'yield':          ('grain',       'SORGHUM, GRAIN - YIELD, MEASURED IN BU / ACRE'),
            'production':     ('grain',       'SORGHUM, GRAIN - PRODUCTION, MEASURED IN BU'),
        },
    },
    'wheat': {
        'file': 'models/Food Grains/us_wheat_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'WHEAT - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'WHEAT - ACRES HARVESTED'),
            'yield':          ('all_classes', 'WHEAT - YIELD, MEASURED IN BU / ACRE'),
            'production':     ('all_classes', 'WHEAT - PRODUCTION, MEASURED IN BU'),
        },
    },
    'cotton': {
        'file': 'models/Cotton/us_cotton_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'COTTON - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'COTTON - ACRES HARVESTED'),
            'yield':          ('all_classes', 'COTTON - YIELD, MEASURED IN LB / ACRE'),
            'production':     ('all_classes', 'COTTON - PRODUCTION, MEASURED IN 480 LB BALES'),
        },
    },
    'safflower': {
        'file': 'models/Oilseeds/us_safflower_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'SAFFLOWER - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'SAFFLOWER - ACRES HARVESTED'),
            'yield':          ('all_classes', 'SAFFLOWER - YIELD, MEASURED IN LB / ACRE'),
            'production':     ('all_classes', 'SAFFLOWER - PRODUCTION, MEASURED IN LB'),
        },
    },
    'flaxseed': {
        'file': 'models/Oilseeds/us_flaxseed_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'FLAXSEED - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'FLAXSEED - ACRES HARVESTED'),
            'yield':          ('all_classes', 'FLAXSEED - YIELD, MEASURED IN BU / ACRE'),
            'production':     ('all_classes', 'FLAXSEED - PRODUCTION, MEASURED IN BU'),
        },
    },
    'mustard': {
        'file': 'models/Oilseeds/us_mustard_production.xlsx',
        'stats': {
            'area_planted':   ('all_classes', 'MUSTARD - ACRES PLANTED'),
            'area_harvested': ('all_classes', 'MUSTARD - ACRES HARVESTED'),
            'yield':          ('all_classes', 'MUSTARD - YIELD, MEASURED IN LB / ACRE'),
            'production':     ('all_classes', 'MUSTARD - PRODUCTION, MEASURED IN LB'),
        },
    },
}


# ============================================================
# Vintage column definitions
# ============================================================
def _any(rel, yr): return True
def _final_year(rel, yr): return rel.year >= yr + 1

AP_COLS = [
    ('PP (Mar)',      ['YEAR - MAR ACREAGE'], _any,        'earliest'),
    ('Acreage (Jun)', ['YEAR - JUN ACREAGE'], _any,        'earliest'),
    ('Aug WASDE',     ['YEAR - AUG FORECAST'], _any,       'earliest'),
    ('Sep',           ['YEAR - SEP FORECAST'], _any,       'earliest'),
    ('Oct',           ['YEAR - OCT FORECAST'], _any,       'earliest'),
    ('Nov',           ['YEAR - NOV FORECAST'], _any,       'earliest'),
    ('Final (Jan)',   ['YEAR'],                _final_year, 'earliest'),
]
SEASON_COLS = [
    ('Aug WASDE',     ['YEAR - AUG FORECAST'], _any,       'earliest'),
    ('Sep',           ['YEAR - SEP FORECAST'], _any,       'earliest'),
    ('Oct',           ['YEAR - OCT FORECAST'], _any,       'earliest'),
    ('Nov',           ['YEAR - NOV FORECAST'], _any,       'earliest'),
    ('Final (Jan)',   ['YEAR'],                _final_year, 'earliest'),
]


# ============================================================
# Styling
# ============================================================
HEADER_FILL = PatternFill(start_color='3C7D22', end_color='3C7D22', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
YEAR_FONT   = Font(name='Calibri', size=10, bold=True)
DATA_FONT   = Font(name='Calibri', size=10)
ALT_FILL    = PatternFill(start_color='F4F8F1', end_color='F4F8F1', fill_type='solid')


# ============================================================
# Core
# ============================================================
def fetch_vintage(commodity_db: str, class_db: str, statistic: str,
                  short_desc: str) -> list:
    sql = """
        SELECT crop_year, reference_period, release_date, value, unit
        FROM silver.crop_production
        WHERE commodity = %s AND class = %s
          AND statistic = %s AND short_desc = %s
          AND agg_level = 'NATIONAL'
          AND value IS NOT NULL
        ORDER BY crop_year, release_date
    """
    out = []
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (commodity_db, class_db, statistic, short_desc))
            for r in cur.fetchall():
                d = dict(r)
                out.append((int(d['crop_year']), d['reference_period'],
                            d['release_date'], float(d['value']), d['unit']))
    return out


def match_value(rows, year, ref_periods, rel_filter, mode):
    for ref_period in ref_periods:
        candidates = [
            (rel, val, unit) for (cy, rp, rel, val, unit) in rows
            if cy == year and rp == ref_period and rel_filter(rel, year)
        ]
        if not candidates:
            continue
        candidates.sort(key=lambda x: x[0])
        chosen = candidates[0] if mode == 'earliest' else candidates[-1]
        rel, val, unit = chosen
        return (val, unit, rel)
    return None


def build_tab(ws, statistic, rows, col_specs, year_ge, year_le):
    units = sorted({r[4] for r in rows})
    unit_label = units[0] if len(units) == 1 else f'mixed: {units}'

    ws.cell(1, 1, value='Year').font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    for i, (label, _, _, _) in enumerate(col_specs, start=2):
        c = ws.cell(1, i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center')

    row = 2
    for year in range(year_ge, year_le + 1):
        c = ws.cell(row, 1, value=year)
        c.font = YEAR_FONT
        if year % 2 == 1:
            c.fill = ALT_FILL
        for i, (_, ref_periods, rel_filter, mode) in enumerate(col_specs, start=2):
            match = match_value(rows, year, ref_periods, rel_filter, mode)
            cell = ws.cell(row, i)
            if year % 2 == 1:
                cell.fill = ALT_FILL
            if match is None:
                continue
            value, unit, rel_date = match
            cell.value = value
            cell.font = DATA_FONT
            cell.number_format = '#,##0.0' if statistic == 'yield' else '#,##0'
        row += 1

    ws.column_dimensions['A'].width = 8
    for c in range(2, 2 + len(col_specs)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    ws.freeze_panes = 'B2'

    foot_row = row + 1
    ws.cell(foot_row, 1, value=f'Units: {unit_label}').font = Font(
        name='Calibri', size=9, italic=True)
    return row - 2, unit_label


def build_meta(ws, file_meta, commodity_name, year_ge, year_le):
    ws['A1'] = 'Meta'; ws['B1'] = 'Value'
    for c in ('A1', 'B1'):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL

    ws['A2'] = 'Commodity'
    ws['B2'] = commodity_name
    ws['A3'] = 'Generated'
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A4'] = 'Source'
    ws['B4'] = 'silver.crop_production (NASS QuickStats / Crop Production / Annual CPS)'
    ws['A5'] = 'Year range'
    ws['B5'] = f'{year_ge} - {year_le}'
    ws['A6'] = 'Layout'
    ws['B6'] = ('One row per crop year, columns = vintage release. '
                'Original release values preserved.')
    ws['A7'] = 'Vintage rule'
    ws['B7'] = ('PP=YEAR-MAR ACREAGE earliest; Acreage=YEAR-JUN ACREAGE earliest; '
                'Aug/Sep/Oct/Nov=YEAR-<MONTH> FORECAST earliest; '
                'Final=YEAR earliest with rel.year>=crop_year+1.')

    r = 9
    ws[f'A{r}'] = 'Tab'
    ws[f'B{r}'] = 'Rows'
    ws[f'C{r}'] = 'Source short_desc'
    ws[f'D{r}'] = 'Unit'
    for c in (f'A{r}', f'B{r}', f'C{r}', f'D{r}'):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
    r += 1
    for tab, info in file_meta.items():
        ws.cell(r, 1, value=tab)
        ws.cell(r, 2, value=info['rows'])
        ws.cell(r, 3, value=info['short_desc'])
        ws.cell(r, 4, value=info['unit'])
        r += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 25


def build_one(commodity_key, cfg, year_ge=2000, year_le=2026):
    out_path = PROJECT_ROOT / cfg['file']
    out_path.parent.mkdir(parents=True, exist_ok=True)
    commodity_db = cfg.get('commodity_db', commodity_key)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    file_meta = {}

    for stat, (class_db, short_desc) in cfg['stats'].items():
        rows = fetch_vintage(commodity_db, class_db, stat, short_desc)
        ws = wb.create_sheet(title=stat)
        col_specs = AP_COLS if stat == 'area_planted' else SEASON_COLS
        n, unit_label = build_tab(ws, stat, rows, col_specs, year_ge, year_le)
        file_meta[stat] = {'rows': n, 'short_desc': short_desc, 'unit': unit_label}

    ws_meta = wb.create_sheet(title='_meta')
    build_meta(ws_meta, file_meta, commodity_key, year_ge, year_le)

    wb.save(out_path)
    log.info(f'Wrote {out_path.relative_to(PROJECT_ROOT)} ({sum(v["rows"] for v in file_meta.values())} data rows)')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--commodity', action='append',
                        choices=list(COMMODITY_CONFIG.keys()),
                        help='Repeatable. Default: all.')
    parser.add_argument('--year-ge', type=int, default=2000)
    parser.add_argument('--year-le', type=int, default=2026)
    args = parser.parse_args()
    commodities = args.commodity or list(COMMODITY_CONFIG.keys())
    for c in commodities:
        build_one(c, COMMODITY_CONFIG[c], args.year_ge, args.year_le)


if __name__ == '__main__':
    main()
