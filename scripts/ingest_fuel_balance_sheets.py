"""
Ingest Fuel Balance Sheet Production Forecasts into Database
Reads monthly production data from analyst-maintained Excel balance sheets
and upserts into silver.fuel_production_forecast.

The allocator reads this table to determine how much fuel each facility
produces (and therefore how much feedstock it needs).

Usage:
    python scripts/ingest_fuel_balance_sheets.py --all
    python scripts/ingest_fuel_balance_sheets.py --fuel biodiesel
    python scripts/ingest_fuel_balance_sheets.py --fuel renewable_diesel --year-range 2024 2028
"""

import argparse
import logging
import os
import sys
from datetime import date, datetime
from pathlib import Path

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv(PROJECT_ROOT / '.env')

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("fuel_ingest")

TEMPLATE_DIR = PROJECT_ROOT / "output" / "balance_sheet_templates"

# Maps fuel key -> (filename, sheet_name, start_year, display_name)
FUEL_FILES = {
    'biodiesel': ('us_biodiesel_bal_sheets.xlsx', 'Biodiesel', 2001, 'BIODIESEL'),
    'renewable_diesel': ('us_renewable_diesel_bal_sheets.xlsx', 'Renewable Diesel', 2011, 'RENEWABLE DIESEL'),
    'saf': ('us_saf_bal_sheets.xlsx', 'SAF', 2020, 'SAF'),
}


def get_db_connection():
    return psycopg2.connect(
        host=os.getenv('RLC_PG_HOST'),
        dbname='rlc_commodities',
        user=os.getenv('RLC_PG_USER'),
        password=os.getenv('RLC_PG_PASSWORD'),
        sslmode='require',
    )


def find_monthly_block(ws, fuel_display, block_keyword):
    """
    Find a monthly block by scanning column A for "US {FUEL} {BLOCK}" title.
    Uses the full block title pattern to avoid matching annual S&D rows.
    Returns (month_start_row, year_header_row) or (None, None).
    """
    pattern = f"US {fuel_display} {block_keyword}".upper()
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and pattern in str(val).upper():
            # Next row should be unit/year headers, then 12 months
            year_header_row = row + 1
            month_start_row = row + 2
            return month_start_row, year_header_row
    return None, None


def read_year_columns(ws, year_header_row):
    """Read year -> column mapping from the year header row."""
    year_cols = {}
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=year_header_row, column=col).value
        if val is not None:
            try:
                year_cols[int(val)] = col
            except (ValueError, TypeError):
                pass
    return year_cols


def read_monthly_block(ws, month_start_row, year_cols, year_range=None):
    """
    Read 12 months of data from a monthly block.
    Returns dict: {(year, month): value}
    """
    data = {}
    for year, col in year_cols.items():
        if year_range and (year < year_range[0] or year > year_range[1]):
            continue
        for m in range(12):
            val = ws.cell(row=month_start_row + m, column=col).value
            if val is not None:
                try:
                    data[(year, m + 1)] = float(val)
                except (ValueError, TypeError):
                    pass
    return data


def ingest_fuel(fuel_key, year_range=None):
    """Ingest one fuel type's production forecasts into the database."""
    filename, sheet_name, default_start, fuel_display = FUEL_FILES[fuel_key]
    filepath = TEMPLATE_DIR / filename

    if not filepath.exists():
        logger.error(f"File not found: {filepath}")
        return 0

    logger.info(f"Ingesting {fuel_key} from {filename}...")

    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]

    # Find production block (matches "US BIODIESEL PRODUCTION" etc.)
    prod_start, prod_year_row = find_monthly_block(ws, fuel_display, 'PRODUCTION')
    if not prod_start:
        logger.error(f"  Could not find PRODUCTION block in {sheet_name}")
        wb.close()
        return 0

    year_cols = read_year_columns(ws, prod_year_row)
    logger.info(f"  Found PRODUCTION block at row {prod_start}, years: {min(year_cols)}..{max(year_cols)}")
    production_data = read_monthly_block(ws, prod_start, year_cols, year_range)

    # Find capacity block (optional)
    cap_start, cap_year_row = find_monthly_block(ws, fuel_display, 'CAPACITY')
    capacity_data = {}
    if cap_start:
        cap_year_cols = read_year_columns(ws, cap_year_row)
        capacity_data = read_monthly_block(ws, cap_start, cap_year_cols, year_range)

    wb.close()

    if not production_data:
        logger.warning(f"  No production data found for {fuel_key}")
        return 0

    # Determine forecast cutoff (current month)
    today = date.today()
    forecast_cutoff = date(today.year, today.month, 1)

    # Upsert into silver.fuel_production_forecast
    conn = get_db_connection()
    cur = conn.cursor()

    upserted = 0
    for (year, month), prod_val in sorted(production_data.items()):
        period = date(year, month, 1)
        is_forecast = period >= forecast_cutoff

        cap_val = capacity_data.get((year, month))
        util_pct = None
        if cap_val and cap_val > 0 and prod_val is not None:
            # Annualize monthly production and compare to capacity
            util_pct = round((prod_val * 12) / cap_val * 100, 1)

        cur.execute("""
            INSERT INTO silver.fuel_production_forecast
                (period, fuel_type, production_mmgal, capacity_mmgy, utilization_pct,
                 is_forecast, source, source_file, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, 'balance_sheet', %s, NOW())
            ON CONFLICT (period, fuel_type) DO UPDATE SET
                production_mmgal = EXCLUDED.production_mmgal,
                capacity_mmgy = EXCLUDED.capacity_mmgy,
                utilization_pct = EXCLUDED.utilization_pct,
                is_forecast = EXCLUDED.is_forecast,
                source = EXCLUDED.source,
                source_file = EXCLUDED.source_file,
                updated_at = NOW()
        """, (period, fuel_key, prod_val, cap_val, util_pct, is_forecast, filename))
        upserted += 1

    conn.commit()
    conn.close()
    logger.info(f"  Upserted {upserted} monthly records for {fuel_key}")
    return upserted


def main():
    parser = argparse.ArgumentParser(description="Ingest fuel balance sheet forecasts into DB")
    parser.add_argument("--fuel", choices=list(FUEL_FILES.keys()),
                        help="Ingest only this fuel type")
    parser.add_argument("--all", action="store_true", help="Ingest all fuel types")
    parser.add_argument("--year-range", nargs=2, type=int, metavar=('START', 'END'),
                        help="Only ingest years in this range (e.g., --year-range 2024 2028)")
    args = parser.parse_args()

    if not args.fuel and not args.all:
        parser.error("Specify --fuel or --all")

    year_range = tuple(args.year_range) if args.year_range else None
    total = 0

    fuels = [args.fuel] if args.fuel else list(FUEL_FILES.keys())
    for fuel_key in fuels:
        try:
            total += ingest_fuel(fuel_key, year_range)
        except Exception as e:
            logger.error(f"Failed to ingest {fuel_key}: {e}", exc_info=True)

    logger.info(f"\nTotal: {total} records upserted across {len(fuels)} fuel types")


if __name__ == '__main__':
    main()
