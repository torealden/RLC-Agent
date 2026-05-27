"""
Add coconut_oil_food_use_subbalance + pk_oil_food_use_subbalance tabs
to us_coconut_balance_sheets.xlsm and us_palm_complex_balance_sheets.xlsm.

Tab layout matches the soybean-complex balance sheet convention:
  Row 1: title
  Row 2: subtitle / methodology note
  Row 3: MY column headers
  Row 4: units note
  Rows 6-10: annual rollup by sub-flow (top section)
  Rows 12+: monthly sections, one per sub-flow:
    Section header, units row, then 12 month rows (Oct-Sep for
    lauric MY per ERS T32 convention), then MY total row.

USDA does NOT publish monthly food use for lauric oils. Until the
USDA ERS Food Expenditure Series collector is built, monthly cells
use a flat 1/12 placeholder allocation of the annual modeled total.
TODO marker noted in the subtitle row.

Sub-flows (Tier 2B, MODELED):
  Confectionery, Baking / Food Service, Food Industrial, Non-food
  Industrial

Idempotent — removes existing tabs before re-adding.
See memory: reference_usda_food_expenditure_reality_check.md
"""

from __future__ import annotations

import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / '.env')
from src.services.database.db_config import get_connection

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger('lauric_food_tab')

HEADER_FILL = PatternFill('solid', fgColor='3C7D22')
PLACEHOLDER_FILL = PatternFill('solid', fgColor='FFF4D6')  # cream
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri')
SECTION_FONT = Font(bold=True, name='Calibri')
ITALIC_FONT = Font(italic=True, color='666666', name='Calibri')

# Lauric MY = Oct-Sep per ERS Table 32
MY_MONTHS = [('Oct', 10), ('Nov', 11), ('Dec', 12),
             ('Jan', 1), ('Feb', 2), ('Mar', 3), ('Apr', 4),
             ('May', 5), ('Jun', 6), ('Jul', 7), ('Aug', 8), ('Sep', 9)]

# Sub-flow order
SUB_FLOWS = [
    ('Confectionery',                                  'confectionery'),
    ('Baking / Food Service',                          'baking_food_service'),
    ('Food Industrial (margarine, shortening)',        'food_industrial'),
    ('Non-food Industrial (soap, cosmetics, oleo)',    'non_food_industrial'),
]

# Section spacing
SECTION_SPACING = 16
FIRST_MONTHLY_SECTION_ROW = 30

TARGETS = [
    {
        'xlsm':      PROJECT_ROOT / 'models' / 'Oilseeds' / 'us_coconut_balance_sheets.xlsm',
        'tab':       'coconut_oil_food_use_subbalance',
        'commodity': 'coconut_oil',
        'title':     'US COCONUT OIL FOOD USE SUB-BALANCE',
        'cleanup':   [],
    },
    {
        'xlsm':      PROJECT_ROOT / 'models' / 'Oilseeds' / 'us_palm_complex_balance_sheets.xlsm',
        'tab':       'pk_oil_food_use_subbalance',
        'commodity': 'palm_kernel_oil',
        'title':     'US PALM KERNEL OIL FOOD USE SUB-BALANCE',
        'cleanup':   ['palm_kernel_oil_food_use_subbalance'],  # too-long legacy name
    },
]


def build(cfg):
    fp = cfg['xlsm']
    if not fp.exists():
        logger.error(f"{fp.name}: not found"); return

    logger.info(f"\n=== {fp.name} ===")

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT marketing_year, sub_flow, allocated_mil_lbs, share_pct
                FROM silver.lauric_food_use_modeled
                WHERE commodity = %s
                ORDER BY marketing_year, sub_flow
            """, (cfg['commodity'],))
            rows = list(cur.fetchall())

    # Aggregate into {(my, sub_flow): allocated_mil_lbs} and {sub_flow: share}
    annual_by_my_flow = {(r['marketing_year'], r['sub_flow']): float(r['allocated_mil_lbs']) if r['allocated_mil_lbs'] is not None else None for r in rows}
    shares = {r['sub_flow']: float(r['share_pct']) for r in rows}

    # MYs available
    my_labels = sorted({r['marketing_year'] for r in rows})
    logger.info(f"  {len(my_labels)} MYs, {len(rows)} sub-flow rows")

    backup = fp.parent / (fp.name + f'.bak_{datetime.now():%Y%m%d_%H%M%S}')
    shutil.copy2(fp, backup)
    logger.info(f"  Backup: {backup.name}")

    wb = openpyxl.load_workbook(fp, keep_vba=fp.name.endswith('.xlsm'))
    for old in cfg['cleanup']:
        if old in wb.sheetnames:
            del wb[old]
            logger.info(f"  Removed legacy: {old}")
    if cfg['tab'] in wb.sheetnames:
        del wb[cfg['tab']]
    ws = wb.create_sheet(cfg['tab'])

    # ── Header rows ──
    ws.cell(1, 1, cfg['title']).font = Font(bold=True, size=14, name='Calibri')
    ws.cell(2, 1, 'Tier 2B (MODELED) — million pounds. Annual = ERS T32 domestic disappearance × assumption shares. Monthly = annual ÷ 12 PLACEHOLDER (TODO: replace with USDA ERS Food Expenditure Series FAH/FAFH seasonality).').font = Font(italic=True, name='Calibri')

    # MY year labels in row 3
    my_cols = {}
    for col_idx, my_label in enumerate(my_labels, start=2):
        c = ws.cell(3, col_idx, my_label)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        my_cols[my_label] = col_idx
    ws.cell(4, 1, '(million pounds, lauric oil basis)').font = ITALIC_FONT

    # ── Annual rollup block ──
    ws.cell(6, 1, f'ANNUAL — MODELED (' + ' / '.join(
        f"{label.split('(')[0].strip()}={shares.get(field, 0)*100:.0f}%"
        for label, field in SUB_FLOWS
    ) + ')').font = SECTION_FONT

    for row_offset, (label, field) in enumerate(SUB_FLOWS):
        r = 7 + row_offset
        ws.cell(r, 1, label)
        for my_label, col_idx in my_cols.items():
            v = annual_by_my_flow.get((my_label, field))
            if v is not None:
                ws.cell(r, col_idx, v)

    # Total row
    total_row = 7 + len(SUB_FLOWS)
    c = ws.cell(total_row, 1, 'Total Domestic Use (= ERS T32)')
    c.font = SECTION_FONT
    for my_label, col_idx in my_cols.items():
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(total_row, col_idx, f'=SUM({col_letter}7:{col_letter}{6+len(SUB_FLOWS)})')
        cell.font = SECTION_FONT

    # ── Monthly sections (PLACEHOLDER: 1/12 of annual) ──
    for s_idx, (label, field) in enumerate(SUB_FLOWS):
        section_row = FIRST_MONTHLY_SECTION_ROW + s_idx * SECTION_SPACING
        ws.cell(section_row, 1,
                f'US {cfg["commodity"].upper().replace("_", " ")} {label.upper()} (MODELED)').font = SECTION_FONT
        ws.cell(section_row + 1, 1,
                '(million pounds — PLACEHOLDER: annual / 12, awaiting FAH/FAFH seasonality)').font = ITALIC_FONT

        for m_idx, (m_label, _m_num) in enumerate(MY_MONTHS):
            r = section_row + 2 + m_idx
            ws.cell(r, 1, m_label)
            for my_label, col_idx in my_cols.items():
                v = annual_by_my_flow.get((my_label, field))
                if v is not None:
                    cell = ws.cell(r, col_idx, round(v / 12.0, 2))
                    cell.fill = PLACEHOLDER_FILL  # mark as modeled placeholder

        total_row = section_row + 2 + 12
        c = ws.cell(total_row, 1, 'Marketing-year Total')
        c.font = SECTION_FONT
        for col_idx in my_cols.values():
            col_letter = get_column_letter(col_idx)
            ws.cell(total_row, col_idx,
                    f'=SUM({col_letter}{section_row+2}:{col_letter}{section_row+13})').font = SECTION_FONT

    ws.column_dimensions['A'].width = 50
    for col_idx in my_cols.values():
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    wb.save(fp)
    logger.info(f"  Saved: {fp.name}")


def main():
    for cfg in TARGETS:
        build(cfg)


if __name__ == '__main__':
    main()
