"""Rescale legacy cells in us_oilseed_crush.xlsm to the post-migration-133 unit.

Problem: migration 133 changed conversion_factor for ~60 attributes across 5
oilseeds. The .bas updater only writes cells where gold.fats_oils_crush_matrix
has a non-null display_value. Months where NASS bronze has no data (canola
pre-2015, gap years, etc.) keep whatever was previously in the cell — typically
ERS Oil Crops Yearbook fills in the OLD scale (000 short tons for meals,
mil lbs for oils). Result: same column shows two scales side by side.

This script identifies every (commodity, attribute, year, month) where the
spreadsheet has a value but the DB does not, and multiplies it by the
mig-133-era rescale ratio (new_factor / old_factor) so the column becomes
internally consistent in the post-migration unit.

Usage:
    python scripts/rescale_oilseed_crush_legacy_cells.py [--dry-run] [--commodity NAME]

Run --dry-run first to see what would change. Then run for real (no flag).
"""

from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path
from typing import Dict, Set, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dotenv import load_dotenv
load_dotenv()

from src.services.database.db_config import get_connection


WORKBOOK = Path(r'C:\dev\RLC-Agent\models\Oilseeds\us_oilseed_crush.xlsm')

# Light gray fill applied to gap cells (those filled from Tore's prior
# estimates workbook rather than NASS bronze). Lets you see at a glance
# which cells are estimates vs which are NASS-sourced.
ESTIMATE_FILL = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')

# Map sheet name -> DB commodity name
SHEET_TO_COMMODITY = {
    'soy_crush':        'soybeans',
    'canola_crush':     'canola',
    'cottonseed_crush': 'cottonseed',
    'sunflower_crush':  'sunflower',
    'peanut_crush':     'peanut',
}

# Rescale ratio = new_factor / old_factor (i.e. multiply legacy cell by this
# to bring it into the post-mig-133 scale).
# Per-attribute rather than per-commodity since the same attribute_code
# follows the same rule across commodities.
RESCALE_RATIOS: Dict[str, float] = {
    # Oils LB->000 lbs (was mil lbs): old 1e-6, new 0.001 -> x1000
    'crude_oil_production':                1000.0,
    'crude_oil_refined':                   1000.0,
    'crude_oil_stocks':                    1000.0,
    'crude_oil_stocks_total':              1000.0,
    'crude_oil_crusher_stocks':            1000.0,
    'crude_oil_inedible_use':              1000.0,
    'crude_oil_production_est':            1000.0,  # sunflower est
    'crude_oil_production_mills':          1.0,     # peanut: factor unchanged (text-only norm)
    'crude_oil_stocks_mills':              1.0,     # peanut: factor unchanged
    'refined_oil_production':              1000.0,
    'refined_oil_stocks':                  1000.0,
    'refined_oil_further_processing':      1000.0,
    'refined_oil_edible_use':              1000.0,
    'refined_oil_inedible_use':            1000.0,
    'oil_offsite_stocks':                  1000.0,  # soybeans

    # Meals TONS->tons (was 000 tons): old 0.001, new 1.0 -> x1000
    'meal_production':                     1000.0,
    'meal_stocks':                         1000.0,
    'meal_animal_feed':                    1000.0,
    'meal_edible_protein':                 1000.0,
    'millfeed_production':                 1000.0,
    'millfeed_stocks':                     1000.0,

    # Peanut meals LB->tons:
    'cake_meal_production':                0.5,    # old 0.001 (000 lbs), new 0.0005 (tons) -> x0.5
    'cake_meal_stocks':                    1000.0, # old 5e-7 (000 ST), new 0.0005 (tons) -> x1000

    # Sunflower derived meal_production_est: old 5e-7, new 0.0005 -> x1000
    'meal_production_est':                 1000.0,

    # Seeds:
    'seeds_crushed':                       1000.0,  # cottonseed: TONS, old 0.001 (000 ST), new 1.0 (tons) -> x1000
                                                    # canola: TONS, old 0.001, new 2.0 -> x2000 — HANDLED BELOW
    'seeds_crushed_est':                   2000.0,  # sunflower: LB, old 5e-7, new 0.001 -> x2000

    # Peanut-block "thousand pounds" -> "000 lbs" text-only normalization
    # (factor was always 0.001). Legacy cells already in the right scale.
    'shelled_peanuts_crushed':             1.0,
    'shelled_crush_farmer_stock_basis':    1.0,
    'edible_usage_total':                  1.0,
    'edible_usage_candy':                  1.0,
    'edible_usage_snacks':                 1.0,
    'edible_usage_peanut_butter':          1.0,
    'edible_usage_other':                  1.0,
    'shelled_usage_all_grades':            1.0,
    'in_shell_usage':                      1.0,
    'farmer_stock_total':                  1.0,
    'shelled_stocks_total':                1.0,
    'shelled_oil_stocks_production':       1.0,
    'roasting_stock_production':           1.0,
    'roasting_stock_in_shell_stocks':      1.0,
}

# Canola is the one commodity where the seeds_crushed factor went TONS->000 lbs (x2000)
# rather than the uniform x1000. Override.
CANOLA_OVERRIDES: Dict[str, float] = {
    'seeds_crushed': 2000.0,
}


def load_db_coverage() -> Dict[Tuple[str, str], Set[Tuple[int, int]]]:
    """For every (commodity, attribute_code) pair, return the set of (year,month)
    tuples where gold.fats_oils_crush_matrix has a non-null display_value.
    Cells outside this set are 'gap cells' that legacy fills occupy."""
    coverage: Dict[Tuple[str, str], Set[Tuple[int, int]]] = {}
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT commodity, attribute_code, year, month
            FROM gold.fats_oils_crush_matrix
            WHERE display_value IS NOT NULL
              AND commodity IN ('soybeans','canola','cottonseed','sunflower','peanut')
        """)
        for r in cur.fetchall():
            k = (r['commodity'], r['attribute_code'])
            coverage.setdefault(k, set()).add((r['year'], r['month']))
    return coverage


def load_header_patterns() -> Dict[Tuple[str, str], str]:
    """Map (commodity, attribute_code) -> header_pattern (used to find the
    spreadsheet column via R3 match)."""
    out = {}
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT commodity, attribute_code, header_pattern
            FROM silver.crush_attribute_reference
            WHERE header_pattern IS NOT NULL AND is_active = TRUE
        """)
        for r in cur.fetchall():
            out[(r['commodity'], r['attribute_code'])] = r['header_pattern']
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--commodity', default=None,
                    help='Limit to one commodity (soybeans, canola, cottonseed, sunflower, peanut)')
    args = ap.parse_args()

    coverage = load_db_coverage()
    patterns = load_header_patterns()
    print(f'DB coverage: {sum(len(v) for v in coverage.values())} (commodity,attr,year,month) tuples')
    print(f'Header patterns loaded: {len(patterns)}')

    wb = load_workbook(WORKBOOK, keep_vba=True, data_only=False)

    total_changes = 0
    total_skipped_formulas = 0

    for sheet_name, commodity in SHEET_TO_COMMODITY.items():
        if args.commodity and args.commodity != commodity:
            continue
        if sheet_name not in wb.sheetnames:
            print(f'  [{sheet_name}] NOT IN WORKBOOK — skipping')
            continue
        ws = wb[sheet_name]
        print(f'\n=== {sheet_name} (commodity={commodity}) ===')

        # Build R3-header -> col map
        r3_to_col = {}
        for c in range(2, ws.max_column + 1):
            h = ws.cell(row=3, column=c).value
            if h:
                r3_to_col[str(h).strip().lower()] = c

        # Iterate attributes for this commodity
        sheet_changes = 0
        for (comm, attr_code), header_pat in patterns.items():
            if comm != commodity:
                continue
            if attr_code not in RESCALE_RATIOS:
                # Attribute not in mig 133 scope; skip
                continue

            ratio = RESCALE_RATIOS[attr_code]
            if commodity == 'canola' and attr_code in CANOLA_OVERRIDES:
                ratio = CANOLA_OVERRIDES[attr_code]
            # NOTE: even ratio==1.0 attrs get fill-shaded if they're gap cells,
            # because Tore wants visual marking of every "estimate" cell (DB has
            # no data, this is from a previous estimates workbook). Falls through.

            # Find the column
            pat_lower = str(header_pat).strip().lower()
            col = r3_to_col.get(pat_lower)
            if col is None:
                # Try partial match (some headers slightly differ)
                for h, c in r3_to_col.items():
                    if h == pat_lower or h.startswith(pat_lower) or pat_lower.startswith(h):
                        col = c
                        break
            if col is None:
                continue  # column not present on this sheet (expected for sub-blocks)

            covered = coverage.get((commodity, attr_code), set())

            attr_changes = 0
            for r in range(5, ws.max_row + 1):
                date_v = ws.cell(row=r, column=1).value
                if date_v is None:
                    continue
                if not hasattr(date_v, 'year'):
                    continue
                y, m = date_v.year, date_v.month

                if (y, m) in covered:
                    continue  # DB has data for this period; .bas writes it correctly

                cell = ws.cell(row=r, column=col)
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str):
                    if v.startswith('='):
                        total_skipped_formulas += 1
                        continue
                    try:
                        v = float(v)
                    except ValueError:
                        continue
                if not isinstance(v, (int, float)):
                    continue

                new_v = v * ratio
                if not args.dry_run:
                    cell.value = new_v
                    cell.fill = ESTIMATE_FILL
                attr_changes += 1

            if attr_changes:
                print(f'  {attr_code:<35} col={col} ratio={ratio:>7.4f}  cells_changed={attr_changes}')
                sheet_changes += attr_changes

        print(f'  TOTAL on {sheet_name}: {sheet_changes} cells')
        total_changes += sheet_changes

    if args.dry_run:
        print(f'\n[DRY RUN] would change {total_changes} cells. Skipped {total_skipped_formulas} formula cells.')
        return

    print(f'\nSaving... ({total_changes} cells changed, {total_skipped_formulas} formula cells skipped)')
    wb.save(WORKBOOK)
    print(f'Saved: {WORKBOOK}')


if __name__ == '__main__':
    main()
