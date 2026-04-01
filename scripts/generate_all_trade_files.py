"""
Generate ALL trade data files for every commodity/country combination.

Each trade file has:
- Annual MY columns from 1993/94 through 2026/27 (34 columns)
- Monthly columns from Jan 2013 through Dec 2026 (168 columns)
- Accumulator formulas in annual columns (SUM of 12 monthly cols for that MY)
- Regional subtotal SUM formulas for all rows
- Standard 217-row country layout matching existing trade files

Usage:
    python scripts/generate_all_trade_files.py
    python scripts/generate_all_trade_files.py --file world_soybean_trade
"""

import argparse
import logging
import os
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("trade_gen")

OUTPUT_DIR = Path("C:/Users/torem/RLC Dropbox/RLC Team Folder/RLC-Models")

# ── Constants ──────────────────────────────────────────────────────
ANNUAL_START_YEAR = 1993
ANNUAL_END_YEAR = 2026
# Monthly columns start from the MY start month of 1993
# (actual start varies by commodity MY — passed per sheet)
MONTHLY_END_YEAR = 2026

NAVY = '1F4E79'
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
DATA_FONT = Font(name='Calibri', size=10)
REGIONAL_FONT = Font(name='Calibri', size=10, bold=True)
TITLE_FONT = Font(name='Calibri', size=11, bold=True, color=NAVY)

REGIONAL_SUMS = {
    4:   (5, 31),
    33:  (34, 45),
    47:  (48, 59),
    61:  (62, 106),
    108: (109, 163),
    165: (166, 210),
}

COUNTRIES = {
    4: "EUROPEAN UNION-27",
    5: "AUSTRIA", 6: "BELGIUM", 7: "BULGARIA", 8: "CROATIA", 9: "CYPRUS",
    10: "CZECH REPUBLIC", 11: "DENMARK", 12: "ESTONIA", 13: "FINLAND",
    14: "FRANCE", 15: "GERMANY", 16: "GREECE", 17: "HUNGARY", 18: "IRELAND",
    19: "ITALY", 20: "LATVIA", 21: "LITHUANIA", 22: "LUXEMBOURG", 23: "MALTA",
    24: "NETHERLANDS", 25: "POLAND", 26: "PORTUGAL", 27: "ROMANIA",
    28: "SLOVAKIA", 29: "SLOVENIA", 30: "SPAIN", 31: "SWEDEN",
    33: "OTHER EUROPE TOTAL",
    34: "ALBANIA", 35: "ANDORRA", 36: "BOSNIA AND HERZEGOVINA", 37: "ICELAND",
    38: "KOSOVO", 39: "MONTENEGRO", 40: "NORTH MACEDONIA", 41: "NORWAY",
    42: "SERBIA", 43: "SWITZERLAND", 44: "TURKEY", 45: "UNITED KINGDOM",
    47: "FORMER SOVIET UNION",
    48: "ARMENIA", 49: "AZERBAIJAN", 50: "BELARUS", 51: "GEORGIA",
    52: "KAZAKHSTAN", 53: "KYRGYZSTAN", 54: "MOLDOVA", 55: "RUSSIA",
    56: "TAJIKISTAN", 57: "TURKMENISTAN", 58: "UKRAINE", 59: "UZBEKISTAN",
    61: "ASIA & OCEANIA",
    62: "AFGHANISTAN", 63: "AUSTRALIA", 64: "BANGLADESH", 65: "BRUNEI",
    66: "BURMA (MYANMAR)", 67: "CAMBODIA", 68: "CHINA", 69: "FIJI",
    70: "FRENCH POLYNESIA", 71: "HONG KONG", 72: "INDIA", 73: "INDONESIA",
    74: "IRAN", 75: "IRAQ", 76: "ISRAEL", 77: "JAPAN",
    78: "JORDAN", 79: "KOREA, NORTH", 80: "KOREA, SOUTH", 81: "KUWAIT",
    82: "LAOS", 83: "LEBANON", 84: "MACAU", 85: "MALAYSIA",
    86: "MALDIVES", 87: "MONGOLIA", 88: "NEPAL", 89: "NEW ZEALAND",
    90: "OMAN", 91: "PAKISTAN", 92: "PAPUA NEW GUINEA", 93: "PHILIPPINES",
    94: "QATAR", 95: "SAMOA", 96: "SAUDI ARABIA", 97: "SINGAPORE",
    98: "SRI LANKA", 99: "SYRIA", 100: "TAIWAN", 101: "THAILAND",
    102: "TIMOR-LESTE", 103: "TONGA", 104: "UNITED ARAB EMIRATES",
    105: "VIETNAM", 106: "YEMEN",
    108: "AFRICA",
    109: "ALGERIA", 110: "ANGOLA", 111: "BENIN", 112: "BOTSWANA",
    113: "BURKINA FASO", 114: "BURUNDI", 115: "CAMEROON",
    116: "CAPE VERDE", 117: "CENTRAL AFRICAN REPUBLIC", 118: "CHAD",
    119: "COMOROS", 120: "CONGO (BRAZZAVILLE)", 121: "CONGO (KINSHASA)",
    122: "DJIBOUTI", 123: "EGYPT", 124: "EQUATORIAL GUINEA",
    125: "ERITREA", 126: "ESWATINI", 127: "ETHIOPIA", 128: "GABON",
    129: "GAMBIA", 130: "GHANA", 131: "GUINEA", 132: "GUINEA-BISSAU",
    133: "IVORY COAST", 134: "KENYA", 135: "LESOTHO", 136: "LIBERIA",
    137: "LIBYA", 138: "MADAGASCAR", 139: "MALAWI", 140: "MALI",
    141: "MAURITANIA", 142: "MAURITIUS", 143: "MOROCCO", 144: "MOZAMBIQUE",
    145: "NAMIBIA", 146: "NIGER", 147: "NIGERIA", 148: "REUNION",
    149: "RWANDA", 150: "SAO TOME AND PRINCIPE", 151: "SENEGAL",
    152: "SEYCHELLES", 153: "SIERRA LEONE", 154: "SOMALIA",
    155: "SOUTH AFRICA", 156: "SOUTH SUDAN", 157: "SUDAN", 158: "TANZANIA",
    159: "TOGO", 160: "TUNISIA", 161: "UGANDA", 162: "ZAMBIA",
    163: "ZIMBABWE",
    165: "WESTERN HEMISPHERE",
    166: "ANTIGUA AND BARBUDA", 167: "ARGENTINA", 168: "ARUBA",
    169: "BAHAMAS", 170: "BARBADOS", 171: "BELIZE", 172: "BERMUDA",
    173: "BOLIVIA", 174: "BRAZIL", 175: "BRITISH VIRGIN ISLANDS",
    176: "CANADA", 177: "CAYMAN ISLANDS", 178: "CHILE", 179: "COLOMBIA",
    180: "COSTA RICA", 181: "CUBA", 182: "CURACAO", 183: "DOMINICA",
    184: "DOMINICAN REPUBLIC", 185: "ECUADOR", 186: "EL SALVADOR",
    187: "FRENCH GUIANA", 188: "GRENADA", 189: "GUADELOUPE",
    190: "GUATEMALA", 191: "GUYANA", 192: "HAITI", 193: "HONDURAS",
    194: "JAMAICA", 195: "MARTINIQUE", 196: "MEXICO", 197: "NICARAGUA",
    198: "PANAMA", 199: "PARAGUAY", 200: "PERU", 201: "ST KITTS AND NEVIS",
    202: "ST LUCIA", 203: "ST VINCENT AND GRENADINES",
    204: "SURINAME", 205: "TRINIDAD AND TOBAGO", 206: "TURKS AND CAICOS",
    207: "UNITED STATES", 208: "URUGUAY", 209: "VENEZUELA",
    210: "VIRGIN ISLANDS US",
    216: "SUM OF REGIONAL TOTALS",
    217: "WORLD TOTAL",
}


def get_my_label(my_start, year):
    """Generate MY label like '1993/94' or '2025/26'."""
    if my_start == 1:
        return str(year)
    end_yr = year + 1
    return f"{year}/{str(end_yr)[-2:]}"


def find_monthly_col_for_month(my_start, my_year, month_offset,
                                first_monthly_col, monthly_start_year, monthly_start_month):
    """
    Find the column number in the monthly section for a given MY month.

    month_offset: 0=first month of MY, 11=last month
    """
    cal_month = ((my_start - 1 + month_offset) % 12) + 1
    cal_year = my_year + (1 if (my_start - 1 + month_offset) >= 12 else 0)

    # Calculate offset from the first monthly column
    months_from_start = (cal_year - monthly_start_year) * 12 + (cal_month - monthly_start_month)
    if months_from_start < 0:
        return None
    return first_monthly_col + months_from_start


def create_trade_sheet(wb, sheet_name, flow, my_start_month):
    """Create one trade data sheet with annual MYs, monthly data, and all formulas."""
    ws = wb.create_sheet(title=sheet_name[:31])

    my_label = {9: 'Sep-Aug', 10: 'Oct-Sep', 6: 'Jun-May', 7: 'Jul-Jun',
                8: 'Aug-Jul', 1: 'Jan-Dec', 2: 'Feb-Jan', 3: 'Mar-Feb',
                4: 'Apr-Mar', 12: 'Dec-Nov'}.get(my_start_month, f'M{my_start_month}')

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"{my_label} MY {flow.upper()}").font = TITLE_FONT

    # Row 2: Headers
    ws.cell(row=2, column=1, value="1,000 MT").font = Font(
        name='Calibri', size=9, italic=True, color='808080')

    col = 2

    # Annual MY columns (1993/94 through 2026/27)
    annual_start_col = col
    for year in range(ANNUAL_START_YEAR, ANNUAL_END_YEAR + 1):
        label = get_my_label(my_start_month, year)
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        col += 1

    annual_end_col = col - 1
    annual_col_count = annual_end_col - annual_start_col + 1

    # Monthly columns start from the MY start month of 1993
    monthly_start_month = my_start_month
    monthly_start_year = ANNUAL_START_YEAR
    first_monthly_col = col
    current_date = date(monthly_start_year, monthly_start_month, 1)
    end_date = date(MONTHLY_END_YEAR, 12, 1)

    while current_date <= end_date:
        cell = ws.cell(row=2, column=col, value=current_date)
        cell.number_format = 'MMM-YY'
        cell.font = Font(name='Calibri', size=9)
        cell.alignment = Alignment(horizontal='center')
        col += 1
        if current_date.month == 12:
            current_date = date(current_date.year + 1, 1, 1)
        else:
            current_date = date(current_date.year, current_date.month + 1, 1)

    last_col = col - 1

    # Countries
    for row_num, country_name in sorted(COUNTRIES.items()):
        cell = ws.cell(row=row_num, column=1, value=country_name)
        if row_num in REGIONAL_SUMS or row_num in (216, 217):
            cell.font = REGIONAL_FONT
        else:
            cell.font = DATA_FONT

    # ── Regional subtotal formulas (all columns) ──────────────────
    for data_col in range(2, last_col + 1):
        cl = get_column_letter(data_col)

        for reg_row, (start, end) in REGIONAL_SUMS.items():
            ws.cell(row=reg_row, column=data_col,
                    value=f"=SUM({cl}{start}:{cl}{end})")

        # Row 216: Sum of regional totals
        ws.cell(row=216, column=data_col,
                value=f"={cl}4+{cl}33+{cl}47+{cl}61+{cl}108+{cl}165")

        # Row 217: World Total — left empty (macro fills this)

    # ── Accumulator formulas in annual columns ────────────────────
    # Each annual MY column = SUM of the 12 monthly columns for that MY
    accum_count = 0
    for my_idx, year in enumerate(range(ANNUAL_START_YEAR, ANNUAL_END_YEAR + 1)):
        annual_col = annual_start_col + my_idx

        # Find the 12 monthly columns for this MY
        monthly_cols = []
        for month_offset in range(12):
            mcol = find_monthly_col_for_month(
                my_start_month, year, month_offset,
                first_monthly_col, monthly_start_year, monthly_start_month)
            if mcol and mcol <= last_col:
                monthly_cols.append(mcol)

        if not monthly_cols:
            continue  # No monthly data for this MY (pre-2013)

        # Write accumulator formula for every data row
        for row in range(4, 218):
            if row in REGIONAL_SUMS or row in (216, 217):
                continue  # Regional rows already have SUM formulas
            if row in (32, 46, 60, 107, 164, 211, 212, 213, 214, 215):
                continue  # Skip blank/unused rows

            # Build SUM formula referencing the monthly columns
            col_refs = '+'.join(f"{get_column_letter(c)}{row}" for c in monthly_cols)
            ws.cell(row=row, column=annual_col, value=f"={col_refs}")
            accum_count += 1

    # Column widths
    ws.column_dimensions['A'].width = 28
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    ws.freeze_panes = 'B3'

    return ws, first_monthly_col, last_col, accum_count


# ── Trade file definitions ──────────────────────────────────────────

TRADE_FILES = {
    'us_grains_trade': {
        'filename': 'us_grains_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Feed Grains' / 'new_models',
        'sheets': [
            ('Corn Exports', 'exports', 9),
            ('Corn Imports', 'imports', 9),
            ('Wheat Exports', 'exports', 6),
            ('Wheat Imports', 'imports', 6),
            ('Sorghum Exports', 'exports', 9),
            ('Sorghum Imports', 'imports', 9),
            ('Barley Exports', 'exports', 6),
            ('Barley Imports', 'imports', 6),
            ('Rice Exports', 'exports', 8),
            ('Rice Imports', 'imports', 8),
        ],
    },
    'us_cotton_trade': {
        'filename': 'us_cotton_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Cotton' / 'new_models',
        'sheets': [
            ('Cotton Exports', 'exports', 8),
            ('Cotton Imports', 'imports', 8),
        ],
    },
    'us_sugar_trade': {
        'filename': 'us_sugar_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Sugar' / 'new_models',
        'sheets': [
            ('Sugar Exports', 'exports', 10),
            ('Sugar Imports', 'imports', 10),
        ],
    },
    'world_soybean_trade': {
        'filename': 'world_soybean_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Oilseeds' / 'new_models',
        'sheets': [(f'{c} {p} {f}', f, ms) for c in
                   ['Brazil','Argentina','China','EU','India','Paraguay',
                    'Uruguay','Canada','Japan','Mexico','Ukraine','Russia']
                   for p in ['Soybeans','Soybean Oil','Soybean Meal']
                   for f, ms in [('Exports', 9 if 'Soy' == p[:3] and 'Oil' not in p and 'Meal' not in p else 10),
                                 ('Imports', 9 if 'Soy' == p[:3] and 'Oil' not in p and 'Meal' not in p else 10)]],
    },
    'world_corn_trade': {
        'filename': 'world_corn_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Feed Grains' / 'new_models',
        'sheets': [(f'{c} Corn {f}', f.lower(), 9) for c in
                   ['Brazil','Argentina','China','Ukraine','EU']
                   for f in ['Exports','Imports']],
    },
    'world_wheat_trade': {
        'filename': 'world_wheat_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Food Grains' / 'new_models',
        'sheets': [(f'{c} Wheat {f}', f.lower(), 6) for c in
                   ['Russia','EU','Canada','Australia','Argentina','Ukraine','India']
                   for f in ['Exports','Imports']],
    },
    'world_rapeseed_trade': {
        'filename': 'world_rapeseed_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Oilseeds' / 'new_models',
        'sheets': [(f'{c} {p} {f}', f.lower(), 7 if 'Rapeseed' == p else 10)
                   for c in ['Canada','EU']
                   for p in ['Rapeseed','Rapeseed Oil','Rapeseed Meal']
                   for f in ['Exports','Imports']],
    },
    'world_palm_trade': {
        'filename': 'world_palm_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Oilseeds' / 'new_models',
        'sheets': [(f'{c} {p} {f}', f.lower(), 1)
                   for c in ['Indonesia','Malaysia','World']
                   for p in ['Palm Oil','Palm Kernel Oil']
                   for f in ['Exports','Imports']],
    },
    'world_sunflower_trade': {
        'filename': 'world_sunflower_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Oilseeds' / 'new_models',
        'sheets': [(f'World {p} {f}', f.lower(), 10)
                   for p in ['Sunflower','Sunflower Oil','Sunflower Meal']
                   for f in ['Exports','Imports']],
    },
    'world_cotton_trade': {
        'filename': 'world_cotton_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Cotton' / 'new_models',
        'sheets': [(f'{c} Cotton {f}', f.lower(), 8)
                   for c in ['Brazil','India','World']
                   for f in ['Exports','Imports']],
    },
    'world_sugar_trade': {
        'filename': 'world_sugar_trade.xlsx',
        'output_dir': OUTPUT_DIR / 'Sugar' / 'new_models',
        'sheets': [(f'{c} Sugar {f}', f.lower(), 10)
                   for c in ['Brazil','World']
                   for f in ['Exports','Imports']],
    },
}


def create_trade_file(name, config):
    """Create a complete trade file."""
    wb = Workbook()
    wb.remove(wb.active)

    output_dir = config['output_dir']
    output_dir.mkdir(parents=True, exist_ok=True)

    total_accum = 0
    for sheet_name, flow, my_start in config['sheets']:
        ws, first_monthly, last_col, accum = create_trade_sheet(
            wb, sheet_name, flow, my_start)
        total_accum += accum

    filepath = output_dir / config['filename']
    wb.save(filepath)
    logger.info(f"Created {filepath.name}: {len(config['sheets'])} sheets, "
                f"{total_accum:,} accumulator formulas")
    return filepath


def main():
    parser = argparse.ArgumentParser(description="Generate trade data files")
    parser.add_argument("--file", help="Generate specific file (e.g., world_soybean_trade)")
    args = parser.parse_args()

    if args.file:
        if args.file in TRADE_FILES:
            create_trade_file(args.file, TRADE_FILES[args.file])
        else:
            logger.error(f"Unknown: {args.file}. Available: {list(TRADE_FILES.keys())}")
    else:
        total = 0
        for name, config in TRADE_FILES.items():
            create_trade_file(name, config)
            total += 1
        logger.info(f"\nGenerated {total} trade files")


if __name__ == '__main__':
    main()
