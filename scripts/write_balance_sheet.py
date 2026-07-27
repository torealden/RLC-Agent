"""Generate a closed ANNUAL balance-sheet workbook from contract flat files.

The reusable sheet-generator (the mechanical template, done programmatically — see the Desktop-bounce
resolution in flat_file_contract_v1.md v1.1). For each commodity it:
  1. copies the flat file's long supply/demand rows into in-workbook ff_<tag>_supply / _demand MIRROR
     tabs (this is why there is no #VALUE! — nothing references a closed external workbook), and
  2. writes a balance-sheet tab whose S&D rows are wired to those mirrors with the VERIFIED idiom
     IF(COUNTIFS=0,"",SUMIFS(value,...,H,MAXIFS(H,...))), bounded $2:$8001, MY matched as integer.

ANNUAL grain (period_type='annual'): one column per marketing year. This closes now because PSD gives
a full internally-consistent annual S&D. Monthly rows already sit in the flat file for later monthly
enrichment (add 12 month rows, same wiring, change the E/F filters). Fundamentals only — no price layer.

Guards written into every tab: two tie-out cells (must read 0) + a non-triviality cell
(SUM(production) — coverage does not go green until > 0).

Run:  python scripts/write_balance_sheet.py brazil-soy
"""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:/dev/RLC-Agent")
OILSEEDS = ROOT / "models" / "Oilseeds"

SUPPLY_SERIES = {"beginning_stocks", "production", "imports", "ending_stocks"}
LONG_COLS = ["commodity", "class", "series", "marketing_year", "period_type", "period",
             "vintage", "vintage_rank", "value", "unit", "source"]

# a build target: country folder, output workbook, and the commodities (with a short mirror tag)
# Each staged Tier-A cell can be ANNUAL-closed now from its *_flat.xlsx (PSD control), before any
# monthly enrichment. Run `python scripts/write_balance_sheet.py <key>` then recalc-verify (win32com).
def _soy(f):   return [("soybeans", "sb"), ("soybean_meal", "sm"), ("soybean_oil", "so")]
def _rape(f):  return [("rapeseed", "rs"), ("rapeseed_meal", "rm"), ("rapeseed_oil", "ro")]
def _sun(f):   return [("sunflowerseed", "us"), ("sunflowerseed_meal", "um"), ("sunflowerseed_oil", "uo")]
def _palm(f):  return [("palm_oil", "po"), ("palm_kernel_oil", "ko")]

TARGETS = {
    "brazil-soy":     dict(folder="Brazil",    out="brazil_soybean_complex_balance_sheet.xlsx",     commodities=[("soybeans","bs"),("soybean_meal","bsm"),("soybean_oil","bso")]),
    "argentina-soy":  dict(folder="Argentina", out="argentina_soybean_complex_balance_sheet.xlsx",  commodities=_soy("Argentina")),
    "eu-rape":        dict(folder="Europe",    out="europe_rapeseed_complex_balance_sheet.xlsx",    commodities=_rape("Europe")),
    "canada-rape":    dict(folder="Canada",    out="canada_rapeseed_complex_balance_sheet.xlsx",    commodities=_rape("Canada")),
    "australia-rape": dict(folder="Australia", out="australia_rapeseed_complex_balance_sheet.xlsx", commodities=_rape("Australia")),
    "russia-rape":    dict(folder="Russia",    out="russia_rapeseed_complex_balance_sheet.xlsx",    commodities=_rape("Russia")),
    "ukraine-sun":    dict(folder="Ukraine",   out="ukraine_sunflower_complex_balance_sheet.xlsx",   commodities=_sun("Ukraine")),
    "russia-sun":     dict(folder="Russia",    out="russia_sunflower_complex_balance_sheet.xlsx",    commodities=_sun("Russia")),
    "argentina-sun":  dict(folder="Argentina", out="argentina_sunflower_complex_balance_sheet.xlsx", commodities=_sun("Argentina")),
    "malaysia-palm":  dict(folder="Malaysia",  out="malaysia_palm_complex_balance_sheet.xlsx",       commodities=_palm("Malaysia")),
    "indonesia-palm": dict(folder="Indonesia", out="indonesia_palm_complex_balance_sheet.xlsx",      commodities=_palm("Indonesia")),
}

GREEN = "3C7D22"; HDR = PatternFill("solid", fgColor="1B2A4A")
TITLE = Font(bold=True, size=13, color="1B2A4A", name="Calibri")
BOLD = Font(bold=True, name="Calibri")


def read_long(flat_path, commodity):
    """Return (supply_rows, demand_rows, annual_MYs) from a flat file's long tabs."""
    wb = openpyxl.load_workbook(flat_path, read_only=True)
    out = {}
    for side in ("supply", "demand"):
        ws = wb[f"{commodity}_{side}"]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[2] is None:
                continue
            rows.append(list(r[:11]))
        out[side] = rows
    wb.close()
    mys = sorted({int(r[3]) for r in out["supply"] if r[4] == "annual"})
    return out["supply"], out["demand"], mys


def write_mirror(wb, tag, side, rows):
    ws = wb.create_sheet(f"ff_{tag}_{side}")
    ws.sheet_state = "hidden"
    for j, h in enumerate(LONG_COLS, 1):
        ws.cell(1, j, h).font = BOLD
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            ws.cell(i, j, v)
    return ws.title


def wired(ff, series, commodity, col):
    """The verified IF/COUNTIFS/SUMIFS/MAXIFS cell against a mirror tab, annual grain."""
    crit = (f'{ff}!$A$2:$A$8001,"{commodity}",{ff}!$B$2:$B$8001,"ALL",'
            f'{ff}!$C$2:$C$8001,"{series}",{ff}!$D$2:$D$8001,{col}$3,'
            f'{ff}!$E$2:$E$8001,"annual",{ff}!$F$2:$F$8001,"ANNUAL"')
    # MAXIFS is Excel 2019+ — openpyxl must write it as _xlfn.MAXIFS or Excel treats it as an
    # undefined name and silently returns 0 (the US reference stores it exactly this way).
    return (f'=IF(COUNTIFS({crit})=0,"",'
            f'SUMIFS({ff}!$I$2:$I$8001,{crit},{ff}!$H$2:$H$8001,_xlfn.MAXIFS({ff}!$H$2:$H$8001,{crit})))')


def build_sheet(wb, commodity, tag, mys, ff_sup, ff_dem, country):
    # Excel caps sheet names at 31 chars. "sunflowerseed_meal_balance_sheet" is 32, so
    # abbreviate the long commodity in the TAB NAME only (A1 title keeps the full name).
    tab_commodity = commodity.replace("sunflowerseed", "sun")
    ws = wb.create_sheet(f"{tab_commodity}_balance_sheet")
    ws["A1"] = f"{country.upper()} {commodity.upper().replace('_',' ')} SUPPLY & DEMAND  (1000 MT, annual)"
    ws["A1"].font = TITLE
    # header row 3 = marketing years (integers, matched directly by the formula)
    ws["A3"] = "Marketing Year"; ws["A3"].font = BOLD
    for k, my in enumerate(mys):
        c = ws.cell(3, 2 + k, my); c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.fill = HDR; c.alignment = Alignment(horizontal="center")
    seed = not (commodity.endswith("_oil") or commodity.endswith("_meal"))
    # (label, kind, series) — kind: wire | sum | memo | tie | nontriv
    layout = [
        ("Beginning Stocks", "wire", "beginning_stocks"),
        ("Production", "wire", "production"),
        ("Imports", "wire", "imports"),
        ("Total Supply", "sum_supply", None),
        ("Crush (memo)", "memo", "crush") if seed else None,
        ("Domestic Use", "wire", "domestic_use"),
        ("Exports", "wire", "exports"),
        ("Total Demand", "sum_demand", None),
        ("Ending Stocks (computed)", "end_calc", None),
        ("Ending Stocks (reported)", "wire", "ending_stocks"),
        ("TIE-OUT  (computed − reported, must = 0)", "tie", None),
        ("CHECK  SUM(production) > 0", "nontriv", None),
    ]
    layout = [x for x in layout if x]
    rmap = {}
    r0 = 5
    for i, (label, kind, series) in enumerate(layout):
        r = r0 + i
        rmap[label] = r
        ws.cell(r, 1, label).font = BOLD if kind in ("sum_supply", "sum_demand", "end_calc", "tie", "nontriv") else Font(name="Calibri")
        for k, my in enumerate(mys):
            col = get_column_letter(2 + k)
            cell = ws.cell(r, 2 + k)
            if kind == "wire" or kind == "memo":
                ff = ff_sup if series in SUPPLY_SERIES else ff_dem
                cell.value = wired(ff, series, commodity, col)
            elif kind == "sum_supply":
                cell.value = f"=SUM({col}{rmap['Beginning Stocks']}:{col}{rmap['Imports']})"
            elif kind == "sum_demand":
                cell.value = f"={col}{rmap['Domestic Use']}+{col}{rmap['Exports']}"
            elif kind == "end_calc":
                cell.value = f"={col}{rmap['Total Supply']}-{col}{rmap['Total Demand']}"
            elif kind == "tie":
                cell.value = (f'=IF({col}{rmap["Ending Stocks (reported)"]}="","",'
                              f'ROUND({col}{rmap["Ending Stocks (computed)"]}-{col}{rmap["Ending Stocks (reported)"]},1))')
        if kind == "nontriv":
            pr = rmap["Production"]
            ws.cell(r, 2, f"=SUM(2:2)".replace("2:2", f"{get_column_letter(2)}{pr}:{get_column_letter(1+len(mys))}{pr}")).font = BOLD
            ws.cell(r, 1).font = Font(bold=True, color=GREEN, name="Calibri")
    ws.column_dimensions["A"].width = 34
    ws.freeze_panes = "B4"
    return ws.title


def build(target_key):
    t = TARGETS[target_key]
    folder = OILSEEDS / t["folder"]
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    summary = []
    for commodity, tag in t["commodities"]:
        flat = folder / f"{t['folder'].lower()}_{commodity}_flat.xlsx"
        if not flat.exists():
            raise SystemExit(f"missing flat file {flat}")
        sup, dem, mys = read_long(flat, commodity)
        ff_sup = write_mirror(wb, tag, "supply", sup)
        ff_dem = write_mirror(wb, tag, "demand", dem)
        build_sheet(wb, commodity, tag, mys, ff_sup, ff_dem, t["folder"])
        summary.append((commodity, len(sup), len(dem), mys[0], mys[-1]))
    out = folder / t["out"]
    wb.save(out)
    return out, summary


def main():
    key = sys.argv[1] if len(sys.argv) > 1 else "brazil-soy"
    out, summary = build(key)
    print(f"Wrote {out}")
    for c, ns, nd, my0, my1 in summary:
        print(f"  {c:14s} mirror: {ns} supply / {nd} demand rows | annual MY {my0}..{my1}")


if __name__ == "__main__":
    main()
