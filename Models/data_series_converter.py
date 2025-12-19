import pandas as pd
from openpyxl import load_workbook

# Load workbook using openpyxl for merged cells and formulas
wb = load_workbook("World Soybean Balance Sheets.xlsx", data_only=True)

output_rows = []  # will collect our flat data rows

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    country = sheet_name
    # Normalize country name (remove generic terms like "Soy Complex")
    country = country.replace("Soy Complex", "").strip()
    
    # Iterate through rows to find series blocks
    max_col = sheet.max_column
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cell_val = row[0]
        if cell_val and isinstance(cell_val, str) and "(" in cell_val:
            # Likely a title row of a data block.
            title = cell_val.strip()
            # Parse commodity and stat from title
            # e.g. "SOYBEAN IMPORTS (1,000 Tonnes)"
            # commodity="Soybean", stat="Imports", unit="1,000 Tonnes"
            if "SOYBEAN" in title or "Soybean" in title:
                if "MEAL" in title.upper():
                    commodity = "Soybean Meal"
                elif "OIL" in title.upper():
                    commodity = "Soybean Oil"
                else:
                    commodity = "Soybean"
            else:
                commodity = "Other"
            # Get unit inside parentheses
            unit = title[title.find("(")+1 : title.find(")")]
            # Get statistic name (remove commodity name and unit part)
            stat = title.split("(")[0].strip()
            # Remove commodity word from stat if present
            stat = stat.replace(commodity.upper(), "").strip().strip("-")
            
            # Check the structure: look ahead to see if months follow
            next_val = sheet.cell(row=i+2, column=1).value  # two rows down, colA
            if next_val and str(next_val).strip().capitalize() in ["January","February","March","April",
                                                                    "May","June","July","August",
                                                                    "September","October","November","December"]:
                # This is a monthly series block
                # Row i+1 is year headers, i+2 to i+13 are months Feb–Jan (assuming 12 months)
                year_row = sheet[i+1]
                # Collect year labels from column 3 onward (skip colA, colB which may be empty or month labels)
                year_labels = [cell.value for cell in year_row[2:]]  # C onwards
                # Now iterate months
                for m in range(0, 12):  # assuming 12 month rows
                    month_name = sheet.cell(row=i+2+m, column=1).value  # month in colA
                    if month_name is None or str(month_name).strip()=="":
                        break  # no month, exit
                    month_name = str(month_name).strip()
                    # Iterate over year columns for this month
                    for col_idx, year_label in enumerate(year_labels, start=3):  # starting at col C=3
                        value = sheet.cell(row=i+2+m, column=col_idx).value
                        if value is None or value == "" or str(value).startswith("#"):
                            continue  # skip missing or error
                        # Determine actual Year (convert label like "09/10" to 2009, etc.)
                        year_str = str(year_label).strip()
                        if "/" in year_str:
                            first_year = year_str.split("/")[0]
                            # Handle two-digit vs four-digit
                            if len(first_year) == 2:
                                yr = int(first_year)
                                year = 1900+yr if yr >= 60 else 2000+yr
                            else:
                                year = int(first_year)
                        else:
                            # If label is single year (e.g. "2010"), take it as int
                            year = int(float(year_label))  # float if read as number
                        # Determine actual calendar year of the month
                        # If month is January and marketing year = X (which started in Feb X),
                        # then January belongs to year+1
                        cal_year = year if month_name != "January" else year+1
                        output_rows.append({
                            "Country": country,
                            "Commodity": commodity,
                            "Statistic": stat,
                            "Unit": unit,
                            "Year": cal_year,
                            "Month": month_name,
                            "Value": value
                        })
            else:
                # This is an annual/multi-series block (Supply & Demand or similar)
                # The current title might be a heading for multiple stat lines.
                # Continue reading subsequent rows until an empty row or another title.
                r = i+1
                # Find the year labels row for this block.
                # Often, the year headers might be in the row above the title or the first row of the sheet.
                # We can scan to the first numeric or year-like values in columns >1.
                # For simplicity, assume the next non-empty row has data and the columns from 3 onwards are years.
                # Get year labels from the first data row that contains numeric entries.
                data_row = sheet[r]
                # Identify the columns that contain years (headers likely not in colA).
                # We find the first row where col3 or col4 has a number or year label.
                # (This part may vary; adjust logic as needed.)
                # ...
                # Here, let's assume row i had the title, and row i+1 onward are data lines with year columns already present.
                year_cells = [cell.value for cell in data_row[2:] if cell.value]  # starting colC
                if all(isinstance(y, str) and any(ch.isdigit() for ch in y) for y in year_cells):
                    # row i+1 is actually a header row (contains year labels)
                    year_labels = [str(y).strip() for y in year_cells]
                    r += 1  # move to first data row
                else:
                    # row i+1 is already the first data row, so the header row is likely above (or at top of sheet)
                    # Use the previously seen header (e.g. top-of-sheet header for years).
                    # For simplicity, skip detailed header-finding in pseudocode.
                    year_labels = []  # assume a global header is known
                # Now iterate through data lines until an empty line or next block title
                while r <= sheet.max_row:
                    first_cell = sheet.cell(row=r, column=1).value
                    if first_cell is None or str(first_cell).strip() == "" or "(" in str(first_cell):
                        # empty or a new block title reached, stop
                        break
                    stat_name = str(first_cell).strip()
                    # Assign correct unit (use block's unit if single, or decide if stat_name implies first or second unit)
                    unit_used = unit
                    if "/" in unit:
                        # e.g. "1,000 Hectares/1,000 Tonnes"
                        units_split = unit.split("/")
                        if "Yield" in stat_name or "Area" in stat_name:
                            unit_used = units_split[0].strip()  # hectares for area, (yield handled separately)
                            if "Yield" in stat_name:
                                # Could define yield unit explicitly, e.g. "Tonnes/Ha"
                                unit_used = "Tonnes per Hectare"
                        else:
                            unit_used = units_split[-1].strip()  # tonnes for production, etc.
                    # Iterate year columns for this stat
                    for col_idx, year_label in enumerate(year_labels, start=3):
                        val = sheet.cell(row=r, column=col_idx).value
                        if val is None or val == "" or str(val).startswith("#"):
                            continue
                        # determine Year similar way as above
                        year = year_label
                        if isinstance(year, str) and "/" in year:
                            first_year = year.split("/")[0]
                            year = int(first_year) if len(first_year)==4 else (1900+int(first_year) if int(first_year)>=60 else 2000+int(first_year))
                        else:
                            year = int(float(year)) if year is not None else None
                        output_rows.append({
                            "Country": country,
                            "Commodity": commodity,
                            "Statistic": stat_name,  # individual stat line name
                            "Unit": unit_used,
                            "Year": year,
                            "Month": None,  # annual
                            "Value": val
                        })
                    r += 1

# Convert output_rows to DataFrame and export to Excel/CSV
df = pd.DataFrame(output_rows)
df.to_excel("Flattened_Soybean_Data.xlsx", index=False)
