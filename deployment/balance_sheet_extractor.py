#!/usr/bin/env python3
"""
Balance Sheet Data Extractor

Specialized extractor for RLC commodity balance sheet Excel files.
Understands marketing year format, section headers, and month-based data.

Key patterns:
- Year columns: Match "2020/21" format (marketing years)
- Historical cutoff: 1965/66 through 2023/24
- Section headers: "SOYBEAN CRUSH" followed by blank row then months
- Tab filter: Only process tabs containing "Complex"

Usage:
    python deployment/balance_sheet_extractor.py --analyze "path/to/file.xlsx"
    python deployment/balance_sheet_extractor.py --extract "path/to/file.xlsx"
    python deployment/balance_sheet_extractor.py --extract-all
    python deployment/balance_sheet_extractor.py --status
"""

import argparse
import sqlite3
import logging
import re
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Project paths
PROJECT_ROOT = Path(__file__).parent.parent
MODELS_DIR = PROJECT_ROOT / "Models"
DATA_DIR = PROJECT_ROOT / "data"
DB_PATH = DATA_DIR / "rlc_commodities.db"

# Marketing year patterns:
# - "2020/21" (4-digit/2-digit)
# - "64/65" or "64/65 " (2-digit/2-digit, possibly with trailing space)
MARKETING_YEAR_PATTERN_4DIGIT = re.compile(r'^(\d{4})/(\d{2})\s*$')
MARKETING_YEAR_PATTERN_2DIGIT = re.compile(r'^(\d{2})/(\d{2})\s*$')

# Historical data range
MIN_YEAR = 1965
MAX_YEAR = 2023  # 2023/24 is the last historical year

# Month names for detecting month-based rows
MONTHS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
          'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
          'january', 'february', 'march', 'april', 'may', 'june',
          'july', 'august', 'september', 'october', 'november', 'december']

# Target files for initial extraction
TARGET_FILES = [
    "US Oilseed Balance Sheets",
    "World Lauric Oils Balance Sheets",
    "World Peanut Balance Sheets",
    "World Rapeseed Balance Sheets",
    "World Soybean Balance Sheets",
    "World Sunflower Balance Sheets",
    "World Vegetable Oil Balance Sheets",
]


@dataclass
class ExtractedData:
    """Container for extracted balance sheet data"""
    source_file: str
    sheet_name: str
    commodity: str
    country: str
    section: str  # e.g., "SOYBEAN CRUSH", "PRODUCTION"
    metric: str   # e.g., "Oct", "Nov" or "Imports", "Exports"
    marketing_year: str  # e.g., "2020/21"
    value: float
    unit: str = ""


@dataclass
class SheetAnalysis:
    """Analysis results for a sheet"""
    name: str
    has_complex: bool
    year_columns: Dict[str, int]  # marketing_year -> column_index
    sections: List[Dict[str, Any]]
    row_count: int
    data_preview: List[Dict]


def is_marketing_year(value: Any) -> Optional[Tuple[int, int]]:
    """
    Check if value is a marketing year like "2020/21" or "64/65".
    Returns (start_year, end_year) tuple or None.
    """
    if value is None:
        return None

    val_str = str(value).strip()

    # Try 4-digit format first: "2020/21"
    match = MARKETING_YEAR_PATTERN_4DIGIT.match(val_str)
    if match:
        start_year = int(match.group(1))
        end_year_short = int(match.group(2))
        end_year = (start_year // 100) * 100 + end_year_short
        if end_year_short < (start_year % 100):
            end_year += 100  # Handle century boundary like 1999/00
        return (start_year, end_year)

    # Try 2-digit format: "64/65"
    match = MARKETING_YEAR_PATTERN_2DIGIT.match(val_str)
    if match:
        start_short = int(match.group(1))
        end_short = int(match.group(2))

        # Convert 2-digit to 4-digit year
        # 64 -> 1964, 00 -> 2000, 25 -> 2025
        if start_short >= 50:
            start_year = 1900 + start_short
        else:
            start_year = 2000 + start_short

        if end_short >= 50:
            end_year = 1900 + end_short
        else:
            end_year = 2000 + end_short

        # Handle wrap-around: 99/00 means 1999/2000
        if end_year < start_year:
            end_year += 100

        return (start_year, end_year)

    return None


def normalize_marketing_year(value: Any) -> Optional[str]:
    """
    Normalize a marketing year to standard format "YYYY/YY".
    Converts "64/65" to "1964/65", "2020/21" stays as is.
    """
    years = is_marketing_year(value)
    if years:
        start_year, end_year = years
        return f"{start_year}/{end_year % 100:02d}"
    return None


def is_historical_year(marketing_year: str) -> bool:
    """Check if marketing year is within historical range."""
    years = is_marketing_year(marketing_year)
    if years:
        start_year = years[0]
        return MIN_YEAR <= start_year <= MAX_YEAR
    return False


def is_month(value: Any) -> bool:
    """Check if value is a month name."""
    if value is None:
        return False
    return str(value).strip().lower() in MONTHS


def detect_commodity_from_filename(filename: str) -> str:
    """Detect commodity from filename."""
    filename_lower = filename.lower()

    if 'soybean' in filename_lower or 'soy' in filename_lower:
        return 'soybeans'
    elif 'rapeseed' in filename_lower or 'canola' in filename_lower:
        return 'rapeseed'
    elif 'sunflower' in filename_lower:
        return 'sunflower'
    elif 'peanut' in filename_lower or 'groundnut' in filename_lower:
        return 'peanuts'
    elif 'lauric' in filename_lower or 'palm' in filename_lower or 'coconut' in filename_lower:
        return 'lauric_oils'
    elif 'vegetable oil' in filename_lower:
        return 'vegetable_oils'
    elif 'corn' in filename_lower:
        return 'corn'
    elif 'wheat' in filename_lower:
        return 'wheat'
    else:
        return 'unknown'


def detect_country_from_sheet(sheet_name: str) -> str:
    """Detect country/region from sheet name."""
    name_lower = sheet_name.lower()

    countries = {
        'us': ['us ', 'u.s.', 'united states', 'usa'],
        'brazil': ['brazil', 'br ', 'brasil'],
        'argentina': ['argentina', 'arg'],
        'china': ['china', 'cn '],
        'eu': ['eu ', 'europe', 'european'],
        'india': ['india'],
        'canada': ['canada', 'can '],
        'australia': ['australia', 'aus'],
        'world': ['world', 'global'],
        'indonesia': ['indonesia'],
        'malaysia': ['malaysia'],
        'ukraine': ['ukraine'],
        'russia': ['russia'],
        'paraguay': ['paraguay'],
    }

    for country, patterns in countries.items():
        if any(p in name_lower for p in patterns):
            return country.upper() if len(country) <= 2 else country.title()

    return 'Unknown'


class BalanceSheetExtractor:
    """Extracts data from RLC balance sheet Excel files."""

    def __init__(self, db_path: Path = DB_PATH):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

    def init_database(self):
        """Initialize database with balance sheet tables."""
        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()

        cursor.executescript("""
            -- Balance sheet data with full provenance
            CREATE TABLE IF NOT EXISTS commodity_balance_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                commodity TEXT NOT NULL,
                country TEXT,
                section TEXT,           -- e.g., "SOYBEAN CRUSH", "PRODUCTION"
                metric TEXT NOT NULL,   -- e.g., "Oct", "Total" or "Production"
                marketing_year TEXT NOT NULL,
                value REAL,
                unit TEXT,
                extracted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(source_file, sheet_name, section, metric, marketing_year)
            );

            CREATE INDEX IF NOT EXISTS idx_cbs_commodity ON commodity_balance_sheets(commodity);
            CREATE INDEX IF NOT EXISTS idx_cbs_country ON commodity_balance_sheets(country);
            CREATE INDEX IF NOT EXISTS idx_cbs_year ON commodity_balance_sheets(marketing_year);
            CREATE INDEX IF NOT EXISTS idx_cbs_section ON commodity_balance_sheets(section);

            -- Track extraction runs
            CREATE TABLE IF NOT EXISTS extraction_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT NOT NULL,
                sheets_processed INTEGER,
                rows_extracted INTEGER,
                status TEXT,
                error_message TEXT,
                started_at TIMESTAMP,
                completed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        conn.commit()
        conn.close()
        logger.info(f"Database initialized: {self.db_path}")

    def analyze_sheet(self, sheet, sheet_name: str) -> SheetAnalysis:
        """Analyze a sheet to understand its structure."""
        has_complex = 'complex' in sheet_name.lower()
        year_columns = {}
        sections = []
        current_section = None

        # Scan first 10 rows for year headers (years may be in row 2 or later)
        # Scan up to 200 columns (balance sheets can have many years)
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            for col_idx in range(1, min(200, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                normalized = normalize_marketing_year(cell_value)
                if normalized and is_historical_year(normalized):
                    year_columns[normalized] = col_idx

        # Scan rows for sections and structure
        for row_idx in range(1, min(100, sheet.max_row + 1)):
            col_a = sheet.cell(row=row_idx, column=1).value
            col_b = sheet.cell(row=row_idx, column=2).value

            if col_a:
                col_a_str = str(col_a).strip()

                # Check for section header (ALL CAPS, not a month)
                if col_a_str.isupper() and len(col_a_str) > 3 and not is_month(col_a_str):
                    current_section = {
                        'name': col_a_str,
                        'start_row': row_idx,
                        'type': 'unknown'
                    }
                    sections.append(current_section)

                # Check if this row has months (monthly data section)
                elif is_month(col_a_str) and current_section:
                    current_section['type'] = 'monthly'

                # Check if row looks like a metric (has values in year columns)
                elif year_columns and current_section:
                    has_numeric = False
                    for year, col in list(year_columns.items())[:3]:
                        val = sheet.cell(row=row_idx, column=col).value
                        if isinstance(val, (int, float)):
                            has_numeric = True
                            break
                    if has_numeric:
                        if current_section['type'] == 'unknown':
                            current_section['type'] = 'annual'

        # Get data preview
        preview = []
        if year_columns:
            sorted_years = sorted(year_columns.keys())[-5:]  # Last 5 years
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                row_data = {'row': row_idx, 'col_a': sheet.cell(row=row_idx, column=1).value}
                for year in sorted_years:
                    col = year_columns[year]
                    row_data[year] = sheet.cell(row=row_idx, column=col).value
                preview.append(row_data)

        return SheetAnalysis(
            name=sheet_name,
            has_complex=has_complex,
            year_columns=year_columns,
            sections=sections,
            row_count=sheet.max_row,
            data_preview=preview
        )

    def analyze_file(self, filepath: Path) -> Dict[str, Any]:
        """Analyze an Excel file and report its structure."""
        logger.info(f"Analyzing: {filepath.name}")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        analysis = {
            'file': filepath.name,
            'commodity': detect_commodity_from_filename(filepath.name),
            'sheets': [],
            'complex_sheets': [],
            'total_year_columns': 0,
            'total_sections': 0
        }

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Skip chart sheets (they don't have data)
            if not hasattr(sheet, 'max_row'):
                logger.info(f"  Skipping chart sheet: {sheet_name}")
                continue

            sheet_analysis = self.analyze_sheet(sheet, sheet_name)

            sheet_info = {
                'name': sheet_name,
                'has_complex': sheet_analysis.has_complex,
                'country': detect_country_from_sheet(sheet_name),
                'year_count': len(sheet_analysis.year_columns),
                'year_range': f"{min(sheet_analysis.year_columns.keys()) if sheet_analysis.year_columns else 'N/A'} to {max(sheet_analysis.year_columns.keys()) if sheet_analysis.year_columns else 'N/A'}",
                'sections': [s['name'] for s in sheet_analysis.sections],
                'section_types': {s['name']: s['type'] for s in sheet_analysis.sections},
                'rows': sheet_analysis.row_count,
                'preview': sheet_analysis.data_preview[:3]
            }

            analysis['sheets'].append(sheet_info)

            if sheet_analysis.has_complex:
                analysis['complex_sheets'].append(sheet_name)

            analysis['total_year_columns'] += len(sheet_analysis.year_columns)
            analysis['total_sections'] += len(sheet_analysis.sections)

        wb.close()

        return analysis

    def extract_sheet(self, sheet, sheet_name: str, source_file: str) -> List[ExtractedData]:
        """Extract data from a single sheet."""
        extracted = []

        sheet_analysis = self.analyze_sheet(sheet, sheet_name)

        if not sheet_analysis.year_columns:
            logger.warning(f"  No year columns found in {sheet_name}")
            return extracted

        commodity = detect_commodity_from_filename(source_file)
        country = detect_country_from_sheet(sheet_name)

        # Filter to only historical years
        historical_years = {y: c for y, c in sheet_analysis.year_columns.items()
                          if is_historical_year(y)}

        if not historical_years:
            logger.warning(f"  No historical years found in {sheet_name}")
            return extracted

        logger.info(f"  Found {len(historical_years)} historical year columns")

        current_section = "General"

        # Process each row
        for row_idx in range(1, sheet.max_row + 1):
            col_a = sheet.cell(row=row_idx, column=1).value

            if col_a is None:
                continue

            col_a_str = str(col_a).strip()

            # Update current section if we hit a section header
            if col_a_str.isupper() and len(col_a_str) > 3 and not is_month(col_a_str):
                # Check if this looks like a section header (next row is blank or has months)
                next_row_a = sheet.cell(row=row_idx + 1, column=1).value if row_idx + 1 <= sheet.max_row else None
                if next_row_a is None or is_month(str(next_row_a).strip() if next_row_a else ''):
                    current_section = col_a_str
                    continue

            # Skip if this is just years in the row header (common pattern)
            if is_marketing_year(col_a_str):
                continue

            # Try to extract values for each year
            has_data = False
            for year, col_idx in historical_years.items():
                cell_value = sheet.cell(row=row_idx, column=col_idx).value

                if cell_value is not None and isinstance(cell_value, (int, float)):
                    has_data = True
                    extracted.append(ExtractedData(
                        source_file=source_file,
                        sheet_name=sheet_name,
                        commodity=commodity,
                        country=country,
                        section=current_section,
                        metric=col_a_str,
                        marketing_year=year,
                        value=float(cell_value)
                    ))

            # Log progress for rows with data
            if has_data and len(extracted) % 1000 == 0:
                logger.info(f"    Extracted {len(extracted)} data points...")

        return extracted

    def extract_file(self, filepath: Path, complex_only: bool = True) -> int:
        """Extract all data from an Excel file."""
        logger.info(f"\n{'='*60}")
        logger.info(f"Extracting: {filepath.name}")
        logger.info(f"{'='*60}")

        start_time = datetime.now()

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        all_extracted = []
        sheets_processed = 0

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Skip chart sheets (they don't have data)
            if not hasattr(sheet, 'max_row'):
                logger.info(f"  Skipping chart sheet: {sheet_name}")
                continue

            # Filter to Complex sheets if requested
            if complex_only and 'complex' not in sheet_name.lower():
                logger.info(f"  Skipping: {sheet_name} (no 'Complex' in name)")
                continue

            logger.info(f"\n  Processing: {sheet_name}")

            extracted = self.extract_sheet(sheet, sheet_name, filepath.name)
            all_extracted.extend(extracted)
            sheets_processed += 1

            logger.info(f"    → Extracted {len(extracted)} data points")

        wb.close()

        # Save to database
        if all_extracted:
            self.save_to_database(all_extracted)

        # Log extraction
        elapsed = (datetime.now() - start_time).total_seconds()
        logger.info(f"\n  Completed in {elapsed:.1f}s")
        logger.info(f"  Sheets processed: {sheets_processed}")
        logger.info(f"  Total data points: {len(all_extracted)}")

        return len(all_extracted)

    def save_to_database(self, data: List[ExtractedData]):
        """Save extracted data to database."""
        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()

        inserted = 0
        for item in data:
            try:
                cursor.execute("""
                    INSERT OR REPLACE INTO commodity_balance_sheets
                    (source_file, sheet_name, commodity, country, section,
                     metric, marketing_year, value, unit)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    item.source_file,
                    item.sheet_name,
                    item.commodity,
                    item.country,
                    item.section,
                    item.metric,
                    item.marketing_year,
                    item.value,
                    item.unit
                ))
                inserted += 1
            except Exception as e:
                logger.warning(f"Failed to insert: {e}")

        conn.commit()
        conn.close()

        logger.info(f"  Saved {inserted} rows to database")

    def extract_all_target_files(self) -> Dict[str, int]:
        """Extract from all target balance sheet files."""
        results = {}

        # Find files matching target patterns
        oilseeds_dir = MODELS_DIR / "Oilseeds"

        if not oilseeds_dir.exists():
            logger.error(f"Oilseeds directory not found: {oilseeds_dir}")
            return results

        for target in TARGET_FILES:
            # Find matching file
            matching_files = list(oilseeds_dir.glob(f"*{target}*.xlsx"))

            if not matching_files:
                logger.warning(f"No file found matching: {target}")
                continue

            filepath = matching_files[0]
            try:
                rows = self.extract_file(filepath, complex_only=True)
                results[filepath.name] = rows
            except Exception as e:
                logger.error(f"Failed to extract {filepath.name}: {e}")
                results[filepath.name] = -1

        return results

    def get_status(self) -> Dict[str, Any]:
        """Get extraction status and database stats."""
        if not self.db_path.exists():
            return {
                'database_exists': False,
                'message': 'Run --extract-all to start extraction'
            }

        conn = sqlite3.connect(str(self.db_path))
        cursor = conn.cursor()

        status = {
            'database_exists': True,
            'database_size_mb': self.db_path.stat().st_size / (1024*1024),
            'tables': {}
        }

        # Get row counts
        try:
            cursor.execute("SELECT COUNT(*) FROM commodity_balance_sheets")
            status['total_data_points'] = cursor.fetchone()[0]
        except:
            status['total_data_points'] = 0

        # Get breakdown by commodity
        try:
            cursor.execute("""
                SELECT commodity, COUNT(*) as count
                FROM commodity_balance_sheets
                GROUP BY commodity
                ORDER BY count DESC
            """)
            status['by_commodity'] = {r[0]: r[1] for r in cursor.fetchall()}
        except:
            status['by_commodity'] = {}

        # Get breakdown by country
        try:
            cursor.execute("""
                SELECT country, COUNT(*) as count
                FROM commodity_balance_sheets
                GROUP BY country
                ORDER BY count DESC
                LIMIT 10
            """)
            status['by_country'] = {r[0]: r[1] for r in cursor.fetchall()}
        except:
            status['by_country'] = {}

        # Get year range
        try:
            cursor.execute("""
                SELECT MIN(marketing_year), MAX(marketing_year)
                FROM commodity_balance_sheets
            """)
            row = cursor.fetchone()
            status['year_range'] = f"{row[0]} to {row[1]}" if row[0] else "No data"
        except:
            status['year_range'] = "Unknown"

        # Get source files
        try:
            cursor.execute("""
                SELECT DISTINCT source_file FROM commodity_balance_sheets
            """)
            status['source_files'] = [r[0] for r in cursor.fetchall()]
        except:
            status['source_files'] = []

        conn.close()

        return status


def find_balance_sheet_files() -> List[Path]:
    """Find all balance sheet Excel files."""
    files = []

    for dir_name in ['Oilseeds', 'Biofuels', 'Feed Grains', 'Fats and Greases']:
        dir_path = MODELS_DIR / dir_name
        if dir_path.exists():
            for f in dir_path.glob("*.xlsx"):
                if 'balance' in f.name.lower() and not f.name.startswith('~'):
                    files.append(f)

    return files


def main():
    parser = argparse.ArgumentParser(description='Balance Sheet Data Extractor')
    parser.add_argument('--analyze', type=str, help='Analyze a specific Excel file')
    parser.add_argument('--extract', type=str, help='Extract data from a specific file')
    parser.add_argument('--extract-all', action='store_true', help='Extract from all target files')
    parser.add_argument('--list', action='store_true', help='List all balance sheet files')
    parser.add_argument('--status', action='store_true', help='Show database status')
    parser.add_argument('--init', action='store_true', help='Initialize database only')
    parser.add_argument('--all-sheets', action='store_true', help='Process all sheets, not just Complex')

    args = parser.parse_args()

    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        return

    extractor = BalanceSheetExtractor()

    if args.init:
        extractor.init_database()
        print(f"Database initialized: {DB_PATH}")
        return

    if args.status:
        status = extractor.get_status()
        print("\n" + "="*60)
        print("  DATABASE STATUS")
        print("="*60)

        if not status['database_exists']:
            print(f"\n  {status['message']}")
        else:
            print(f"\n  Database: {DB_PATH}")
            print(f"  Size: {status['database_size_mb']:.2f} MB")
            print(f"  Total data points: {status.get('total_data_points', 0):,}")
            print(f"  Year range: {status.get('year_range', 'N/A')}")

            print("\n  By Commodity:")
            for comm, count in status.get('by_commodity', {}).items():
                print(f"    {comm}: {count:,}")

            print("\n  By Country (top 10):")
            for country, count in status.get('by_country', {}).items():
                print(f"    {country}: {count:,}")

            print("\n  Source Files:")
            for f in status.get('source_files', []):
                print(f"    - {f}")
        print()
        return

    if args.list:
        files = find_balance_sheet_files()
        print(f"\nFound {len(files)} balance sheet files:\n")
        for f in sorted(files):
            print(f"  - {f.relative_to(PROJECT_ROOT)}")
        print()
        return

    if args.analyze:
        filepath = Path(args.analyze)
        if not filepath.exists():
            # Try relative to project root
            filepath = PROJECT_ROOT / args.analyze
        if not filepath.exists():
            # Try in Models/Oilseeds
            filepath = MODELS_DIR / "Oilseeds" / args.analyze

        if not filepath.exists():
            print(f"File not found: {args.analyze}")
            return

        analysis = extractor.analyze_file(filepath)

        print("\n" + "="*60)
        print(f"  FILE ANALYSIS: {analysis['file']}")
        print("="*60)
        print(f"\n  Detected Commodity: {analysis['commodity']}")
        print(f"  Total Sheets: {len(analysis['sheets'])}")
        print(f"  Complex Sheets: {len(analysis['complex_sheets'])}")

        if analysis['complex_sheets']:
            print(f"\n  Complex Sheets to Process:")
            for name in analysis['complex_sheets']:
                print(f"    ✓ {name}")

        print(f"\n  Sheet Details:")
        for sheet in analysis['sheets']:
            marker = "✓" if sheet['has_complex'] else "○"
            print(f"\n    {marker} {sheet['name']}")
            print(f"      Country: {sheet['country']}")
            print(f"      Years: {sheet['year_count']} columns ({sheet['year_range']})")
            print(f"      Sections: {', '.join(sheet['sections'][:5])}" +
                  (f"... +{len(sheet['sections'])-5} more" if len(sheet['sections']) > 5 else ""))

            if sheet['preview']:
                print(f"      Preview (first 3 rows):")
                for row in sheet['preview']:
                    preview_str = json.dumps(row, default=str)[:100]
                    print(f"        {preview_str}...")
        print()
        return

    if args.extract:
        filepath = Path(args.extract)
        if not filepath.exists():
            filepath = PROJECT_ROOT / args.extract
        if not filepath.exists():
            filepath = MODELS_DIR / "Oilseeds" / args.extract

        if not filepath.exists():
            print(f"File not found: {args.extract}")
            return

        extractor.init_database()
        rows = extractor.extract_file(filepath, complex_only=not args.all_sheets)
        print(f"\nExtracted {rows:,} data points to {DB_PATH}")
        return

    if args.extract_all:
        extractor.init_database()
        results = extractor.extract_all_target_files()

        print("\n" + "="*60)
        print("  EXTRACTION COMPLETE")
        print("="*60)

        total = sum(r for r in results.values() if r > 0)
        for filename, rows in results.items():
            status = f"{rows:,} rows" if rows >= 0 else "FAILED"
            print(f"  {filename}: {status}")

        print(f"\n  Total: {total:,} data points")
        print(f"  Database: {DB_PATH}")
        print()
        return

    # Default: show help
    parser.print_help()


if __name__ == '__main__':
    main()
