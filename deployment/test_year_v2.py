#!/usr/bin/env python3
"""Quick test to debug year detection - v2"""
import openpyxl
import re

PATTERN_2DIGIT = re.compile(r'^(\d{2})/(\d{2})\s*$')
PATTERN_4DIGIT = re.compile(r'^(\d{4})/(\d{2})\s*$')

MIN_YEAR = 1965
MAX_YEAR = 2023

def is_marketing_year(value):
    if value is None:
        return None
    val_str = str(value).strip()

    match = PATTERN_4DIGIT.match(val_str)
    if match:
        start_year = int(match.group(1))
        end_year_short = int(match.group(2))
        end_year = (start_year // 100) * 100 + end_year_short
        if end_year_short < (start_year % 100):
            end_year += 100
        return (start_year, end_year)

    match = PATTERN_2DIGIT.match(val_str)
    if match:
        start_short = int(match.group(1))
        end_short = int(match.group(2))
        if start_short >= 50:
            start_year = 1900 + start_short
        else:
            start_year = 2000 + start_short
        if end_short >= 50:
            end_year = 1900 + end_short
        else:
            end_year = 2000 + end_short
        if end_year < start_year:
            end_year += 100
        return (start_year, end_year)

    return None

def normalize_marketing_year(value):
    years = is_marketing_year(value)
    if years:
        start_year, end_year = years
        return f"{start_year}/{end_year % 100:02d}"
    return None

def is_historical_year(marketing_year):
    years = is_marketing_year(marketing_year)
    if years:
        start_year = years[0]
        return MIN_YEAR <= start_year <= MAX_YEAR
    return False

filepath = 'Models/Oilseeds/World Soybean Balance Sheets.xlsx'

print("=" * 60)
print("YEAR DETECTION DEBUG - V2")
print("=" * 60)

wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
sheet = wb['Brazil Soy Complex']

print(f"\nSheet: Brazil Soy Complex")
print(f"max_row: {sheet.max_row}, max_column: {sheet.max_column}")

# Test the full pipeline on a few year values
print("\n--- Testing full pipeline on sample years ---")
test_values = ['64/65 ', '65/66 ', '99/00 ', '23/24 ']
for val in test_values:
    normalized = normalize_marketing_year(val)
    is_hist = is_historical_year(normalized) if normalized else False
    print(f"  '{val}' -> '{normalized}' -> historical={is_hist}")

# Now test what the analyze_sheet loop would find
print("\n--- Simulating analyze_sheet loop ---")
year_columns = {}
for row_idx in range(1, min(11, sheet.max_row + 1)):
    for col_idx in range(1, min(200, sheet.max_column + 1)):
        cell_value = sheet.cell(row=row_idx, column=col_idx).value
        normalized = normalize_marketing_year(cell_value)
        if normalized and is_historical_year(normalized):
            year_columns[normalized] = col_idx

print(f"Found {len(year_columns)} year columns that pass filter")
if year_columns:
    sorted_years = sorted(year_columns.keys())
    print(f"Range: {sorted_years[0]} to {sorted_years[-1]}")
    print(f"First 5: {sorted_years[:5]}")
    print(f"Last 5: {sorted_years[-5:]}")
else:
    print("NO YEARS FOUND!")
    print("\nDebugging row 2:")
    for col in range(1, 21):
        val = sheet.cell(row=2, column=col).value
        if val:
            normalized = normalize_marketing_year(val)
            is_hist = is_historical_year(normalized) if normalized else None
            print(f"  Col {col}: '{val}' -> '{normalized}' -> hist={is_hist}")

wb.close()
print("\n" + "=" * 60)
