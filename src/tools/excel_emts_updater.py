"""
EPA EMTS Monthly RIN Generation → Excel Updater
=================================================
Reads downloaded EPA CSV and updates EMTS Data.xlsx workbook.

Source: EPA EMTS interactive table at
https://www.epa.gov/fuels-registration-reporting-and-compliance-help/rins-generated-transactions

Usage:
    python excel_emts_updater.py --file "path/to/EMTS Data.xlsx" --csv "path/to/monthly_rin_generation.csv"
    python excel_emts_updater.py --file "..." --csv "..." --latest 12
    python excel_emts_updater.py --file "..." --csv "..." --all

Can also be called from VBA (see EMTSDataUpdater.bas for full implementation)
"""

import argparse
import csv
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import logging

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CONFIGURATION
# =============================================================================

LOG_FILE = Path(__file__).parent / 'emts_updater.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# =============================================================================
# TAB CONFIGURATION
# =============================================================================

# Each tab's sheet name, header row count, and first data row
TAB_CONFIG = {
    'D3_gallon': {'tab': 'Monthly D3 Gallon Data', 'data_start_row': 4, 'value_field': 'volume'},
    'D3_rin':    {'tab': 'Monthly D3 RIN Data',    'data_start_row': 4, 'value_field': 'rins'},
    'D4_gallon': {'tab': 'Monthly D4 Gallon Data', 'data_start_row': 2, 'value_field': 'volume'},
    'D4_rin':    {'tab': 'Monthly D4 RIN Data',    'data_start_row': 4, 'value_field': 'rins'},
    'D5_gallon': {'tab': 'Monthly D5 Gallon Data', 'data_start_row': 4, 'value_field': 'volume'},
    'D5_rin':    {'tab': 'Monthly D5 RIN Data',    'data_start_row': 4, 'value_field': 'rins'},
    'D6_gallon': {'tab': 'Monthly D6 Gallon Data', 'data_start_row': 4, 'value_field': 'volume'},
    'D6_rin':    {'tab': 'Monthly D6 RIN Data',    'data_start_row': 4, 'value_field': 'rins'},
    'D7_gallon': {'tab': 'Monthly D7 Gallon Data', 'data_start_row': 4, 'value_field': 'volume'},
    'D7_rin':    {'tab': 'Monthly D7 RIN Data',    'data_start_row': 4, 'value_field': 'rins'},
}


# =============================================================================
# COLUMN MAPPINGS: (EPA Fuel Category, EPA Producer Type) → Column Letter
#
# Every data column in the spreadsheet is listed. Multiple EPA fuel category
# names that map to the same column handle name evolution (e.g. CNG → CNG-CNG).
# Values mapping to the same column in the same month are SUMMED.
#
# Formula columns, Total columns, and Annual Cumulative columns are excluded.
# =============================================================================

# --- D3: Cellulosic Biofuel ---
# Gallon and RIN tabs share the same column layout
D3_MAP = {
    # Cellulosic Ethanol (cols B-C)
    ('Cellulosic Ethanol (EV 1.0)', 'Domestic'):  'B',
    ('Cellulosic Ethanol (EV 1.0)', 'Importer'):  'C',
    # Cellulosic Renewable Gasoline Blendstock (cols E-F)
    ('Cellulosic Renewable Gasoline Blendstock (EV application required)', 'Domestic'):  'E',
    ('Cellulosic Renewable Gasoline Blendstock (EV application required)', 'Importer'):  'F',
    # Renewable Compressed Natural Gas (cols H-I)
    # "Renewable Compressed Natural Gas - CNG" is a 2025 rename of the original
    ('Renewable Compressed Natural Gas', 'Domestic'):       'H',
    ('Renewable Compressed Natural Gas - CNG', 'Domestic'): 'H',
    ('Renewable Compressed Natural Gas', 'Importer'):       'I',
    # Renewable Liquefied Natural Gas (cols K-L)
    # "Renewable Natural Gas - RNG" replaces "Renewable Liquefied Natural Gas" starting 2024
    ('Renewable Liquefied Natural Gas', 'Domestic'):  'K',
    ('Renewable Natural Gas - RNG', 'Domestic'):      'K',
    ('Renewable Liquefied Natural Gas', 'Importer'):  'L',
    ('Renewable Natural Gas - RNG', 'Importer'):      'L',
}

# --- D4 Gallon: Biomass-based Diesel (1-row header, data row 2) ---
D4_GALLON_MAP = {
    # Biodiesel (cols B-D), col E = Total formula
    ('Biodiesel (EV 1.5)', 'Domestic'):            'B',
    ('Biodiesel (EV 1.5)', 'Foreign Generation'):  'C',
    ('Biodiesel (EV 1.5)', 'Importer'):            'D',
    # NERD summary cols F-I are formulas pulling from breakout cols S-AD
    # Renewable Jet Fuel - Foreign Gen manual (col K)
    ('Renewable Jet Fuel (EV 1.6)', 'Foreign Generation'): 'K',
    # Renewable Heating Oil - Domestic manual (col M)
    ('Renewable Heating Oil (EV 1.6)', 'Domestic'): 'M',
    # NERD Domestic Breakout (cols S-U), col V = Total formula
    ('Non-ester Renewable Diesel (EV 1.5)', 'Domestic'): 'S',
    ('Non-ester Renewable Diesel (EV 1.6)', 'Domestic'): 'T',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Domestic'): 'U',
    # NERD Foreign Generation Breakout (cols W-Y), col Z = Total formula
    ('Non-ester Renewable Diesel (EV 1.5)', 'Foreign Generation'): 'W',
    ('Non-ester Renewable Diesel (EV 1.6)', 'Foreign Generation'): 'X',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Foreign Generation'): 'Y',
    # NERD Importer Breakout (cols AA-AC), col AD = Total formula
    ('Non-ester Renewable Diesel (EV 1.5)', 'Importer'): 'AA',
    ('Non-ester Renewable Diesel (EV 1.6)', 'Importer'): 'AB',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Importer'): 'AC',
    # RJF Domestic Breakout (cols AE-AF), col AG = Total formula
    ('Renewable Jet Fuel (EV 1.6)', 'Domestic'): 'AE',
    ('Renewable Jet Fuel (EV 1.0)', 'Domestic'):  'AF',
}

# --- D4 RIN: Biomass-based Diesel (3-row header, data row 4) ---
# Different layout from D4 Gallon: RJF Foreign Gen → col AH, RHO → col AI
D4_RIN_MAP = {
    # Biodiesel (cols B-D), col E = Total formula
    ('Biodiesel (EV 1.5)', 'Domestic'):            'B',
    ('Biodiesel (EV 1.5)', 'Foreign Generation'):  'C',
    ('Biodiesel (EV 1.5)', 'Importer'):            'D',
    # NERD summary cols F-I are formulas
    # RJF col J = formula from breakout, col K = Importer (no EPA data), col L = formula
    # RHO cols M-N are formulas
    # NERD Domestic Breakout (cols S-U), col V = Total formula
    ('Non-ester Renewable Diesel (EV 1.5)', 'Domestic'): 'S',
    ('Non-ester Renewable Diesel (EV 1.6)', 'Domestic'): 'T',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Domestic'): 'U',
    # NERD Foreign Generation Breakout (cols W-Y), col Z = Total formula
    ('Non-ester Renewable Diesel (EV 1.5)', 'Foreign Generation'): 'W',
    ('Non-ester Renewable Diesel (EV 1.6)', 'Foreign Generation'): 'X',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Foreign Generation'): 'Y',
    # NERD Importer Breakout (cols AA-AC), col AD = Total formula
    ('Non-ester Renewable Diesel (EV 1.5)', 'Importer'): 'AA',
    ('Non-ester Renewable Diesel (EV 1.6)', 'Importer'): 'AB',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Importer'): 'AC',
    # RJF Domestic Breakout (cols AE-AF), col AG = Total formula
    ('Renewable Jet Fuel (EV 1.6)', 'Domestic'): 'AE',
    ('Renewable Jet Fuel (EV 1.0)', 'Domestic'):  'AF',
    # RJF Foreign Gen manual (col AH) — different position from D4 Gallon
    ('Renewable Jet Fuel (EV 1.6)', 'Foreign Generation'): 'AH',
    # RHO Domestic manual enter (col AI) — different position from D4 Gallon
    ('Renewable Heating Oil (EV 1.6)', 'Domestic'): 'AI',
}

# --- D5: Advanced Biofuel ---
# Gallon and RIN tabs share the same column layout
D5_MAP = {
    # Non-Cellulosic Ethanol (cols B-C), col D = Total formula
    ('Non-cellulosic Ethanol (EV 1.0)', 'Domestic'):  'B',
    ('Non-cellulosic Ethanol (EV 1.0)', 'Importer'):  'C',
    # NERD col E = formula from breakout (=AB4), col F = Importer (no EPA data), col G = formula
    # Renewable Heating Oil (cols H-I), col J = Total formula
    ('Renewable Heating Oil (EV 1.6)', 'Domestic'): 'H',
    # Biogas (cols K-L), col M = Total formula
    ('Biogas (77,000 Btu LHV/1 gallon)', 'Domestic'): 'K',
    # Naphtha col N = formula from breakout (=AE4), col O = Importer (no EPA data), col P = formula
    # Renewable CNG (cols Q-R), col S = Total formula
    # "Renewable Compressed Natural Gas - CNG" is 2025 rename
    # "Renewable Natural Gas - RNG" also maps here (D5 has no LNG column)
    ('Renewable Compressed Natural Gas', 'Domestic'):       'Q',
    ('Renewable Compressed Natural Gas - CNG', 'Domestic'): 'Q',
    ('Renewable Natural Gas - RNG', 'Domestic'):            'Q',
    # LPG (cols T-U), col V = Total formula
    ('LPG (EV 1.1)', 'Domestic'): 'T',
    # D5 Totals cols W-Y = formulas
    # NERD Breakout Domestic (cols Z-AA), col AB = Total formula
    ('Non-ester Renewable Diesel (EV 1.6)', 'Domestic'): 'Z',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Domestic'): 'AA',
    # Naphtha Breakout Domestic manual (cols AC-AD), col AE = Total formula
    ('Naphtha (EV 1.4)', 'Domestic'): 'AC',
    ('Naphtha (EV 1.5)', 'Domestic'): 'AD',
}

# --- D6: Renewable Fuel / Conventional Ethanol ---
# Gallon and RIN tabs share the same column layout
D6_MAP = {
    # Biodiesel (cols B-C), col D = Total formula
    ('Biodiesel (EV 1.5)', 'Domestic'):  'B',
    ('Biodiesel (EV 1.5)', 'Importer'):  'C',
    # Non-cellulosic Ethanol (cols E-F), col G = Total formula
    ('Non-cellulosic Ethanol (EV 1.0)', 'Domestic'):  'E',
    ('Non-cellulosic Ethanol (EV 1.0)', 'Importer'):  'F',
    # Butanol (cols H-I), col J = Total formula
    ('Butanol (EV 1.3)', 'Domestic'): 'H',
    # Non-ester Renewable Diesel (cols K-M), col N = Total formula
    # D6 NERD has 4 columns: Dom, Foreign Gen, Importer, Total
    ('Non-ester Renewable Diesel (EV 1.7)', 'Domestic'):            'K',
    ('Non-ester Renewable Diesel (EV 1.7)', 'Foreign Generation'):  'L',
    # Renewable Jet Fuel (cols O-Q), col R = Total formula
    # D6 RJF has 4 columns: Dom, Foreign Gen, Importer, Total
    ('Renewable Jet Fuel (EV 1.7)', 'Domestic'): 'O',
    # Renewable Gasoline (cols S-T), col U = Total formula
    ('Renewable Gasoline (EV 1.5)', 'Domestic'): 'S',
}

# --- D7: Cellulosic Diesel ---
# Gallon and RIN tabs share the same column layout
D7_MAP = {
    # Summary cols B-G are formulas pulling from breakout cols K-P
    # Cellulosic Diesel Domestic Breakout (col K = EV App Required, col L = Other)
    ('Cellulosic Diesel (EV application required)', 'Domestic'): 'K',
    # Cellulosic Heating Oil Importer Breakout (col N = EV App Required, col O = Other)
    ('Cellulosic Heating Oil (EV application required)', 'Importer'): 'N',
}

# Master map: maps (d_code, tab_type) → column mapping dict
COLUMN_MAPS = {
    ('D3', 'gallon'): D3_MAP,
    ('D3', 'rin'):    D3_MAP,       # Same layout as gallon
    ('D4', 'gallon'): D4_GALLON_MAP,
    ('D4', 'rin'):    D4_RIN_MAP,   # Different layout from gallon
    ('D5', 'gallon'): D5_MAP,
    ('D5', 'rin'):    D5_MAP,       # Same layout as gallon
    ('D6', 'gallon'): D6_MAP,
    ('D6', 'rin'):    D6_MAP,       # Same layout as gallon
    ('D7', 'gallon'): D7_MAP,
    ('D7', 'rin'):    D7_MAP,       # Same layout as gallon
}

# Maps tab config key → (d_code, tab_type) for looking up column maps
TAB_TO_DCODE = {
    'D3_gallon': ('D3', 'gallon'), 'D3_rin': ('D3', 'rin'),
    'D4_gallon': ('D4', 'gallon'), 'D4_rin': ('D4', 'rin'),
    'D5_gallon': ('D5', 'gallon'), 'D5_rin': ('D5', 'rin'),
    'D6_gallon': ('D6', 'gallon'), 'D6_rin': ('D6', 'rin'),
    'D7_gallon': ('D7', 'gallon'), 'D7_rin': ('D7', 'rin'),
}


# =============================================================================
# CSV READER
# =============================================================================

def parse_number(s: str) -> int:
    """Parse a number string that may be quoted with embedded commas."""
    if not s:
        return 0
    return int(s.replace(',', '').replace('"', '').strip())


def read_epa_csv(csv_path: str) -> List[Dict]:
    """
    Read EPA monthly RIN generation CSV.

    Returns list of dicts with keys:
        year, month, producer_type, d_code, fuel_category, rins, volume
    """
    records = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                record = {
                    'year': int(row['RIN Year'].replace('"', '').strip()),
                    'month': int(row['Month'].replace('"', '').strip()),
                    'producer_type': row['Producer Type'].replace('"', '').strip(),
                    'd_code': row['Fuel (D Code)'].replace('"', '').strip(),
                    'fuel_category': row['Fuel Category'].replace('"', '').strip(),
                    'rins': parse_number(row['RINs']),
                    'volume': parse_number(row['Volume (Gal.)']),
                }
                records.append(record)
            except (ValueError, KeyError) as e:
                logger.warning(f"Skipping malformed row: {row} — {e}")
    return records


# =============================================================================
# AGGREGATION
# =============================================================================

def aggregate_data(
    records: List[Dict],
    d_code: str,
    column_map: Dict[Tuple[str, str], str],
    value_field: str,
    month_filter: set = None,
) -> Dict[Tuple[int, int], Dict[str, int]]:
    """
    Aggregate EPA CSV records into column-level values per month.

    Args:
        records: Parsed CSV records
        d_code: Filter to this D-code (e.g. 'D4')
        column_map: (fuel_category, producer_type) → column_letter
        value_field: 'volume' or 'rins'
        month_filter: Optional set of (year, month) tuples to include

    Returns:
        {(year, month): {column_letter: summed_value}}
    """
    result = defaultdict(lambda: defaultdict(int))
    unmapped = set()

    for rec in records:
        if rec['d_code'] != d_code:
            continue

        ym = (rec['year'], rec['month'])
        if month_filter and ym not in month_filter:
            continue

        key = (rec['fuel_category'], rec['producer_type'])
        col = column_map.get(key)

        if col is None:
            unmapped.add(key)
            continue

        result[ym][col] += rec[value_field]

    if unmapped:
        for fc, pt in sorted(unmapped):
            logger.warning(f"Unmapped {d_code} combination: fuel_category='{fc}', producer_type='{pt}'")

    return dict(result)


# =============================================================================
# EXCEL WRITER
# =============================================================================

def find_row_for_date(ws, year: int, month: int, data_start_row: int) -> int:
    """
    Find the row in column A matching the given year/month.
    Column A contains datetime values (1st of month).
    Returns row number, or 0 if not found.
    """
    last_row = ws.max_row
    for row in range(data_start_row, last_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val is None:
            continue
        if isinstance(cell_val, datetime):
            if cell_val.year == year and cell_val.month == month:
                return row
        elif isinstance(cell_val, (int, float)) and cell_val > 30000:
            # Excel date serial number
            try:
                dt = datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(cell_val))
                if dt.year == year and dt.month == month:
                    return row
            except (ValueError, OverflowError):
                pass
    return 0


def update_tab(
    wb,
    tab_name: str,
    aggregated_data: Dict[Tuple[int, int], Dict[str, int]],
    data_start_row: int,
) -> int:
    """
    Write aggregated data to a worksheet tab.

    Returns number of cells updated.
    """
    if tab_name not in wb.sheetnames:
        logger.warning(f"Sheet '{tab_name}' not found in workbook — skipping")
        return 0

    ws = wb[tab_name]
    cells_updated = 0

    for (year, month), col_values in sorted(aggregated_data.items()):
        row = find_row_for_date(ws, year, month, data_start_row)
        if row == 0:
            logger.debug(f"No row found for {year}-{month:02d} in '{tab_name}'")
            continue

        for col_letter, value in col_values.items():
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=row, column=col_idx, value=value)
            cells_updated += 1

    return cells_updated


# =============================================================================
# MAIN
# =============================================================================

def get_latest_months(records: List[Dict], n: int) -> set:
    """Get the set of (year, month) tuples for the latest N months in the data."""
    all_months = sorted({(r['year'], r['month']) for r in records}, reverse=True)
    return set(all_months[:n])


def main():
    parser = argparse.ArgumentParser(
        description='EPA EMTS Monthly RIN Generation → Excel Updater'
    )
    parser.add_argument('--file', required=True, help='Path to EMTS Data.xlsx workbook')
    parser.add_argument('--csv', required=True, help='Path to EPA monthly_rin_generation.csv')
    parser.add_argument('--latest', type=int, default=6,
                        help='Update only the latest N months (default: 6)')
    parser.add_argument('--all', action='store_true',
                        help='Update all months in the CSV (overrides --latest)')
    args = parser.parse_args()

    # Validate paths
    csv_path = Path(args.csv)
    xl_path = Path(args.file)

    if not csv_path.exists():
        logger.error(f"CSV file not found: {csv_path}")
        sys.exit(1)
    if not xl_path.exists():
        logger.error(f"Excel file not found: {xl_path}")
        sys.exit(1)

    # Read CSV
    logger.info(f"Reading EPA CSV: {csv_path}")
    records = read_epa_csv(str(csv_path))
    logger.info(f"Loaded {len(records)} records")

    if not records:
        logger.error("No records found in CSV")
        sys.exit(1)

    # Determine month filter
    if args.all:
        month_filter = None
        filter_desc = "all months"
    else:
        month_filter = get_latest_months(records, args.latest)
        filter_desc = f"latest {args.latest} months"
    logger.info(f"Update scope: {filter_desc}")

    # Date range in data
    all_months = sorted({(r['year'], r['month']) for r in records})
    first_ym = all_months[0]
    last_ym = all_months[-1]
    logger.info(f"CSV data range: {first_ym[0]}-{first_ym[1]:02d} to {last_ym[0]}-{last_ym[1]:02d}")

    # Open workbook (preserve formulas by not using data_only)
    logger.info(f"Opening workbook: {xl_path}")
    wb = load_workbook(str(xl_path))

    # Process each tab
    total_cells = 0
    tab_results = []

    for tab_key, config in TAB_CONFIG.items():
        d_code, tab_type = TAB_TO_DCODE[tab_key]
        column_map = COLUMN_MAPS[(d_code, tab_type)]
        value_field = config['value_field']
        tab_name = config['tab']
        data_start_row = config['data_start_row']

        # Aggregate data for this tab
        agg = aggregate_data(records, d_code, column_map, value_field, month_filter)

        if not agg:
            tab_results.append((tab_name, 0, 0))
            continue

        # Write to worksheet
        cells = update_tab(wb, tab_name, agg, data_start_row)
        months_written = len(agg)
        total_cells += cells
        tab_results.append((tab_name, cells, months_written))

        logger.info(f"  {tab_name}: {cells} cells across {months_written} months")

    # Save workbook
    logger.info(f"Saving workbook...")
    wb.save(str(xl_path))
    wb.close()

    # Print summary
    print("\n" + "=" * 60)
    print("EMTS Data Update Complete")
    print("=" * 60)
    print(f"CSV: {csv_path.name}")
    print(f"Workbook: {xl_path.name}")
    print(f"Scope: {filter_desc}")
    print(f"Data range: {first_ym[0]}-{first_ym[1]:02d} to {last_ym[0]}-{last_ym[1]:02d}")
    print("-" * 60)
    for tab_name, cells, months in tab_results:
        status = f"{cells:>6} cells, {months:>3} months" if cells > 0 else "  (no data)"
        print(f"  {tab_name:<30} {status}")
    print("-" * 60)
    print(f"  Total cells updated: {total_cells}")
    print("=" * 60)

    logger.info(f"Done. Total cells updated: {total_cells}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
