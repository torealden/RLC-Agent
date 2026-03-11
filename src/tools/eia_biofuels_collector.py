"""
EIA Monthly Biofuels Feedstock & Capacity Collector
====================================================
Downloads and loads EIA Form 819 data from xlsx files into PostgreSQL.

Source: https://www.eia.gov/biofuels/update/
  - table1.xlsx = Operable production capacity (MMGY)
  - table2.xlsx = Feedstock consumption (million lbs/month, 4 sheets)

Usage:
    python eia_biofuels_collector.py --download
    python eia_biofuels_collector.py --table1 path/to/table1.xlsx --table2 path/to/table2.xlsx

Once loaded, use Ctrl+Shift+D in the feedstock workbook to pull into Excel.
"""

import os
import sys
import re
import logging
import argparse

import requests
import openpyxl
import psycopg2

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'rlc_commodities',
    'user': 'postgres',
    'password': 'SoupBoss1',
}

TABLE1_URL = "https://www.eia.gov/biofuels/update/table1.xlsx"
TABLE2_URL = "https://www.eia.gov/biofuels/update/table2.xlsx"

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '..', '..', 'data', 'eia_biofuels')
LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'eia_biofuels_collector.log')

MONTH_NAMES = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4,
    'May': 5, 'June': 6, 'July': 7, 'August': 8,
    'September': 9, 'October': 10, 'November': 11, 'December': 12,
}

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Feedstock name normalization
# ---------------------------------------------------------------------------
def normalize_feedstock_name(raw_name):
    """Strip footnote superscripts and normalize EIA feedstock labels."""
    if not raw_name or not isinstance(raw_name, str):
        return raw_name
    name = raw_name.strip()
    # Strip trailing footnote numbers: "Tallow3" -> "Tallow", "Yellow Grease4" -> "Yellow Grease"
    name = re.sub(r'[\d,]+$', '', name).strip()
    # Normalize specific names to match reference mapping
    name_map = {
        'Canola oil': 'Canola Oil',
        'Corn oil': 'Corn Oil',
        'Soybean oil': 'Soybean Oil',
        'Soybean Oil': 'Soybean Oil',
        'Other vegetable oils': 'Other Vegetable Oil',
        'Yellow Grease': 'Yellow Grease',
        'Agriculture and forestry residues': 'Ag Forestry Residues',
        'Dedicated energy crops': 'Energy Crops',
        'Municipal solid waste': 'Municipal Solid Waste',
        'Yard and food waste': 'Yard Food Waste',
        'Oil from Algae': 'Algae Oil',
        'Other biofuel feedstocks (NESOI)': 'Other NESOI',
    }
    return name_map.get(name, name)


def normalize_plant_type(raw_type):
    """Normalize plant type sub-header."""
    if not raw_type or not isinstance(raw_type, str):
        return 'total'
    t = raw_type.strip().lower()
    if 'biodiesel' in t:
        return 'biodiesel'
    elif 'renewable diesel' in t:
        return 'renewable_diesel'
    elif 'total' in t:
        return 'total'
    return 'total'


# ---------------------------------------------------------------------------
# XLSX Parsing: Table 1 (Capacity)
# ---------------------------------------------------------------------------
def parse_table1(filepath):
    """Parse EIA Table 1 (capacity) xlsx.

    Structure:
      Row 3: Headers — A='Period', B='Biodiesel', D='Fuel Ethanol', F='RD & Other'
      Row 5+: Year header (int), then month rows with values in B, D, F
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    records = []

    # Column mapping: col_index -> biofuel_type (from row 3 headers)
    col_map = {}
    for cell in ws[3]:
        if cell.value and isinstance(cell.value, str) and cell.value.strip() != 'Period':
            # Strip footnote refs
            name = re.sub(r'[\d,]+$', '', cell.value).strip()
            col_map[cell.column] = name

    logger.info(f"Table 1 biofuel types: {list(col_map.values())}")

    current_year = None
    for row in ws.iter_rows(min_row=4, max_col=ws.max_column):
        a_val = row[0].value
        if a_val is None:
            continue

        # Year header row
        if isinstance(a_val, (int, float)) and 2000 <= a_val <= 2100:
            current_year = int(a_val)
            continue

        # Month data row
        if isinstance(a_val, str) and a_val.strip() in MONTH_NAMES and current_year:
            month = MONTH_NAMES[a_val.strip()]
            for col_idx, biofuel_type in col_map.items():
                cell = ws.cell(row=row[0].row, column=col_idx)
                val = cell.value
                if val is not None and isinstance(val, (int, float)):
                    records.append({
                        'year': current_year,
                        'month': month,
                        'biofuel_type': biofuel_type,
                        'capacity_mmgy': float(val),
                    })

    wb.close()
    return records


# ---------------------------------------------------------------------------
# XLSX Parsing: Table 2 (Feedstocks)
# ---------------------------------------------------------------------------
def parse_table2(filepath):
    """Parse all 4 sheets of EIA Table 2 (feedstocks)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == 'table_2c':
            records = parse_table2c(ws, sheet_name)
        else:
            records = parse_table2_simple(ws, sheet_name)
        all_records.extend(records)
        logger.info(f"  {sheet_name}: {len(records)} records parsed")

    wb.close()
    return all_records


def parse_table2_simple(ws, sheet_name):
    """Parse table_2a, table_2b, table_2d — simple structure with no plant_type split.

    Structure:
      Row 4: Headers — feedstock names in odd columns (C, E, G, I, K, M)
      Row 6+: Year header (int), then month rows with values
    """
    records = []

    # Build column -> feedstock_name map from row 4
    col_map = {}
    for cell in ws[4]:
        if cell.value and isinstance(cell.value, str) and cell.value.strip() != 'Period':
            name = normalize_feedstock_name(cell.value)
            if name:
                col_map[cell.column] = name

    if not col_map:
        logger.warning(f"  {sheet_name}: No feedstock columns found in row 4")
        return records

    logger.info(f"  {sheet_name}: Feedstocks: {list(col_map.values())}")

    # Handle "Other" appearing in multiple sheets — disambiguate
    other_label = {
        'table_2a': 'Other Ag',
        'table_2b': 'Other Waste',
        'table_2d': 'Other Recycled',
    }
    for col_idx in col_map:
        if col_map[col_idx] == 'Other' and sheet_name in other_label:
            col_map[col_idx] = other_label[sheet_name]

    current_year = None
    for row in ws.iter_rows(min_row=5, max_col=ws.max_column):
        a_val = row[0].value
        if a_val is None:
            continue

        if isinstance(a_val, (int, float)) and 2000 <= a_val <= 2100:
            current_year = int(a_val)
            continue

        if isinstance(a_val, str) and a_val.strip() in MONTH_NAMES and current_year:
            month = MONTH_NAMES[a_val.strip()]
            for col_idx, feedstock_name in col_map.items():
                cell = ws.cell(row=row[0].row, column=col_idx)
                val = cell.value

                is_withheld = isinstance(val, str) and val.strip().upper() == 'W'
                is_no_data = isinstance(val, str) and val.strip() == '-'

                quantity = None
                if not is_withheld and not is_no_data and val is not None:
                    try:
                        quantity = float(val)
                    except (ValueError, TypeError):
                        pass

                records.append({
                    'year': current_year,
                    'month': month,
                    'source_sheet': sheet_name,
                    'feedstock_name': feedstock_name,
                    'plant_type': 'total',
                    'quantity_mil_lbs': quantity,
                    'is_withheld': is_withheld,
                    'is_no_data': is_no_data,
                })

    return records


def parse_table2c(ws, sheet_name):
    """Parse table_2c (Vegetable Oils) — has plant_type sub-headers.

    Structure:
      Row 4: Vegetable oil type names (merged cells): C4='Canola oil', I4='Corn oil', O4='Soybean Oil'
      Row 5: Plant type sub-headers: C5='Total', E5='Biodiesel Plants', G5='Renewable Diesel Plants', ...
             Plus U5='Other vegetable oils5' (standalone, no split)
      Row 7+: Year header (int), then month rows
    """
    records = []

    # Build column -> (feedstock_name, plant_type) map from rows 4-5
    col_map = {}

    # Row 4: merged cells with veg oil type names
    # Get merged cell ranges for row 4
    veg_oil_ranges = {}  # col_start -> oil_name
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_row == 4 and merge_range.max_row == 4:
            val = ws.cell(merge_range.min_row, merge_range.min_col).value
            if val:
                name = normalize_feedstock_name(val)
                for c in range(merge_range.min_col, merge_range.max_col + 1):
                    veg_oil_ranges[c] = name

    # Also check non-merged cells in row 4
    for cell in ws[4]:
        if cell.value and cell.column not in veg_oil_ranges:
            name = normalize_feedstock_name(cell.value)
            if name and name != 'Period':
                veg_oil_ranges[cell.column] = name

    # Row 5: plant type sub-headers
    for cell in ws[5]:
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip()
            if val == 'Period':
                continue

            col = cell.column
            # Determine the parent veg oil name
            parent_name = veg_oil_ranges.get(col)

            if parent_name:
                pt = normalize_plant_type(val)
                col_map[col] = (parent_name, pt)
            else:
                # Standalone column (e.g., "Other vegetable oils5")
                name = normalize_feedstock_name(val)
                col_map[col] = (name, 'total')

    logger.info(f"  {sheet_name}: Column map: {col_map}")

    current_year = None
    for row_cells in ws.iter_rows(min_row=6, max_col=ws.max_column):
        a_val = row_cells[0].value
        if a_val is None:
            continue

        if isinstance(a_val, (int, float)) and 2000 <= a_val <= 2100:
            current_year = int(a_val)
            continue

        if isinstance(a_val, str) and a_val.strip() in MONTH_NAMES and current_year:
            month = MONTH_NAMES[a_val.strip()]
            for col_idx, (feedstock_name, plant_type) in col_map.items():
                cell = ws.cell(row=row_cells[0].row, column=col_idx)
                val = cell.value

                is_withheld = isinstance(val, str) and val.strip().upper() == 'W'
                is_no_data = isinstance(val, str) and val.strip() == '-'

                quantity = None
                if not is_withheld and not is_no_data and val is not None:
                    try:
                        quantity = float(val)
                    except (ValueError, TypeError):
                        pass

                records.append({
                    'year': current_year,
                    'month': month,
                    'source_sheet': sheet_name,
                    'feedstock_name': feedstock_name,
                    'plant_type': plant_type,
                    'quantity_mil_lbs': quantity,
                    'is_withheld': is_withheld,
                    'is_no_data': is_no_data,
                })

    return records


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------
def download_file(url, dest_path):
    """Download file from URL."""
    logger.info(f"Downloading {url}")
    resp = requests.get(url, timeout=30, headers={
        'User-Agent': 'Mozilla/5.0 (RLC-Agent Data Collector)'
    })
    resp.raise_for_status()
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with open(dest_path, 'wb') as f:
        f.write(resp.content)
    logger.info(f"  Saved to {dest_path} ({len(resp.content):,} bytes)")
    return dest_path


# ---------------------------------------------------------------------------
# Database Loading
# ---------------------------------------------------------------------------
UPSERT_FEEDSTOCK_SQL = """
    INSERT INTO bronze.eia_feedstock_monthly
        (year, month, source_sheet, feedstock_name, plant_type,
         quantity_mil_lbs, is_withheld, is_no_data, source_file, collected_at)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
    ON CONFLICT (year, month, source_sheet, feedstock_name, plant_type)
    DO UPDATE SET
        quantity_mil_lbs = EXCLUDED.quantity_mil_lbs,
        is_withheld = EXCLUDED.is_withheld,
        is_no_data = EXCLUDED.is_no_data,
        source_file = EXCLUDED.source_file,
        collected_at = NOW()
    RETURNING (xmax = 0) AS is_insert
"""

UPSERT_CAPACITY_SQL = """
    INSERT INTO bronze.eia_capacity_monthly
        (year, month, biofuel_type, capacity_mmgy, source_file, collected_at)
    VALUES (%s, %s, %s, %s, %s, NOW())
    ON CONFLICT (year, month, biofuel_type)
    DO UPDATE SET
        capacity_mmgy = EXCLUDED.capacity_mmgy,
        source_file = EXCLUDED.source_file,
        collected_at = NOW()
    RETURNING (xmax = 0) AS is_insert
"""


def save_feedstock_records(records, source_file):
    """Upsert feedstock records into bronze.eia_feedstock_monthly."""
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()
    inserted = updated = errors = 0

    for rec in records:
        try:
            cursor.execute(UPSERT_FEEDSTOCK_SQL, (
                rec['year'], rec['month'], rec['source_sheet'],
                rec['feedstock_name'], rec['plant_type'],
                rec['quantity_mil_lbs'], rec['is_withheld'], rec['is_no_data'],
                source_file,
            ))
            result = cursor.fetchone()
            if result and result[0]:
                inserted += 1
            else:
                updated += 1
        except Exception as e:
            logger.error(f"Error saving feedstock record {rec}: {e}")
            errors += 1
            conn.rollback()

    conn.commit()
    cursor.close()
    conn.close()
    return inserted, updated, errors


def save_capacity_records(records, source_file):
    """Upsert capacity records into bronze.eia_capacity_monthly."""
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()
    inserted = updated = errors = 0

    for rec in records:
        try:
            cursor.execute(UPSERT_CAPACITY_SQL, (
                rec['year'], rec['month'], rec['biofuel_type'],
                rec['capacity_mmgy'], source_file,
            ))
            result = cursor.fetchone()
            if result and result[0]:
                inserted += 1
            else:
                updated += 1
        except Exception as e:
            logger.error(f"Error saving capacity record {rec}: {e}")
            errors += 1
            conn.rollback()

    conn.commit()
    cursor.close()
    conn.close()
    return inserted, updated, errors


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description='Download and load EIA biofuels data (table1 + table2)')
    parser.add_argument('--table1', help='Path to table1.xlsx (capacity)')
    parser.add_argument('--table2', help='Path to table2.xlsx (feedstocks)')
    parser.add_argument('--download', action='store_true',
                        help='Download files from EIA website first')
    parser.add_argument('--download-dir', default=DATA_DIR,
                        help='Directory to save downloaded files')
    args = parser.parse_args()

    table1_path = args.table1
    table2_path = args.table2

    if args.download:
        os.makedirs(args.download_dir, exist_ok=True)
        table1_path = download_file(TABLE1_URL,
                                    os.path.join(args.download_dir, 'table1.xlsx'))
        table2_path = download_file(TABLE2_URL,
                                    os.path.join(args.download_dir, 'table2.xlsx'))

    if not table1_path and not table2_path:
        logger.error("No input files specified. Use --download or --table1/--table2.")
        sys.exit(1)

    # --- Process Table 1 (Capacity) ---
    if table1_path:
        if not os.path.isfile(table1_path):
            logger.error(f"File not found: {table1_path}")
            sys.exit(1)

        logger.info(f"Parsing Table 1 (capacity): {table1_path}")
        cap_records = parse_table1(table1_path)
        logger.info(f"  Parsed {len(cap_records)} capacity records")

        if cap_records:
            years = sorted(set(r['year'] for r in cap_records))
            types = sorted(set(r['biofuel_type'] for r in cap_records))
            logger.info(f"  Years: {years}")
            logger.info(f"  Biofuel types: {types}")

            source = os.path.basename(table1_path)
            ins, upd, err = save_capacity_records(cap_records, source)
            logger.info(f"  Capacity: {ins} inserted, {upd} updated, {err} errors")
            print(f"\nCapacity: {len(cap_records)} records -> {ins} inserted, {upd} updated, {err} errors")

    # --- Process Table 2 (Feedstocks) ---
    if table2_path:
        if not os.path.isfile(table2_path):
            logger.error(f"File not found: {table2_path}")
            sys.exit(1)

        logger.info(f"Parsing Table 2 (feedstocks): {table2_path}")
        feed_records = parse_table2(table2_path)
        logger.info(f"  Total: {len(feed_records)} feedstock records parsed")

        if feed_records:
            sheets = sorted(set(r['source_sheet'] for r in feed_records))
            names = sorted(set(r['feedstock_name'] for r in feed_records))
            types = sorted(set(r['plant_type'] for r in feed_records))
            numeric = sum(1 for r in feed_records if r['quantity_mil_lbs'] is not None)
            withheld = sum(1 for r in feed_records if r['is_withheld'])
            no_data = sum(1 for r in feed_records if r['is_no_data'])

            logger.info(f"  Sheets: {sheets}")
            logger.info(f"  Feedstocks: {names}")
            logger.info(f"  Plant types: {types}")
            logger.info(f"  Values: {numeric} numeric, {withheld} withheld, {no_data} no-data")

            source = os.path.basename(table2_path)
            ins, upd, err = save_feedstock_records(feed_records, source)
            logger.info(f"  Feedstock: {ins} inserted, {upd} updated, {err} errors")
            print(f"Feedstock: {len(feed_records)} records -> {ins} inserted, {upd} updated, {err} errors")

            if err > 0:
                sys.exit(1)


if __name__ == '__main__':
    main()
