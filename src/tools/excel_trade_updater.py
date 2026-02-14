"""
Excel Trade Data Updater
========================
Updates Excel trade spreadsheets with data from PostgreSQL database.

Usage:
    python excel_trade_updater.py --file "path/to/workbook.xlsx" --sheet "Soybean Exports" --start 2024-10 --end 2024-12
    python excel_trade_updater.py --file "path/to/workbook.xlsx" --sheet "Soybean Exports" --months "2024-10,2024-11,2024-12"
    python excel_trade_updater.py --file "path/to/workbook.xlsx" --sheet "Soybean Exports" --all  # Update all available months

Can also be called from VBA (see TradeUpdater.bas for full implementation)
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import logging

# Third-party imports
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    print("ERROR: psycopg2 not installed. Run: pip install psycopg2-binary")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CONFIGURATION
# =============================================================================

DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'rlc_commodities',
    'user': 'postgres',
    'password': 'SoupBoss1'  # Consider using environment variable
}

# Spreadsheet structure constants
HEADER_ROW = 2  # Row with date headers
DATA_START_ROW = 4  # First row with country data
COUNTRY_COLUMN = 1  # Column A has country names

# Logging setup - log to file for debugging VBA calls
LOG_FILE = Path(__file__).parent / 'trade_updater.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()  # Also print to console
    ]
)
logger = logging.getLogger(__name__)


# =============================================================================
# DATABASE FUNCTIONS
# =============================================================================

def get_db_connection():
    """Create database connection."""
    return psycopg2.connect(**DB_CONFIG)


def get_latest_available_months(num_months: int = 3) -> List[Tuple[int, int]]:
    """
    Get the most recent N months of available data from the database.

    Args:
        num_months: Number of months to return (default 3)

    Returns:
        List of (year, month) tuples, most recent first
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
        SELECT DISTINCT year, month
        FROM bronze.census_trade
        ORDER BY year DESC, month DESC
        LIMIT %s
    """

    cursor.execute(query, (num_months,))
    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    # Return in chronological order (oldest first)
    return list(reversed(rows))


def get_commodity_from_sheet_name(sheet_name: str) -> Tuple[str, str]:
    """
    Determine commodity group and flow type from sheet name.

    Returns:
        Tuple of (commodity_group, flow_type)
    """
    sheet_lower = sheet_name.lower()

    # Determine commodity
    if 'soybean' in sheet_lower and 'meal' in sheet_lower:
        commodity = 'SOYBEAN_MEAL'
    elif 'soybean' in sheet_lower and 'oil' in sheet_lower:
        commodity = 'SOYBEAN_OIL'
    elif 'soybean' in sheet_lower or 'soy ' in sheet_lower:
        commodity = 'SOYBEANS'
    elif 'corn' in sheet_lower or 'maize' in sheet_lower:
        commodity = 'CORN'
    elif 'wheat' in sheet_lower:
        commodity = 'WHEAT'
    elif 'ddgs' in sheet_lower:
        commodity = 'DDGS'
    elif 'canola' in sheet_lower or 'rapeseed' in sheet_lower:
        commodity = 'CANOLA'
    elif 'cottonseed' in sheet_lower and 'meal' in sheet_lower:
        commodity = 'COTTONSEED_MEAL'
    elif 'cottonseed' in sheet_lower and 'oil' in sheet_lower:
        commodity = 'COTTONSEED_OIL'
    elif 'cottonseed' in sheet_lower:
        commodity = 'COTTONSEED'
    else:
        commodity = 'UNKNOWN'

    # Determine flow type (lowercase to match gold layer: exports, imports)
    if 'import' in sheet_lower:
        flow = 'imports'
    elif 'export' in sheet_lower:
        flow = 'exports'
    else:
        flow = 'exports'  # Default to exports

    return commodity, flow


def fetch_trade_data(
    commodity_group: str,
    flow: str,
    months: List[Tuple[int, int]]
) -> Dict:
    """
    Fetch trade data from database for specified months.

    Args:
        commodity_group: e.g., 'SOYBEANS'
        flow: 'EXPORT' or 'IMPORT'
        months: List of (year, month) tuples

    Returns:
        Dict with structure: {(year, month): {country_name: quantity}}
    """
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # Build month filter
    month_conditions = []
    for year, month in months:
        month_conditions.append(f"(year = {year} AND month = {month})")
    month_filter = " OR ".join(month_conditions)

    query = f"""
        SELECT
            country_name,
            region,
            spreadsheet_row,
            is_regional_total,
            year,
            month,
            quantity,
            display_unit
        FROM gold.trade_export_matrix
        WHERE commodity_group = %s
          AND flow = %s
          AND ({month_filter})
        ORDER BY region_sort_order, country_sort_order, year, month
    """

    cursor.execute(query, (commodity_group, flow))
    rows = cursor.fetchall()

    # Organize data by month -> country -> quantity
    data = {}
    for row in rows:
        key = (row['year'], row['month'])
        if key not in data:
            data[key] = {}
        data[key][row['country_name']] = {
            'quantity': row['quantity'],
            'row': row['spreadsheet_row'],
            'is_regional': row['is_regional_total']
        }

    cursor.close()
    conn.close()

    return data


def fetch_my_totals(
    commodity_group: str,
    flow: str,
    marketing_years: List[str]
) -> Dict:
    """
    Fetch marketing year totals for accumulator columns.

    Args:
        commodity_group: e.g., 'SOYBEANS'
        flow: 'EXPORT' or 'IMPORT'
        marketing_years: List of MY strings like ['23/24', '24/25']

    Returns:
        Dict with structure: {marketing_year: {country_name: quantity}}
    """
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # Convert short MY format to full format if needed
    my_list = []
    for my in marketing_years:
        if '/' in my and len(my) == 5:  # '23/24' format
            # Convert to full year format for query
            my_list.append(my)

    if not my_list:
        return {}

    placeholders = ','.join(['%s'] * len(my_list))

    query = f"""
        SELECT
            country_name,
            region,
            spreadsheet_row,
            is_regional_total,
            marketing_year,
            quantity,
            display_unit
        FROM gold.trade_my_totals_by_country
        WHERE commodity_group = %s
          AND flow = %s
          AND marketing_year IN ({placeholders})
        ORDER BY region_sort_order, country_sort_order, marketing_year_end
    """

    cursor.execute(query, [commodity_group, flow] + my_list)
    rows = cursor.fetchall()

    # Organize data
    data = {}
    for row in rows:
        my = row['marketing_year']
        if my not in data:
            data[my] = {}
        data[my][row['country_name']] = {
            'quantity': row['quantity'],
            'row': row['spreadsheet_row'],
            'is_regional': row['is_regional_total']
        }

    cursor.close()
    conn.close()

    return data


def fetch_country_row_mapping() -> Dict[str, int]:
    """Fetch country name to spreadsheet row mapping from database."""
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    query = """
        SELECT country_name, spreadsheet_row
        FROM silver.trade_country_reference
        WHERE spreadsheet_row IS NOT NULL
        ORDER BY region_sort_order, country_sort_order
    """

    cursor.execute(query)
    rows = cursor.fetchall()

    mapping = {row['country_name']: row['spreadsheet_row'] for row in rows}

    cursor.close()
    conn.close()

    return mapping


# =============================================================================
# EXCEL FUNCTIONS
# =============================================================================

def find_column_for_date(ws, target_date: datetime) -> Optional[int]:
    """
    Find the column number for a given date in row 2.

    Args:
        ws: openpyxl worksheet
        target_date: datetime object for the month to find

    Returns:
        Column number (1-indexed) or None if not found
    """
    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=HEADER_ROW, column=col).value
        if cell_value is None:
            continue

        # Handle different date formats
        if isinstance(cell_value, datetime):
            if cell_value.year == target_date.year and cell_value.month == target_date.month:
                return col
        elif isinstance(cell_value, str):
            # Try parsing string date
            try:
                parsed = datetime.strptime(cell_value, '%Y-%m-%d')
                if parsed.year == target_date.year and parsed.month == target_date.month:
                    return col
            except ValueError:
                pass

    return None


def find_column_for_my(ws, marketing_year: str) -> Optional[int]:
    """
    Find the column number for a marketing year accumulator column.

    Args:
        ws: openpyxl worksheet
        marketing_year: String like '23/24' or '2023/24'

    Returns:
        Column number (1-indexed) or None if not found
    """
    # Normalize MY format
    if len(marketing_year) == 5:  # '23/24'
        my_short = marketing_year
    else:
        # Extract short format
        parts = marketing_year.split('/')
        my_short = parts[0][-2:] + '/' + parts[1][-2:]

    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=HEADER_ROW, column=col).value
        if cell_value is None:
            continue

        if isinstance(cell_value, str):
            # Normalize and compare
            cell_clean = cell_value.strip()
            if cell_clean == my_short or cell_clean == marketing_year:
                return col

    return None


def find_row_for_country(ws, country_name: str, country_col: int = 1) -> Optional[int]:
    """
    Find the row number for a country name.

    Args:
        ws: openpyxl worksheet
        country_name: Country name to find
        country_col: Column containing country names (default 1 = A)

    Returns:
        Row number (1-indexed) or None if not found
    """
    country_upper = country_name.upper().strip()

    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=country_col).value
        if cell_value is None:
            continue
        if cell_value.upper().strip() == country_upper:
            return row

    return None


def update_worksheet(
    ws,
    monthly_data: Dict,
    my_data: Dict = None,
    use_db_rows: bool = True
) -> Tuple[int, int]:
    """
    Update worksheet with trade data.

    Args:
        ws: openpyxl worksheet
        monthly_data: Dict from fetch_trade_data()
        my_data: Optional dict from fetch_my_totals()
        use_db_rows: If True, use row numbers from database. If False, search by country name.

    Returns:
        Tuple of (cells_updated, cells_not_found)
    """
    cells_updated = 0
    cells_not_found = 0

    # Update monthly data
    for (year, month), country_data in monthly_data.items():
        # Find column for this month
        target_date = datetime(year, month, 1)
        col = find_column_for_date(ws, target_date)

        if col is None:
            logger.warning(f"Column not found for {year}-{month:02d}")
            cells_not_found += len(country_data)
            continue

        # Update each country
        for country_name, data in country_data.items():
            if use_db_rows and data.get('row'):
                row = data['row']
            else:
                row = find_row_for_country(ws, country_name)

            if row is None:
                logger.warning(f"Row not found for country: {country_name}")
                cells_not_found += 1
                continue

            # Write the value
            quantity = data['quantity']
            if quantity is not None:
                ws.cell(row=row, column=col).value = float(quantity)
                cells_updated += 1
            else:
                cells_not_found += 1

    # Update marketing year totals if provided
    if my_data:
        for marketing_year, country_data in my_data.items():
            col = find_column_for_my(ws, marketing_year)

            if col is None:
                logger.warning(f"Column not found for MY: {marketing_year}")
                cells_not_found += len(country_data)
                continue

            for country_name, data in country_data.items():
                if use_db_rows and data.get('row'):
                    row = data['row']
                else:
                    row = find_row_for_country(ws, country_name)

                if row is None:
                    logger.warning(f"Row not found for country: {country_name}")
                    cells_not_found += 1
                    continue

                quantity = data['quantity']
                if quantity is not None:
                    ws.cell(row=row, column=col).value = float(quantity)
                    cells_updated += 1

    return cells_updated, cells_not_found


# =============================================================================
# MAIN FUNCTIONS
# =============================================================================

def parse_month_string(month_str: str) -> Tuple[int, int]:
    """Parse month string like '2024-10' to (year, month) tuple."""
    parts = month_str.strip().split('-')
    return int(parts[0]), int(parts[1])


def get_months_in_range(start: str, end: str) -> List[Tuple[int, int]]:
    """Get list of (year, month) tuples between start and end."""
    start_year, start_month = parse_month_string(start)
    end_year, end_month = parse_month_string(end)

    months = []
    current_year, current_month = start_year, start_month

    while (current_year, current_month) <= (end_year, end_month):
        months.append((current_year, current_month))
        current_month += 1
        if current_month > 12:
            current_month = 1
            current_year += 1

    return months


def update_trade_sheet(
    file_path: str,
    sheet_name: str,
    months: List[Tuple[int, int]],
    update_my: bool = True
) -> Dict:
    """
    Main function to update a trade spreadsheet.

    Args:
        file_path: Path to Excel workbook
        sheet_name: Name of worksheet to update
        months: List of (year, month) tuples to update
        update_my: Whether to also update marketing year accumulator columns

    Returns:
        Dict with update statistics
    """
    logger.info(f"Opening workbook: {file_path}")
    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    logger.info(f"Updating sheet: {sheet_name}")

    # Determine commodity and flow from sheet name
    commodity, flow = get_commodity_from_sheet_name(sheet_name)
    logger.info(f"Detected commodity: {commodity}, flow: {flow}")

    if commodity == 'UNKNOWN':
        logger.warning("Could not determine commodity from sheet name. Using SOYBEANS as default.")
        commodity = 'SOYBEANS'

    # Fetch monthly data
    logger.info(f"Fetching data for {len(months)} months...")
    monthly_data = fetch_trade_data(commodity, flow, months)
    logger.info(f"Retrieved data for {len(monthly_data)} months")

    # Fetch MY totals if requested
    my_data = None
    if update_my:
        # Determine which MYs are affected
        marketing_years = set()
        for year, month in months:
            if month >= 9:
                my = f"{year % 100:02d}/{(year + 1) % 100:02d}"
            else:
                my = f"{(year - 1) % 100:02d}/{year % 100:02d}"
            marketing_years.add(my)

        if marketing_years:
            logger.info(f"Fetching MY totals for: {marketing_years}")
            my_data = fetch_my_totals(commodity, flow, list(marketing_years))

    # Update the worksheet (use_db_rows=False to search by country name)
    cells_updated, cells_not_found = update_worksheet(ws, monthly_data, my_data, use_db_rows=False)

    # Save the workbook
    logger.info(f"Saving workbook...")
    wb.save(file_path)
    wb.close()

    results = {
        'file': file_path,
        'sheet': sheet_name,
        'commodity': commodity,
        'flow': flow,
        'months_requested': len(months),
        'months_with_data': len(monthly_data),
        'cells_updated': cells_updated,
        'cells_not_found': cells_not_found
    }

    logger.info(f"Update complete: {cells_updated} cells updated, {cells_not_found} not found")
    return results


# =============================================================================
# TEMPLATE GENERATOR
# =============================================================================

def create_trade_template(
    output_path: str,
    commodity_group: str,
    flow: str
) -> str:
    """
    Create a new Excel template with country names from the database.

    Args:
        output_path: Path for the new Excel file
        commodity_group: e.g., 'SOYBEANS'
        flow: 'EXPORTS' or 'IMPORTS'

    Returns:
        Path to created file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    logger.info(f"Creating template for {commodity_group} {flow}")

    # Fetch country structure from database
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    query = """
        SELECT
            country_name,
            region,
            region_sort_order,
            country_sort_order,
            spreadsheet_row,
            is_regional_total
        FROM silver.trade_country_reference
        WHERE is_active = TRUE
        ORDER BY region_sort_order, country_sort_order
    """
    cursor.execute(query)
    countries = cursor.fetchall()
    cursor.close()
    conn.close()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{commodity_group.replace('_', ' ').title()} {flow.title()}"

    # Styles
    header_font = Font(bold=True)
    regional_font = Font(bold=True, color="FFFFFF")
    regional_fill = PatternFill(start_color="1B4D4D", end_color="1B4D4D", fill_type="solid")
    border = Side(style='thin', color='CCCCCC')
    cell_border = Border(left=border, right=border, top=border, bottom=border)

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"US {commodity_group.replace('_', ' ').title()} {flow.title()}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Row 2: Headers (will have dates added later)
    ws.cell(row=2, column=1, value="Country")
    ws.cell(row=2, column=1).font = header_font

    # Row 3: Units placeholder
    ws.cell(row=3, column=1, value="")

    # Add countries starting at row 4
    for country in countries:
        row = country['spreadsheet_row']
        if row is None:
            continue

        cell = ws.cell(row=row, column=1, value=country['country_name'])
        cell.border = cell_border

        # Style regional totals
        if country['is_regional_total']:
            cell.font = regional_font
            cell.fill = regional_fill

    # Set column width
    ws.column_dimensions['A'].width = 30

    # Add some placeholder date columns (these will be populated by data)
    for col in range(2, 15):
        ws.cell(row=2, column=col, value=f"Month {col-1}")
        ws.cell(row=2, column=col).font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Save
    wb.save(output_path)
    wb.close()

    logger.info(f"Template created: {output_path}")
    return output_path


# =============================================================================
# CLI INTERFACE
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Update Excel trade spreadsheets with database data',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Update specific month range
    python excel_trade_updater.py --file "trade.xlsx" --sheet "Soybean Exports" --start 2024-10 --end 2024-12

    # Update specific months
    python excel_trade_updater.py --file "trade.xlsx" --sheet "Soybean Exports" --months "2024-10,2024-11"

    # Update all available months in database
    python excel_trade_updater.py --file "trade.xlsx" --sheet "Soybean Exports" --all

    # Update latest 3 months of available data (for Ctrl+I quick update)
    python excel_trade_updater.py --file "trade.xlsx" --sheet "Soybean Exports" --latest

    # Create a new template with correct country names from database
    python excel_trade_updater.py --create-template "Soybean_Exports.xlsx" --commodity SOYBEANS --flow EXPORTS
        """
    )

    parser.add_argument('--file', '-f', help='Path to Excel workbook')
    parser.add_argument('--sheet', '-s', help='Worksheet name to update')
    parser.add_argument('--create-template', help='Create new template file with this name')
    parser.add_argument('--commodity', help='Commodity for template (SOYBEANS, CORN, WHEAT, etc.)')
    parser.add_argument('--flow', help='Flow type for template (EXPORTS or IMPORTS)')
    parser.add_argument('--start', help='Start month (YYYY-MM format)')
    parser.add_argument('--end', help='End month (YYYY-MM format)')
    parser.add_argument('--months', help='Comma-separated list of months (YYYY-MM,YYYY-MM,...)')
    parser.add_argument('--all', action='store_true', help='Update all available months')
    parser.add_argument('--latest', type=int, nargs='?', const=3, default=None,
                        help='Update latest N months of available data (default: 3)')
    parser.add_argument('--no-my', action='store_true', help='Skip marketing year accumulator updates')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Handle template creation mode
    if args.create_template:
        if not args.commodity or not args.flow:
            parser.error("--create-template requires --commodity and --flow")
        create_trade_template(args.create_template, args.commodity.upper(), args.flow.upper())
        print(f"Template created: {args.create_template}")
        sys.exit(0)

    # Validate arguments for update mode
    if not args.file or not args.sheet:
        parser.error("--file and --sheet are required for updating")

    if not args.all and not args.latest and not args.months and not (args.start and args.end):
        parser.error("Must specify --all, --latest, --months, or both --start and --end")

    # Determine months to update
    if args.all:
        # Query database for available months
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT year, month FROM bronze.census_trade ORDER BY year, month")
        months = cursor.fetchall()
        cursor.close()
        conn.close()
    elif args.latest:
        # Get latest N months from database
        months = get_latest_available_months(args.latest)
        logger.info(f"Using latest {args.latest} months from database")
    elif args.months:
        months = [parse_month_string(m) for m in args.months.split(',')]
    else:
        months = get_months_in_range(args.start, args.end)

    if not months:
        logger.error("No months to update")
        sys.exit(1)

    logger.info(f"Will update {len(months)} months: {months[0]} to {months[-1]}")

    # Validate file exists
    file_path = Path(args.file)
    if not file_path.exists():
        logger.error(f"File not found: {file_path}")
        sys.exit(1)

    # Run update
    try:
        results = update_trade_sheet(
            str(file_path),
            args.sheet,
            months,
            update_my=not args.no_my
        )

        print("\n" + "="*50)
        print("UPDATE COMPLETE")
        print("="*50)
        for key, value in results.items():
            print(f"  {key}: {value}")
        print("="*50)

    except Exception as e:
        logger.exception(f"Error updating spreadsheet: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
