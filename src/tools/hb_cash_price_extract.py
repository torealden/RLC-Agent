"""
HB Weekly Cash Price Extract

Reads config.hb_price_mapping, queries bronze/silver tables, and produces
the weekly price table for the HB report. Can also write prices directly
into the hb_cash_price.xlsx spreadsheet.

Usage:
    python src/tools/hb_cash_price_extract.py                     # Current week, dry-run
    python src/tools/hb_cash_price_extract.py --write              # Write to spreadsheet
    python src/tools/hb_cash_price_extract.py --date 2026-02-28   # Specific report date
    python src/tools/hb_cash_price_extract.py --output out.xlsx    # Write to separate file
"""

import os
import sys
import json
import argparse
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / '.env')
except ImportError:
    pass

try:
    import psycopg2
    import psycopg2.extras
    PSYCOPG2_AVAILABLE = True
except ImportError:
    PSYCOPG2_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

HB_SPREADSHEET_PATH = PROJECT_ROOT / 'data' / 'hb_cash_price.xlsx'


def get_connection():
    """Get database connection."""
    if not PSYCOPG2_AVAILABLE:
        raise ImportError("psycopg2 required")
    return psycopg2.connect(
        host=os.environ.get('DB_HOST', 'localhost'),
        port=os.environ.get('DB_PORT', 5432),
        database=os.environ.get('DB_NAME', 'rlc_commodities'),
        user=os.environ.get('DB_USER', 'postgres'),
        password=os.environ.get('DB_PASSWORD'),
    )


def extract_hb_prices(report_date: date = None) -> Dict[str, Any]:
    """
    Extract all HB prices by reading config.hb_price_mapping and querying
    the appropriate source tables for current, week-ago, and year-ago prices.

    Args:
        report_date: Reference date (default: today)

    Returns:
        Dict with keys:
            'prices': list of dicts with hb_row_label, sheet_row, this_week,
                      last_week, year_ago, unit, source_type
            'missing': list of labels with no price found
            'report_date': the reference date used
    """
    report_date = report_date or date.today()
    week_ago = report_date - timedelta(days=7)
    year_ago = report_date - timedelta(days=365)

    conn = get_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Load all active HB mappings
    cur.execute("""
        SELECT * FROM config.hb_price_mapping
        WHERE is_active = TRUE
        ORDER BY hb_sheet_row
    """)
    mappings = cur.fetchall()

    prices = []
    missing = []

    for m in mappings:
        row_label = m['hb_row_label']
        source_type = m['source_type']
        sheet_row = m['hb_sheet_row']
        carry_weeks = m['carry_forward_weeks'] or 1

        if source_type == 'ams_structured':
            this_week = _find_ams_price(cur, m, report_date, carry_weeks)
            last_week_val = _find_ams_price(cur, m, week_ago, carry_weeks)
            year_ago_val = _find_ams_price(cur, m, year_ago, carry_weeks)
        elif source_type == 'futures':
            this_week = _find_futures_price(cur, m, report_date)
            last_week_val = _find_futures_price(cur, m, week_ago)
            year_ago_val = _find_futures_price(cur, m, year_ago)
        else:
            this_week = None
            last_week_val = None
            year_ago_val = None

        entry = {
            'hb_row_label': row_label,
            'hb_sheet_row': sheet_row,
            'this_week': this_week,
            'last_week': last_week_val,
            'year_ago': year_ago_val,
            'unit': m['unit'],
            'source_type': source_type,
        }
        prices.append(entry)

        if this_week is None:
            missing.append(row_label)

    conn.close()

    return {
        'prices': prices,
        'missing': missing,
        'report_date': str(report_date),
    }


def _find_ams_price(
    cur, mapping: Dict, target_date: date, carry_weeks: int
) -> Optional[float]:
    """
    Find the best-matching AMS price record near the target date.

    Looks back up to carry_weeks * 7 days for the most recent matching record.
    """
    lookback_days = carry_weeks * 7 + 3  # small buffer
    slug_id = mapping['slug_id']
    price_field = mapping['price_field'] or 'price_avg'

    # Build WHERE conditions for matching
    conditions = ["slug_id = %(slug_id)s"]
    params = {
        'slug_id': slug_id,
        'target_date': target_date,
        'lookback': target_date - timedelta(days=lookback_days),
    }

    if mapping['match_commodity']:
        conditions.append("LOWER(commodity) LIKE %(match_commodity)s")
        params['match_commodity'] = f"%{mapping['match_commodity'].lower()}%"

    if mapping['match_location']:
        conditions.append("LOWER(location) LIKE %(match_location)s")
        params['match_location'] = f"%{mapping['match_location'].lower()}%"

    if mapping['match_grade']:
        conditions.append("LOWER(grade) LIKE %(match_grade)s")
        params['match_grade'] = f"%{mapping['match_grade'].lower()}%"

    if mapping['match_section']:
        # MARS API stores sub-section names (e.g., "Mills and Processors",
        # "Country Elevators") in delivery_point, not report_section.
        # Search both fields so the mapping works regardless.
        conditions.append(
            "(LOWER(report_section) LIKE %(match_section)s"
            " OR LOWER(delivery_point) LIKE %(match_section)s)"
        )
        params['match_section'] = f"%{mapping['match_section'].lower()}%"

    if mapping.get('match_delivery_period'):
        conditions.append("delivery_period = %(match_delivery_period)s")
        params['match_delivery_period'] = mapping['match_delivery_period']

    where = " AND ".join(conditions)

    # Map price_field to actual column
    valid_fields = {'price', 'price_low', 'price_high', 'price_avg', 'price_mostly'}
    col = price_field if price_field in valid_fields else 'price_avg'

    query = f"""
        SELECT {col}
        FROM bronze.ams_price_record
        WHERE {where}
          AND report_date <= %(target_date)s
          AND report_date >= %(lookback)s
          AND {col} IS NOT NULL
        ORDER BY report_date DESC
        LIMIT 1
    """

    try:
        cur.execute(query, params)
        row = cur.fetchone()
        if row and row[0] is not None:
            return float(row[0])
    except Exception as e:
        logger.warning(f"Error querying AMS price for {mapping['hb_row_label']}: {e}")
        cur.connection.rollback()

    return None


def _find_futures_price(
    cur, mapping: Dict, target_date: date
) -> Optional[float]:
    """
    Find the front-month (non-spot) futures settlement price near target_date.

    Used for Rice (ZR) and Milk Class III (DC).
    """
    symbol = mapping.get('futures_symbol')
    if not symbol:
        return None

    try:
        cur.execute("""
            SELECT settlement
            FROM silver.futures_price
            WHERE symbol = %(symbol)s
              AND trade_date <= %(target_date)s
              AND trade_date >= %(lookback)s
              AND contract_date > %(target_date)s
              AND settlement IS NOT NULL
            ORDER BY contract_date ASC, trade_date DESC
            LIMIT 1
        """, {
            'symbol': symbol,
            'target_date': target_date,
            'lookback': target_date - timedelta(days=7),
        })
        row = cur.fetchone()
        if row and row[0] is not None:
            return float(row[0])
    except Exception as e:
        logger.warning(f"Error querying futures for {symbol}: {e}")
        cur.connection.rollback()

    return None


def update_hb_spreadsheet(
    prices: List[Dict],
    output_path: Path = None,
    report_date: date = None,
) -> str:
    """
    Write extracted prices into the HB cash price spreadsheet.

    Writes into Sheet2 columns D (this week), E (last week), F (year ago).

    Args:
        prices: List of price dicts from extract_hb_prices()
        output_path: Output file path (default: overwrite original)
        report_date: Date to set in cell D3

    Returns:
        Path to the written file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required to write spreadsheets")

    source_path = HB_SPREADSHEET_PATH
    out_path = output_path or source_path

    wb = openpyxl.load_workbook(str(source_path))
    ws = wb['Sheet2']

    # Set report date in D3
    if report_date:
        ws['D3'] = datetime.combine(report_date, datetime.min.time())

    # Write prices
    written = 0
    for p in prices:
        row = p['hb_sheet_row']
        if row < 4 or row > 60:
            continue

        if p['this_week'] is not None:
            ws.cell(row=row, column=4, value=p['this_week'])  # D
            written += 1
        if p['last_week'] is not None:
            ws.cell(row=row, column=5, value=p['last_week'])  # E
        if p['year_ago'] is not None:
            ws.cell(row=row, column=6, value=p['year_ago'])   # F

    wb.save(str(out_path))
    logger.info(f"Wrote {written} prices to {out_path}")
    return str(out_path)


def print_price_table(result: Dict):
    """Print a formatted price table to stdout."""
    prices = result['prices']
    missing = result['missing']

    print(f"\nHB Weekly Price Extract — Report Date: {result['report_date']}")
    print(f"{'='*85}")
    print(f"{'Row':>4}  {'Label':<50}  {'This Wk':>10}  {'Last Wk':>10}  {'Yr Ago':>10}")
    print(f"{'-'*85}")

    for p in prices:
        tw = f"{p['this_week']:.4f}" if p['this_week'] is not None else '---'
        lw = f"{p['last_week']:.4f}" if p['last_week'] is not None else '---'
        ya = f"{p['year_ago']:.4f}" if p['year_ago'] is not None else '---'
        label = p['hb_row_label'][:48]
        marker = ' *' if p['this_week'] is None else ''
        print(f"{p['hb_sheet_row']:>4}  {label:<50}  {tw:>10}  {lw:>10}  {ya:>10}{marker}")

    print(f"\n{'='*85}")
    print(f"Total: {len(prices)} rows, {len(prices) - len(missing)} with prices, {len(missing)} missing")

    if missing:
        print(f"\nMissing prices:")
        for m in missing:
            print(f"  - {m}")


def main():
    parser = argparse.ArgumentParser(description='HB Weekly Cash Price Extract')
    parser.add_argument('--date', help='Report date (YYYY-MM-DD), default today')
    parser.add_argument('--write', action='store_true', help='Write to HB spreadsheet')
    parser.add_argument('--output', '-o', help='Output file path (default: overwrite original)')
    parser.add_argument('--json', action='store_true', help='Output as JSON')
    parser.add_argument('--verbose', '-v', action='store_true')
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s %(levelname)s %(message)s',
    )

    report_date = datetime.strptime(args.date, '%Y-%m-%d').date() if args.date else None

    result = extract_hb_prices(report_date)

    if args.json:
        print(json.dumps(result, indent=2, default=str))
    else:
        print_price_table(result)

    if args.write:
        out_path = Path(args.output) if args.output else None
        saved = update_hb_spreadsheet(
            result['prices'],
            output_path=out_path,
            report_date=report_date or date.today(),
        )
        print(f"\nSpreadsheet written to: {saved}")


if __name__ == '__main__':
    main()
