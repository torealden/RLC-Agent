"""
Update models/Biofuels/us_liquid_fuel_and_biofuel_production.xlsx from DB.

Pulls monthly production, stocks, AND domestic-use data and writes them into
the workbook's three flat-file sheets ("Production", "Stocks", "Domestic Use").
All three sheets share an identical layout:

  Row 3 headers:
    A=thousand gallons  B=Biodiesel  C=Renewable Diesel  D=Co-Processing
    E=Sustainable Aviation Fuel  F=Ethanol  H=Diesel  I=Jet Fuel  J=Gasoline
    L=Glycerin  M=Fatty Acid Methyl Ester  N=Renewable Naptha
    O=Renewable Propane  P=Soap Stock  Q=Methyl Acetate

  Rows 4+ : column A = date (monthly), 1990-01 onward.

Production source:    gold.us_liquid_fuel_production_monthly (EMTS RIN-based).
Stocks source:        gold.us_liquid_fuel_stocks_monthly (EIA monthly biofuels).
Domestic Use source:  gold.us_liquid_fuel_domestic_use_monthly (EIA apparent consumption).

Columns currently driven (production):
  B Biodiesel, C Renewable Diesel, D Co-Processing, E SAF, F Ethanol,
  N Renewable Naphtha, O Renewable Propane

Columns currently driven (stocks):
  B Biodiesel (combined_bd_rd − renewable_diesel), C Renewable Diesel, F Ethanol

Columns currently driven (domestic use):
  B Biodiesel (apparent: prod + imp − exp), C Renewable Diesel (apparent),
  F Ethanol (blender_input)

Left untouched (no public source or not separable):
  H Diesel, I Jet Fuel, J Gasoline           ← EIA petroleum, not yet wired
  L Glycerin, M FAME, P Soap Stock, Q Methyl Acetate  ← no source
  D/E/N/O on stocks + domestic use            ← EIA does not split these

Usage:
  python src/tools/update_us_liquid_fuel_production.py                          # full refresh, all sheets
  python src/tools/update_us_liquid_fuel_production.py --months 12              # last 12 months
  python src/tools/update_us_liquid_fuel_production.py --sheet domestic_use     # only one sheet
  python src/tools/update_us_liquid_fuel_production.py --dry-run                # report only
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")

import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl

XLSX = ROOT / "models" / "Biofuels" / "us_liquid_fuel_and_biofuel_production.xlsx"

HEADER_ROW = 3
DATA_START_ROW = 4

# Column-letter → DB field. Same letter set for both sheets; if a column has
# no source for one of the two views, that view simply skips it.
PRODUCTION_COLUMN_MAP = {
    "B": "biodiesel_kgal",
    "C": "renewable_diesel_kgal",
    "D": "co_processing_kgal",
    "E": "saf_kgal",
    "F": "ethanol_kgal",
    "N": "renewable_naphtha_kgal",
    "O": "renewable_propane_kgal",
}

STOCKS_COLUMN_MAP = {
    "B": "biodiesel_kgal",
    "C": "renewable_diesel_kgal",
    "F": "ethanol_kgal",
}

DOMESTIC_USE_COLUMN_MAP = {
    "B": "biodiesel_kgal",
    "C": "renewable_diesel_kgal",
    "F": "ethanol_kgal",
}

PRICES_COLUMN_MAP = {
    "B": "biodiesel_usd_gal",
    "C": "renewable_diesel_usd_gal",  # currently NULL pending source review
    "F": "ethanol_usd_gal",            # currently NULL
    "H": "diesel_usd_gal",
    "I": "jet_fuel_usd_gal",           # currently NULL
    "J": "gasoline_usd_gal",
}

SHEET_SPECS = {
    "production": {
        "sheet_name": "Production",
        "view":       "gold.us_liquid_fuel_production_monthly",
        "column_map": PRODUCTION_COLUMN_MAP,
        "cadence":    "monthly",
    },
    "stocks": {
        "sheet_name": "Stocks",
        "view":       "gold.us_liquid_fuel_stocks_monthly",
        "column_map": STOCKS_COLUMN_MAP,
        "cadence":    "monthly",
    },
    "domestic_use": {
        "sheet_name": "Domestic Use",
        "view":       "gold.us_liquid_fuel_domestic_use_monthly",
        "column_map": DOMESTIC_USE_COLUMN_MAP,
        "cadence":    "monthly",
    },
    "prices": {
        "sheet_name": "Prices - Daily",
        "view":       "gold.us_liquid_fuel_prices_daily",
        "column_map": PRICES_COLUMN_MAP,
        "cadence":    "daily",
    },
}
# Note: "Price - Monthly" sheet is maintained by Tore via Excel AVERAGEIFS over
# the "Prices - Daily" sheet — no DB-side update needed.


def get_conn():
    return psycopg2.connect(
        host=os.getenv("RLC_PG_HOST"), port=os.getenv("RLC_PG_PORT", "5432"),
        dbname=os.getenv("RLC_PG_DB", "rlc_commodities"),
        user=os.getenv("RLC_PG_USER"), password=os.getenv("RLC_PG_PASSWORD"),
        sslmode="require",
    )


def fetch_view(view: str, months: int | None, cadence: str = "monthly") -> list[dict]:
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if cadence == "daily":
        # Daily views key on price_date, not year/month.
        if months:
            # months interpreted as ~30-day window approximation.
            cur.execute(
                f"SELECT * FROM {view} WHERE price_date >= CURRENT_DATE - INTERVAL '%s months' ORDER BY price_date",
                (months,),
            )
        else:
            cur.execute(f"SELECT * FROM {view} ORDER BY price_date")
    else:
        if months:
            cur.execute(
                f"SELECT * FROM {view} ORDER BY year DESC, month DESC LIMIT %s",
                (months,),
            )
        else:
            cur.execute(f"SELECT * FROM {view} ORDER BY year, month")
    rows = cur.fetchall()
    conn.close()
    return rows


def build_date_to_row_index(ws) -> dict:
    """Walk column A and build (year, month) → spreadsheet row number."""
    idx = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=1):
        cell = row[0]
        if cell.value is None:
            break
        try:
            idx[(cell.value.year, cell.value.month)] = cell.row
        except AttributeError:
            continue
    return idx


def update_sheet(wb, spec: dict, months: int | None) -> dict:
    """Update a single sheet from its gold view. Returns per-column write counts."""
    sheet_name = spec["sheet_name"]
    cadence = spec.get("cadence", "monthly")
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] Sheet {sheet_name!r} not found in workbook. Available: {wb.sheetnames}")
        return {}

    ws = wb[sheet_name]
    print(f"\n--- Updating [{sheet_name}] from {spec['view']} ({cadence}) ---")

    rows = fetch_view(spec["view"], months, cadence=cadence)
    print(f"  DB returned {len(rows)} rows")

    cells_per_col = {col: 0 for col in spec["column_map"]}

    if cadence == "daily":
        # Daily sheet: write dates to col A starting at row 4, then write prices.
        for i, row_data in enumerate(rows):
            sheet_row = DATA_START_ROW + i
            ws.cell(row=sheet_row, column=1).value = row_data["price_date"]
            for col_letter, db_field in spec["column_map"].items():
                val = row_data.get(db_field)
                if val is None:
                    continue
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                ws.cell(row=sheet_row, column=col_idx).value = float(val)
                cells_per_col[col_letter] += 1
        if rows:
            print(f"  Dates written: {len(rows)} (rows {DATA_START_ROW} → {DATA_START_ROW + len(rows) - 1})")
    else:
        # Monthly sheet: existing logic — read col A dates as index, write to matching rows.
        date_idx = build_date_to_row_index(ws)
        print(f"  Date index: {len(date_idx)} months "
              f"({min(date_idx.keys()) if date_idx else None} → "
              f"{max(date_idx.keys()) if date_idx else None})")
        missing_dates = 0
        for row_data in rows:
            yr, mo = row_data["year"], row_data["month"]
            if (yr, mo) not in date_idx:
                missing_dates += 1
                continue
            sheet_row = date_idx[(yr, mo)]
            for col_letter, db_field in spec["column_map"].items():
                val = row_data.get(db_field)
                if val is None:
                    continue
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                ws.cell(row=sheet_row, column=col_idx).value = float(val)
                cells_per_col[col_letter] += 1
        if missing_dates:
            print(f"  [WARN] {missing_dates} DB months had no matching spreadsheet row")

    print(f"  Cells written by column:")
    for col, db_field in spec["column_map"].items():
        print(f"    {col} ({db_field:30s}): {cells_per_col[col]}")
    print(f"  Total cells written: {sum(cells_per_col.values())}")

    return cells_per_col


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--months", type=int, default=None,
                    help="Only update the latest N months (default: all)")
    ap.add_argument("--sheet", choices=["production", "stocks", "domestic_use", "prices", "all"], default="all",
                    help="Which sheet(s) to update (default: all)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Report what would be written but don't save")
    args = ap.parse_args()

    if not XLSX.exists():
        print(f"[ERROR] Spreadsheet not found: {XLSX}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {XLSX}")
    wb = openpyxl.load_workbook(XLSX)

    sheets_to_update = ["production", "stocks", "domestic_use", "prices"] if args.sheet == "all" else [args.sheet]
    for key in sheets_to_update:
        update_sheet(wb, SHEET_SPECS[key], args.months)

    if args.dry_run:
        print("\n[DRY RUN] not saving")
    else:
        print(f"\nSaving {XLSX}...")
        wb.save(XLSX)
        print("Done.")


if __name__ == "__main__":
    main()
