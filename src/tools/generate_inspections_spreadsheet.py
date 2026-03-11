"""
FGIS Export Inspections Spreadsheet Generator

Queries bronze.fgis_inspections_history aggregations and produces an Excel
workbook with tabs by commodity, rows = countries, columns = months.
Matches the Census export trade spreadsheet layout.

Output: output/reports/fgis_inspections_{commodity}_{period}.xlsx

Usage:
    # Generate for current marketing year
    python src/tools/generate_inspections_spreadsheet.py

    # Specific marketing years
    python src/tools/generate_inspections_spreadsheet.py --my 2526 2425

    # Calendar year mode
    python src/tools/generate_inspections_spreadsheet.py --calendar-year 2025

    # Specific commodities
    python src/tools/generate_inspections_spreadsheet.py --grains corn soybeans
"""

import sys
import os
import argparse
import logging
from pathlib import Path
from datetime import datetime, date
from typing import List, Optional, Dict

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from src.services.database.db_config import get_connection as get_db_connection, get_engine

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)-8s %(message)s')
logger = logging.getLogger(__name__)

OUTPUT_DIR = PROJECT_ROOT / "output" / "reports"

# Marketing year start months by grain
MY_START_MONTH = {
    'corn': 9,       # Sep
    'soybeans': 9,    # Sep
    'wheat': 6,       # Jun
    'sorghum': 9,     # Sep
    'barley': 6,      # Jun
    'oats': 6,        # Jun
    'rice': 8,        # Aug
}

# Wheat subclasses to break out
WHEAT_CLASSES = {
    'HRW': 'Wheat HRW',
    'SRW': 'Wheat SRW',
    'HRS': 'Wheat HRS',
    'DUR': 'Wheat Durum',
    'SW': 'Wheat White',
    'SWW': 'Wheat White',
    'HWW': 'Wheat White',
    'WHT': 'Wheat White',
}

# Month names for column headers
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Top destinations to always include (even if zero)
TOP_DESTINATIONS = [
    'CHINA', 'JAPAN', 'MEXICO', 'S KOREA', 'TAIWAN',
    'COLOMBIA', 'EGYPT', 'PHILIPPINES', 'INDONESIA',
    'CANADA', 'UN KINGDOM', 'SPAIN', 'GUATEMALA',
    'NETHERLANDS', 'MOROCCO', 'PERU', 'NIGERIA',
    'DOMINICAN RP', 'COSTA RICA', 'EL SALVADOR',
    'HONDURAS', 'ECUADOR', 'VENEZUELA', 'BRAZIL',
    'VIETNAM', 'ALGERIA', 'ITALY', 'SAUDI ARABIA',
    'THAILAND', 'BANGLADESH', 'IRAQ', 'ISRAEL',
]


def get_monthly_data(grain: str, marketing_years: List[str] = None,
                     calendar_year: int = None,
                     wheat_class: str = None) -> pd.DataFrame:
    """Query monthly inspections data for a grain.

    Returns DataFrame with columns: year, month, destination, metric_tons
    """
    conditions = ["type_service IN ('IW', 'I')"]
    params: Dict = {}

    if wheat_class:
        conditions.append("UPPER(grain) = 'WHEAT'")
        conditions.append("grain_class = %(wheat_class)s")
        params['wheat_class'] = wheat_class
    else:
        conditions.append("UPPER(grain) = %(grain)s")
        params['grain'] = grain.upper()
        if grain.lower() == 'wheat':
            # All wheat combined (no class filter)
            pass

    if marketing_years:
        placeholders = ', '.join(f'%(my_{i})s' for i in range(len(marketing_years)))
        conditions.append(f"marketing_year IN ({placeholders})")
        for i, my in enumerate(marketing_years):
            params[f'my_{i}'] = my
    elif calendar_year:
        conditions.append("EXTRACT(YEAR FROM cert_date) = %(cal_year)s")
        params['cal_year'] = calendar_year

    where = ' AND '.join(conditions)

    sql = f"""
        SELECT
            EXTRACT(YEAR FROM cert_date)::int AS year,
            EXTRACT(MONTH FROM cert_date)::int AS month,
            destination,
            marketing_year,
            SUM(metric_tons) AS metric_tons,
            SUM(bushels_1000) AS bushels_1000
        FROM bronze.fgis_inspections_history
        WHERE {where}
        GROUP BY 1, 2, 3, 4
        ORDER BY 1, 2
    """

    engine = get_engine()
    df = pd.read_sql(sql, engine, params=params)
    return df


def build_pivot(df: pd.DataFrame, marketing_years: List[str] = None,
                calendar_year: int = None, grain: str = 'corn') -> pd.DataFrame:
    """Pivot monthly data: rows = destinations, columns = year-month."""
    if df.empty:
        return pd.DataFrame()

    # Create year-month label
    df['ym'] = df.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}", axis=1)

    # Pivot: destination x year-month
    pivot = df.pivot_table(
        index='destination',
        columns='ym',
        values='metric_tons',
        aggfunc='sum',
        fill_value=0
    )

    # Sort columns chronologically
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)

    # Add row total
    pivot['TOTAL'] = pivot.sum(axis=1)

    # Sort by total descending
    pivot = pivot.sort_values('TOTAL', ascending=False)

    # Add column totals row
    totals = pivot.sum(axis=0)
    totals.name = 'TOTAL'
    pivot = pd.concat([pivot, totals.to_frame().T])

    return pivot


def write_sheet(wb: Workbook, sheet_name: str, pivot: pd.DataFrame, grain: str):
    """Write a pivot table to an Excel sheet with formatting."""
    if pivot.empty:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = f"No data for {sheet_name}"
        return

    ws = wb.create_sheet(sheet_name)

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    total_fill = PatternFill('solid', fgColor='D6E4F0')
    total_font = Font(bold=True, size=11)
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws['A1'] = f"FGIS Export Inspections — {sheet_name} (Metric Tons)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(pivot.columns) + 1, 15))

    # Column headers (row 3)
    start_row = 3
    ws.cell(row=start_row, column=1, value='Destination').font = header_font_white
    ws.cell(row=start_row, column=1).fill = header_fill
    ws.cell(row=start_row, column=1).border = thin_border

    for col_idx, col_name in enumerate(pivot.columns, start=2):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for row_idx, (dest, row_data) in enumerate(pivot.iterrows(), start=start_row + 1):
        is_total_row = (dest == 'TOTAL')

        cell = ws.cell(row=row_idx, column=1, value=dest)
        cell.font = total_font if is_total_row else Font(size=11)
        cell.border = thin_border
        if is_total_row:
            cell.fill = total_fill

        for col_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val != 0 else '')
            cell.number_format = number_fmt
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            if is_total_row:
                cell.fill = total_fill
                cell.font = total_font

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_idx in range(2, len(pivot.columns) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Freeze panes
    ws.freeze_panes = 'B4'


def generate_workbook(
    marketing_years: List[str] = None,
    calendar_year: int = None,
    grains: List[str] = None,
    include_wheat_classes: bool = True
) -> Path:
    """Generate the full inspections spreadsheet.

    Returns path to output file.
    """
    if grains is None:
        grains = ['corn', 'soybeans', 'wheat', 'sorghum']

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    period_label = ''
    if marketing_years:
        period_label = f"MY{'_'.join(marketing_years)}"
    elif calendar_year:
        period_label = f"CY{calendar_year}"
    else:
        period_label = f"ALL"

    for grain in grains:
        logger.info(f"Generating {grain.title()} tab...")
        df = get_monthly_data(grain, marketing_years, calendar_year)
        pivot = build_pivot(df, marketing_years, calendar_year, grain)
        write_sheet(wb, grain.title(), pivot, grain)

        # Wheat subclasses
        if grain.lower() == 'wheat' and include_wheat_classes:
            for wc_code, wc_label in [('HRW', 'Wheat HRW'), ('SRW', 'Wheat SRW'),
                                       ('HRS', 'Wheat HRS'), ('DUWH', 'Wheat Durum'),
                                       ('SWW', 'Wheat White')]:
                logger.info(f"  Generating {wc_label} tab...")
                df_wc = get_monthly_data('wheat', marketing_years, calendar_year, wheat_class=wc_code)
                if not df_wc.empty:
                    pivot_wc = build_pivot(df_wc, marketing_years, calendar_year, grain)
                    write_sheet(wb, wc_label, pivot_wc, grain)

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d')
    filename = f"fgis_inspections_{period_label}_{timestamp}.xlsx"
    output_path = OUTPUT_DIR / filename
    wb.save(str(output_path))
    logger.info(f"Saved: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Generate FGIS Inspections Spreadsheet')
    parser.add_argument('--my', nargs='+', help='Marketing years (e.g. 2526 2425)')
    parser.add_argument('--calendar-year', type=int, help='Calendar year (e.g. 2025)')
    parser.add_argument('--grains', nargs='+', default=['corn', 'soybeans', 'wheat', 'sorghum'],
                        help='Grains to include')
    parser.add_argument('--no-wheat-classes', action='store_true',
                        help='Skip wheat subclass tabs')
    args = parser.parse_args()

    # Default: current and previous marketing year
    marketing_years = args.my
    calendar_year = args.calendar_year

    if not marketing_years and not calendar_year:
        # Default to recent marketing years
        now = date.today()
        yr = now.year % 100  # 2-digit
        # Current MY: e.g. if it's March 2026, current corn MY is 2526
        current_my = f"{yr-1:02d}{yr:02d}"
        prev_my = f"{yr-2:02d}{yr-1:02d}"
        marketing_years = [prev_my, current_my]
        logger.info(f"Defaulting to marketing years: {marketing_years}")

    output_path = generate_workbook(
        marketing_years=marketing_years,
        calendar_year=calendar_year,
        grains=args.grains,
        include_wheat_classes=not args.no_wheat_classes
    )

    print(f"\nOutput: {output_path}")


if __name__ == '__main__':
    main()
