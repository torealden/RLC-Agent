"""
Create Excel template files for RIN and Feedstock data updaters.

Generates:
  1. us_rin_data.xlsx - RIN data workbook (4 sheets)
  2. us_feedstock_data.xlsx - EIA feedstock/energy workbook (2 sheets)

Save as .xlsm after importing VBA modules.
"""

import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# Output directory
OUT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "..", "domain_knowledge", "spreadsheet_samples", "biofuel sheets"
)

# ── Shared styling ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True)
UNIT_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
DATA_FONT = Font(name="Calibri", size=11)
CATEGORY_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
UNIT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

NUMBER_FMT = '#,##0'
PCT_FMT = '0.0%'


def style_header_row(ws, row, max_col, font=HEADER_FONT, fill=HEADER_FILL):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        if fill == HEADER_FILL:
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_unit_row(ws, row, max_col):
    """Apply unit row styling."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = UNIT_FONT
        cell.fill = UNIT_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def set_column_widths(ws, widths):
    """Set column widths from a dict of {col_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def freeze_panes(ws, cell_ref):
    """Freeze panes at the given cell reference."""
    ws.freeze_panes = cell_ref


# ── RIN Workbook ────────────────────────────────────────────────────────────

def create_rin_monthly(wb):
    """Sheet 1: RIN Monthly - monthly time series."""
    ws = wb.active
    ws.title = "RIN Monthly"

    # Headers (Row 3)
    headers = [
        "Date",
        "D3 Cellulosic", "D4 BBD", "D5 Advanced", "D6 Renewable", "D7 Cel. Diesel",
        "Total RINs",
        "D3 Volume", "D4 Volume", "D5 Volume", "D6 Volume", "D7 Volume",
        "Total Volume"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Units (Row 4)
    units = [
        "",
        "RINs", "RINs", "RINs", "RINs", "RINs",
        "RINs",
        "Gallons", "Gallons", "Gallons", "Gallons", "Gallons",
        "Gallons"
    ]
    for col, u in enumerate(units, 1):
        ws.cell(row=4, column=col, value=u)

    # Title (Row 1)
    ws.cell(row=1, column=1, value="EPA RFS - Monthly RIN Generation by D-Code")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    # Style header and unit rows
    max_col = len(headers)
    style_header_row(ws, 3, max_col)
    style_unit_row(ws, 4, max_col)

    # Data rows: Jan 2010 through Dec 2026
    row = 5
    for year in range(2010, 2027):
        for month in range(1, 13):
            dt = date(year, month, 1)
            ws.cell(row=row, column=1, value=dt)
            ws.cell(row=row, column=1).number_format = 'MMM YYYY'
            ws.cell(row=row, column=1).alignment = LEFT_ALIGN

            # Total RINs formula (Col G = SUM B:F)
            ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
            ws.cell(row=row, column=7).number_format = NUMBER_FMT

            # Total Volume formula (Col M = SUM H:L)
            ws.cell(row=row, column=13, value=f"=SUM(H{row}:L{row})")
            ws.cell(row=row, column=13).number_format = NUMBER_FMT

            # Number format for data columns
            for c in range(2, 7):
                ws.cell(row=row, column=c).number_format = NUMBER_FMT
            for c in range(8, 13):
                ws.cell(row=row, column=c).number_format = NUMBER_FMT

            # Alternating row fill
            if year % 2 == 0:
                for c in range(1, max_col + 1):
                    ws.cell(row=row, column=c).fill = ALT_FILL

            # Border
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER

            row += 1

    # Column widths
    set_column_widths(ws, {
        'A': 14, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16,
        'G': 16, 'H': 16, 'I': 16, 'J': 16, 'K': 16, 'L': 16, 'M': 16
    })

    freeze_panes(ws, "B5")
    return ws


def create_annual_generation(wb):
    """Sheet 2: Annual Generation - annual totals."""
    ws = wb.create_sheet("Annual Generation")

    # Title
    ws.cell(row=1, column=1, value="EPA RFS - Annual RIN Generation Summary")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    # Headers (Row 3)
    headers = [
        "Year",
        "D3 Cellulosic", "D4 BBD", "D5 Advanced", "D6 Renewable",
        "D7 Cel. Diesel", "Total RINs", "Total Advanced"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Units (Row 4)
    units = ["", "RINs", "RINs", "RINs", "RINs", "RINs", "RINs", "RINs"]
    for col, u in enumerate(units, 1):
        ws.cell(row=4, column=col, value=u)

    max_col = len(headers)
    style_header_row(ws, 3, max_col)
    style_unit_row(ws, 4, max_col)

    # Data rows: 2010-2026
    row = 5
    for year in range(2010, 2027):
        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=1).alignment = CENTER

        for c in range(2, max_col + 1):
            ws.cell(row=row, column=c).number_format = NUMBER_FMT
            ws.cell(row=row, column=c).border = THIN_BORDER

        ws.cell(row=row, column=1).border = THIN_BORDER

        if year % 2 == 0:
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = ALT_FILL

        row += 1

    set_column_widths(ws, {
        'A': 10, 'B': 18, 'C': 18, 'D': 18, 'E': 18,
        'F': 18, 'G': 18, 'H': 18
    })

    freeze_panes(ws, "B5")


def create_rin_balance(wb):
    """Sheet 3: RIN Balance - annual gen/ret/avail by D-code."""
    ws = wb.create_sheet("RIN Balance")

    # Title
    ws.cell(row=1, column=1, value="EPA RFS - Annual RIN Balance (Generated / Retired / Available)")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    # Category headers (Row 2) - merged groups
    categories = [
        (1, 1, ""),
        (2, 4, "D3 Cellulosic"),
        (5, 7, "D4 Biomass-Based Diesel"),
        (8, 10, "D5 Advanced"),
        (11, 13, "D6 Renewable"),
        (14, 16, "Totals"),
    ]
    for start_col, end_col, label in categories:
        ws.cell(row=2, column=start_col, value=label)
        if start_col != end_col:
            ws.merge_cells(
                start_row=2, start_column=start_col,
                end_row=2, end_column=end_col
            )
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=2, column=c)
            cell.font = CATEGORY_FONT
            cell.fill = CATEGORY_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # Sub-headers (Row 3)
    sub_headers = [
        "Year",
        "Generated", "Retired", "Available",
        "Generated", "Retired", "Available",
        "Generated", "Retired", "Available",
        "Generated", "Retired", "Available",
        "Generated", "Retired", "Available",
    ]
    for col, h in enumerate(sub_headers, 1):
        ws.cell(row=3, column=col, value=h)

    max_col = len(sub_headers)
    style_header_row(ws, 3, max_col)

    # Units (Row 4)
    for col in range(1, max_col + 1):
        ws.cell(row=4, column=col, value="RINs" if col > 1 else "")
    style_unit_row(ws, 4, max_col)

    # Data rows: 2010-2026
    row = 5
    for year in range(2010, 2027):
        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=1).alignment = CENTER

        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).number_format = NUMBER_FMT
            ws.cell(row=row, column=c).border = THIN_BORDER

        if year % 2 == 0:
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = ALT_FILL

        row += 1

    set_column_widths(ws, {
        'A': 10, 'B': 16, 'C': 16, 'D': 16,
        'E': 16, 'F': 16, 'G': 16,
        'H': 16, 'I': 16, 'J': 16,
        'K': 16, 'L': 16, 'M': 16,
        'N': 16, 'O': 16, 'P': 16,
    })

    freeze_panes(ws, "B5")


def create_d4_fuel_mix(wb):
    """Sheet 4: D4 Fuel Mix - D4 fuel production breakdown."""
    ws = wb.create_sheet("D4 Fuel Mix")

    # Title
    ws.cell(row=1, column=1, value="EPA RFS - D4 (BBD) Fuel Production by Type")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    # Category headers (Row 2)
    vol_cats = [
        (1, 1, ""),
        (2, 7, "Physical Volume (Gallons)"),
        (8, 13, "RIN Quantity"),
    ]
    for start_col, end_col, label in vol_cats:
        ws.cell(row=2, column=start_col, value=label)
        if start_col != end_col:
            ws.merge_cells(
                start_row=2, start_column=start_col,
                end_row=2, end_column=end_col
            )
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=2, column=c)
            cell.font = CATEGORY_FONT
            cell.fill = CATEGORY_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # Headers (Row 3)
    headers = [
        "Year",
        "Biodiesel", "RD (EV 1.7)", "RD (EV 1.6)", "Ren Jet", "Other", "Total Vol",
        "Biodiesel", "RD (EV 1.7)", "RD (EV 1.6)", "Ren Jet", "Other", "Total RINs"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Units (Row 4)
    units = [
        "",
        "Gallons", "Gallons", "Gallons", "Gallons", "Gallons", "Gallons",
        "RINs", "RINs", "RINs", "RINs", "RINs", "RINs"
    ]
    for col, u in enumerate(units, 1):
        ws.cell(row=4, column=col, value=u)

    max_col = len(headers)
    style_header_row(ws, 3, max_col)
    style_unit_row(ws, 4, max_col)

    # Data rows: 2010-2026
    row = 5
    for year in range(2010, 2027):
        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=1).alignment = CENTER

        # Total Volume formula (Col G = SUM B:F)
        ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
        ws.cell(row=row, column=7).number_format = NUMBER_FMT

        # Total RINs formula (Col M = SUM H:L)
        ws.cell(row=row, column=13, value=f"=SUM(H{row}:L{row})")
        ws.cell(row=row, column=13).number_format = NUMBER_FMT

        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).number_format = NUMBER_FMT
            ws.cell(row=row, column=c).border = THIN_BORDER

        if year % 2 == 0:
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = ALT_FILL

        row += 1

    set_column_widths(ws, {
        'A': 10, 'B': 16, 'C': 16, 'D': 16, 'E': 14, 'F': 14, 'G': 16,
        'H': 16, 'I': 16, 'J': 16, 'K': 14, 'L': 14, 'M': 16
    })

    freeze_panes(ws, "B5")


# ── Feedstock Workbook ──────────────────────────────────────────────────────

def generate_fridays(start_date, end_date):
    """Generate all Friday dates between start and end."""
    # Find the first Friday on or after start_date
    d = start_date
    while d.weekday() != 4:  # 4 = Friday
        d += timedelta(days=1)

    dates = []
    while d <= end_date:
        dates.append(d)
        d += timedelta(days=7)
    return dates


def create_ethanol_weekly(wb):
    """Sheet 1: Ethanol Weekly."""
    ws = wb.active
    ws.title = "Ethanol Weekly"

    # Title
    ws.cell(row=1, column=1, value="EIA - Weekly Ethanol Data")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    # Headers (Row 3)
    headers = [
        "Week Ending",
        "Production", "Stocks", "Blender Input", "Balance",
        "Prod YoY", "Stocks YoY"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Units (Row 4)
    units = ["", "kbd", "kb", "kbd", "kbd", "%", "%"]
    for col, u in enumerate(units, 1):
        ws.cell(row=4, column=col, value=u)

    max_col = len(headers)
    style_header_row(ws, 3, max_col)
    style_unit_row(ws, 4, max_col)

    # Generate Friday dates from Jan 2020 through Dec 2026
    fridays = generate_fridays(date(2020, 1, 1), date(2026, 12, 31))

    row = 5
    for i, friday in enumerate(fridays):
        ws.cell(row=row, column=1, value=friday)
        ws.cell(row=row, column=1).number_format = 'M/D/YYYY'
        ws.cell(row=row, column=1).alignment = LEFT_ALIGN

        # Number formats for data columns
        for c in range(2, 5):
            ws.cell(row=row, column=c).number_format = '#,##0'
        ws.cell(row=row, column=5).number_format = '#,##0'

        # YoY formulas (compare to same row 52 weeks ago)
        if row >= 5 + 52:
            ws.cell(row=row, column=6, value=f"=IF(B{row - 52}<>0,(B{row}-B{row - 52})/B{row - 52},\"\")")
            ws.cell(row=row, column=6).number_format = '0.0%'
            ws.cell(row=row, column=7, value=f"=IF(C{row - 52}<>0,(C{row}-C{row - 52})/C{row - 52},\"\")")
            ws.cell(row=row, column=7).number_format = '0.0%'

        # Border
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

        # Alternating year fill
        if friday.year % 2 == 0:
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = ALT_FILL

        row += 1

    set_column_widths(ws, {
        'A': 14, 'B': 14, 'C': 14, 'D': 16, 'E': 14, 'F': 12, 'G': 12
    })

    freeze_panes(ws, "B5")


def create_petroleum_weekly(wb):
    """Sheet 2: Petroleum Weekly."""
    ws = wb.create_sheet("Petroleum Weekly")

    # Title
    ws.cell(row=1, column=1, value="EIA - Weekly Petroleum Supply Data")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    # Category headers (Row 2)
    cats = [
        (1, 1, ""),
        (2, 6, "Stocks"),
        (7, 8, "Production & Imports"),
        (9, 10, "Refinery"),
        (11, 11, "Gasoline"),
    ]
    for start_col, end_col, label in cats:
        ws.cell(row=2, column=start_col, value=label)
        if start_col != end_col:
            ws.merge_cells(
                start_row=2, start_column=start_col,
                end_row=2, end_column=end_col
            )
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=2, column=c)
            cell.font = CATEGORY_FONT
            cell.fill = CATEGORY_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # Headers (Row 3)
    headers = [
        "Week Ending",
        "Crude Total", "Crude ex-SPR", "SPR", "Gasoline", "Distillate",
        "Crude Prod", "Crude Imports",
        "Ref Inputs", "Ref Util",
        "Days Supply"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Units (Row 4)
    units = [
        "",
        "kb", "kb", "kb", "kb", "kb",
        "kbd", "kbd",
        "kbd", "%",
        "days"
    ]
    for col, u in enumerate(units, 1):
        ws.cell(row=4, column=col, value=u)

    max_col = len(headers)
    style_header_row(ws, 3, max_col)
    style_unit_row(ws, 4, max_col)

    # Generate Friday dates from Jan 2020 through Dec 2026
    fridays = generate_fridays(date(2020, 1, 1), date(2026, 12, 31))

    row = 5
    for friday in fridays:
        ws.cell(row=row, column=1, value=friday)
        ws.cell(row=row, column=1).number_format = 'M/D/YYYY'
        ws.cell(row=row, column=1).alignment = LEFT_ALIGN

        # Number formats
        for c in range(2, 9):
            ws.cell(row=row, column=c).number_format = '#,##0'
        ws.cell(row=row, column=9).number_format = '#,##0'
        ws.cell(row=row, column=10).number_format = '0.0'
        ws.cell(row=row, column=11).number_format = '0.0'

        # Border
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

        if friday.year % 2 == 0:
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = ALT_FILL

        row += 1

    set_column_widths(ws, {
        'A': 14, 'B': 14, 'C': 14, 'D': 12, 'E': 14, 'F': 14,
        'G': 14, 'H': 14, 'I': 14, 'J': 10, 'K': 12
    })

    freeze_panes(ws, "B5")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    # ── RIN Workbook ──
    print("Creating us_rin_data.xlsx...")
    wb_rin = Workbook()
    create_rin_monthly(wb_rin)
    create_annual_generation(wb_rin)
    create_rin_balance(wb_rin)
    create_d4_fuel_mix(wb_rin)

    rin_path = os.path.join(OUT_DIR, "us_rin_data.xlsx")
    wb_rin.save(rin_path)
    print(f"  Saved: {rin_path}")
    print(f"  Sheets: {wb_rin.sheetnames}")

    # ── Feedstock Workbook ──
    print("Creating us_feedstock_data.xlsx...")
    wb_feed = Workbook()
    create_ethanol_weekly(wb_feed)
    create_petroleum_weekly(wb_feed)

    feed_path = os.path.join(OUT_DIR, "us_feedstock_data.xlsx")
    wb_feed.save(feed_path)
    print(f"  Saved: {feed_path}")
    print(f"  Sheets: {wb_feed.sheetnames}")

    print("\nDone! Next steps:")
    print("  1. Open each .xlsx in Excel")
    print("  2. Alt+F11 > Tools > References > check 'Microsoft ActiveX Data Objects 6.1 Library'")
    print("  3. Import the matching .bas file (RINUpdaterSQL.bas or FeedstockUpdaterSQL.bas)")
    print("  4. Paste the matching WorkbookEvents code into ThisWorkbook")
    print("  5. Save As .xlsm (macro-enabled workbook)")
    print("  6. Close and reopen — shortcuts will auto-assign")


if __name__ == "__main__":
    main()
