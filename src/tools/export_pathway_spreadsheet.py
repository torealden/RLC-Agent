"""Export EPA pathway determination data to Excel spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2
import psycopg2.extras
from pathlib import Path

DB_CONFIG = {
    'host': 'localhost', 'port': 5432, 'database': 'rlc_commodities',
    'user': 'postgres', 'password': 'SoupBoss1',
}

OUTPUT_DIR = Path(__file__).resolve().parent.parent.parent / 'data' / 'epa_pathways'


def main():
    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cursor.execute("""
        SELECT
            i.id,
            i.determination_name,
            i.category,
            i.determination_date,
            i.fuel_type AS index_fuel_type,
            i.feedstock AS index_feedstock,
            i.d_code AS index_d_code,
            d.company_name,
            d.facility_city,
            d.facility_state,
            d.recipient_name,
            d.recipient_title,
            d.mailing_address,
            ARRAY_TO_STRING(d.fuel_types, '; ') AS parsed_fuel_types,
            ARRAY_TO_STRING(d.feedstocks, '; ') AS parsed_feedstocks,
            d.d_code AS parsed_d_code,
            d.production_process,
            ARRAY_TO_STRING(d.process_energy_sources, '; ') AS energy_sources,
            d.ghg_reduction_pct,
            d.lifecycle_ghg_gco2e_mj,
            d.ghg_baseline_gco2e_mj,
            d.pathway_name,
            d.table1_row_reference,
            d.parse_confidence,
            d.page_count,
            i.pdf_filename
        FROM bronze.epa_pathway_index i
        LEFT JOIN bronze.epa_pathway_detail d ON d.pathway_index_id = i.id
        ORDER BY i.category DESC, i.determination_date DESC
    """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    index_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    needs_review_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    error_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        ('ID', 6),
        ('Determination Name', 30),
        ('Category', 10),
        ('Date', 12),
        ('Fuel Type (Index)', 35),
        ('Feedstock (Index)', 30),
        ('D-Code (Index)', 12),
        ('Company Name', 35),
        ('Facility City', 20),
        ('State', 8),
        ('Contact Name', 25),
        ('Contact Title', 30),
        ('Mailing Address', 35),
        ('Fuel Types (Parsed)', 35),
        ('Feedstocks (Parsed)', 30),
        ('D-Code (Parsed)', 12),
        ('Production Process', 30),
        ('Energy Sources', 25),
        ('GHG Reduction %', 14),
        ('Lifecycle GHG (gCO2e/MJ)', 20),
        ('GHG Baseline (gCO2e/MJ)', 20),
        ('Pathway Name', 35),
        ('Table 1 Row', 10),
        ('Parse Confidence', 14),
        ('Pages', 7),
        ('PDF Filename', 50),
    ]

    def write_sheet(ws, data_rows, sheet_headers):
        for col, (name, width) in enumerate(sheet_headers, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(sheet_headers))}1'

        for row_idx, row in enumerate(data_rows, 2):
            data = [
                row['id'],
                row['determination_name'],
                row['category'].upper(),
                row['determination_date'],
                row['index_fuel_type'],
                row['index_feedstock'],
                row['index_d_code'],
                row['company_name'],
                row['facility_city'],
                row['facility_state'],
                row['recipient_name'],
                row['recipient_title'],
                row['mailing_address'],
                row['parsed_fuel_types'],
                row['parsed_feedstocks'],
                row['parsed_d_code'],
                row['production_process'],
                row['energy_sources'],
                float(row['ghg_reduction_pct']) if row['ghg_reduction_pct'] else None,
                float(row['lifecycle_ghg_gco2e_mj']) if row['lifecycle_ghg_gco2e_mj'] else None,
                float(row['ghg_baseline_gco2e_mj']) if row['ghg_baseline_gco2e_mj'] else None,
                row['pathway_name'],
                row['table1_row_reference'],
                row['parse_confidence'],
                row['page_count'],
                row['pdf_filename'],
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

                confidence = row['parse_confidence']
                if 8 <= col <= 23:
                    if confidence == 'low':
                        cell.fill = error_fill
                    elif confidence == 'medium':
                        cell.fill = needs_review_fill
                elif 5 <= col <= 7:
                    cell.fill = index_fill

            if data[3]:
                ws.cell(row=row_idx, column=4).number_format = 'YYYY-MM-DD'
            if data[18]:
                ws.cell(row=row_idx, column=19).number_format = '0.0'

    # Sheet 1: All Plants
    ws1 = wb.active
    ws1.title = 'All Plants'
    write_sheet(ws1, rows, headers)

    # Sheet 2: Non-EP3 Only
    ws2 = wb.create_sheet('Non-EP3 (BD-RD-SAF)')
    non_ep3 = [r for r in rows if r['category'] == 'non_ep3']
    write_sheet(ws2, non_ep3, headers)

    # Sheet 3: Legend
    ws3 = wb.create_sheet('Legend')
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 80

    legend_data = [
        ('EPA Pathway Determination Letters', '', True, None),
        ('', '', False, None),
        ('Data Source', 'https://www.epa.gov/renewable-fuel-standard/approved-pathways-renewable-fuel', False, None),
        ('Total Determinations', str(len(rows)), False, None),
        ('Non-EP3 (BD, RD, SAF, Novel)', str(len(non_ep3)), False, None),
        ('EP3 (Corn/Sorghum Ethanol)', str(len(rows) - len(non_ep3)), False, None),
        ('', '', False, None),
        ('Color Coding', '', True, None),
        ('Blue columns (E-G)', 'Index data from HTML table - always available', False, index_fill),
        ('No fill', 'High-confidence parsed data from PDF', False, None),
        ('Yellow fill', 'Medium-confidence parse - some fields may need review', False, needs_review_fill),
        ('Orange fill', 'Low-confidence parse - PDF was scanned or unusual format. Please fill in.', False, error_fill),
        ('', '', False, None),
        ('PDF Location', r'C:\dev\RLC-Agent\data\epa_pathways\pdfs\\', True, None),
        ('', 'The PDF Filename column (Z) maps to files in this folder.', False, None),
        ('', 'Orange-highlighted rows need manual review of the PDF.', False, None),
        ('', '', False, None),
        ('Categories', '', True, None),
        ('NON_EP3', 'Facility-specific approvals for novel fuels/feedstocks (biodiesel, RD, SAF, CNG, DME)', False, None),
        ('EP3', 'Efficient Producer Petition - corn/sorghum ethanol plants demonstrating GHG compliance', False, None),
        ('', '', False, None),
        ('D-Codes', '', True, None),
        ('D3', 'Cellulosic Biofuel (60% GHG reduction minimum)', False, None),
        ('D4', 'Biomass-Based Diesel (50% GHG reduction minimum)', False, None),
        ('D5', 'Advanced Biofuel (50% GHG reduction minimum)', False, None),
        ('D6', 'Renewable Fuel / Conventional (20% GHG reduction minimum)', False, None),
        ('D7', 'Cellulosic Diesel (60% GHG reduction minimum)', False, None),
    ]

    for i, (label, desc, bold, fill) in enumerate(legend_data, 1):
        cell_a = ws3.cell(row=i, column=1, value=label)
        cell_b = ws3.cell(row=i, column=2, value=desc)
        if bold:
            cell_a.font = Font(bold=True, size=12 if i == 1 else 11)
        if fill:
            cell_a.fill = fill

    output_path = OUTPUT_DIR / 'epa_pathway_plants.xlsx'
    wb.save(str(output_path))

    print(f"Saved: {output_path}")
    print(f"Total rows: {len(rows)}")
    print(f"  Non-EP3: {len(non_ep3)}")
    print(f"  EP3: {len(rows) - len(non_ep3)}")
    low = sum(1 for r in rows if r['parse_confidence'] == 'low')
    med = sum(1 for r in rows if r['parse_confidence'] == 'medium')
    print(f"  Needs review: {low} orange (low) + {med} yellow (medium)")


if __name__ == '__main__':
    main()
