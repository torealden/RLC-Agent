"""
Populate the monthly Imports and Exports blocks of the US biodiesel and
renewable diesel balance sheet workbooks from gold.biofuel_trade_split.

Source view: gold.biofuel_trade_split (mig 082-086) — heuristic BD/RD allocation
of HS 3826 Census trade. Output in million gallons (workbook unit).

Workbooks:
  output/balance_sheet_templates/us_biodiesel_bal_sheets.xlsx        → BIODIESEL
  output/balance_sheet_templates/us_renewable_diesel_bal_sheets.xlsx → RENEWABLE_DIESEL

Workbook layout (Year × Month grid):
  Row 3:   year headers, col B = 2001, col AE = 2030
  Rows 42-53: IMPORTS Jan-Dec (row 54 is annual SUM formula)
  Rows 58-69: EXPORTS Jan-Dec (row 70 is annual SUM formula)

Behavior:
  - Clears any stale "='[1]...'" external-link references in the trade rows.
  - Writes monthly data only where gold.biofuel_trade_split has it
    (Census starts Jan 2013). Pre-2013 cells are left untouched to preserve
    legacy entries the workbook already contains.
  - Unit: gold view emits gallons; we write to workbook as million gallons (÷1e6).

Usage:
  python src/tools/populate_biofuel_bal_sheet_trade.py             # both workbooks
  python src/tools/populate_biofuel_bal_sheet_trade.py --commodity biodiesel
  python src/tools/populate_biofuel_bal_sheet_trade.py --dry-run
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")

import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl

YEAR_HEADER_ROW = 3
PRODUCTION_JAN_ROW = 26  # Jan; Feb=27, ..., Dec=37
IMPORTS_JAN_ROW = 42     # Jan; Feb=43, ..., Dec=53
EXPORTS_JAN_ROW = 58

WORKBOOKS = {
    "biodiesel": {
        "path":         ROOT / "output" / "balance_sheet_templates" / "us_biodiesel_bal_sheets.xlsx",
        "sheet":        "Biodiesel",
        "commodity":    "BIODIESEL",
        "trade_source": "biofuel_split",   # gold.biofuel_trade_split union gold.trade_export_mapped
    },
    "renewable_diesel": {
        "path":         ROOT / "output" / "balance_sheet_templates" / "us_renewable_diesel_bal_sheets.xlsx",
        "sheet":        "Renewable Diesel",
        "commodity":    "RENEWABLE_DIESEL",
        "trade_source": "biofuel_split",
    },
    "saf": {
        "path":         ROOT / "output" / "balance_sheet_templates" / "us_saf_bal_sheets.xlsx",
        "sheet":        "SAF",
        "commodity":    "SAF",
        "trade_source": "saf_candidates",   # gold.saf_trade_candidates (price-threshold view)
        "populate_production": True,         # also fill Production block from EMTS
    },
}


def _get_conn():
    return psycopg2.connect(
        host=os.getenv("RLC_PG_HOST"), port=os.getenv("RLC_PG_PORT", "5432"),
        dbname=os.getenv("RLC_PG_DB", "rlc_commodities"),
        user=os.getenv("RLC_PG_USER"), password=os.getenv("RLC_PG_PASSWORD"),
        sslmode="require",
    )


def fetch_trade_biofuel_split(commodity: str) -> dict[tuple[int, int, str], float]:
    """BD/RD: combine HS 3826 split + HS 2710.20 direct."""
    conn = _get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        WITH hs_3826_split AS (
            SELECT year, month, flow, SUM(quantity_gal) / 1e6 AS mil_gal
            FROM gold.biofuel_trade_split
            WHERE commodity_split = %(commodity)s
            GROUP BY year, month, flow
        ),
        hs_271020 AS (
            SELECT year, month, LOWER(flow) AS flow,
                   SUM(quantity_converted) / 1000.0 AS mil_gal
            FROM gold.trade_export_mapped
            WHERE hs_code LIKE '271020%%' AND commodity_group = %(commodity)s
              AND NOT is_regional_total AND country_code <> '-'
            GROUP BY year, month, flow
        )
        SELECT year, month, flow, SUM(mil_gal) AS mil_gal FROM (
            SELECT * FROM hs_3826_split UNION ALL SELECT * FROM hs_271020
        ) c GROUP BY year, month, flow ORDER BY year, month, flow
        """,
        {"commodity": commodity},
    )
    return {(r["year"], r["month"], r["flow"]): float(r["mil_gal"]) for r in cur.fetchall()}


def fetch_trade_saf() -> dict[tuple[int, int, str], float]:
    """SAF: price-threshold candidates across HS 3826 + 2710.19.11 + 2710.20."""
    conn = _get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT year, month, flow, SUM(quantity_gal) / 1e6 AS mil_gal
        FROM gold.saf_trade_candidates
        GROUP BY year, month, flow ORDER BY year, month, flow
        """
    )
    return {(r["year"], r["month"], r["flow"]): float(r["mil_gal"]) for r in cur.fetchall()}


def fetch_saf_production() -> dict[tuple[int, int], float]:
    """SAF production from EMTS, in mil gal."""
    conn = _get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT year, month, saf_kgal / 1000.0 AS mil_gal
        FROM gold.us_liquid_fuel_production_monthly
        WHERE saf_kgal IS NOT NULL
        ORDER BY year, month
        """
    )
    return {(r["year"], r["month"]): float(r["mil_gal"]) for r in cur.fetchall()}


def fetch_trade(commodity: str, trade_source: str) -> dict[tuple[int, int, str], float]:
    """Dispatch to the right fetcher per workbook."""
    if trade_source == "saf_candidates":
        return fetch_trade_saf()
    return fetch_trade_biofuel_split(commodity)


def build_year_col_map(ws) -> dict[int, int]:
    """Map year integer → column letter index."""
    out: dict[int, int] = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(YEAR_HEADER_ROW, c).value
        if isinstance(v, int):
            out[v] = c
    return out


def clear_stale_links(ws, row_start: int) -> int:
    """Clear external-link references in 12 month rows for all columns. Returns count cleared."""
    cleared = 0
    for r in range(row_start, row_start + 12):
        for c in range(2, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("='[1]"):
                ws.cell(r, c).value = None
                cleared += 1
    return cleared


def populate(spec: dict, trade: dict[tuple[int, int, str], float], dry_run: bool) -> tuple[int, int]:
    """Write trade into a workbook; returns (imports_written, exports_written)."""
    path = spec["path"]
    if not path.exists():
        print(f"[ERROR] Workbook not found: {path}")
        return (0, 0)

    print(f"\n--- Populating {path.name} [{spec['sheet']}] for commodity {spec['commodity']} ---")
    wb = openpyxl.load_workbook(path)
    if spec["sheet"] not in wb.sheetnames:
        print(f"[ERROR] Sheet {spec['sheet']!r} not in {wb.sheetnames}")
        return (0, 0)
    ws = wb[spec["sheet"]]

    year_cols = build_year_col_map(ws)
    print(f"  Year columns mapped: {min(year_cols)} → {max(year_cols)} ({len(year_cols)} years)")

    n_cleared_imp = clear_stale_links(ws, IMPORTS_JAN_ROW)
    n_cleared_exp = clear_stale_links(ws, EXPORTS_JAN_ROW)
    if n_cleared_imp or n_cleared_exp:
        print(f"  Cleared stale external-link references: {n_cleared_imp} imports, {n_cleared_exp} exports")

    imp_written = exp_written = 0
    skipped_years: set[int] = set()
    for (yr, mo, flow), mil_gal in trade.items():
        if yr not in year_cols:
            skipped_years.add(yr)
            continue
        col = year_cols[yr]
        if flow == "imports":
            row = IMPORTS_JAN_ROW + (mo - 1)
            ws.cell(row, col).value = mil_gal
            imp_written += 1
        elif flow == "exports":
            row = EXPORTS_JAN_ROW + (mo - 1)
            ws.cell(row, col).value = mil_gal
            exp_written += 1

    print(f"  Imports cells written: {imp_written}")
    print(f"  Exports cells written: {exp_written}")
    if skipped_years:
        print(f"  Skipped years not in header: {sorted(skipped_years)}")

    # Optional: populate the Production block (currently SAF only)
    if spec.get("populate_production"):
        prod = fetch_saf_production()
        prod_written = 0
        for (yr, mo), mil_gal in prod.items():
            if yr not in year_cols:
                continue
            col = year_cols[yr]
            row = PRODUCTION_JAN_ROW + (mo - 1)
            ws.cell(row, col).value = mil_gal
            prod_written += 1
        print(f"  Production cells written: {prod_written}")

    if dry_run:
        print("  [DRY RUN] not saving")
    else:
        wb.save(path)
        print(f"  Saved {path.name}")

    return (imp_written, exp_written)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commodity", choices=list(WORKBOOKS.keys()) + ["all"], default="all")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    keys = list(WORKBOOKS.keys()) if args.commodity == "all" else [args.commodity]
    for key in keys:
        spec = WORKBOOKS[key]
        trade = fetch_trade(spec["commodity"], spec.get("trade_source", "biofuel_split"))
        print(f"\nFetched {len(trade)} (year, month, flow) rows for {spec['commodity']}")
        populate(spec, trade, args.dry_run)


if __name__ == "__main__":
    main()
