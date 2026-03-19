"""
Generate Weekly "Cash Prices - MMDDYYYY.xlsx" for HB Report

Copies the formatting-preserved template and fills in prices from the database
using extract_hb_prices(). Works headless (scheduler) or interactive (.bat).

Usage:
    python src/tools/generate_cash_prices.py                    # Today's date
    python src/tools/generate_cash_prices.py --date 2026-03-05  # Specific date
    python src/tools/generate_cash_prices.py --output path.xlsx  # Custom output
"""

import os
import sys
import argparse
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from copy import copy

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / '.env')
except ImportError:
    pass

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

from src.tools.hb_cash_price_extract import extract_hb_prices

TEMPLATE_PATH = PROJECT_ROOT / 'templates' / 'Cash Prices - template.xlsx'
OUTPUT_DIR = PROJECT_ROOT / 'output' / 'reports'

# Rows containing price data (D/E/F columns to clear and fill)
PRICE_ROW_MIN = 5
PRICE_ROW_MAX = 52


def generate_cash_prices(report_date: date = None, output_path: Path = None) -> Path:
    """
    Generate the weekly Cash Prices spreadsheet.

    Args:
        report_date: Reference date (default: today)
        output_path: Custom output path (default: data/reports/Cash Prices - MMDDYYYY.xlsx)

    Returns:
        Path to the generated file
    """
    report_date = report_date or date.today()
    week_ago = report_date - timedelta(days=7)
    year_ago = report_date - timedelta(days=365)

    # Determine output path
    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        filename = f"Cash Prices - {report_date.strftime('%m%d%Y')}.xlsx"
        output_path = OUTPUT_DIR / filename

    # Load template (preserves all formatting, labels, column widths)
    if not TEMPLATE_PATH.exists():
        print(f"ERROR: Template not found at {TEMPLATE_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
    ws = wb.worksheets[0]  # Sheet name has trailing space, use index

    # Clear old price values in D/E/F (rows 5-52), preserve formatting
    for row in range(PRICE_ROW_MIN, PRICE_ROW_MAX + 1):
        for col in (4, 5, 6):  # D, E, F
            cell = ws.cell(row=row, column=col)
            cell.value = None

    # Write dates in row 3
    ws['D3'] = datetime.combine(report_date, datetime.min.time())
    ws['E3'] = datetime.combine(week_ago, datetime.min.time())
    ws['F3'] = datetime.combine(year_ago, datetime.min.time())

    # Extract prices from database
    print(f"Querying prices for report date {report_date}...")
    result = extract_hb_prices(report_date)
    prices = result['prices']
    missing = result['missing']

    # Write prices into spreadsheet
    filled = 0
    for p in prices:
        row = p['hb_sheet_row']
        if row < PRICE_ROW_MIN or row > PRICE_ROW_MAX:
            continue

        if p['this_week'] is not None:
            ws.cell(row=row, column=4, value=p['this_week'])
            filled += 1
        if p['last_week'] is not None:
            ws.cell(row=row, column=5, value=p['last_week'])
        if p['year_ago'] is not None:
            ws.cell(row=row, column=6, value=p['year_ago'])

    # Save
    wb.save(str(output_path))

    # Print summary
    total = len(prices)
    blank = len(missing)
    print(f"\n{'='*60}")
    print(f"  Cash Prices - {report_date.strftime('%m/%d/%Y')}")
    print(f"{'='*60}")
    print(f"  Filled:  {filled}/{total} rows")
    print(f"  Blank:   {blank} rows (manual entry needed)")
    print(f"  Output:  {output_path}")
    print(f"{'='*60}")

    if missing:
        print(f"\n  Manual entry needed:")
        for m in missing:
            print(f"    - {m}")

    print()
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Generate weekly Cash Prices spreadsheet for HB report'
    )
    parser.add_argument(
        '--date', '-d',
        help='Report date (YYYY-MM-DD), default: today'
    )
    parser.add_argument(
        '--output', '-o',
        help='Custom output file path'
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable debug logging'
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format='%(asctime)s %(levelname)s %(message)s',
    )

    report_date = (
        datetime.strptime(args.date, '%Y-%m-%d').date()
        if args.date else None
    )
    output_path = Path(args.output) if args.output else None

    generate_cash_prices(report_date, output_path)


if __name__ == '__main__':
    main()
