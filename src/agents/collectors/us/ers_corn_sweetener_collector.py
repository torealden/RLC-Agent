"""
ERS Sugar & Sweeteners Yearbook — corn sweetener supply & use collector.

Source: USDA ERS "U.S. corn sweetener supply and use" workbook on the
Sugar and Sweeteners Yearbook Tables data-product page. Provides HFCS-42/55,
glucose, and dextrose production (the F/G/I/J columns of the us_grain_crush
corn_products tab). All values are 1,000 short tons, dry basis.

Tables parsed:
  Table 29 — HFCS total production, quarterly + annual (rows=period, cols=years)
  Table 30 — HFCS-42 / HFCS-55 / total production, annual (rows=years, wide cols)
  Table 37 — Dextrose supply & use, annual (rows=years; Production = col 1)
  Table 38 — Glucose supply & use, annual (rows=years; Production = col 1)

Runtime URL resolution: the /media/<id>/ file URLs change per release, so we
scrape the data-product page for the current 'corn sweetener supply and use'
xlsx link rather than hardcoding it (mirrors ers_feed_grains_collector).

Pipeline: XLSX -> bronze.corn_products_raw (long, vintage-stamped).
"""

import io
import os
import re
import sys
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import requests
import openpyxl
from dotenv import load_dotenv

load_dotenv(PROJECT_ROOT / ".env")
logger = logging.getLogger(__name__)

ERS_PAGE = "https://www.ers.usda.gov/data-products/sugar-and-sweeteners-yearbook-tables"
ERS_BASE = "https://www.ers.usda.gov"
# Fallback if page scrape fails (URL as of 2026-05 release).
FALLBACK_URL = "https://www.ers.usda.gov/media/5147/us-corn-sweetener-supply-and-use.xlsx"
HEADERS = {"User-Agent": "Mozilla/5.0 (RLC-Agent data collector)"}

_QTR = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}


class ERSCornSweetenerCollector:
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
        self.session = requests.Session()
        self.session.headers.update(HEADERS)

    # ---- URL resolution + download ----------------------------------------
    def resolve_url(self) -> str:
        try:
            r = self.session.get(ERS_PAGE, timeout=60)
            if r.status_code == 200:
                # find the corn-sweetener supply-and-use xlsx link
                for m in re.finditer(r'href="([^"]+\.xlsx[^"]*)"', r.text, re.I):
                    href = m.group(1)
                    if "corn-sweetener-supply-and-use" in href.lower():
                        return href if href.startswith("http") else ERS_BASE + href
        except requests.RequestException as e:
            self.logger.warning(f"ERS page scrape failed: {e}")
        self.logger.info("Using fallback corn-sweetener URL")
        return FALLBACK_URL

    def download(self) -> Tuple[Optional[bytes], Optional[str]]:
        url = self.resolve_url()
        try:
            r = self.session.get(url, timeout=90)
            if r.status_code == 200:
                return r.content, url
            self.logger.warning(f"ERS download HTTP {r.status_code}: {url}")
        except requests.RequestException as e:
            self.logger.warning(f"ERS download error: {e}")
        return None, url

    # ---- parse -------------------------------------------------------------
    def parse(self, raw: bytes) -> Tuple[List[Dict], Optional[date]]:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        release = self._release_date(wb)
        recs: List[Dict] = []
        if "Table29" in wb.sheetnames:
            recs += self._parse_period_rows(wb["Table29"], "Table 29", "hfcs_total")
        if "Table30" in wb.sheetnames:
            recs += self._parse_table30(wb["Table30"])
        if "Table37" in wb.sheetnames:
            recs += self._parse_annual_col(wb["Table37"], "Table 37", "dextrose")
        if "Table38" in wb.sheetnames:
            recs += self._parse_annual_col(wb["Table38"], "Table 38", "glucose")
        wb.close()
        for r in recs:
            r["vintage"] = release
            r["source_release_date"] = release
        return recs, release

    def _release_date(self, wb) -> Optional[date]:
        try:
            ws = wb["Contents"]
            for row in ws.iter_rows(values_only=True):
                for c in row:
                    if c and "last updated" in str(c).lower():
                        m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", str(c))
                        if m:
                            return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        except Exception:
            pass
        return None

    def _parse_period_rows(self, ws, table: str, product: str) -> List[Dict]:
        """Table 29: row labels = period (Q1-Q4, Fiscal/Calendar year), cols = years."""
        rows = list(ws.iter_rows(values_only=True))
        hdr = rows[1]                       # year header
        years = [(i, int(v)) for i, v in enumerate(hdr) if isinstance(v, (int, float)) and 1900 < v < 2100]
        out = []
        for row in rows[2:]:
            label = str(row[0] or "").strip()
            if not label:
                continue
            ptype, qtr = self._classify_period(label)
            if not ptype:
                continue
            for i, yr in years:
                v = row[i] if i < len(row) else None
                if isinstance(v, (int, float)):
                    out.append(dict(source_table=table, product=product, measure="production",
                                    period_type=ptype, period_label=label, year=yr,
                                    quarter=qtr, raw_value=float(v)))
        return out

    def _parse_table30(self, ws) -> List[Dict]:
        """Table 30: rows = years; cols 2,3,4 = Production HFCS-42 / HFCS-55 / total."""
        rows = list(ws.iter_rows(values_only=True))
        out = []
        prodmap = [(2, "hfcs_42"), (3, "hfcs_55"), (4, "hfcs_total")]
        for row in rows[2:]:
            yr = row[0]
            if not isinstance(yr, (int, float)) or not (1900 < yr < 2100):
                continue
            for idx, prod in prodmap:
                v = row[idx] if idx < len(row) else None
                if isinstance(v, (int, float)):
                    out.append(dict(source_table="Table 30", product=prod, measure="production",
                                    period_type="annual_calendar", period_label="Calendar year",
                                    year=int(yr), quarter=None, raw_value=float(v)))
        return out

    def _parse_annual_col(self, ws, table: str, product: str) -> List[Dict]:
        """Table 37/38: rows = years; col 1 = Production."""
        rows = list(ws.iter_rows(values_only=True))
        out = []
        for row in rows[2:]:
            yr = row[0]
            if not isinstance(yr, (int, float)) or not (1900 < yr < 2100):
                continue
            v = row[1] if len(row) > 1 else None
            if isinstance(v, (int, float)):
                out.append(dict(source_table=table, product=product, measure="production",
                                period_type="annual_calendar", period_label="Calendar year",
                                year=int(yr), quarter=None, raw_value=float(v)))
        return out

    def _classify_period(self, label: str) -> Tuple[Optional[str], Optional[int]]:
        u = label.upper()
        for q, n in _QTR.items():
            if u.startswith(q):
                return "quarterly", n
        if "FISCAL" in u:
            return "annual_fiscal", None
        if "CALENDAR" in u:
            return "annual_calendar", None
        return None, None

    # ---- persist + collect -------------------------------------------------
    def save_to_bronze(self, recs: List[Dict], source_url: str) -> int:
        import psycopg2
        conn = psycopg2.connect(
            host=os.environ.get("RLC_PG_HOST"), port=os.environ.get("RLC_PG_PORT", 5432),
            database=os.environ.get("RLC_PG_DATABASE", "rlc_commodities"),
            user=os.environ.get("RLC_PG_USER", "postgres"),
            password=os.environ.get("RLC_PG_PASSWORD"),
            sslmode=os.environ.get("RLC_PG_SSLMODE", "require"))
        n = 0
        try:
            cur = conn.cursor()
            for r in recs:
                cur.execute("""
                    INSERT INTO bronze.corn_products_raw
                        (source, source_url, source_release_date, source_table, product,
                         measure, period_type, period_label, year, quarter, raw_value,
                         raw_unit, vintage)
                    VALUES ('ERS_SUGAR_SWEETENERS',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                            '1000 short tons, dry',%s)
                    ON CONFLICT (source_table, product, measure, period_type, year,
                        COALESCE(quarter,0), COALESCE(vintage,'1900-01-01')) DO UPDATE SET
                        raw_value=EXCLUDED.raw_value, pull_ts=NOW()
                """, (source_url, r.get("source_release_date"), r["source_table"], r["product"],
                      r["measure"], r["period_type"], r.get("period_label"), r["year"],
                      r.get("quarter"), r["raw_value"], r.get("vintage")))
                n += 1
            conn.commit()
        finally:
            conn.close()
        return n

    def collect(self, **kwargs):
        from dataclasses import dataclass, field as dc_field
        @dataclass
        class _Result:
            success: bool = False
            source: str = "ers_corn_sweetener"
            records_fetched: int = 0
            error_message: Optional[str] = None
            warnings: list = dc_field(default_factory=list)
            collected_at: datetime = dc_field(default_factory=datetime.now)
            data_as_of: Optional[str] = None
        res = _Result()
        raw, url = self.download()
        if not raw:
            res.error_message = "could not download ERS corn sweetener workbook"
            return res
        recs, release = self.parse(raw)
        if not recs:
            res.error_message = "no records parsed"
            return res
        res.records_fetched = self.save_to_bronze(recs, url)
        res.data_as_of = str(release) if release else None
        res.success = res.records_fetched > 0
        return res


def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s %(levelname)s %(message)s")
    c = ERSCornSweetenerCollector()
    res = c.collect()
    print(f"success={res.success} rows={res.records_fetched} release={res.data_as_of} err={res.error_message}")


if __name__ == "__main__":
    main()
