"""Step 7 -- sheet blocks and formula-criteria mining. Depth, on the oils/fats chain only.

This is the step that closes Q1 end to end, and it works because of one measured fact: the
balance-sheet formulas carry the flat-file series key as string literals.

    soyoil_balance_sheet!B37
    =IF(COUNTIFS(ff_sbo_supply!$A$2:$A$8001,"soybean_oil",
                 ff_sbo_supply!$B$2:$B$8001,"ALL",
                 ff_sbo_supply!$C$2:$C$8001,"production",
                 ff_sbo_supply!$D$2:$D$8001,LEFT(B$3,4)*1,
                 ff_sbo_supply!$E$2:$E$8001,"cal_month",
                 ff_sbo_supply!$F$2:$F$8001,"M10")=0, '[2]Census Crush'!$K135, SUMIFS(...))

`"soybean_oil"` / `"ALL"` / `"production"` / `"cal_month"` / `"M10"` are commodity / class /
series / period_type / period -- columns 1,2,3,5,6 of the flat-file contract. So the
spreadsheet-to-series edge is auto-extractable, not something a human has to declare. That was
the open risk in the design and this closes it.

Two anchors, both stable, neither positional (D3):
  * the block title in column A -- `US SOYBEAN OIL PRODUCTION` at r35, 16-row pitch. This is
    the same anchor `SoyOilRepointToFlatFile.bas` binds by, and its header says why:
    "read it; never count rows."
  * the flat-file series key mined from the criteria.

Cell addresses appear only in edge evidence, as observations for sizing blast radius.

Scope is D6: the vegetable-oil / fats-and-greases / BBD feedstock chain. Breadth without depth
still answers Q3; depth without breadth answers nothing outside one chain. Do both, at
different grains.
"""

from __future__ import annotations

import re
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]

IN_SCOPE_DIRS = (
    "models/Oilseeds/United States",
    "models/Fats and Greases",
    "models/Biofuels",
)
EXCLUDE = re.compile(r"(?:^|/)(Archive|archive)/|backup[_.]?\d{6,}|_old\b|conflicted copy", re.I)

# Flat-file contract section 2, columns 1-9. Verified against
# us_soybean_complex_bal_sheets.xlsm: $A=commodity ... $F=period, $H=vintage_rank, $I=value.
# Not checked on any other workbook -- if a second workbook uses a different layout, the
# mined keys for it will be wrong in a way that looks right. See the not-verified list.
COLUMN_MAP = {
    "A": "commodity", "B": "class", "C": "series", "D": "marketing_year",
    "E": "period_type", "F": "period", "G": "vintage", "H": "vintage_rank", "I": "value",
}

# Only these become part of the series identity. marketing_year / period vary row by row
# across a block; they are not what distinguishes one series from another.
IDENTITY_COLS = ("commodity", "class", "series")

# A criterion literal can be an Excel comparison operator rather than a value -- `"<>"` means
# "not blank". The first pass mined 20 `class=<>` series out of rfs_data.xlsm before this
# filter existed. An operator is not an identity.
OPERATOR_LITERAL = re.compile(r"^\s*(<>|<=|>=|[<>=]|\*|\?)\s*[\d.]*\s*$")

# tab!$C$2:$C$8001,"literal"   and   tab!$C:$C,"literal"
CRITERION_RE = re.compile(
    r"""(?:'([^']+)'|([A-Za-z0-9_.]+))!          # tab name, optionally quoted
        \$([A-Z]{1,3})(?:\$?\d+)?                 # column letter (row part ignored -- D3)
        :\$?[A-Z]{1,3}(?:\$?\d+)?                 # ... : end of range
        \s*,\s*"([^"]*)"                          # the literal criterion
    """,
    re.X,
)

# '[2]Census Crush'!$K135  or  [1]biodiesel_monthly!...
EXTLINK_RE = re.compile(r"\[(\d+)\]")

TITLE_RE = re.compile(r"^[A-Z][A-Z0-9 &/,'()%.\-]{7,}$")

# How many filled cells may sit to the right of a column-A string before it stops being a
# block title and starts being a data label. Trade sheets carry country names in column A --
# CHINA, SOUTH KOREA -- with a decade of numbers beside them, and the first pass called all
# 177 of them blocks. A real block title heads an otherwise empty row.
MAX_FILLED_RIGHT_OF_TITLE = 1
TITLE_SCAN_WIDTH = 12


def _in_scope(rel: str) -> bool:
    if EXCLUDE.search(rel):
        return False
    return any(rel.startswith(d) for d in IN_SCOPE_DIRS)


def extract(conn, store, external_link_index: dict[str, list[str]]) -> dict:
    import openpyxl

    stats = {
        "workbooks_in_scope": 0, "workbooks_failed": 0, "sheets_scanned": 0,
        "blocks": 0, "formula_cells": 0, "criteria_mined": 0,
        "flat_file_series": 0, "binds_to_edges": 0, "extlink_block_edges": 0,
        "cells_read": 0,
    }
    failures: list[str] = []

    targets = []
    for pat in ("*.xlsm", "*.xlsx"):
        for p in ROOT.rglob(pat):
            rel = p.relative_to(ROOT).as_posix()
            if p.name.startswith("~$") or not _in_scope(rel):
                continue
            targets.append((p, rel))

    for path, rel in sorted(targets, key=lambda t: t[1]):
        wb_key = f"wb:{rel}"
        if not store.has_node(wb_key):
            # step 6 owns workbook nodes; re-create a minimal one so edges can attach when
            # step 7 is run on its own.
            store.add_node("workbook", wb_key, label=path.name,
                           properties={"path": rel}, extraction_method="xlsx_formula",
                           confidence=1.00)
        stats["workbooks_in_scope"] += 1
        t0 = time.time()
        try:
            wbk = openpyxl.load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:  # noqa: BLE001
            stats["workbooks_failed"] += 1
            failures.append(f"{rel}: {type(exc).__name__}: {exc}")
            continue

        ext_targets = external_link_index.get(wb_key, [])

        try:
            for sheet_name in wbk.sheetnames:
                ws = wbk[sheet_name]
                if not hasattr(ws, "iter_rows"):
                    continue  # chartsheet -- no cells, nothing to mine
                ws_key = f"{wb_key}#{sheet_name}"
                store.add_node("worksheet", ws_key, label=sheet_name,
                               properties={"workbook": rel, "sheet": sheet_name},
                               extraction_method="xlsx_formula", confidence=1.00)
                store.add_edge(wb_key, "DEFINES", ws_key,
                               extraction_method="xlsx_formula", confidence=1.00)
                stats["sheets_scanned"] += 1

                blocks: list[tuple[int, str]] = []      # (row, title)
                # (block_row, flat_tab, identity tuple) -> [cell count, sample cell]
                bindings: dict[tuple[int, str, tuple], list] = {}
                ext_hits: dict[tuple[int, int], list] = {}

                for row in ws.iter_rows():
                    filled_right = sum(
                        1 for c in row[1:TITLE_SCAN_WIDTH]
                        if c.value is not None and str(c.value).strip() != ""
                    )
                    for cell in row:
                        v = cell.value
                        if v is None:
                            continue
                        stats["cells_read"] += 1
                        if cell.column == 1 and isinstance(v, str):
                            s = v.strip()
                            if TITLE_RE.match(s) and filled_right <= MAX_FILLED_RIGHT_OF_TITLE:
                                blocks.append((cell.row, s))
                        if not (isinstance(v, str) and v.startswith("=")):
                            continue
                        stats["formula_cells"] += 1

                        block_row = _owning_block(blocks, cell.row)

                        crits: dict[str, dict[str, str]] = {}
                        for m in CRITERION_RE.finditer(v):
                            tab = m.group(1) or m.group(2)
                            col = COLUMN_MAP.get(m.group(3))
                            lit = m.group(4)
                            if col is None or not lit or OPERATOR_LITERAL.match(lit):
                                continue
                            crits.setdefault(tab, {})[col] = lit
                            stats["criteria_mined"] += 1

                        for tab, kv in crits.items():
                            ident = tuple((c, kv[c]) for c in IDENTITY_COLS if c in kv)
                            # commodity is column A of the flat-file contract and is present
                            # on every real key. Without it we are looking at some other
                            # SUMIFS pattern that happens to use columns A-C, and inventing a
                            # flat_file_series for it would be fiction.
                            if not ident or "commodity" not in dict(ident):
                                continue
                            k = (block_row, tab, ident)
                            entry = bindings.setdefault(k, [0, cell.coordinate, kv])
                            entry[0] += 1

                        for m in EXTLINK_RE.finditer(v):
                            idx = int(m.group(1))
                            if 1 <= idx <= len(ext_targets):
                                k2 = (block_row, idx)
                                e = ext_hits.setdefault(k2, [0, cell.coordinate])
                                e[0] += 1

                block_key_by_row = {}
                for brow, title in blocks:
                    bkey = f"{ws_key}#{title}"
                    block_key_by_row[brow] = bkey
                    store.add_node(
                        "sheet_block", bkey, label=title,
                        properties={"workbook": rel, "sheet": sheet_name, "title": title,
                                    "observed_row": brow},
                        extraction_method="xlsx_formula", confidence=0.90,
                    )
                    store.add_edge(ws_key, "DEFINES", bkey,
                                   extraction_method="xlsx_formula", confidence=0.90)
                    stats["blocks"] += 1

                for (brow, tab, ident), (n_cells, sample, kv) in bindings.items():
                    target = block_key_by_row.get(brow, ws_key)
                    ffs_key = f"{wb_key}#{tab}#" + ",".join(f"{k}={v}" for k, v in ident)
                    store.add_node(
                        "flat_file_series", ffs_key,
                        label=",".join(v for _, v in ident),
                        properties={"workbook": rel, "tab": tab, "keys": dict(ident),
                                    "criteria_seen": kv},
                        extraction_method="xlsx_formula", confidence=0.70,
                    )
                    stats["flat_file_series"] += 1
                    # Data flows out of the flat-file tab into the block.
                    store.add_edge(
                        ffs_key, "BINDS_TO", target,
                        evidence={"cells": n_cells, "sample": sample},
                        extraction_method="xlsx_formula", confidence=0.70,
                    )
                    stats["binds_to_edges"] += 1

                for (brow, idx), (n_cells, sample) in ext_hits.items():
                    tkey = ext_targets[idx - 1]
                    if not tkey or not store.has_node(tkey):
                        continue
                    target = block_key_by_row.get(brow, ws_key)
                    store.add_edge(
                        tkey, "LINKS_TO", target,
                        properties={"link_index": idx},
                        evidence={"cells": n_cells, "sample": sample},
                        extraction_method="xlsx_formula", confidence=1.00,
                    )
                    stats["extlink_block_edges"] += 1
        finally:
            wbk.close()

        stats.setdefault("per_workbook_seconds", {})[rel] = round(time.time() - t0, 1)

    stats["failures"] = failures
    return stats


def _owning_block(blocks: list[tuple[int, str]], row: int) -> int:
    """The block a cell belongs to is the nearest title at or above it. Blocks are found in
    row order because iter_rows walks in row order."""
    owner = 0
    for brow, _ in blocks:
        if brow <= row:
            owner = brow
        else:
            break
    return owner
